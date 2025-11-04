# app/services/processing_service.py
"""Core processing service for financial analysis and Excel report generation."""

from __future__ import annotations

import io
from typing import Optional, List, Tuple
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from ..data.data_utils import DEFAULT_CONFIG, EXCEL_PROCESSING, FILE_PROCESSING
from ..data.excel_processing import (
    extract_subsidiary_name_from_bytes, process_financial_tab_from_bytes,
    apply_excel_formatting_ws, _add_revenue_analysis_to_sheet, _add_consolidated_months_analysis_to_sheet,
    _add_month_to_month_analysis_to_sheet
)
from ..analysis.anomaly_detection import build_anoms_python_mode, build_anoms_ai_mode
from ..analysis.revenue_analysis import analyze_comprehensive_revenue_impact_from_bytes, analyze_revenue_variance_comprehensive
from ..analysis.revenue_variance_excel import _add_revenue_variance_analysis_to_sheet
from ..utils.logging_config import get_logger

logger = get_logger(__name__)

def process_all(
    files: list[tuple[str, bytes]],
    corr_rules: Optional[pd.DataFrame] = None,
    season_rules: Optional[pd.DataFrame] = None,
    CONFIG: dict = DEFAULT_CONFIG,
    progress_callback=None
) -> bytes | tuple[bytes, list[tuple[str, bytes]]]:
    """
    Unified processing function that handles both Python and AI analysis modes.

    Args:
        files: List of (filename, bytes) tuples
        corr_rules: Correlation rules DataFrame (Python mode only)
        season_rules: Seasonality rules DataFrame (Python mode only)
        CONFIG: Configuration dictionary
        progress_callback: Progress callback function (AI mode only)

    Returns:
        bytes: Excel file bytes (Python mode)
        tuple[bytes, list[tuple[str, bytes]]]: Excel bytes and debug files (AI mode)
    """
    use_ai = CONFIG.get("use_llm_analysis", False)

    if use_ai:
        return process_all_ai_mode(files, CONFIG, progress_callback)
    else:
        return process_all_python_mode(files, corr_rules, season_rules, CONFIG)

def process_all_python_mode(
    files: list[tuple[str, bytes]],
    corr_rules: Optional[pd.DataFrame] = None,
    season_rules: Optional[pd.DataFrame] = None,
    CONFIG: dict = DEFAULT_CONFIG
) -> bytes:
    """Python rule-based analysis mode."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Anomalies Summary"
    all_anoms: list[pd.DataFrame] = []

    # default empty rules if None
    corr_rules = corr_rules if corr_rules is not None else pd.DataFrame()
    season_rules = season_rules if season_rules is not None else pd.DataFrame()

    for fname, xl_bytes in files:
        sub = extract_subsidiary_name_from_bytes(xl_bytes, fname)

        # Be forgiving if a sheet is missing
        bs_df, bs_cols = pd.DataFrame(), []
        pl_df, pl_cols = pd.DataFrame(), []
        try:
            bs_df, bs_cols = process_financial_tab_from_bytes(xl_bytes, "BS Breakdown", "BS", sub)
        except Exception:
            pass
        try:
            pl_df, pl_cols = process_financial_tab_from_bytes(xl_bytes, "PL Breakdown", "PL", sub)
        except Exception:
            pass

        anoms = build_anoms_python_mode(sub, bs_df, bs_cols, pl_df, pl_cols, corr_rules, season_rules, CONFIG)
        if anoms is not None and not anoms.empty:
            all_anoms.append(anoms)

    # Safe concat (even if no anomalies/files)
    if all_anoms:
        anom_df = pd.concat(all_anoms, ignore_index=True)
    else:
        anom_df = pd.DataFrame(columns=[
            "Subsidiary","Account","Period","Pct Change","Abs Change (VND)",
            "Trigger(s)","Suggested likely cause","Status","Notes"
        ])

    for r in dataframe_to_rows(anom_df, index=False, header=True):
        ws.append(r)
    apply_excel_formatting_ws(ws, anom_df, CONFIG)

    # === ADD CLEANED SHEETS FOR EACH FILE ===
    logger.info("📊 Adding cleaned BS and PL sheets...")
    try:
        for idx, (fname, xl_bytes) in enumerate(files):
            sub = extract_subsidiary_name_from_bytes(xl_bytes, fname)
            file_prefix = f"{sub}_{idx+1}" if len(files) > 1 else sub

            # Add cleaned Balance Sheet
            try:
                bs_df, bs_cols = process_financial_tab_from_bytes(xl_bytes, "BS Breakdown", "BS", sub)
                if not bs_df.empty:
                    bs_ws = wb.create_sheet(title=f"{file_prefix}_BS_Cleaned"[:CONFIG["max_sheet_name_length"]])
                    for r in dataframe_to_rows(bs_df, index=False, header=True):
                        bs_ws.append(r)
                    logger.info(f"✅ Added cleaned BS sheet for {sub}")
            except Exception as e:
                logger.warning(f"⚠️  Could not add cleaned BS sheet for {sub}: {e}")

            # Add cleaned Profit & Loss Sheet
            try:
                pl_df, pl_cols = process_financial_tab_from_bytes(xl_bytes, "PL Breakdown", "PL", sub)
                if not pl_df.empty:
                    pl_ws = wb.create_sheet(title=f"{file_prefix}_PL_Cleaned"[:CONFIG["max_sheet_name_length"]])
                    for r in dataframe_to_rows(pl_df, index=False, header=True):
                        pl_ws.append(r)
                    logger.info(f"✅ Added cleaned PL sheet for {sub}")
            except Exception as e:
                logger.warning(f"⚠️  Could not add cleaned PL sheet for {sub}: {e}")

    except Exception as e:
        logger.error(f"⚠️  Error adding cleaned sheets: {e}", exc_info=True)

    # === ADD CONSOLIDATED MONTHS ANALYSIS SHEETS ===
    logger.info("📊 Adding Consolidated Months Analysis sheets...")
    try:
        # Run comprehensive revenue variance analysis for the first file
        if files:
            first_file_name, first_file_bytes = files[0]

            # Run both analyses
            logger.info("🚀 Running comprehensive revenue variance analysis...")
            variance_analysis = analyze_revenue_variance_comprehensive(first_file_bytes, first_file_name, CONFIG)

            logger.info("📊 Running comprehensive revenue impact analysis...")
            revenue_analysis = analyze_comprehensive_revenue_impact_from_bytes(first_file_bytes, first_file_name, CONFIG)

            # Create consolidated Months Analysis sheet (all months)
            logger.info("📝 Creating 'Months Analysis' sheet (all months)...")
            months_ws = wb.create_sheet(title="Months Analysis")
            _add_consolidated_months_analysis_to_sheet(months_ws, revenue_analysis, variance_analysis)
            logger.info("✅ Months Analysis sheet added successfully")

            # Create Month to Month Analysis sheet (latest 2 months only)
            logger.info("📝 Creating 'Month to Month Analysis' sheet (latest 2 months)...")
            month_to_month_ws = wb.create_sheet(title="Month to Month Analysis")
            _add_month_to_month_analysis_to_sheet(month_to_month_ws, revenue_analysis, variance_analysis, first_file_bytes, first_file_name)
            logger.info("✅ Month to Month Analysis sheet added successfully")

    except Exception as e:
        logger.error(f"⚠️  Months Analysis sheets creation failed: {e}", exc_info=True)
        # Continue without months analysis if it fails


    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

def process_all_ai_mode(
    files: list[tuple[str, bytes]],
    CONFIG: dict = DEFAULT_CONFIG,
    progress_callback=None
) -> tuple[bytes, list[tuple[str, bytes]]]:
    """AI-powered analysis mode using same Excel format as Python mode."""
    logger.info("🚀 ===== STARTING AI VARIANCE ANALYSIS PROCESSING =====")
    logger.info(f"📥 Processing {len(files)} Excel file(s) for AI analysis")
    logger.info(f"🤖 LLM Model: {CONFIG.get('llm_model', 'gpt-4o')}")
    logger.info(f"🔧 AI-Only Mode: {CONFIG.get('use_llm_analysis', True)}")

    # Collect variance flags in Python mode format
    all_variance_flags: list[dict] = []
    all_file_data: list[dict] = []
    debug_files: list[tuple[str, bytes]] = []
    logger.info("✅ Starting AI analysis processing...")

    # === MULTI-FILE PROCESSING LOOP ===
    logger.info(f"🔄 Starting processing loop for {len(files)} file(s)...")

    for file_idx, (fname, xl_bytes) in enumerate(files, 1):
        # Calculate progress range for this file (30% to 80% of total)
        file_start = 30 + ((file_idx - 1) * 50 // len(files))
        file_end = 30 + (file_idx * 50 // len(files))

        if progress_callback:
            progress_callback(file_start, f"Processing file {file_idx}/{len(files)}: {fname}")

        logger.info(f"📁 ===== PROCESSING FILE {file_idx}/{len(files)} =====")
        logger.info(f"📄 File: {fname}")
        logger.info(f"📏 File Size: {len(xl_bytes):,} bytes ({len(xl_bytes)/CONFIG['bytes_per_kb']:.1f} KB)")

        if progress_callback:
            progress_callback(file_start + 2, f"Extracting subsidiary name from {fname}")

        logger.info("🏢 Extracting subsidiary name...")
        sub = extract_subsidiary_name_from_bytes(xl_bytes, fname)
        logger.info(f"✅ Subsidiary: '{sub}'")

        if progress_callback:
            progress_callback(file_start + 5, f"Starting AI analysis for {sub}")

        # === AI ANALYSIS (Returns Python mode format) ===
        logger.info(f"🤖 Starting AI analysis for '{sub}'...")
        from ..analysis.llm_analyzer import LLMFinancialAnalyzer

        llm_analyzer = LLMFinancialAnalyzer(CONFIG.get("llm_model", "gpt-4o"), progress_callback=progress_callback, initial_progress=file_start + 5)
        variance_flags = llm_analyzer.analyze_raw_excel_file(xl_bytes, fname, sub, CONFIG)

        if progress_callback:
            progress_callback(file_end - 5, f"AI analysis complete for {sub}")

        logger.info(f"✅ AI analysis completed for '{sub}'")
        logger.info(f"   • Variance flags detected: {len(variance_flags)}")

        # Convert variance flags to Python mode format
        for flag in variance_flags:
            all_variance_flags.append(flag)

        # Also collect BS/PL data for raw data sheets
        try:
            bs_df, bs_cols = process_financial_tab_from_bytes(xl_bytes, "BS Breakdown", "BS", sub)
            pl_df, pl_cols = process_financial_tab_from_bytes(xl_bytes, "PL Breakdown", "PL", sub)

            all_file_data.append({
                'subsidiary': sub,
                'filename': fname,
                'bs_df': bs_df,
                'pl_df': pl_df
            })
            logger.info(f"✅ Collected BS/PL data for {sub}")
        except Exception as e:
            logger.warning(f"⚠️  Could not collect BS/PL data for {sub}: {e}")

        logger.info(f"✅ File '{fname}' processing completed")

    # === EXCEL GENERATION (Using Python mode format) ===
    logger.info("📊 ===== GENERATING EXCEL OUTPUT (Python Mode Format) =====")

    # Import Python mode's Excel generation function
    from ..analysis.variance_analysis_pipeline import create_excel_output

    logger.info(f"   • Total variance flags: {len(all_variance_flags)}")
    logger.info(f"   • Files processed: {len(all_file_data)}")

    # Use Python mode's Excel generation
    xlsx_bytes = create_excel_output(all_file_data, all_variance_flags)

    final_size = len(xlsx_bytes)
    logger.info("✅ Excel file generated successfully (Python mode format)")
    logger.info(f"   • Output size: {final_size:,} bytes ({final_size/1024:.1f} KB)")

    logger.info("🎉 ===== AI VARIANCE ANALYSIS COMPLETED =====")
    return xlsx_bytes, debug_files