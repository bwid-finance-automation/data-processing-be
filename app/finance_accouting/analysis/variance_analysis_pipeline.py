#!/usr/bin/env python3
"""
Variance Analysis Pipeline - 21 Rules

This module processes financial data (BS and PL) and applies 21 variance analysis rules
to detect anomalies, process breakdowns, and compliance issues.

Rules:
- 8 Critical Rules (🔴): Data integrity / Process breakdown
- 9 Review Rules (🟡): Material movements requiring explanation
- 4 Info Rules (🟢): Optimization / Expected patterns

Adapted for FastAPI backend integration.
"""

import re
import io
from typing import List, Tuple, Optional
from pathlib import Path
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import warnings
warnings.filterwarnings('ignore')


# ============================================================================
# CONFIGURATION AND CONSTANTS
# ============================================================================

# Statistical thresholds
STD_DEV_THRESHOLD = 2.0  # 2 standard deviations for outlier detection
MONTHS_3M_AVG = 3  # 3-month average baseline
MONTHS_6M_AVG = 6  # 6-month average baseline

# Materiality thresholds (VND)
MATERIALITY_THRESHOLDS = {
    'vat_input_threshold': 10_000_000_000,  # 10B VND for E3
    'retained_earnings_tolerance': 100_000_000,  # 100M VND for D2
    'ar_growth_threshold': 0.30,  # 30% AR growth for E5
}

# Month tokens for detection
MONTH_TOKENS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
MONTHS = ["jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec"]

# Days in each month (non-leap year)
DAYS_IN_MONTH = {
    'Jan': 31, 'Feb': 28, 'Mar': 31, 'Apr': 30,
    'May': 31, 'Jun': 30, 'Jul': 31, 'Aug': 31,
    'Sep': 30, 'Oct': 31, 'Nov': 30, 'Dec': 31
}

# Quarter-start months for E4
QUARTER_START_MONTHS = ['Jan', 'Apr', 'Jul', 'Oct']

# Regex patterns for BS and PL month detection
BS_PAT = re.compile(r'^\s*as\s*of\s*(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[\.\-\s]*(\d{2,4})\s*$', re.I)
PL_PAT = re.compile(r'^\s*(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[\.\-\s]*(\d{2,4})\s*$', re.I)


# ============================================================================
# EXCEL PROCESSING FUNCTIONS
# ============================================================================

def extract_current_period_from_row4(xl_file, sheet_name):
    """Extract current period from Row 4 (index 3) of Excel sheet."""
    try:
        df = pd.read_excel(xl_file, sheet_name=sheet_name, header=None, nrows=5, dtype=str)
        row4_text = str(df.iloc[3, 0]).strip()  # Row 4, Column A

        matches = re.findall(r'(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\s+(\d{4})', row4_text, re.I)

        if matches:
            month_name, year = matches[-1]
            month_num = MONTHS.index(month_name.lower()) + 1
            return int(year), month_num, f"{month_name.title()} {year}"
        else:
            return None, None, None
    except Exception:
        return None, None, None


def filter_months_up_to_current(month_cols, current_year, current_month):
    """Filter month columns to only include months up to current period."""
    if current_year is None or current_month is None:
        return month_cols

    filtered_months = []
    for month_col in month_cols:
        m = re.search(r'(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\s+(\d{4})', month_col, re.I)
        if m:
            month_name = m.group(1).lower()
            year = int(m.group(2))
            month_num = MONTHS.index(month_name) + 1

            if (year < current_year) or (year == current_year and month_num <= current_month):
                filtered_months.append(month_col)

    return filtered_months


def normalize_period_label(label):
    """Normalize period labels to 'Mon YYYY' format"""
    if label is None: return ""
    s = str(label).strip()
    if s == "": return ""
    try:
        s_clean = re.sub(r'^\s*(as\s*of|tinh\s*den|tính\s*đến|den\s*ngay|đến\s*ngày)\s*', '', s, flags=re.I)
        m = re.search(r'\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[^\w]?[\s\-\.]*([12]\d{3}|\d{2})\b', s_clean, flags=re.I)
        if m:
            mon, yr = m.group(1), m.group(2)
            yr = int(yr)
            yr = yr+2000 if yr < 100 else yr
            return f"{mon.title()} {yr}"
        m = re.search(r'\b(1[0-2]|0?[1-9])[./\-](\d{4})\b', s_clean)
        if m:
            mon = int(m.group(1))
            yr = int(m.group(2))
            return f"{MONTHS[mon-1].title()} {yr}"
        m = re.search(r'\b(\d{4})[./\-](1[0-2]|0?[1-9])\b', s_clean)
        if m:
            yr = int(m.group(1))
            mon = int(m.group(2))
            return f"{MONTHS[mon-1].title()} {yr}"
        m_year = re.search(r'(20\d{2}|19\d{2})', s_clean)
        m_mon = re.search(r'\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\b', s_clean, flags=re.I)
        if m_year and m_mon:
            yr = int(m_year.group(1))
            mon = m_mon.group(0)
            return f"{mon.title()} {yr}"
    except Exception:
        pass
    return s


def month_key(label):
    """Convert month label to sortable tuple (year, month_num)"""
    n = normalize_period_label(label)
    m = re.search(r'(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\s+(\d{4})', n, re.I)
    if not m: return (9999, 99)
    y = int(m.group(2))
    mi = MONTHS.index(m.group(1).lower()) + 1
    return (y, mi)


def detect_header_row(xl, sheet):
    """Find the row containing 'Financial row' header"""
    try:
        probe = pd.read_excel(xl, sheet_name=sheet, header=None, nrows=40)
        for i in range(len(probe)):
            row_values = probe.iloc[i].astype(str).str.strip().str.lower()
            if any("financial row" in v for v in row_values):
                return i
    except Exception:
        pass
    return 0


def process_financial_tab(xl_file, sheet_name, mode, subsidiary):
    """Process BS or PL sheet and extract data."""
    try:
        current_year, current_month, current_period_str = extract_current_period_from_row4(xl_file, sheet_name)
        header_row = detect_header_row(xl_file, sheet_name)
        df_raw = pd.read_excel(xl_file, sheet_name=sheet_name, header=None, dtype=str)
        headers = df_raw.iloc[header_row].tolist()
        data_start = header_row + 2
        df = df_raw.iloc[data_start:].copy()
        df.columns = [f'col_{i}' for i in range(len(df.columns))]

        if mode == 'BS':
            df = df.rename(columns={
                'col_0': 'Financial_Row',
                'col_1': 'Entity',
                'col_2': 'Entity_Line',
                'col_3': 'Account_Line'
            })
            filter_col = 'Account_Line'
            month_start_col = 4
        else:  # PL
            df = df.rename(columns={
                'col_0': 'Financial_Row',
                'col_1': 'Entity',
                'col_2': 'Account_Line'
            })
            filter_col = 'Account_Line'
            month_start_col = 3

        df['Financial_Row'] = df['Financial_Row'].fillna(method='ffill')
        df = df[df[filter_col].notna() & (df[filter_col].astype(str).str.strip() != '')].copy()

        month_row = df_raw.iloc[header_row + 1]
        month_cols = []
        month_col_indices = []
        for i in range(month_start_col, len(month_row)):
            val = str(month_row.iloc[i]).strip()
            if BS_PAT.match(val) or PL_PAT.match(val):
                normalized = normalize_period_label(val)
                val_lower = val.lower()
                if not any(keyword in val_lower for keyword in ['ytd', 'ltm', 'amount ytd', 'amount ltm', 'year to date', 'last twelve']):
                    if re.match(r'^(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\s+\d{4}$', normalized, re.I):
                        month_cols.append(normalized)
                        month_col_indices.append(i)

        for idx, month in zip(month_col_indices, month_cols):
            df = df.rename(columns={f'col_{idx}': month})

        for month in month_cols:
            if month in df.columns:
                series = df[month].astype(str)
                series = (series
                    .str.replace("\u00a0","", regex=False)
                    .str.replace(",","", regex=False)
                    .str.replace(r"\((.*)\)", r"-\1", regex=True)
                    .str.replace(r"[^0-9\.\-]", "", regex=True)
                )
                df[month] = pd.to_numeric(series, errors="coerce").fillna(0.0)

        final_cols = ['Financial_Row', 'Account_Line'] + month_cols
        df = df[[col for col in final_cols if col in df.columns]].copy()

        df = df.rename(columns={
            'Financial_Row': 'Account Code',
            'Account_Line': 'Account Line'
        })

        result = df.groupby(['Account Code', 'Account Line'], as_index=False).agg({
            **{month: 'sum' for month in month_cols if month in df.columns}
        })

        result['Account Name'] = result['Account Code']
        final_column_order = ['Account Code', 'Account Line', 'Account Name'] + [m for m in month_cols if m in result.columns]
        result = result[final_column_order]
        month_cols = sorted(month_cols, key=month_key)
        filtered_month_cols = filter_months_up_to_current(month_cols, current_year, current_month)

        keep_cols = ['Account Code', 'Account Line', 'Account Name'] + filtered_month_cols
        result = result[[col for col in keep_cols if col in result.columns]]

        return result, filtered_month_cols

    except Exception as e:
        print(f"Error processing {sheet_name} for {subsidiary}: {e}")
        return pd.DataFrame(), []


def extract_subsidiary_name(xl_file):
    """Extract subsidiary name from cell A2"""
    try:
        wb = load_workbook(xl_file, read_only=True, data_only=True)
        for sheet_name in ["BS Breakdown", "PL Breakdown"]:
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                cell_value = sheet["A2"].value
                if isinstance(cell_value, str) and ":" in cell_value:
                    return cell_value.split(":")[-1].strip()
        wb.close()
    except Exception:
        pass
    return "Subsidiary"


# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def extract_month_from_column(col_name):
    """Extract month name from column header like 'Jan 2025'"""
    col_str = str(col_name).replace('As of ', '').strip()
    for month in MONTH_TOKENS:
        if month in col_str:
            return month
    return None


def normalize_interest_for_calendar(value, month_name):
    """Normalize interest to 30-day basis for comparison"""
    if month_name not in DAYS_IN_MONTH:
        return value
    days = DAYS_IN_MONTH[month_name]
    if days == 30:
        return value
    return (value / days) * 30


def get_account_pattern_data(df, pattern):
    """Get rows matching account code pattern"""
    if df is None or df.empty:
        return pd.DataFrame()

    if pattern.endswith('xxx'):
        prefix = pattern[:-3]
        return df[df['Account Code'].astype(str).str.startswith(prefix)].copy()
    else:
        return df[df['Account Code'].astype(str) == pattern].copy()


def get_month_cols(df):
    """Extract month columns from dataframe"""
    if df is None or df.empty:
        return []

    month_cols = []
    for col in df.columns:
        if isinstance(col, str) and re.match(r'^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{4}$', col):
            month_cols.append(col)

    return sorted(month_cols, key=month_key)


# ============================================================================
# RULE IMPLEMENTATIONS - CRITICAL (🔴)
# ============================================================================

def check_rule_A1(bs_df, pl_df):
    """A1 - Asset capitalized but depreciation not started"""
    flags = []
    if bs_df is None or pl_df is None or bs_df.empty or pl_df.empty:
        return flags

    ip_accounts = get_account_pattern_data(bs_df, '217xxx')
    dep_accounts = get_account_pattern_data(pl_df, '632xxx')
    if ip_accounts.empty or dep_accounts.empty:
        return flags

    month_cols = get_month_cols(bs_df)
    if len(month_cols) < 2:
        return flags

    for i in range(1, len(month_cols)):
        prev_month = month_cols[i-1]
        curr_month = month_cols[i]

        ip_change = ip_accounts[curr_month].sum() - ip_accounts[prev_month].sum()
        dep_change = dep_accounts[curr_month].sum() - dep_accounts[prev_month].sum()

        if ip_change > 0 and dep_change <= 0:
            flags.append({
                'Rule_ID': 'A1',
                'Priority': '🔴 Critical',
                'Issue': 'Asset capitalized but depreciation not started',
                'Accounts': '217xxx ↔ 632xxx',
                'Period': f"{prev_month} → {curr_month}",
                'Reason': f'IP increased by {ip_change:,.0f} VND but Depreciation changed by {dep_change:,.0f} VND',
                'Flag_Trigger': 'IP↑ BUT Depreciation ≤ previous'
            })
    return flags


def check_rule_A2(bs_df, pl_df):
    """A2 - Loan drawdown but interest not recorded"""
    flags = []
    if bs_df is None or pl_df is None or bs_df.empty or pl_df.empty:
        return flags

    loan_accounts = get_account_pattern_data(bs_df, '341xxx')
    interest_expense = get_account_pattern_data(pl_df, '635xxx')
    cip_interest = get_account_pattern_data(bs_df, '241xxx')
    if loan_accounts.empty:
        return flags

    month_cols = get_month_cols(bs_df)
    if len(month_cols) < 2:
        return flags

    for i in range(1, len(month_cols)):
        prev_month = month_cols[i-1]
        curr_month = month_cols[i]

        prev_month_name = extract_month_from_column(prev_month)
        curr_month_name = extract_month_from_column(curr_month)

        loan_change = loan_accounts[curr_month].sum() - loan_accounts[prev_month].sum()

        interest_prev_raw = interest_expense[prev_month].sum() if not interest_expense.empty else 0
        interest_curr_raw = interest_expense[curr_month].sum() if not interest_expense.empty else 0
        cip_prev_raw = cip_interest[prev_month].sum() if not cip_interest.empty else 0
        cip_curr_raw = cip_interest[curr_month].sum() if not cip_interest.empty else 0

        interest_prev_adj = normalize_interest_for_calendar(interest_prev_raw, prev_month_name)
        interest_curr_adj = normalize_interest_for_calendar(interest_curr_raw, curr_month_name)
        cip_prev_adj = normalize_interest_for_calendar(cip_prev_raw, prev_month_name)
        cip_curr_adj = normalize_interest_for_calendar(cip_curr_raw, curr_month_name)

        interest_change = (interest_curr_adj + cip_curr_adj) - (interest_prev_adj + cip_prev_adj)

        if loan_change > 0 and interest_change <= 0:
            flags.append({
                'Rule_ID': 'A2',
                'Priority': '🔴 Critical',
                'Issue': 'Loan drawdown but interest not recorded',
                'Accounts': '341xxx ↔ 635xxx + 241xxx',
                'Period': f"{prev_month} → {curr_month}",
                'Reason': f'Loan increased by {loan_change:,.0f} VND but day-adjusted interest changed by {interest_change:,.0f} VND',
                'Flag_Trigger': 'Loan↑ BUT Day-adjusted Interest ≤ previous'
            })
    return flags


def check_rule_A3(bs_df, pl_df):
    """A3 - Capex incurred but VAT not recorded"""
    flags = []
    if bs_df is None or bs_df.empty:
        return flags

    ip_accounts = get_account_pattern_data(bs_df, '217xxx')
    cip_accounts = get_account_pattern_data(bs_df, '241xxx')
    vat_accounts = get_account_pattern_data(bs_df, '133xxx')
    if (ip_accounts.empty and cip_accounts.empty) or vat_accounts.empty:
        return flags

    month_cols = get_month_cols(bs_df)
    if len(month_cols) < 2:
        return flags

    for i in range(1, len(month_cols)):
        prev_month = month_cols[i-1]
        curr_month = month_cols[i]

        ip_change = ip_accounts[curr_month].sum() - ip_accounts[prev_month].sum() if not ip_accounts.empty else 0
        cip_change = cip_accounts[curr_month].sum() - cip_accounts[prev_month].sum() if not cip_accounts.empty else 0
        total_asset_change = ip_change + cip_change
        vat_change = vat_accounts[curr_month].sum() - vat_accounts[prev_month].sum()

        if total_asset_change > 0 and vat_change <= 0:
            flags.append({
                'Rule_ID': 'A3',
                'Priority': '🔴 Critical',
                'Issue': 'Capex incurred but VAT not recorded',
                'Accounts': '217/241 ↔ 133xxx',
                'Period': f"{prev_month} → {curr_month}",
                'Reason': f'Assets (IP+CIP) increased by {total_asset_change:,.0f} VND but VAT input changed by {vat_change:,.0f} VND',
                'Flag_Trigger': 'Assets↑ BUT VAT input ≤ previous'
            })
    return flags


def check_rule_A4(bs_df, pl_df):
    """A4 - Cash movement disconnected from interest"""
    flags = []
    if bs_df is None or pl_df is None or bs_df.empty or pl_df.empty:
        return flags

    cash_111 = get_account_pattern_data(bs_df, '111xxx')
    cash_112 = get_account_pattern_data(bs_df, '112xxx')
    cash_accounts = pd.concat([cash_111, cash_112], ignore_index=True)
    interest_income = get_account_pattern_data(pl_df, '515xxx')
    if cash_accounts.empty or interest_income.empty:
        return flags

    month_cols = get_month_cols(bs_df)
    if len(month_cols) < 2:
        return flags

    for i in range(1, len(month_cols)):
        prev_month = month_cols[i-1]
        curr_month = month_cols[i]

        prev_month_name = extract_month_from_column(prev_month)
        curr_month_name = extract_month_from_column(curr_month)

        cash_change = cash_accounts[curr_month].sum() - cash_accounts[prev_month].sum()

        interest_prev_adj = normalize_interest_for_calendar(interest_income[prev_month].sum(), prev_month_name)
        interest_curr_adj = normalize_interest_for_calendar(interest_income[curr_month].sum(), curr_month_name)
        interest_change = interest_curr_adj - interest_prev_adj

        if (cash_change > 0 and interest_change < 0) or (cash_change < 0 and interest_change > 0):
            flags.append({
                'Rule_ID': 'A4',
                'Priority': '🔴 Critical',
                'Issue': 'Cash movement disconnected from interest',
                'Accounts': '111/112 ↔ 515xxx',
                'Period': f"{prev_month} → {curr_month}",
                'Reason': f'Cash changed by {cash_change:,.0f} VND but day-adjusted interest changed oppositely by {interest_change:,.0f} VND',
                'Flag_Trigger': 'Cash↑ BUT Interest↓ OR Cash↓ BUT Interest↑'
            })
    return flags


def check_rule_A5(bs_df, pl_df):
    """A5 - Lease termination but broker asset not written off"""
    flags = []
    if bs_df is None or pl_df is None or bs_df.empty or pl_df.empty:
        return flags

    revenue_accounts = get_account_pattern_data(pl_df, '511xxx')
    broker_accounts = get_account_pattern_data(bs_df, '242xxx')
    selling_expense = get_account_pattern_data(pl_df, '641xxx')
    if revenue_accounts.empty or broker_accounts.empty:
        return flags

    month_cols = get_month_cols(pl_df)
    if len(month_cols) < 1:
        return flags

    for i, month in enumerate(month_cols):
        revenue_total = revenue_accounts[month].sum()

        broker_prev_month_idx = max(0, i-1)
        broker_change = broker_accounts[month].sum() - broker_accounts[month_cols[broker_prev_month_idx]].sum() if not broker_accounts.empty else 0
        selling_change = selling_expense[month].sum() - selling_expense[month_cols[broker_prev_month_idx]].sum() if not selling_expense.empty else 0

        if revenue_total <= 0 and broker_change == 0 and selling_change == 0:
            flags.append({
                'Rule_ID': 'A5',
                'Priority': '🔴 Critical',
                'Issue': 'Lease termination but broker asset not written off',
                'Accounts': '511xxx ↔ 242xxx ↔ 641xxx',
                'Period': month,
                'Reason': f'Revenue is {revenue_total:,.0f} VND but broker asset (242) and selling expense (641) unchanged',
                'Flag_Trigger': 'Revenue ≤ 0 BUT 242 unchanged AND 641 unchanged'
            })
    return flags


def check_rule_A7(bs_df, pl_df):
    """A7 - Asset disposal but accumulated depreciation not written off"""
    flags = []
    if bs_df is None or bs_df.empty:
        return flags

    ip_cost = bs_df[bs_df['Account Name'].astype(str).str.contains('cost|historical', na=False, regex=True, case=False)]
    ip_cost = ip_cost[ip_cost['Account Code'].astype(str).str.startswith('217')]
    ip_accum_dep = bs_df[bs_df['Account Name'].astype(str).str.contains('accum|depreciation', na=False, regex=True, case=False)]
    ip_accum_dep = ip_accum_dep[ip_accum_dep['Account Code'].astype(str).str.startswith('217')]

    if ip_cost.empty or ip_accum_dep.empty:
        return flags

    month_cols = get_month_cols(bs_df)
    if len(month_cols) < 2:
        return flags

    for i in range(1, len(month_cols)):
        prev_month = month_cols[i-1]
        curr_month = month_cols[i]

        cost_change = ip_cost[curr_month].sum() - ip_cost[prev_month].sum()
        accum_dep_change = ip_accum_dep[curr_month].sum() - ip_accum_dep[prev_month].sum()

        if cost_change < 0 and accum_dep_change >= 0:
            flags.append({
                'Rule_ID': 'A7',
                'Priority': '🔴 Critical',
                'Issue': 'Asset disposal but accumulated depreciation not written off',
                'Accounts': '217xxx (cost) ↔ 217xxx (accum dep)',
                'Period': f"{prev_month} → {curr_month}",
                'Reason': f'IP cost decreased by {abs(cost_change):,.0f} VND but accumulated depreciation changed by {accum_dep_change:,.0f} VND',
                'Flag_Trigger': 'IP cost↓ BUT Accumulated depreciation unchanged'
            })
    return flags


def check_rule_D1(bs_df, pl_df):
    """D1 - Balance sheet imbalance"""
    flags = []
    if bs_df is None or bs_df.empty:
        return flags

    month_cols = get_month_cols(bs_df)
    bs_df['Account_Line_Str'] = bs_df['Account Line'].astype(str)
    bs_df['First_Digit'] = bs_df['Account_Line_Str'].str[0]

    type_1 = bs_df[bs_df['First_Digit'] == '1']  # Assets
    type_2 = bs_df[bs_df['First_Digit'] == '2']  # Liabilities
    type_3 = bs_df[bs_df['First_Digit'] == '3']  # Liabilities (subtract)
    type_4 = bs_df[bs_df['First_Digit'] == '4']  # Equity (subtract)

    for month in month_cols:
        total_1 = type_1[month].sum() if not type_1.empty else 0
        total_2 = type_2[month].sum() if not type_2.empty else 0
        total_3 = type_3[month].sum() if not type_3.empty else 0
        total_4 = type_4[month].sum() if not type_4.empty else 0

        balance = total_1 + total_2 - total_3 - total_4

        if abs(balance) > 1:
            flags.append({
                'Rule_ID': 'D1',
                'Priority': '🔴 Critical',
                'Issue': 'Balance sheet imbalance',
                'Accounts': 'Account Lines: 1xx + 2xx - 3xx - 4xx',
                'Period': month,
                'Reason': f'Balance sheet formula (1+2-3-4) = {balance:,.0f} VND (should be 0)',
                'Flag_Trigger': '1 + 2 - 3 - 4 ≠ 0'
            })
    return flags


def check_rule_E1(bs_df, pl_df):
    """E1 - Negative asset balance"""
    flags = []
    if bs_df is None or bs_df.empty:
        return flags

    assets = bs_df[bs_df['Account Code'].astype(str).str.match(r'^1\d{2}', na=False)]
    month_cols = get_month_cols(bs_df)

    for month in month_cols:
        for _, row in assets.iterrows():
            value = row[month]
            if value < 0:
                flags.append({
                    'Rule_ID': 'E1',
                    'Priority': '🔴 Critical',
                    'Issue': 'Negative asset balance',
                    'Accounts': f"{row['Account Code']} - {row['Account Name']}",
                    'Period': month,
                    'Reason': f'Asset account {row["Account Code"]} has negative balance: {value:,.0f} VND',
                    'Flag_Trigger': 'Any asset < 0'
                })
    return flags


# ============================================================================
# RULE IMPLEMENTATIONS - REVIEW (🟡)
# ============================================================================

def check_rule_B1(bs_df, pl_df):
    """B1 - Rental revenue volatility"""
    flags = []
    if pl_df is None or pl_df.empty:
        return flags

    rental_revenue = pl_df[pl_df['Account Code'].astype(str).str.contains('511710001', na=False)]
    if rental_revenue.empty:
        return flags

    month_cols = get_month_cols(pl_df)
    if len(month_cols) < MONTHS_6M_AVG:
        return flags

    for _, row in rental_revenue.iterrows():
        values = [row[month] for month in month_cols]
        series = pd.Series(values, index=month_cols)

        for i in range(MONTHS_6M_AVG, len(month_cols)):
            curr_month = month_cols[i]
            curr_value = series.iloc[i]
            baseline_window = series.iloc[i-MONTHS_6M_AVG:i]
            mean_6m = baseline_window.mean()
            std_6m = baseline_window.std()
            threshold = std_6m * STD_DEV_THRESHOLD
            deviation = abs(curr_value - mean_6m)

            if deviation > threshold and threshold > 0:
                flags.append({
                    'Rule_ID': 'B1',
                    'Priority': '🟡 Review',
                    'Issue': 'Rental revenue volatility',
                    'Accounts': '511710001',
                    'Period': curr_month,
                    'Reason': f'Rental revenue {curr_value:,.0f} VND deviates by {deviation:,.0f} VND from 6M average {mean_6m:,.0f} VND',
                    'Flag_Trigger': 'abs(Current - Avg) > 2σ'
                })
    return flags


def check_rule_B2(bs_df, pl_df):
    """B2 - Depreciation changes without asset movement"""
    flags = []
    if bs_df is None or pl_df is None or bs_df.empty or pl_df.empty:
        return flags

    depreciation = pl_df[pl_df['Account Code'].astype(str).str.contains('632100002', na=False)]
    ip_accounts = get_account_pattern_data(bs_df, '217xxx')
    if depreciation.empty or ip_accounts.empty:
        return flags

    month_cols = get_month_cols(pl_df)
    if len(month_cols) < MONTHS_6M_AVG + 1:
        return flags

    dep_values = [depreciation[month].sum() for month in month_cols]
    dep_series = pd.Series(dep_values, index=month_cols)
    ip_values = [ip_accounts[month].sum() for month in month_cols]
    ip_series = pd.Series(ip_values, index=month_cols)

    for i in range(MONTHS_6M_AVG, len(month_cols)):
        curr_month = month_cols[i]
        curr_dep = dep_series.iloc[i]
        baseline_window = dep_series.iloc[i-MONTHS_6M_AVG:i]
        mean_dep = baseline_window.mean()
        std_dep = baseline_window.std()
        threshold = std_dep * STD_DEV_THRESHOLD
        deviation = abs(curr_dep - mean_dep)
        ip_change = ip_series.iloc[i] - ip_series.iloc[i-1]

        if deviation > threshold and threshold > 0 and ip_change <= 0:
            flags.append({
                'Rule_ID': 'B2',
                'Priority': '🟡 Review',
                'Issue': 'Depreciation changes without asset movement',
                'Accounts': '632100002 + 217xxx',
                'Period': curr_month,
                'Reason': f'Depreciation {curr_dep:,.0f} VND deviates by {deviation:,.0f} VND but IP unchanged',
                'Flag_Trigger': 'Depreciation deviates > 2σ AND IP unchanged'
            })
    return flags


def check_rule_B3(bs_df, pl_df):
    """B3 - Amortization changes"""
    flags = []
    if pl_df is None or pl_df.empty:
        return flags

    amortization = pl_df[pl_df['Account Code'].astype(str).str.contains('632100001', na=False)]
    if amortization.empty:
        return flags

    month_cols = get_month_cols(pl_df)
    if len(month_cols) < MONTHS_6M_AVG:
        return flags

    for _, row in amortization.iterrows():
        values = [row[month] for month in month_cols]
        series = pd.Series(values, index=month_cols)

        for i in range(MONTHS_6M_AVG, len(month_cols)):
            curr_month = month_cols[i]
            curr_value = series.iloc[i]
            baseline_window = series.iloc[i-MONTHS_6M_AVG:i]
            mean = baseline_window.mean()
            std = baseline_window.std()
            threshold = std * STD_DEV_THRESHOLD
            deviation = abs(curr_value - mean)

            if deviation > threshold and threshold > 0:
                flags.append({
                    'Rule_ID': 'B3',
                    'Priority': '🟡 Review',
                    'Issue': 'Amortization changes',
                    'Accounts': '632100001',
                    'Period': curr_month,
                    'Reason': f'Amortization {curr_value:,.0f} VND deviates by {deviation:,.0f} VND from 6M average',
                    'Flag_Trigger': 'abs(Current - Avg) > 2σ'
                })
    return flags


def check_rule_C1(bs_df, pl_df):
    """C1 - Gross margin compression"""
    flags = []
    if pl_df is None or pl_df.empty:
        return flags

    revenue_accounts = get_account_pattern_data(pl_df, '511xxx')
    cogs_accounts = get_account_pattern_data(pl_df, '632xxx')
    if revenue_accounts.empty or cogs_accounts.empty:
        return flags

    month_cols = get_month_cols(pl_df)
    if len(month_cols) < MONTHS_6M_AVG:
        return flags

    gm_percentages = []
    for month in month_cols:
        revenue = revenue_accounts[month].sum()
        cogs = cogs_accounts[month].sum()
        gm_pct = ((revenue - cogs) / revenue * 100) if revenue != 0 else 0
        gm_percentages.append(gm_pct)

    gm_series = pd.Series(gm_percentages, index=month_cols)

    for i in range(MONTHS_6M_AVG, len(month_cols)):
        curr_month = month_cols[i]
        curr_gm = gm_series.iloc[i]
        baseline_window = gm_series.iloc[i-MONTHS_6M_AVG:i]
        mean_gm = baseline_window.mean()
        std_gm = baseline_window.std()
        threshold = std_gm * STD_DEV_THRESHOLD
        drop = mean_gm - curr_gm

        if drop > threshold and threshold > 0:
            flags.append({
                'Rule_ID': 'C1',
                'Priority': '🟡 Review',
                'Issue': 'Gross margin compression',
                'Accounts': 'Revenue (511) and COGS (632)',
                'Period': curr_month,
                'Reason': f'Gross margin dropped to {curr_gm:.2f}% from baseline {mean_gm:.2f}%',
                'Flag_Trigger': 'GM% drops > 2σ'
            })
    return flags


def check_rule_C2(bs_df, pl_df):
    """C2 - Unbilled reimbursable expenses"""
    flags = []
    if pl_df is None or pl_df.empty:
        return flags

    cogs_641 = get_account_pattern_data(pl_df, '641xxx')
    cogs_632 = get_account_pattern_data(pl_df, '632xxx')
    reimbursable_cogs = pd.concat([cogs_641, cogs_632], ignore_index=True)
    revenue_accounts = get_account_pattern_data(pl_df, '511xxx')
    if reimbursable_cogs.empty or revenue_accounts.empty:
        return flags

    month_cols = get_month_cols(pl_df)
    if len(month_cols) < 2:
        return flags

    for i in range(1, len(month_cols)):
        prev_month = month_cols[i-1]
        curr_month = month_cols[i]

        cogs_change = reimbursable_cogs[curr_month].sum() - reimbursable_cogs[prev_month].sum()
        revenue_change = revenue_accounts[curr_month].sum() - revenue_accounts[prev_month].sum()

        if cogs_change > 0 and revenue_change <= 0:
            flags.append({
                'Rule_ID': 'C2',
                'Priority': '🟡 Review',
                'Issue': 'Unbilled reimbursable expenses',
                'Accounts': '641xxx/632xxx ↔ 511xxx',
                'Period': f"{prev_month} → {curr_month}",
                'Reason': f'Reimbursable COGS increased by {cogs_change:,.0f} VND but revenue changed by {revenue_change:,.0f} VND',
                'Flag_Trigger': 'Reimbursable COGS↑ BUT Revenue unchanged'
            })
    return flags


def check_rule_D2(bs_df, pl_df):
    """D2 - Retained earnings reconciliation break"""
    flags = []
    if bs_df is None or pl_df is None or bs_df.empty or pl_df.empty:
        return flags

    re_421a = bs_df[bs_df['Account Line'].astype(str) == '421']
    re_421b = bs_df[bs_df['Account Line'].astype(str) == '4211']

    if re_421a.empty:
        return flags

    bs_month_cols = get_month_cols(bs_df)
    pl_month_cols = get_month_cols(pl_df)

    if len(bs_month_cols) < 2:
        return flags

    pl_components = {}
    for line in ['1', '11', '21', '22', '23', '25', '26', '31', '32', '51', '52']:
        pl_components[line] = pl_df[pl_df['Account Line'].astype(str) == line]

    for i in range(1, len(bs_month_cols)):
        prev_month = bs_month_cols[i-1]
        curr_month = bs_month_cols[i]

        opening_re_421a = re_421a[prev_month].sum() if not re_421a.empty else 0
        opening_re_421b = re_421b[prev_month].sum() if not re_421b.empty else 0
        opening_re = opening_re_421a + opening_re_421b

        closing_re_421a = re_421a[curr_month].sum() if not re_421a.empty else 0
        closing_re_421b = re_421b[curr_month].sum() if not re_421b.empty else 0
        closing_re = closing_re_421a + closing_re_421b

        if curr_month not in pl_month_cols:
            continue

        pl_1 = pl_components['1'][curr_month].sum() if not pl_components['1'].empty else 0
        pl_11 = pl_components['11'][curr_month].sum() if not pl_components['11'].empty else 0
        pl_21 = pl_components['21'][curr_month].sum() if not pl_components['21'].empty else 0
        pl_22 = pl_components['22'][curr_month].sum() if not pl_components['22'].empty else 0
        pl_23 = pl_components['23'][curr_month].sum() if not pl_components['23'].empty else 0
        pl_25 = pl_components['25'][curr_month].sum() if not pl_components['25'].empty else 0
        pl_26 = pl_components['26'][curr_month].sum() if not pl_components['26'].empty else 0
        pl_31 = pl_components['31'][curr_month].sum() if not pl_components['31'].empty else 0
        pl_32 = pl_components['32'][curr_month].sum() if not pl_components['32'].empty else 0
        pl_51 = pl_components['51'][curr_month].sum() if not pl_components['51'].empty else 0
        pl_52 = pl_components['52'][curr_month].sum() if not pl_components['52'].empty else 0

        profit_loss = pl_1 - pl_11 + pl_21 - pl_23 - pl_22 - pl_25 - pl_26 + pl_31 - pl_32 - pl_51 - pl_52
        expected_closing_re = opening_re + profit_loss
        difference = abs(closing_re - expected_closing_re)

        if difference > MATERIALITY_THRESHOLDS['retained_earnings_tolerance']:
            flags.append({
                'Rule_ID': 'D2',
                'Priority': '🟡 Review',
                'Issue': 'Retained earnings reconciliation break',
                'Accounts': 'Account Lines: 421a+421b (BS) and P/L components',
                'Period': f"{prev_month} → {curr_month}",
                'Reason': f'Closing RE {closing_re:,.0f} VND differs from expected {expected_closing_re:,.0f} VND by {difference:,.0f} VND',
                'Flag_Trigger': 'abs(Closing RE - (Opening RE + P/L)) > 100M'
            })
    return flags


def check_rule_E2(bs_df, pl_df):
    """E2 - CIP stagnation"""
    flags = []
    if bs_df is None or bs_df.empty:
        return flags

    cip_accounts = get_account_pattern_data(bs_df, '241xxx')
    ip_accounts = get_account_pattern_data(bs_df, '217xxx')
    if cip_accounts.empty or ip_accounts.empty:
        return flags

    month_cols = get_month_cols(bs_df)
    if len(month_cols) < 4:
        return flags

    for i in range(3, len(month_cols)):
        curr_month = month_cols[i]
        months_to_check = month_cols[i-3:i+1]
        cip_values = [cip_accounts[month].sum() for month in months_to_check]

        if len(set(cip_values)) == 1 and cip_values[0] > 0:
            ip_change = ip_accounts[months_to_check[-1]].sum() - ip_accounts[months_to_check[0]].sum()

            if ip_change <= 0:
                flags.append({
                    'Rule_ID': 'E2',
                    'Priority': '🟡 Review',
                    'Issue': 'CIP stagnation',
                    'Accounts': '241xxx (CIP) ↔ 217xxx (IP)',
                    'Period': f"{months_to_check[0]} → {curr_month}",
                    'Reason': f'CIP unchanged at {cip_values[0]:,.0f} VND for {len(months_to_check)} months',
                    'Flag_Trigger': 'CIP flat >3 months AND IP unchanged'
                })
                break
    return flags


def check_rule_E5(bs_df, pl_df):
    """E5 - Trade Receivables aging"""
    flags = []
    if bs_df is None or pl_df is None or bs_df.empty or pl_df.empty:
        return flags

    ar_accounts = get_account_pattern_data(bs_df, '131xxx')
    revenue_accounts = get_account_pattern_data(pl_df, '511xxx')
    if ar_accounts.empty or revenue_accounts.empty:
        return flags

    month_cols = get_month_cols(bs_df)
    if len(month_cols) < 4:
        return flags

    for i in range(3, len(month_cols)):
        start_month = month_cols[i-3]
        curr_month = month_cols[i]

        ar_start = ar_accounts[start_month].sum()
        ar_end = ar_accounts[curr_month].sum()
        ar_growth_pct = ((ar_end - ar_start) / ar_start * 100) if ar_start != 0 else 0

        revenue_start = revenue_accounts[start_month].sum()
        revenue_end = revenue_accounts[curr_month].sum()
        revenue_growth_pct = ((revenue_end - revenue_start) / revenue_start * 100) if revenue_start != 0 else 0

        if ar_growth_pct > (MATERIALITY_THRESHOLDS['ar_growth_threshold'] * 100) and revenue_growth_pct <= 0:
            flags.append({
                'Rule_ID': 'E5',
                'Priority': '🟡 Review',
                'Issue': 'Trade Receivables aging',
                'Accounts': '131xxx (AR) ↔ 511xxx (Revenue)',
                'Period': f"{start_month} → {curr_month}",
                'Reason': f'AR grew {ar_growth_pct:.1f}% over 3 months but revenue changed by {revenue_growth_pct:.1f}%',
                'Flag_Trigger': 'AR growth >> Revenue growth over 3 months'
            })
    return flags


def check_rule_E6(bs_df, pl_df):
    """E6 - Accounts Payable compression"""
    flags = []
    if bs_df is None or pl_df is None or bs_df.empty or pl_df.empty:
        return flags

    ap_accounts = get_account_pattern_data(bs_df, '331xxx')
    cogs_accounts = get_account_pattern_data(pl_df, '632xxx')
    if ap_accounts.empty or cogs_accounts.empty:
        return flags

    month_cols = get_month_cols(bs_df)
    if len(month_cols) < 2:
        return flags

    for i in range(1, len(month_cols)):
        prev_month = month_cols[i-1]
        curr_month = month_cols[i]

        ap_change = ap_accounts[curr_month].sum() - ap_accounts[prev_month].sum()
        cogs_change = cogs_accounts[curr_month].sum() - cogs_accounts[prev_month].sum()

        if ap_change < 0 and cogs_change >= 0:
            flags.append({
                'Rule_ID': 'E6',
                'Priority': '🟡 Review',
                'Issue': 'Accounts Payable compression',
                'Accounts': '331xxx (AP) ↔ 632xxx (COGS)',
                'Period': f"{prev_month} → {curr_month}",
                'Reason': f'AP decreased by {abs(ap_change):,.0f} VND but COGS changed by {cogs_change:,.0f} VND',
                'Flag_Trigger': 'AP↓ BUT COGS unchanged/increased'
            })
    return flags


# ============================================================================
# RULE IMPLEMENTATIONS - INFO (🟢)
# ============================================================================

def check_rule_E3(bs_df, pl_df):
    """E3 - Large VAT input"""
    flags = []
    if bs_df is None or bs_df.empty:
        return flags

    vat_accounts = get_account_pattern_data(bs_df, '133xxx')
    if vat_accounts.empty:
        return flags

    month_cols = get_month_cols(bs_df)

    for month in month_cols:
        vat_balance = vat_accounts[month].sum()

        if vat_balance > MATERIALITY_THRESHOLDS['vat_input_threshold']:
            flags.append({
                'Rule_ID': 'E3',
                'Priority': '🟢 Info',
                'Issue': 'Large VAT input - tax refund opportunity',
                'Accounts': '133xxx (VAT input)',
                'Period': month,
                'Reason': f'VAT input balance is {vat_balance:,.0f} VND (threshold: 10B VND)',
                'Flag_Trigger': 'VAT input > 10B'
            })
    return flags


def check_rule_E4(bs_df, pl_df):
    """E4 - Interest payable quarterly pattern"""
    flags = []
    if bs_df is None or bs_df.empty:
        return flags

    interest_payable = get_account_pattern_data(bs_df, '335xxx')
    if interest_payable.empty:
        return flags

    month_cols = get_month_cols(bs_df)
    if len(month_cols) < 2:
        return flags

    for i in range(1, len(month_cols)):
        prev_month = month_cols[i-1]
        curr_month = month_cols[i]
        curr_month_name = extract_month_from_column(curr_month)

        if curr_month_name in QUARTER_START_MONTHS:
            interest_change = interest_payable[curr_month].sum() - interest_payable[prev_month].sum()

            if interest_change >= 0:
                flags.append({
                    'Rule_ID': 'E4',
                    'Priority': '🟢 Info',
                    'Issue': 'Interest payable quarterly pattern validation',
                    'Accounts': '335xxx (Interest payable)',
                    'Period': f"{prev_month} → {curr_month}",
                    'Reason': f'Quarter-start month {curr_month_name} but interest payable changed by {interest_change:,.0f} VND',
                    'Flag_Trigger': 'Q-start month BUT Interest payable unchanged/increased'
                })
    return flags


def check_rule_F1(bs_df, pl_df):
    """F1 - Trade receivables quarterly billing"""
    flags = []
    if bs_df is None or bs_df.empty:
        return flags

    ar_accounts = bs_df[bs_df['Account Code'].astype(str).str.contains('131100001', na=False)]
    if ar_accounts.empty:
        return flags

    month_cols = get_month_cols(bs_df)
    if len(month_cols) < 2:
        return flags

    for i in range(1, len(month_cols)):
        prev_month = month_cols[i-1]
        curr_month = month_cols[i]
        curr_month_name = extract_month_from_column(curr_month)

        if curr_month_name in QUARTER_START_MONTHS:
            ar_change = ar_accounts[curr_month].sum() - ar_accounts[prev_month].sum()

            if ar_change > 0:
                flags.append({
                    'Rule_ID': 'F1',
                    'Priority': '🟢 Info',
                    'Issue': 'Trade receivables - quarterly billing (normal pattern)',
                    'Accounts': '131100001 (Trade receivables)',
                    'Period': f"{prev_month} → {curr_month}",
                    'Reason': f'AR increased by {ar_change:,.0f} VND at quarter-start month {curr_month_name}',
                    'Flag_Trigger': 'Document expected pattern'
                })
    return flags


def check_rule_F2(bs_df, pl_df):
    """F2 - Unbilled revenue quarter-end peak"""
    flags = []
    if bs_df is None or bs_df.empty:
        return flags

    unbilled_revenue = bs_df[bs_df['Account Code'].astype(str).str.contains('138900003', na=False)]
    if unbilled_revenue.empty:
        return flags

    month_cols = get_month_cols(bs_df)
    quarter_end_months = ['Mar', 'Jun', 'Sep', 'Dec']
    if len(month_cols) < 2:
        return flags

    for i in range(1, len(month_cols)):
        prev_month = month_cols[i-1]
        curr_month = month_cols[i]
        curr_month_name = extract_month_from_column(curr_month)

        if curr_month_name in quarter_end_months:
            unbilled_change = unbilled_revenue[curr_month].sum() - unbilled_revenue[prev_month].sum()

            if unbilled_change > 0:
                flags.append({
                    'Rule_ID': 'F2',
                    'Priority': '🟢 Info',
                    'Issue': 'Unbilled revenue - quarter-end peak (normal pattern)',
                    'Accounts': '138900003 (Unbilled revenue)',
                    'Period': f"{prev_month} → {curr_month}",
                    'Reason': f'Unbilled revenue increased by {unbilled_change:,.0f} VND at quarter-end month {curr_month_name}',
                    'Flag_Trigger': 'Document expected pattern'
                })
    return flags


# ============================================================================
# MAIN RULE EXECUTION
# ============================================================================

def run_all_variance_rules(bs_df, pl_df):
    """Run all 21 variance analysis rules"""
    all_flags = []

    # Critical Rules (🔴)
    all_flags.extend(check_rule_A1(bs_df, pl_df))
    all_flags.extend(check_rule_A2(bs_df, pl_df))
    all_flags.extend(check_rule_A3(bs_df, pl_df))
    all_flags.extend(check_rule_A4(bs_df, pl_df))
    all_flags.extend(check_rule_A5(bs_df, pl_df))
    all_flags.extend(check_rule_A7(bs_df, pl_df))
    all_flags.extend(check_rule_D1(bs_df, pl_df))
    all_flags.extend(check_rule_E1(bs_df, pl_df))

    # Review Rules (🟡)
    all_flags.extend(check_rule_B1(bs_df, pl_df))
    all_flags.extend(check_rule_B2(bs_df, pl_df))
    all_flags.extend(check_rule_B3(bs_df, pl_df))
    all_flags.extend(check_rule_C1(bs_df, pl_df))
    all_flags.extend(check_rule_C2(bs_df, pl_df))
    all_flags.extend(check_rule_D2(bs_df, pl_df))
    all_flags.extend(check_rule_E2(bs_df, pl_df))
    all_flags.extend(check_rule_E5(bs_df, pl_df))
    all_flags.extend(check_rule_E6(bs_df, pl_df))

    # Info Rules (🟢)
    all_flags.extend(check_rule_E3(bs_df, pl_df))
    all_flags.extend(check_rule_E4(bs_df, pl_df))
    all_flags.extend(check_rule_F1(bs_df, pl_df))
    all_flags.extend(check_rule_F2(bs_df, pl_df))

    return all_flags


# ============================================================================
# OUTPUT FUNCTIONS
# ============================================================================

def create_excel_output(all_file_data, combined_flags):
    """
    Create Excel output with raw data sheets and flagged variances.

    Args:
        all_file_data: List of dicts with keys: 'subsidiary', 'filename', 'bs_df', 'pl_df', 'flags'
        combined_flags: Combined list of all flags from all files

    Returns:
        bytes: Excel file
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Sheet 1: Variance Flags (with separators between files)
        if not combined_flags:
            flags_df = pd.DataFrame([{
                'File': 'N/A',
                'Rule_ID': 'N/A',
                'Priority': 'N/A',
                'Issue': 'No variance flags detected',
                'Accounts': 'N/A',
                'Period': 'N/A',
                'Reason': 'All checks passed',
                'Flag_Trigger': 'N/A'
            }])
        else:
            flags_df = pd.DataFrame(combined_flags)
            column_order = ['File', 'Rule_ID', 'Priority', 'Issue', 'Accounts', 'Period', 'Reason', 'Flag_Trigger']
            flags_df = flags_df[column_order]

            priority_order = {'🔴 Critical': 0, '🟡 Review': 1, '🟢 Info': 2}
            flags_df['Priority_Rank'] = flags_df['Priority'].map(priority_order)
            flags_df = flags_df.sort_values(['File', 'Priority_Rank', 'Rule_ID']).drop('Priority_Rank', axis=1)

            # Add empty rows between different files (if multiple files)
            if len(all_file_data) > 1:
                # Group by File and insert empty rows
                grouped_dfs = []
                for file_name in flags_df['File'].unique():
                    file_group = flags_df[flags_df['File'] == file_name].copy()
                    grouped_dfs.append(file_group)
                    # Add empty separator row (except after last group)
                    if file_name != flags_df['File'].unique()[-1]:
                        empty_row = pd.DataFrame([{col: '' for col in flags_df.columns}])
                        grouped_dfs.append(empty_row)

                flags_df = pd.concat(grouped_dfs, ignore_index=True)

        flags_df.to_excel(writer, sheet_name='Variance Flags', index=False)

        # Format Variance Flags sheet
        workbook = writer.book
        worksheet = writer.sheets['Variance Flags']

        critical_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        review_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        info_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        separator_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        current_file = None
        for idx, row in flags_df.iterrows():
            excel_row = idx + 2

            # Add separator row when file changes (for multiple files)
            if len(all_file_data) > 1 and current_file is not None and row.get('File') != current_file:
                # This row is a file boundary - make it a separator
                for cell in worksheet[excel_row]:
                    cell.fill = separator_fill

            current_file = row.get('File')
            priority = row.get('Priority', '')

            if '🔴' in priority:
                fill = critical_fill
            elif '🟡' in priority:
                fill = review_fill
            elif '🟢' in priority:
                fill = info_fill
            else:
                continue

            for cell in worksheet[excel_row]:
                cell.fill = fill
                cell.alignment = Alignment(vertical='top', wrap_text=True)

        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 80)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Add raw data sheets for each file
        for file_data in all_file_data:
            subsidiary = file_data['subsidiary']
            filename = file_data['filename']
            bs_df = file_data['bs_df']
            pl_df = file_data['pl_df']

            # Create sheet names (truncate if needed to fit Excel limits)
            bs_sheet_name = f"{subsidiary}_BS"[:31]
            pl_sheet_name = f"{subsidiary}_PL"[:31]

            # Make sheet names unique if there are duplicates
            sheet_counter = 1
            original_bs_name = bs_sheet_name
            original_pl_name = pl_sheet_name

            while bs_sheet_name in writer.sheets:
                bs_sheet_name = f"{original_bs_name[:28]}_{sheet_counter}"
                sheet_counter += 1

            sheet_counter = 1
            while pl_sheet_name in writer.sheets:
                pl_sheet_name = f"{original_pl_name[:28]}_{sheet_counter}"
                sheet_counter += 1

            # Write BS data
            if not bs_df.empty:
                bs_df.to_excel(writer, sheet_name=bs_sheet_name, index=False)
                bs_ws = writer.sheets[bs_sheet_name]

                # Format header
                for cell in bs_ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                # Auto-adjust column widths
                for column in bs_ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    bs_ws.column_dimensions[column_letter].width = adjusted_width

            # Write PL data
            if not pl_df.empty:
                pl_df.to_excel(writer, sheet_name=pl_sheet_name, index=False)
                pl_ws = writer.sheets[pl_sheet_name]

                # Format header
                for cell in pl_ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                # Auto-adjust column widths
                for column in pl_ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    pl_ws.column_dimensions[column_letter].width = adjusted_width

    return output.getvalue()


# ============================================================================
# MAIN PROCESSING FUNCTION FOR BACKEND
# ============================================================================

def process_variance_analysis(files: List[Tuple[str, bytes]]) -> bytes:
    """
    Main function to process variance analysis for backend integration.
    Supports multiple files with separators and raw data sheets.

    Args:
        files: List of tuples (filename, file_bytes)

    Returns:
        bytes: Excel file with variance flags and raw data
    """
    if not files:
        raise ValueError("No files provided for analysis")

    all_file_data = []
    combined_flags = []

    # Process each file
    for filename, file_bytes in files:
        try:
            # Convert bytes to file-like object
            file_obj = io.BytesIO(file_bytes)

            # Extract subsidiary name
            subsidiary = extract_subsidiary_name(file_obj)

            # Process BS and PL sheets
            file_obj.seek(0)  # Reset file pointer
            bs_df, _ = process_financial_tab(file_obj, "BS Breakdown", "BS", subsidiary)

            file_obj.seek(0)  # Reset file pointer
            pl_df, _ = process_financial_tab(file_obj, "PL Breakdown", "PL", subsidiary)

            if bs_df.empty and pl_df.empty:
                print(f"Warning: No data could be extracted from {filename}")
                continue

            # Run variance rules
            flags = run_all_variance_rules(bs_df, pl_df)

            # Add filename to each flag for identification
            for flag in flags:
                flag['File'] = filename

            # Store file data
            all_file_data.append({
                'subsidiary': subsidiary,
                'filename': filename,
                'bs_df': bs_df,
                'pl_df': pl_df,
                'flags': flags
            })

            # Add to combined flags
            combined_flags.extend(flags)

        except Exception as e:
            print(f"Error processing file {filename}: {e}")
            # Continue with next file instead of failing completely
            continue

    if not all_file_data:
        raise ValueError("No data could be extracted from any of the provided files")

    # Create output Excel with all data
    xlsx_bytes = create_excel_output(all_file_data, combined_flags)

    return xlsx_bytes
