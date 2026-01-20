"""
NTM Variance Calculator Service
Calculates variances between periods and generates output Excel files.
"""
import pandas as pd
from pathlib import Path
from typing import Dict, Any, List, Optional
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from app.shared.utils.logging_config import get_logger
from ..models.ntm_ebitda_models import (
    ProjectNTMSummary,
    NTMVarianceResult,
    NTMAnalysisSummary,
    LeaseChange,
    LeaseRecord,
    ChangeType,
    AnalysisConfig,
)

logger = get_logger(__name__)


class NTMVarianceCalculator:
    """
    Calculator for NTM EBITDA variance analysis.
    Compares NTM data between two periods and generates reports.
    """

    # Excel styling
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    VARIANCE_POSITIVE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    VARIANCE_NEGATIVE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    SIGNIFICANT_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def __init__(self, config: Optional[AnalysisConfig] = None):
        """
        Initialize the variance calculator.

        Args:
            config: Optional analysis configuration
        """
        self.config = config or AnalysisConfig()

    def calculate_variance(
        self,
        previous: Dict[str, ProjectNTMSummary],
        current: Dict[str, ProjectNTMSummary]
    ) -> NTMAnalysisSummary:
        """
        Calculate variance between previous and current periods.

        Args:
            previous: Dict of project name -> ProjectNTMSummary for previous period
            current: Dict of project name -> ProjectNTMSummary for current period

        Returns:
            NTMAnalysisSummary with variance results
        """
        logger.info(f"Calculating variance: {len(previous)} prev projects, {len(current)} curr projects")

        results = []
        all_projects = set(previous.keys()) | set(current.keys())

        for project_name in sorted(all_projects):
            prev_data = previous.get(project_name)
            curr_data = current.get(project_name)

            result = NTMVarianceResult(
                project_name=project_name,
                project_code=prev_data.project_code if prev_data else (curr_data.project_code if curr_data else ""),
                stake=prev_data.stake if prev_data else (curr_data.stake if curr_data else 1.0),
            )

            # Set previous values
            if prev_data:
                result.revenue_previous = prev_data.revenue_ntm
                result.opex_previous = prev_data.opex_ntm
                result.sga_previous = prev_data.sga_ntm
                result.ebitda_previous = prev_data.ebitda_ntm
                result.leases_previous = prev_data.leases

            # Set current values
            if curr_data:
                result.revenue_current = curr_data.revenue_ntm
                result.opex_current = curr_data.opex_ntm
                result.sga_current = curr_data.sga_ntm
                result.ebitda_current = curr_data.ebitda_ntm
                result.leases_current = curr_data.leases

            # Calculate variances
            result.calculate_variances()

            # Detect lease-level changes
            if prev_data and curr_data:
                result.lease_changes = self._detect_lease_changes(
                    prev_data.leases,
                    curr_data.leases
                )

            results.append(result)

        # Create summary
        summary = NTMAnalysisSummary(
            results=results,
            fx_rate=self.config.fx_rate,
            variance_threshold=self.config.variance_threshold,
        )
        summary.calculate_totals()

        logger.info(f"Variance calculation complete: {len(results)} projects, {len(summary.significant_results)} significant")

        return summary

    def _detect_lease_changes(
        self,
        previous_leases: List[LeaseRecord],
        current_leases: List[LeaseRecord]
    ) -> List[LeaseChange]:
        """
        Detect changes between previous and current lease lists.

        Args:
            previous_leases: List of leases from previous period
            current_leases: List of leases from current period

        Returns:
            List of LeaseChange objects describing the changes
        """
        changes = []

        # Build tenant dictionaries
        prev_by_tenant = {l.tenant_name: l for l in previous_leases if l.tenant_name}
        curr_by_tenant = {l.tenant_name: l for l in current_leases if l.tenant_name}

        all_tenants = set(prev_by_tenant.keys()) | set(curr_by_tenant.keys())

        for tenant in all_tenants:
            prev_lease = prev_by_tenant.get(tenant)
            curr_lease = curr_by_tenant.get(tenant)

            if prev_lease and not curr_lease:
                # Terminated
                changes.append(LeaseChange(
                    tenant_name=tenant,
                    change_type=ChangeType.TERMINATED,
                    gla_sqm=prev_lease.gla_sqm,
                    previous_ntm=prev_lease.total_ntm,
                    current_ntm=0,
                    variance=-prev_lease.total_ntm,
                    lease_start=prev_lease.lease_start_date,
                    lease_end=prev_lease.lease_end_date,
                    term_months=prev_lease.term_months,
                    description=f"Terminated: {tenant} ({prev_lease.gla_sqm:,.0f} sqm)"
                ))

            elif curr_lease and not prev_lease:
                # New
                changes.append(LeaseChange(
                    tenant_name=tenant,
                    change_type=ChangeType.NEW,
                    gla_sqm=curr_lease.gla_sqm,
                    previous_ntm=0,
                    current_ntm=curr_lease.total_ntm,
                    variance=curr_lease.total_ntm,
                    lease_start=curr_lease.lease_start_date,
                    lease_end=curr_lease.lease_end_date,
                    term_months=curr_lease.term_months,
                    description=f"New: {tenant} ({curr_lease.gla_sqm:,.0f} sqm, {curr_lease.term_months}M)"
                ))

            elif prev_lease and curr_lease:
                variance = curr_lease.total_ntm - prev_lease.total_ntm

                # Determine change type based on GLA and NTM changes
                if abs(variance) < self.config.min_lease_ntm_for_comment * 1e9:
                    # Minor change - skip
                    continue

                if curr_lease.gla_sqm > prev_lease.gla_sqm * 1.1:
                    change_type = ChangeType.EXPANDED
                    desc = f"Expanded: {tenant} ({prev_lease.gla_sqm:,.0f} -> {curr_lease.gla_sqm:,.0f} sqm)"
                elif curr_lease.gla_sqm < prev_lease.gla_sqm * 0.9:
                    change_type = ChangeType.REDUCED
                    desc = f"Reduced: {tenant} ({prev_lease.gla_sqm:,.0f} -> {curr_lease.gla_sqm:,.0f} sqm)"
                elif curr_lease.lease_end_date != prev_lease.lease_end_date:
                    change_type = ChangeType.RENEWED
                    desc = f"Renewed: {tenant}"
                else:
                    change_type = ChangeType.TIMING_SHIFT
                    desc = f"Timing shift: {tenant}"

                changes.append(LeaseChange(
                    tenant_name=tenant,
                    change_type=change_type,
                    gla_sqm=curr_lease.gla_sqm,
                    previous_ntm=prev_lease.total_ntm,
                    current_ntm=curr_lease.total_ntm,
                    variance=variance,
                    lease_start=curr_lease.lease_start_date,
                    lease_end=curr_lease.lease_end_date,
                    term_months=curr_lease.term_months,
                    description=desc
                ))

        # Sort by absolute variance (most impactful first)
        changes.sort(key=lambda c: abs(c.variance), reverse=True)

        return changes

    def generate_output_excel(
        self,
        summary: NTMAnalysisSummary,
        output_path: str,
        previous_period: str = "Previous",
        current_period: str = "Current"
    ) -> str:
        """
        Generate output Excel file with variance analysis.

        Args:
            summary: NTMAnalysisSummary with variance results
            output_path: Path for output file
            previous_period: Label for previous period
            current_period: Label for current period

        Returns:
            Path to generated file
        """
        logger.info(f"Generating output Excel: {output_path}")

        wb = Workbook()
        ws = wb.active
        ws.title = f"NTM EBITDA {current_period}"

        # Define columns matching BW standard format
        headers = [
            "Project Name",
            "Stake",
            f"Revenue NTM ($mn) {previous_period}",
            f"Revenue NTM ($mn) {current_period}",
            "Revenue Var",
            "Revenue Var %",
            f"OPEX NTM {previous_period}",
            f"OPEX NTM {current_period}",
            "OPEX Var",
            f"SG&A NTM {previous_period}",
            f"SG&A NTM {current_period}",
            "SG&A Var",
            f"EBITDA NTM {previous_period}",
            f"EBITDA NTM {current_period}",
            "EBITDA Var",
            "EBITDA Var %",
            "Commentary",
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = self.BORDER

        # Write data
        for row_idx, result in enumerate(summary.results, 2):
            # Convert to USD millions
            fx = self.config.fx_rate
            revenue_prev_usd = result.revenue_previous / fx / 1e6
            revenue_curr_usd = result.revenue_current / fx / 1e6
            revenue_var_usd = result.revenue_variance / fx / 1e6

            opex_prev_usd = result.opex_previous / fx / 1e6
            opex_curr_usd = result.opex_current / fx / 1e6
            opex_var_usd = result.opex_variance / fx / 1e6

            sga_prev_usd = result.sga_previous / fx / 1e6
            sga_curr_usd = result.sga_current / fx / 1e6
            sga_var_usd = result.sga_variance / fx / 1e6

            ebitda_prev_usd = result.ebitda_previous / fx / 1e6
            ebitda_curr_usd = result.ebitda_current / fx / 1e6
            ebitda_var_usd = result.ebitda_variance / fx / 1e6

            row_data = [
                result.project_name,
                result.stake,
                revenue_prev_usd,
                revenue_curr_usd,
                revenue_var_usd,
                result.revenue_variance_pct,
                opex_prev_usd,
                opex_curr_usd,
                opex_var_usd,
                sga_prev_usd,
                sga_curr_usd,
                sga_var_usd,
                ebitda_prev_usd,
                ebitda_curr_usd,
                ebitda_var_usd,
                result.ebitda_variance_pct,
                result.commentary,
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = self.BORDER

                # Apply number formatting
                if col in [3, 4, 5, 7, 8, 9, 10, 11, 12, 13, 14, 15]:
                    cell.number_format = '#,##0.00'
                elif col in [6, 16]:
                    cell.number_format = '0.0%'

                # Highlight significant variances
                if col == 16 and result.is_significant(self.config.variance_threshold):
                    cell.fill = self.SIGNIFICANT_FILL

                # Color variance columns
                if col in [5, 15]:  # Revenue Var, EBITDA Var
                    if isinstance(value, (int, float)):
                        if value > 0:
                            cell.fill = self.VARIANCE_POSITIVE_FILL
                        elif value < 0:
                            cell.fill = self.VARIANCE_NEGATIVE_FILL

        # Add totals row
        total_row = len(summary.results) + 2
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)

        fx = self.config.fx_rate
        totals = [
            summary.total_revenue_previous / fx / 1e6,
            summary.total_revenue_current / fx / 1e6,
            summary.total_revenue_variance / fx / 1e6,
            summary.total_revenue_variance / summary.total_revenue_previous if summary.total_revenue_previous else 0,
            summary.total_opex_previous / fx / 1e6,
            summary.total_opex_current / fx / 1e6,
            summary.total_opex_variance / fx / 1e6,
            summary.total_sga_previous / fx / 1e6,
            summary.total_sga_current / fx / 1e6,
            summary.total_sga_variance / fx / 1e6,
            summary.total_ebitda_previous / fx / 1e6,
            summary.total_ebitda_current / fx / 1e6,
            summary.total_ebitda_variance / fx / 1e6,
            summary.total_ebitda_variance_pct,
        ]

        for col, value in enumerate(totals, 3):
            cell = ws.cell(row=total_row, column=col, value=value)
            cell.font = Font(bold=True)
            cell.border = self.BORDER
            if col in [3, 4, 5, 7, 8, 9, 10, 11, 12, 13, 14, 15]:
                cell.number_format = '#,##0.00'
            elif col in [6, 16]:
                cell.number_format = '0.0%'

        # Adjust column widths
        column_widths = [25, 8, 15, 15, 12, 10, 15, 15, 12, 15, 15, 12, 15, 15, 12, 10, 50]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

        # Save
        wb.save(output_path)
        logger.info(f"Output Excel saved: {output_path}")

        return output_path

    def generate_summary_statistics(self, summary: NTMAnalysisSummary) -> Dict[str, Any]:
        """
        Generate summary statistics for the analysis.

        Args:
            summary: NTMAnalysisSummary with variance results

        Returns:
            Dict with summary statistics
        """
        fx = self.config.fx_rate

        return {
            "total_projects": len(summary.results),
            "significant_variances": len(summary.significant_results),
            "projects_increased": len(summary.projects_with_increase),
            "projects_decreased": len(summary.projects_with_decrease),
            "portfolio_totals": {
                "revenue": {
                    "previous_usd_mn": summary.total_revenue_previous / fx / 1e6,
                    "current_usd_mn": summary.total_revenue_current / fx / 1e6,
                    "variance_usd_mn": summary.total_revenue_variance / fx / 1e6,
                },
                "ebitda": {
                    "previous_usd_mn": summary.total_ebitda_previous / fx / 1e6,
                    "current_usd_mn": summary.total_ebitda_current / fx / 1e6,
                    "variance_usd_mn": summary.total_ebitda_variance / fx / 1e6,
                    "variance_pct": summary.total_ebitda_variance_pct,
                },
            },
            "variance_threshold": summary.variance_threshold,
            "fx_rate": fx,
        }
