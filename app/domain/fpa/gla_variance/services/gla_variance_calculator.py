"""
GLA Variance Calculator Service
Calculates variances between previous and current GLA data and generates output.
"""
import pandas as pd
from typing import Dict, List, Tuple, Optional
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from app.shared.utils.logging_config import get_logger
from ..models.gla_models import (
    GLAVarianceResult,
    GLAAnalysisSummary,
    ProjectGLASummary,
    ProductType,
    TenantChange
)

logger = get_logger(__name__)


class GLAVarianceCalculator:
    """
    Calculates GLA variances and generates output reports.
    """

    def __init__(self):
        self.variance_threshold = 0.01  # Report variances > 0.01 sqm

    def _calculate_tenant_changes(
        self,
        previous_summary: Optional[ProjectGLASummary],
        current_summary: Optional[ProjectGLASummary]
    ) -> List[TenantChange]:
        """
        Calculate tenant-level changes between two periods.

        Returns:
            List of TenantChange objects describing what changed
        """
        changes = []

        # Build tenant dicts for comparison
        prev_tenants = {}
        curr_tenants = {}

        if previous_summary and previous_summary.tenants:
            for t in previous_summary.tenants:
                prev_tenants[t.tenant_name] = t.gla_sqm

        if current_summary and current_summary.tenants:
            for t in current_summary.tenants:
                curr_tenants[t.tenant_name] = t.gla_sqm

        # Find all unique tenants
        all_tenants = set(prev_tenants.keys()) | set(curr_tenants.keys())

        for tenant in all_tenants:
            prev_gla = prev_tenants.get(tenant, 0)
            curr_gla = curr_tenants.get(tenant, 0)
            variance = curr_gla - prev_gla

            if abs(variance) < self.variance_threshold:
                continue  # Skip unchanged

            # Determine change type
            if prev_gla == 0 and curr_gla > 0:
                change_type = "new"
            elif prev_gla > 0 and curr_gla == 0:
                change_type = "terminated"
            elif variance > 0:
                change_type = "expanded"
            else:
                change_type = "reduced"

            changes.append(TenantChange(
                tenant_name=tenant,
                previous_gla=prev_gla,
                current_gla=curr_gla,
                variance=variance,
                change_type=change_type
            ))

        # Sort by absolute variance (largest impact first)
        changes.sort(key=lambda x: abs(x.variance), reverse=True)

        return changes

    def calculate_variance(
        self,
        handover_previous: Dict[Tuple[str, str], ProjectGLASummary],
        handover_current: Dict[Tuple[str, str], ProjectGLASummary],
        committed_previous: Dict[Tuple[str, str], ProjectGLASummary],
        committed_current: Dict[Tuple[str, str], ProjectGLASummary]
    ) -> GLAAnalysisSummary:
        """
        Calculate variance between previous and current periods.

        Args:
            handover_previous: Handover GLA data for previous period
            handover_current: Handover GLA data for current period
            committed_previous: Committed GLA data for previous period
            committed_current: Committed GLA data for current period

        Returns:
            GLAAnalysisSummary with all variance results
        """
        logger.info("Calculating GLA variances...")

        # Get all unique project-type combinations
        all_keys = set()
        all_keys.update(handover_previous.keys())
        all_keys.update(handover_current.keys())
        all_keys.update(committed_previous.keys())
        all_keys.update(committed_current.keys())

        logger.info(f"Found {len(all_keys)} unique project-type combinations")

        results: List[GLAVarianceResult] = []

        for key in sorted(all_keys):
            project_name, product_type = key

            # Get values for each data type and period
            hp = handover_previous.get(key)
            hc = handover_current.get(key)
            cp = committed_previous.get(key)
            cc = committed_current.get(key)

            # Determine region (from any available source)
            region = ""
            for summary in [hp, hc, cp, cc]:
                if summary and summary.region:
                    region = summary.region
                    break

            # Create variance result
            result = GLAVarianceResult(
                project_name=project_name,
                product_type=product_type,
                region=region,
                handover_previous=hp.gla_sqm if hp else 0.0,
                handover_current=hc.gla_sqm if hc else 0.0,
                committed_previous=cp.gla_sqm if cp else 0.0,
                committed_current=cc.gla_sqm if cc else 0.0,
                # New attributes - Handover sheet
                monthly_gross_rent_previous=hp.monthly_gross_rent if hp else 0.0,
                monthly_gross_rent_current=hc.monthly_gross_rent if hc else 0.0,
                monthly_rate_previous=hp.monthly_rate if hp else 0.0,
                monthly_rate_current=hc.monthly_rate if hc else 0.0,
                # New attributes - Committed sheet
                months_to_expire_previous=cp.months_to_expire if cp else 0.0,
                months_to_expire_current=cc.months_to_expire if cc else 0.0,
                months_to_expire_x_gla_previous=cp.months_to_expire_x_committed_gla if cp else 0.0,
                months_to_expire_x_gla_current=cc.months_to_expire_x_committed_gla if cc else 0.0,
            )

            result.calculate_variances()

            # Calculate tenant-level changes
            result.handover_tenant_changes = self._calculate_tenant_changes(hp, hc)
            result.committed_tenant_changes = self._calculate_tenant_changes(cp, cc)

            # Generate notes for significant variances (placeholder - will be populated by AI)
            result.handover_note = self._generate_variance_note(
                result.handover_variance, 'handover'
            )
            result.committed_note = self._generate_variance_note(
                result.committed_variance, 'committed'
            )

            results.append(result)

        # Create summary and calculate totals
        summary = GLAAnalysisSummary(results=results)
        summary.calculate_totals()

        logger.info(f"Variance calculation complete. {len(results)} results generated.")
        return summary

    def _generate_variance_note(self, variance: float, data_type: str) -> str:
        """Generate explanatory note for variance (placeholder for manual input)"""
        if abs(variance) < self.variance_threshold:
            return ""
        # Return empty for now - notes will be added manually or by AI
        return ""

    def generate_output_excel(
        self,
        summary: GLAAnalysisSummary,
        output_path: str,
        previous_period: str = "Previous",
        current_period: str = "Current"
    ) -> str:
        """
        Generate Excel output file with variance analysis results.

        Args:
            summary: GLAAnalysisSummary with all results
            output_path: Path to save the output file
            previous_period: Label for previous period (e.g., "Oct 2025")
            current_period: Label for current period (e.g., "Nov 2025")

        Returns:
            Path to generated file
        """
        logger.info(f"Generating output Excel file: {output_path}")

        wb = Workbook()
        ws = wb.active
        ws.title = "GLA Variance Analysis"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        subheader_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        variance_positive_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        variance_negative_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        total_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Column headers - Order: Committed GLA -> WALE -> Handover GLA -> Gross Rent
        headers = [
            "Project Name",
            "Product Type",
            "Region",
            # Committed GLA (sqm)
            f"Committed GLA\n(sqm)\n{previous_period}",
            f"Committed GLA\n(sqm)\n{current_period}",
            "Var",
            "Note",
            # WALE (Weighted Average Lease Expiry)
            f"WALE\n{previous_period}",
            f"WALE\n{current_period}",
            "Var",
            "Note",
            # Handover GLA (sqm)
            f"Handover GLA\n(sqm)\n{previous_period}",
            f"Handover GLA\n(sqm)\n{current_period}",
            "Var",
            "Note",
            # Gross rent (sqm)
            f"Gross rent\n(sqm)\n{previous_period}",
            f"Gross rent\n(sqm)\n{current_period}",
            "Var",
            "Note",
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Write data rows
        row_num = 2
        for result in summary.results:
            values = [
                result.project_name,
                result.product_type,
                result.region,
                # Committed GLA (cols 4-7)
                result.committed_previous,
                result.committed_current,
                result.committed_variance,
                result.committed_note,
                # WALE (cols 8-11)
                result.months_to_expire_previous,
                result.months_to_expire_current,
                result.months_to_expire_variance,
                result.wale_note,
                # Handover GLA (cols 12-15)
                result.handover_previous,
                result.handover_current,
                result.handover_variance,
                result.handover_note,
                # Gross rent (cols 16-19)
                result.monthly_gross_rent_previous,
                result.monthly_gross_rent_current,
                result.monthly_gross_rent_variance,
                result.gross_rent_note,
            ]

            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = border

                # Format numbers - Committed GLA columns (4, 5)
                if col in [4, 5]:
                    cell.number_format = '#,##0.00'
                # Format numbers - WALE columns (8, 9)
                elif col in [8, 9]:
                    cell.number_format = '#,##0.00'
                # Format numbers - Handover GLA columns (12, 13)
                elif col in [12, 13]:
                    cell.number_format = '#,##0.00'
                # Format numbers - Gross rent columns (16, 17)
                elif col in [16, 17]:
                    cell.number_format = '#,##0.00'
                # Variance columns with color (6, 10, 14, 18)
                elif col in [6, 10, 14, 18]:
                    cell.number_format = '#,##0.00'
                    if isinstance(value, (int, float)):
                        if value > 0:
                            cell.fill = variance_positive_fill
                        elif value < 0:
                            cell.fill = variance_negative_fill

            row_num += 1

        # Add totals section
        row_num += 1  # Empty row
        num_cols = len(headers)

        # Total RBF
        total_rbf_row = row_num
        ws.cell(row=row_num, column=1, value="Total RBF").font = Font(bold=True)
        # Committed GLA totals (cols 4-6)
        ws.cell(row=row_num, column=4, value=summary.total_rbf_committed_previous).number_format = '#,##0.00'
        ws.cell(row=row_num, column=5, value=summary.total_rbf_committed_current).number_format = '#,##0.00'
        ws.cell(row=row_num, column=6, value=summary.total_rbf_committed_variance).number_format = '#,##0.00'
        # Handover GLA totals (cols 12-14)
        ws.cell(row=row_num, column=12, value=summary.total_rbf_handover_previous).number_format = '#,##0.00'
        ws.cell(row=row_num, column=13, value=summary.total_rbf_handover_current).number_format = '#,##0.00'
        ws.cell(row=row_num, column=14, value=summary.total_rbf_handover_variance).number_format = '#,##0.00'
        for col in range(1, num_cols + 1):
            ws.cell(row=row_num, column=col).fill = total_fill
            ws.cell(row=row_num, column=col).border = border

        row_num += 1

        # Total RBW
        ws.cell(row=row_num, column=1, value="Total RBW").font = Font(bold=True)
        # Committed GLA totals (cols 4-6)
        ws.cell(row=row_num, column=4, value=summary.total_rbw_committed_previous).number_format = '#,##0.00'
        ws.cell(row=row_num, column=5, value=summary.total_rbw_committed_current).number_format = '#,##0.00'
        ws.cell(row=row_num, column=6, value=summary.total_rbw_committed_variance).number_format = '#,##0.00'
        # Handover GLA totals (cols 12-14)
        ws.cell(row=row_num, column=12, value=summary.total_rbw_handover_previous).number_format = '#,##0.00'
        ws.cell(row=row_num, column=13, value=summary.total_rbw_handover_current).number_format = '#,##0.00'
        ws.cell(row=row_num, column=14, value=summary.total_rbw_handover_variance).number_format = '#,##0.00'
        for col in range(1, num_cols + 1):
            ws.cell(row=row_num, column=col).fill = total_fill
            ws.cell(row=row_num, column=col).border = border

        row_num += 1

        # Total Portfolio
        ws.cell(row=row_num, column=1, value="Total Portfolio").font = Font(bold=True)
        # Committed GLA totals (cols 4-6)
        ws.cell(row=row_num, column=4, value=summary.total_portfolio_committed_previous).number_format = '#,##0.00'
        ws.cell(row=row_num, column=5, value=summary.total_portfolio_committed_current).number_format = '#,##0.00'
        ws.cell(row=row_num, column=6, value=summary.total_portfolio_committed_variance).number_format = '#,##0.00'
        # Handover GLA totals (cols 12-14)
        ws.cell(row=row_num, column=12, value=summary.total_portfolio_handover_previous).number_format = '#,##0.00'
        ws.cell(row=row_num, column=13, value=summary.total_portfolio_handover_current).number_format = '#,##0.00'
        ws.cell(row=row_num, column=14, value=summary.total_portfolio_handover_variance).number_format = '#,##0.00'
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
            cell.border = border
            cell.font = Font(bold=True)

        # Adjust column widths - Order: Committed GLA -> WALE -> Handover GLA -> Gross Rent
        column_widths = [
            25,  # Project Name
            12,  # Product Type
            10,  # Region
            # Committed GLA (cols 4-7)
            15,  # Committed GLA Previous
            15,  # Committed GLA Current
            10,  # Var
            25,  # Note
            # WALE (cols 8-11)
            10,  # WALE Previous
            10,  # WALE Current
            10,  # Var
            25,  # Note
            # Handover GLA (cols 12-15)
            15,  # Handover GLA Previous
            15,  # Handover GLA Current
            10,  # Var
            25,  # Note
            # Gross rent (cols 16-19)
            15,  # Gross rent Previous
            15,  # Gross rent Current
            10,  # Var
            25,  # Note
        ]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

        # Set row height for header
        ws.row_dimensions[1].height = 40

        # Save workbook
        wb.save(output_path)
        logger.info(f"Output file saved: {output_path}")

        return output_path

    def generate_summary_statistics(self, summary: GLAAnalysisSummary) -> Dict:
        """
        Generate summary statistics for API response.
        """
        # Count variances
        positive_committed = sum(1 for r in summary.results if r.committed_variance > 0)
        negative_committed = sum(1 for r in summary.results if r.committed_variance < 0)
        unchanged_committed = sum(1 for r in summary.results if r.committed_variance == 0)

        positive_handover = sum(1 for r in summary.results if r.handover_variance > 0)
        negative_handover = sum(1 for r in summary.results if r.handover_variance < 0)
        unchanged_handover = sum(1 for r in summary.results if r.handover_variance == 0)

        return {
            "total_projects": len(summary.results),
            "committed": {
                "increased": positive_committed,
                "decreased": negative_committed,
                "unchanged": unchanged_committed,
                "total_variance": summary.total_portfolio_committed_variance,
                "total_previous": summary.total_portfolio_committed_previous,
                "total_current": summary.total_portfolio_committed_current
            },
            "handover": {
                "increased": positive_handover,
                "decreased": negative_handover,
                "unchanged": unchanged_handover,
                "total_variance": summary.total_portfolio_handover_variance,
                "total_previous": summary.total_portfolio_handover_previous,
                "total_current": summary.total_portfolio_handover_current
            },
            "by_type": {
                "rbf": {
                    "committed_variance": summary.total_rbf_committed_variance,
                    "handover_variance": summary.total_rbf_handover_variance
                },
                "rbw": {
                    "committed_variance": summary.total_rbw_committed_variance,
                    "handover_variance": summary.total_rbw_handover_variance
                }
            }
        }
