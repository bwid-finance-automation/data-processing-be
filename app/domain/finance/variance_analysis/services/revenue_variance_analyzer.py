"""
Revenue Variance Analyzer - AI-powered MoM variance analysis with UFL linkage.

This module implements comprehensive revenue variance analysis for BW Industrial,
comparing monthly revenue data with Unit for Lease (UFL) activity to explain
revenue movements and flag unusual items.

Uses AI (OpenAI/Claude) for:
- Dynamic data location and parsing
- Deep reasoning about variance drivers
- Executive narrative generation
- Recommendations

Output: 14-sheet Excel with Flags, Executive Summary, AI Insights, and 11 monthly comparison sheets.
"""

import io
import os
import re
import json
import logging
import xml.etree.ElementTree as ET
from typing import Dict, List, Tuple, Optional, Any
from dataclasses import dataclass, field
from datetime import datetime
from collections import defaultdict

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from openai import OpenAI
from anthropic import Anthropic

logger = logging.getLogger(__name__)

# =============================================================================
# AI SYSTEM PROMPTS FOR EACH ANALYSIS STEP
# =============================================================================

# Step 1: AI analyzes each monthly variance
AI_MONTHLY_ANALYSIS_PROMPT = """You are a senior financial analyst for BW Industrial Development JSC analyzing monthly revenue variance.

## Context
You are analyzing revenue changes from {prior_month} to {current_month}.

## Your Task
Analyze the revenue data and UFL (Unit for Lease) activity to:
1. Identify the TOP drivers of variance (both increases and decreases)
2. Match revenue changes to UFL activity (handovers = new leases, terminations = ended leases)
3. Flag items that need investigation (revenue changes without matching UFL records)
4. Provide a concise narrative explaining the month's variance

## Revenue Types
- **Leasing (511710001)**: Core rental income - changes indicate new leases or terminations
- **Service/Mgmt Fee (511600xxx)**: Management fees - typically follows leasing
- **Utilities (511800001)**: Pass-through utility charges
- **Others (511800002)**: Miscellaneous revenue

## UFL Matching Rules
- Revenue INCREASE should have UFL handover within 0-2 months prior
- Revenue DECREASE should have UFL termination within 0-1 month prior
- REVERSAL = prior month negative, current month positive (billing correction)
- NEGATIVE NO UFL = revenue went negative without termination record

## Output Format (JSON)
{{
    "period": "{current_month} vs {prior_month}",
    "narrative": "2-3 sentence summary explaining the variance drivers",
    "net_variance_explanation": "Why did total revenue change by X amount?",
    "top_increases": [
        {{
            "tenant": "tenant name",
            "subsidiary": "code",
            "variance_b": 1.5,
            "ufl_match": "NEW LEASE" or "RATE ADJUSTMENT" or "UNKNOWN",
            "ufl_date": "Nov 2025" or null,
            "gla": 5000 or null,
            "explanation": "Why this tenant's revenue increased"
        }}
    ],
    "top_decreases": [
        {{
            "tenant": "tenant name",
            "subsidiary": "code",
            "variance_b": -2.0,
            "ufl_match": "TERMINATED" or "PARTIAL TERMINATION" or "UNKNOWN",
            "termination_date": "Oct 2025" or null,
            "gla": 3000 or null,
            "explanation": "Why this tenant's revenue decreased"
        }}
    ],
    "flags": [
        {{
            "flag_type": "INCREASE NO UFL" or "DECREASE NO UFL" or "NEGATIVE NO UFL" or "REVERSAL",
            "tenant": "name",
            "subsidiary": "code",
            "variance_b": 0.5,
            "reason": "Why this is flagged and what to investigate"
        }}
    ],
    "stream_analysis": {{
        "leasing": {{"variance_b": 10.0, "driver": "explanation"}},
        "service_fee": {{"variance_b": 1.0, "driver": "explanation"}},
        "utilities": {{"variance_b": 0.5, "driver": "explanation"}},
        "others": {{"variance_b": 0.1, "driver": "explanation"}}
    }}
}}

IMPORTANT:
- Use specific tenant names and subsidiaries from the data
- Amounts in billions (B) VND
- Only flag material items (>100M VND = 0.1B variance)
- Be specific about UFL matching - did you find a matching handover/termination?
"""

# Step 2: AI analyzes flags across all months
AI_FLAG_ANALYSIS_PROMPT = """You are a senior financial analyst reviewing flagged items from revenue variance analysis.

## Context
These flags represent revenue changes that don't have matching UFL (Unit for Lease) records.
This could indicate:
- Missing UFL records in NetSuite
- Rate adjustments or expansions (not captured in UFL)
- Billing corrections or accruals
- Data quality issues

## Your Task
1. Prioritize flags by materiality and risk
2. Group similar flags to identify patterns
3. Recommend specific actions for each HIGH priority flag
4. Identify if certain subsidiaries or tenants have recurring issues

## Output Format (JSON)
{{
    "total_flags": 50,
    "by_type": {{
        "INCREASE NO UFL": {{"count": 20, "total_variance_b": 15.5}},
        "DECREASE NO UFL": {{"count": 15, "total_variance_b": -12.0}},
        "NEGATIVE NO UFL": {{"count": 10, "total_variance_b": -5.0}},
        "REVERSAL": {{"count": 5, "total_variance_b": 2.0}}
    }},
    "high_priority": [
        {{
            "tenant": "name",
            "subsidiary": "code",
            "flag_type": "type",
            "variance_b": 5.0,
            "periods_affected": ["Nov vs Oct", "Dec vs Nov"],
            "risk_assessment": "Why this is high priority",
            "recommended_action": "Specific action to take"
        }}
    ],
    "patterns_identified": [
        "Pattern 1: Description of recurring issue",
        "Pattern 2: Another pattern"
    ],
    "subsidiaries_with_issues": [
        {{"subsidiary": "BB6", "flag_count": 10, "concern": "What's the issue"}}
    ],
    "overall_data_quality": "Assessment of data completeness between revenue and UFL systems"
}}
"""

# Step 3: AI generates executive summary and recommendations
AI_EXECUTIVE_SUMMARY_PROMPT = """You are a CFO-level financial analyst preparing an executive briefing on revenue variance.

## Context
You have analyzed a full year of monthly revenue data with UFL linkage for BW Industrial Development JSC.

## Your Task
1. Summarize the year's revenue performance in 3-5 sentences
2. Identify the TOP 5 drivers of variance for the year
3. Assess how well revenue changes correlate with UFL activity
4. Provide actionable recommendations for management
5. Note any data quality concerns

## Output Format (JSON)
{{
    "executive_summary": "3-5 sentence overview for C-suite",
    "year_highlights": {{
        "total_revenue_b": 350.0,
        "yoy_change_pct": 15.0,
        "best_month": "November",
        "worst_month": "December",
        "key_trend": "Revenue growth driven by new handovers in H2"
    }},
    "top_5_drivers": [
        {{
            "rank": 1,
            "driver": "What drove the variance",
            "impact_b": 25.0,
            "tenants_involved": ["Tenant A", "Tenant B"],
            "explanation": "Detailed explanation"
        }}
    ],
    "ufl_correlation_score": "HIGH/MEDIUM/LOW",
    "ufl_correlation_explanation": "How well do revenue changes match UFL activity?",
    "data_quality_assessment": {{
        "score": "GOOD/FAIR/POOR",
        "issues": ["Issue 1", "Issue 2"],
        "impact": "How data quality affects analysis reliability"
    }},
    "recommendations": [
        {{
            "priority": "HIGH/MEDIUM/LOW",
            "recommendation": "Specific actionable recommendation",
            "rationale": "Why this matters",
            "owner": "Finance/Operations/IT"
        }}
    ],
    "risks_and_concerns": [
        "Risk 1: Description",
        "Risk 2: Description"
    ]
}}

IMPORTANT:
- This is for executive consumption - be clear and actionable
- Use specific numbers and percentages
- Prioritize what matters most
- Be honest about data quality issues
"""

# Revenue type classification based on account codes
REVENUE_TYPE_MAP = {
    "511710001": "Leasing",
    "511600001": "Service/Mgmt Fee",
    "511600002": "Service/Mgmt Fee",
    "511600003": "Service/Mgmt Fee",
    "511600004": "Service/Mgmt Fee",
    "511600005": "Service/Mgmt Fee",
    "511600006": "Service/Mgmt Fee",
    "511600007": "Service/Mgmt Fee",
    "511800001": "Utilities",
    "511800002": "Others",
}

# Variance threshold in VND (100 million = 0.1B)
VARIANCE_THRESHOLD = 100_000_000
# GLA threshold for UFL NO REVENUE flag (sqm)
GLA_THRESHOLD = 5000

MONTH_NAMES = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


@dataclass
class RevenueRecord:
    """Single revenue record from NetSuite export."""
    subsidiary: str = ""
    tenant: str = ""
    tenant_code: str = ""
    account_code: str = ""
    revenue_type: str = ""
    monthly_amounts: Dict[int, float] = field(default_factory=dict)  # month (1-12) -> amount
    ltm: float = 0.0


@dataclass
class UFLRecord:
    """Unit for Lease record from NetSuite export."""
    tenant: str = ""
    tenant_code: str = ""
    subsidiary: str = ""
    status: str = ""  # Handed Over, Terminated, Voided, Open
    start_date: Optional[datetime] = None
    termination_date: Optional[datetime] = None
    gla: float = 0.0
    unit_name: str = ""


@dataclass
class VarianceFlag:
    """A flagged item requiring investigation."""
    period: str  # e.g., "Dec vs Nov"
    flag_type: str  # INCREASE NO UFL, DECREASE NO UFL, NEGATIVE NO UFL, REVERSAL, UFL NO REVENUE
    subsidiary: str
    tenant: str
    prior_amount: float  # in billions
    current_amount: float  # in billions
    variance: float  # in billions
    detail: str


@dataclass
class MonthlyComparison:
    """Data for a single month-over-month comparison."""
    current_month: int
    prior_month: int
    current_month_name: str
    prior_month_name: str

    # Level 1: By Revenue Stream
    stream_breakdown: List[Dict[str, Any]] = field(default_factory=list)
    net_change: float = 0.0
    net_change_pct: float = 0.0

    # Level 2: Leasing drill-down
    leasing_variance: float = 0.0
    leasing_increases: List[Dict[str, Any]] = field(default_factory=list)
    leasing_decreases: List[Dict[str, Any]] = field(default_factory=list)
    total_increases: float = 0.0
    total_decreases: float = 0.0

    # Flags for this period
    flags: List[VarianceFlag] = field(default_factory=list)


class RevenueVarianceAnalyzer:
    """
    AI-powered revenue variance analyzer with UFL linkage.
    """

    def __init__(self, api_key: Optional[str] = None, model: Optional[str] = None):
        """
        Initialize analyzer with optional AI configuration.

        Args:
            api_key: OpenAI or Anthropic API key (defaults to env var)
            model: Model to use (e.g., "gpt-4o", "claude-3-5-sonnet-20241022")
        """
        self.ns = {'ss': 'urn:schemas-microsoft-com:office:spreadsheet'}
        self.revenue_records: List[RevenueRecord] = []
        self.ufl_records: List[UFLRecord] = []
        self.year = 2025  # Default year, will be detected from data

        # AI Configuration
        self.api_key = api_key or os.getenv("OPENAI_API_KEY") or os.getenv("ANTHROPIC_API_KEY")
        self.model = model or os.getenv("AI_MODEL", "gpt-4o")

        # Determine API type
        self.is_claude = "claude" in self.model.lower() if self.model else False
        self._total_usage = {"input_tokens": 0, "output_tokens": 0, "total_tokens": 0}

        # Initialize AI client
        self.ai_client = None
        if self.api_key:
            try:
                if self.is_claude:
                    self.ai_client = Anthropic(api_key=self.api_key)
                    logger.info(f"Initialized Claude client with model: {self.model}")
                else:
                    self.ai_client = OpenAI(api_key=self.api_key)
                    logger.info(f"Initialized OpenAI client with model: {self.model}")
            except Exception as e:
                logger.warning(f"Failed to initialize AI client: {e}")
                self.ai_client = None
        else:
            logger.warning("No API key found - AI analysis will be skipped")

    def get_and_reset_usage(self) -> Dict[str, Any]:
        """Return accumulated token usage and reset counters."""
        usage = dict(self._total_usage)
        usage["model"] = self.model
        usage["provider"] = "anthropic" if self.is_claude else "openai"
        self._total_usage = {"input_tokens": 0, "output_tokens": 0, "total_tokens": 0}
        return usage

    def analyze(self, revenue_bytes: bytes, ufl_bytes: bytes) -> bytes:
        """
        Perform full AI-powered revenue variance analysis.

        The entire analysis is driven by AI at each step:
        1. Parse data (Python - structured XML)
        2. For EACH month: AI analyzes variance, matches UFL, identifies flags
        3. AI analyzes all flags across months to identify patterns
        4. AI generates executive summary and recommendations

        Args:
            revenue_bytes: Raw bytes of RevenueBreakdown XML-Excel file
            ufl_bytes: Raw bytes of UnitForLeaseList XML-Excel file

        Returns:
            Excel file bytes with 14 sheets (including AI Insights)
        """
        logger.info("=" * 60)
        logger.info("STARTING AI-POWERED REVENUE VARIANCE ANALYSIS")
        logger.info("=" * 60)

        if not self.ai_client:
            logger.error("AI client not available - cannot proceed with AI analysis")
            raise ValueError("AI analysis requires API key. Set OPENAI_API_KEY or ANTHROPIC_API_KEY environment variable.")

        # Step 1: Parse revenue data (Python - structured XML parsing)
        logger.info("\nSTEP 1: Parsing Revenue Data...")
        self.revenue_records = self._parse_revenue_data(revenue_bytes)
        logger.info(f"   Parsed {len(self.revenue_records)} revenue records")

        # Step 2: Parse UFL data (Python - structured XML parsing)
        logger.info("\nSTEP 2: Parsing UFL Data...")
        self.ufl_records = self._parse_ufl_data(ufl_bytes)
        logger.info(f"   Parsed {len(self.ufl_records)} UFL records")

        # Step 3: Calculate monthly totals for context
        logger.info("\nSTEP 3: Calculating Monthly Totals...")
        monthly_totals = self._calculate_monthly_totals()
        logger.info(f"   Calculated totals for 12 months")

        # Step 4: AI ANALYZES EACH MONTHLY COMPARISON
        logger.info("\nSTEP 4: AI Analyzing Each Monthly Variance...")
        logger.info("   AI will analyze each month-over-month comparison with UFL linkage")
        monthly_ai_analyses = []
        all_ai_flags = []

        for month in range(12, 1, -1):  # Dec vs Nov, Nov vs Oct, ..., Feb vs Jan
            prior_month = month - 1
            current_name = MONTH_NAMES[month - 1]
            prior_name = MONTH_NAMES[prior_month - 1]

            logger.info(f"   AI analyzing {current_name} vs {prior_name}...")

            # Build context for this month's analysis
            monthly_context = self._build_monthly_context(month, prior_month, monthly_totals)

            # Call AI for this month's analysis
            ai_result = self._ai_analyze_monthly_variance(
                current_name, prior_name, monthly_context
            )
            monthly_ai_analyses.append(ai_result)

            # Collect flags from AI analysis
            if "flags" in ai_result:
                for flag in ai_result["flags"]:
                    all_ai_flags.append({
                        "period": f"{current_name} vs {prior_name}",
                        **flag
                    })

            logger.info(f"   {current_name} vs {prior_name}: {len(ai_result.get('flags', []))} flags identified")

        # Step 5: Check for UFL NO REVENUE flags (large handovers without revenue)
        logger.info("\nSTEP 5: Checking for UFL NO REVENUE flags...")
        ufl_no_revenue_flags = self._check_ufl_no_revenue()
        for flag in ufl_no_revenue_flags:
            all_ai_flags.append({
                "period": "Full Year",
                "flag_type": flag.flag_type,
                "tenant": flag.tenant,
                "subsidiary": flag.subsidiary,
                "variance_b": flag.variance,
                "reason": flag.detail
            })
        logger.info(f"   Found {len(ufl_no_revenue_flags)} UFL NO REVENUE flags")

        # Step 6: AI ANALYZES ALL FLAGS
        logger.info(f"\nSTEP 6: AI Analyzing All {len(all_ai_flags)} Flags...")
        flag_analysis = self._ai_analyze_all_flags(all_ai_flags)
        logger.info(f"   Flag analysis complete")

        # Step 7: AI GENERATES EXECUTIVE SUMMARY
        logger.info("\nSTEP 7: AI Generating Executive Summary...")
        executive_summary = self._ai_generate_executive_summary(
            monthly_totals, monthly_ai_analyses, flag_analysis
        )
        logger.info(f"   Executive summary generated")

        # Step 8: Build comparison objects from AI analyses
        logger.info("\nSTEP 8: Building Output Structures...")
        comparisons = self._build_comparisons_from_ai(monthly_ai_analyses)
        all_flags = self._build_flags_from_ai(all_ai_flags)

        # Step 9: Combine all AI insights
        ai_insights = {
            "executive_summary": executive_summary.get("executive_summary", ""),
            "year_highlights": executive_summary.get("year_highlights", {}),
            "key_drivers": executive_summary.get("top_5_drivers", []),
            "monthly_trends": executive_summary.get("year_highlights", {}).get("key_trend", ""),
            "flag_priorities": flag_analysis.get("high_priority", []),
            "ufl_correlation_analysis": executive_summary.get("ufl_correlation_explanation", ""),
            "data_quality_notes": executive_summary.get("data_quality_assessment", {}).get("issues", []),
            "recommendations": executive_summary.get("recommendations", []),
            "flag_analysis": flag_analysis,
            "monthly_analyses": monthly_ai_analyses
        }

        # Step 10: Generate Excel output with AI Insights
        logger.info("\nSTEP 9: Generating Excel Output...")
        xlsx_bytes = self._generate_excel(comparisons, all_flags, monthly_totals, ai_insights)

        logger.info("\n" + "=" * 60)
        logger.info("AI-POWERED REVENUE VARIANCE ANALYSIS COMPLETE")
        logger.info("=" * 60)
        return xlsx_bytes

    def _build_monthly_context(self, current_month: int, prior_month: int,
                               monthly_totals: Dict[int, float]) -> str:
        """Build context string for a single month's AI analysis."""
        current_name = MONTH_NAMES[current_month - 1]
        prior_name = MONTH_NAMES[prior_month - 1]

        lines = [
            f"# Revenue Variance Analysis: {current_name} vs {prior_name} {self.year}",
            "",
            "## Revenue Totals",
            f"- {prior_name} Total: {monthly_totals.get(prior_month, 0) / 1e9:.2f}B VND",
            f"- {current_name} Total: {monthly_totals.get(current_month, 0) / 1e9:.2f}B VND",
            f"- Net Change: {(monthly_totals.get(current_month, 0) - monthly_totals.get(prior_month, 0)) / 1e9:+.2f}B VND",
            "",
            "## Revenue by Tenant (Top 30 Changes)",
        ]

        # Get revenue changes by tenant for this month pair
        tenant_changes = self._get_tenant_changes(current_month, prior_month)
        tenant_changes.sort(key=lambda x: abs(x["variance"]), reverse=True)

        for tc in tenant_changes[:30]:
            lines.append(
                f"- {tc['subsidiary']} | {tc['tenant'][:40]} | "
                f"Type: {tc['revenue_type']} | "
                f"{prior_name}: {tc['prior']/1e9:.2f}B | "
                f"{current_name}: {tc['current']/1e9:.2f}B | "
                f"Δ: {tc['variance']/1e9:+.2f}B"
            )

        # Add UFL activity for matching
        lines.extend([
            "",
            f"## UFL Activity (Handovers & Terminations near {current_name})",
            "",
            "### Handovers (New Leases):",
        ])

        # Get handovers within 0-2 months of current month
        relevant_handovers = [
            u for u in self.ufl_records
            if u.status == "Handed Over" and u.start_date
            and u.start_date.year == self.year
            and u.start_date.month in [current_month, prior_month, max(1, prior_month - 1)]
        ]
        for u in relevant_handovers[:20]:
            lines.append(
                f"- {u.subsidiary} | {u.tenant_code} {u.tenant[:30]} | "
                f"Start: {u.start_date.strftime('%b %Y')} | GLA: {u.gla:,.0f} sqm"
            )

        lines.append("\n### Terminations:")
        relevant_terminations = [
            u for u in self.ufl_records
            if u.status == "Terminated" and u.termination_date
            and u.termination_date.year == self.year
            and u.termination_date.month in [current_month, prior_month]
        ]
        for u in relevant_terminations[:20]:
            lines.append(
                f"- {u.subsidiary} | {u.tenant_code} {u.tenant[:30]} | "
                f"Term: {u.termination_date.strftime('%b %Y')} | GLA: {u.gla:,.0f} sqm"
            )

        return "\n".join(lines)

    def _get_tenant_changes(self, current_month: int, prior_month: int) -> List[Dict]:
        """Get revenue changes by tenant between two months."""
        changes = []

        # Group revenue by subsidiary + tenant
        tenant_data = defaultdict(lambda: {"prior": 0, "current": 0, "type": ""})

        for r in self.revenue_records:
            key = (r.subsidiary, r.tenant, r.tenant_code)
            tenant_data[key]["prior"] += r.monthly_amounts.get(prior_month, 0)
            tenant_data[key]["current"] += r.monthly_amounts.get(current_month, 0)
            tenant_data[key]["type"] = r.revenue_type

        for (sub, tenant, code), data in tenant_data.items():
            variance = data["current"] - data["prior"]
            if abs(variance) > 0:  # Include all non-zero changes
                changes.append({
                    "subsidiary": sub,
                    "tenant": tenant,
                    "tenant_code": code,
                    "revenue_type": data["type"],
                    "prior": data["prior"],
                    "current": data["current"],
                    "variance": variance
                })

        return changes

    def _ai_analyze_monthly_variance(self, current_month: str, prior_month: str,
                                     context: str) -> Dict[str, Any]:
        """Call AI to analyze a single month's variance."""
        prompt = AI_MONTHLY_ANALYSIS_PROMPT.format(
            current_month=current_month,
            prior_month=prior_month
        )

        try:
            if self.is_claude:
                response = self.ai_client.messages.create(
                    model=self.model,
                    max_tokens=4000,
                    system=prompt,
                    messages=[{"role": "user", "content": context}]
                )
                result = response.content[0].text
                self._total_usage["input_tokens"] += getattr(response.usage, 'input_tokens', 0)
                self._total_usage["output_tokens"] += getattr(response.usage, 'output_tokens', 0)
            else:
                params = {
                    "model": self.model,
                    "messages": [
                        {"role": "system", "content": prompt},
                        {"role": "user", "content": context}
                    ],
                    "max_tokens": 4000,
                }
                if not self.model.startswith("gpt-5"):
                    params["temperature"] = 0.2
                response = self.ai_client.chat.completions.create(**params)
                result = response.choices[0].message.content
                self._total_usage["input_tokens"] += getattr(response.usage, 'prompt_tokens', 0)
                self._total_usage["output_tokens"] += getattr(response.usage, 'completion_tokens', 0)

            self._total_usage["total_tokens"] = self._total_usage["input_tokens"] + self._total_usage["output_tokens"]
            return self._parse_ai_response(result)
        except Exception as e:
            logger.error(f"AI monthly analysis failed for {current_month} vs {prior_month}: {e}")
            return {
                "period": f"{current_month} vs {prior_month}",
                "narrative": f"AI analysis error: {str(e)}",
                "flags": [],
                "top_increases": [],
                "top_decreases": [],
                "stream_analysis": {}
            }

    def _ai_analyze_all_flags(self, all_flags: List[Dict]) -> Dict[str, Any]:
        """Call AI to analyze all flags and identify patterns."""
        if not all_flags:
            return {"total_flags": 0, "high_priority": [], "patterns_identified": []}

        # Build flag context
        lines = [
            "# All Flagged Items for Analysis",
            f"Total Flags: {len(all_flags)}",
            "",
            "## Flag Details (sorted by variance magnitude):",
        ]

        sorted_flags = sorted(all_flags, key=lambda x: abs(x.get("variance_b", 0)), reverse=True)
        for f in sorted_flags[:50]:  # Top 50 flags
            lines.append(
                f"- [{f.get('flag_type', 'UNKNOWN')}] {f.get('period', '')} | "
                f"{f.get('subsidiary', '')} | {f.get('tenant', '')[:30]} | "
                f"Variance: {f.get('variance_b', 0):+.2f}B | "
                f"Reason: {f.get('reason', 'N/A')[:50]}"
            )

        context = "\n".join(lines)

        try:
            if self.is_claude:
                response = self.ai_client.messages.create(
                    model=self.model,
                    max_tokens=3000,
                    system=AI_FLAG_ANALYSIS_PROMPT,
                    messages=[{"role": "user", "content": context}]
                )
                result = response.content[0].text
                self._total_usage["input_tokens"] += getattr(response.usage, 'input_tokens', 0)
                self._total_usage["output_tokens"] += getattr(response.usage, 'output_tokens', 0)
            else:
                params = {
                    "model": self.model,
                    "messages": [
                        {"role": "system", "content": AI_FLAG_ANALYSIS_PROMPT},
                        {"role": "user", "content": context}
                    ],
                    "max_tokens": 3000,
                }
                if not self.model.startswith("gpt-5"):
                    params["temperature"] = 0.2
                response = self.ai_client.chat.completions.create(**params)
                result = response.choices[0].message.content
                self._total_usage["input_tokens"] += getattr(response.usage, 'prompt_tokens', 0)
                self._total_usage["output_tokens"] += getattr(response.usage, 'completion_tokens', 0)

            self._total_usage["total_tokens"] = self._total_usage["input_tokens"] + self._total_usage["output_tokens"]
            return self._parse_ai_response(result)
        except Exception as e:
            logger.error(f"AI flag analysis failed: {e}")
            return {
                "total_flags": len(all_flags),
                "high_priority": [],
                "patterns_identified": [f"Analysis error: {str(e)}"],
                "overall_data_quality": "Unable to assess"
            }

    def _ai_generate_executive_summary(self, monthly_totals: Dict[int, float],
                                       monthly_analyses: List[Dict],
                                       flag_analysis: Dict) -> Dict[str, Any]:
        """Call AI to generate executive summary."""
        # Build comprehensive context
        lines = [
            "# Full Year Revenue Variance Summary",
            f"Year: {self.year}",
            "",
            "## Monthly Revenue (VND Billions):",
        ]

        for month in range(1, 13):
            total = monthly_totals.get(month, 0) / 1e9
            lines.append(f"- {MONTH_NAMES[month-1]}: {total:.2f}B")

        ytd = sum(monthly_totals.values()) / 1e9
        lines.append(f"\nYTD Total: {ytd:.2f}B")

        # Add monthly analysis summaries
        lines.append("\n## Monthly Analysis Summaries:")
        for analysis in monthly_analyses:
            period = analysis.get("period", "Unknown")
            narrative = analysis.get("narrative", "No narrative")[:200]
            lines.append(f"\n### {period}")
            lines.append(narrative)

        # Add flag summary
        lines.append("\n## Flag Summary:")
        lines.append(f"Total Flags: {flag_analysis.get('total_flags', 0)}")
        lines.append(f"High Priority: {len(flag_analysis.get('high_priority', []))}")
        patterns = flag_analysis.get("patterns_identified", [])
        if patterns:
            lines.append("Patterns: " + "; ".join(patterns[:3]))

        context = "\n".join(lines)

        try:
            if self.is_claude:
                response = self.ai_client.messages.create(
                    model=self.model,
                    max_tokens=4000,
                    system=AI_EXECUTIVE_SUMMARY_PROMPT,
                    messages=[{"role": "user", "content": context}]
                )
                result = response.content[0].text
                self._total_usage["input_tokens"] += getattr(response.usage, 'input_tokens', 0)
                self._total_usage["output_tokens"] += getattr(response.usage, 'output_tokens', 0)
            else:
                params = {
                    "model": self.model,
                    "messages": [
                        {"role": "system", "content": AI_EXECUTIVE_SUMMARY_PROMPT},
                        {"role": "user", "content": context}
                    ],
                    "max_tokens": 4000,
                }
                if not self.model.startswith("gpt-5"):
                    params["temperature"] = 0.3
                response = self.ai_client.chat.completions.create(**params)
                result = response.choices[0].message.content
                self._total_usage["input_tokens"] += getattr(response.usage, 'prompt_tokens', 0)
                self._total_usage["output_tokens"] += getattr(response.usage, 'completion_tokens', 0)

            self._total_usage["total_tokens"] = self._total_usage["input_tokens"] + self._total_usage["output_tokens"]
            return self._parse_ai_response(result)
        except Exception as e:
            logger.error(f"AI executive summary failed: {e}")
            return {
                "executive_summary": f"AI analysis error: {str(e)}",
                "recommendations": []
            }

    def _build_comparisons_from_ai(self, monthly_analyses: List[Dict]) -> List[MonthlyComparison]:
        """Build MonthlyComparison objects combining AI analysis with actual parsed data."""
        comparisons = []

        for i, analysis in enumerate(monthly_analyses):
            current_month = 12 - i  # Dec, Nov, Oct, ...
            prior_month = current_month - 1

            comp = MonthlyComparison(
                current_month=current_month,
                prior_month=prior_month,
                current_month_name=MONTH_NAMES[current_month - 1],
                prior_month_name=MONTH_NAMES[prior_month - 1]
            )

            # Calculate ACTUAL stream breakdown from parsed revenue data
            stream_data = self._calculate_stream_breakdown(current_month, prior_month)
            comp.stream_breakdown = stream_data["breakdown"]
            comp.net_change = stream_data["net_change"]
            comp.net_change_pct = stream_data["net_change_pct"]

            # Calculate ACTUAL leasing drill-down with UFL linkage
            leasing_data = self._calculate_leasing_drilldown(current_month, prior_month)
            comp.leasing_variance = leasing_data["variance"]
            comp.leasing_increases = leasing_data["increases"]
            comp.leasing_decreases = leasing_data["decreases"]
            comp.total_increases = leasing_data["total_increases"]
            comp.total_decreases = leasing_data["total_decreases"]
            comp.flags = leasing_data["flags"]

            # Enrich with AI explanations if available
            ai_increases = {inc.get("tenant", "").upper(): inc for inc in analysis.get("top_increases", [])}
            ai_decreases = {dec.get("tenant", "").upper(): dec for dec in analysis.get("top_decreases", [])}

            for inc in comp.leasing_increases:
                tenant_key = inc.get("tenant", "").upper()
                if tenant_key in ai_increases:
                    ai_data = ai_increases[tenant_key]
                    inc["ai_explanation"] = ai_data.get("explanation", "")
                    # Use AI's UFL match if it found one and we didn't
                    if inc.get("ufl_action") == "-" and ai_data.get("ufl_match"):
                        inc["ufl_action"] = ai_data.get("ufl_match", "-")
                        inc["ufl_month"] = ai_data.get("ufl_date", "")

            for dec in comp.leasing_decreases:
                tenant_key = dec.get("tenant", "").upper()
                if tenant_key in ai_decreases:
                    ai_data = ai_decreases[tenant_key]
                    dec["ai_explanation"] = ai_data.get("explanation", "")
                    if dec.get("ufl_action") == "-" and ai_data.get("ufl_match"):
                        dec["ufl_action"] = ai_data.get("ufl_match", "-")
                        dec["term_month"] = ai_data.get("termination_date", "")

            comparisons.append(comp)

        return comparisons

    def _build_flags_from_ai(self, ai_flags: List[Dict]) -> List[VarianceFlag]:
        """Build VarianceFlag objects from AI flag results."""
        flags = []

        for f in ai_flags:
            flag = VarianceFlag(
                period=f.get("period", ""),
                flag_type=f.get("flag_type", "UNKNOWN"),
                subsidiary=f.get("subsidiary", ""),
                tenant=f.get("tenant", ""),
                prior_amount=0,
                current_amount=0,
                variance=f.get("variance_b", 0),
                detail=f.get("reason", "")
            )
            flags.append(flag)

        return flags

    # =============================================================================
    # AI RESPONSE PARSING
    # =============================================================================

    def _parse_ai_response(self, response: str) -> Dict[str, Any]:
        """Parse AI JSON response with robust error handling."""
        try:
            # Try to extract JSON from response
            json_match = response
            if "```json" in response:
                json_match = response.split("```json")[1].split("```")[0].strip()
            elif "```" in response:
                json_match = response.split("```")[1].split("```")[0].strip()

            # Clean common JSON issues from AI output
            # Remove trailing commas before } or ]
            import re
            json_match = re.sub(r',\s*}', '}', json_match)
            json_match = re.sub(r',\s*]', ']', json_match)
            # Remove any BOM or special characters
            json_match = json_match.strip().lstrip('\ufeff')

            return json.loads(json_match)
        except (json.JSONDecodeError, IndexError) as e:
            logger.warning(f"Failed to parse AI JSON response: {e}")
            # Try a more aggressive cleanup
            try:
                # Find JSON-like structure
                import re
                json_pattern = re.search(r'\{[\s\S]*\}', response)
                if json_pattern:
                    cleaned = json_pattern.group(0)
                    cleaned = re.sub(r',\s*}', '}', cleaned)
                    cleaned = re.sub(r',\s*]', ']', cleaned)
                    return json.loads(cleaned)
            except Exception:
                pass
            # Fallback: return raw response as summary
            return {
                "narrative": response[:1000] if len(response) > 1000 else response,
                "flags": [],
                "top_increases": [],
                "top_decreases": [],
                "stream_analysis": {}
            }

    # =============================================================================
    # DATA PARSING METHODS (structured XML parsing - not AI dependent)
    # =============================================================================

    def _parse_revenue_data(self, file_bytes: bytes) -> List[RevenueRecord]:
        """Parse RevenueBreakdown XML-Excel file."""
        records = []

        try:
            content = file_bytes.decode('utf-8')
            root = ET.fromstring(content)
        except ET.ParseError as e:
            logger.error(f"Failed to parse revenue XML: {e}")
            raise ValueError(f"Invalid XML format: {e}")

        worksheet = root.find('.//ss:Worksheet', self.ns)
        if worksheet is None:
            raise ValueError("No worksheet found in revenue file")

        table = worksheet.find('ss:Table', self.ns)
        if table is None:
            raise ValueError("No table found in revenue worksheet")

        rows = table.findall('ss:Row', self.ns)

        # Detect year from header
        self._detect_year(rows)

        # Find the data start row (after headers)
        data_start_idx = self._find_data_start(rows)

        current_subsidiary = ""
        current_account_code = ""

        for row in rows[data_start_idx:]:
            cells = row.findall('ss:Cell', self.ns)
            cell_values = self._extract_cell_values_with_index(cells)

            if not cell_values:
                continue

            first_val = cell_values.get(0, "").strip()

            # Check if this is an account code row (e.g., "511710001 - Revenue...")
            if first_val and re.match(r'^\d{9}\s*-', first_val):
                match = re.match(r'^(\d{9})', first_val)
                if match:
                    current_account_code = match.group(1)
                continue

            # Check if this is a Total row - skip
            if first_val.startswith("Total -"):
                continue

            # Check if this is a data row (empty first cell, subsidiary in col 1)
            if first_val == "" and cell_values.get(1, ""):
                subsidiary = cell_values.get(1, "").strip()
                if subsidiary and subsidiary != "- No Entity -":
                    current_subsidiary = subsidiary

                tenant = cell_values.get(4, "").strip()  # Entity name contains code like "C00000259 SPX"

                # Skip intercompany entries (IE/ prefix)
                if "IE/" in tenant:
                    continue

                # Extract tenant code from tenant name using regex
                # Format: "C00000259 SPX" or "S00001552 VIETMY"
                tenant_code = ""
                code_match = re.search(r'([CS]\d+)', tenant)
                if code_match:
                    tenant_code = code_match.group(1)

                # Get revenue type from account code
                revenue_type = REVENUE_TYPE_MAP.get(current_account_code, "Others")

                # Parse monthly amounts (columns 7-18 are Jan-Dec)
                monthly_amounts = {}
                for month in range(1, 13):
                    col_idx = 6 + month  # Column 7 is Jan (month 1)
                    val = cell_values.get(col_idx, "")
                    try:
                        amount = float(val) if val else 0.0
                        monthly_amounts[month] = amount
                    except (ValueError, TypeError):
                        monthly_amounts[month] = 0.0

                # Get LTM (last column before LTM is column 19, LTM is 20)
                ltm = 0.0
                try:
                    ltm_val = cell_values.get(19, "") or cell_values.get(20, "")
                    ltm = float(ltm_val) if ltm_val else 0.0
                except (ValueError, TypeError):
                    pass

                # Skip if no activity
                if ltm == 0 and all(v == 0 for v in monthly_amounts.values()):
                    continue

                record = RevenueRecord(
                    subsidiary=current_subsidiary,
                    tenant=tenant,
                    tenant_code=tenant_code,
                    account_code=current_account_code,
                    revenue_type=revenue_type,
                    monthly_amounts=monthly_amounts,
                    ltm=ltm
                )
                records.append(record)

        return records

    def _parse_ufl_data(self, file_bytes: bytes) -> List[UFLRecord]:
        """Parse UnitForLeaseList XML-Excel file."""
        records = []

        try:
            content = file_bytes.decode('utf-8')
            root = ET.fromstring(content)
        except ET.ParseError as e:
            logger.error(f"Failed to parse UFL XML: {e}")
            raise ValueError(f"Invalid XML format: {e}")

        worksheet = root.find('.//ss:Worksheet', self.ns)
        if worksheet is None:
            raise ValueError("No worksheet found in UFL file")

        table = worksheet.find('ss:Table', self.ns)
        if table is None:
            raise ValueError("No table found in UFL worksheet")

        rows = table.findall('ss:Row', self.ns)

        if len(rows) < 2:
            return records

        # First row is header
        header_row = rows[0]
        headers = self._extract_header_row(header_row)
        header_map = {h.lower(): i for i, h in enumerate(headers)}

        for row in rows[1:]:
            cells = row.findall('ss:Cell', self.ns)
            values = self._extract_cell_values_with_index(cells)

            # Get values by header position
            tenant = self._get_by_header(values, header_map, "tenant", "")
            status = self._get_by_header(values, header_map, "unit for lease status", "")
            subsidiary = self._get_by_header(values, header_map, "subsidiary", "")
            gla_str = self._get_by_header(values, header_map, "gla for lease", "0")
            start_date_str = self._get_by_header(values, header_map, "start leasing date", "")
            term_date_str = self._get_by_header(values, header_map, "termination date", "")
            unit_name = self._get_by_header(values, header_map, "unit", "")

            # Extract tenant code
            tenant_code = ""
            code_match = re.search(r'([CS]\d+)', tenant)
            if code_match:
                tenant_code = code_match.group(1)

            # Parse GLA
            try:
                gla = float(gla_str) if gla_str else 0.0
            except (ValueError, TypeError):
                gla = 0.0

            # Parse dates
            start_date = self._parse_date(start_date_str)
            term_date = self._parse_date(term_date_str)

            # Only keep relevant statuses
            if status in ["Handed Over", "Terminated"]:
                record = UFLRecord(
                    tenant=tenant,
                    tenant_code=tenant_code,
                    subsidiary=subsidiary,
                    status=status,
                    start_date=start_date,
                    termination_date=term_date,
                    gla=gla,
                    unit_name=unit_name
                )
                records.append(record)

        return records

    def _build_monthly_comparisons(self) -> List[MonthlyComparison]:
        """Build all 11 monthly comparisons (Dec vs Nov through Feb vs Jan)."""
        comparisons = []

        # Dec vs Nov, Nov vs Oct, ..., Feb vs Jan
        for current_month in range(12, 1, -1):  # 12 down to 2
            prior_month = current_month - 1

            comp = MonthlyComparison(
                current_month=current_month,
                prior_month=prior_month,
                current_month_name=MONTH_NAMES[current_month - 1],
                prior_month_name=MONTH_NAMES[prior_month - 1]
            )

            # Calculate Level 1: By Revenue Stream
            stream_data = self._calculate_stream_breakdown(current_month, prior_month)
            comp.stream_breakdown = stream_data["breakdown"]
            comp.net_change = stream_data["net_change"]
            comp.net_change_pct = stream_data["net_change_pct"]

            # Calculate Level 2: Leasing drill-down with UFL linkage
            leasing_data = self._calculate_leasing_drilldown(current_month, prior_month)
            comp.leasing_variance = leasing_data["variance"]
            comp.leasing_increases = leasing_data["increases"]
            comp.leasing_decreases = leasing_data["decreases"]
            comp.total_increases = leasing_data["total_increases"]
            comp.total_decreases = leasing_data["total_decreases"]
            comp.flags = leasing_data["flags"]

            comparisons.append(comp)

        return comparisons

    def _calculate_stream_breakdown(self, current_month: int, prior_month: int) -> Dict[str, Any]:
        """Calculate variance by revenue stream."""
        streams = ["Leasing", "Service/Mgmt Fee", "Utilities", "Others"]
        breakdown = []

        total_prior = 0.0
        total_current = 0.0

        for stream in streams:
            stream_records = [r for r in self.revenue_records if r.revenue_type == stream]

            prior_sum = sum(r.monthly_amounts.get(prior_month, 0) for r in stream_records)
            current_sum = sum(r.monthly_amounts.get(current_month, 0) for r in stream_records)
            variance = current_sum - prior_sum

            total_prior += prior_sum
            total_current += current_sum

            breakdown.append({
                "stream": stream,
                "prior": prior_sum,
                "current": current_sum,
                "variance": variance
            })

        net_change = total_current - total_prior
        net_change_pct = (net_change / abs(total_prior) * 100) if total_prior != 0 else 0

        # Calculate % of net for each stream
        for item in breakdown:
            if net_change != 0:
                item["pct_of_net"] = (item["variance"] / net_change * 100)
            else:
                item["pct_of_net"] = 0

        # Add total row
        breakdown.append({
            "stream": "TOTAL",
            "prior": total_prior,
            "current": total_current,
            "variance": net_change,
            "pct_of_net": 100
        })

        return {
            "breakdown": breakdown,
            "net_change": net_change,
            "net_change_pct": net_change_pct
        }

    def _calculate_leasing_drilldown(self, current_month: int, prior_month: int) -> Dict[str, Any]:
        """Calculate leasing drill-down with UFL linkage."""
        leasing_records = [r for r in self.revenue_records if r.revenue_type == "Leasing"]

        # Aggregate by subsidiary + tenant
        tenant_data = defaultdict(lambda: {"prior": 0, "current": 0, "subsidiary": "", "tenant": "", "tenant_code": ""})

        for r in leasing_records:
            key = (r.subsidiary, r.tenant_code or r.tenant)
            tenant_data[key]["prior"] += r.monthly_amounts.get(prior_month, 0)
            tenant_data[key]["current"] += r.monthly_amounts.get(current_month, 0)
            tenant_data[key]["subsidiary"] = r.subsidiary
            tenant_data[key]["tenant"] = r.tenant
            tenant_data[key]["tenant_code"] = r.tenant_code

        increases = []
        decreases = []
        flags = []
        period_name = f"{MONTH_NAMES[current_month - 1]} vs {MONTH_NAMES[prior_month - 1]}"

        for key, data in tenant_data.items():
            variance = data["current"] - data["prior"]

            if abs(variance) < VARIANCE_THRESHOLD:
                continue

            # Check for REVERSAL (prior negative, current positive or zero)
            is_reversal = data["prior"] < 0 and data["current"] >= 0

            # UFL matching
            ufl_action = "-"
            ufl_month = "-"
            ufl_gla = ""
            flag_type = None

            tenant_code = data["tenant_code"]

            if variance > 0:  # Increase
                if is_reversal:
                    flag_type = "REVERSAL"
                else:
                    # Look for UFL handover
                    ufl_match = self._find_ufl_handover(tenant_code, current_month)
                    if ufl_match:
                        ufl_action = "NEW LEASE"
                        ufl_month = MONTH_NAMES[ufl_match.start_date.month - 1] if ufl_match.start_date else "-"
                        ufl_gla = f"{ufl_match.gla:,.0f}" if ufl_match.gla else ""
                    else:
                        flag_type = "INCREASE NO UFL"

                increases.append({
                    "subsidiary": data["subsidiary"],
                    "tenant": data["tenant"],
                    "tenant_code": tenant_code,
                    "prior": data["prior"],
                    "current": data["current"],
                    "variance": variance,
                    "ufl_action": ufl_action,
                    "ufl_month": ufl_month,
                    "gla": ufl_gla,
                    "flag": flag_type or ""
                })

            else:  # Decrease
                # Check for negative revenue (straight-line reversal)
                if data["current"] < 0:
                    # Look for UFL termination
                    ufl_match = self._find_ufl_termination(tenant_code, current_month)
                    if ufl_match:
                        ufl_action = "TERMINATED"
                        ufl_month = MONTH_NAMES[ufl_match.termination_date.month - 1] if ufl_match.termination_date else "-"
                        ufl_gla = f"{ufl_match.gla:,.0f}" if ufl_match.gla else ""
                    else:
                        flag_type = "NEGATIVE NO UFL"
                else:
                    # Regular decrease
                    ufl_match = self._find_ufl_termination(tenant_code, current_month)
                    if ufl_match:
                        ufl_action = "TERMINATED"
                        ufl_month = MONTH_NAMES[ufl_match.termination_date.month - 1] if ufl_match.termination_date else "-"
                        ufl_gla = f"{ufl_match.gla:,.0f}" if ufl_match.gla else ""
                    else:
                        flag_type = "DECREASE NO UFL"

                decreases.append({
                    "subsidiary": data["subsidiary"],
                    "tenant": data["tenant"],
                    "tenant_code": tenant_code,
                    "prior": data["prior"],
                    "current": data["current"],
                    "variance": variance,
                    "ufl_action": ufl_action,
                    "ufl_month": ufl_month,
                    "gla": ufl_gla,
                    "flag": flag_type or ""
                })

            # Create flag if needed
            if flag_type:
                flags.append(VarianceFlag(
                    period=period_name,
                    flag_type=flag_type,
                    subsidiary=data["subsidiary"],
                    tenant=data["tenant"],
                    prior_amount=data["prior"] / 1e9,  # Convert to billions
                    current_amount=data["current"] / 1e9,
                    variance=variance / 1e9,
                    detail=self._get_flag_detail(flag_type, current_month)
                ))

        # Sort by variance magnitude
        increases.sort(key=lambda x: abs(x["variance"]), reverse=True)
        decreases.sort(key=lambda x: abs(x["variance"]), reverse=True)

        # Calculate totals
        total_variance = sum(d["current"] - d["prior"] for d in tenant_data.values())
        total_increases = sum(i["variance"] for i in increases)
        total_decreases = sum(d["variance"] for d in decreases)

        return {
            "variance": total_variance,
            "increases": increases[:15],  # Top 15
            "decreases": decreases[:15],
            "total_increases": total_increases,
            "total_decreases": total_decreases,
            "flags": flags
        }

    def _find_ufl_handover(self, tenant_code: str, current_month: int) -> Optional[UFLRecord]:
        """Find UFL handover for a tenant within 0-2 months of current month."""
        if not tenant_code:
            return None

        for ufl in self.ufl_records:
            if ufl.tenant_code != tenant_code:
                continue
            if ufl.status != "Handed Over":
                continue
            if not ufl.start_date:
                continue

            # Check if start date is within 0-2 months of current month
            if ufl.start_date.year == self.year:
                month_diff = current_month - ufl.start_date.month
                if 0 <= month_diff <= 2:
                    return ufl

        return None

    def _find_ufl_termination(self, tenant_code: str, current_month: int) -> Optional[UFLRecord]:
        """Find UFL termination for a tenant within 0-1 month of current month."""
        if not tenant_code:
            return None

        for ufl in self.ufl_records:
            if ufl.tenant_code != tenant_code:
                continue
            if ufl.status != "Terminated":
                continue
            if not ufl.termination_date:
                continue

            # Check if termination date is within 0-1 month of current month
            if ufl.termination_date.year == self.year:
                month_diff = current_month - ufl.termination_date.month
                if 0 <= month_diff <= 1:
                    return ufl

        return None

    def _check_ufl_no_revenue(self) -> List[VarianceFlag]:
        """Check for large UFL handovers without corresponding revenue increase."""
        flags = []

        # Group UFL handovers by tenant code and month
        handovers = [ufl for ufl in self.ufl_records
                     if ufl.status == "Handed Over"
                     and ufl.start_date
                     and ufl.start_date.year == self.year
                     and ufl.gla >= GLA_THRESHOLD]

        for ufl in handovers:
            handover_month = ufl.start_date.month

            # Check if there's revenue increase within 3 months
            has_revenue_increase = False
            for check_month in range(handover_month, min(handover_month + 3, 13)):
                prior_month = check_month - 1 if check_month > 1 else 1

                # Find revenue for this tenant
                tenant_revenue = [r for r in self.revenue_records
                                  if r.tenant_code == ufl.tenant_code
                                  and r.revenue_type == "Leasing"]

                for r in tenant_revenue:
                    current = r.monthly_amounts.get(check_month, 0)
                    prior = r.monthly_amounts.get(prior_month, 0)
                    if current > prior + VARIANCE_THRESHOLD:
                        has_revenue_increase = True
                        break

                if has_revenue_increase:
                    break

            if not has_revenue_increase:
                flags.append(VarianceFlag(
                    period=f"{MONTH_NAMES[handover_month - 1]} Handover",
                    flag_type="UFL NO REVENUE",
                    subsidiary=ufl.subsidiary,
                    tenant=ufl.tenant,
                    prior_amount=0,
                    current_amount=0,
                    variance=0,
                    detail=f"UFL handover {ufl.gla:,.0f} sqm but no revenue increase within 3 months"
                ))

        return flags

    def _calculate_monthly_totals(self) -> Dict[int, float]:
        """Calculate total revenue for each month."""
        totals = {}
        for month in range(1, 13):
            total = sum(r.monthly_amounts.get(month, 0) for r in self.revenue_records)
            totals[month] = total
        return totals

    def _get_flag_detail(self, flag_type: str, current_month: int) -> str:
        """Get detail message for a flag type."""
        details = {
            "INCREASE NO UFL": f"No UFL handover found (checked 0-2 months prior)",
            "DECREASE NO UFL": f"No UFL termination found (checked 0-1 month prior)",
            "NEGATIVE NO UFL": f"Revenue went negative but no termination found",
            "REVERSAL": "Prior month negative, now positive - billing correction",
            "UFL NO REVENUE": "UFL handover but no corresponding revenue increase"
        }
        return details.get(flag_type, "")

    def _detect_year(self, rows: List) -> None:
        """Detect year from header rows."""
        for row in rows[:10]:
            cells = row.findall('ss:Cell', self.ns)
            for cell in cells:
                data = cell.find('ss:Data', self.ns)
                if data is not None and data.text:
                    # Look for patterns like "Jan 2025" or "From Jan 2025"
                    match = re.search(r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s*(\d{4})', data.text)
                    if match:
                        self.year = int(match.group(2))
                        logger.info(f"Detected year: {self.year}")
                        return

    def _find_data_start(self, rows: List) -> int:
        """Find where data rows start."""
        for idx, row in enumerate(rows):
            cells = row.findall('ss:Cell', self.ns)
            values = self._extract_cell_values_with_index(cells)
            first_val = values.get(0, "").strip()

            # Data starts after account code rows
            if first_val and re.match(r'^\d{9}\s*-', first_val):
                return idx

        return 10  # Default

    def _extract_cell_values_with_index(self, cells: List) -> Dict[int, str]:
        """Extract cell values handling sparse cells with Index attribute."""
        values = {}
        col_idx = 0

        for cell in cells:
            # Check for Index attribute
            idx_attr = cell.get('{urn:schemas-microsoft-com:office:spreadsheet}Index')
            if idx_attr:
                col_idx = int(idx_attr) - 1  # Convert to 0-based

            data = cell.find('ss:Data', self.ns)
            if data is not None and data.text:
                values[col_idx] = data.text.strip()
            else:
                values[col_idx] = ""

            col_idx += 1

        return values

    def _extract_header_row(self, row) -> List[str]:
        """Extract header values from a row."""
        cells = row.findall('ss:Cell', self.ns)
        headers = []
        col_idx = 0

        for cell in cells:
            idx_attr = cell.get('{urn:schemas-microsoft-com:office:spreadsheet}Index')
            if idx_attr:
                # Fill gaps with empty strings
                target_idx = int(idx_attr) - 1
                while col_idx < target_idx:
                    headers.append("")
                    col_idx += 1

            data = cell.find('ss:Data', self.ns)
            if data is not None and data.text:
                headers.append(data.text.strip())
            else:
                headers.append("")
            col_idx += 1

        return headers

    def _get_by_header(self, values: Dict[int, str], header_map: Dict[str, int], header: str, default: str = "") -> str:
        """Get value by header name."""
        idx = header_map.get(header.lower())
        if idx is not None:
            return values.get(idx, default)
        return default

    def _parse_date(self, date_str: str) -> Optional[datetime]:
        """Parse ISO date string."""
        if not date_str:
            return None
        try:
            # Format: 2025-11-08T00:00:00
            return datetime.fromisoformat(date_str.replace("T", " ").split(".")[0])
        except ValueError:
            return None

    def _generate_excel(self, comparisons: List[MonthlyComparison],
                        all_flags: List[VarianceFlag],
                        monthly_totals: Dict[int, float],
                        ai_insights: Optional[Dict[str, Any]] = None) -> bytes:
        """Generate the 14-sheet Excel output with AI Insights."""
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Sheet 1: Flags - Unusual Items
            self._write_flags_sheet(writer, all_flags)

            # Sheet 2: Executive Summary
            self._write_executive_summary(writer, monthly_totals, all_flags)

            # Sheet 3: AI Insights
            self._write_ai_insights_sheet(writer, ai_insights or {})

            # Sheets 4-14: Monthly comparisons
            for comp in comparisons:
                sheet_name = f"{comp.current_month_name} vs {comp.prior_month_name}"
                self._write_monthly_sheet(writer, comp, sheet_name)

        return output.getvalue()

    def _write_ai_insights_sheet(self, writer, ai_insights: Dict[str, Any]):
        """Write the AI Insights sheet."""
        ws = writer.book.create_sheet("AI Insights", 2)

        # Styles
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=14)
        title_font = Font(bold=True, size=18, color='2F5496')
        section_font = Font(bold=True, size=12, color='2F5496')
        wrap_alignment = Alignment(wrap_text=True, vertical='top')

        # Title
        ws['A1'] = "AI-POWERED REVENUE VARIANCE INSIGHTS"
        ws['A1'].font = title_font
        ws.merge_cells('A1:H1')

        ws['A2'] = f"Generated by {self.model or 'AI'} | Analysis Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A2'].font = Font(italic=True, color='666666')

        row = 4

        # Executive Summary
        ws.cell(row=row, column=1, value="EXECUTIVE SUMMARY")
        ws.cell(row=row, column=1).font = section_font
        ws.cell(row=row, column=1).fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        ws.merge_cells(f'A{row}:H{row}')
        row += 1

        summary = ai_insights.get("executive_summary", "No AI summary available")
        ws.cell(row=row, column=1, value=summary)
        ws.cell(row=row, column=1).alignment = wrap_alignment
        ws.merge_cells(f'A{row}:H{row}')
        ws.row_dimensions[row].height = 80
        row += 2

        # Key Drivers
        ws.cell(row=row, column=1, value="KEY VARIANCE DRIVERS")
        ws.cell(row=row, column=1).font = section_font
        ws.cell(row=row, column=1).fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        ws.merge_cells(f'A{row}:H{row}')
        row += 1

        key_drivers = ai_insights.get("key_drivers", [])
        if key_drivers:
            # Headers
            driver_headers = ["Rank", "Driver", "Impact", "Explanation"]
            for col, header in enumerate(driver_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = header_fill
                cell.font = Font(bold=True, color='FFFFFF')
            row += 1

            for driver in key_drivers:
                # Handle both dict and string drivers
                if isinstance(driver, dict):
                    rank = driver.get("rank", "")
                    driver_text = driver.get("driver", "")
                    # Impact might be a number or string like "25.0B" or {"impact_b": 25.0}
                    impact = driver.get("impact", "") or driver.get("impact_b", "")
                    if isinstance(impact, (int, float)):
                        impact = f"{impact:.1f}B VND"
                    explanation = driver.get("explanation", "")
                    # tenants_involved might be a list
                    tenants = driver.get("tenants_involved", [])
                    if tenants and isinstance(tenants, list):
                        explanation = f"{explanation} (Tenants: {', '.join(str(t) for t in tenants[:3])})"
                else:
                    rank = ""
                    driver_text = str(driver)
                    impact = ""
                    explanation = ""

                ws.cell(row=row, column=1, value=str(rank))
                ws.cell(row=row, column=2, value=str(driver_text)[:50] if driver_text else "")
                ws.cell(row=row, column=3, value=str(impact))
                ws.cell(row=row, column=4, value=str(explanation)[:200] if explanation else "")
                ws.cell(row=row, column=4).alignment = wrap_alignment
                row += 1
        else:
            ws.cell(row=row, column=1, value="No key drivers identified")
            row += 1
        row += 1

        # Monthly Trends
        ws.cell(row=row, column=1, value="MONTHLY TRENDS")
        ws.cell(row=row, column=1).font = section_font
        ws.cell(row=row, column=1).fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        ws.merge_cells(f'A{row}:H{row}')
        row += 1

        trends = ai_insights.get("monthly_trends", "N/A")
        ws.cell(row=row, column=1, value=trends)
        ws.cell(row=row, column=1).alignment = wrap_alignment
        ws.merge_cells(f'A{row}:H{row}')
        ws.row_dimensions[row].height = 60
        row += 2

        # Flag Priorities
        ws.cell(row=row, column=1, value="FLAG INVESTIGATION PRIORITIES")
        ws.cell(row=row, column=1).font = section_font
        ws.cell(row=row, column=1).fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        ws.merge_cells(f'A{row}:H{row}')
        row += 1

        flag_priorities = ai_insights.get("flag_priorities", [])
        if flag_priorities:
            flag_headers = ["Priority", "Flag Type", "Tenant", "Recommended Action"]
            for col, header in enumerate(flag_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = header_fill
                cell.font = Font(bold=True, color='FFFFFF')
            row += 1

            for flag in flag_priorities:
                # Handle both dict flags and potential string flags
                if isinstance(flag, dict):
                    priority = str(flag.get("priority", ""))
                    flag_type = str(flag.get("flag_type", ""))
                    tenant = str(flag.get("tenant", ""))
                    # AI might use "recommended_action" or "action"
                    action = str(flag.get("recommended_action", "") or flag.get("action", ""))
                else:
                    priority = ""
                    flag_type = ""
                    tenant = str(flag)
                    action = ""

                ws.cell(row=row, column=1, value=priority)
                # Color code priority
                if priority == "HIGH":
                    ws.cell(row=row, column=1).fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                elif priority == "MEDIUM":
                    ws.cell(row=row, column=1).fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                else:
                    ws.cell(row=row, column=1).fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

                ws.cell(row=row, column=2, value=flag_type)
                ws.cell(row=row, column=3, value=tenant[:40] if tenant else "")
                ws.cell(row=row, column=4, value=action[:100] if action else "")
                row += 1
        else:
            ws.cell(row=row, column=1, value="No priority flags identified")
            row += 1
        row += 1

        # UFL Correlation Analysis
        ws.cell(row=row, column=1, value="UFL CORRELATION ANALYSIS")
        ws.cell(row=row, column=1).font = section_font
        ws.cell(row=row, column=1).fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        ws.merge_cells(f'A{row}:H{row}')
        row += 1

        ufl_analysis = ai_insights.get("ufl_correlation_analysis", "N/A")
        ws.cell(row=row, column=1, value=ufl_analysis)
        ws.cell(row=row, column=1).alignment = wrap_alignment
        ws.merge_cells(f'A{row}:H{row}')
        ws.row_dimensions[row].height = 60
        row += 2

        # Data Quality Notes
        ws.cell(row=row, column=1, value="DATA QUALITY NOTES")
        ws.cell(row=row, column=1).font = section_font
        ws.cell(row=row, column=1).fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        ws.merge_cells(f'A{row}:H{row}')
        row += 1

        data_notes = ai_insights.get("data_quality_notes", "N/A")
        # Handle list of notes
        if isinstance(data_notes, list):
            data_notes_str = "\n".join(f"• {note}" for note in data_notes) if data_notes else "No data quality issues noted"
        else:
            data_notes_str = str(data_notes) if data_notes else "N/A"
        ws.cell(row=row, column=1, value=data_notes_str)
        ws.cell(row=row, column=1).alignment = wrap_alignment
        ws.merge_cells(f'A{row}:H{row}')
        ws.row_dimensions[row].height = max(40, len(data_notes) * 20 if isinstance(data_notes, list) else 40)
        row += 2

        # Recommendations
        ws.cell(row=row, column=1, value="RECOMMENDATIONS")
        ws.cell(row=row, column=1).font = section_font
        ws.cell(row=row, column=1).fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        ws.merge_cells(f'A{row}:H{row}')
        row += 1

        recommendations = ai_insights.get("recommendations", [])
        if recommendations:
            for i, rec in enumerate(recommendations, 1):
                # Handle both string and dict recommendations
                if isinstance(rec, dict):
                    priority = rec.get("priority", "")
                    recommendation = rec.get("recommendation", "")
                    rationale = rec.get("rationale", "")
                    rec_text = f"[{priority}] {recommendation}"
                    if rationale:
                        rec_text += f" - {rationale}"
                else:
                    rec_text = str(rec)
                ws.cell(row=row, column=1, value=f"{i}. {rec_text}")
                ws.cell(row=row, column=1).alignment = wrap_alignment
                ws.merge_cells(f'A{row}:H{row}')
                row += 1
        else:
            ws.cell(row=row, column=1, value="No specific recommendations")
            row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15

    def _write_flags_sheet(self, writer, flags: List[VarianceFlag]):
        """Write the Flags - Unusual Items sheet."""
        ws = writer.book.create_sheet("Flags - Unusual Items", 0)

        # Styles
        header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=14)
        title_font = Font(bold=True, size=16)

        # Title
        ws['A1'] = "REVENUE VS UFL MISMATCH FLAGS"
        ws['A1'].font = title_font
        ws['A2'] = "Leasing Revenue (511) Only - Items requiring investigation"

        # Legend
        ws['A4'] = "FLAG TYPES:"
        ws['A4'].font = Font(bold=True)

        flag_legend = [
            ("INCREASE NO UFL", "Revenue increased but no UFL handover found", "FFEB9C"),
            ("DECREASE NO UFL", "Revenue decreased but no UFL termination found", "FFEB9C"),
            ("NEGATIVE NO UFL", "Revenue went negative but no termination found", "FFC7CE"),
            ("REVERSAL", "Prior month negative, now positive - billing correction", "B4C6E7"),
            ("UFL NO REVENUE", "UFL handover but no corresponding revenue increase", "C6EFCE"),
        ]

        for i, (flag_type, desc, color) in enumerate(flag_legend):
            ws.cell(row=5+i, column=1, value=flag_type)
            ws.cell(row=5+i, column=1).fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            ws.cell(row=5+i, column=2, value=desc)

        ws['A11'] = "NOTE: Flags apply to LEASING (511) revenue only - other revenue types do not have UFL linkage"

        # Summary counts
        flag_counts = defaultdict(int)
        for f in flags:
            flag_counts[f.flag_type] += 1

        ws['A13'] = f"TOTAL FLAGS: {len(flags)}"
        ws['A13'].font = Font(bold=True)

        row = 14
        for flag_type, count in sorted(flag_counts.items()):
            ws.cell(row=row, column=1, value=f"  {flag_type}: {count}")
            row += 1

        # Detail table
        detail_start = row + 2
        headers = ["Period", "Flag Type", "Subsidiary", "Tenant", "Prior (B)", "Current (B)", "Δ (B)", "Detail"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=detail_start, column=col, value=header)
            cell.fill = header_fill
            cell.font = Font(bold=True, color='FFFFFF')

        for i, flag in enumerate(flags):
            row = detail_start + 1 + i
            ws.cell(row=row, column=1, value=flag.period)
            ws.cell(row=row, column=2, value=flag.flag_type)
            ws.cell(row=row, column=3, value=flag.subsidiary)
            ws.cell(row=row, column=4, value=flag.tenant)
            ws.cell(row=row, column=5, value=round(flag.prior_amount, 2) if flag.prior_amount else "-")
            ws.cell(row=row, column=6, value=round(flag.current_amount, 2) if flag.current_amount else "-")
            ws.cell(row=row, column=7, value=round(flag.variance, 2))
            ws.cell(row=row, column=8, value=flag.detail)

            # Color based on flag type
            color_map = {
                "INCREASE NO UFL": "FFEB9C",
                "DECREASE NO UFL": "FFEB9C",
                "NEGATIVE NO UFL": "FFC7CE",
                "REVERSAL": "B4C6E7",
                "UFL NO REVENUE": "C6EFCE"
            }
            fill_color = color_map.get(flag.flag_type, "FFFFFF")
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 50

    def _write_executive_summary(self, writer, monthly_totals: Dict[int, float], flags: List[VarianceFlag]):
        """Write the Executive Summary sheet."""
        ws = writer.book.create_sheet("Executive Summary", 1)

        # Title
        ws['A1'] = "BW INDUSTRIAL - REVENUE VARIANCE ANALYSIS"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A2'] = f"MoM Waterfall | {self.year} | UFL Linkage for Leasing (511)"

        # Monthly Revenue table
        ws['A5'] = "MONTHLY REVENUE (VND Billion)"
        ws['A5'].font = Font(bold=True, size=12)

        # Headers
        header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        ws['A7'] = ""
        for i, month in enumerate(MONTH_NAMES):
            cell = ws.cell(row=7, column=i+2, value=month)
            cell.fill = header_fill
            cell.font = header_font

        # Revenue row
        ws['A8'] = "Revenue"
        for i in range(1, 13):
            val = monthly_totals.get(i, 0) / 1e9  # Convert to billions
            ws.cell(row=8, column=i+1, value=round(val, 1))

        # MoM Change row
        ws['A9'] = "MoM Δ"
        ws.cell(row=9, column=2, value="-")  # Jan has no prior
        for i in range(2, 13):
            prior = monthly_totals.get(i-1, 0)
            current = monthly_totals.get(i, 0)
            change = (current - prior) / 1e9
            cell = ws.cell(row=9, column=i+1, value=round(change, 1))

            # Color code
            if change > 0:
                cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            elif change < 0:
                cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        # Set column widths
        for i in range(1, 14):
            ws.column_dimensions[chr(64+i)].width = 10

    def _write_monthly_sheet(self, writer, comp: MonthlyComparison, sheet_name: str):
        """Write a monthly comparison sheet."""
        ws = writer.book.create_sheet(sheet_name)

        # Styles
        header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        section_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        increase_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        decrease_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
        flag_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        reversal_fill = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')

        # Title
        ws['A1'] = f"VARIANCE WATERFALL: {comp.current_month_name} vs {comp.prior_month_name}"
        ws['A1'].font = Font(bold=True, size=14)

        flag_count = len(comp.flags)
        ws['A2'] = f"Net Change: {comp.net_change/1e9:+.1f}B VND ({comp.net_change_pct:+.1f}%) | Flags: {flag_count}"

        # Level 1: By Revenue Stream
        ws['A4'] = "LEVEL 1: BY REVENUE STREAM"
        ws['A4'].font = Font(bold=True, size=12)

        level1_headers = ["Revenue Stream", f"{comp.prior_month_name} (B)", f"{comp.current_month_name} (B)", "Variance (B)", "% of Net"]
        for col, header in enumerate(level1_headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        for i, stream_data in enumerate(comp.stream_breakdown):
            row = 6 + i
            ws.cell(row=row, column=1, value=stream_data["stream"])
            ws.cell(row=row, column=2, value=round(stream_data["prior"]/1e9, 2) if stream_data["prior"] else "")
            ws.cell(row=row, column=3, value=round(stream_data["current"]/1e9, 2) if stream_data["current"] else "")
            ws.cell(row=row, column=4, value=round(stream_data["variance"]/1e9, 2))
            ws.cell(row=row, column=5, value=f"{stream_data['pct_of_net']:+.0f}%")

            # Bold the TOTAL row
            if stream_data["stream"] == "TOTAL":
                for col in range(1, 6):
                    ws.cell(row=row, column=col).font = Font(bold=True)

        # Level 2: Leasing Drill Down
        level2_start = 6 + len(comp.stream_breakdown) + 2
        ws.cell(row=level2_start, column=1, value="LEVEL 2: LEASING (511) DRILL DOWN - UFL Linkage")
        ws.cell(row=level2_start, column=1).font = Font(bold=True, size=12)

        ws.cell(row=level2_start + 2, column=1, value=f"Leasing: {comp.leasing_variance/1e9:+.2f}B")
        ws.cell(row=level2_start + 2, column=1).fill = section_fill
        ws.cell(row=level2_start + 2, column=1).font = Font(bold=True, color='FFFFFF')

        # Increases section
        inc_start = level2_start + 3
        ws.cell(row=inc_start, column=1, value=f"  (+) INCREASES: {comp.total_increases/1e9:+.2f}B")
        ws.cell(row=inc_start, column=1).font = Font(bold=True)

        inc_headers = ["#", "Sub", "Tenant", comp.prior_month_name, comp.current_month_name, "Δ (B)", "UFL Action", "UFL Month", "GLA", "Flag"]
        for col, header in enumerate(inc_headers, 1):
            cell = ws.cell(row=inc_start + 1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        for i, inc in enumerate(comp.leasing_increases):
            row = inc_start + 2 + i
            ws.cell(row=row, column=1, value=i+1)
            ws.cell(row=row, column=2, value=inc.get("subsidiary", ""))
            tenant = inc.get("tenant", "")
            ws.cell(row=row, column=3, value=tenant[:30] + ".." if len(tenant) > 30 else tenant)
            prior = inc.get("prior", 0)
            ws.cell(row=row, column=4, value=round(prior/1e9, 2) if prior else "-")
            current = inc.get("current", 0)
            ws.cell(row=row, column=5, value=round(current/1e9, 2) if current else "")
            variance = inc.get("variance", 0)
            ws.cell(row=row, column=6, value=round(variance/1e9, 2) if variance else 0)
            ufl_action = inc.get("ufl_action", "-")
            ws.cell(row=row, column=7, value=ufl_action)
            ws.cell(row=row, column=8, value=inc.get("ufl_month", ""))
            ws.cell(row=row, column=9, value=inc.get("gla", ""))
            flag = inc.get("flag", "")
            ws.cell(row=row, column=10, value=flag)

            # Color UFL Action
            if ufl_action == "NEW LEASE":
                ws.cell(row=row, column=7).fill = increase_fill

            # Color flags
            if flag == "REVERSAL":
                ws.cell(row=row, column=10).fill = reversal_fill
            elif flag:
                ws.cell(row=row, column=10).fill = flag_fill

        # Decreases section
        dec_start = inc_start + 2 + len(comp.leasing_increases) + 2
        ws.cell(row=dec_start, column=1, value=f"  (-) DECREASES: {comp.total_decreases/1e9:.2f}B")
        ws.cell(row=dec_start, column=1).font = Font(bold=True)

        dec_headers = ["#", "Sub", "Tenant", comp.prior_month_name, comp.current_month_name, "Δ (B)", "UFL Action", "Term Month", "GLA", "Flag"]
        for col, header in enumerate(dec_headers, 1):
            cell = ws.cell(row=dec_start + 1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        for i, dec in enumerate(comp.leasing_decreases):
            row = dec_start + 2 + i
            ws.cell(row=row, column=1, value=i+1)
            ws.cell(row=row, column=2, value=dec.get("subsidiary", ""))
            tenant = dec.get("tenant", "")
            ws.cell(row=row, column=3, value=tenant[:30] + ".." if len(tenant) > 30 else tenant)
            prior = dec.get("prior", 0)
            ws.cell(row=row, column=4, value=round(prior/1e9, 2) if prior else "-")
            current = dec.get("current", 0)
            ws.cell(row=row, column=5, value=round(current/1e9, 2) if current else "-")
            variance = dec.get("variance", 0)
            ws.cell(row=row, column=6, value=round(variance/1e9, 2) if variance else 0)
            ws.cell(row=row, column=7, value=dec.get("ufl_action", "-"))
            # Handle both "term_month" and "ufl_month" keys
            ws.cell(row=row, column=8, value=dec.get("term_month", "") or dec.get("ufl_month", ""))
            ws.cell(row=row, column=9, value=dec.get("gla", ""))
            ws.cell(row=row, column=10, value=dec.get("flag", ""))

            # Color UFL Action
            if dec.get("ufl_action") == "TERMINATED":
                ws.cell(row=row, column=7).fill = decrease_fill

            # Color flags
            if dec.get("flag"):
                ws.cell(row=row, column=10).fill = flag_fill

        # Set column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 10
        ws.column_dimensions['J'].width = 18


def analyze_revenue_variance(
    revenue_bytes: bytes,
    ufl_bytes: bytes,
    api_key: Optional[str] = None,
    model: Optional[str] = None,
    return_usage: bool = False,
) -> Any:
    """
    AI-powered revenue variance analysis with UFL linkage.

    Args:
        revenue_bytes: Raw bytes of RevenueBreakdown file
        ufl_bytes: Raw bytes of UnitForLeaseList file
        api_key: Optional API key (defaults to OPENAI_API_KEY or ANTHROPIC_API_KEY env var)
        model: Optional model name (defaults to AI_MODEL env var or "gpt-4o")
        return_usage: If True, return (bytes, usage_dict) tuple

    Returns:
        Excel file bytes, or (bytes, usage_dict) tuple if return_usage=True
    """
    analyzer = RevenueVarianceAnalyzer(api_key=api_key, model=model)
    xlsx_bytes = analyzer.analyze(revenue_bytes, ufl_bytes)
    if return_usage:
        return xlsx_bytes, analyzer.get_and_reset_usage()
    return xlsx_bytes
