#!/usr/bin/env python3
"""
Variance Analysis Pipeline - 21 Rules

This module processes financial data (BS and PL) and applies 21 variance analysis rules
to detect anomalies, process breakdowns, and compliance issues.

Rules:
- 8 Critical Rules (🔴): Data integrity / Process breakdown
- 9 Review Rules (🟡): Material movements requiring explanation
- 4 Info Rules (🟢): Optimization / Expected patterns

Adapted for FastAPI backend integration.
"""

import re
import io
from typing import List, Tuple, Optional, Dict, Any
from pathlib import Path
from datetime import datetime
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import warnings
warnings.filterwarnings('ignore')

# Import loan interest rate parser for enhanced A2
from app.domain.finance.variance_analysis.services.loan_interest_parser import (
    parse_loan_interest_file,
    build_entity_rate_lookup,
    get_applicable_rate,
    get_rate_for_period,
    calculate_expected_interest,
    find_matching_entity,
    generate_a2_explanation,
    get_days_in_month,
    MONTH_ABBR_TO_NUM
)


# ============================================================================
# CONFIGURATION AND CONSTANTS
# ============================================================================

# Statistical thresholds
STD_DEV_THRESHOLD = 2.0  # 2 standard deviations for outlier detection
MONTHS_3M_AVG = 3  # 3-month average baseline
MONTHS_6M_AVG = 6  # 6-month average baseline

# Materiality thresholds (VND)
MATERIALITY_THRESHOLDS = {
    'vat_input_threshold': 10_000_000_000,  # 10B VND for E3
    'retained_earnings_tolerance': 100_000_000,  # 100M VND for D2
    'ar_growth_threshold': 0.30,  # 30% AR growth for E5
}

# Month tokens for detection
MONTH_TOKENS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
MONTHS = ["jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec"]

# Days in each month (non-leap year)
DAYS_IN_MONTH = {
    'Jan': 31, 'Feb': 28, 'Mar': 31, 'Apr': 30,
    'May': 31, 'Jun': 30, 'Jul': 31, 'Aug': 31,
    'Sep': 30, 'Oct': 31, 'Nov': 30, 'Dec': 31
}

# Quarter-start months for E4
QUARTER_START_MONTHS = ['Jan', 'Apr', 'Jul', 'Oct']

# Regex patterns for BS and PL month detection (expanded for edge cases)
BS_PAT = re.compile(
    r'^\s*(as\s*of\s*|tinh\s*den\s*|tính\s*đến\s*|den\s*ngay\s*|đến\s*ngày\s*)?' +
    r'(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[\.\-\s,]*(\d{2,4})\s*$',
    re.I
)
PL_PAT = re.compile(
    r'^\s*(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[\.\-\s,]*(\d{2,4})\s*$',
    re.I
)


# ============================================================================
# EXCEL PROCESSING FUNCTIONS
# ============================================================================

def extract_current_period_from_row4(xl_file, sheet_name):
    """Extract current period from Row 4 (index 3) of Excel sheet with fallbacks."""
    try:
        df = pd.read_excel(xl_file, sheet_name=sheet_name, header=None, nrows=10, dtype=str)

        # Try multiple rows (common locations for period info)
        rows_to_check = [3, 2, 4, 1, 5]  # Row 4 first, then other common locations

        for row_idx in rows_to_check:
            if row_idx >= len(df):
                continue

            # Check multiple columns (sometimes period is in B, C, etc.)
            for col_idx in range(min(5, len(df.columns))):
                try:
                    cell_text = str(df.iloc[row_idx, col_idx]).strip()

                    # Enhanced patterns to find dates
                    # Pattern 1: "From Jan 2025 to Sep 2025" or "End of Sep 2025"
                    matches = re.findall(r'(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\s+(\d{4})', cell_text, re.I)
                    if matches:
                        month_name, year = matches[-1]  # Take the last match
                        month_num = MONTHS.index(month_name.lower()) + 1
                        return int(year), month_num, f"{month_name.title()} {year}"

                    # Pattern 2: MM/YYYY or MM-YYYY
                    match = re.search(r'(\d{1,2})[/\-](\d{4})', cell_text)
                    if match:
                        month_num = int(match.group(1))
                        year = int(match.group(2))
                        if 1 <= month_num <= 12:
                            month_name = MONTHS[month_num - 1]
                            return year, month_num, f"{month_name.title()} {year}"

                    # Pattern 3: YYYY/MM or YYYY-MM
                    match = re.search(r'(\d{4})[/\-](\d{1,2})', cell_text)
                    if match:
                        year = int(match.group(1))
                        month_num = int(match.group(2))
                        if 1 <= month_num <= 12:
                            month_name = MONTHS[month_num - 1]
                            return year, month_num, f"{month_name.title()} {year}"

                except Exception:
                    continue

    except Exception as e:
        print(f"Warning: Could not extract current period from {sheet_name}: {e}")

    return None, None, None


def filter_months_up_to_current(month_cols, current_year, current_month):
    """Filter month columns to only include months up to current period."""
    if current_year is None or current_month is None:
        return month_cols

    filtered_months = []
    for month_col in month_cols:
        m = re.search(r'(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\s+(\d{4})', month_col, re.I)
        if m:
            month_name = m.group(1).lower()
            year = int(m.group(2))
            month_num = MONTHS.index(month_name) + 1

            if (year < current_year) or (year == current_year and month_num <= current_month):
                filtered_months.append(month_col)

    return filtered_months


def normalize_period_label(label):
    """Normalize period labels to 'Mon YYYY' format"""
    if label is None: return ""
    s = str(label).strip()
    if s == "": return ""
    try:
        s_clean = re.sub(r'^\s*(as\s*of|tinh\s*den|tính\s*đến|den\s*ngay|đến\s*ngày)\s*', '', s, flags=re.I)
        m = re.search(r'\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[^\w]?[\s\-\.]*([12]\d{3}|\d{2})\b', s_clean, flags=re.I)
        if m:
            mon, yr = m.group(1), m.group(2)
            yr = int(yr)
            yr = yr+2000 if yr < 100 else yr
            return f"{mon.title()} {yr}"
        m = re.search(r'\b(1[0-2]|0?[1-9])[./\-](\d{4})\b', s_clean)
        if m:
            mon = int(m.group(1))
            yr = int(m.group(2))
            return f"{MONTHS[mon-1].title()} {yr}"
        m = re.search(r'\b(\d{4})[./\-](1[0-2]|0?[1-9])\b', s_clean)
        if m:
            yr = int(m.group(1))
            mon = int(m.group(2))
            return f"{MONTHS[mon-1].title()} {yr}"
        m_year = re.search(r'(20\d{2}|19\d{2})', s_clean)
        m_mon = re.search(r'\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\b', s_clean, flags=re.I)
        if m_year and m_mon:
            yr = int(m_year.group(1))
            mon = m_mon.group(0)
            return f"{mon.title()} {yr}"
    except Exception:
        pass
    return s


def month_key(label):
    """Convert month label to sortable tuple (year, month_num)"""
    n = normalize_period_label(label)
    m = re.search(r'(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\s+(\d{4})', n, re.I)
    if not m: return (9999, 99)
    y = int(m.group(2))
    mi = MONTHS.index(m.group(1).lower()) + 1
    return (y, mi)


def detect_header_row(xl, sheet):
    """Find the row containing 'Financial row' header with multiple fallback patterns"""
    try:
        probe = pd.read_excel(xl, sheet_name=sheet, header=None, nrows=50, dtype=str)

        # Common header patterns to search for
        header_patterns = [
            "financial row",
            "financial item",
            "account",
            "line item",
            "description",
            "chỉ tiêu",  # Vietnamese: indicator/metric
            "khoản mục",  # Vietnamese: item/account
        ]

        for i in range(len(probe)):
            row_values = probe.iloc[i].astype(str).str.strip().str.lower()

            # Check for specific "Financial Row" pattern first (most reliable)
            if any("financial row" in v for v in row_values):
                return i

            # Check for other header patterns
            for pattern in header_patterns:
                if pattern == "financial row":
                    continue  # Already checked above
                if any(pattern in v for v in row_values):
                    return i

            # Also check if row has "Entity" as a column header (not in company description)
            # Must be a short value to avoid matching company descriptions
            if any(v == "entity" or v == "subsidiary" for v in row_values):
                return i

    except Exception as e:
        print(f"Warning: Header detection failed for {sheet}: {e}")

    # Default fallback: row 5 (common in many templates)
    return 5


def process_financial_tab(xl_file, sheet_name, mode, subsidiary):
    """Process BS or PL sheet and extract data."""
    try:
        current_year, current_month, current_period_str = extract_current_period_from_row4(xl_file, sheet_name)
        header_row = detect_header_row(xl_file, sheet_name)
        df_raw = pd.read_excel(xl_file, sheet_name=sheet_name, header=None, dtype=str)
        headers = df_raw.iloc[header_row].tolist()
        # Data typically starts 3 rows after header (header_row + month_header_row + empty_row + data)
        # Row 7 (header) -> Row 8 (month headers) -> Row 9 (empty) -> Row 10 (data)
        data_start = header_row + 3
        df = df_raw.iloc[data_start:].copy()
        df.columns = [f'col_{i}' for i in range(len(df.columns))]

        # === FLEXIBLE COLUMN STRUCTURE DETECTION ===
        # Instead of hardcoding column positions, detect them from the header row
        headers_lower = [str(h).lower().strip() for h in headers]

        print(f"DEBUG: ========== COLUMN STRUCTURE DETECTION for {sheet_name} ==========")
        print(f"DEBUG: Headers (first 10): {headers[:10]}")

        # Find Financial Row column
        financial_row_col = None
        for i, h in enumerate(headers_lower):
            if any(keyword in h for keyword in ['financial row', 'account code', 'financial code', 'mã tk']):
                financial_row_col = i
                print(f"DEBUG: Found 'Financial Row' at column {i}: '{headers[i]}'")
                break

        if financial_row_col is None:
            financial_row_col = 0  # fallback
            print(f"DEBUG: Using default col 0 for Financial Row")

        # Find Account Line column
        account_line_col = None
        for i, h in enumerate(headers_lower):
            # More flexible matching for Account Line variations
            if any(keyword in h for keyword in ['account line', 'account name', 'description', 'tên tk', 'diễn giải']):
                account_line_col = i
                print(f"DEBUG: Found 'Account Line' at column {i}: '{headers[i]}'")
                break
            # Also check for patterns like "Account (Line): Mã số - Code"
            if 'account' in h and ('line' in h or 'mã số' in h or 'code' in h):
                account_line_col = i
                print(f"DEBUG: Found 'Account Line' (pattern match) at column {i}: '{headers[i]}'")
                break

        if account_line_col is None:
            # Fallback based on mode
            account_line_col = 3 if mode == 'BS' else 2
            print(f"DEBUG: Using default col {account_line_col} for Account Line (mode={mode})")

        # Find Entity columns (optional)
        entity_cols = []
        for i, h in enumerate(headers_lower):
            if 'entity' in h or 'subsidiary' in h or 'đơn vị' in h:
                entity_cols.append(i)

        # Determine where month columns likely start (after the last structural column)
        structural_cols = [financial_row_col, account_line_col] + entity_cols
        month_start_col = max(structural_cols) + 1

        print(f"DEBUG: Financial Row column: {financial_row_col}")
        print(f"DEBUG: Account Line column: {account_line_col}")
        print(f"DEBUG: Entity columns: {entity_cols}")
        print(f"DEBUG: Month columns expected to start at: {month_start_col}")

        # Rename columns based on detection
        df = df.rename(columns={f'col_{financial_row_col}': 'Financial_Row'})
        df = df.rename(columns={f'col_{account_line_col}': 'Account_Line'})

        for i, entity_idx in enumerate(entity_cols):
            df = df.rename(columns={f'col_{entity_idx}': f'Entity_{i}'})

        filter_col = 'Account_Line'

        # Forward fill Financial Row and filter empty rows
        if 'Financial_Row' in df.columns:
            df['Financial_Row'] = df['Financial_Row'].fillna(method='ffill')

        # Filter rows: Keep rows with Account_Line OR rows that are total rows
        # Total rows have keywords like "TỔNG CỘNG", "Total", etc. in Financial_Row
        print(f"DEBUG: Filtering rows - filter_col={filter_col}, filter_col in df.columns={filter_col in df.columns}, Financial_Row in df.columns={'Financial_Row' in df.columns}")
        if filter_col in df.columns and 'Financial_Row' in df.columns:
            # Keep rows that either:
            # 1. Have Account_Line (detail rows)
            # 2. OR have total keywords in Financial_Row (total rows)
            has_account_line = df[filter_col].notna() & (df[filter_col].astype(str).str.strip() != '')
            is_total_row = df['Financial_Row'].astype(str).str.contains(
                'TỔNG CỘNG|Total|TOTAL|Tổng cộng',
                case=False,
                na=False,
                regex=True
            )
            print(f"DEBUG: Rows with Account_Line: {has_account_line.sum()}")
            print(f"DEBUG: Rows with total keywords: {is_total_row.sum()}")
            print(f"DEBUG: Total rows to keep: {(has_account_line | is_total_row).sum()}")
            df = df[has_account_line | is_total_row].copy()
            print(f"DEBUG: After filtering, df has {len(df)} rows")
        elif filter_col in df.columns:
            print(f"DEBUG: Using fallback filtering (Financial_Row not in columns)")
            df = df[df[filter_col].notna() & (df[filter_col].astype(str).str.strip() != '')].copy()
        else:
            print(f"WARNING: {filter_col} column not found after renaming!")

        # === FLEXIBLE MONTH COLUMN DETECTION ===
        # Step 1: Try to detect where month columns actually start by checking ALL columns
        # (not just from month_start_col) since structure may vary

        month_cols = []
        month_col_indices = []
        month_row_found = None

        # Try ALL rows in the header area (not just near header_row)
        rows_to_try = list(range(0, min(20, len(df_raw))))  # Check first 20 rows thoroughly

        print(f"DEBUG: ========== MONTH HEADER DETECTION for {sheet_name} ==========")
        print(f"DEBUG: Header row detected at: {header_row}")
        print(f"DEBUG: Expected month_start_col (based on mode): {month_start_col}")
        print(f"DEBUG: Will scan first 20 rows thoroughly to find month headers")

        for row_idx in rows_to_try:
            if row_idx >= len(df_raw):
                continue

            test_row = df_raw.iloc[row_idx]

            # Show full row for debugging
            sample_vals = [str(test_row.iloc[i]).strip() if i < len(test_row) else 'N/A' for i in range(min(15, len(test_row)))]

            temp_month_cols = []
            temp_month_indices = []

            # Check EVERY column (from 0 onwards), not just month_start_col
            for i in range(0, len(test_row)):
                val = str(test_row.iloc[i]).strip()

                if val == 'nan' or val == '':
                    continue

                # Try both patterns
                bs_match = BS_PAT.match(val)
                pl_match = PL_PAT.match(val)

                if bs_match or pl_match:
                    normalized = normalize_period_label(val)
                    val_lower = val.lower()

                    # Exclude YTD/LTM columns
                    if not any(keyword in val_lower for keyword in ['ytd', 'ltm', 'amount ytd', 'amount ltm', 'year to date', 'last twelve']):
                        if re.match(r'^(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\s+\d{4}$', normalized, re.I):
                            temp_month_cols.append(normalized)
                            temp_month_indices.append(i)

            if temp_month_cols:
                month_cols = temp_month_cols
                month_col_indices = temp_month_indices
                month_row_found = row_idx
                print(f"DEBUG: ✓✓✓ FOUND {len(month_cols)} month columns in ROW {row_idx}")
                print(f"DEBUG: Month columns: {month_cols[:8]}")
                print(f"DEBUG: Column indices: {month_col_indices[:8]}")
                print(f"DEBUG: Full row {row_idx}: {sample_vals}")
                break
            else:
                # Only print rows that have non-empty values
                if any(v not in ['nan', ''] for v in sample_vals):
                    print(f"DEBUG: Row {row_idx:2d} [{len(temp_month_cols)} months]: {sample_vals}")

        if not month_cols:
            print(f"DEBUG: ✗✗✗ NO MONTH COLUMNS FOUND in {sheet_name}")
            print(f"DEBUG: Dumping detailed first 20 rows for manual inspection:")
            for idx in range(min(20, len(df_raw))):
                row_sample = [str(df_raw.iloc[idx, col]).strip() for col in range(min(15, len(df_raw.columns)))]
                print(f"  Row {idx:2d}: {row_sample}")

        for idx, month in zip(month_col_indices, month_cols):
            df = df.rename(columns={f'col_{idx}': month})

        for month in month_cols:
            if month in df.columns:
                series = df[month].astype(str)
                series = (series
                    .str.replace("\u00a0","", regex=False)
                    .str.replace(",","", regex=False)
                    .str.replace(r"\((.*)\)", r"-\1", regex=True)
                    .str.replace(r"[^0-9\.\-]", "", regex=True)
                )
                df[month] = pd.to_numeric(series, errors="coerce").fillna(0.0)

        final_cols = ['Financial_Row', 'Account_Line'] + month_cols
        df = df[[col for col in final_cols if col in df.columns]].copy()

        df = df.rename(columns={
            'Financial_Row': 'Account Code',
            'Account_Line': 'Account Line'
        })

        # Check if we have any data to process
        if df.empty:
            print(f"Warning: No data rows found in {sheet_name} for {subsidiary}")
            return pd.DataFrame(), []

        # Check if we have month columns
        available_month_cols = [month for month in month_cols if month in df.columns]
        if not available_month_cols:
            print(f"Warning: No month columns found in {sheet_name} for {subsidiary}")
            return pd.DataFrame(), []

        # Perform aggregation only if we have month columns
        # Split dataframe into detail rows (with Account Line) and total rows (without Account Line)
        detail_rows = df[df['Account Line'].notna()].copy()
        total_rows = df[df['Account Line'].isna()].copy()

        print(f"DEBUG: Before aggregation - detail_rows: {len(detail_rows)}, total_rows: {len(total_rows)}")

        # Aggregate detail rows
        agg_dict = {month: 'sum' for month in available_month_cols}
        result_detail = detail_rows.groupby(['Account Code', 'Account Line'], as_index=False).agg(agg_dict)
        result_detail['Account Name'] = result_detail['Account Code']

        # For total rows, keep them as-is (no aggregation needed, they're already totals)
        if not total_rows.empty:
            total_rows['Account Name'] = total_rows['Account Code']
            # Combine detail and total rows
            result = pd.concat([result_detail, total_rows], ignore_index=True)
        else:
            result = result_detail

        print(f"DEBUG: After aggregation - result has {len(result)} rows")

        final_column_order = ['Account Code', 'Account Line', 'Account Name'] + available_month_cols
        result = result[final_column_order]

        month_cols = sorted(available_month_cols, key=month_key)
        filtered_month_cols = filter_months_up_to_current(month_cols, current_year, current_month)

        keep_cols = ['Account Code', 'Account Line', 'Account Name'] + filtered_month_cols
        result = result[[col for col in keep_cols if col in result.columns]]

        return result, filtered_month_cols

    except Exception as e:
        print(f"Error processing {sheet_name} for {subsidiary}: {e}")
        return pd.DataFrame(), []


def find_sheet_by_pattern(wb, patterns):
    """Find sheet that matches any of the given patterns (case-insensitive)"""
    sheet_names_lower = {name.lower(): name for name in wb.sheetnames}

    for pattern in patterns:
        pattern_lower = pattern.lower()
        # Exact match
        if pattern_lower in sheet_names_lower:
            return sheet_names_lower[pattern_lower]

        # Partial match (contains pattern)
        for lower_name, actual_name in sheet_names_lower.items():
            if pattern_lower in lower_name or lower_name in pattern_lower:
                return actual_name

    return None


def extract_subsidiary_name(xl_file):
    """Extract subsidiary name from cell A2 with multiple fallback strategies"""
    try:
        wb = load_workbook(xl_file, read_only=True, data_only=True)

        # Try common sheet name variations
        # IMPORTANT: Include both "BS Breakdown" (with space) and "BSbreakdown" (no space)
        # Priority: Try breakdown sheets first, then standalone sheets
        bs_patterns = ["BS Breakdown", "BSbreakdown", "BS breakdown", "bs breakdown",
                       "balance sheet breakdown", "BẢNG CÂN ĐỐI KẾ TOÁN", "Balance Sheet", "BS"]
        pl_patterns = ["PL Breakdown", "PLBreakdown", "PL breakdown", "pl breakdown",
                       "Profit Loss", "Income Statement", "BÁO CÁO KẾT QUẢ KINH DOANH", "P&L", "P/L", "PL"]

        # Try to find any sheet
        for patterns in [bs_patterns, pl_patterns]:
            sheet_name = find_sheet_by_pattern(wb, patterns)
            if sheet_name:
                sheet = wb[sheet_name]

                # Try A2 first (most common)
                cell_value = sheet["A2"].value
                if isinstance(cell_value, str) and ":" in cell_value:
                    wb.close()
                    return cell_value.split(":")[-1].strip()

                # Try A1 as fallback
                cell_value = sheet["A1"].value
                if isinstance(cell_value, str) and ":" in cell_value:
                    wb.close()
                    return cell_value.split(":")[-1].strip()

        wb.close()
    except Exception:
        pass

    return "Subsidiary"


def extract_entity_from_file(xl_file) -> Optional[str]:
    """
    Extract entity name from Row 1 (index 1) of BS/PL Breakdown file.

    Row 1 format: "Parent Company : BWID : VC1 : VC2 : ENTITY_NAME"
    Entity is always the LAST part after ":"

    Args:
        xl_file: File-like object or path to Excel file

    Returns:
        Entity name, or None if not found
    """
    try:
        wb = load_workbook(xl_file, read_only=True, data_only=True)

        # Try common sheet name variations
        bs_patterns = ["BS Breakdown", "BSbreakdown", "BS breakdown", "bs breakdown",
                       "balance sheet breakdown", "BẢNG CÂN ĐỐI KẾ TOÁN", "Balance Sheet", "BS"]
        pl_patterns = ["PL Breakdown", "PLBreakdown", "PL breakdown", "pl breakdown",
                       "Profit Loss", "Income Statement", "BÁO CÁO KẾT QUẢ KINH DOANH", "P&L", "P/L", "PL"]

        for patterns in [bs_patterns, pl_patterns]:
            sheet_name = find_sheet_by_pattern(wb, patterns)
            if sheet_name:
                sheet = wb[sheet_name]

                # Row 2 (index 1 in 0-based, cell A2) contains "Parent Company : ... : ENTITY"
                cell_value = sheet["A2"].value
                if isinstance(cell_value, str) and ":" in cell_value:
                    # Split by ":" and get the LAST part (entity name)
                    parts = cell_value.split(":")
                    entity = parts[-1].strip()
                    wb.close()
                    return entity if entity else None

        wb.close()
    except Exception as e:
        print(f"Error extracting entity from Row 1: {e}")

    return None


def parse_month_year_from_column(col_name: str) -> Tuple[Optional[str], Optional[int]]:
    """
    Parse month name and year from column header.

    Args:
        col_name: Column name like 'Jan 2025' or 'As of Sep 2025'

    Returns:
        Tuple of (month_name, year) e.g., ('Sep', 2025)
    """
    col_str = str(col_name).replace('As of ', '').strip()

    for month in MONTH_TOKENS:
        if month in col_str:
            # Extract year
            year_match = re.search(r'(\d{4})', col_str)
            if year_match:
                return month, int(year_match.group(1))
            return month, None

    return None, None


# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def extract_month_from_column(col_name):
    """Extract month name from column header like 'Jan 2025'"""
    col_str = str(col_name).replace('As of ', '').strip()
    for month in MONTH_TOKENS:
        if month in col_str:
            return month
    return None


def normalize_interest_for_calendar(value, month_name):
    """Normalize interest to 30-day basis for comparison"""
    if month_name not in DAYS_IN_MONTH:
        return value
    days = DAYS_IN_MONTH[month_name]
    if days == 30:
        return value
    return (value / days) * 30


def get_account_pattern_data(df, pattern):
    """Get rows matching account code pattern"""
    if df is None or df.empty:
        return pd.DataFrame()

    if pattern.endswith('xxx'):
        prefix = pattern[:-3]
        return df[df['Account Code'].astype(str).str.startswith(prefix)].copy()
    else:
        return df[df['Account Code'].astype(str) == pattern].copy()


def get_month_cols(df):
    """Extract month columns from dataframe"""
    if df is None or df.empty:
        return []

    month_cols = []
    for col in df.columns:
        if isinstance(col, str) and re.match(r'^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{4}$', col):
            month_cols.append(col)

    return sorted(month_cols, key=month_key)


# ============================================================================
# RULE IMPLEMENTATIONS - CRITICAL (🔴)
# ============================================================================

def check_rule_A1(bs_df, pl_df):
    """A1 - Asset capitalized but depreciation not started

    Checks if Investment Property increased but depreciation/amortization did not increase.
    Only looks at actual D&A accounts (632100001 amortization, 632100002 depreciation),
    NOT the total 632xxx which includes other cost of goods sold items.
    """
    flags = []
    if bs_df is None or pl_df is None or bs_df.empty or pl_df.empty:
        return flags

    ip_accounts = get_account_pattern_data(bs_df, '217xxx')

    # Get ONLY depreciation and amortization accounts, not all 632xxx
    # 632100001 = Amortization (Land Use Rights)
    # 632100002 = Depreciation (Buildings/Assets)
    dep_amort_accounts = pl_df[
        pl_df['Account Line'].astype(str).isin(['632100001', '632100002'])
    ].copy()

    if ip_accounts.empty or dep_amort_accounts.empty:
        return flags

    month_cols = get_month_cols(bs_df)
    if len(month_cols) < 2:
        return flags

    for i in range(1, len(month_cols)):
        prev_month = month_cols[i-1]
        curr_month = month_cols[i]

        ip_change = ip_accounts[curr_month].sum() - ip_accounts[prev_month].sum()
        dep_prev = dep_amort_accounts[prev_month].sum() if prev_month in dep_amort_accounts.columns else 0
        dep_curr = dep_amort_accounts[curr_month].sum() if curr_month in dep_amort_accounts.columns else 0
        dep_change = dep_curr - dep_prev

        # Flag if IP increased but D&A did not increase
        if ip_change > 0 and dep_change <= 0:
            flags.append({
                'Rule_ID': 'A1',
                'Priority': '🔴 Critical',
                'Issue': 'Asset capitalized but depreciation not started',
                'Accounts': '217xxx ↔ 632100001/632100002',
                'Period': f"{prev_month} → {curr_month}",
                'Reason': f'IP increased by {ip_change:,.0f} VND but D&A changed by {dep_change:,.0f} VND',
                'Flag_Trigger': 'IP↑ BUT D&A ≤ previous'
            })
    return flags


def check_rule_A2(bs_df, pl_df):
    """A2 - Loan drawdown but interest not recorded"""
    flags = []
    if bs_df is None or pl_df is None or bs_df.empty or pl_df.empty:
        return flags

    loan_accounts = get_account_pattern_data(bs_df, '341xxx')
    interest_expense = get_account_pattern_data(pl_df, '635xxx')
    cip_interest = get_account_pattern_data(bs_df, '241xxx')
    if loan_accounts.empty:
        return flags

    month_cols = get_month_cols(bs_df)
    if len(month_cols) < 2:
        return flags

    for i in range(1, len(month_cols)):
        prev_month = month_cols[i-1]
        curr_month = month_cols[i]

        prev_month_name = extract_month_from_column(prev_month)
        curr_month_name = extract_month_from_column(curr_month)

        loan_change = loan_accounts[curr_month].sum() - loan_accounts[prev_month].sum()

        interest_prev_raw = interest_expense[prev_month].sum() if not interest_expense.empty else 0
        interest_curr_raw = interest_expense[curr_month].sum() if not interest_expense.empty else 0
        cip_prev_raw = cip_interest[prev_month].sum() if not cip_interest.empty else 0
        cip_curr_raw = cip_interest[curr_month].sum() if not cip_interest.empty else 0

        interest_prev_adj = normalize_interest_for_calendar(interest_prev_raw, prev_month_name)
        interest_curr_adj = normalize_interest_for_calendar(interest_curr_raw, curr_month_name)
        cip_prev_adj = normalize_interest_for_calendar(cip_prev_raw, prev_month_name)
        cip_curr_adj = normalize_interest_for_calendar(cip_curr_raw, curr_month_name)

        interest_change = (interest_curr_adj + cip_curr_adj) - (interest_prev_adj + cip_prev_adj)

        if loan_change > 0 and interest_change <= 0:
            flags.append({
                'Rule_ID': 'A2',
                'Priority': '🔴 Critical',
                'Issue': 'Loan drawdown but interest not recorded',
                'Accounts': '341xxx ↔ 635xxx + 241xxx',
                'Period': f"{prev_month} → {curr_month}",
                'Reason': f'Loan increased by {loan_change:,.0f} VND but day-adjusted interest changed by {interest_change:,.0f} VND',
                'Flag_Trigger': 'Loan↑ BUT Day-adjusted Interest ≤ previous'
            })
    return flags


def check_rule_A2_enhanced(
    bs_df: pd.DataFrame,
    pl_df: pd.DataFrame,
    entity_name: Optional[str],
    loan_rate_lookup: Optional[Dict[str, List[Tuple[datetime, float]]]]
) -> List[Dict[str, Any]]:
    """
    A2 Enhanced - Loan Interest Analysis with ERP Rate Data

    Performs deep-dive analysis of loan/interest relationships using
    interest rate data from ERP system.

    Logic:
    - If loan increased but interest didn't → Missing interest accrual
    - If loan unchanged but interest changed → Check if rate changed
    - Calculate expected interest: Principal × Rate × (Days/365)
    - Compare expected vs actual and generate explanation

    Args:
        bs_df: Balance Sheet DataFrame
        pl_df: P&L DataFrame
        entity_name: Entity name extracted from file (for rate lookup)
        loan_rate_lookup: Dict mapping entity names to rate history

    Returns:
        List of enhanced A2 variance flags
    """
    flags = []

    if bs_df is None or pl_df is None or bs_df.empty or pl_df.empty:
        return flags

    # Get loan and interest data
    loan_accounts = get_account_pattern_data(bs_df, '341xxx')
    interest_expense = get_account_pattern_data(pl_df, '635xxx')
    cip_interest = get_account_pattern_data(bs_df, '241xxx')

    if loan_accounts.empty:
        return flags

    month_cols = get_month_cols(bs_df)
    if len(month_cols) < 2:
        return flags

    # Find matching entity in loan rate lookup
    matched_entity = None
    entity_rates = []

    if entity_name and loan_rate_lookup:
        available_entities = list(loan_rate_lookup.keys())
        matched_entity = find_matching_entity(entity_name, available_entities)
        if matched_entity:
            entity_rates = loan_rate_lookup[matched_entity]

    # Analyze each month transition
    for i in range(1, len(month_cols)):
        prev_month = month_cols[i-1]
        curr_month = month_cols[i]

        # Parse month/year for calculations
        prev_month_name, prev_year = parse_month_year_from_column(prev_month)
        curr_month_name, curr_year = parse_month_year_from_column(curr_month)

        if not curr_month_name or not curr_year:
            continue

        # Calculate loan balances and changes
        loan_prev = loan_accounts[prev_month].sum()
        loan_curr = loan_accounts[curr_month].sum()
        loan_change = loan_curr - loan_prev

        # Calculate interest (raw values)
        interest_prev_raw = interest_expense[prev_month].sum() if not interest_expense.empty else 0
        interest_curr_raw = interest_expense[curr_month].sum() if not interest_expense.empty else 0
        cip_prev_raw = cip_interest[prev_month].sum() if not cip_interest.empty else 0
        cip_curr_raw = cip_interest[curr_month].sum() if not cip_interest.empty else 0

        total_interest_prev = abs(interest_prev_raw) + abs(cip_prev_raw)
        total_interest_curr = abs(interest_curr_raw) + abs(cip_curr_raw)
        interest_change = total_interest_curr - total_interest_prev

        # Get interest rates for the period
        current_rate = None
        previous_rate = None
        rate_change_date = None
        expected_interest = None

        if entity_rates:
            # Build datetime for period boundaries
            prev_month_num = MONTH_ABBR_TO_NUM.get(prev_month_name, 1) if prev_month_name else 1
            curr_month_num = MONTH_ABBR_TO_NUM.get(curr_month_name, 1)

            period_start = datetime(prev_year or curr_year, prev_month_num, 1)
            # End of current month
            days_in_curr = get_days_in_month(curr_month_name, curr_year)
            period_end = datetime(curr_year, curr_month_num, days_in_curr)

            current_rate, previous_rate, rate_change_date = get_rate_for_period(
                entity_rates, period_start, period_end
            )

            # Calculate expected interest for current month
            if current_rate and loan_curr > 0:
                expected_interest = calculate_expected_interest(
                    loan_curr, current_rate, curr_month_name, curr_year
                )

        # Determine if we should flag and what explanation to provide
        should_flag = False
        flag_priority = '🔴 Critical'
        flag_issue = 'Loan drawdown but interest not recorded'

        # Scenario 1: Loan increased but interest didn't increase (basic A2)
        if loan_change > 0 and interest_change <= 0:
            should_flag = True
            if previous_rate is not None and current_rate != previous_rate:
                # Rate changed - might explain the discrepancy
                flag_priority = '🟡 Review'
                flag_issue = 'Loan increased but interest pattern changed - rate adjustment detected'
            else:
                flag_issue = 'Loan drawdown but interest not recorded'

        # Scenario 2: Loan unchanged but interest changed significantly
        elif abs(loan_change) < 1_000_000_000 and abs(interest_change) > 100_000_000:
            # Loan stable but interest changed by > 100M
            if previous_rate is not None and current_rate != previous_rate:
                should_flag = True
                flag_priority = '🟢 Info'
                flag_issue = 'Interest change due to rate adjustment'

        # Scenario 3: Expected vs Actual variance
        elif expected_interest is not None and total_interest_curr > 0:
            variance = abs(total_interest_curr - expected_interest)
            variance_pct = variance / expected_interest if expected_interest > 0 else 0

            if variance_pct > 0.10 and variance > 100_000_000:  # >10% and >100M variance
                should_flag = True
                flag_priority = '🟡 Review'
                flag_issue = 'Interest variance from expected calculation'

        if should_flag:
            # Generate explanation
            explanation = generate_a2_explanation(
                loan_change=loan_change,
                interest_change=interest_change,
                current_rate=current_rate,
                previous_rate=previous_rate,
                rate_change_date=rate_change_date,
                expected_interest=expected_interest,
                actual_interest=total_interest_curr
            )

            # Format rate for display
            def fmt_rate(r):
                return f"{r*100:.2f}%" if r else "N/A"

            flag = {
                'Rule_ID': 'A2',
                'Priority': flag_priority,
                'Issue': flag_issue,
                'Accounts': '341xxx ↔ 635xxx + 241xxx',
                'Period': f"{prev_month} → {curr_month}",
                'Loan_Change': f'{loan_change:,.0f}',
                'Interest_Change': f'{interest_change:,.0f}',
                'Current_Rate': fmt_rate(current_rate),
                'Previous_Rate': fmt_rate(previous_rate) if previous_rate else 'N/A',
                'Rate_Effective_Date': rate_change_date.strftime('%Y-%m-%d') if rate_change_date else 'N/A',
                'Expected_Interest': f'{expected_interest:,.0f}' if expected_interest else 'N/A',
                'Actual_Interest': f'{total_interest_curr:,.0f}',
                'Variance': f'{(total_interest_curr - expected_interest):,.0f}' if expected_interest else 'N/A',
                'Reason': explanation,
                'Flag_Trigger': f'Enhanced A2 analysis with rate data'
            }

            # Add entity info if available
            if matched_entity:
                flag['Entity_Matched'] = matched_entity

            flags.append(flag)

    return flags


def check_rule_A3(bs_df, pl_df):
    """A3 - Capex incurred but VAT not recorded"""
    flags = []
    if bs_df is None or bs_df.empty:
        return flags

    ip_accounts = get_account_pattern_data(bs_df, '217xxx')
    cip_accounts = get_account_pattern_data(bs_df, '241xxx')
    vat_accounts = get_account_pattern_data(bs_df, '133xxx')
    if (ip_accounts.empty and cip_accounts.empty) or vat_accounts.empty:
        return flags

    month_cols = get_month_cols(bs_df)
    if len(month_cols) < 2:
        return flags

    for i in range(1, len(month_cols)):
        prev_month = month_cols[i-1]
        curr_month = month_cols[i]

        ip_change = ip_accounts[curr_month].sum() - ip_accounts[prev_month].sum() if not ip_accounts.empty else 0
        cip_change = cip_accounts[curr_month].sum() - cip_accounts[prev_month].sum() if not cip_accounts.empty else 0
        total_asset_change = ip_change + cip_change
        vat_change = vat_accounts[curr_month].sum() - vat_accounts[prev_month].sum()

        if total_asset_change > 0 and vat_change <= 0:
            flags.append({
                'Rule_ID': 'A3',
                'Priority': '🔴 Critical',
                'Issue': 'Capex incurred but VAT not recorded',
                'Accounts': '217/241 ↔ 133xxx',
                'Period': f"{prev_month} → {curr_month}",
                'Reason': f'Assets (IP+CIP) increased by {total_asset_change:,.0f} VND but VAT input changed by {vat_change:,.0f} VND',
                'Flag_Trigger': 'Assets↑ BUT VAT input ≤ previous'
            })
    return flags


def check_rule_A4(bs_df, pl_df):
    """A4 - Cash movement disconnected from interest"""
    flags = []
    if bs_df is None or pl_df is None or bs_df.empty or pl_df.empty:
        return flags

    cash_111 = get_account_pattern_data(bs_df, '111xxx')
    cash_112 = get_account_pattern_data(bs_df, '112xxx')
    cash_accounts = pd.concat([cash_111, cash_112], ignore_index=True)
    interest_income = get_account_pattern_data(pl_df, '515xxx')
    if cash_accounts.empty or interest_income.empty:
        return flags

    month_cols = get_month_cols(bs_df)
    if len(month_cols) < 2:
        return flags

    for i in range(1, len(month_cols)):
        prev_month = month_cols[i-1]
        curr_month = month_cols[i]

        prev_month_name = extract_month_from_column(prev_month)
        curr_month_name = extract_month_from_column(curr_month)

        cash_change = cash_accounts[curr_month].sum() - cash_accounts[prev_month].sum()

        interest_prev_adj = normalize_interest_for_calendar(interest_income[prev_month].sum(), prev_month_name)
        interest_curr_adj = normalize_interest_for_calendar(interest_income[curr_month].sum(), curr_month_name)
        interest_change = interest_curr_adj - interest_prev_adj

        if (cash_change > 0 and interest_change < 0) or (cash_change < 0 and interest_change > 0):
            flags.append({
                'Rule_ID': 'A4',
                'Priority': '🔴 Critical',
                'Issue': 'Cash movement disconnected from interest',
                'Accounts': '111/112 ↔ 515xxx',
                'Period': f"{prev_month} → {curr_month}",
                'Reason': f'Cash changed by {cash_change:,.0f} VND but day-adjusted interest changed oppositely by {interest_change:,.0f} VND',
                'Flag_Trigger': 'Cash↑ BUT Interest↓ OR Cash↓ BUT Interest↑'
            })
    return flags


def check_rule_A5(bs_df, pl_df):
    """A5 - Lease termination but broker asset not written off"""
    flags = []
    if bs_df is None or pl_df is None or bs_df.empty or pl_df.empty:
        return flags

    revenue_accounts = get_account_pattern_data(pl_df, '511xxx')
    broker_accounts = get_account_pattern_data(bs_df, '242xxx')
    selling_expense = get_account_pattern_data(pl_df, '641xxx')
    if revenue_accounts.empty or broker_accounts.empty:
        return flags

    month_cols = get_month_cols(pl_df)
    if len(month_cols) < 1:
        return flags

    for i, month in enumerate(month_cols):
        revenue_total = revenue_accounts[month].sum()

        broker_prev_month_idx = max(0, i-1)
        broker_change = broker_accounts[month].sum() - broker_accounts[month_cols[broker_prev_month_idx]].sum() if not broker_accounts.empty else 0
        selling_change = selling_expense[month].sum() - selling_expense[month_cols[broker_prev_month_idx]].sum() if not selling_expense.empty else 0

        if revenue_total <= 0 and broker_change == 0 and selling_change == 0:
            flags.append({
                'Rule_ID': 'A5',
                'Priority': '🔴 Critical',
                'Issue': 'Lease termination but broker asset not written off',
                'Accounts': '511xxx ↔ 242xxx ↔ 641xxx',
                'Period': month,
                'Reason': f'Revenue is {revenue_total:,.0f} VND but broker asset (242) and selling expense (641) unchanged',
                'Flag_Trigger': 'Revenue ≤ 0 BUT 242 unchanged AND 641 unchanged'
            })
    return flags


def check_rule_A7(bs_df, pl_df):
    """A7 - Asset disposal but accumulated depreciation not written off"""
    flags = []
    if bs_df is None or bs_df.empty:
        return flags

    ip_cost = bs_df[bs_df['Account Name'].astype(str).str.contains('cost|historical', na=False, regex=True, case=False)]
    ip_cost = ip_cost[ip_cost['Account Code'].astype(str).str.startswith('217')]
    ip_accum_dep = bs_df[bs_df['Account Name'].astype(str).str.contains('accum|depreciation', na=False, regex=True, case=False)]
    ip_accum_dep = ip_accum_dep[ip_accum_dep['Account Code'].astype(str).str.startswith('217')]

    if ip_cost.empty or ip_accum_dep.empty:
        return flags

    month_cols = get_month_cols(bs_df)
    if len(month_cols) < 2:
        return flags

    for i in range(1, len(month_cols)):
        prev_month = month_cols[i-1]
        curr_month = month_cols[i]

        cost_change = ip_cost[curr_month].sum() - ip_cost[prev_month].sum()
        accum_dep_change = ip_accum_dep[curr_month].sum() - ip_accum_dep[prev_month].sum()

        if cost_change < 0 and accum_dep_change >= 0:
            flags.append({
                'Rule_ID': 'A7',
                'Priority': '🔴 Critical',
                'Issue': 'Asset disposal but accumulated depreciation not written off',
                'Accounts': '217xxx (cost) ↔ 217xxx (accum dep)',
                'Period': f"{prev_month} → {curr_month}",
                'Reason': f'IP cost decreased by {abs(cost_change):,.0f} VND but accumulated depreciation changed by {accum_dep_change:,.0f} VND',
                'Flag_Trigger': 'IP cost↓ BUT Accumulated depreciation unchanged'
            })
    return flags


def check_rule_D1(bs_df, pl_df):
    """D1 - Balance sheet imbalance

    CORRECT Formula: Total Assets = Total Liabilities + Equity

    For consolidated files with inter-company eliminations, we use the total rows.
    For single-entity files, we sum by account line category (1xx, 2xx, 3xx, 4xx).

    Vietnamese Chart of Accounts:
    - 1xx: Assets (add)
    - 2xx: Liabilities (add)
    - 3xx: Contra-Liabilities (subtract from right side)
    - 4xx: Equity (subtract from right side)

    Balance Equation: Assets + Liabilities = Contra-Liabilities + Equity
    """
    flags = []
    if bs_df is None or bs_df.empty:
        return flags

    month_cols = get_month_cols(bs_df)
    if not month_cols:
        return flags

    # Create a copy to avoid SettingWithCopyWarning
    bs_df_copy = bs_df.copy()

    # NEW APPROACH: Classify accounts by SECTION, not by account code first digit
    # Reason: Some files (like BNH) have 2xx codes in Assets section (e.g., 231 Fixed Assets)
    # and 1xx codes in Liabilities section (e.g., 136 Unearned Revenue)
    #
    # Strategy: Find the "TỔNG CỘNG TÀI SẢN" row (Total Assets) to split sections
    # - Everything BEFORE this row with Account Line = Assets
    # - Everything AFTER this row with Account Line = Liabilities+Equity

    # Find the Total Assets row
    print(f"DEBUG: D1 - Searching for Total Assets marker in {len(bs_df_copy)} rows")
    print(f"DEBUG: D1 - Columns: {bs_df_copy.columns.tolist()}")
    print(f"DEBUG: D1 - Sample Account Code values: {bs_df_copy['Account Code'].head(10).tolist()}")

    total_assets_marker_rows = bs_df_copy[
        (bs_df_copy['Account Code'].astype(str).str.contains(
            r'TỔNG CỘNG TÀI SẢN$|^Total Assets\s*$',
            case=False,
            na=False,
            regex=True
        )) &
        (bs_df_copy['Account Line'].isna())  # Total row has no Account Line
    ]

    print(f"DEBUG: D1 - Found {len(total_assets_marker_rows)} Total Assets markers")

    if not total_assets_marker_rows.empty:
        # Use TOTAL ROWS approach - directly compare the total row values
        # This is more reliable than summing account lines
        print(f"DEBUG: D1 - Using total rows approach (found Total Assets marker)")

        total_assets_row = total_assets_marker_rows.iloc[-1]

        # Find Total Liabilities+Equity marker
        total_liab_eq_marker_rows = bs_df_copy[
            (bs_df_copy['Account Code'].astype(str).str.contains(
                r'TỔNG CỘNG NGUỒN VỐN|Total.*Liabilities.*Equity|Total.*Equity',
                case=False,
                na=False,
                regex=True
            )) &
            (bs_df_copy['Account Line'].isna())
        ]

        if not total_liab_eq_marker_rows.empty:
            total_liab_eq_row = total_liab_eq_marker_rows.iloc[-1]

            print(f"DEBUG: D1 - Found both total rows, comparing their values")

            for month in month_cols:
                total_assets = total_assets_row[month] if pd.notna(total_assets_row[month]) else 0
                total_liab_equity = total_liab_eq_row[month] if pd.notna(total_liab_eq_row[month]) else 0

                print(f"DEBUG: D1 - Month {month}: Assets={total_assets:,.0f}, Liab+Eq={total_liab_equity:,.0f}")

                balance = total_assets - total_liab_equity

                # Tolerance: 100M VND
                if abs(balance) > 100_000_000:
                    flags.append({
                        'Rule_ID': 'D1',
                        'Priority': '🔴 Critical',
                        'Issue': 'Balance sheet imbalance',
                        'Accounts': 'Total Assets vs Total Liabilities+Equity',
                        'Period': month,
                        'Total_Assets': f'{total_assets:,.0f}',
                        'Total_Liabilities_Equity': f'{total_liab_equity:,.0f}',
                        'Balance': f'{balance:,.0f}',
                        'Reason': f'Total Assets - Total Liabilities+Equity = {balance:,.0f} VND (should be ~0). Assets={total_assets:,.0f}, Liab+Equity={total_liab_equity:,.0f}',
                        'Flag_Trigger': 'Total Assets ≠ Total Liabilities+Equity'
                    })
        else:
            print(f"DEBUG: D1 - Could not find Total Liabilities+Equity marker, falling back to account code approach")
    else:
        # Fallback to traditional account code approach if no Total Assets marker found
        print(f"DEBUG: D1 - Using traditional account code approach (no Total Assets marker found)")

        bs_detail = bs_df_copy[bs_df_copy['Account Line'].notna()].copy()

        if bs_detail.empty:
            print(f"DEBUG: D1 - No detail rows with Account Line found")
            return flags

        bs_detail['Account_Line_Str'] = bs_detail['Account Line'].astype(str).str.strip()
        bs_detail['First_Digit'] = bs_detail['Account_Line_Str'].str[0]

        type_1 = bs_detail[bs_detail['First_Digit'] == '1']  # Assets (1xx)
        type_2 = bs_detail[bs_detail['First_Digit'] == '2']  # Liabilities (2xx)
        type_3 = bs_detail[bs_detail['First_Digit'] == '3']  # Contra-Liabilities (3xx)
        type_4 = bs_detail[bs_detail['First_Digit'] == '4']  # Equity (4xx)

        for month in month_cols:
            total_1 = type_1[month].sum() if not type_1.empty else 0
            total_2 = type_2[month].sum() if not type_2.empty else 0
            total_3 = type_3[month].sum() if not type_3.empty else 0
            total_4 = type_4[month].sum() if not type_4.empty else 0

            # CORRECT FORMULA: 1 + 2 = 3 + 4, rearranged as 1 + 2 - 3 - 4 = 0
            balance = total_1 + total_2 - total_3 - total_4

            # Tolerance: 100M VND
            if abs(balance) > 100_000_000:
                flags.append({
                    'Rule_ID': 'D1',
                    'Priority': '🔴 Critical',
                    'Issue': 'Balance sheet imbalance',
                    'Accounts': 'Account Lines: 1xx + 2xx = 3xx + 4xx',
                    'Period': month,
                    'Type_1_Assets': f'{total_1:,.0f}',
                    'Type_2_Liabilities': f'{total_2:,.0f}',
                    'Type_3_Contra_Liabilities': f'{total_3:,.0f}',
                    'Type_4_Equity': f'{total_4:,.0f}',
                    'Balance': f'{balance:,.0f}',
                    'Reason': f'Balance sheet formula (1+2-3-4) = {balance:,.0f} VND (should be ~0). Assets={total_1:,.0f}, Liabilities={total_2:,.0f}, Contra-Liabilities={total_3:,.0f}, Equity={total_4:,.0f}',
                    'Flag_Trigger': '(1 + 2) - (3 + 4) ≠ 0'
                })

    return flags


def check_rule_E1(bs_df, pl_df):
    """E1 - Negative Net Book Value (NBV)

    Checks that Cost > Accumulated Depreciation (NBV > 0) for:
    - Account Line 222 (cost) > 223 (accumulated depreciation)
    - Account Line 228 (cost) > 229 (accumulated depreciation)
    - Account Line 231 (cost) > 232 (accumulated depreciation)
    """
    flags = []
    if bs_df is None or bs_df.empty:
        return flags

    month_cols = get_month_cols(bs_df)

    # Define cost vs accumulated depreciation pairs based on Account Line
    nbv_pairs = [
        ('222', '223', 'Asset Class 222/223'),
        ('228', '229', 'Asset Class 228/229'),
        ('231', '232', 'Asset Class 231/232')
    ]

    for cost_line, accum_dep_line, asset_class in nbv_pairs:
        # Get accounts by Account Line (not Account Code)
        cost_accounts = bs_df[bs_df['Account Line'].astype(str) == cost_line]
        accum_dep_accounts = bs_df[bs_df['Account Line'].astype(str) == accum_dep_line]

        if cost_accounts.empty and accum_dep_accounts.empty:
            continue

        for month in month_cols:
            cost_total = cost_accounts[month].sum() if not cost_accounts.empty else 0
            accum_dep_total = accum_dep_accounts[month].sum() if not accum_dep_accounts.empty else 0

            # NBV = Cost - Accumulated Depreciation (note: accum_dep is typically negative)
            nbv = cost_total + accum_dep_total  # Adding because accum_dep is negative

            # Flag if NBV is negative (cost < abs(accumulated depreciation))
            if nbv < 0:
                flags.append({
                    'Rule_ID': 'E1',
                    'Priority': '🔴 Critical',
                    'Issue': 'Negative Net Book Value (NBV)',
                    'Accounts': f'Account Line {cost_line} (cost) vs {accum_dep_line} (accum dep)',
                    'Period': month,
                    'Cost': f'{cost_total:,.0f}',
                    'Accumulated_Depreciation': f'{accum_dep_total:,.0f}',
                    'NBV': f'{nbv:,.0f}',
                    'Reason': f'{asset_class}: Cost={cost_total:,.0f} VND, Accum Dep={accum_dep_total:,.0f} VND, NBV={nbv:,.0f} VND (should be > 0)',
                    'Flag_Trigger': f'NBV ({cost_line} + {accum_dep_line}) < 0'
                })

    return flags


# ============================================================================
# RULE IMPLEMENTATIONS - REVIEW (🟡)
# ============================================================================

def check_rule_B1(bs_df, pl_df):
    """B1 - Rental revenue volatility"""
    flags = []
    if pl_df is None or pl_df.empty:
        return flags

    rental_revenue = pl_df[pl_df['Account Code'].astype(str).str.contains('511710001', na=False)]
    if rental_revenue.empty:
        return flags

    month_cols = get_month_cols(pl_df)
    if len(month_cols) < MONTHS_6M_AVG:
        return flags

    for _, row in rental_revenue.iterrows():
        values = [row[month] for month in month_cols]
        series = pd.Series(values, index=month_cols)

        for i in range(MONTHS_6M_AVG, len(month_cols)):
            curr_month = month_cols[i]
            curr_value = series.iloc[i]
            baseline_window = series.iloc[i-MONTHS_6M_AVG:i]
            mean_6m = baseline_window.mean()
            std_6m = baseline_window.std()
            threshold = std_6m * STD_DEV_THRESHOLD
            deviation = abs(curr_value - mean_6m)

            if deviation > threshold and threshold > 0:
                flags.append({
                    'Rule_ID': 'B1',
                    'Priority': '🟡 Review',
                    'Issue': 'Rental revenue volatility',
                    'Accounts': '511710001',
                    'Period': curr_month,
                    'Reason': f'Rental revenue {curr_value:,.0f} VND deviates by {deviation:,.0f} VND from 6M average {mean_6m:,.0f} VND',
                    'Flag_Trigger': 'abs(Current - Avg) > 2σ'
                })
    return flags


def check_rule_B2(bs_df, pl_df):
    """B2 - Depreciation changes without asset movement"""
    flags = []
    if bs_df is None or pl_df is None or bs_df.empty or pl_df.empty:
        return flags

    depreciation = pl_df[pl_df['Account Code'].astype(str).str.contains('632100002', na=False)]
    ip_accounts = get_account_pattern_data(bs_df, '217xxx')
    if depreciation.empty or ip_accounts.empty:
        return flags

    month_cols = get_month_cols(pl_df)
    if len(month_cols) < MONTHS_6M_AVG + 1:
        return flags

    dep_values = [depreciation[month].sum() for month in month_cols]
    dep_series = pd.Series(dep_values, index=month_cols)
    ip_values = [ip_accounts[month].sum() for month in month_cols]
    ip_series = pd.Series(ip_values, index=month_cols)

    for i in range(MONTHS_6M_AVG, len(month_cols)):
        curr_month = month_cols[i]
        curr_dep = dep_series.iloc[i]
        baseline_window = dep_series.iloc[i-MONTHS_6M_AVG:i]
        mean_dep = baseline_window.mean()
        std_dep = baseline_window.std()
        threshold = std_dep * STD_DEV_THRESHOLD
        deviation = abs(curr_dep - mean_dep)
        ip_change = ip_series.iloc[i] - ip_series.iloc[i-1]

        if deviation > threshold and threshold > 0 and ip_change <= 0:
            flags.append({
                'Rule_ID': 'B2',
                'Priority': '🟡 Review',
                'Issue': 'Depreciation changes without asset movement',
                'Accounts': '632100002 + 217xxx',
                'Period': curr_month,
                'Reason': f'Depreciation {curr_dep:,.0f} VND deviates by {deviation:,.0f} VND but IP unchanged',
                'Flag_Trigger': 'Depreciation deviates > 2σ AND IP unchanged'
            })
    return flags


def check_rule_B3(bs_df, pl_df):
    """B3 - Amortization changes"""
    flags = []
    if pl_df is None or pl_df.empty:
        return flags

    amortization = pl_df[pl_df['Account Code'].astype(str).str.contains('632100001', na=False)]
    if amortization.empty:
        return flags

    month_cols = get_month_cols(pl_df)
    if len(month_cols) < MONTHS_6M_AVG:
        return flags

    for _, row in amortization.iterrows():
        values = [row[month] for month in month_cols]
        series = pd.Series(values, index=month_cols)

        for i in range(MONTHS_6M_AVG, len(month_cols)):
            curr_month = month_cols[i]
            curr_value = series.iloc[i]
            baseline_window = series.iloc[i-MONTHS_6M_AVG:i]
            mean = baseline_window.mean()
            std = baseline_window.std()
            threshold = std * STD_DEV_THRESHOLD
            deviation = abs(curr_value - mean)

            if deviation > threshold and threshold > 0:
                flags.append({
                    'Rule_ID': 'B3',
                    'Priority': '🟡 Review',
                    'Issue': 'Amortization changes',
                    'Accounts': '632100001',
                    'Period': curr_month,
                    'Reason': f'Amortization {curr_value:,.0f} VND deviates by {deviation:,.0f} VND from 6M average',
                    'Flag_Trigger': 'abs(Current - Avg) > 2σ'
                })
    return flags


def check_rule_C1(bs_df, pl_df):
    """C1 - Gross margin by revenue stream

    Checks gross profit margin for rechargeable services:
    - Utilities revenue vs expenses
    - Service charges vs direct expenses
    - Others revenue vs other expenses

    Revenue streams:
    - 511800001: Revenue: Utilities
    - 511800002: Revenue: Others
    - 511600001: Revenue: Service charge - Third parties
    - 511600005: Revenue: Operation management fee (Share-mgt services charge)

    Direct expenses (should be recharged to clients):
    - 632100008: Direct Expenses: Repair & Maintenance - Tenant's request
    - 632100011: Direct Expenses: Free factories' utility expenses
    - 632100015: Direct Expenses: Construction requested by tenants
    - 632199999: Direct Expenses: Others

    Note: Leasing revenue (rental income) vs depreciation/amortization is IGNORED
    """
    flags = []
    if pl_df is None or pl_df.empty:
        return flags

    month_cols = get_month_cols(pl_df)
    if len(month_cols) < 2:
        return flags

    # Define revenue stream groups (excluding rental/leasing revenue)
    revenue_stream_groups = [
        {
            'name': 'Utilities',
            'revenue_accounts': ['511800001'],
            'expense_accounts': ['632100011'],  # Free factories' utility expenses
            'description': 'Utilities Revenue vs Utility Expenses'
        },
        {
            'name': 'Service Charges',
            'revenue_accounts': ['511600001', '511600005'],  # Service charge + Operation management fee
            'expense_accounts': ['632100008', '632100015'],  # Repair & Maintenance + Construction
            'description': 'Service Charges vs Direct Service Expenses'
        },
        {
            'name': 'Other Revenue',
            'revenue_accounts': ['511800002'],  # Revenue: Others
            'expense_accounts': ['632199999'],  # Direct Expenses: Others
            'description': 'Other Revenue vs Other Direct Expenses'
        }
    ]

    for stream in revenue_stream_groups:
        # Get revenue accounts for this stream
        revenue_data = None
        for acc in stream['revenue_accounts']:
            acc_data = pl_df[pl_df['Account Code'].astype(str).str.contains(acc, na=False)]
            if not acc_data.empty:
                if revenue_data is None:
                    revenue_data = acc_data
                else:
                    revenue_data = pd.concat([revenue_data, acc_data])

        # Get expense accounts for this stream
        expense_data = None
        for acc in stream['expense_accounts']:
            acc_data = pl_df[pl_df['Account Code'].astype(str).str.contains(acc, na=False)]
            if not acc_data.empty:
                if expense_data is None:
                    expense_data = acc_data
                else:
                    expense_data = pd.concat([expense_data, acc_data])

        if revenue_data is None or revenue_data.empty:
            continue

        # Need at least 6 months of data to establish baseline
        if len(month_cols) < MONTHS_6M_AVG:
            continue

        # Calculate gross margin percentage for each month
        gm_percentages = []
        for month in month_cols:
            revenue = revenue_data[month].sum() if revenue_data is not None else 0
            expense = expense_data[month].sum() if expense_data is not None else 0

            # Calculate gross margin percentage
            if revenue != 0:
                gm_pct = ((revenue - expense) / revenue * 100)
            else:
                gm_pct = 0

            gm_percentages.append({
                'month': month,
                'revenue': revenue,
                'expense': expense,
                'gm_pct': gm_pct
            })

        # Compare current months against 6-month baseline
        for i in range(MONTHS_6M_AVG, len(gm_percentages)):
            curr_data = gm_percentages[i]
            baseline_window = [gm_percentages[j]['gm_pct'] for j in range(i - MONTHS_6M_AVG, i)]

            # Skip if no revenue in current month
            if curr_data['revenue'] == 0:
                continue

            mean_gm = pd.Series(baseline_window).mean()
            std_gm = pd.Series(baseline_window).std()
            threshold = std_gm * STD_DEV_THRESHOLD

            # Calculate change in gross margin
            gm_change = curr_data['gm_pct'] - mean_gm

            # Flag if gross margin changes significantly (either drop OR increase)
            if abs(gm_change) > threshold and threshold > 0:
                direction = "dropped" if gm_change < 0 else "increased"

                flags.append({
                    'Rule_ID': 'C1',
                    'Priority': '🟡 Review',
                    'Issue': f'Gross margin {direction} for {stream["name"]}',
                    'Accounts': f'{stream["description"]}',
                    'Period': curr_data['month'],
                    'Revenue': f'{curr_data["revenue"]:,.0f}',
                    'Expenses': f'{curr_data["expense"]:,.0f}',
                    'Current_GM': f'{curr_data["gm_pct"]:.2f}%',
                    'Baseline_GM': f'{mean_gm:.2f}%',
                    'GM_Change': f'{gm_change:+.2f}%',
                    'Reason': f'{stream["name"]}: GM {direction} to {curr_data["gm_pct"]:.2f}% from baseline {mean_gm:.2f}% (Revenue={curr_data["revenue"]:,.0f}, Expenses={curr_data["expense"]:,.0f})',
                    'Flag_Trigger': f'GM% change > 2σ ({abs(gm_change):.2f}% vs threshold {threshold:.2f}%)'
                })

    return flags


def check_rule_C2(bs_df, pl_df):
    """C2 - Unbilled reimbursable expenses"""
    flags = []
    if pl_df is None or pl_df.empty:
        return flags

    cogs_641 = get_account_pattern_data(pl_df, '641xxx')
    cogs_632 = get_account_pattern_data(pl_df, '632xxx')
    reimbursable_cogs = pd.concat([cogs_641, cogs_632], ignore_index=True)
    revenue_accounts = get_account_pattern_data(pl_df, '511xxx')
    if reimbursable_cogs.empty or revenue_accounts.empty:
        return flags

    month_cols = get_month_cols(pl_df)
    if len(month_cols) < 2:
        return flags

    for i in range(1, len(month_cols)):
        prev_month = month_cols[i-1]
        curr_month = month_cols[i]

        cogs_change = reimbursable_cogs[curr_month].sum() - reimbursable_cogs[prev_month].sum()
        revenue_change = revenue_accounts[curr_month].sum() - revenue_accounts[prev_month].sum()

        if cogs_change > 0 and revenue_change <= 0:
            flags.append({
                'Rule_ID': 'C2',
                'Priority': '🟡 Review',
                'Issue': 'Unbilled reimbursable expenses',
                'Accounts': '641xxx/632xxx ↔ 511xxx',
                'Period': f"{prev_month} → {curr_month}",
                'Reason': f'Reimbursable COGS increased by {cogs_change:,.0f} VND but revenue changed by {revenue_change:,.0f} VND',
                'Flag_Trigger': 'Reimbursable COGS↑ BUT Revenue unchanged'
            })
    return flags


def check_rule_D2(bs_df, pl_df):
    """D2 - Retained earnings reconciliation break"""
    flags = []
    if bs_df is None or pl_df is None or bs_df.empty or pl_df.empty:
        return flags

    re_421a = bs_df[bs_df['Account Line'].astype(str) == '421']
    re_421b = bs_df[bs_df['Account Line'].astype(str) == '4211']

    if re_421a.empty:
        return flags

    bs_month_cols = get_month_cols(bs_df)
    pl_month_cols = get_month_cols(pl_df)

    if len(bs_month_cols) < 2:
        return flags

    pl_components = {}
    for line in ['1', '11', '21', '22', '23', '25', '26', '31', '32', '51', '52']:
        pl_components[line] = pl_df[pl_df['Account Line'].astype(str) == line]

    for i in range(1, len(bs_month_cols)):
        prev_month = bs_month_cols[i-1]
        curr_month = bs_month_cols[i]

        opening_re_421a = re_421a[prev_month].sum() if not re_421a.empty else 0
        opening_re_421b = re_421b[prev_month].sum() if not re_421b.empty else 0
        opening_re = opening_re_421a + opening_re_421b

        closing_re_421a = re_421a[curr_month].sum() if not re_421a.empty else 0
        closing_re_421b = re_421b[curr_month].sum() if not re_421b.empty else 0
        closing_re = closing_re_421a + closing_re_421b

        if curr_month not in pl_month_cols:
            continue

        pl_1 = pl_components['1'][curr_month].sum() if not pl_components['1'].empty else 0
        pl_11 = pl_components['11'][curr_month].sum() if not pl_components['11'].empty else 0
        pl_21 = pl_components['21'][curr_month].sum() if not pl_components['21'].empty else 0
        pl_22 = pl_components['22'][curr_month].sum() if not pl_components['22'].empty else 0
        pl_23 = pl_components['23'][curr_month].sum() if not pl_components['23'].empty else 0
        pl_25 = pl_components['25'][curr_month].sum() if not pl_components['25'].empty else 0
        pl_26 = pl_components['26'][curr_month].sum() if not pl_components['26'].empty else 0
        pl_31 = pl_components['31'][curr_month].sum() if not pl_components['31'].empty else 0
        pl_32 = pl_components['32'][curr_month].sum() if not pl_components['32'].empty else 0
        pl_51 = pl_components['51'][curr_month].sum() if not pl_components['51'].empty else 0
        pl_52 = pl_components['52'][curr_month].sum() if not pl_components['52'].empty else 0

        profit_loss = pl_1 - pl_11 + pl_21 - pl_23 - pl_22 - pl_25 - pl_26 + pl_31 - pl_32 - pl_51 - pl_52
        expected_closing_re = opening_re + profit_loss
        difference = abs(closing_re - expected_closing_re)

        if difference > MATERIALITY_THRESHOLDS['retained_earnings_tolerance']:
            flags.append({
                'Rule_ID': 'D2',
                'Priority': '🟡 Review',
                'Issue': 'Retained earnings reconciliation break',
                'Accounts': 'Account Lines: 421a+421b (BS) and P/L components',
                'Period': f"{prev_month} → {curr_month}",
                'Reason': f'Closing RE {closing_re:,.0f} VND differs from expected {expected_closing_re:,.0f} VND by {difference:,.0f} VND',
                'Flag_Trigger': 'abs(Closing RE - (Opening RE + P/L)) > 100M'
            })
    return flags


def check_rule_E2(bs_df, pl_df):
    """E2 - CIP stagnation"""
    flags = []
    if bs_df is None or bs_df.empty:
        return flags

    cip_accounts = get_account_pattern_data(bs_df, '241xxx')
    ip_accounts = get_account_pattern_data(bs_df, '217xxx')
    if cip_accounts.empty or ip_accounts.empty:
        return flags

    month_cols = get_month_cols(bs_df)
    if len(month_cols) < 4:
        return flags

    for i in range(3, len(month_cols)):
        curr_month = month_cols[i]
        months_to_check = month_cols[i-3:i+1]
        cip_values = [cip_accounts[month].sum() for month in months_to_check]

        if len(set(cip_values)) == 1 and cip_values[0] > 0:
            ip_change = ip_accounts[months_to_check[-1]].sum() - ip_accounts[months_to_check[0]].sum()

            if ip_change <= 0:
                flags.append({
                    'Rule_ID': 'E2',
                    'Priority': '🟡 Review',
                    'Issue': 'CIP stagnation',
                    'Accounts': '241xxx (CIP) ↔ 217xxx (IP)',
                    'Period': f"{months_to_check[0]} → {curr_month}",
                    'Reason': f'CIP unchanged at {cip_values[0]:,.0f} VND for {len(months_to_check)} months',
                    'Flag_Trigger': 'CIP flat >3 months AND IP unchanged'
                })
                break
    return flags


def check_rule_E5(bs_df, pl_df):
    """E5 - Trade Receivables aging"""
    flags = []
    if bs_df is None or pl_df is None or bs_df.empty or pl_df.empty:
        return flags

    ar_accounts = get_account_pattern_data(bs_df, '131xxx')
    revenue_accounts = get_account_pattern_data(pl_df, '511xxx')
    if ar_accounts.empty or revenue_accounts.empty:
        return flags

    month_cols = get_month_cols(bs_df)
    if len(month_cols) < 4:
        return flags

    for i in range(3, len(month_cols)):
        start_month = month_cols[i-3]
        curr_month = month_cols[i]

        ar_start = ar_accounts[start_month].sum()
        ar_end = ar_accounts[curr_month].sum()
        ar_growth_pct = ((ar_end - ar_start) / ar_start * 100) if ar_start != 0 else 0

        revenue_start = revenue_accounts[start_month].sum()
        revenue_end = revenue_accounts[curr_month].sum()
        revenue_growth_pct = ((revenue_end - revenue_start) / revenue_start * 100) if revenue_start != 0 else 0

        if ar_growth_pct > (MATERIALITY_THRESHOLDS['ar_growth_threshold'] * 100) and revenue_growth_pct <= 0:
            flags.append({
                'Rule_ID': 'E5',
                'Priority': '🟡 Review',
                'Issue': 'Trade Receivables aging',
                'Accounts': '131xxx (AR) ↔ 511xxx (Revenue)',
                'Period': f"{start_month} → {curr_month}",
                'Reason': f'AR grew {ar_growth_pct:.1f}% over 3 months but revenue changed by {revenue_growth_pct:.1f}%',
                'Flag_Trigger': 'AR growth >> Revenue growth over 3 months'
            })
    return flags


def check_rule_E6(bs_df, pl_df):
    """E6 - Accounts Payable compression"""
    flags = []
    if bs_df is None or pl_df is None or bs_df.empty or pl_df.empty:
        return flags

    ap_accounts = get_account_pattern_data(bs_df, '331xxx')
    cogs_accounts = get_account_pattern_data(pl_df, '632xxx')
    if ap_accounts.empty or cogs_accounts.empty:
        return flags

    month_cols = get_month_cols(bs_df)
    if len(month_cols) < 2:
        return flags

    for i in range(1, len(month_cols)):
        prev_month = month_cols[i-1]
        curr_month = month_cols[i]

        ap_change = ap_accounts[curr_month].sum() - ap_accounts[prev_month].sum()
        cogs_change = cogs_accounts[curr_month].sum() - cogs_accounts[prev_month].sum()

        if ap_change < 0 and cogs_change >= 0:
            flags.append({
                'Rule_ID': 'E6',
                'Priority': '🟡 Review',
                'Issue': 'Accounts Payable compression',
                'Accounts': '331xxx (AP) ↔ 632xxx (COGS)',
                'Period': f"{prev_month} → {curr_month}",
                'Reason': f'AP decreased by {abs(ap_change):,.0f} VND but COGS changed by {cogs_change:,.0f} VND',
                'Flag_Trigger': 'AP↓ BUT COGS unchanged/increased'
            })
    return flags


# ============================================================================
# RULE IMPLEMENTATIONS - INFO (🟢)
# ============================================================================

def check_rule_E3(bs_df, pl_df):
    """E3 - Large VAT input - tax refund opportunity

    Checks specific account 133100002 (not all 133xxx accounts)
    """
    flags = []
    if bs_df is None or bs_df.empty:
        return flags

    # Check for specific account 133100002
    vat_accounts = get_account_pattern_data(bs_df, '133100002')
    if vat_accounts.empty:
        return flags

    month_cols = get_month_cols(bs_df)

    for month in month_cols:
        vat_balance = vat_accounts[month].sum()

        if vat_balance > MATERIALITY_THRESHOLDS['vat_input_threshold']:
            flags.append({
                'Rule_ID': 'E3',
                'Priority': '🟢 Info',
                'Issue': 'Large VAT input - tax refund opportunity',
                'Accounts': '133100002 (VAT input)',
                'Period': month,
                'Reason': f'VAT input balance is {vat_balance:,.0f} VND (threshold: 10B VND)',
                'Flag_Trigger': 'VAT input > 10B'
            })
    return flags


def check_rule_E4(bs_df, pl_df):
    """E4 - Interest payable quarterly pattern"""
    flags = []
    if bs_df is None or bs_df.empty:
        return flags

    interest_payable = get_account_pattern_data(bs_df, '335xxx')
    if interest_payable.empty:
        return flags

    month_cols = get_month_cols(bs_df)
    if len(month_cols) < 2:
        return flags

    for i in range(1, len(month_cols)):
        prev_month = month_cols[i-1]
        curr_month = month_cols[i]
        curr_month_name = extract_month_from_column(curr_month)

        if curr_month_name in QUARTER_START_MONTHS:
            interest_change = interest_payable[curr_month].sum() - interest_payable[prev_month].sum()

            if interest_change >= 0:
                flags.append({
                    'Rule_ID': 'E4',
                    'Priority': '🟢 Info',
                    'Issue': 'Interest payable quarterly pattern validation',
                    'Accounts': '335xxx (Interest payable)',
                    'Period': f"{prev_month} → {curr_month}",
                    'Reason': f'Quarter-start month {curr_month_name} but interest payable changed by {interest_change:,.0f} VND',
                    'Flag_Trigger': 'Q-start month BUT Interest payable unchanged/increased'
                })
    return flags


def check_rule_F1(bs_df, pl_df):
    """F1 - Trade receivables quarterly billing"""
    flags = []
    if bs_df is None or bs_df.empty:
        return flags

    ar_accounts = bs_df[bs_df['Account Code'].astype(str).str.contains('131100001', na=False)]
    if ar_accounts.empty:
        return flags

    month_cols = get_month_cols(bs_df)
    if len(month_cols) < 2:
        return flags

    for i in range(1, len(month_cols)):
        prev_month = month_cols[i-1]
        curr_month = month_cols[i]
        curr_month_name = extract_month_from_column(curr_month)

        if curr_month_name in QUARTER_START_MONTHS:
            ar_change = ar_accounts[curr_month].sum() - ar_accounts[prev_month].sum()

            if ar_change > 0:
                flags.append({
                    'Rule_ID': 'F1',
                    'Priority': '🟢 Info',
                    'Issue': 'Trade receivables - quarterly billing (normal pattern)',
                    'Accounts': '131100001 (Trade receivables)',
                    'Period': f"{prev_month} → {curr_month}",
                    'Reason': f'AR increased by {ar_change:,.0f} VND at quarter-start month {curr_month_name}',
                    'Flag_Trigger': 'Document expected pattern'
                })
    return flags


def check_rule_F2(bs_df, pl_df):
    """F2 - Unbilled revenue quarter-end peak"""
    flags = []
    if bs_df is None or bs_df.empty:
        return flags

    unbilled_revenue = bs_df[bs_df['Account Code'].astype(str).str.contains('138900003', na=False)]
    if unbilled_revenue.empty:
        return flags

    month_cols = get_month_cols(bs_df)
    quarter_end_months = ['Mar', 'Jun', 'Sep', 'Dec']
    if len(month_cols) < 2:
        return flags

    for i in range(1, len(month_cols)):
        prev_month = month_cols[i-1]
        curr_month = month_cols[i]
        curr_month_name = extract_month_from_column(curr_month)

        if curr_month_name in quarter_end_months:
            unbilled_change = unbilled_revenue[curr_month].sum() - unbilled_revenue[prev_month].sum()

            if unbilled_change > 0:
                flags.append({
                    'Rule_ID': 'F2',
                    'Priority': '🟢 Info',
                    'Issue': 'Unbilled revenue - quarter-end peak (normal pattern)',
                    'Accounts': '138900003 (Unbilled revenue)',
                    'Period': f"{prev_month} → {curr_month}",
                    'Reason': f'Unbilled revenue increased by {unbilled_change:,.0f} VND at quarter-end month {curr_month_name}',
                    'Flag_Trigger': 'Document expected pattern'
                })
    return flags


def check_rule_F3(bs_df, pl_df):
    """F3 - P&L line items with material month-over-month fluctuation

    Flags any P&L account that fluctuates month-over-month by:
    - More than 5% change (percentage threshold)
    - AND more than 1 billion VND (absolute threshold)
    """
    flags = []
    if pl_df is None or pl_df.empty:
        return flags

    month_cols = get_month_cols(pl_df)
    if len(month_cols) < 2:
        return flags

    # Thresholds
    PERCENTAGE_THRESHOLD = 0.05  # 5%
    ABSOLUTE_THRESHOLD = 1_000_000_000  # 1 billion VND

    # Check each P&L account
    for idx, row in pl_df.iterrows():
        account_code = row.get('Account Code', 'N/A')
        account_line = row.get('Account Line', 'N/A')

        # Compare each month with previous month
        for i in range(1, len(month_cols)):
            prev_month = month_cols[i-1]
            curr_month = month_cols[i]

            prev_value = row[prev_month]
            curr_value = row[curr_month]

            # Skip if previous value is zero (can't calculate percentage)
            if prev_value == 0:
                continue

            # Calculate absolute change and percentage change
            absolute_change = abs(curr_value - prev_value)
            percentage_change = abs((curr_value - prev_value) / prev_value)

            # Flag if BOTH thresholds are exceeded
            if absolute_change > ABSOLUTE_THRESHOLD and percentage_change > PERCENTAGE_THRESHOLD:
                direction = "increased" if curr_value > prev_value else "decreased"

                flags.append({
                    'Rule_ID': 'F3',
                    'Priority': '🟡 Review',
                    'Issue': f'Material P&L fluctuation - {account_line}',
                    'Accounts': f'{account_code} ({account_line})',
                    'Period': f"{prev_month} → {curr_month}",
                    'Previous_Value': f'{prev_value:,.0f}',
                    'Current_Value': f'{curr_value:,.0f}',
                    'Absolute_Change': f'{absolute_change:,.0f}',
                    'Percentage_Change': f'{percentage_change:.2%}',
                    'Reason': f'{account_code} {direction} from {prev_value:,.0f} to {curr_value:,.0f} VND (change: {absolute_change:,.0f} VND, {percentage_change:.2%})',
                    'Flag_Trigger': f'Change > 5% AND > 1B VND'
                })

    return flags


# ============================================================================
# MAIN RULE EXECUTION
# ============================================================================

def run_all_variance_rules(
    bs_df,
    pl_df,
    bs_df_full=None,
    entity_name: Optional[str] = None,
    loan_rate_lookup: Optional[Dict[str, List[Tuple[datetime, float]]]] = None
):
    """Run all 22 variance analysis rules

    Args:
        bs_df: Entity-specific BS data
        pl_df: Entity-specific PL data
        bs_df_full: Full BS data (not split by entity) for D1 balance check
        entity_name: Entity name for enhanced A2 lookup (optional)
        loan_rate_lookup: Dict mapping entity names to interest rate history (optional)
    """
    all_flags = []

    # Critical Rules (🔴)
    all_flags.extend(check_rule_A1(bs_df, pl_df))

    # Use enhanced A2 if loan rate data is available, otherwise use basic A2
    if loan_rate_lookup:
        all_flags.extend(check_rule_A2_enhanced(bs_df, pl_df, entity_name, loan_rate_lookup))
    else:
        all_flags.extend(check_rule_A2(bs_df, pl_df))

    all_flags.extend(check_rule_A3(bs_df, pl_df))
    all_flags.extend(check_rule_A4(bs_df, pl_df))
    all_flags.extend(check_rule_A5(bs_df, pl_df))
    all_flags.extend(check_rule_A7(bs_df, pl_df))
    # D1 uses full BS dataframe if available (for file-level balance check)
    all_flags.extend(check_rule_D1(bs_df_full if bs_df_full is not None and not bs_df_full.empty else bs_df, pl_df))
    all_flags.extend(check_rule_E1(bs_df, pl_df))

    # Review Rules (🟡)
    all_flags.extend(check_rule_B1(bs_df, pl_df))
    all_flags.extend(check_rule_B2(bs_df, pl_df))
    all_flags.extend(check_rule_B3(bs_df, pl_df))
    all_flags.extend(check_rule_C1(bs_df, pl_df))
    all_flags.extend(check_rule_C2(bs_df, pl_df))
    all_flags.extend(check_rule_D2(bs_df, pl_df))
    all_flags.extend(check_rule_E2(bs_df, pl_df))
    all_flags.extend(check_rule_E5(bs_df, pl_df))
    all_flags.extend(check_rule_E6(bs_df, pl_df))
    all_flags.extend(check_rule_F3(bs_df, pl_df))  # P&L material fluctuation

    # Info Rules (🟢)
    all_flags.extend(check_rule_E3(bs_df, pl_df))
    all_flags.extend(check_rule_E4(bs_df, pl_df))
    all_flags.extend(check_rule_F1(bs_df, pl_df))
    all_flags.extend(check_rule_F2(bs_df, pl_df))

    return all_flags


# ============================================================================
# OUTPUT FUNCTIONS
# ============================================================================

def create_excel_output(all_file_data, combined_flags):
    """
    Create Excel output with raw data sheets and flagged variances.

    Args:
        all_file_data: List of dicts with keys: 'subsidiary', 'filename', 'bs_df', 'pl_df', 'flags'
        combined_flags: Combined list of all flags from all files

    Returns:
        bytes: Excel file
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Sheet 1: Variance Flags (with separators between files)
        if not combined_flags:
            flags_df = pd.DataFrame([{
                'File': 'N/A',
                'Rule_ID': 'N/A',
                'Priority': 'N/A',
                'Issue': 'No variance flags detected',
                'Accounts': 'N/A',
                'Period': 'N/A',
                'Reason': 'All checks passed',
                'Flag_Trigger': 'N/A'
            }])
        else:
            flags_df = pd.DataFrame(combined_flags)

            # Define column order - include enhanced A2 columns if present
            base_columns = ['File', 'Rule_ID', 'Priority', 'Issue', 'Accounts', 'Period']
            enhanced_a2_columns = [
                'Loan_Change', 'Interest_Change', 'Current_Rate', 'Previous_Rate',
                'Rate_Effective_Date', 'Expected_Interest', 'Actual_Interest', 'Variance',
                'Entity_Matched'
            ]
            end_columns = ['Reason', 'Flag_Trigger']

            # Build column order based on what columns exist in the data
            column_order = base_columns.copy()
            for col in enhanced_a2_columns:
                if col in flags_df.columns:
                    column_order.append(col)
            column_order.extend(end_columns)

            # Only include columns that exist in the dataframe
            column_order = [col for col in column_order if col in flags_df.columns]
            flags_df = flags_df[column_order]

            priority_order = {'🔴 Critical': 0, '🟡 Review': 1, '🟢 Info': 2}
            flags_df['Priority_Rank'] = flags_df['Priority'].map(priority_order)
            flags_df = flags_df.sort_values(['File', 'Priority_Rank', 'Rule_ID']).drop('Priority_Rank', axis=1)

            # Add empty rows between different files (if multiple files)
            if len(all_file_data) > 1:
                # Group by File and insert empty rows
                grouped_dfs = []
                for file_name in flags_df['File'].unique():
                    file_group = flags_df[flags_df['File'] == file_name].copy()
                    grouped_dfs.append(file_group)
                    # Add empty separator row (except after last group)
                    if file_name != flags_df['File'].unique()[-1]:
                        empty_row = pd.DataFrame([{col: '' for col in flags_df.columns}])
                        grouped_dfs.append(empty_row)

                flags_df = pd.concat(grouped_dfs, ignore_index=True)

        flags_df.to_excel(writer, sheet_name='Variance Flags', index=False)

        # Format Variance Flags sheet
        workbook = writer.book
        worksheet = writer.sheets['Variance Flags']

        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        separator_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        # Priority-based row colors (light versions for readability)
        critical_fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')  # Light red
        review_fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')    # Light yellow
        info_fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')      # Light green

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Find the Priority column index
        priority_col_idx = None
        for col_idx, col_name in enumerate(flags_df.columns):
            if col_name == 'Priority':
                priority_col_idx = col_idx
                break

        current_file = None
        for idx, row in flags_df.iterrows():
            excel_row = idx + 2

            # Add separator row when file changes (for multiple files)
            if len(all_file_data) > 1 and current_file is not None and row.get('File') != current_file:
                # This row is a file boundary - make it a separator
                for cell in worksheet[excel_row]:
                    cell.fill = separator_fill

            current_file = row.get('File')

            # Get priority value and apply appropriate color
            priority_value = str(row.get('Priority', '')).lower()

            # Determine fill color based on priority
            row_fill = None
            if 'critical' in priority_value:
                row_fill = critical_fill
            elif 'review' in priority_value:
                row_fill = review_fill
            elif 'info' in priority_value:
                row_fill = info_fill

            # Apply styling to all cells in the row
            for cell in worksheet[excel_row]:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                if row_fill:
                    cell.fill = row_fill

        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 80)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Add raw data sheets for each file
        for file_data in all_file_data:
            subsidiary = file_data['subsidiary']
            filename = file_data['filename']
            bs_df = file_data['bs_df']
            pl_df = file_data['pl_df']

            # Create sheet names (truncate if needed to fit Excel limits)
            bs_sheet_name = f"{subsidiary}_BS"[:31]
            pl_sheet_name = f"{subsidiary}_PL"[:31]

            # Make sheet names unique if there are duplicates
            sheet_counter = 1
            original_bs_name = bs_sheet_name
            original_pl_name = pl_sheet_name

            while bs_sheet_name in writer.sheets:
                bs_sheet_name = f"{original_bs_name[:28]}_{sheet_counter}"
                sheet_counter += 1

            sheet_counter = 1
            while pl_sheet_name in writer.sheets:
                pl_sheet_name = f"{original_pl_name[:28]}_{sheet_counter}"
                sheet_counter += 1

            # Write BS data
            if not bs_df.empty:
                bs_df.to_excel(writer, sheet_name=bs_sheet_name, index=False)
                bs_ws = writer.sheets[bs_sheet_name]

                # Format header
                for cell in bs_ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                # Auto-adjust column widths
                for column in bs_ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    bs_ws.column_dimensions[column_letter].width = adjusted_width

            # Write PL data
            if not pl_df.empty:
                pl_df.to_excel(writer, sheet_name=pl_sheet_name, index=False)
                pl_ws = writer.sheets[pl_sheet_name]

                # Format header
                for cell in pl_ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                # Auto-adjust column widths
                for column in pl_ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    pl_ws.column_dimensions[column_letter].width = adjusted_width

    return output.getvalue()


# ============================================================================
# MAIN PROCESSING FUNCTION FOR BACKEND
# ============================================================================

def process_variance_analysis(
    files: List[Tuple[str, bytes]],
    loan_interest_file: Optional[Tuple[str, bytes]] = None
) -> bytes:
    """
    Main function to process variance analysis for backend integration.
    Supports multiple files with separators and raw data sheets.

    Args:
        files: List of tuples (filename, file_bytes) for BS/PL Breakdown files
        loan_interest_file: Optional tuple (filename, file_bytes) for ERP Loan Interest Rate file

    Returns:
        bytes: Excel file with variance flags and raw data
    """
    if not files:
        raise ValueError("No files provided for analysis")

    all_file_data = []
    combined_flags = []

    # Parse loan interest rate file if provided
    loan_rate_lookup = None
    if loan_interest_file:
        try:
            loan_filename, loan_bytes = loan_interest_file
            print(f"📊 Parsing loan interest rate file: {loan_filename}")
            loan_df = parse_loan_interest_file(loan_bytes)
            loan_rate_lookup = build_entity_rate_lookup(loan_df)
            print(f"✅ Loaded interest rates for {len(loan_rate_lookup)} entities")
        except Exception as e:
            print(f"⚠️ Warning: Could not parse loan interest file: {e}")
            print("   Continuing with basic A2 analysis...")
            loan_rate_lookup = None

    # Process each file
    for filename, file_bytes in files:
        try:
            # Convert bytes to file-like object
            file_obj = io.BytesIO(file_bytes)

            # Extract subsidiary name
            subsidiary = extract_subsidiary_name(file_obj)

            # Find actual sheet names (case-insensitive and flexible)
            file_obj.seek(0)
            temp_wb = load_workbook(file_obj, read_only=True, data_only=True)

            # IMPORTANT: Include both "BS Breakdown" (with space) and "BSbreakdown" (no space)
            # Priority: Try breakdown sheets first, then standalone sheets
            bs_patterns = ["BS Breakdown", "BSbreakdown", "BS breakdown", "bs breakdown",
                           "balance sheet breakdown", "BẢNG CÂN ĐỐI KẾ TOÁN", "Balance Sheet", "BS"]
            pl_patterns = ["PL Breakdown", "PLBreakdown", "PL breakdown", "pl breakdown",
                           "Profit Loss", "Income Statement", "BÁO CÁO KẾT QUẢ KINH DOANH", "P&L", "P/L", "PL"]

            bs_sheet_name = find_sheet_by_pattern(temp_wb, bs_patterns)
            pl_sheet_name = find_sheet_by_pattern(temp_wb, pl_patterns)
            temp_wb.close()

            if not bs_sheet_name and not pl_sheet_name:
                print(f"Warning: No BS or PL sheets found in {filename}")
                print(f"Available sheets: {', '.join(temp_wb.sheetnames)}")
                continue

            # Process BS and PL sheets with found names
            bs_df = pd.DataFrame()
            pl_df = pd.DataFrame()

            if bs_sheet_name:
                file_obj.seek(0)  # Reset file pointer
                bs_df, _ = process_financial_tab(file_obj, bs_sheet_name, "BS", subsidiary)

            if pl_sheet_name:
                file_obj.seek(0)  # Reset file pointer
                pl_df, _ = process_financial_tab(file_obj, pl_sheet_name, "PL", subsidiary)

            if bs_df.empty and pl_df.empty:
                print(f"Warning: No data could be extracted from {filename}")
                continue

            # For D1 balance check, we need the FULL BS sheet (not split by entity)
            # because balance sheets should balance at the file level
            bs_df_full = pd.DataFrame()
            if bs_sheet_name:
                file_obj.seek(0)
                bs_df_full, _ = process_financial_tab(file_obj, bs_sheet_name, "BS", "ALL_ENTITIES")

            # Extract entity name for enhanced A2 analysis (from Row 1)
            entity_name = None
            if loan_rate_lookup:
                file_obj.seek(0)
                entity_name = extract_entity_from_file(file_obj)
                if entity_name:
                    print(f"📌 Extracted entity: {entity_name}")

            # Run variance rules (pass both entity-specific and full BS dataframes)
            # If loan_rate_lookup is available, enhanced A2 will be used
            flags = run_all_variance_rules(
                bs_df, pl_df, bs_df_full,
                entity_name=entity_name,
                loan_rate_lookup=loan_rate_lookup
            )

            # Add filename to each flag for identification
            for flag in flags:
                flag['File'] = filename

            # Store file data
            all_file_data.append({
                'subsidiary': subsidiary,
                'filename': filename,
                'bs_df': bs_df,
                'pl_df': pl_df,
                'flags': flags
            })

            # Add to combined flags
            combined_flags.extend(flags)

        except Exception as e:
            error_msg = str(e)
            print(f"❌ Error processing file '{filename}': {error_msg}")

            # Provide user-friendly error messages based on error type
            if "find sheet" in error_msg.lower() or "sheet" in error_msg.lower():
                user_friendly_error = f"❌ File '{filename}': Missing required sheets. Please ensure your file contains both 'BS Breakdown' and 'PL Breakdown' sheets (case-insensitive)."
            elif "header" in error_msg.lower() or "financial row" in error_msg.lower():
                user_friendly_error = f"❌ File '{filename}': Could not find header row with 'Financial Row' or 'Account Code'. Please check the file structure."
            elif "month" in error_msg.lower() or "period" in error_msg.lower():
                user_friendly_error = f"❌ File '{filename}': Could not detect month columns. Please ensure month headers are in format 'Jan 2025', 'Feb 2025', etc."
            elif "empty" in error_msg.lower():
                user_friendly_error = f"❌ File '{filename}': The file contains no data after the header rows. Please ensure your sheets have financial data."
            elif "account" in error_msg.lower():
                user_friendly_error = f"❌ File '{filename}': Could not find 'Account Code' or 'Account Line' columns. Please check your file format."
            else:
                user_friendly_error = f"❌ File '{filename}': {error_msg}"

            print(user_friendly_error)

            # Continue with next file instead of failing completely
            continue

    if not all_file_data:
        raise ValueError(
            "❌ Could not process any files. Please check that:\n"
            "  • Files are valid Excel files (.xlsx or .xls)\n"
            "  • Files contain 'BS Breakdown' and 'PL Breakdown' sheets\n"
            "  • Sheets have headers with 'Financial Row' or 'Account Code'\n"
            "  • Month columns are in format 'Jan 2025', 'Feb 2025', etc.\n"
            "  • Files contain financial data below the headers"
        )

    # Create output Excel with all data
    xlsx_bytes = create_excel_output(all_file_data, combined_flags)

    return xlsx_bytes
