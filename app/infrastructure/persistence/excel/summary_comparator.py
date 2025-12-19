# excel_processor/summary_comparator.py
import pandas as pd
import xlwings as xw
from typing import Dict, List, Tuple, Set, Optional
from datetime import datetime, timedelta
import os

class SummaryComparator:
    """
    Handles comparison between T7 (previous month) and T8 (current month) Summary files
    to identify and highlight changes in the Summary file.
    """
    
    def __init__(self):
        # Key columns for entity matching (predefined entity mapping keys)
        self.entity_key_columns = ['Unit name', 'Tenant ID', 'Tenant']
        
        # Document Number column for primary comparison
        self.document_number_column = 'Document Number'
        
        # Tracked columns for change detection (when comparing same Document Number + Type + Item)
        self.tracked_columns = {
            'GLA': 'GLA',
            'Start date (for model)': 'Start date (for model)',
            'End date (for model)': 'End date (for model)',
            'Rent (USD)_Item': 'Rent USD_Item (for model)',
            'Rent (VND)_Item': 'Rent VND_Item (for model)',
            'Escalation rate': 'Escalation rate (for model)',
            'Service charge': 'Service charge (for model)',
            'Broker? (Yes/No)': 'Broker? (Yes/No)',
            'UFL Status': 'UFL Status'
        }

    
    def generate_detailed_change_log(self, summary_old_path: str, summary_new_path: str, log_file_path: str = None) -> str:
        """
        Generate a detailed log file showing row-level changes with specific column modifications
        
        Args:
            summary_old_path: Path to previous period Summary file
            summary_new_path: Path to current period Summary file
            log_file_path: Optional custom path for log file. If None, auto-generates based on input files
        
        Returns:
            Path to the generated log file
        """
        import json
        from datetime import datetime
        
        # Auto-generate log file path if not provided
        if log_file_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            old_name = os.path.splitext(os.path.basename(summary_old_path))[0]
            new_name = os.path.splitext(os.path.basename(summary_new_path))[0]
            log_file_path = f"summary_comparison_{old_name}_vs_{new_name}_{timestamp}.log"
        
        print(f"📝 Generating detailed change log: {log_file_path}")
        
        try:
            # Load both files
            old_df = self.load_summary_file(summary_old_path)
            new_df = self.load_summary_file(summary_new_path)
            
            # Create lookup for old data
            old_lookup = {}
            for idx, row in old_df.iterrows():
                key = self.normalize_row_key(row)
                if key and key != '||':
                    old_lookup[key] = row
            
            # Detailed change tracking
            detailed_changes = []
            summary_stats = {
                'total_new_rows': len(new_df),
                'total_old_rows': len(old_df),
                'changed_rows': 0,
                'new_rows': 0,
                'column_change_counts': {}
            }
            
            # Compare each row in new file
            for idx, new_row in new_df.iterrows():
                new_key = self.normalize_row_key(new_row)
                excel_row = idx + 2  # Excel row number (1-indexed + header)
                
                if not new_key or new_key == '||':
                    continue
                
                row_changes = {
                    'excel_row': excel_row,
                    'key_columns': {
                        'Unit name': str(new_row.get('Unit name', '')),
                        'Tenant ID': str(new_row.get('Tenant ID', '')),
                        'Tenant': str(new_row.get('Tenant', ''))
                    },
                    'change_type': '',
                    'modified_columns': []
                }
                
                if new_key in old_lookup:
                    old_row = old_lookup[new_key]
                    has_changes = False
                    
                    # Check each tracked column for changes
                    for new_col, old_col in self.tracked_columns.items():
                        if new_col in new_df.columns and old_col in old_df.columns:
                            new_val = self._normalize_value(new_row.get(new_col), new_col)
                            old_val = self._normalize_value(old_row.get(old_col), old_col)
                            
                            if new_val != old_val:
                                has_changes = True
                                
                                # Track raw values for better readability
                                raw_new_val = str(new_row.get(new_col, ''))
                                raw_old_val = str(old_row.get(old_col, ''))
                                
                                column_change = {
                                    'column': new_col,
                                    'old_value': raw_old_val,
                                    'new_value': raw_new_val,
                                    'normalized_old': old_val,
                                    'normalized_new': new_val
                                }
                                row_changes['modified_columns'].append(column_change)
                                
                                # Update stats
                                if new_col not in summary_stats['column_change_counts']:
                                    summary_stats['column_change_counts'][new_col] = 0
                                summary_stats['column_change_counts'][new_col] += 1
                    
                    if has_changes:
                        row_changes['change_type'] = 'MODIFIED'
                        detailed_changes.append(row_changes)
                        summary_stats['changed_rows'] += 1
                else:
                    # New row
                    row_changes['change_type'] = 'NEW'
                    # Add all column values for new rows
                    for col in self.tracked_columns.keys():
                        if col in new_df.columns:
                            raw_val = str(new_row.get(col, ''))
                            if raw_val.strip():  # Only include non-empty values
                                column_info = {
                                    'column': col,
                                    'old_value': '',
                                    'new_value': raw_val,
                                    'normalized_old': '',
                                    'normalized_new': self._normalize_value(new_row.get(col), col)
                                }
                                row_changes['modified_columns'].append(column_info)
                    
                    detailed_changes.append(row_changes)
                    summary_stats['new_rows'] += 1
            
            # Generate log content
            log_content = self._generate_log_content(
                summary_old_path, summary_new_path, detailed_changes, summary_stats
            )
            
            # Write log file
            with open(log_file_path, 'w', encoding='utf-8') as f:
                f.write(log_content)
            
            print(f"   ✅ Change log generated: {log_file_path}")
            print(f"   📊 Summary: {summary_stats['changed_rows']} modified, {summary_stats['new_rows']} new rows")
            
            return log_file_path
            
        except Exception as e:
            print(f"   ❌ Error generating change log: {e}")
            raise
    
    def _generate_log_content(self, old_path: str, new_path: str, changes: list, stats: dict) -> str:
        """Generate formatted log content"""
        from datetime import datetime
        
        content = []
        content.append("=" * 100)
        content.append("SUMMARY FILE COMPARISON - DETAILED CHANGE LOG")
        content.append("=" * 100)
        content.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        content.append(f"Previous File: {old_path}")
        content.append(f"Current File:  {new_path}")
        content.append("")
        
        # Summary statistics
        content.append("📊 SUMMARY STATISTICS")
        content.append("-" * 50)
        content.append(f"Total rows in previous file: {stats['total_old_rows']}")
        content.append(f"Total rows in current file:  {stats['total_new_rows']}")
        content.append(f"Modified rows: {stats['changed_rows']}")
        content.append(f"New rows:      {stats['new_rows']}")
        content.append(f"Total changes: {stats['changed_rows'] + stats['new_rows']}")
        content.append("")
        
        # Column change statistics
        if stats['column_change_counts']:
            content.append("📈 COLUMN CHANGE FREQUENCY")
            content.append("-" * 50)
            for col, count in sorted(stats['column_change_counts'].items(), key=lambda x: x[1], reverse=True):
                content.append(f"{col}: {count} changes")
            content.append("")
        
        # Detailed changes
        content.append("📝 DETAILED CHANGES")
        content.append("-" * 50)
        
        if not changes:
            content.append("No changes detected.")
        else:
            for i, change in enumerate(changes, 1):
                content.append(f"\n[{i}] ROW {change['excel_row']} - {change['change_type']}")
                content.append(f"    Key: Unit='{change['key_columns']['Unit name']}', "
                             f"TenantID='{change['key_columns']['Tenant ID']}', "
                             f"Tenant='{change['key_columns']['Tenant']}'")
                
                if change['modified_columns']:
                    content.append("    Modified Columns:")
                    for col_change in change['modified_columns']:
                        if change['change_type'] == 'NEW':
                            content.append(f"      • {col_change['column']}: '{col_change['new_value']}'")
                        else:
                            content.append(f"      • {col_change['column']}: '{col_change['old_value']}' → '{col_change['new_value']}'")
                content.append("")
        
        content.append("=" * 100)
        content.append("END OF CHANGE LOG")
        content.append("=" * 100)
        
        return "\n".join(content)
    
    def load_summary_file(self, file_path: str) -> pd.DataFrame:
        """Load summary file with error handling"""
        try:
            # Try different engines based on file extension
            if file_path.endswith('.xlsx'):
                engines = ['openpyxl', 'xlrd']
            elif file_path.endswith('.xls'):
                engines = ['xlrd', 'openpyxl']
            else:
                engines = ['openpyxl', 'xlrd']
            
            for engine in engines:
                try:
                    df = pd.read_excel(file_path, engine=engine)
                    print(f"   Loaded {file_path} using {engine} engine")
                    return df
                except Exception as e:
                    print(f"   Failed to load with {engine}: {e}")
                    continue
            
            raise Exception(f"Failed to load {file_path} with any engine")
            
        except Exception as e:
            print(f"   Error loading {file_path}: {e}")
            raise
    
    def normalize_row_key(self, row: pd.Series) -> str:
        """Create normalized key for row matching - for backward compatibility"""
        key_parts = []
        for col in self.entity_key_columns:
            if col in row:
                value = str(row[col]).strip().lower() if pd.notna(row[col]) else ''
                key_parts.append(value)
            else:
                key_parts.append('')
        return '|'.join(key_parts)

    
    def normalize_entity_key(self, row: pd.Series) -> str:
        """Create normalized key for entity matching using predefined keys"""
        key_parts = []
        for col in self.entity_key_columns:
            if col in row:
                value = str(row[col]).strip().lower() if pd.notna(row[col]) else ''
                key_parts.append(value)
            else:
                key_parts.append('')
        return '|'.join(key_parts)
    
    def get_latest_record_by_date_created(self, records: List[pd.Series], date_created_col: str = 'Date created') -> pd.Series:
        """
        Get record with latest date_created from a list of records with same entity key.
        This handles the requirement: for duplicate items, use the one with latest date_created.
        
        Args:
            records: List of pandas Series (rows) with the same entity key
            date_created_col: Column name to check for date created
            
        Returns:
            The record with the latest date_created
        """
        if len(records) == 1:
            return records[0]
        
        latest_record = records[0]
        latest_date = None
        
        # Parse date_created for the first record
        if date_created_col in latest_record:
            latest_date = self._parse_date_created(latest_record[date_created_col])
        
        # Compare with other records
        for record in records[1:]:
            if date_created_col in record:
                current_date = self._parse_date_created(record[date_created_col])
                
                # If current record has a later date_created, use it
                if current_date and (latest_date is None or current_date > latest_date):
                    latest_record = record
                    latest_date = current_date
        
        return latest_record
    
    def compare_summary_files_with_latest_date_created(self, summary_old_path: str, summary_new_path: str) -> Dict[str, any]:
        """
        NEW LOGIC: Compare summary files using latest date_created rule to avoid duplicates.
        
        For both Old and New files:
        - Group items by entity key (Unit name, Tenant ID, Tenant) 
        - Within each group, take the record with the latest date_created
        - Then compare these deduplicated records
        
        Args:
            summary_old_path: Path to previous period Summary file
            summary_new_path: Path to current period Summary file
        
        Returns:
            Dict with:
            - 'changed_cells': dict mapping row indices to sets of changed column names
            - 'changed_rows': set of row indices
            - 'new_rows': set of entirely new row indices
        """
        print(f"Comparing Summary files using latest date_created rule:")
        print(f"   Previous: {summary_old_path}")
        print(f"   Current:  {summary_new_path}")
        
        # Load both files
        old_df = self.load_summary_file(summary_old_path)
        new_df = self.load_summary_file(summary_new_path)
        
        print(f"   Previous rows: {len(old_df)}, Current rows: {len(new_df)}")
        
        # Find date_created column in both files
        old_date_created_col = None
        new_date_created_col = None
        
        for col in old_df.columns:
            if 'date' in col.lower() and 'created' in col.lower():
                old_date_created_col = col
                break
        
        for col in new_df.columns:
            if 'date' in col.lower() and 'created' in col.lower():
                new_date_created_col = col
                break
        
        if not old_date_created_col or not new_date_created_col:
            print(f"   Warning: Date created column not found. Falling back to regular comparison.")
            return self.compare_summary_files(summary_old_path, summary_new_path)
        
        print(f"   Using date_created columns: Old='{old_date_created_col}', New='{new_date_created_col}'")
        
        # Group OLD file records by entity key and get latest date_created for each group
        old_entity_groups = {}
        for idx, row in old_df.iterrows():
            entity_key = self.normalize_entity_key(row)
            if entity_key and entity_key != '||':
                if entity_key not in old_entity_groups:
                    old_entity_groups[entity_key] = []
                old_entity_groups[entity_key].append(row)
        
        # Get latest record for each entity in old file
        old_latest_lookup = {}
        for entity_key, records in old_entity_groups.items():
            latest_record = self.get_latest_record_by_date_created(records, old_date_created_col)
            old_latest_lookup[entity_key] = latest_record
        
        print(f"   Old file: {len(old_entity_groups)} entity groups -> {len(old_latest_lookup)} latest records")
        
        # Group NEW file records by entity key and get latest date_created for each group
        new_entity_groups = {}
        new_row_indices = {}  # Track original row indices
        
        for idx, row in new_df.iterrows():
            entity_key = self.normalize_entity_key(row)
            if entity_key and entity_key != '||':
                if entity_key not in new_entity_groups:
                    new_entity_groups[entity_key] = []
                    new_row_indices[entity_key] = []
                new_entity_groups[entity_key].append(row)
                new_row_indices[entity_key].append(idx + 2)  # Excel row number
        
        # Get latest record for each entity in new file and track row numbers
        new_latest_lookup = {}
        latest_row_mapping = {}  # entity_key -> excel_row_num
        
        for entity_key, records in new_entity_groups.items():
            latest_record = self.get_latest_record_by_date_created(records, new_date_created_col)
            new_latest_lookup[entity_key] = latest_record
            
            # Find the row index of the latest record
            for i, record in enumerate(records):
                if latest_record.equals(record):
                    latest_row_mapping[entity_key] = new_row_indices[entity_key][i]
                    break
        
        print(f"   New file: {len(new_entity_groups)} entity groups -> {len(new_latest_lookup)} latest records")
        
        # Compare latest records
        changed_rows = set()
        changed_cells = {}
        new_rows = set()
        
        for entity_key, new_latest in new_latest_lookup.items():
            excel_row = latest_row_mapping[entity_key]
            
            if entity_key in old_latest_lookup:
                # Compare latest records
                old_latest = old_latest_lookup[entity_key]
                changed_columns = set()
                
                for new_col, old_col in self.tracked_columns.items():
                    if new_col in new_df.columns and old_col in old_df.columns:
                        new_val = self._normalize_value(new_latest.get(new_col), new_col)
                        old_val = self._normalize_value(old_latest.get(old_col), old_col)
                        
                        if new_val != old_val:
                            changed_columns.add(new_col)
                
                if changed_columns:
                    changed_rows.add(excel_row)
                    changed_cells[excel_row] = changed_columns
                    print(f"   Entity '{entity_key}' has changes in row {excel_row}")
            else:
                # Entirely new entity
                new_rows.add(excel_row)
                changed_rows.add(excel_row)
                
                # Mark all tracked columns as new for highlighting
                all_columns = set(self.tracked_columns.keys())
                all_columns = {col for col in all_columns if col in new_df.columns}
                changed_cells[excel_row] = all_columns
                
                print(f"   New entity '{entity_key}' in row {excel_row}")
        
        print(f"   Found {len(new_rows)} entirely new entities")
        print(f"   Found {len(changed_rows)} total rows with changes")
        print(f"   Deduplication applied: Latest date_created used for each entity")
        
        return {
            'changed_rows': changed_rows,
            'changed_cells': changed_cells,
            'new_rows': new_rows
        }
    
    def get_document_number(self, row: pd.Series) -> str:
        """Extract Document Number from row"""
        if self.document_number_column in row:
            doc_num = str(row[self.document_number_column]).strip() if pd.notna(row[self.document_number_column]) else ''
            return doc_num
        return ''
    
    def compare_summary_files_by_document_number(self, summary_old_path: str, summary_new_path: str) -> Dict[str, any]:
        """
        Compare summary files using Document Number as the primary key identifier.
        
        NEW LOGIC:
        - Use Document Number for comparison (each Document Number is treated as unique record)
        - Highlight newly added Document Numbers (entire rows)
        - Map entities using predefined keys (Unit name, Tenant ID, Tenant)
        - Within each mapped entity, compare the list of Document Numbers
        
        Args:
            summary_old_path: Path to previous period Summary file
            summary_new_path: Path to current period Summary file
        
        Returns:
            Dict with:
            - 'new_document_numbers': set of row indices with newly added Document Numbers
            - 'changed_cells': dict mapping row indices to sets of changed column names
            - 'changed_rows': set of row indices (for backward compatibility)
            - 'entity_mapping': dict mapping entities between old and new files
        """
        print(f"🔧 Comparing Summary files using Document Number logic:")
        print(f"   Previous: {summary_old_path}")
        print(f"   Current:  {summary_new_path}")
        
        # Load both files
        old_df = self.load_summary_file(summary_old_path)
        new_df = self.load_summary_file(summary_new_path)
        
        print(f"   Previous rows: {len(old_df)}, Current rows: {len(new_df)}")
        
        # Check if Document Number column exists
        if self.document_number_column not in new_df.columns:
            print(f"   ⚠️ Warning: '{self.document_number_column}' column not found in new file. Using fallback logic.")
            return self.compare_summary_files(summary_old_path, summary_new_path)
        
        if self.document_number_column not in old_df.columns:
            print(f"   ⚠️ Warning: '{self.document_number_column}' column not found in old file. Treating all as new.")
            # All rows in new file are considered new
            new_document_numbers = set(range(2, len(new_df) + 2))  # Excel row numbers
            return {
                'new_document_numbers': new_document_numbers,
                'changed_cells': {},
                'changed_rows': new_document_numbers,
                'entity_mapping': {}
            }
        
        # Create lookups
        # 1. Document Number lookup for old file
        old_doc_lookup = {}
        for idx, row in old_df.iterrows():
            doc_num = self.get_document_number(row)
            if doc_num:
                old_doc_lookup[doc_num] = row
        
        # 2. Entity mapping lookup for old file
        old_entity_lookup = {}
        for idx, row in old_df.iterrows():
            entity_key = self.normalize_entity_key(row)
            if entity_key and entity_key != '||':
                if entity_key not in old_entity_lookup:
                    old_entity_lookup[entity_key] = []
                doc_num = self.get_document_number(row)
                if doc_num:
                    old_entity_lookup[entity_key].append(doc_num)
        
        print(f"   Previous Document Numbers: {len(old_doc_lookup)}")
        print(f"   Previous Entities: {len(old_entity_lookup)}")
        
        # Track results
        new_document_numbers = set()
        changed_cells = {}
        changed_rows = set()
        entity_mapping = {}
        
        # Compare each row in new file
        for idx, new_row in new_df.iterrows():
            row_num = idx + 2  # Excel row number (1-indexed + header)
            doc_num = self.get_document_number(new_row)
            entity_key = self.normalize_entity_key(new_row)
            
            if not doc_num:
                continue
            
            # Check if this Document Number exists in old file
            if doc_num in old_doc_lookup:
                # Document Number exists - check for changes in tracked columns
                old_row = old_doc_lookup[doc_num]
                changed_columns = set()
                
                for new_col, old_col in self.tracked_columns.items():
                    if new_col in new_df.columns and old_col in old_df.columns:
                        new_val = self._normalize_value(new_row.get(new_col), new_col)
                        old_val = self._normalize_value(old_row.get(old_col), old_col)
                        
                        if new_val != old_val:
                            changed_columns.add(new_col)
                
                if changed_columns:
                    changed_rows.add(row_num)
                    changed_cells[row_num] = changed_columns
                    print(f"   📝 Document Number '{doc_num}' has changes in row {row_num}")
            
            else:
                # New Document Number - highlight entire row
                new_document_numbers.add(row_num)
                changed_rows.add(row_num)
                
                # Mark all tracked columns as new for highlighting
                all_columns = set(self.tracked_columns.keys())
                # Only include columns that actually exist in the new dataframe
                all_columns = {col for col in all_columns if col in new_df.columns}
                changed_cells[row_num] = all_columns
                
                print(f"   🆕 New Document Number '{doc_num}' in row {row_num}")
                
                # Track entity mapping for this new Document Number
                if entity_key and entity_key != '||':
                    if entity_key not in entity_mapping:
                        entity_mapping[entity_key] = {
                            'entity_info': {
                                'Unit name': str(new_row.get('Unit name', '')),
                                'Tenant ID': str(new_row.get('Tenant ID', '')),
                                'Tenant': str(new_row.get('Tenant', ''))
                            },
                            'old_document_numbers': old_entity_lookup.get(entity_key, []),
                            'new_document_numbers': []
                        }
                    entity_mapping[entity_key]['new_document_numbers'].append(doc_num)
        
        print(f"   🎯 Found {len(new_document_numbers)} newly added Document Numbers")
        print(f"   📊 Found {len(changed_rows)} total rows with changes")
        print(f"   🏢 Mapped {len(entity_mapping)} entities with new Document Numbers")
        
        return {
            'new_document_numbers': new_document_numbers,
            'changed_cells': changed_cells,
            'changed_rows': changed_rows,
            'entity_mapping': entity_mapping
        }
    
    def compare_summary_files(self, summary_old_path: str, summary_new_path: str) -> Dict[str, any]:
        """
        Compare summary_old vs summary_new and return specific cells that changed
        
        Args:
            summary_old_path: Path to previous period Summary file
            summary_new_path: Path to current period Summary file
        
        Returns:
            Dict with 'changed_cells': dict mapping row indices to sets of changed column names,
            and 'changed_rows': set of row indices (for backward compatibility)
        """
        print(f"Comparing Summary files:")
        print(f"   Previous: {summary_old_path}")
        print(f"   Current:  {summary_new_path}")
        
        # Load both files
        old_df = self.load_summary_file(summary_old_path)
        new_df = self.load_summary_file(summary_new_path)
        
        print(f"   Previous rows: {len(old_df)}, Current rows: {len(new_df)}")
        
        # Create lookup for old data
        old_lookup = {}
        for idx, row in old_df.iterrows():
            key = self.normalize_row_key(row)
            if key and key != '||':  # Skip empty keys
                old_lookup[key] = row
        
        print(f"   Previous lookup created with {len(old_lookup)} unique keys")
        
        changed_rows = set()
        changed_cells = {}  # row_num -> set of changed column names
        
        # Compare each row in new file with corresponding row in old file
        for idx, new_row in new_df.iterrows():
            new_key = self.normalize_row_key(new_row)
            row_num = idx + 2  # +2 for Excel row number (1-indexed + header)
            
            if not new_key or new_key == '||':
                continue
                
            if new_key in old_lookup:
                old_row = old_lookup[new_key]
                
                # Check each tracked column for changes
                changed_columns = set()
                for new_col, old_col in self.tracked_columns.items():
                    if new_col in new_df.columns and old_col in old_df.columns:
                        new_val = self._normalize_value(new_row.get(new_col), new_col)
                        old_val = self._normalize_value(old_row.get(old_col), old_col)
                        
                        if new_val != old_val:
                            changed_columns.add(new_col)
                
                if changed_columns:
                    changed_rows.add(row_num)
                    changed_cells[row_num] = changed_columns
            else:
                # New row in current file - mark all tracked columns as changed
                changed_rows.add(row_num)
                changed_columns = set(self.tracked_columns.keys())
                # Only include columns that actually exist in the new dataframe
                changed_columns = {col for col in changed_columns if col in new_df.columns}
                changed_cells[row_num] = changed_columns
                print(f"   Row {row_num} is new in current file")
        
        print(f"   Found {len(changed_rows)} changed/new rows in current file")
        print(f"   Found {sum(len(cols) for cols in changed_cells.values())} individual cell changes")
        
        return {
            'changed_rows': changed_rows,
            'changed_cells': changed_cells
        }

    def compare_summary_files_with_oldest_match(self, summary_old_path: str, summary_new_path: str) -> Dict[str, any]:
        """
        Compare each row in new Excel file against oldest matching record in old file
        using key mapping ['Unit name', 'Tenant ID', 'Tenant'] and Date created column.
        
        For Document Numbers that don't exist in old file, find oldest record with same key mapping.
        Highlight only changed cells (light blue) or entire new rows (yellow).
        
        Args:
            summary_old_path: Path to previous period Summary file
            summary_new_path: Path to current period Summary file
        
        Returns:
            Dict with 'changed_cells': dict mapping row indices to sets of changed column names,
            'changed_rows': set of row indices, and 'new_rows': set of entirely new row indices
        """
        print(f"Comparing Summary files with oldest match logic:")
        print(f"   Previous: {summary_old_path}")
        print(f"   Current:  {summary_new_path}")
        
        # Load both files
        old_df = self.load_summary_file(summary_old_path)
        new_df = self.load_summary_file(summary_new_path)
        
        print(f"   Previous rows: {len(old_df)}, Current rows: {len(new_df)}")
        
        # Ensure Date created column exists
        date_created_col = 'Date created'
        if date_created_col not in old_df.columns:
            print(f"   Warning: '{date_created_col}' column not found in old file")
            # Fallback to regular comparison
            return self.compare_summary_files(summary_old_path, summary_new_path)
        
        # Group old records by entity key and find oldest in each group
        old_grouped = {}
        for idx, row in old_df.iterrows():
            entity_key = self.normalize_entity_key(row)
            if not entity_key or entity_key == '||':
                continue
                
            # Parse date created
            date_created = self._parse_date_created(row.get(date_created_col))
            if date_created is None:
                continue
                
            # Store all records for this entity key
            if entity_key not in old_grouped:
                old_grouped[entity_key] = []
            old_grouped[entity_key].append({
                'row': row,
                'date_created': date_created,
                'document_number': self.get_document_number(row)
            })
        
        # Find oldest record for each entity key
        old_oldest_lookup = {}
        old_document_lookup = {}  # For direct document number matches
        
        for entity_key, records in old_grouped.items():
            # Sort by date created to find oldest
            records.sort(key=lambda x: x['date_created'])
            oldest_record = records[0]
            old_oldest_lookup[entity_key] = oldest_record['row']
            
            # Also create document number lookup for direct matches
            for record in records:
                doc_num = record['document_number']
                if doc_num:
                    old_document_lookup[doc_num] = record['row']
        
        print(f"   Created lookup with {len(old_oldest_lookup)} entity groups")
        print(f"   Created document lookup with {len(old_document_lookup)} documents")
        
        changed_rows = set()
        changed_cells = {}  # row_num -> set of changed column names
        new_rows = set()    # Entirely new rows (no matching entity key)
        
        # Compare each row in new file
        for idx, new_row in new_df.iterrows():
            row_num = idx + 2  # +2 for Excel row number (1-indexed + header)
            new_doc_num = self.get_document_number(new_row)
            entity_key = self.normalize_entity_key(new_row)
            
            if not entity_key or entity_key == '||':
                continue
            
            # First, try to find direct document number match
            comparison_row = None
            if new_doc_num and new_doc_num in old_document_lookup:
                comparison_row = old_document_lookup[new_doc_num]
                print(f"   Row {row_num}: Found direct document match for '{new_doc_num}'")
            
            # If no direct document match, use oldest record with same entity key
            elif entity_key in old_oldest_lookup:
                comparison_row = old_oldest_lookup[entity_key]
                print(f"   Row {row_num}: Using oldest record for entity key (doc '{new_doc_num}' not found)")
            
            # If no matching entity key at all, mark as entirely new
            else:
                new_rows.add(row_num)
                print(f"   Row {row_num}: Entirely new entity key")
                continue
            
            # Compare cells between new row and comparison row
            changed_columns = set()
            for new_col, old_col in self.tracked_columns.items():
                if new_col in new_df.columns and old_col in old_df.columns:
                    new_val = self._normalize_value(new_row.get(new_col), new_col)
                    old_val = self._normalize_value(comparison_row.get(old_col), old_col)
                    
                    if new_val != old_val:
                        changed_columns.add(new_col)
            
            if changed_columns:
                changed_rows.add(row_num)
                changed_cells[row_num] = changed_columns
        
        print(f"   Found {len(changed_rows)} rows with cell changes")
        print(f"   Found {len(new_rows)} entirely new rows")
        print(f"   Found {sum(len(cols) for cols in changed_cells.values())} individual cell changes")
        
        return {
            'changed_rows': changed_rows,
            'changed_cells': changed_cells,
            'new_rows': new_rows
        }
    
    def _parse_date_created(self, date_value) -> datetime:
        """Parse date created value to datetime object for comparison"""
        if pd.isna(date_value):
            return None
            
        if isinstance(date_value, datetime):
            return date_value
            
        try:
            date_str = str(date_value).strip()
            
            # Try common date formats
            for date_format in ['%m/%d/%Y', '%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d']:
                try:
                    return datetime.strptime(date_str, date_format)
                except ValueError:
                    continue
            
            # If no format works, return None
            return None
            
        except Exception:
            return None
    
    def _normalize_value(self, value, column_name: str = '') -> str:
        """Normalize value for comparison based on column type"""
        if pd.isna(value):
            return ''
        
        # Handle datetime objects
        if isinstance(value, datetime):
            return value.strftime('%m/%d/%Y')
        
        # Convert to string and normalize
        str_val = str(value).strip()
        
        # Handle date columns
        if 'date' in column_name.lower():
            try:
                # Try to parse and reformat date
                if '/' in str_val:
                    # Try mm/dd/yyyy format
                    dt = datetime.strptime(str_val, '%m/%d/%Y')
                elif '-' in str_val:
                    # Try yyyy-mm-dd format
                    dt = datetime.strptime(str_val, '%Y-%m-%d')
                else:
                    return str_val.lower()
                return dt.strftime('%m/%d/%Y')
            except:
                return str_val.lower()
        
        # Handle numeric columns (Rent, Service charge, Escalation rate)
        elif any(keyword in column_name.lower() for keyword in ['rent', 'service charge', 'escalation rate']):
            try:
                # Remove common formatting characters
                clean_val = str_val.replace(',', '').replace('$', '').replace('%', '').strip()
                if clean_val and clean_val.replace('.', '').replace('-', '').isdigit():
                    # Convert to float for consistent comparison
                    num_val = float(clean_val)
                    return str(num_val)
                else:
                    return str_val.lower()
            except:
                return str_val.lower()
        
        # Handle boolean columns (Broker)
        elif 'broker' in column_name.lower():
            normalized = str_val.lower()
            # Normalize common boolean representations
            if normalized in ['yes', 'y', 'true', '1']:
                return 'yes'
            elif normalized in ['no', 'n', 'false', '0']:
                return 'no'
            else:
                return normalized
        
        # Handle GLA (numeric)
        elif 'gla' in column_name.lower():
            try:
                clean_val = str_val.replace(',', '').strip()
                if clean_val and clean_val.replace('.', '').isdigit():
                    return str(float(clean_val))
                else:
                    return str_val.lower()
            except:
                return str_val.lower()
        
        # Default: return normalized string
        return str_val.lower()
    
    def apply_highlighting_to_summary(self, summary_path: str, comparison_result: Dict[str, any]) -> bool:
        """
        Apply highlighting to specific changed cells in the Summary file
        
        Args:
            summary_path: Path to the T8 Summary file
            comparison_result: Result from compare_summary_files containing changed_cells and changed_rows
        
        Returns:
            True if highlighting was successful, False otherwise
        """
        changed_cells = comparison_result.get('changed_cells', {})
        changed_rows = comparison_result.get('changed_rows', set())
        
        if not changed_cells and not changed_rows:
            print("   No cells to highlight in Summary file")
            return True
        
        try:
            print(f"Applying cell-specific highlighting to Summary file")
            print(f"   Rows with changes: {len(changed_rows)}")
            print(f"   Total changed cells: {sum(len(cols) for cols in changed_cells.values())}")
            
            # Open the file with xlwings
            app = xw.App(visible=False, add_book=False)
            try:
                wb = app.books.open(summary_path)
                sheet = wb.sheets[0]  # Assume first sheet
                
                # Get column mapping for the sheet
                header_row = sheet.range('1:1').value
                if not header_row:
                    print("   Could not read header row")
                    return False
                
                # Create column name to column number mapping
                col_mapping = {}
                for col_idx, col_name in enumerate(header_row):
                    if col_name:
                        col_mapping[str(col_name).strip()] = col_idx + 1
                
                highlighted_cells = 0
                for row_num, changed_column_names in changed_cells.items():
                    try:
                        for col_name in changed_column_names:
                            # Find the column number for this column name
                            col_num = None
                            
                            # Direct match first
                            if col_name in col_mapping:
                                col_num = col_mapping[col_name]
                            else:
                                # Try fuzzy matching for column names
                                for header_col, header_col_num in col_mapping.items():
                                    if col_name.lower().strip() in header_col.lower().strip() or \
                                       header_col.lower().strip() in col_name.lower().strip():
                                        col_num = header_col_num
                                        break # Note
                            
                            if col_num:
                                # Highlight the specific cell with light blue background
                                cell = sheet.range((row_num, col_num))
                                cell.color = (173, 216, 230)  # Light blue for changed cells
                                highlighted_cells += 1
                            else:
                                print(f"   Could not find column '{col_name}' in header row")
                    
                    except Exception as e:
                        print(f"   Failed to highlight cells in row {row_num}: {e}")
                
                # Save the file
                wb.save()
                print(f"   Saved Summary file with {highlighted_cells} highlighted cells")
                
                return True
                
            finally:
                wb.close()
                app.quit()
                
        except Exception as e:
            print(f"   Error applying highlighting to Summary file: {e}")
            return False

    def apply_enhanced_highlighting_to_summary(self, summary_path: str, comparison_result: Dict[str, any]) -> bool:
        """
        Apply enhanced highlighting to Summary file with different colors for different change types:
        - Light blue for changed cells
        - Yellow for entirely new rows (no matching entity key)
        
        Args:
            summary_path: Path to the T8 Summary file
            comparison_result: Result from compare_summary_files_with_oldest_match
        
        Returns:
            True if highlighting was successful, False otherwise
        """
        changed_cells = comparison_result.get('changed_cells', {})
        changed_rows = comparison_result.get('changed_rows', set())
        new_rows = comparison_result.get('new_rows', set())
        
        if not changed_cells and not changed_rows and not new_rows:
            print("   No cells to highlight in Summary file")
            return True
        
        try:
            print(f"Applying enhanced highlighting to Summary file")
            print(f"   Rows with cell changes: {len(changed_rows)}")
            print(f"   Entirely new rows: {len(new_rows)}")
            print(f"   Total changed cells: {sum(len(cols) for cols in changed_cells.values())}")
            
            # Open the file with xlwings
            app = xw.App(visible=False, add_book=False)
            try:
                wb = app.books.open(summary_path)
                sheet = wb.sheets[0]  # Assume first sheet
                
                # Get column mapping for the sheet
                header_row = sheet.range('1:1').value
                if not header_row:
                    print("   Could not read header row")
                    return False
                
                # Create column name to column number mapping
                col_mapping = {}
                for col_idx, col_name in enumerate(header_row):
                    if col_name:
                        col_mapping[str(col_name).strip()] = col_idx + 1
                
                highlighted_cells = 0
                highlighted_rows = 0
                
                # Highlight changed cells with light blue
                for row_num, changed_column_names in changed_cells.items():
                    try:
                        for col_name in changed_column_names:
                            # Find the column number for this column name
                            col_num = None
                            
                            # Direct match first
                            if col_name in col_mapping:
                                col_num = col_mapping[col_name]
                            else:
                                # Try fuzzy matching for column names
                                for header_col, header_col_num in col_mapping.items():
                                    if col_name.lower().strip() in header_col.lower().strip() or \
                                       header_col.lower().strip() in col_name.lower().strip():
                                        col_num = header_col_num
                                        break
                            
                            if col_num:
                                # Highlight the specific cell with light blue background
                                cell = sheet.range((row_num, col_num))
                                cell.color = (173, 216, 230)  # Light blue for changed cells
                                highlighted_cells += 1
                            else:
                                print(f"   Could not find column '{col_name}' in header row")
                    
                    except Exception as e:
                        print(f"   Failed to highlight cells in row {row_num}: {e}")
                
                # Highlight entirely new rows with yellow
                for row_num in new_rows:
                    try:
                        # Highlight the entire row with yellow background
                        last_col = len(header_row)
                        row_range = sheet.range((row_num, 1), (row_num, last_col))
                        row_range.color = (255, 255, 0)  # Yellow for entirely new rows
                        highlighted_rows += 1
                        
                    except Exception as e:
                        print(f"   Failed to highlight row {row_num}: {e}")
                
                # Save the file
                wb.save()
                print(f"   Saved Summary file with {highlighted_cells} highlighted cells and {highlighted_rows} highlighted rows")
                
                return True
                
            finally:
                wb.close()
                app.quit()
                
        except Exception as e:
            print(f"   Error applying enhanced highlighting to Summary file: {e}\"")
            return False
    
    def process_summary_comparison_with_latest_date_created(self, summary_old_path: str, summary_new_path: str, 
                                                          generate_log: bool = True, log_file_path: str = None) -> bool:
        """
        Complete workflow using NEW latest date_created deduplication logic.
        
        This method implements the updated requirements:
        1. Group items by entity key (Unit name, Tenant ID, Tenant)
        2. For each group, use the record with the latest date_created
        3. Compare these deduplicated records to avoid duplicate processing
        4. Highlight changes with appropriate colors
        
        Args:
            summary_old_path: Path to previous period Summary file
            summary_new_path: Path to current period Summary file
            generate_log: Whether to generate detailed change log file
            log_file_path: Optional custom path for log file
        
        Returns:
            True if process completed successfully
        """
        try:
            # Validate files exist
            if not os.path.exists(summary_old_path):
                print(f"   Previous file not found: {summary_old_path}")
                return False
            
            if not os.path.exists(summary_new_path):
                print(f"   Current file not found: {summary_new_path}")
                return False
            
            print(f"Starting latest date_created deduplication comparison workflow")
            
            # Use the new latest date_created comparison logic
            comparison_results = self.compare_summary_files_with_latest_date_created(summary_old_path, summary_new_path)
            
            changed_rows = comparison_results['changed_rows']
            changed_cells = comparison_results['changed_cells']
            new_rows = comparison_results['new_rows']
            
            # Generate detailed log if requested
            if generate_log:
                log_path = self.generate_latest_date_created_change_log(
                    summary_old_path, summary_new_path, comparison_results, log_file_path
                )
                print(f"   Detailed log saved: {log_path}")
            
            # Apply enhanced highlighting (blue cells for changes, yellow rows for new entities)
            success = self.apply_enhanced_highlighting_to_summary(summary_new_path, comparison_results)
            
            if success:
                print(f"   Latest date_created comparison completed successfully!")
                print(f"   Entirely new entities: {len(new_rows)}")
                print(f"   Total rows with changes: {len(changed_rows)}")
                print(f"   Individual cell changes: {sum(len(cols) for cols in changed_cells.values())}")
                if generate_log:
                    print(f"   Detailed change log: {log_path}")
            
            return success
            
        except Exception as e:
            print(f"   Error in latest date_created comparison process: {e}")
            return False
    
    def generate_latest_date_created_change_log(self, summary_old_path: str, summary_new_path: str, 
                                              comparison_results: Dict, log_file_path: str = None) -> str:
        """
        Generate a detailed log file for latest date_created comparison
        
        Args:
            summary_old_path: Path to previous period Summary file
            summary_new_path: Path to current period Summary file
            comparison_results: Results from compare_summary_files_with_latest_date_created
            log_file_path: Optional custom path for log file
        
        Returns:
            Path to the generated log file
        """
        from datetime import datetime
        
        # Auto-generate log file path if not provided
        if log_file_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            old_name = os.path.splitext(os.path.basename(summary_old_path))[0]
            new_name = os.path.splitext(os.path.basename(summary_new_path))[0]
            log_file_path = f"latest_date_created_comparison_{old_name}_vs_{new_name}_{timestamp}.log"
        
        print(f"Generating latest date_created-based change log: {log_file_path}")
        
        try:
            # Load files for detailed analysis
            old_df = self.load_summary_file(summary_old_path)
            new_df = self.load_summary_file(summary_new_path)
            
            changed_rows = comparison_results['changed_rows']
            changed_cells = comparison_results['changed_cells']
            new_rows = comparison_results['new_rows']
            
            # Generate log content
            content = []
            content.append("=" * 100)
            content.append("LATEST DATE_CREATED DEDUPLICATION SUMMARY COMPARISON LOG")
            content.append("=" * 100)
            content.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            content.append(f"Previous File: {summary_old_path}")
            content.append(f"Current File:  {summary_new_path}")
            content.append(f"Deduplication Rule: Latest date_created for each entity key")
            content.append(f"Entity Keys: {', '.join(self.entity_key_columns)}")
            content.append("")
            
            # Summary statistics
            content.append("📊 SUMMARY STATISTICS")
            content.append("-" * 50)
            content.append(f"Total rows in previous file: {len(old_df)}")
            content.append(f"Total rows in current file:  {len(new_df)}")
            content.append(f"Entirely new entities: {len(new_rows)}")
            content.append(f"Entities with changes: {len(changed_rows) - len(new_rows)}")
            content.append(f"Total rows with changes: {len(changed_rows)}")
            content.append(f"Total individual cell changes: {sum(len(cols) for cols in changed_cells.values())}")
            content.append("")
            
            # New entities
            if new_rows:
                content.append("🆕 ENTIRELY NEW ENTITIES")
                content.append("-" * 50)
                
                for i, row_num in enumerate(sorted(new_rows), 1):
                    try:
                        # Get row data (convert from Excel row number to DataFrame index)
                        df_idx = row_num - 2
                        if df_idx < len(new_df):
                            row_data = new_df.iloc[df_idx]
                            
                            content.append(f"[{i}] ROW {row_num} - NEW ENTITY")
                            content.append(f"    Entity: Unit='{row_data.get('Unit name', '')}', "
                                         f"TenantID='{row_data.get('Tenant ID', '')}', "
                                         f"Tenant='{row_data.get('Tenant', '')}'")
                            content.append(f"    Date Created: '{row_data.get('Date created', '')}'")
                            content.append(f"    Highlighting: Yellow row")
                            content.append("")
                    except Exception as e:
                        content.append(f"[{i}] ROW {row_num} - Error reading data: {e}")
                        content.append("")
            
            # Changed entities (existing but modified)
            changed_entities = {row: cols for row, cols in changed_cells.items() if row not in new_rows}
            if changed_entities:
                content.append("📝 ENTITIES WITH CHANGES")
                content.append("-" * 50)
                
                for i, (row_num, changed_columns) in enumerate(sorted(changed_entities.items()), 1):
                    try:
                        # Get row data
                        df_idx = row_num - 2
                        if df_idx < len(new_df):
                            row_data = new_df.iloc[df_idx]
                            
                            content.append(f"[{i}] ROW {row_num} - MODIFIED ENTITY")
                            content.append(f"    Entity: Unit='{row_data.get('Unit name', '')}', "
                                         f"TenantID='{row_data.get('Tenant ID', '')}', "
                                         f"Tenant='{row_data.get('Tenant', '')}'")
                            content.append(f"    Date Created: '{row_data.get('Date created', '')}'")
                            content.append(f"    Modified Columns: {', '.join(changed_columns)}")
                            content.append(f"    Highlighting: Blue cells")
                            content.append("")
                    except Exception as e:
                        content.append(f"[{i}] ROW {row_num} - Error reading data: {e}")
                        content.append("")
            
            content.append("=" * 100)
            content.append("END OF LATEST DATE_CREATED COMPARISON LOG")
            content.append("=" * 100)
            
            # Write log file
            with open(log_file_path, 'w', encoding='utf-8') as f:
                f.write("\n".join(content))
            
            print(f"   Latest date_created change log generated: {log_file_path}")
            
            return log_file_path
            
        except Exception as e:
            print(f"   Error generating latest date_created change log: {e}")
            raise

    
    def apply_document_number_highlighting_to_summary(self, summary_path: str, comparison_result: Dict[str, any]) -> bool:
        """
        Apply highlighting to newly added Document Numbers (entire rows) in the Summary file
        
        Args:
            summary_path: Path to the T8 Summary file
            comparison_result: Result from compare_summary_files_by_document_number
        
        Returns:
            True if highlighting was successful, False otherwise
        """
        new_document_numbers = comparison_result.get('new_document_numbers', set())
        
        if not new_document_numbers:
            print("   No new document numbers to highlight in Summary file")
            return True
        
        try:
            print(f"🎨 Applying Document Number-based highlighting to Summary file")
            print(f"   Newly added Document Numbers: {len(new_document_numbers)}")
            
            # Open the file with xlwings
            app = xw.App(visible=False, add_book=False)
            try:
                wb = app.books.open(summary_path)
                sheet = wb.sheets[0]  # Assume first sheet
                
                # Get column mapping for the sheet
                header_row = sheet.range('1:1').value
                if not header_row:
                    print("   Could not read header row")
                    return False
                
                highlighted_rows = 0
                
                # Highlight newly added Document Numbers (entire rows) only
                for row_num in new_document_numbers:
                    try:
                        # Highlight entire row with light yellow background for new Document Numbers
                        row_range = sheet.range((row_num, 1), (row_num, len(header_row)))
                        row_range.color = (255, 255, 180)  # Light yellow for new Document Numbers
                        highlighted_rows += 1
                        print(f"   🆕 Highlighted entire row {row_num} (new Document Number)")
                    except Exception as e:
                        print(f"   Failed to highlight row {row_num}: {e}")
                
                # Save the file
                wb.save()
                print(f"   ✅ Saved Summary file with {highlighted_rows} highlighted rows")
                
                return True
                
            finally:
                wb.close()
                app.quit()
                
        except Exception as e:
            print(f"   ❌ Error applying Document Number highlighting to Summary file: {e}")
            return False

    
    def process_summary_comparison_with_document_numbers(self, summary_old_path: str, summary_new_path: str, generate_log: bool = True, log_file_path: str = None) -> bool:
        """
        Complete workflow using NEW Document Number-based comparison logic
        
        This method implements the new requirements:
        1. Use Document Number as the key identifier
        2. Highlight newly added Document Numbers (entire rows)
        3. Map entities using predefined keys
        4. Simplify review process by only flagging genuinely new entries
        
        Args:
            summary_old_path: Path to previous period Summary file
            summary_new_path: Path to current period Summary file
            generate_log: Whether to generate detailed change log file
            log_file_path: Optional custom path for log file
        
        Returns:
            True if process completed successfully
        """
        try:
            # Validate files exist
            if not os.path.exists(summary_old_path):
                print(f"   Previous file not found: {summary_old_path}")
                return False
            
            if not os.path.exists(summary_new_path):
                print(f"   Current file not found: {summary_new_path}")
                return False
            
            print(f"🔧 Starting Document Number-based comparison workflow")
            
            # Use the new Document Number-based comparison logic
            comparison_results = self.compare_summary_files_by_document_number(summary_old_path, summary_new_path)
            
            new_document_numbers = comparison_results['new_document_numbers']
            changed_rows = comparison_results['changed_rows-']
            changed_cells = comparison_results['changed_cells']
            entity_mapping = comparison_results['entity_mapping']
            
            # Generate detailed log if requested (using new format)
            if generate_log:
                log_path = self.generate_document_number_change_log(
                    summary_old_path, summary_new_path, comparison_results, log_file_path
                )
                print(f"   📝 Detailed log saved: {log_path}")
            
            # Apply Document Number-based highlighting
            success = self.apply_document_number_highlighting_to_summary(summary_new_path, comparison_results)
            
            if success:
                print(f"   🎉 Document Number-based comparison completed successfully!")
                print(f"   🆕 Newly added Document Numbers: {len(new_document_numbers)}")
                print(f"   📊 Total rows with changes: {len(changed_rows)}")
                print(f"   🏢 Entities with new Document Numbers: {len(entity_mapping)}")
                if generate_log:
                    print(f"   📝 Detailed change log: {log_path}")
            
            return success
            
        except Exception as e:
            print(f"   ❌ Error in Document Number-based comparison process: {e}")
            return False

    
    def generate_document_number_change_log(self, summary_old_path: str, summary_new_path: str, comparison_results: Dict, log_file_path: str = None) -> str:
        """
        Generate a detailed log file for Document Number-based comparison
        
        Args:
            summary_old_path: Path to previous period Summary file
            summary_new_path: Path to current period Summary file
            comparison_results: Results from compare_summary_files_by_document_number
            log_file_path: Optional custom path for log file
        
        Returns:
            Path to the generated log file
        """
        from datetime import datetime
        
        # Auto-generate log file path if not provided
        if log_file_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            old_name = os.path.splitext(os.path.basename(summary_old_path))[0]
            new_name = os.path.splitext(os.path.basename(summary_new_path))[0]
            log_file_path = f"document_number_comparison_{old_name}_vs_{new_name}_{timestamp}.log"
        
        print(f"📝 Generating Document Number-based change log: {log_file_path}")
        
        try:
            # Load files for detailed analysis
            old_df = self.load_summary_file(summary_old_path)
            new_df = self.load_summary_file(summary_new_path)
            
            new_document_numbers = comparison_results['new_document_numbers']
            entity_mapping = comparison_results['entity_mapping']
            changed_cells = comparison_results['changed_cells']
            
            # Generate log content
            content = []
            content.append("=" * 100)
            content.append("DOCUMENT NUMBER-BASED SUMMARY COMPARISON LOG")
            content.append("=" * 100)
            content.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            content.append(f"Previous File: {summary_old_path}")
            content.append(f"Current File:  {summary_new_path}")
            content.append("")
            
            # Summary statistics
            content.append("📊 SUMMARY STATISTICS")
            content.append("-" * 50)
            content.append(f"Total rows in previous file: {len(old_df)}")
            content.append(f"Total rows in current file:  {len(new_df)}")
            content.append(f"Newly added Document Numbers: {len(new_document_numbers)}")
            content.append(f"Entities with new Document Numbers: {len(entity_mapping)}")
            content.append(f"Total rows with changes: {len(changed_cells)}")
            content.append("")
            
            # Entity mapping summary
            if entity_mapping:
                content.append("🏢 ENTITY MAPPING SUMMARY")
                content.append("-" * 50)
                for entity_key, mapping_info in entity_mapping.items():
                    entity_info = mapping_info['entity_info']
                    old_docs = mapping_info['old_document_numbers']
                    new_docs = mapping_info['new_document_numbers']
                    
                    content.append(f"Entity: {entity_info['Unit name']} | {entity_info['Tenant ID']} | {entity_info['Tenant']}")
                    content.append(f"  Previous Document Numbers: {len(old_docs)} ({', '.join(old_docs) if old_docs else 'None'})")
                    content.append(f"  New Document Numbers: {len(new_docs)} ({', '.join(new_docs)})")
                    content.append("")
            
            # Detailed new Document Numbers
            content.append("🆕 NEWLY ADDED DOCUMENT NUMBERS")
            content.append("-" * 50)
            
            if not new_document_numbers:
                content.append("No new Document Numbers found.")
            else:
                for i, row_num in enumerate(sorted(new_document_numbers), 1):
                    try:
                        # Get row data (convert from Excel row number to DataFrame index)
                        df_idx = row_num - 2
                        if df_idx < len(new_df):
                            row_data = new_df.iloc[df_idx]
                            doc_num = self.get_document_number(row_data)
                            
                            content.append(f"[{i}] ROW {row_num} - NEW DOCUMENT NUMBER")
                            content.append(f"    Document Number: '{doc_num}'")
                            content.append(f"    Entity: Unit='{row_data.get('Unit name', '')}', "
                                         f"TenantID='{row_data.get('Tenant ID', '')}', "
                                         f"Tenant='{row_data.get('Tenant', '')}'")
                            
                            # Show key values for new Document Number
                            content.append("    Key Values:")
                            for col in self.tracked_columns.keys():
                                if col in new_df.columns:
                                    value = str(row_data.get(col, ''))
                                    if value.strip():
                                        content.append(f"      • {col}: '{value}'")
                            content.append("")
                    except Exception as e:
                        content.append(f"[{i}] ROW {row_num} - Error reading data: {e}")
                        content.append("")
            
            # Changes in existing Document Numbers
            existing_doc_changes = {row: cols for row, cols in changed_cells.items() 
                                  if row not in new_document_numbers}
            
            if existing_doc_changes:
                content.append("📝 CHANGES IN EXISTING DOCUMENT NUMBERS")
                content.append("-" * 50)
                
                for i, (row_num, changed_columns) in enumerate(sorted(existing_doc_changes.items()), 1):
                    try:
                        # Get row data
                        df_idx = row_num - 2
                        if df_idx < len(new_df):
                            row_data = new_df.iloc[df_idx]
                            doc_num = self.get_document_number(row_data)
                            
                            content.append(f"[{i}] ROW {row_num} - MODIFIED DOCUMENT NUMBER")
                            content.append(f"    Document Number: '{doc_num}'")
                            content.append(f"    Entity: Unit='{row_data.get('Unit name', '')}', "
                                         f"TenantID='{row_data.get('Tenant ID', '')}', "
                                         f"Tenant='{row_data.get('Tenant', '')}'")
                            content.append(f"    Modified Columns: {', '.join(changed_columns)}")
                            content.append("")
                    except Exception as e:
                        content.append(f"[{i}] ROW {row_num} - Error reading data: {e}")
                        content.append("")
            
            content.append("=" * 100)
            content.append("END OF DOCUMENT NUMBER COMPARISON LOG")
            content.append("=" * 100)
            
            # Write log file
            with open(log_file_path, 'w', encoding='utf-8') as f:
                f.write("\n".join(content))
            
            print(f"   ✅ Document Number change log generated: {log_file_path}")
            
            return log_file_path
            
        except Exception as e:
            print(f"   ❌ Error generating Document Number change log: {e}")
            raise
    
    def process_summary_comparison(self, summary_old_path: str, summary_new_path: str, generate_log: bool = True, log_file_path: str = None) -> bool:
        """
        Complete workflow: compare summary_old vs summary_new and highlight changes in summary_new
        
        Args:
            summary_old_path: Path to previous period Summary file
            summary_new_path: Path to current period Summary file
            generate_log: Whether to generate detailed change log file
            log_file_path: Optional custom path for log file
        
        Returns:
            True if process completed successfully
        """
        try:
            # Validate files exist
            if not os.path.exists(summary_old_path):
                print(f"   Previous file not found: {summary_old_path}")
                return False
            
            if not os.path.exists(summary_new_path):
                print(f"   Current file not found: {summary_new_path}")
                return False
            
            # Generate detailed log if requested
            if generate_log:
                log_path = self.generate_detailed_change_log(summary_old_path, summary_new_path, log_file_path)
                print(f"   Detailed log saved: {log_path}")
            
            # Compare files for highlighting (now returns detailed comparison results)
            comparison_results = self.compare_summary_files(summary_old_path, summary_new_path)
            changed_rows = comparison_results['changed_rows']
            changed_cells = comparison_results['changed_cells']
            
            # Apply cell-specific highlighting
            success = self.apply_highlighting_to_summary(summary_new_path, comparison_results)
            
            if success:
                print(f"   Summary comparison completed successfully!")
                print(f"   Total rows with changes: {len(changed_rows)}")
                print(f"   Total individual cells highlighted: {sum(len(cols) for cols in changed_cells.values())}")
                if generate_log:
                    print(f"   Detailed change log: {log_path}")
            
            return success
            
        except Exception as e:
            print(f"   Error in summary comparison process: {e}")
            return False

    def process_enhanced_summary_comparison(self, summary_old_path: str, summary_new_path: str, 
                                          output_dir: str = None) -> bool:
        """
        Process summary comparison with enhanced logic using oldest matching records
        and cell-level highlighting with different colors for different change types.
        
        Args:
            summary_old_path: Path to previous period Summary file
            summary_new_path: Path to current period Summary file  
            output_dir: Optional output directory for change logs
        
        Returns:
            True if comparison was successful, False otherwise
        """
        try:
            print(f"\n=== Enhanced Summary Comparison ===")
            
            # Perform enhanced comparison
            comparison_result = self.compare_summary_files_with_oldest_match(
                summary_old_path, summary_new_path
            )
            
            # Apply enhanced highlighting
            highlighting_success = self.apply_enhanced_highlighting_to_summary(
                summary_new_path, comparison_result
            )
            
            if not highlighting_success:
                print("   Warning: Failed to apply highlighting")
            
            # Generate change log if output directory is provided
            if output_dir:
                log_success = self.generate_enhanced_change_log(
                    summary_old_path, summary_new_path, comparison_result, output_dir
                )
                if not log_success:
                    print("   Warning: Failed to generate change log")
            
            return True
            
        except Exception as e:
            print(f"   Error in enhanced summary comparison: {e}")
            return False
    
    def generate_enhanced_change_log(self, summary_old_path: str, summary_new_path: str, 
                                   comparison_result: Dict[str, any], output_dir: str) -> bool:
        """
        Generate detailed change log for enhanced comparison results.
        
        Args:
            summary_old_path: Path to previous period Summary file
            summary_new_path: Path to current period Summary file
            comparison_result: Result from compare_summary_files_with_oldest_match
            output_dir: Directory to save the change log
        
        Returns:
            True if log generation was successful, False otherwise
        """
        try:
            # Load the new file to get row data
            new_df = self.load_summary_file(summary_new_path)
            
            # Create timestamp for log filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            log_filename = f"enhanced_summary_changes_{timestamp}.log"
            log_path = os.path.join(output_dir, log_filename)
            
            # Ensure output directory exists
            os.makedirs(output_dir, exist_ok=True)
            
            with open(log_path, 'w', encoding='utf-8') as log_file:
                log_file.write(f"Enhanced Summary Comparison Change Log\n")
                log_file.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                log_file.write(f"Previous File: {summary_old_path}\n")
                log_file.write(f"Current File: {summary_new_path}\n")
                log_file.write(f"Comparison Method: Oldest matching record by entity key\n")
                log_file.write(f"Entity Key Columns: {', '.join(self.entity_key_columns)}\n")
                log_file.write("=" * 80 + "\n\n")
                
                changed_cells = comparison_result.get('changed_cells', {})
                new_rows = comparison_result.get('new_rows', set())
                
                # Log cell-level changes
                if changed_cells:
                    log_file.write(f"CELL-LEVEL CHANGES ({len(changed_cells)} rows):\n")
                    log_file.write("-" * 50 + "\n")
                    
                    for row_num in sorted(changed_cells.keys()):
                        if row_num - 2 < len(new_df):  # Convert back to 0-indexed
                            row_data = new_df.iloc[row_num - 2]
                            entity_key = self.normalize_entity_key(row_data)
                            doc_num = self.get_document_number(row_data)
                            
                            log_file.write(f"Row {row_num}: {entity_key}\n")
                            log_file.write(f"  Document Number: {doc_num or 'N/A'}\n")
                            log_file.write(f"  Changed Columns: {', '.join(sorted(changed_cells[row_num]))}\n")
                            log_file.write(f"  Highlighting: Light blue cells\n\n")
                
                # Log entirely new rows
                if new_rows:
                    log_file.write(f"ENTIRELY NEW ROWS ({len(new_rows)} rows):\n")
                    log_file.write("-" * 50 + "\n")
                    
                    for row_num in sorted(new_rows):
                        if row_num - 2 < len(new_df):  # Convert back to 0-indexed
                            row_data = new_df.iloc[row_num - 2]
                            entity_key = self.normalize_entity_key(row_data)
                            doc_num = self.get_document_number(row_data)
                            
                            log_file.write(f"Row {row_num}: {entity_key}\n")
                            log_file.write(f"  Document Number: {doc_num or 'N/A'}\n")
                            log_file.write(f"  Status: No matching entity key in previous file\n")
                            log_file.write(f"  Highlighting: Yellow row\n\n")
                
                # Summary
                total_changes = len(changed_cells) + len(new_rows)
                total_cell_changes = sum(len(cols) for cols in changed_cells.values())
                
                log_file.write("SUMMARY:\n")
                log_file.write("-" * 20 + "\n")
                log_file.write(f"Total affected rows: {total_changes}\n")
                log_file.write(f"Rows with cell changes: {len(changed_cells)}\n")
                log_file.write(f"Entirely new rows: {len(new_rows)}\n")
                log_file.write(f"Individual cell changes: {total_cell_changes}\n")
            
            print(f"   Enhanced change log saved to: {log_path}")
            return True
            
        except Exception as e:
            print(f"   Error generating enhanced change log: {e}")
            return False
    
    def _is_within_90_days(self, date_str: str) -> bool:
        """Check if the given date is 90 days before today or any day after today"""
        parsed_date = self._parse_date_flexible(date_str)
        if parsed_date is None:
            return False

        today = datetime.now()
        lower_bound = today - timedelta(days=90)

        return parsed_date >= lower_bound
        
    def _parse_date_flexible(self, date_str: str) -> Optional[datetime]:
        """Parse date from multiple formats commonly found in Excel"""
        if not date_str or pd.isna(date_str):
            return None
            
        date_str = str(date_str).strip()
        if not date_str or date_str.lower() in ['', 'nan', 'none', 'null']:
            return None
        
        # For summary data, prioritize mm/dd/yyyy format but also support other common formats
        date_formats = [
            "%m/%d/%Y",              # mm/dd/yyyy (US format: month/day/year) - PRIORITY for summary
            "%m/%d/%y",              # mm/dd/yy (US format with 2-digit year)
            "%Y-%m-%d %H:%M:%S",     # 2022-10-30 00:00:00 (datetime format from summary)
            "%Y-%m-%d",              # 2022-10-30 (ISO date format)
            "%d-%b-%y",              # 8-Mar-24
            "%d-%B-%y",              # 8-March-24
            "%d-%b-%Y",              # 8-Mar-2024
            "%d-%B-%Y",              # 8-March-2024
            "%Y/%m/%d",              # 2023/03/08
        ]
        
        # Try each format in order - mm/dd/yyyy will be tried first
        for fmt in date_formats:
            try:
                return datetime.strptime(date_str, fmt)
            except ValueError:
                continue
        
        try:
            # Force US locale interpretation for mm/dd/yyyy
            parsed = pd.to_datetime(date_str, errors='coerce')
            if pd.notna(parsed):
               return parsed.to_pydatetime()
        except:
            pass
        
        # Handle Excel serial dates
        try:
            float_val = float(date_str)
            if 1 <= float_val <= 100000:
                base_date = datetime(1899, 12, 30)
                return base_date + timedelta(days=float_val)
        except:
            pass
        
        return None

    def _get_entity_column_mapping(self, summary_headers: List[str]) -> Dict[str, str]:
        """
        Map summary columns to entity model column positions.
        Starting from "Leasing period" column (Item2) instead of column 1.
        
        Returns:
            Dictionary mapping entity columns to summary columns
        """
        # Entity model columns starting from "Leasing period" position (column 5 onwards)
        entity_columns = [
            'Item2',                     # 5  - Always "Leasing period" (START HERE)
            'Note',                      # 6  - Always "Committed"
            'Region',                    # 7  - From summary
            'Product type',              # 8  - From summary
            'Factory code',              # 9  - From summary (Unit name)
            'Tenant code',               # 10 - From summary (Tenant ID)
            'Tenant name',               # 11 - From summary (Tenant)
            'Existing/New/Exp/Renew',    # 12 - Fixed as "New"
            'Tenant Industry',           # 13 - From summary (Industry)
            'Desk',                      # 14 - Empty
            'Broker',                    # 15 - Empty
            'GLA',                       # 16 - From summary
            'Rent',                      # 17 - From summary (VND)
            'EX rate',                   # 18 - Empty
            'Rent',                      # 19 - From summary (USD) 
            'Deposit',                   # 20 - From summary
            'Growth rate (Act)',         # 21 - From summary (Escalation rate)
            'IPA date',                  # 22 - Empty
            'PLC date',                  # 23 - Empty
            'Handover',                  # 24 - Fixed as "N"
            'Start date',                # 25 - From summary
            'End date',                  # 26 - From summary
            'Term',                      # 27 - From summary
            'Payment term',              # 28 - From summary
            'Fitting out',               # 29 - From summary (Fitout & rent-free)
            'Rent free',                 # 30 - Empty
            'Service charge',            # 31 - From summary
            'Commission',                # 32 - From summary (Commission fee)
            'Frequently'                 # 33 - Empty (END HERE)
        ]
        
        # Map to available summary columns
        column_mapping = {}
        
        # Fixed values for entity model
        fixed_values = {
            'Item2': 'Leasing period',                        # Fixed value (STARTING COLUMN)
            'Note': 'Committed',                              # Fixed value  
            'Existing/New/Exp/Renew': 'New',                  # Fixed as "New"
            'Desk': '',                                       # Empty
            'Broker': '',                                     # Empty
            'EX rate': '',                                    # Empty
            'IPA date': '',                                   # Empty
            'PLC date': '',                                   # Empty
            'Handover': 'N',                                  # Fixed as "N"
            'Rent free': '',                                  # Empty
            'Frequently': ''                                  # Empty (ENDING COLUMN)
        }
        
        # Map summary columns to entity columns
        summary_to_entity = {
            # Map summary columns to entity positions (starting from Item2)
            'Region': 'Region',                              # 7
            'Product Type': 'Product type',                  # 8
            'Unit name': 'Factory code',                     # 9
            'Tenant ID': 'Tenant code',                      # 10
            'Tenant': 'Tenant name',                         # 11
            'Industry': 'Tenant Industry',                   # 13
            'GLA': 'GLA',                                    # 16
            'Rent (VND)_Item': 'Rent',                       # 17 (VND)
            'Rent (USD)_Item': 'Rent',                       # 19 (USD) - will be handled specially
            'Deposit': 'Deposit',                            # 20
            'Escalation rate': 'Growth rate (Act)',          # 21
            'Start date (for model)': 'Start date',          # 25
            'End date (for model)': 'End date',              # 26
            'Term': 'Term',                                  # 27
            'Payment term': 'Payment term',                  # 28
            'Fitout & rent-free (months)': 'Fitting out',    # 29
            'Service charge': 'Service charge',              # 31
            'Commission fee': 'Commission'                   # 32
        }
        
        # Add fixed values
        for entity_col, fixed_val in fixed_values.items():
            column_mapping[entity_col] = fixed_val
        
        # Add mapped columns
        for summary_col, entity_col in summary_to_entity.items():
            if summary_col in summary_headers:
                column_mapping[entity_col] = summary_col
        
        # Handle special case for Rent USD (column 19)
        if 'Rent (USD)_Item' in summary_headers:
            column_mapping['Rent_USD'] = 'Rent (USD)_Item'  # Special key for USD rent
        
        return column_mapping, entity_columns

    def _format_row_for_entity_model(self, row_data: pd.Series, headers: List[str]) -> List:
        """
        Format row data for output sheet using entity model structure.
        Creates columns starting from "Leasing period" column only.
        
        Args:
            row_data: Pandas Series containing row data
            headers: Column headers from summary
            
        Returns:
            Formatted row data as list starting from "Leasing period" column
        """
        column_mapping, entity_columns = self._get_entity_column_mapping(headers)
        formatted_row = []
        
        for i, entity_col in enumerate(entity_columns):
            if entity_col in column_mapping:
                mapped_value = column_mapping[entity_col]
                
                # Handle fixed values
                if mapped_value in ['Leasing period', 'Committed', 'New', 'N', '']:
                    formatted_row.append(mapped_value)
                    
                # Handle special case for Rent USD (column 19 in original, now adjusted index)
                elif entity_col == 'Rent' and 'Rent (USD)_Item' in entity_col:
                    if 'Rent_USD' in column_mapping:
                        usd_col = column_mapping['Rent_USD']  
                        if usd_col in headers:
                            value = row_data.get(usd_col, '')
                            formatted_row.append(self._format_numeric_value(value))
                        else:
                            formatted_row.append('')
                    else:
                        formatted_row.append('')
                
                # Handle dynamic values from row data
                elif mapped_value in headers:
                    value = row_data.get(mapped_value, '')
                    
                    # Format dates to MM/DD/YYYY
                    if 'date' in entity_col.lower() and value and str(value).strip():
                        parsed_date = self._parse_date_flexible(str(value))
                        if parsed_date:
                            formatted_row.append(parsed_date.strftime('%m/%d/%Y'))
                        else:
                            formatted_row.append(str(value))
                    
                    # Format numeric columns
                    elif entity_col in ['GLA', 'Rent', 'Deposit', 'Growth rate (Act)', 'Term', 'Payment term', 
                                       'Fitting out', 'Service charge', 'Commission']:
                        formatted_row.append(self._format_numeric_value(value))
                    
                    # Default: convert to string
                    else:
                        formatted_row.append(str(value) if pd.notna(value) and str(value).strip() != '' else '')
                else:
                    formatted_row.append('')
            else:
                # Column not mapped, add empty
                formatted_row.append('')
        
        return formatted_row  # Limit to exactly 33 columns
    
    def _format_numeric_value(self, value) -> str:
        """Helper method to format numeric values consistently"""
        if value and str(value).strip():
            try:
                val_str = str(value).replace(',', '').strip()
                if val_str and val_str != '' and val_str.lower() != 'nan':
                    val_float = float(val_str)
                    if val_float.is_integer():
                        return str(int(val_float))
                    else:
                        return str(val_float)
            except (ValueError, TypeError):
                pass
        return ''
    
    def apply_highlighting_to_summary_with_90_day_rule(self, summary_old_path: str, summary_new_path: str,
                                                       output_path: str = None) -> str:
        """
        Apply highlighting to a COPY of the summary new file (does not modify original):
        - YELLOW rows: NEW Document Number + Type + Item combinations (within 90 days)
        - BLUE cells: UPDATE cells in existing combinations
        - Respects 90-day filtering rule

        Args:
            summary_old_path: Path to old summary file for comparison
            summary_new_path: Path to the summary new file (will NOT be modified)
            output_path: Path where highlighted copy should be saved (if None, auto-generates)

        Returns:
            Path to the highlighted output file
        """
        print(f"Creating highlighted copy of: {summary_new_path}")

        try:
            # Create output path if not provided
            if output_path is None:
                base_name = os.path.splitext(os.path.basename(summary_new_path))[0]
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_path = f"highlighted_{timestamp}_{base_name}.xlsx"

            # Create a copy of the file to highlight (DO NOT MODIFY ORIGINAL)
            import shutil
            shutil.copy2(summary_new_path, output_path)
            print(f"   Created copy: {output_path}")

            # First get comparison results with 90-day rule
            comparison_results = self.compare_summary_files_by_document_item_key(
                summary_old_path,
                summary_new_path,
                apply_90_day_filter=True
            )

            new_rows_indices = comparison_results['new_rows_indices']
            changed_cells = comparison_results['changed_cells']

            if not new_rows_indices and not changed_cells:
                print("   No changes found - no highlighting needed")
                return output_path

            # Open THE COPY (not the original) with openpyxl for highlighting
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill

            try:
                wb = load_workbook(output_path)
                sheet = wb.active  # Get the active sheet

                print(f"   Highlighting {len(new_rows_indices)} new rows (yellow)")
                print(f"   Highlighting {len(changed_cells)} rows with cell changes (blue cells)")

                # Get total columns for full row highlighting
                last_col = sheet.max_column

                # Define fill colors
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow for new rows
                blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Light blue for changed cells

                # Apply yellow highlighting to new rows (full rows)
                for df_idx in new_rows_indices:
                    excel_row = df_idx + 2  # Convert pandas index to Excel row (+2 for header)
                    try:
                        for col in range(1, last_col + 1):
                            cell = sheet.cell(row=excel_row, column=col)
                            cell.fill = yellow_fill
                    except Exception as e:
                        print(f"   Warning: Could not highlight row {excel_row}: {e}")

                # Load summary file once for column mapping
                summary_df = self.load_summary_file(summary_new_path)
                headers = list(summary_df.columns)

                # Apply blue highlighting to changed cells
                for excel_row, changed_columns in changed_cells.items():
                    try:
                        for col_name in changed_columns:
                            if col_name in headers:
                                col_idx = headers.index(col_name) + 1  # +1 for Excel 1-based indexing
                                cell = sheet.cell(row=excel_row, column=col_idx)
                                cell.fill = blue_fill
                    except Exception as e:
                        print(f"   Warning: Could not highlight cells in row {excel_row}: {e}")

                # Save and close
                wb.save(output_path)
                wb.close()

                print("   Highlighting applied successfully!")
                return output_path

            except Exception as e:
                print(f"   Error during highlighting: {e}")
                if 'wb' in locals():
                    try:
                        wb.close()
                    except:
                        pass
                # Return the output path even if highlighting failed (file was copied)
                return output_path

        except Exception as e:
            print(f"   Error applying highlighting: {e}")
            raise

    def compare_summary_files_by_document_item_key(self, summary_old_path: str, summary_new_path: str, 
                                                  apply_90_day_filter: bool = True) -> Dict[str, any]:
        """
        NEW LOGIC: Compare summary files using Document Number + Type + Item as composite key

        This is the most accurate logic for identifying:
        - NEW ROWS: Document Number + Type + Item combinations that don't exist in old file (within 90-day rule)
        - UPDATE ROWS: Document Number + Type + Item combinations that exist but have changes in tracked columns

        Args:
            summary_old_path: Path to previous period Summary file
            summary_new_path: Path to current period Summary file
            apply_90_day_filter: Apply 90-day filtering rule on new rows (default True)

        Returns:
            Dict with:
            - 'new_rows_indices': list of DataFrame indices for completely new combinations
            - 'update_rows_indices': list of DataFrame indices for existing combinations with changes
            - 'unchanged_rows_indices': list of DataFrame indices for unchanged combinations
        """
        print(f"Comparing Summary files using Document Number + Type + Item key:")
        print(f"   Previous: {summary_old_path}")
        print(f"   Current:  {summary_new_path}")
        
        # Load both files
        old_df = self.load_summary_file(summary_old_path)
        new_df = self.load_summary_file(summary_new_path)
        
        print(f"   Previous rows: {len(old_df)}, Current rows: {len(new_df)}")

        # Use tracked columns for comparison (9 columns total)
        column_mapping = self.tracked_columns
        
        # Create Document Number + Type + Item key for old file lookup
        old_lookup = {}
        for idx, row in old_df.iterrows():
            doc_num = str(row.get('Document Number', '')).strip()
            type_val = str(row.get('Type', '')).strip()
            item = str(row.get('Item', '')).strip()

            if doc_num and type_val and item and doc_num != 'nan' and type_val != 'nan' and item != 'nan':
                key = f"{doc_num}|{type_val}|{item}"
                old_lookup[key] = row

        print(f"   Created old lookup with {len(old_lookup)} Document Number + Type + Item combinations")
        
        # Find Start date column for 90-day filtering
        start_date_col = None
        if apply_90_day_filter:
            for col in new_df.columns:
                col_lower = col.lower().strip()
                if any(keyword in col_lower for keyword in ['start']):
                    if any(keyword in col_lower for keyword in ['date']):
                        start_date_col = col
                        break
            
            if start_date_col:
                print(f"   Using '{start_date_col}' column for 90-day filtering")
            else:
                print("   No Start date column found - skipping 90-day filter")
                apply_90_day_filter = False
        
        # Analyze each row in new file
        new_rows_indices = []
        update_rows_indices = []
        unchanged_rows_indices = []
        filtered_out_count = 0
        changed_cells_details = {}  # Track changed cells for highlighting
        
        for idx, new_row in new_df.iterrows():
            doc_num = str(new_row.get('Document Number', '')).strip()
            type_val = str(new_row.get('Type', '')).strip()
            item = str(new_row.get('Item', '')).strip()

            if not doc_num or not type_val or not item or doc_num == 'nan' or type_val == 'nan' or item == 'nan':
                continue

            key = f"{doc_num}|{type_val}|{item}"
            
            if key in old_lookup:
                # Existing combination - check for changes
                old_row = old_lookup[key]
                has_changes = False
                changed_columns = []
                
                for new_col, old_col in column_mapping.items():
                    if new_col in new_df.columns and old_col in old_df.columns:
                        new_val = self._normalize_value(new_row.get(new_col), new_col)
                        old_val = self._normalize_value(old_row.get(old_col), old_col)
                        
                        if new_val != old_val:
                            has_changes = True
                            changed_columns.append(new_col)
                
                if has_changes:
                    update_rows_indices.append(idx)
                    # Track changed cells for highlighting (convert idx to Excel row number)
                    excel_row = idx + 2  # +2 because pandas is 0-indexed, Excel starts at 1, and has header
                    changed_cells_details[excel_row] = changed_columns
                else:
                    unchanged_rows_indices.append(idx)
            else:
                # New combination - apply 90-day filter if enabled
                if apply_90_day_filter and start_date_col:
                    start_date_val = new_row.get(start_date_col)
                    if self._is_within_90_days(str(start_date_val)):
                        new_rows_indices.append(idx)
                    else:
                        filtered_out_count += 1
                else:
                    # No 90-day filter or no start date column
                    new_rows_indices.append(idx)
        
        print(f"   NEW Document Number + Type + Item combinations: {len(new_rows_indices)}")
        print(f"   UPDATED existing combinations: {len(update_rows_indices)}")
        print(f"   UNCHANGED combinations: {len(unchanged_rows_indices)}")
        if apply_90_day_filter and filtered_out_count > 0:
            print(f"   Filtered out by 90-day rule: {filtered_out_count}")
        
        return {
            'new_rows_indices': new_rows_indices,
            'update_rows_indices': update_rows_indices,
            'unchanged_rows_indices': unchanged_rows_indices,
            'changed_cells': changed_cells_details
        }
    
    def generate_excel_output_files(self, summary_old_path: str, summary_new_path: str, 
                                   output_file_path: str = None, use_document_item_key: bool = True) -> str:
        """
        Generate Excel output file with 'new_rows' and 'update_rows' sheets
        
        Args:
            summary_old_path: Path to previous period Summary file
            summary_new_path: Path to current period Summary file
            output_file_path: Optional output file path. If None, auto-generates based on input files
            use_document_item_key: Use Document Number + Type + Item key comparison (NEW RECOMMENDED)
        
        Returns:
            Path to the generated Excel file
        """
        from datetime import datetime
        
        # Auto-generate output file path if not provided
        if output_file_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            old_name = os.path.splitext(os.path.basename(summary_old_path))[0]
            new_name = os.path.splitext(os.path.basename(summary_new_path))[0]
            output_file_path = f"comparison_output_{old_name}_vs_{new_name}_{timestamp}.xlsx"
        
        print(f"Generating Excel output file: {output_file_path}")
        
        try:
            # Load both files
            old_df = self.load_summary_file(summary_old_path)
            new_df = self.load_summary_file(summary_new_path)
            
            # Use the NEW Document Number + Type + Item key logic
            if use_document_item_key:
                print("   Using NEW Document Number + Type + Item key comparison logic")
                comparison_results = self.compare_summary_files_by_document_item_key(summary_old_path, summary_new_path)
                new_rows_indices = comparison_results['new_rows_indices']
                update_rows_indices = comparison_results['update_rows_indices']
            else:
                # Fallback to old logic
                print("   Using legacy Document Number-only comparison logic")
                comparison_results = self.compare_summary_files_by_document_number(summary_old_path, summary_new_path)
                new_document_numbers = comparison_results['new_document_numbers']
                changed_cells = comparison_results['changed_cells']
                
                # Convert to indices
                new_rows_indices = [row_num - 2 for row_num in new_document_numbers if row_num - 2 >= 0 and row_num - 2 < len(new_df)]
                update_rows_indices = [row_num - 2 for row_num, changed_columns in changed_cells.items() 
                                     if row_num not in new_document_numbers and row_num - 2 >= 0 and row_num - 2 < len(new_df)]
            
            # Create new_rows and update_rows DataFrames with entity model formatting
            new_rows_df = new_df.iloc[new_rows_indices] if new_rows_indices else pd.DataFrame(columns=new_df.columns)
            update_rows_df = new_df.iloc[update_rows_indices] if update_rows_indices else pd.DataFrame(columns=new_df.columns)
            
            print(f"   New rows: {len(new_rows_df)}")
            print(f"   Update rows: {len(update_rows_df)}")
            
            # Apply entity model formatting to new_rows (similar to entity processing)
            if not new_rows_df.empty:
                print("   Applying entity model formatting to new rows...")
                summary_headers = list(new_rows_df.columns)
                
                # Get entity column mapping and order
                column_mapping, entity_columns = self._get_entity_column_mapping(summary_headers)
                
                formatted_new_rows = []
                for idx, row in new_rows_df.iterrows():
                    formatted_row = self._format_row_for_entity_model(row, summary_headers)
                    formatted_new_rows.append(formatted_row)
                
                # Convert to DataFrame with entity column order for easy copy-paste
                new_rows_df = pd.DataFrame(formatted_new_rows, columns=entity_columns)
                print(f"   Formatted {len(new_rows_df)} new rows with entity model structure")
            
            # Create Excel file with multiple sheets
            with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
                # Write new_rows sheet with entity model formatting
                new_rows_df.to_excel(writer, sheet_name='new_rows', index=False)
                
                # Write update_rows sheet (standard formatting)
                update_rows_df.to_excel(writer, sheet_name='update_rows', index=False)                
                # Add a summary sheet with detailed info
                summary_data = {
                    'Metric': [
                        'Total Rows in Previous File',
                        'Total Rows in Current File',
                        'New Rows (Document Number + Type + Item not in old, within 90 days)',
                        'Updated Rows (existing combinations with changes)',
                        'Unchanged Rows',
                        'Comparison Method',
                        'Key Used',
                        '90-Day Filter Applied',
                        'New Rows Format',
                        'Generated At'
                    ],
                    'Value': [
                        len(old_df),
                        len(new_df),
                        len(new_rows_df),
                        len(update_rows_df),
                        len(new_df) - len(new_rows_df) - len(update_rows_df),
                        'Document Number + Type + Item Key with 90-day filter' if use_document_item_key else 'Document Number Only',
                        'Document Number + Type + Item' if use_document_item_key else 'Document Number',
                        'Yes (+-90 days from today)' if use_document_item_key else 'No',
                        'Entity Model Formatting Applied' if use_document_item_key else 'Standard',
                        datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    ]
                }
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='summary', index=False)
            
            print(f"   Excel output file generated: {output_file_path}")
            return output_file_path
            
        except Exception as e:
            print(f"   Error generating Excel output file: {e}")
            raise

    def _extract_month_from_filename(self, filename: str) -> str:
        """
        Extract month identifier from filename for output naming.
        Looks for patterns like T9, T10, T11, or month names.

        Args:
            filename: The filename to extract month from

        Returns:
            Month identifier string (e.g., 'T9', 'T10', 'Nov')
        """
        import re

        # Try to find T+number pattern (e.g., T9, T10, T11)
        t_match = re.search(r'[Tt](\d{1,2})', filename)
        if t_match:
            return f"T{t_match.group(1)}"

        # Try month abbreviations
        month_patterns = [
            (r'[Jj]an', 'Jan'), (r'[Ff]eb', 'Feb'), (r'[Mm]ar', 'Mar'),
            (r'[Aa]pr', 'Apr'), (r'[Mm]ay', 'May'), (r'[Jj]un', 'Jun'),
            (r'[Jj]ul', 'Jul'), (r'[Aa]ug', 'Aug'), (r'[Ss]ep', 'Sep'),
            (r'[Oo]ct', 'Oct'), (r'[Nn]ov', 'Nov'), (r'[Dd]ec', 'Dec')
        ]

        for pattern, month in month_patterns:
            if re.search(pattern, filename):
                return month

        # Fallback: use first part of filename
        base = os.path.splitext(os.path.basename(filename))[0]
        return base[:10] if len(base) > 10 else base

    def _get_project_phase_statistics(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Generate statistics grouped by Project and Phase.

        Args:
            df: DataFrame containing the data

        Returns:
            DataFrame with Project/Phase statistics
        """
        # Find Project and Phase columns (case-insensitive)
        project_col = None
        phase_col = None

        for col in df.columns:
            col_lower = col.lower().strip()
            if 'project' in col_lower and project_col is None:
                project_col = col
            if 'phase' in col_lower and phase_col is None:
                phase_col = col

        if not project_col:
            print("   Warning: No 'Project' column found for statistics")
            return pd.DataFrame()

        # Create grouping columns list
        group_cols = [project_col]
        if phase_col:
            group_cols.append(phase_col)

        # Group and count
        try:
            stats = df.groupby(group_cols).size().reset_index(name='Row Count')

            # Add GLA sum if available
            gla_col = None
            for col in df.columns:
                if 'gla' in col.lower():
                    gla_col = col
                    break

            if gla_col:
                gla_stats = df.groupby(group_cols)[gla_col].sum().reset_index()
                gla_stats.columns = group_cols + ['Total GLA']
                stats = stats.merge(gla_stats, on=group_cols, how='left')

            return stats
        except Exception as e:
            print(f"   Warning: Could not generate Project/Phase statistics: {e}")
            return pd.DataFrame()

    def generate_unified_comparison_output(self, summary_old_path: str, summary_new_path: str,
                                          output_dir: str = None) -> str:
        """
        Generate a UNIFIED comparison output file with all improvements:

        1. MERGED FILE: Single output file with highlighted data + comparison sheets
        2. ENHANCED HIGHLIGHTING:
           - Light yellow background for entire updated rows
           - Different color (orange/darker) for specific changed columns
           - Yellow for new rows
        3. PROJECT/PHASE STATISTICS: Summary table broken down by Project and Phase
        4. ENHANCED NEW_ROWS SHEET: Includes Document Number, Item, Type columns
        5. SMART FILENAME: Auto-named as "Leasing SS_Comparison [PrevMonth] – [CurrMonth].xlsx"

        Args:
            summary_old_path: Path to previous period Summary file
            summary_new_path: Path to current period Summary file
            output_dir: Directory to save output file (defaults to current directory)

        Returns:
            Path to the generated unified Excel file
        """
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        import shutil

        print("=" * 60)
        print("GENERATING UNIFIED COMPARISON OUTPUT (v2.0)")
        print("=" * 60)

        try:
            # Load both files
            old_df = self.load_summary_file(summary_old_path)
            new_df = self.load_summary_file(summary_new_path)

            print(f"   Previous file: {len(old_df)} rows")
            print(f"   Current file: {len(new_df)} rows")

            # === IMPROVEMENT 5: Smart filename with month range ===
            old_month = self._extract_month_from_filename(summary_old_path)
            new_month = self._extract_month_from_filename(summary_new_path)

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"Leasing SS_Comparison {old_month} – {new_month}_{timestamp}.xlsx"

            if output_dir:
                output_path = os.path.join(output_dir, output_filename)
            else:
                output_path = output_filename

            print(f"   Output filename: {output_filename}")

            # Get comparison results
            comparison_results = self.compare_summary_files_by_document_item_key(
                summary_old_path, summary_new_path, apply_90_day_filter=True
            )

            new_rows_indices = comparison_results['new_rows_indices']
            update_rows_indices = comparison_results['update_rows_indices']
            unchanged_rows_indices = comparison_results['unchanged_rows_indices']
            changed_cells = comparison_results['changed_cells']

            print(f"   New rows: {len(new_rows_indices)}")
            print(f"   Updated rows: {len(update_rows_indices)}")
            print(f"   Unchanged rows: {len(unchanged_rows_indices)}")

            # === IMPROVEMENT 1: Create merged file (copy of new file as base) ===
            shutil.copy2(summary_new_path, output_path)
            print("   Created base file from current month data")

            # Open the file for modifications
            wb = load_workbook(output_path)

            # Rename first sheet to 'Highlighted Data'
            ws_main = wb.active
            ws_main.title = 'Highlighted Data'

            # === IMPROVEMENT 2: Enhanced highlighting ===
            # Define colors
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # New rows
            blue_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")  # Updated row background (light blue)
            red_font = Font(color="FF0000", bold=True)  # Changed cells - red text
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Header

            last_col = ws_main.max_column

            # Apply yellow highlighting to NEW rows (entire row)
            print(f"   Applying yellow highlight to {len(new_rows_indices)} new rows...")
            for df_idx in new_rows_indices:
                excel_row = df_idx + 2  # Convert to Excel row
                try:
                    for col in range(1, last_col + 1):
                        cell = ws_main.cell(row=excel_row, column=col)
                        cell.fill = yellow_fill
                except Exception as e:
                    print(f"   Warning: Could not highlight new row {excel_row}: {e}")

            # Apply enhanced highlighting to UPDATED rows
            # Blue background for entire row, red text for changed cells
            print(f"   Applying enhanced highlight to {len(update_rows_indices)} updated rows...")
            headers = list(new_df.columns)

            for df_idx in update_rows_indices:
                excel_row = df_idx + 2
                try:
                    # First apply blue background to entire row
                    for col in range(1, last_col + 1):
                        cell = ws_main.cell(row=excel_row, column=col)
                        cell.fill = blue_fill

                    # Then apply red text to changed cells
                    if excel_row in changed_cells:
                        for col_name in changed_cells[excel_row]:
                            if col_name in headers:
                                col_idx = headers.index(col_name) + 1
                                cell = ws_main.cell(row=excel_row, column=col_idx)
                                cell.font = red_font
                except Exception as e:
                    print(f"   Warning: Could not highlight updated row {excel_row}: {e}")

            # === IMPROVEMENT 4: Create enhanced new_rows sheet with Document Number, Item, Type ===
            ws_new_rows = wb.create_sheet('new_rows')

            # Define columns to include (ensure Document Number, Item, Type are first)
            priority_cols = ['Document Number', 'Type', 'Item']
            other_cols = [col for col in new_df.columns if col not in priority_cols]
            ordered_cols = priority_cols + other_cols

            # Filter to only existing columns
            ordered_cols = [col for col in ordered_cols if col in new_df.columns]

            # Write headers
            for col_idx, col_name in enumerate(ordered_cols, 1):
                cell = ws_new_rows.cell(row=1, column=col_idx, value=col_name)
                cell.fill = header_fill
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal='center')

            # Write new rows data
            new_rows_df = new_df.iloc[new_rows_indices] if new_rows_indices else pd.DataFrame()
            for row_idx, (df_idx, row) in enumerate(new_rows_df.iterrows(), 2):
                for col_idx, col_name in enumerate(ordered_cols, 1):
                    value = row.get(col_name, '')
                    if pd.isna(value):
                        value = ''
                    ws_new_rows.cell(row=row_idx, column=col_idx, value=value)

            print(f"   Created new_rows sheet with {len(new_rows_df)} rows")

            # === Create update_rows sheet ===
            ws_update_rows = wb.create_sheet('update_rows')

            # Write headers
            for col_idx, col_name in enumerate(ordered_cols, 1):
                cell = ws_update_rows.cell(row=1, column=col_idx, value=col_name)
                cell.fill = header_fill
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal='center')

            # Write update rows data with changed cells highlighted
            update_rows_df = new_df.iloc[update_rows_indices] if update_rows_indices else pd.DataFrame()
            for row_idx, (df_idx, row) in enumerate(update_rows_df.iterrows(), 2):
                excel_row_in_main = df_idx + 2  # Row number in main sheet for changed_cells lookup
                changed_cols_for_row = changed_cells.get(excel_row_in_main, [])

                for col_idx, col_name in enumerate(ordered_cols, 1):
                    value = row.get(col_name, '')
                    if pd.isna(value):
                        value = ''
                    cell = ws_update_rows.cell(row=row_idx, column=col_idx, value=value)

                    # Highlight changed cells with red text
                    if col_name in changed_cols_for_row:
                        cell.font = red_font

            print(f"   Created update_rows sheet with {len(update_rows_df)} rows")

            # === IMPROVEMENT 3: Add Project/Phase statistics to summary sheet ===
            ws_summary = wb.create_sheet('Summary')

            # Overall statistics
            summary_data = [
                ['COMPARISON SUMMARY', ''],
                ['', ''],
                ['Previous File', os.path.basename(summary_old_path)],
                ['Current File', os.path.basename(summary_new_path)],
                ['', ''],
                ['Total Rows (Previous)', len(old_df)],
                ['Total Rows (Current)', len(new_df)],
                ['New Rows', len(new_rows_indices)],
                ['Updated Rows', len(update_rows_indices)],
                ['Unchanged Rows', len(unchanged_rows_indices)],
                ['', ''],
                ['Comparison Method', 'Document Number + Type + Item'],
                ['90-Day Filter', 'Applied'],
                ['Generated At', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ]

            for row_idx, row_data in enumerate(summary_data, 1):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                    if row_idx == 1:
                        cell.font = Font(bold=True, size=14)

            # Add Project/Phase statistics
            current_row = len(summary_data) + 3

            # Statistics for NEW rows
            ws_summary.cell(row=current_row, column=1, value='NEW ROWS BY PROJECT/PHASE').font = Font(bold=True, size=12)
            current_row += 1

            if new_rows_indices:
                new_rows_stats = self._get_project_phase_statistics(new_df.iloc[new_rows_indices])
                if not new_rows_stats.empty:
                    # Write headers
                    for col_idx, col_name in enumerate(new_rows_stats.columns, 1):
                        cell = ws_summary.cell(row=current_row, column=col_idx, value=col_name)
                        cell.fill = header_fill
                        cell.font = Font(bold=True, color="FFFFFF")
                    current_row += 1

                    # Write data
                    for _, row in new_rows_stats.iterrows():
                        for col_idx, value in enumerate(row, 1):
                            ws_summary.cell(row=current_row, column=col_idx, value=value)
                        current_row += 1
                else:
                    ws_summary.cell(row=current_row, column=1, value='No Project/Phase data available')
                    current_row += 1
            else:
                ws_summary.cell(row=current_row, column=1, value='No new rows')
                current_row += 1

            current_row += 2

            # Statistics for UPDATE rows
            ws_summary.cell(row=current_row, column=1, value='UPDATED ROWS BY PROJECT/PHASE').font = Font(bold=True, size=12)
            current_row += 1

            if update_rows_indices:
                update_rows_stats = self._get_project_phase_statistics(new_df.iloc[update_rows_indices])
                if not update_rows_stats.empty:
                    # Write headers
                    for col_idx, col_name in enumerate(update_rows_stats.columns, 1):
                        cell = ws_summary.cell(row=current_row, column=col_idx, value=col_name)
                        cell.fill = header_fill
                        cell.font = Font(bold=True, color="FFFFFF")
                    current_row += 1

                    # Write data
                    for _, row in update_rows_stats.iterrows():
                        for col_idx, value in enumerate(row, 1):
                            ws_summary.cell(row=current_row, column=col_idx, value=value)
                        current_row += 1
                else:
                    ws_summary.cell(row=current_row, column=1, value='No Project/Phase data available')
                    current_row += 1
            else:
                ws_summary.cell(row=current_row, column=1, value='No updated rows')
                current_row += 1

            # Add legend
            current_row += 2
            ws_summary.cell(row=current_row, column=1, value='HIGHLIGHTING LEGEND').font = Font(bold=True, size=12)
            current_row += 1

            # Legend item 1: Yellow for new rows
            cell = ws_summary.cell(row=current_row, column=1, value='Yellow Row')
            cell.fill = yellow_fill
            ws_summary.cell(row=current_row, column=2, value='New row (Document Number + Type + Item not in previous file)')
            current_row += 1

            # Legend item 2: Blue for updated rows
            cell = ws_summary.cell(row=current_row, column=1, value='Blue Row')
            cell.fill = blue_fill
            ws_summary.cell(row=current_row, column=2, value='Updated row (existing combination with changes)')
            current_row += 1

            # Legend item 3: Red text for changed cells
            cell = ws_summary.cell(row=current_row, column=1, value='Red Text')
            cell.fill = blue_fill
            cell.font = red_font
            ws_summary.cell(row=current_row, column=2, value='Specific cell that changed value')
            current_row += 1

            # Adjust column widths
            ws_summary.column_dimensions['A'].width = 30
            ws_summary.column_dimensions['B'].width = 60

            # Save the workbook
            wb.save(output_path)
            wb.close()

            print("=" * 60)
            print(f"UNIFIED OUTPUT GENERATED: {output_path}")
            print("=" * 60)

            return output_path

        except Exception as e:
            print(f"   Error generating unified comparison output: {e}")
            import traceback
            traceback.print_exc()
            raise