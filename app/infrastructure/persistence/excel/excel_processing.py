# app/excel_processing.py
"""Excel processing and formatting functions."""

from __future__ import annotations

import io
import re
import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from .data_utils import (
    detect_header_row, normalize_financial_col, promote_row8,
    fill_down_assign, coerce_numeric, aggregate_totals, DEFAULT_CONFIG
)

# ---------------------------
# Helpers for grouped sheets
# ---------------------------

def _short_title(name: str, CONFIG: dict = DEFAULT_CONFIG) -> str:
    # Excel sheet name rules
    name = re.sub(r'[:\\/?*\[\]]', '-', name).strip()
    return name[:CONFIG.get("max_sheet_name_length", 31)]

def _format_pct(pct):
    return "N/A" if pct is None else f"{pct:+.1f}%"

def _write_account_block(
    ws: Worksheet,
    row: int,
    account_name: str,
    account_data: dict,
    *,
    section_font,
    header_font,
    header_fill,
    thin_border,
    format_vnd,
    cost_mode: bool,  # True for 632/641/642 (up = bad/red), False for 511 (up = good/green)
) -> int:
    # Account header
    ws[f"A{row}"] = f"Account: {account_name}"
    ws[f"A{row}"].font = section_font
    row += 1

    biggest = account_data.get("biggest_change")
    if biggest:
        ws[f"A{row}"] = "Biggest Change"
        ws[f"B{row}"] = f"{biggest.get('from','')} → {biggest.get('to','')}"
        ws[f"C{row}"] = format_vnd(biggest.get('change', 0))
        ws[f"D{row}"] = _format_pct(biggest.get('pct_change'))
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = thin_border
        row += 1

    # MoM table
    headers = ["Period", "Prev (VND)", "Curr (VND)", "Δ (VND)", "Δ (%)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    row += 1

    for ch in account_data.get("changes", []):
        ws[f"A{row}"] = f"{ch.get('from','')} → {ch.get('to','')}"
        ws[f"B{row}"] = format_vnd(ch.get('prev_val', 0))
        ws[f"C{row}"] = format_vnd(ch.get('curr_val', 0))
        ws[f"D{row}"] = format_vnd(ch.get('change', 0))
        ws[f"E{row}"] = _format_pct(ch.get('pct_change'))

        delta = ch.get('change', 0)
        # polarity: revenue up = good; cost up = bad
        if delta != 0:
            good = (delta > 0) if not cost_mode else (delta < 0)
            fill = PatternFill(
                start_color=("E8F5E8" if good else "FFEBEE"),
                end_color=("E8F5E8" if good else "FFEBEE"),
                fill_type="solid",
            )
        else:
            fill = None

        for col in range(1, 6):
            c = ws.cell(row=row, column=col)
            if fill:
                c.fill = fill
            c.border = thin_border
        row += 1

    # Top entity impacts
    impacts = account_data.get("customer_impacts") or account_data.get("entity_impacts") or []
    if impacts:
        row += 1
        ws[f"A{row}"] = "Top Entity Impacts:"
        ws[f"A{row}"].font = section_font
        row += 1

        heads = ["Entity", "Δ (VND)", "Δ (%)", "Prev", "Curr"]
        for col, h in enumerate(heads, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row += 1

        for imp in impacts:
            ws[f"A{row}"] = imp.get("entity", "")
            ws[f"B{row}"] = format_vnd(imp.get("change", 0))
            ws[f"C{row}"] = _format_pct(imp.get("pct_change"))
            ws[f"D{row}"] = format_vnd(imp.get("prev_val", 0))
            ws[f"E{row}"] = format_vnd(imp.get("curr_val", 0))

            delta = imp.get('change', 0)
            if delta != 0:
                good = (delta > 0) if not cost_mode else (delta < 0)
                fill = PatternFill(
                    start_color=("E8F5E8" if good else "FFEBEE"),
                    end_color=("E8F5E8" if good else "FFEBEE"),
                    fill_type="solid",
                )
            else:
                fill = None

            for col in range(1, 6):
                c = ws.cell(row=row, column=col)
                if fill:
                    c.fill = fill
                c.border = thin_border
            row += 1

    row += 2
    return row

def _autofit_ws(ws: Worksheet):
    for column in ws.columns:
        maxlen = 0
        col_letter = column[0].column_letter
        for cell in column:
            try:
                maxlen = max(maxlen, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(maxlen + 2, 50)

def create_group_sheet(
    wb,
    title: str,
    accounts: dict,
    *,
    section_font,
    header_font,
    header_fill,
    thin_border,
    format_vnd,
    cost_mode: bool,
):
    title = _short_title(title)
    ws = wb.create_sheet(title)
    row = 1
    ws[f"A{row}"] = title.upper()
    ws[f"A{row}"].font = section_font
    row += 2

    # sort by |biggest change| desc
    sorted_items = []
    for acct, data in accounts.items():
        big = data.get("biggest_change") or {}
        sorted_items.append((acct, abs(big.get("change", 0)), data))
    sorted_items.sort(key=lambda t: t[1], reverse=True)

    for acct, _, data in sorted_items:
        row = _write_account_block(
            ws, row, acct, data,
            section_font=section_font,
            header_font=header_font,
            header_fill=header_fill,
            thin_border=thin_border,
            format_vnd=format_vnd,
            cost_mode=cost_mode,
        )

    _autofit_ws(ws)

# --------------------------------
# Sheet load/normalize utilities
# --------------------------------

def process_financial_tab_from_bytes(
    xl_bytes: bytes,
    sheet_name: str,
    mode: str,
    subsidiary: str,
) -> tuple[pd.DataFrame, list[str]]:
    """Load and clean one sheet ('BS Breakdown' or 'PL Breakdown') from in-memory bytes."""
    header_row = detect_header_row(xl_bytes, sheet_name)
    df = pd.read_excel(io.BytesIO(xl_bytes), sheet_name=sheet_name, header=header_row, dtype=str)
    df = normalize_financial_col(df)
    df, month_cols = promote_row8(df, mode, subsidiary)
    df = fill_down_assign(df)
    df = coerce_numeric(df, month_cols)
    keep_cols = ["Account Code","Account Name","RowHadOwnCode","IsTotal"] + [c for c in month_cols if c in df.columns]
    df = df[keep_cols]
    totals = aggregate_totals(df, month_cols)
    return totals, month_cols

def extract_subsidiary_name_from_bytes(xl_bytes: bytes, fallback_filename: str) -> str:
    """Try to find a name on A2 of BS/PL sheets like 'Subsidiary: XYZ'. Fallback to filename stem."""
    try:
        from app.shared.utils.sheet_detection import detect_sheets_from_bytes

        # Use fuzzy matching to find BS/PL sheets
        bs_sheet, pl_sheet = detect_sheets_from_bytes(xl_bytes)

        wb = load_workbook(io.BytesIO(xl_bytes), read_only=True, data_only=True)

        # Try BS sheet first, then PL sheet
        for sheet_name in [bs_sheet, pl_sheet]:
            if sheet_name and sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                cell_value = sheet["A2"].value
                if isinstance(cell_value, str) and ":" in cell_value:
                    wb.close()
                    return cell_value.split(":")[-1].strip()
        wb.close()
    except Exception:
        pass
    # fallback: filename before first underscore or dot
    stem = fallback_filename.rsplit("/", 1)[-1]
    stem = stem.split("\\")[-1]
    stem = stem.split(".")[0]
    return stem.split("_")[0] if "_" in stem else stem

# -----------------------------------------------------------------------------
# Excel formatting (IN-MEMORY, works on a worksheet not a saved file)
# -----------------------------------------------------------------------------

def apply_excel_formatting_ws(ws, anomaly_df: pd.DataFrame, CONFIG: dict) -> None:
    """Apply simple conditional fills directly on the 'Anomalies Summary' worksheet."""
    try:
        critical_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        warning_fill  = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        header_fill   = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        ai_fill       = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")

        # Header
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True)

        # Find indexes
        headers = [c.value for c in ws[1]]
        try:
            abs_idx = headers.index("Abs Change (VND)") + 1
            trig_idx = headers.index("Trigger(s)") + 1
            status_idx = headers.index("Status") + 1
        except ValueError:
            return

        # Rows
        for row_idx in range(2, ws.max_row + 1):
            try:
                abs_change = ws.cell(row=row_idx, column=abs_idx).value or 0
                trigger = str(ws.cell(row=row_idx, column=trig_idx).value or "")
                status = str(ws.cell(row=row_idx, column=status_idx).value or "")

                fill = None
                if "AI Analysis" in status:
                    fill = ai_fill
                elif abs_change >= CONFIG.get("materiality_vnd", 1000000000) * 5:
                    fill = critical_fill
                elif "Correlation break" in trigger or abs_change >= CONFIG.get("materiality_vnd", 1000000000) * 2:
                    fill = warning_fill

                if fill:
                    for col_idx in range(1, len(headers) + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = fill
            except Exception:
                continue
    except Exception:
        # Formatting should never break
        pass

# -------------------------------------------
# Main: write summary sheet + grouped sheets
# -------------------------------------------

def _add_revenue_analysis_to_sheet(ws, revenue_analysis: dict):
    """Add revenue analysis data to an Excel worksheet in a structured format."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    section_font = Font(bold=True, size=12, color="2F5597")
    currency_font = Font(name="Arial", size=10)

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def format_vnd(amount):
        if isinstance(amount, (int, float)) and not pd.isna(amount):
            return f"{amount:,.0f} VND"
        return "N/A"

    row = 1

    # Title
    ws[f"A{row}"] = "COMPREHENSIVE REVENUE ANALYSIS"
    ws[f"A{row}"].font = Font(bold=True, size=16, color="2F5597")
    row += 2

    # Executive Summary
    if revenue_analysis.get('summary'):
        summary = revenue_analysis['summary']
        ws[f"A{row}"] = "EXECUTIVE SUMMARY"
        ws[f"A{row}"].font = section_font
        row += 1

        summary_data = [
            ["Subsidiary", revenue_analysis.get('subsidiary', 'N/A')],
            ["Months Analyzed", len(revenue_analysis.get('months_analyzed', []))],
            ["Revenue Accounts", summary.get('total_accounts', 0)],
            ["Latest Total Revenue", format_vnd(summary.get('total_revenue_latest', 0))],
            ["Latest Gross Margin %", f"{summary.get('gross_margin_latest', 0):.1f}%" if summary.get('gross_margin_latest') else 'N/A'],
            ["SG&A 641* Accounts", summary.get('total_sga_641_accounts', 0)],
            ["SG&A 642* Accounts", summary.get('total_sga_642_accounts', 0)],
            ["Latest Total SG&A", format_vnd(summary.get('total_sga_latest', 0))],
            ["Latest SG&A Ratio %", f"{summary.get('sga_ratio_latest', 0):.1f}%" if summary.get('sga_ratio_latest') else 'N/A']
        ]

        for label, value in summary_data:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            ws[f"A{row}"].font = Font(bold=True)
            row += 1
        row += 1

    # Risk Assessment
    if revenue_analysis.get('risk_assessment'):
        ws[f"A{row}"] = "RISK ASSESSMENT"
        ws[f"A{row}"].font = section_font
        row += 1

        # Headers
        headers = ["Period", "Risk Level", "Type", "Description"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row += 1

        for risk in revenue_analysis['risk_assessment']:
            ws[f"A{row}"] = risk.get('period', '')
            ws[f"B{row}"] = risk.get('risk_level', '')
            ws[f"C{row}"] = risk.get('type', '')
            ws[f"D{row}"] = risk.get('description', '')

            # Color code risk levels
            risk_level = risk.get('risk_level', '').lower()
            if risk_level == 'high':
                fill_color = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
            elif risk_level == 'medium':
                fill_color = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
            else:
                fill_color = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")

            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.fill = fill_color
                cell.border = thin_border
            row += 1
        row += 1

    # Total Revenue Trend
    if revenue_analysis.get('total_revenue_analysis', {}).get('changes'):
        ws[f"A{row}"] = "TOTAL REVENUE TREND (511*)"
        ws[f"A{row}"].font = section_font
        row += 1

        headers = ["Period", "Previous Value", "Current Value", "Change (VND)", "Change (%)", "Account Breakdown"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row += 1

        for change in revenue_analysis['total_revenue_analysis']['changes']:
            period = f"{change.get('from', '')} → {change.get('to', '')}"
            ws[f"A{row}"] = period
            ws[f"B{row}"] = format_vnd(change.get('prev_value', 0))
            ws[f"C{row}"] = format_vnd(change.get('curr_value', 0))
            ws[f"D{row}"] = format_vnd(change.get('change', 0))
            ws[f"E{row}"] = f"{change.get('pct_change', 0):+.1f}%"
            ws[f"F{row}"] = change.get('account_breakdown', 'No account breakdown available')

            change_val = change.get('change', 0)
            if change_val > 0:
                fill_color = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
            elif change_val < 0:
                fill_color = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
            else:
                fill_color = None

            for col in range(1, 7):  # Now 6 columns instead of 5
                cell = ws.cell(row=row, column=col)
                if fill_color:
                    cell.fill = fill_color
                cell.border = thin_border
            row += 1
        row += 1

        # Total COGS Trend (632*)
    if revenue_analysis.get('total_632_trend', {}).get('changes'):
        ws[f"A{row}"] = "TOTAL COGS TREND (632*)"
        ws[f"A{row}"].font = section_font
        row += 1

        headers = ["Period", "Previous Value", "Current Value", "Change (VND)", "Change (%)", "Account Breakdown"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row += 1

        for change in revenue_analysis['total_632_trend']['changes']:
            ws[f"A{row}"] = f"{change.get('from', '')} → {change.get('to', '')}"
            ws[f"B{row}"] = format_vnd(change.get('prev_value', 0))
            ws[f"C{row}"] = format_vnd(change.get('curr_value', 0))
            ws[f"D{row}"] = format_vnd(change.get('change', 0))
            ws[f"E{row}"] = f"{change.get('pct_change', 0):+.1f}%"
            ws[f"F{row}"] = change.get('account_breakdown', 'No account breakdown available')

            delta = change.get('change', 0)
            if delta > 0:
                fill_color = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")  # Cost up = bad
            elif delta < 0:
                fill_color = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")  # Cost down = good
            else:
                fill_color = None

            for col in range(1, 7):  # Now 6 columns instead of 5
                cell = ws.cell(row=row, column=col)
                if fill_color:
                    cell.fill = fill_color
                cell.border = thin_border
            row += 1
        row += 1

    # Total SG&A 641 Trend
    if revenue_analysis.get('total_641_trend', {}).get('changes'):
        ws[f"A{row}"] = "TOTAL SG&A TREND (641*)"
        ws[f"A{row}"].font = section_font
        row += 1

        headers = ["Period", "Previous Value", "Current Value", "Change (VND)", "Change (%)", "Account Breakdown"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row += 1

        for change in revenue_analysis['total_641_trend']['changes']:
            ws[f"A{row}"] = f"{change.get('from', '')} → {change.get('to', '')}"
            ws[f"B{row}"] = format_vnd(change.get('prev_value', 0))
            ws[f"C{row}"] = format_vnd(change.get('curr_value', 0))
            ws[f"D{row}"] = format_vnd(change.get('change', 0))
            ws[f"E{row}"] = f"{change.get('pct_change', 0):+.1f}%"
            ws[f"F{row}"] = change.get('account_breakdown', 'No account breakdown available')

            delta = change.get('change', 0)
            fill_color = PatternFill(start_color="FFEBEE" if delta > 0 else "E8F5E8", end_color="FFEBEE" if delta > 0 else "E8F5E8", fill_type="solid") if delta != 0 else None
            for col in range(1, 7):  # Now 6 columns instead of 5
                cell = ws.cell(row=row, column=col)
                if fill_color:
                    cell.fill = fill_color
                cell.border = thin_border
            row += 1
        row += 1

    # Total SG&A 642 Trend
    if revenue_analysis.get('total_642_trend', {}).get('changes'):
        ws[f"A{row}"] = "TOTAL SG&A TREND (642*)"
        ws[f"A{row}"].font = section_font
        row += 1

        headers = ["Period", "Previous Value", "Current Value", "Change (VND)", "Change (%)", "Account Breakdown"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row += 1

        for change in revenue_analysis['total_642_trend']['changes']:
            ws[f"A{row}"] = f"{change.get('from', '')} → {change.get('to', '')}"
            ws[f"B{row}"] = format_vnd(change.get('prev_value', 0))
            ws[f"C{row}"] = format_vnd(change.get('curr_value', 0))
            ws[f"D{row}"] = format_vnd(change.get('change', 0))
            ws[f"E{row}"] = f"{change.get('pct_change', 0):+.1f}%"
            ws[f"F{row}"] = change.get('account_breakdown', 'No account breakdown available')

            delta = change.get('change', 0)
            fill_color = PatternFill(start_color="FFEBEE" if delta > 0 else "E8F5E8", end_color="FFEBEE" if delta > 0 else "E8F5E8", fill_type="solid") if delta != 0 else None
            for col in range(1, 7):  # Now 6 columns instead of 5
                cell = ws.cell(row=row, column=col)
                if fill_color:
                    cell.fill = fill_color
                cell.border = thin_border
            row += 1
        row += 1


    # Gross Margin Analysis
    if revenue_analysis.get('gross_margin_analysis', {}).get('trend'):
        ws[f"A{row}"] = "GROSS MARGIN ANALYSIS"
        ws[f"A{row}"].font = section_font
        row += 1

        headers = ["Month", "Revenue", "Cost", "Gross Margin %", "Change from Previous"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row += 1

        prev_margin = None
        for margin_data in revenue_analysis['gross_margin_analysis']['trend']:
            ws[f"A{row}"] = margin_data.get('month', '')
            ws[f"B{row}"] = format_vnd(margin_data.get('revenue', 0))
            ws[f"C{row}"] = format_vnd(margin_data.get('cost', 0))
            ws[f"D{row}"] = f"{margin_data.get('gross_margin_pct', 0):.1f}%"

            current_margin = margin_data.get('gross_margin_pct', 0)
            if prev_margin is not None:
                change = current_margin - prev_margin
                ws[f"E{row}"] = f"{change:+.1f}pp"
                if change > 0:
                    ws.cell(row=row, column=5).fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                elif change < 0:
                    ws.cell(row=row, column=5).fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
            else:
                ws[f"E{row}"] = "N/A"

            prev_margin = current_margin

            for col in range(1, 6):
                ws.cell(row=row, column=col).border = thin_border
            row += 1
        row += 1

    # Utility Analysis (if available)
    if revenue_analysis.get('utility_analysis'):
        ws[f"A{row}"] = "UTILITY REVENUE VS COST ANALYSIS"
        ws[f"A{row}"].font = section_font
        row += 1

        if revenue_analysis['utility_analysis'].get('available') and revenue_analysis['utility_analysis'].get('margins'):
            headers = ["Month", "Utility Revenue", "Utility Cost", "Margin (VND)", "Margin %"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
            row += 1

            for margin in revenue_analysis['utility_analysis']['margins']:
                ws[f"A{row}"] = margin.get('month', '')
                ws[f"B{row}"] = format_vnd(margin.get('revenue', 0))
                ws[f"C{row}"] = format_vnd(margin.get('cost', 0))
                ws[f"D{row}"] = format_vnd(margin.get('revenue', 0) - margin.get('cost', 0))
                ws[f"E{row}"] = f"{margin.get('margin_pct', 0):.1f}%"

                margin_pct = margin.get('margin_pct', 0)
                if margin_pct >= 0:
                    fill_color = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                else:
                    fill_color = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

                for col in range(1, 6):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = fill_color
                    cell.border = thin_border
                row += 1
        else:
            ws[f"A{row}"] = "Utility accounts not found in the data."
            ws[f"A{row}"].font = Font(italic=True, color="666666")
            row += 1
        row += 1

    # Combined SG&A Analysis (641 + 642)
    if revenue_analysis.get('combined_sga_analysis', {}).get('ratio_trend'):
        ws[f"A{row}"] = "COMBINED SG&A ANALYSIS (641* + 642*)"
        ws[f"A{row}"].font = section_font
        row += 1

        combined_analysis = revenue_analysis['combined_sga_analysis']
        ws[f"A{row}"] = f"Total 641* Accounts: {combined_analysis.get('total_641_accounts', 0)}"
        row += 1
        ws[f"A{row}"] = f"Total 642* Accounts: {combined_analysis.get('total_642_accounts', 0)}"
        row += 2

        ws[f"A{row}"] = "SG&A Ratio Trend (% of Revenue)"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1

        headers = ["Month", "Revenue", "641* Total", "642* Total", "Total SG&A", "SG&A Ratio %", "Change"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row += 1

        prev_ratio = None
        for sga_data in combined_analysis['ratio_trend']:
            ws[f"A{row}"] = sga_data.get('month', '')
            ws[f"B{row}"] = format_vnd(sga_data.get('revenue', 0))
            ws[f"C{row}"] = format_vnd(sga_data.get('sga_641_total', 0))
            ws[f"D{row}"] = format_vnd(sga_data.get('sga_642_total', 0))
            ws[f"E{row}"] = format_vnd(sga_data.get('total_sga', 0))
            ws[f"F{row}"] = f"{sga_data.get('sga_ratio_pct', 0):.1f}%"

            current_ratio = sga_data.get('sga_ratio_pct', 0)
            if prev_ratio is not None:
                change = current_ratio - prev_ratio
                ws[f"G{row}"] = f"{change:+.1f}pp"
                if change > 2:
                    ws.cell(row=row, column=7).fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
                elif change < -2:
                    ws.cell(row=row, column=7).fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
            else:
                ws[f"G{row}"] = "N/A"

            prev_ratio = current_ratio

            for col in range(1, 8):
                ws.cell(row=row, column=col).border = thin_border
            row += 1
        row += 1

    # Auto-fit main sheet
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # === GROUPED DETAIL SHEETS ===
    wb = ws.parent  # workbook

    if revenue_analysis.get('revenue_by_account'):
        create_group_sheet(
            wb, "511 Accounts",
            revenue_analysis['revenue_by_account'],
            section_font=section_font,
            header_font=header_font,
            header_fill=header_fill,
            thin_border=thin_border,
            format_vnd=format_vnd,
            cost_mode=False,  # revenue polarity
        )

    if revenue_analysis.get('cogs_632_analysis'):
        create_group_sheet(
            wb, "632 Accounts",
            revenue_analysis['cogs_632_analysis'],
            section_font=section_font,
            header_font=header_font,
            header_fill=header_fill,
            thin_border=thin_border,
            format_vnd=format_vnd,
            cost_mode=True,
        )

    if revenue_analysis.get('sga_641_analysis'):
        create_group_sheet(
            wb, "641 Accounts",
            revenue_analysis['sga_641_analysis'],
            section_font=section_font,
            header_font=header_font,
            header_fill=header_fill,
            thin_border=thin_border,
            format_vnd=format_vnd,
            cost_mode=True,
        )

    if revenue_analysis.get('sga_642_analysis'):
        create_group_sheet(
            wb, "642 Accounts",
            revenue_analysis['sga_642_analysis'],
            section_font=section_font,
            header_font=header_font,
            header_fill=header_fill,
            thin_border=thin_border,
            format_vnd=format_vnd,
            cost_mode=True,
        )


def _add_consolidated_months_analysis_to_sheet(ws, revenue_analysis: dict, variance_analysis: dict):
    """
    Add a consolidated 'Months Analysis' sheet combining:
    - Revenue Variance Analysis
    - Legacy Revenue Analysis
    - 511 Accounts
    - 632 Accounts
    - 641 Accounts
    - 642 Accounts

    All in one comprehensive sheet split by clear sections.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    section_font = Font(bold=True, size=14, color="2F5597")
    subsection_font = Font(bold=True, size=12, color="1F4E79")
    insight_font = Font(bold=True, size=11, color="D83B01")

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def format_vnd(amount):
        if isinstance(amount, (int, float)) and not pd.isna(amount):
            return f"{amount:,.0f} VND"
        return "N/A"

    def format_pct(pct):
        if isinstance(pct, (int, float)) and not pd.isna(pct):
            return f"{pct:+.1f}%"
        return "N/A"

    row = 1

    # ========================================
    # MAIN TITLE
    # ========================================
    ws[f"A{row}"] = "COMPREHENSIVE MONTHS ANALYSIS"
    ws[f"A{row}"].font = Font(bold=True, size=18, color="2F5597")
    ws.merge_cells(f"A{row}:F{row}")
    row += 2

    # ========================================
    # SECTION 1: REVENUE VARIANCE ANALYSIS
    # ========================================
    ws[f"A{row}"] = "═" * 100
    row += 1
    ws[f"A{row}"] = "SECTION 1: REVENUE VARIANCE ANALYSIS"
    ws[f"A{row}"].font = section_font
    ws.merge_cells(f"A{row}:F{row}")
    row += 1
    ws[f"A{row}"] = "═" * 100
    row += 2

    # Executive Summary from Variance Analysis
    if variance_analysis:
        ws[f"A{row}"] = "Executive Summary"
        ws[f"A{row}"].font = subsection_font
        row += 1

        summary_data = [
            ["Subsidiary", variance_analysis.get('subsidiary', 'N/A')],
            ["File", variance_analysis.get('filename', 'N/A')],
            ["Months Analyzed", len(variance_analysis.get('months_analyzed', []))],
            ["Revenue Streams Identified", variance_analysis.get('analysis_summary', {}).get('total_revenue_streams', 0)],
            ["Variance Periods", variance_analysis.get('analysis_summary', {}).get('total_variance_periods', 0)],
            ["Accounts with Vendor Impact", variance_analysis.get('analysis_summary', {}).get('accounts_with_vendor_impact', 0)]
        ]

        for label, value in summary_data:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            ws[f"A{row}"].font = Font(bold=True)
            row += 1
        row += 1

        # Key Insights
        if variance_analysis.get('key_insights'):
            ws[f"A{row}"] = "Key Insights"
            ws[f"A{row}"].font = subsection_font
            row += 1

            for insight in variance_analysis['key_insights']:
                ws[f"A{row}"] = f"• {insight}"
                ws[f"A{row}"].font = insight_font
                ws.merge_cells(f"A{row}:F{row}")
                row += 1
            row += 1

        # Total Revenue Month-over-Month Changes
        if variance_analysis.get('total_revenue_analysis', {}).get('month_over_month_changes'):
            ws[f"A{row}"] = "1.1 Total Revenue Month-over-Month Analysis"
            ws[f"A{row}"].font = subsection_font
            row += 1

            headers = ["Period From", "Period To", "Previous Revenue", "Current Revenue", "Absolute Change", "% Change", "Direction"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
            row += 1

            changes = variance_analysis['total_revenue_analysis']['month_over_month_changes']
            for change in changes:
                ws[f"A{row}"] = change.get('period_from', '')
                ws[f"B{row}"] = change.get('period_to', '')
                ws[f"C{row}"] = format_vnd(change.get('previous_revenue', 0))
                ws[f"D{row}"] = format_vnd(change.get('current_revenue', 0))
                ws[f"E{row}"] = format_vnd(change.get('absolute_change', 0))
                ws[f"F{row}"] = format_pct(change.get('percentage_change', 0))
                ws[f"G{row}"] = change.get('change_direction', '')

                direction = change.get('change_direction', '').lower()
                if 'increase' in direction:
                    fill_color = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                elif 'decrease' in direction:
                    fill_color = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
                else:
                    fill_color = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

                for col in range(1, 8):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = fill_color
                    cell.border = thin_border
                row += 1
            row += 2

        # Revenue Stream Analysis
        if variance_analysis.get('revenue_stream_analysis', {}).get('streams'):
            ws[f"A{row}"] = "1.2 Revenue Stream Analysis (Contribution Code '01' Only)"
            ws[f"A{row}"].font = subsection_font
            row += 1

            streams = variance_analysis['revenue_stream_analysis']['streams']

            for stream_name, stream_data in streams.items():
                ws[f"A{row}"] = f"Stream: {stream_name}"
                ws[f"A{row}"].font = Font(bold=True, color="1F4E79")
                ws.merge_cells(f"A{row}:F{row}")
                row += 1

                ws[f"A{row}"] = f"Total Entities: {stream_data.get('total_entities', 0)}"
                row += 1

                if stream_data.get('month_changes'):
                    headers = ["Period From", "Period To", "Previous Value", "Current Value", "Change", "% Change"]
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=row, column=col, value=header)
                        cell.font = header_font
                        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                        cell.border = thin_border
                    row += 1

                    for change in stream_data['month_changes']:
                        ws[f"A{row}"] = change.get('period_from', '')
                        ws[f"B{row}"] = change.get('period_to', '')
                        ws[f"C{row}"] = format_vnd(change.get('previous_value', 0))
                        ws[f"D{row}"] = format_vnd(change.get('current_value', 0))
                        ws[f"E{row}"] = format_vnd(change.get('absolute_change', 0))
                        ws[f"F{row}"] = format_pct(change.get('percentage_change', 0))

                        for col in range(1, 7):
                            cell = ws.cell(row=row, column=col)
                            cell.border = thin_border
                        row += 1
                row += 1
            row += 1

        # Vendor/Customer Impact Analysis
        if variance_analysis.get('vendor_customer_impact', {}).get('detailed_analysis'):
            ws[f"A{row}"] = "1.3 Vendor/Customer Impact Analysis (Net Effect Breakdown)"
            ws[f"A{row}"].font = subsection_font
            row += 1

            detailed_analysis = variance_analysis['vendor_customer_impact']['detailed_analysis']

            for account_name, account_analysis in detailed_analysis.items():
                ws[f"A{row}"] = f"Account: {account_name}"
                ws[f"A{row}"].font = Font(bold=True, color="1F4E79")
                ws.merge_cells(f"A{row}:F{row}")
                row += 1

                if account_analysis.get('period_impacts'):
                    for period_impact in account_analysis['period_impacts']:
                        ws[f"A{row}"] = f"Period: {period_impact.get('period_from', '')} → {period_impact.get('period_to', '')}"
                        ws[f"A{row}"].font = Font(bold=True)
                        row += 1

                        net_explanation = period_impact.get('net_effect_explanation', '')
                        ws[f"A{row}"] = f"Net Effect: {net_explanation}"
                        ws[f"A{row}"].font = Font(bold=True, color="D83B01")
                        ws.merge_cells(f"A{row}:F{row}")
                        row += 1

                        total_positive = period_impact.get('total_positive_change', 0)
                        total_negative = period_impact.get('total_negative_change', 0)
                        ws[f"A{row}"] = f"Total Increases: {format_vnd(total_positive)} | Total Decreases: {format_vnd(total_negative)} | Net: {format_vnd(total_positive + total_negative)}"
                        ws[f"A{row}"].font = Font(italic=True)
                        ws.merge_cells(f"A{row}:F{row}")
                        row += 1
                        row += 1

                        # Positive Contributors
                        if period_impact.get('positive_contributors'):
                            ws[f"A{row}"] = "POSITIVE CONTRIBUTORS (Increases)"
                            ws[f"A{row}"].font = Font(bold=True, color="0F7B0F")
                            row += 1

                            headers = ["Entity", "Previous Value", "Current Value", "Increase (+)", "% Change", "% of Total Change"]
                            for col, header in enumerate(headers, 1):
                                cell = ws.cell(row=row, column=col, value=header)
                                cell.font = header_font
                                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                                cell.border = thin_border
                            row += 1

                            for entity in period_impact['positive_contributors']:
                                ws[f"A{row}"] = entity.get('entity', '')
                                ws[f"B{row}"] = format_vnd(entity.get('previous_value', 0))
                                ws[f"C{row}"] = format_vnd(entity.get('current_value', 0))
                                ws[f"D{row}"] = f"+{entity.get('absolute_change', 0):,.0f} VND"
                                ws[f"E{row}"] = format_pct(entity.get('percentage_change', 0))
                                ws[f"F{row}"] = format_pct(entity.get('contribution_to_period_change', 0))

                                fill_color = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                                for col in range(1, 7):
                                    cell = ws.cell(row=row, column=col)
                                    cell.fill = fill_color
                                    cell.border = thin_border
                                row += 1
                            row += 1

                        # Negative Contributors
                        if period_impact.get('negative_contributors'):
                            ws[f"A{row}"] = "NEGATIVE CONTRIBUTORS (Decreases)"
                            ws[f"A{row}"].font = Font(bold=True, color="C5504B")
                            row += 1

                            headers = ["Entity", "Previous Value", "Current Value", "Decrease (-)", "% Change", "% of Total Change"]
                            for col, header in enumerate(headers, 1):
                                cell = ws.cell(row=row, column=col, value=header)
                                cell.font = header_font
                                cell.fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
                                cell.border = thin_border
                            row += 1

                            for entity in period_impact['negative_contributors']:
                                ws[f"A{row}"] = entity.get('entity', '')
                                ws[f"B{row}"] = format_vnd(entity.get('previous_value', 0))
                                ws[f"C{row}"] = format_vnd(entity.get('current_value', 0))
                                ws[f"D{row}"] = f"{entity.get('absolute_change', 0):,.0f} VND"
                                ws[f"E{row}"] = format_pct(entity.get('percentage_change', 0))
                                ws[f"F{row}"] = format_pct(entity.get('contribution_to_period_change', 0))

                                fill_color = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
                                for col in range(1, 7):
                                    cell = ws.cell(row=row, column=col)
                                    cell.fill = fill_color
                                    cell.border = thin_border
                                row += 1
                            row += 1
                row += 1

    # ========================================
    # SECTION 2: 511 REVENUE ACCOUNTS
    # ========================================
    row += 2
    ws[f"A{row}"] = "═" * 100
    row += 1
    ws[f"A{row}"] = "SECTION 2: 511 REVENUE ACCOUNTS DETAILED ANALYSIS"
    ws[f"A{row}"].font = section_font
    ws.merge_cells(f"A{row}:F{row}")
    row += 1
    ws[f"A{row}"] = "═" * 100
    row += 2

    if revenue_analysis.get('revenue_by_account'):
        row = _add_account_group_section(ws, row, revenue_analysis['revenue_by_account'],
                                        "Revenue Accounts (511*)", False,
                                        header_font, header_fill, subsection_font, thin_border, format_vnd)

    # ========================================
    # SECTION 3: 632 COGS ACCOUNTS
    # ========================================
    row += 2
    ws[f"A{row}"] = "═" * 100
    row += 1
    ws[f"A{row}"] = "SECTION 3: 632 COGS ACCOUNTS DETAILED ANALYSIS"
    ws[f"A{row}"].font = section_font
    ws.merge_cells(f"A{row}:F{row}")
    row += 1
    ws[f"A{row}"] = "═" * 100
    row += 2

    if revenue_analysis.get('cogs_632_analysis'):
        row = _add_account_group_section(ws, row, revenue_analysis['cogs_632_analysis'],
                                        "COGS Accounts (632*)", True,
                                        header_font, header_fill, subsection_font, thin_border, format_vnd)

    # ========================================
    # SECTION 4: 641 SG&A ACCOUNTS
    # ========================================
    row += 2
    ws[f"A{row}"] = "═" * 100
    row += 1
    ws[f"A{row}"] = "SECTION 4: 641 SG&A ACCOUNTS DETAILED ANALYSIS"
    ws[f"A{row}"].font = section_font
    ws.merge_cells(f"A{row}:F{row}")
    row += 1
    ws[f"A{row}"] = "═" * 100
    row += 2

    if revenue_analysis.get('sga_641_analysis'):
        row = _add_account_group_section(ws, row, revenue_analysis['sga_641_analysis'],
                                        "SG&A Accounts (641*)", True,
                                        header_font, header_fill, subsection_font, thin_border, format_vnd)

    # ========================================
    # SECTION 5: 642 SG&A ACCOUNTS
    # ========================================
    row += 2
    ws[f"A{row}"] = "═" * 100
    row += 1
    ws[f"A{row}"] = "SECTION 5: 642 SG&A ACCOUNTS DETAILED ANALYSIS"
    ws[f"A{row}"].font = section_font
    ws.merge_cells(f"A{row}:F{row}")
    row += 1
    ws[f"A{row}"] = "═" * 100
    row += 2

    if revenue_analysis.get('sga_642_analysis'):
        row = _add_account_group_section(ws, row, revenue_analysis['sga_642_analysis'],
                                        "SG&A Accounts (642*)", True,
                                        header_font, header_fill, subsection_font, thin_border, format_vnd)

    # ========================================
    # SECTION 6: SUMMARY METRICS & TRENDS
    # ========================================
    row += 2
    ws[f"A{row}"] = "═" * 100
    row += 1
    ws[f"A{row}"] = "SECTION 6: SUMMARY METRICS & TRENDS"
    ws[f"A{row}"].font = section_font
    ws.merge_cells(f"A{row}:F{row}")
    row += 1
    ws[f"A{row}"] = "═" * 100
    row += 2

    # Gross Margin Analysis
    if revenue_analysis.get('gross_margin_analysis', {}).get('trend'):
        ws[f"A{row}"] = "6.1 Gross Margin Analysis"
        ws[f"A{row}"].font = subsection_font
        row += 1

        headers = ["Month", "Revenue", "Cost", "Gross Margin %", "Change from Previous"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row += 1

        prev_margin = None
        for margin_data in revenue_analysis['gross_margin_analysis']['trend']:
            ws[f"A{row}"] = margin_data.get('month', '')
            ws[f"B{row}"] = format_vnd(margin_data.get('revenue', 0))
            ws[f"C{row}"] = format_vnd(margin_data.get('cost', 0))
            ws[f"D{row}"] = f"{margin_data.get('gross_margin_pct', 0):.1f}%"

            current_margin = margin_data.get('gross_margin_pct', 0)
            if prev_margin is not None:
                change = current_margin - prev_margin
                ws[f"E{row}"] = f"{change:+.1f}pp"
                if change > 0:
                    ws.cell(row=row, column=5).fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                elif change < 0:
                    ws.cell(row=row, column=5).fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
            else:
                ws[f"E{row}"] = "N/A"

            prev_margin = current_margin

            for col in range(1, 6):
                ws.cell(row=row, column=col).border = thin_border
            row += 1
        row += 2

    # Combined SG&A Analysis
    if revenue_analysis.get('combined_sga_analysis', {}).get('ratio_trend'):
        ws[f"A{row}"] = "6.2 Combined SG&A Analysis (641* + 642*)"
        ws[f"A{row}"].font = subsection_font
        row += 1

        combined_analysis = revenue_analysis['combined_sga_analysis']
        ws[f"A{row}"] = f"Total 641* Accounts: {combined_analysis.get('total_641_accounts', 0)}"
        row += 1
        ws[f"A{row}"] = f"Total 642* Accounts: {combined_analysis.get('total_642_accounts', 0)}"
        row += 2

        headers = ["Month", "Revenue", "641* Total", "642* Total", "Total SG&A", "SG&A Ratio %", "Change"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row += 1

        prev_ratio = None
        for sga_data in combined_analysis['ratio_trend']:
            ws[f"A{row}"] = sga_data.get('month', '')
            ws[f"B{row}"] = format_vnd(sga_data.get('revenue', 0))
            ws[f"C{row}"] = format_vnd(sga_data.get('sga_641_total', 0))
            ws[f"D{row}"] = format_vnd(sga_data.get('sga_642_total', 0))
            ws[f"E{row}"] = format_vnd(sga_data.get('total_sga', 0))
            ws[f"F{row}"] = f"{sga_data.get('sga_ratio_pct', 0):.1f}%"

            current_ratio = sga_data.get('sga_ratio_pct', 0)
            if prev_ratio is not None:
                change = current_ratio - prev_ratio
                ws[f"G{row}"] = f"{change:+.1f}pp"
                if change > 2:
                    ws.cell(row=row, column=7).fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
                elif change < -2:
                    ws.cell(row=row, column=7).fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
            else:
                ws[f"G{row}"] = "N/A"

            prev_ratio = current_ratio

            for col in range(1, 8):
                ws.cell(row=row, column=col).border = thin_border
            row += 1
        row += 1

    # Auto-fit columns (skip merged cells)
    from openpyxl.cell.cell import MergedCell
    for column in ws.columns:
        max_length = 0
        column_letter = None
        for cell in column:
            # Skip merged cells
            if isinstance(cell, MergedCell):
                continue
            if column_letter is None:
                column_letter = cell.column_letter
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        if column_letter:
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column_letter].width = adjusted_width


def _add_account_group_section(ws, start_row, accounts_data, title, cost_mode,
                               header_font, header_fill, subsection_font, thin_border, format_vnd):
    """Helper function to add an account group section to the consolidated sheet."""
    row = start_row

    for account_name, account_data in accounts_data.items():
        ws[f"A{row}"] = f"Account: {account_name}"
        ws[f"A{row}"].font = subsection_font
        ws.merge_cells(f"A{row}:F{row}")
        row += 1

        # Month-over-Month Changes
        if account_data.get('changes'):
            headers = ["Period", "Previous Value", "Current Value", "Change (VND)", "Change (%)"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
            row += 1

            for change in account_data['changes']:
                ws[f"A{row}"] = f"{change.get('from', '')} → {change.get('to', '')}"
                ws[f"B{row}"] = format_vnd(change.get('prev_val', 0))
                ws[f"C{row}"] = format_vnd(change.get('curr_val', 0))
                ws[f"D{row}"] = format_vnd(change.get('change', 0))
                ws[f"E{row}"] = f"{change.get('pct_change', 0):+.1f}%"

                delta = change.get('change', 0)
                if delta != 0:
                    good = (delta > 0) if not cost_mode else (delta < 0)
                    fill = PatternFill(
                        start_color=("E8F5E8" if good else "FFEBEE"),
                        end_color=("E8F5E8" if good else "FFEBEE"),
                        fill_type="solid"
                    )
                else:
                    fill = None

                for col in range(1, 6):
                    cell = ws.cell(row=row, column=col)
                    if fill:
                        cell.fill = fill
                    cell.border = thin_border
                row += 1
            row += 1

        # Top Entity Impacts
        impacts = account_data.get('customer_impacts') or account_data.get('entity_impacts') or []
        if impacts:
            ws[f"A{row}"] = "Top Entity Impacts:"
            ws[f"A{row}"].font = Font(bold=True, color="1F4E79")
            row += 1

            headers = ["Entity", "Previous Value", "Current Value", "Change", "% Change"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                cell.border = thin_border
            row += 1

            for impact in impacts:
                ws[f"A{row}"] = impact.get('entity', '')
                ws[f"B{row}"] = format_vnd(impact.get('prev_val', 0))
                ws[f"C{row}"] = format_vnd(impact.get('curr_val', 0))
                ws[f"D{row}"] = format_vnd(impact.get('change', 0))
                ws[f"E{row}"] = f"{impact.get('pct_change', 0):+.1f}%"

                for col in range(1, 6):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                row += 1
            row += 1

        row += 1  # Space between accounts

    return row


def _filter_to_latest_two_months(data_dict: dict, target_months: list) -> dict:
    """
    Filter analysis data to only include the last two months.

    Args:
        data_dict: Dictionary containing analysis data
        target_months: List of two month strings to filter for (e.g., ['Jun-2024', 'Jul-2024'])

    Returns:
        Filtered dictionary with only the target months
    """
    if not target_months or len(target_months) < 2:
        return data_dict

    prev_month, curr_month = target_months[0], target_months[1]
    filtered = {}

    for key, value in data_dict.items():
        if isinstance(value, dict):
            # Recursively filter nested dictionaries
            if 'changes' in value:
                # Filter changes to only include the target period
                filtered_changes = [
                    change for change in value.get('changes', [])
                    if change.get('from') == prev_month and change.get('to') == curr_month
                ]
                filtered[key] = {**value, 'changes': filtered_changes}

                # Update biggest_change if it exists
                if filtered_changes and 'biggest_change' in value:
                    filtered[key]['biggest_change'] = max(
                        filtered_changes,
                        key=lambda x: abs(x.get('change', 0))
                    ) if filtered_changes else None
            else:
                filtered[key] = value
        else:
            filtered[key] = value

    return filtered


def _add_month_to_month_analysis_to_sheet(ws, revenue_analysis: dict, variance_analysis: dict, xl_bytes: bytes, filename: str):
    """
    Add a 'Month to Month Analysis' sheet focusing only on the last two consecutive months.

    This is similar to the consolidated months analysis but filters data to show only:
    - Current month vs Previous month comparison
    - No historical trends, just the latest change

    Args:
        ws: Worksheet to write to
        revenue_analysis: Full revenue analysis dictionary
        variance_analysis: Full variance analysis dictionary
        xl_bytes: Excel file bytes to read row 4 from
        filename: Filename for subsidiary extraction
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from app.shared.utils.logging_config import get_logger

    logger = get_logger(__name__)

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    section_font = Font(bold=True, size=14, color="2F5597")
    subsection_font = Font(bold=True, size=12, color="1F4E79")
    insight_font = Font(bold=True, size=11, color="D83B01")

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def format_vnd(amount):
        if isinstance(amount, (int, float)) and not pd.isna(amount):
            return f"{amount:,.0f} VND"
        return "N/A"

    def format_pct(pct):
        if isinstance(pct, (int, float)) and not pd.isna(pct):
            return f"{pct:+.1f}%"
        return "N/A"

    # Read row 4 to find "End of [Month]" text and determine which month to analyze
    logger.info("🔍 Reading row 4 to find 'End of [Month]' text...")

    try:
        # Load the Excel file using pandas
        import io
        from app.shared.utils.sheet_detection import detect_sheets_from_bytes

        # Detect PL sheet using fuzzy matching
        _, pl_sheet = detect_sheets_from_bytes(xl_bytes)

        if not pl_sheet:
            logger.warning("⚠️  Could not detect PL sheet for month-to-month analysis")
            # Fall back to default sheet name
            pl_sheet = "PL Breakdown"

        pl_df = pd.read_excel(io.BytesIO(xl_bytes), sheet_name=pl_sheet, header=None)

        # Read row 4 (index 3 in pandas, 0-indexed)
        row_4_values = pl_df.iloc[3].fillna('').astype(str).tolist()
        logger.info(f"🔍 Row 4 values (first 10): {row_4_values[:10]}")

        # Find month pattern in row 4
        # Can be either "End of [Month]" or "From ... to [Month]"
        MONTH_TOKENS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        MONTH_ORDER = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

        target_month_name = None
        for cell_value in row_4_values:
            cell_str = str(cell_value).strip()
            cell_lower = cell_str.lower()

            # Check for "End of [Month]" pattern
            if "end of" in cell_lower:
                for month_tok in MONTH_TOKENS:
                    if month_tok in cell_str:
                        target_month_name = month_tok
                        logger.info(f"✅ Found 'End of' text in row 4: '{cell_str}'")
                        logger.info(f"✅ Extracted target month: {target_month_name}")
                        break
            # Check for "to [Month]" pattern (e.g., "From Jan 2025 to May 2025")
            elif "to" in cell_lower:
                # Find the month after "to"
                parts = cell_str.split("to")
                if len(parts) >= 2:
                    after_to = parts[-1].strip()
                    for month_tok in MONTH_TOKENS:
                        if month_tok in after_to:
                            target_month_name = month_tok
                            logger.info(f"✅ Found 'to [Month]' text in row 4: '{cell_str}'")
                            logger.info(f"✅ Extracted target month: {target_month_name}")
                            break

            if target_month_name:
                break

        if not target_month_name:
            logger.error("⚠️  Could not find month pattern in row 4")
            logger.error(f"   Searched for 'End of [Month]' or 'to [Month]' patterns")
            ws[f"A1"] = "ERROR: Could not find month information in row 4"
            ws[f"A1"].font = Font(bold=True, color="FF0000", size=14)
            return

        # Find the previous month
        target_month_idx = MONTH_ORDER.index(target_month_name)
        if target_month_idx == 0:
            logger.error("⚠️  Target month is January - no previous month in same year")
            ws[f"A1"] = "ERROR: Cannot analyze month-to-month for January (no previous month)"
            ws[f"A1"].font = Font(bold=True, color="FF0000", size=14)
            return

        prev_month_name = MONTH_ORDER[target_month_idx - 1]
        logger.info(f"✅ Previous month: {prev_month_name}")

        # Now match these month names to the actual month strings in revenue_analysis
        # The months_analyzed will have format like "Jan 2025", "Feb 2025", etc.
        all_months = revenue_analysis.get('months_analyzed', [])
        logger.info(f"🔍 Available months in analysis: {all_months}")

        curr_month = None
        prev_month = None

        for month_str in all_months:
            if target_month_name in month_str:
                curr_month = month_str
            if prev_month_name in month_str:
                prev_month = month_str

        if not curr_month or not prev_month:
            logger.error(f"⚠️  Could not match extracted months to analysis data")
            logger.error(f"   Looking for: {prev_month_name} and {target_month_name}")
            logger.error(f"   Available: {all_months}")
            ws[f"A1"] = f"ERROR: Could not find {prev_month_name} and {target_month_name} in analysis data"
            ws[f"A1"].font = Font(bold=True, color="FF0000", size=14)
            return

        logger.info(f"✅ Matched months for analysis: {prev_month} → {curr_month}")

    except Exception as e:
        logger.error(f"⚠️  Error reading row 4 from Excel: {e}", exc_info=True)
        ws[f"A1"] = f"ERROR: Failed to read row 4 from Excel file: {str(e)}"
        ws[f"A1"].font = Font(bold=True, color="FF0000", size=14)
        return

    target_months = [prev_month, curr_month]
    logger.info(f"✅ Selected months for Month to Month Analysis: {prev_month} → {curr_month}")

    row = 1

    # ========================================
    # MAIN TITLE
    # ========================================
    ws[f"A{row}"] = "MONTH TO MONTH ANALYSIS"
    ws[f"A{row}"].font = Font(bold=True, size=18, color="2F5597")
    ws.merge_cells(f"A{row}:F{row}")
    row += 1

    ws[f"A{row}"] = f"Analysis Period: {prev_month} → {curr_month}"
    ws[f"A{row}"].font = Font(bold=True, size=12, color="1F4E79")
    ws.merge_cells(f"A{row}:F{row}")
    row += 2

    # ========================================
    # SECTION 1: REVENUE VARIANCE ANALYSIS
    # ========================================
    ws[f"A{row}"] = "═" * 100
    row += 1
    ws[f"A{row}"] = "SECTION 1: REVENUE VARIANCE ANALYSIS"
    ws[f"A{row}"].font = section_font
    ws.merge_cells(f"A{row}:F{row}")
    row += 1
    ws[f"A{row}"] = "═" * 100
    row += 2

    # Executive Summary
    if variance_analysis:
        ws[f"A{row}"] = "Executive Summary"
        ws[f"A{row}"].font = subsection_font
        row += 1

        summary_data = [
            ["Subsidiary", variance_analysis.get('subsidiary', 'N/A')],
            ["File", variance_analysis.get('filename', 'N/A')],
            ["Analysis Period", f"{prev_month} → {curr_month}"],
            ["Revenue Streams Analyzed", variance_analysis.get('analysis_summary', {}).get('total_revenue_streams', 0)],
        ]

        for label, value in summary_data:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            ws[f"A{row}"].font = Font(bold=True)
            row += 1
        row += 1

        # Total Revenue Change (only for target period)
        if variance_analysis.get('total_revenue_analysis', {}).get('month_over_month_changes'):
            ws[f"A{row}"] = "1.1 Total Revenue Change"
            ws[f"A{row}"].font = subsection_font
            row += 1

            headers = ["Period From", "Period To", "Previous Revenue", "Current Revenue", "Absolute Change", "% Change", "Direction"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
            row += 1

            # Filter to only the target period
            all_changes = variance_analysis['total_revenue_analysis']['month_over_month_changes']
            target_change = [c for c in all_changes if c.get('period_from') == prev_month and c.get('period_to') == curr_month]

            for change in target_change:
                ws[f"A{row}"] = change.get('period_from', '')
                ws[f"B{row}"] = change.get('period_to', '')
                ws[f"C{row}"] = format_vnd(change.get('previous_revenue', 0))
                ws[f"D{row}"] = format_vnd(change.get('current_revenue', 0))
                ws[f"E{row}"] = format_vnd(change.get('absolute_change', 0))
                ws[f"F{row}"] = format_pct(change.get('percentage_change', 0))
                ws[f"G{row}"] = change.get('change_direction', '')

                direction = change.get('change_direction', '').lower()
                if 'increase' in direction:
                    fill_color = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                elif 'decrease' in direction:
                    fill_color = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
                else:
                    fill_color = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

                for col in range(1, 8):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = fill_color
                    cell.border = thin_border
                row += 1
            row += 2

        # Revenue Stream Analysis (filtered)
        if variance_analysis.get('revenue_stream_analysis', {}).get('streams'):
            ws[f"A{row}"] = "1.2 Revenue Stream Analysis (Contribution Code '01' Only)"
            ws[f"A{row}"].font = subsection_font
            row += 1

            streams = variance_analysis['revenue_stream_analysis']['streams']

            for stream_name, stream_data in streams.items():
                # Filter month changes to target period
                filtered_changes = [
                    c for c in stream_data.get('month_changes', [])
                    if c.get('period_from') == prev_month and c.get('period_to') == curr_month
                ]

                if not filtered_changes:
                    continue  # Skip streams with no data for this period

                ws[f"A{row}"] = f"Stream: {stream_name}"
                ws[f"A{row}"].font = Font(bold=True, color="1F4E79")
                ws.merge_cells(f"A{row}:F{row}")
                row += 1

                ws[f"A{row}"] = f"Total Entities: {stream_data.get('total_entities', 0)}"
                row += 1

                headers = ["Period From", "Period To", "Previous Value", "Current Value", "Change", "% Change"]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                    cell.border = thin_border
                row += 1

                for change in filtered_changes:
                    ws[f"A{row}"] = change.get('period_from', '')
                    ws[f"B{row}"] = change.get('period_to', '')
                    ws[f"C{row}"] = format_vnd(change.get('previous_value', 0))
                    ws[f"D{row}"] = format_vnd(change.get('current_value', 0))
                    ws[f"E{row}"] = format_vnd(change.get('absolute_change', 0))
                    ws[f"F{row}"] = format_pct(change.get('percentage_change', 0))

                    for col in range(1, 7):
                        cell = ws.cell(row=row, column=col)
                        cell.border = thin_border
                    row += 1
                row += 1
            row += 1

        # Vendor/Customer Impact Analysis (filtered)
        if variance_analysis.get('vendor_customer_impact', {}).get('detailed_analysis'):
            ws[f"A{row}"] = "1.3 Vendor/Customer Impact Analysis"
            ws[f"A{row}"].font = subsection_font
            row += 1

            detailed_analysis = variance_analysis['vendor_customer_impact']['detailed_analysis']

            for account_name, account_analysis in detailed_analysis.items():
                # Filter to target period
                filtered_impacts = [
                    p for p in account_analysis.get('period_impacts', [])
                    if p.get('period_from') == prev_month and p.get('period_to') == curr_month
                ]

                if not filtered_impacts:
                    continue

                ws[f"A{row}"] = f"Account: {account_name}"
                ws[f"A{row}"].font = Font(bold=True, color="1F4E79")
                ws.merge_cells(f"A{row}:F{row}")
                row += 1

                for period_impact in filtered_impacts:
                    ws[f"A{row}"] = f"Period: {period_impact.get('period_from', '')} → {period_impact.get('period_to', '')}"
                    ws[f"A{row}"].font = Font(bold=True)
                    row += 1

                    net_explanation = period_impact.get('net_effect_explanation', '')
                    ws[f"A{row}"] = f"Net Effect: {net_explanation}"
                    ws[f"A{row}"].font = Font(bold=True, color="D83B01")
                    ws.merge_cells(f"A{row}:F{row}")
                    row += 1

                    total_positive = period_impact.get('total_positive_change', 0)
                    total_negative = period_impact.get('total_negative_change', 0)
                    ws[f"A{row}"] = f"Total Increases: {format_vnd(total_positive)} | Total Decreases: {format_vnd(total_negative)} | Net: {format_vnd(total_positive + total_negative)}"
                    ws[f"A{row}"].font = Font(italic=True)
                    ws.merge_cells(f"A{row}:F{row}")
                    row += 1
                    row += 1

                    # Positive Contributors
                    if period_impact.get('positive_contributors'):
                        ws[f"A{row}"] = "POSITIVE CONTRIBUTORS (Increases)"
                        ws[f"A{row}"].font = Font(bold=True, color="0F7B0F")
                        row += 1

                        headers = ["Entity", "Previous Value", "Current Value", "Increase (+)", "% Change", "% of Total Change"]
                        for col, header in enumerate(headers, 1):
                            cell = ws.cell(row=row, column=col, value=header)
                            cell.font = header_font
                            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                            cell.border = thin_border
                        row += 1

                        for entity in period_impact['positive_contributors']:
                            ws[f"A{row}"] = entity.get('entity', '')
                            ws[f"B{row}"] = format_vnd(entity.get('previous_value', 0))
                            ws[f"C{row}"] = format_vnd(entity.get('current_value', 0))
                            ws[f"D{row}"] = f"+{entity.get('absolute_change', 0):,.0f} VND"
                            ws[f"E{row}"] = format_pct(entity.get('percentage_change', 0))
                            ws[f"F{row}"] = format_pct(entity.get('contribution_to_period_change', 0))

                            fill_color = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                            for col in range(1, 7):
                                cell = ws.cell(row=row, column=col)
                                cell.fill = fill_color
                                cell.border = thin_border
                            row += 1
                        row += 1

                    # Negative Contributors
                    if period_impact.get('negative_contributors'):
                        ws[f"A{row}"] = "NEGATIVE CONTRIBUTORS (Decreases)"
                        ws[f"A{row}"].font = Font(bold=True, color="C5504B")
                        row += 1

                        headers = ["Entity", "Previous Value", "Current Value", "Decrease (-)", "% Change", "% of Total Change"]
                        for col, header in enumerate(headers, 1):
                            cell = ws.cell(row=row, column=col, value=header)
                            cell.font = header_font
                            cell.fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
                            cell.border = thin_border
                        row += 1

                        for entity in period_impact['negative_contributors']:
                            ws[f"A{row}"] = entity.get('entity', '')
                            ws[f"B{row}"] = format_vnd(entity.get('previous_value', 0))
                            ws[f"C{row}"] = format_vnd(entity.get('current_value', 0))
                            ws[f"D{row}"] = f"{entity.get('absolute_change', 0):,.0f} VND"
                            ws[f"E{row}"] = format_pct(entity.get('percentage_change', 0))
                            ws[f"F{row}"] = format_pct(entity.get('contribution_to_period_change', 0))

                            fill_color = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
                            for col in range(1, 7):
                                cell = ws.cell(row=row, column=col)
                                cell.fill = fill_color
                                cell.border = thin_border
                            row += 1
                        row += 1
                row += 1

    # ========================================
    # SECTION 2: 511 REVENUE ACCOUNTS
    # ========================================
    row += 2
    ws[f"A{row}"] = "═" * 100
    row += 1
    ws[f"A{row}"] = "SECTION 2: 511 REVENUE ACCOUNTS"
    ws[f"A{row}"].font = section_font
    ws.merge_cells(f"A{row}:F{row}")
    row += 1
    ws[f"A{row}"] = "═" * 100
    row += 2

    if revenue_analysis.get('revenue_by_account'):
        filtered_511 = _filter_to_latest_two_months(revenue_analysis['revenue_by_account'], target_months)
        row = _add_account_group_section(ws, row, filtered_511, "Revenue Accounts (511*)", False,
                                        header_font, header_fill, subsection_font, thin_border, format_vnd)

    # ========================================
    # SECTION 3: 632 COGS ACCOUNTS
    # ========================================
    row += 2
    ws[f"A{row}"] = "═" * 100
    row += 1
    ws[f"A{row}"] = "SECTION 3: 632 COGS ACCOUNTS"
    ws[f"A{row}"].font = section_font
    ws.merge_cells(f"A{row}:F{row}")
    row += 1
    ws[f"A{row}"] = "═" * 100
    row += 2

    if revenue_analysis.get('cogs_632_analysis'):
        filtered_632 = _filter_to_latest_two_months(revenue_analysis['cogs_632_analysis'], target_months)
        row = _add_account_group_section(ws, row, filtered_632, "COGS Accounts (632*)", True,
                                        header_font, header_fill, subsection_font, thin_border, format_vnd)

    # ========================================
    # SECTION 4: 641 SG&A ACCOUNTS
    # ========================================
    row += 2
    ws[f"A{row}"] = "═" * 100
    row += 1
    ws[f"A{row}"] = "SECTION 4: 641 SG&A ACCOUNTS"
    ws[f"A{row}"].font = section_font
    ws.merge_cells(f"A{row}:F{row}")
    row += 1
    ws[f"A{row}"] = "═" * 100
    row += 2

    if revenue_analysis.get('sga_641_analysis'):
        filtered_641 = _filter_to_latest_two_months(revenue_analysis['sga_641_analysis'], target_months)
        row = _add_account_group_section(ws, row, filtered_641, "SG&A Accounts (641*)", True,
                                        header_font, header_fill, subsection_font, thin_border, format_vnd)

    # ========================================
    # SECTION 5: 642 SG&A ACCOUNTS
    # ========================================
    row += 2
    ws[f"A{row}"] = "═" * 100
    row += 1
    ws[f"A{row}"] = "SECTION 5: 642 SG&A ACCOUNTS"
    ws[f"A{row}"].font = section_font
    ws.merge_cells(f"A{row}:F{row}")
    row += 1
    ws[f"A{row}"] = "═" * 100
    row += 2

    if revenue_analysis.get('sga_642_analysis'):
        filtered_642 = _filter_to_latest_two_months(revenue_analysis['sga_642_analysis'], target_months)
        row = _add_account_group_section(ws, row, filtered_642, "SG&A Accounts (642*)", True,
                                        header_font, header_fill, subsection_font, thin_border, format_vnd)

    # ========================================
    # SECTION 6: KEY METRICS FOR PERIOD
    # ========================================
    row += 2
    ws[f"A{row}"] = "═" * 100
    row += 1
    ws[f"A{row}"] = "SECTION 6: KEY METRICS FOR PERIOD"
    ws[f"A{row}"].font = section_font
    ws.merge_cells(f"A{row}:F{row}")
    row += 1
    ws[f"A{row}"] = "═" * 100
    row += 2

    # Gross Margin for the two months
    if revenue_analysis.get('gross_margin_analysis', {}).get('trend'):
        ws[f"A{row}"] = "6.1 Gross Margin"
        ws[f"A{row}"].font = subsection_font
        row += 1

        # Filter to only the two target months
        all_trend = revenue_analysis['gross_margin_analysis']['trend']
        target_trend = [t for t in all_trend if t.get('month') in target_months]

        headers = ["Month", "Revenue", "Cost", "Gross Margin %", "Change from Previous"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row += 1

        prev_margin = None
        for margin_data in target_trend:
            ws[f"A{row}"] = margin_data.get('month', '')
            ws[f"B{row}"] = format_vnd(margin_data.get('revenue', 0))
            ws[f"C{row}"] = format_vnd(margin_data.get('cost', 0))
            ws[f"D{row}"] = f"{margin_data.get('gross_margin_pct', 0):.1f}%"

            current_margin = margin_data.get('gross_margin_pct', 0)
            if prev_margin is not None:
                change = current_margin - prev_margin
                ws[f"E{row}"] = f"{change:+.1f}pp"
                if change > 0:
                    ws.cell(row=row, column=5).fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                elif change < 0:
                    ws.cell(row=row, column=5).fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
            else:
                ws[f"E{row}"] = "N/A"

            prev_margin = current_margin

            for col in range(1, 6):
                ws.cell(row=row, column=col).border = thin_border
            row += 1
        row += 2

    # Combined SG&A for the two months
    if revenue_analysis.get('combined_sga_analysis', {}).get('ratio_trend'):
        ws[f"A{row}"] = "6.2 Combined SG&A (641* + 642*)"
        ws[f"A{row}"].font = subsection_font
        row += 1

        combined_analysis = revenue_analysis['combined_sga_analysis']
        ws[f"A{row}"] = f"Total 641* Accounts: {combined_analysis.get('total_641_accounts', 0)}"
        row += 1
        ws[f"A{row}"] = f"Total 642* Accounts: {combined_analysis.get('total_642_accounts', 0)}"
        row += 2

        # Filter to target months
        all_sga_trend = combined_analysis['ratio_trend']
        target_sga_trend = [s for s in all_sga_trend if s.get('month') in target_months]

        headers = ["Month", "Revenue", "641* Total", "642* Total", "Total SG&A", "SG&A Ratio %", "Change"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row += 1

        prev_ratio = None
        for sga_data in target_sga_trend:
            ws[f"A{row}"] = sga_data.get('month', '')
            ws[f"B{row}"] = format_vnd(sga_data.get('revenue', 0))
            ws[f"C{row}"] = format_vnd(sga_data.get('sga_641_total', 0))
            ws[f"D{row}"] = format_vnd(sga_data.get('sga_642_total', 0))
            ws[f"E{row}"] = format_vnd(sga_data.get('total_sga', 0))
            ws[f"F{row}"] = f"{sga_data.get('sga_ratio_pct', 0):.1f}%"

            current_ratio = sga_data.get('sga_ratio_pct', 0)
            if prev_ratio is not None:
                change = current_ratio - prev_ratio
                ws[f"G{row}"] = f"{change:+.1f}pp"
                if change > 2:
                    ws.cell(row=row, column=7).fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
                elif change < -2:
                    ws.cell(row=row, column=7).fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
            else:
                ws[f"G{row}"] = "N/A"

            prev_ratio = current_ratio

            for col in range(1, 8):
                ws.cell(row=row, column=col).border = thin_border
            row += 1
        row += 1

    # Auto-fit columns (skip merged cells)
    from openpyxl.cell.cell import MergedCell
    for column in ws.columns:
        max_length = 0
        column_letter = None
        for cell in column:
            # Skip merged cells
            if isinstance(cell, MergedCell):
                continue
            if column_letter is None:
                column_letter = cell.column_letter
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        if column_letter:
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column_letter].width = adjusted_width
