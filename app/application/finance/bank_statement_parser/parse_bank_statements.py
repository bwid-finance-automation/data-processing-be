"""Use case for parsing bank statements (batch processing)."""

from typing import List, Tuple, Dict, Any, Optional
from datetime import datetime
import pandas as pd
import csv
import math
import json
import re
from io import BytesIO, StringIO
from collections import defaultdict


def _safe_int(value, default=0) -> int:
    """Safely convert value to int, handling NaN, None, and invalid values."""
    if value is None:
        return default
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return default
    try:
        return int(value)
    except (ValueError, TypeError):
        return default


def _safe_float(value, default=0.0) -> float:
    """Safely convert value to float, handling NaN, None, and invalid values."""
    if value is None:
        return default
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return default
    try:
        return float(value)
    except (ValueError, TypeError):
        return default


def _is_valid_transaction(tx) -> bool:
    """
    Check if a transaction is valid (has meaningful data).

    A transaction is considered invalid if:
    - No transaction_id (SỐ CT) AND no debit AND no credit AND no date
    - All key fields are empty/zero

    Returns:
        True if transaction is valid, False otherwise
    """
    has_tx_id = tx.transaction_id and str(tx.transaction_id).strip()
    has_debit = tx.debit is not None and tx.debit != 0
    has_credit = tx.credit is not None and tx.credit != 0
    has_date = tx.date is not None

    # Transaction must have at least transaction_id OR (debit/credit AND date)
    if has_tx_id:
        return True
    if (has_debit or has_credit) and has_date:
        return True

    return False


from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.domain.finance.bank_statement_parser.models import BankStatement, BankTransaction, BankBalance
from .bank_parsers.parser_factory import ParserFactory
from .gemini_ocr_service import GeminiOCRService
from app.shared.utils.logging_config import get_logger

logger = get_logger(__name__)


class ParseBankStatementsUseCase:
    """
    Use case for batch parsing bank statements.

    Handles:
    - Auto-detection of banks
    - Parsing transactions and balances
    - Aggregating results from multiple files
    """

    def __init__(self):
        self.parser_factory = ParserFactory

    def _reconcile_balances_with_transactions(
        self,
        all_balances: List[BankBalance],
        all_transactions: List[BankTransaction]
    ) -> List[BankBalance]:
        """
        Backfill missing or zero balances using transaction totals.

        This helps PDF/OCR runs where opening/closing balances are missing or
        mis-placed by Gemini, but transaction lines are complete.
        """
        aggregates = defaultdict(lambda: {"debit": 0.0, "credit": 0.0})

        for tx in all_transactions:
            key = f"{tx.bank_name}_{tx.acc_no}"
            aggregates[key]["debit"] += tx.debit or 0.0
            aggregates[key]["credit"] += tx.credit or 0.0

        reconciled: List[BankBalance] = []
        for bal in all_balances:
            key = f"{bal.bank_name}_{bal.acc_no}"
            totals = aggregates.get(key, {"debit": 0.0, "credit": 0.0})

            opening = bal.opening_balance or 0.0
            closing = bal.closing_balance or 0.0
            total_debit = totals["debit"]
            total_credit = totals["credit"]
            # For bank accounts (Asset): Debit increases balance, Credit decreases
            # Net change = Debit - Credit (money in minus money out)
            net_change = total_debit - total_credit

            calculated_closing = opening + net_change
            calculated_opening = closing - net_change

            # If opening is missing/zero but we have closing + movement, backfill opening
            if (opening == 0 or opening is None) and (closing != 0 or net_change != 0):
                opening = calculated_opening

            # If closing is missing/zero, or equals opening despite movement, recompute
            if closing == 0 and (opening != 0 or net_change != 0):
                closing = calculated_closing
            elif (total_debit or total_credit) and abs(closing - opening) < 0.01:
                closing = calculated_closing

            # If no movement and closing is still zero, mirror opening
            if closing == 0 and opening != 0 and not (total_debit or total_credit):
                closing = opening

            reconciled.append(BankBalance(
                bank_name=bal.bank_name,
                acc_no=bal.acc_no,
                currency=bal.currency,
                opening_balance=opening,
                closing_balance=closing,
                statement_date=bal.statement_date  # Preserve statement period end date
            ))

        return reconciled

    def _update_statement_balances(
        self,
        statements: List[BankStatement],
        reconciled_balances: List[BankBalance]
    ) -> None:
        """Update statement.balance objects to reflect reconciled values."""
        balance_map = {
            (bal.bank_name, bal.acc_no): bal for bal in reconciled_balances
        }

        for statement in statements:
            if statement.balance:
                key = (statement.balance.bank_name, statement.balance.acc_no)
                if key in balance_map:
                    statement.balance = balance_map[key]

    def execute(
        self,
        files: List[Tuple[str, bytes]]
    ) -> Dict[str, Any]:
        """
        Parse multiple bank statement files.

        Args:
            files: List of (filename, file_bytes) tuples

        Returns:
            Dictionary with:
            - statements: List of parsed statements
            - all_transactions: Combined list of transactions
            - all_balances: Combined list of balances
            - summary: Processing summary
        """
        statements: List[BankStatement] = []
        all_transactions: List[BankTransaction] = []
        all_balances: List[BankBalance] = []

        successful = 0
        failed = 0
        failed_files = []

        for file_name, file_bytes in files:
            try:
                # Auto-detect bank
                parser = self.parser_factory.get_parser(file_bytes)

                if parser is None:
                    failed += 1
                    failed_files.append({
                        "file_name": file_name,
                        "error": "Bank not recognized"
                    })
                    continue

                # Parse transactions
                transactions = parser.parse_transactions(file_bytes, file_name)

                # Parse balances
                balance = parser.parse_balances(file_bytes, file_name)

                # Create statement
                statement = BankStatement(
                    bank_name=parser.bank_name,
                    file_name=file_name,
                    balance=balance,
                    transactions=transactions
                )

                statements.append(statement)
                all_transactions.extend(transactions)
                if balance:
                    all_balances.append(balance)

                successful += 1

            except Exception as e:
                failed += 1
                failed_files.append({
                    "file_name": file_name,
                    "error": str(e)
                })

        reconciled_balances = self._reconcile_balances_with_transactions(all_balances, all_transactions)
        self._update_statement_balances(statements, reconciled_balances)

        return {
            "statements": statements,
            "all_transactions": all_transactions,
            "all_balances": reconciled_balances,
            "summary": {
                "total_files": len(files),
                "successful": successful,
                "failed": failed,
                "failed_files": failed_files,
                "total_transactions": len(all_transactions),
                "total_balances": len(reconciled_balances),
                "total_accounts": len(reconciled_balances)  # Alias for UI display
            }
        }

    def execute_from_text(
        self,
        text_inputs: List[Tuple[str, str, str]]
    ) -> Dict[str, Any]:
        """
        Parse multiple bank statements from OCR text.

        Args:
            text_inputs: List of (file_name, ocr_text, bank_code) tuples
                        bank_code can be None for auto-detection

        Returns:
            Dictionary with same structure as execute():
            - statements: List of parsed statements
            - all_transactions: Combined list of transactions
            - all_balances: Combined list of balances
            - summary: Processing summary
        """
        statements: List[BankStatement] = []
        all_transactions: List[BankTransaction] = []
        all_balances: List[BankBalance] = []

        successful = 0
        failed = 0
        failed_files = []

        for file_name, ocr_text, bank_code in text_inputs:
            try:
                logger.info(f"Processing file: {file_name} (bank_code={bank_code})")
                # Debug: Log first 500 chars of OCR text
                logger.debug(f"OCR text preview for {file_name}: {ocr_text[:500]}...")
                parser = None

                # If bank_code provided, get parser by name
                if bank_code:
                    parser = self.parser_factory.get_parser_by_name(bank_code)
                    if parser is None:
                        failed += 1
                        failed_files.append({
                            "file_name": file_name,
                            "error": f"Unknown bank code: {bank_code}"
                        })
                        continue

                    # Validate parser supports text parsing
                    if not parser.can_parse_text(ocr_text):
                        failed += 1
                        failed_files.append({
                            "file_name": file_name,
                            "error": f"Parser {parser.bank_name} does not support OCR text parsing"
                        })
                        continue
                else:
                    # Auto-detect bank from text
                    parser = self.parser_factory.get_parser_for_text(ocr_text)
                    if parser is None:
                        failed += 1
                        failed_files.append({
                            "file_name": file_name,
                            "error": "Bank not recognized from OCR text"
                        })
                        continue

                # Parse transactions from text
                transactions = parser.parse_transactions_from_text(ocr_text, file_name)
                logger.info(f"  -> Detected bank: {parser.bank_name}, transactions: {len(transactions)}")

                # Parse ALL balances from text (supports multiple accounts per PDF)
                balances = parser.parse_all_balances_from_text(ocr_text, file_name)
                logger.info(f"  -> Balances: {len(balances)}, acc_no: {balances[0].acc_no if balances else 'N/A'}")

                # Create statement (use first balance for backward compatibility)
                statement = BankStatement(
                    bank_name=parser.bank_name,
                    file_name=file_name,
                    balance=balances[0] if balances else None,
                    transactions=transactions
                )

                statements.append(statement)
                all_transactions.extend(transactions)
                all_balances.extend(balances)

                successful += 1
                logger.info(f"  -> Success! Total transactions so far: {len(all_transactions)}")

            except Exception as e:
                failed += 1
                failed_files.append({
                    "file_name": file_name,
                    "error": str(e)
                })
                logger.error(f"  -> Exception parsing {file_name}: {e}")

        logger.info(f"Batch processing complete: {successful} successful, {failed} failed, {len(all_transactions)} total transactions")
        reconciled_balances = self._reconcile_balances_with_transactions(all_balances, all_transactions)
        self._update_statement_balances(statements, reconciled_balances)

        return {
            "statements": statements,
            "all_transactions": all_transactions,
            "all_balances": reconciled_balances,
            "summary": {
                "total_files": len(text_inputs),
                "successful": successful,
                "failed": failed,
                "failed_files": failed_files,
                "total_transactions": len(all_transactions),
                "total_balances": len(reconciled_balances),
                "total_accounts": len(reconciled_balances)  # Alias for UI display
            }
        }

    def execute_from_json(
        self,
        json_inputs: List[Tuple[str, str, Optional[str]]]
    ) -> Dict[str, Any]:
        """
        Parse bank statements from structured JSON (Gemini AI output).

        This is more accurate than text parsing because Gemini can visually
        understand table structures and align data correctly.

        Args:
            json_inputs: List of (file_name, json_content, bank_code) tuples
                        json_content can be raw JSON or markdown-wrapped (```json...```)
                        bank_code can be None (will be extracted from JSON if available)

        Expected JSON structure from Gemini:
        {
            "bank_name": "KBANK",
            "account_number": "102001530257",
            "currency": "VND",
            "opening_balance": 91419267.00,
            "closing_balance": 305348290.00,
            "transactions": [
                {
                    "date": "04/11/2025",
                    "time": "11:47 AM",
                    "description": "DDA Withdrawal by Transfer...",
                    "debit": 33000.00,
                    "credit": null,
                    "balance": 91386267.00,
                    "transaction_id": "TF50"
                }
            ]
        }

        Returns:
            Dictionary with same structure as execute():
            - statements: List of parsed statements
            - all_transactions: Combined list of transactions
            - all_balances: Combined list of balances
            - summary: Processing summary
        """
        statements: List[BankStatement] = []
        all_transactions: List[BankTransaction] = []
        all_balances: List[BankBalance] = []

        successful = 0
        failed = 0
        failed_files = []

        for file_name, json_content, bank_code in json_inputs:
            try:
                # Clean markdown code blocks if present
                clean_content = json_content.strip()
                if clean_content.startswith("```"):
                    # Remove ```json or ``` at start
                    clean_content = re.sub(r'^```(?:json)?\s*', '', clean_content)
                    # Remove ``` at end
                    clean_content = re.sub(r'\s*```$', '', clean_content)

                # Parse JSON
                data = json.loads(clean_content)

                # Extract bank info
                extracted_bank_name = data.get("bank_name") or data.get("bank") or bank_code or "UNKNOWN"
                acc_no = str(data.get("account_number") or data.get("acc_no") or "")
                currency = data.get("currency") or "VND"

                # Extract balance info
                opening_balance = _safe_float(data.get("opening_balance") or data.get("openning_balance") or 0)
                closing_balance = _safe_float(data.get("closing_balance") or 0)

                # Create balance object
                balance = BankBalance(
                    bank_name=extracted_bank_name,
                    acc_no=acc_no,
                    currency=currency,
                    opening_balance=opening_balance,
                    closing_balance=closing_balance
                )
                all_balances.append(balance)

                # Parse transactions
                transactions = []
                tx_list = data.get("transactions") or data.get("details") or []

                for tx_data in tx_list:
                    # Parse date
                    date_str = tx_data.get("date") or tx_data.get("trans_date") or ""
                    tx_date = None
                    if date_str:
                        # Try multiple date formats
                        for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"]:
                            try:
                                tx_date = datetime.strptime(date_str, fmt).date()
                                break
                            except ValueError:
                                continue

                    # Parse amounts
                    debit = _safe_float(tx_data.get("debit"))
                    credit = _safe_float(tx_data.get("credit"))

                    # Handle case where amount is in single field with type indicator
                    if "amount" in tx_data and "type" in tx_data:
                        amount = _safe_float(tx_data.get("amount"))
                        tx_type = tx_data.get("type", "").upper()
                        if tx_type in ["D", "DEBIT", "DR"]:
                            debit = amount
                            credit = 0
                        elif tx_type in ["C", "CREDIT", "CR"]:
                            credit = amount
                            debit = 0

                    tx = BankTransaction(
                        bank_name=extracted_bank_name,
                        acc_no=acc_no,
                        debit=debit if debit else None,
                        credit=credit if credit else None,
                        date=tx_date,
                        description=tx_data.get("description") or tx_data.get("desc") or "",
                        currency=currency,
                        transaction_id=str(tx_data.get("transaction_id") or tx_data.get("trans_id") or ""),
                        beneficiary_bank=tx_data.get("beneficiary_bank") or "",
                        beneficiary_acc_no=tx_data.get("beneficiary_acc_no") or tx_data.get("partner_account") or "",
                        beneficiary_acc_name=tx_data.get("beneficiary_acc_name") or tx_data.get("partner") or ""
                    )
                    transactions.append(tx)

                all_transactions.extend(transactions)

                # Create statement
                statement = BankStatement(
                    bank_name=extracted_bank_name,
                    file_name=file_name,
                    balance=balance,
                    transactions=transactions
                )
                statements.append(statement)
                successful += 1

                logger.info(f"Parsed JSON from {file_name}: {len(transactions)} transactions, bank={extracted_bank_name}")

            except json.JSONDecodeError as e:
                failed += 1
                failed_files.append({
                    "file_name": file_name,
                    "error": f"Invalid JSON format: {str(e)}"
                })
            except Exception as e:
                failed += 1
                failed_files.append({
                    "file_name": file_name,
                    "error": f"Failed to parse JSON: {str(e)}"
                })

        reconciled_balances = self._reconcile_balances_with_transactions(all_balances, all_transactions)
        self._update_statement_balances(statements, reconciled_balances)

        return {
            "statements": statements,
            "all_transactions": all_transactions,
            "all_balances": reconciled_balances,
            "summary": {
                "total_files": len(json_inputs),
                "successful": successful,
                "failed": failed,
                "failed_files": failed_files,
                "total_transactions": len(all_transactions),
                "total_balances": len(reconciled_balances),
                "total_accounts": len(reconciled_balances)  # Alias for UI display
            }
        }

    async def execute_from_pdf(
        self,
        pdf_inputs: List[Tuple[str, bytes, Optional[str], Optional[str]]],
        sequential: bool = False,
        max_concurrent: Optional[int] = None
    ) -> Dict[str, Any]:
        """
        Parse multiple bank statements from PDF files using Gemini OCR (async).

        This method is async to support non-blocking OCR processing with
        Redis caching and concurrent API calls.

        Args:
            pdf_inputs: List of (file_name, pdf_bytes, bank_code, password) tuples
                        bank_code can be None for auto-detection
                        password can be None for non-encrypted PDFs
            sequential: If True, process files one by one instead of concurrently.
                       Recommended when uploading many files to avoid rate limits.
            max_concurrent: Maximum concurrent requests (default: 3).
                           Only used when sequential=False.

        Returns:
            Dictionary with same structure as execute():
            - statements: List of parsed statements
            - all_transactions: Combined list of transactions
            - all_balances: Combined list of balances
            - summary: Processing summary (includes ocr_results for debugging)
            - ai_usage: AI API usage metrics (tokens, processing time, etc.)
        """
        # Initialize Gemini OCR service
        gemini_service = GeminiOCRService()

        # Step 1: Extract text from all PDFs using Gemini (async with caching)
        pdf_files = [(file_name, pdf_bytes, password) for file_name, pdf_bytes, _, password in pdf_inputs]
        ocr_results, ocr_metrics = await gemini_service.extract_text_from_pdf_batch(
            pdf_files,
            sequential=sequential,
            max_concurrent=max_concurrent
        )

        # Step 2: Build text_inputs for execute_from_text
        text_inputs: List[Tuple[str, str, str]] = []
        ocr_failed = []

        for i, (file_name, ocr_text, ocr_error) in enumerate(ocr_results):
            bank_code = pdf_inputs[i][2]  # Get bank_code from original input (index 2)

            if ocr_error:
                ocr_failed.append({
                    "file_name": file_name,
                    "error": f"OCR failed: {ocr_error}"
                })
            elif not ocr_text.strip():
                ocr_failed.append({
                    "file_name": file_name,
                    "error": "OCR returned empty text"
                })
            else:
                text_inputs.append((file_name, ocr_text, bank_code or ""))

        # Step 3: Parse the extracted text using existing execute_from_text
        if text_inputs:
            result = self.execute_from_text(text_inputs)
        else:
            result = {
                "statements": [],
                "all_transactions": [],
                "all_balances": [],
                "summary": {
                    "total_files": 0,
                    "successful": 0,
                    "failed": 0,
                    "failed_files": [],
                    "total_transactions": 0,
                    "total_balances": 0
                }
            }

        # Step 4: Merge OCR failures into the result
        result["summary"]["total_files"] = len(pdf_inputs)
        result["summary"]["failed"] += len(ocr_failed)
        result["summary"]["failed_files"].extend(ocr_failed)

        # Step 5: Add AI usage metrics to result
        result["ai_usage"] = ocr_metrics.to_dict()

        return result

    def export_to_excel(
        self,
        all_transactions: List[BankTransaction],
        all_balances: List[BankBalance]
    ) -> bytes:
        """
        Export parsed data to Excel format matching your schema.

        Creates Excel with sheets:
        1. Transactions - All transactions in standardized format
        2. Balances - All balance information
        3. Summary - Summary by bank and account

        Args:
            all_transactions: List of all transactions
            all_balances: List of all balances

        Returns:
            Excel file as bytes
        """
        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # ========== Handle Empty Data ==========
            if not all_transactions and not all_balances:
                # Create an info sheet when there's no data
                df_info = pd.DataFrame([{
                    "Message": "No transactions or balances found",
                    "Possible Reasons": "Bank not recognized from OCR text, or parser doesn't support this bank format"
                }])
                df_info.to_excel(writer, sheet_name="Info", index=False)
                output.seek(0)
                return output.read()

            # ========== Sheet 1: Transactions ==========
            if all_transactions:
                tx_data = []
                for tx in all_transactions:
                    # Skip invalid transactions (no TX ID, no amounts, no date)
                    if not _is_valid_transaction(tx):
                        continue
                        tx_data.append({
                        "Bank Name": tx.bank_name,
                        "Acc No": str(tx.acc_no) if tx.acc_no else "",
                        "Debit": tx.debit,
                        "Credit": tx.credit,
                        "Date": tx.date,
                        "Description": tx.description,
                        "Currency": tx.currency,
                        "Transaction ID": tx.transaction_id,
                        "Beneficiary Acc No": str(tx.beneficiary_acc_no) if tx.beneficiary_acc_no else "",
                        "Beneficiary Acc Name": tx.beneficiary_acc_name,
                        "Beneficiary Bank": tx.beneficiary_bank
                    })

                # Create DataFrame with explicit column order
                df_transactions = pd.DataFrame(tx_data, columns=[
                    "Bank Name",
                    "Acc No",
                    "Debit",
                    "Credit",
                    "Date",
                    "Description",
                    "Currency",
                    "Transaction ID",
                    "Beneficiary Acc No",
                    "Beneficiary Acc Name",
                    "Beneficiary Bank"
                ])
                df_transactions.to_excel(writer, sheet_name="Transactions", index=False)

                # Convert Transactions sheet to Table for formula references
                ws_transactions = writer.sheets["Transactions"]

                # Format account number columns as text to prevent scientific notation
                for row_idx in range(2, len(df_transactions) + 2):
                    # Column B (Acc No) - index 2
                    ws_transactions.cell(row=row_idx, column=2).number_format = '@'
                    # Column I (Beneficiary Acc No) - index 9
                    ws_transactions.cell(row=row_idx, column=9).number_format = '@'

                from openpyxl.worksheet.table import Table, TableStyleInfo
                tab = Table(displayName="Transactions_All_banks", ref=f"A1:K{len(df_transactions) + 1}")
                style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                     showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                tab.tableStyleInfo = style
                ws_transactions.add_table(tab)

            # ========== Sheet 2: Balances ==========
            if all_balances:
                # Build set of account keys that have VALID transactions
                tx_account_keys = set()
                for tx in all_transactions:
                    if _is_valid_transaction(tx):
                        tx_account_keys.add(f"{tx.bank_name}_{tx.acc_no}")

                bal_data = []
                for bal in all_balances:
                    key = f"{bal.bank_name}_{bal.acc_no}"

                    # Skip accounts with no valid transactions
                    # If OCR shows no transactions for an account, don't include it in balance export
                    has_transactions = key in tx_account_keys
                    if not has_transactions:
                        continue

                    bal_data.append({
                        "Bank Name": bal.bank_name,
                        "Acc No": str(bal.acc_no) if bal.acc_no else "",
                        "Currency": bal.currency,
                        "Opening Balance": bal.opening_balance,
                        "Closing Balance": bal.closing_balance,
                        "Debit": "",  # Will be filled with formula
                        "Credit": "",  # Will be filled with formula
                        "Checking": "",  # Will be filled with formula
                        "Remark": ""  # Will be filled with formula
                    })

                # Create DataFrame with explicit column order
                df_balances = pd.DataFrame(bal_data, columns=[
                    "Bank Name",
                    "Acc No",
                    "Currency",
                    "Opening Balance",
                    "Closing Balance",
                    "Debit",
                    "Credit",
                    "Checking",
                    "Remark"
                ])
                df_balances.to_excel(writer, sheet_name="Balances", index=False)

                # Add formulas to Balances sheet
                ws_balances = writer.sheets["Balances"]

                # Format Acc No column as text to prevent scientific notation
                for row_idx in range(2, len(df_balances) + 2):
                    ws_balances.cell(row=row_idx, column=2).number_format = '@'

                for row_idx in range(2, len(df_balances) + 2):  # Start from row 2 (after header)
                    # Debit formula: =SUMIFS(Transactions_All_banks[Debit];Transactions_All_banks[Acc No];[@[Acc No]];Transactions_All_banks[Bank Name];[@[Bank Name]])
                    ws_balances[f'F{row_idx}'] = f'=SUMIFS(Transactions_All_banks[Debit],Transactions_All_banks[Acc No],B{row_idx},Transactions_All_banks[Bank Name],A{row_idx})'

                    # Credit formula: =SUMIFS(Transactions_All_banks[Credit];Transactions_All_banks[Acc No];[@[Acc No]];Transactions_All_banks[Bank Name];[@[Bank Name]])
                    ws_balances[f'G{row_idx}'] = f'=SUMIFS(Transactions_All_banks[Credit],Transactions_All_banks[Acc No],B{row_idx},Transactions_All_banks[Bank Name],A{row_idx})'

                    # Checking formula: =[@[Opening Balance]]+[@Debit]-[@Credit]-[@[Closing Balance]]
                    ws_balances[f'H{row_idx}'] = f'=D{row_idx}+F{row_idx}-G{row_idx}-E{row_idx}'

                    # Remark formula: =IF([@Checking]=0;"Reconciled";"Mismatched")
                    ws_balances[f'I{row_idx}'] = f'=IF(H{row_idx}=0,"Reconciled","Mismatched")'

            # ========== Sheet 3: Summary ==========
            summary_data = []

            # Group by bank and account
            from collections import defaultdict
            by_account = defaultdict(lambda: {
                "transactions": 0,
                "total_debit": 0.0,
                "total_credit": 0.0,
                "opening": 0.0,
                "closing": 0.0
            })

            for tx in all_transactions:
                # Skip invalid transactions
                if not _is_valid_transaction(tx):
                    continue
                key = f"{tx.bank_name}_{tx.acc_no}"
                by_account[key]["transactions"] += 1
                by_account[key]["total_debit"] += tx.debit or 0.0
                by_account[key]["total_credit"] += tx.credit or 0.0

            for bal in all_balances:
                key = f"{bal.bank_name}_{bal.acc_no}"
                by_account[key]["opening"] = bal.opening_balance
                by_account[key]["closing"] = bal.closing_balance

            for key, stats in by_account.items():
                bank_name, acc_no = key.split("_", 1)
                summary_data.append({
                    "Bank Name": bank_name,
                    "Acc No": acc_no,
                    "Transaction Count": stats["transactions"],
                    "Total Debit": stats["total_debit"],
                    "Total Credit": stats["total_credit"],
                    "Opening Balance": stats["opening"],
                    "Closing Balance": stats["closing"],
                    "Calculated Closing": stats["opening"] - stats["total_debit"] + stats["total_credit"]
                })

            if summary_data:
                df_summary = pd.DataFrame(summary_data)
                df_summary.to_excel(writer, sheet_name="Summary", index=False)

        output.seek(0)
        return output.read()

    def export_to_netsuite_balance_csv(
        self,
        all_transactions: List[BankTransaction],
        all_balances: List[BankBalance]
    ) -> Tuple[bytes, Dict[str, str]]:
        """
        Export Balance CSV for NetSuite import.

        Creates CSV with 10 columns:
        External ID, Name (*), Bank Account Number (*), Bank code (*),
        Openning Balance (*), Closing Balance (*), Total Debit (*),
        Total Credit (*), Currency (*), Date (*)

        Args:
            all_transactions: List of all transactions
            all_balances: List of all balances

        Returns:
            Tuple of (CSV file as bytes, dict mapping bank_acc to external_id)
        """
        output = StringIO()
        writer = csv.writer(output)

        # Write header row (10 columns)
        headers = [
            "External ID",
            "Name (*)",
            "Bank Account Number (*)",
            "Bank code (*)",
            "Openning Balance (*)",
            "Closing Balance (*)",
            "Total Debit (*)",
            "Total Credit (*)",
            "Currency (*)",
            "Date (*)"
        ]
        writer.writerow(headers)

        # Get current date for External ID format
        now = datetime.now()
        date_suffix = now.strftime("%m%d%y")  # MMDDYY format

        # Aggregate VALID transactions by bank_name + acc_no
        tx_aggregates = defaultdict(lambda: {"total_debit": 0.0, "total_credit": 0.0, "max_date": None, "count": 0})
        for tx in all_transactions:
            # Skip invalid transactions
            if not _is_valid_transaction(tx):
                continue
            key = f"{tx.bank_name}_{tx.acc_no}"
            tx_aggregates[key]["total_debit"] += _safe_float(tx.debit)
            tx_aggregates[key]["total_credit"] += _safe_float(tx.credit)
            tx_aggregates[key]["count"] += 1
            if tx.date:
                if tx_aggregates[key]["max_date"] is None or tx.date > tx_aggregates[key]["max_date"]:
                    tx_aggregates[key]["max_date"] = tx.date

        # Map to store balance external IDs for linking with details
        balance_external_ids = {}

        seq = 0
        for bal in all_balances:
            key = f"{bal.bank_name}_{bal.acc_no}"

            # Skip accounts with no valid transactions
            # If OCR shows no transactions for an account, don't include it in balance export
            agg_check = tx_aggregates.get(key, {"count": 0})
            has_transactions = agg_check.get("count", 0) > 0
            if not has_transactions:
                continue

            seq += 1

            # External ID format: External ID{MMDDYY}_{SEQ:04d}
            external_id = f"External ID{date_suffix}_{seq:04d}"
            balance_external_ids[key] = external_id

            # Get aggregated totals
            agg = tx_aggregates.get(key, {"total_debit": 0, "total_credit": 0, "max_date": None})

            # Get date from transactions, or statement_date from balance, or use current date
            if agg["max_date"]:
                tx_date = agg["max_date"]
                date_str_yyyymmdd = tx_date.strftime("%Y%m%d")
                # Format: M/D/YYYY (no leading zeros)
                formatted_date = f"{tx_date.month}/{tx_date.day}/{tx_date.year}"
            elif bal.statement_date:
                # Use statement period end date from parsed file
                tx_date = bal.statement_date
                date_str_yyyymmdd = tx_date.strftime("%Y%m%d")
                formatted_date = f"{tx_date.month}/{tx_date.day}/{tx_date.year}"
            else:
                date_str_yyyymmdd = now.strftime("%Y%m%d")
                formatted_date = f"{now.month}/{now.day}/{now.year}"

            # Currency default
            currency = bal.currency if bal.currency else "VND"

            # Name format: BS/{BankCode}/{Currency}-{AccNo}/{YYYYMMDD}
            name = f"BS/{bal.bank_name}/{currency}-{bal.acc_no}/{date_str_yyyymmdd}"

            row = [
                external_id,
                name,
                bal.acc_no,
                bal.bank_name,
                _safe_int(bal.opening_balance),
                _safe_int(bal.closing_balance),
                agg["total_debit"],
                agg["total_credit"],
                currency,
                formatted_date
            ]
            writer.writerow(row)

        # Get CSV content and encode with UTF-8 BOM
        csv_content = output.getvalue()
        bom = b'\xef\xbb\xbf'
        return bom + csv_content.encode('utf-8'), balance_external_ids

    def export_to_netsuite_details_csv(
        self,
        all_transactions: List[BankTransaction],
        balance_external_ids: Dict[str, str]
    ) -> bytes:
        """
        Export Details CSV for NetSuite import.

        Creates CSV with 18 columns:
        External ID, External ID BSM Daily, Bank Statement Daily, Name (*),
        Bank Code (*), Bank Account Number (*), TRANS ID, Trans Date (*),
        Description (*), Currency(*), DEBIT (*), CREDIT (*), Amount, Type,
        Balance, PARTNER, PARTNER ACCOUNT, PARTNER BANK ID

        Args:
            all_transactions: List of all transactions
            balance_external_ids: Dict mapping bank_acc key to Balance External ID

        Returns:
            CSV file as bytes (UTF-8 with BOM)
        """
        output = StringIO()
        writer = csv.writer(output)

        # Write header row (18 columns)
        headers = [
            "External ID",
            "External ID BSM Daily",
            "Bank Statement Daily",
            "Name (*)",
            "Bank Code (*)",
            "Bank Account Number (*)",
            "TRANS ID",
            "Trans Date (*)",
            "Description (*)",
            "Currency(*)",
            "DEBIT (*)",
            "CREDIT (*)",
            "Amount",
            "Type",
            "Balance",
            "PARTNER",
            "PARTNER ACCOUNT",
            "PARTNER BANK ID"
        ]
        writer.writerow(headers)

        # Get current date for External ID format
        now = datetime.now()
        date_suffix = now.strftime("%m%d%y")  # MMDDYY format

        seq = 0
        for tx in all_transactions:
            # Skip invalid transactions
            if not _is_valid_transaction(tx):
                continue

            seq += 1

            # External ID format: External Idline_{MMDDYY}_{SEQ:04d}
            external_id = f"External Idline_{date_suffix}_{seq:04d}"

            # Link to parent Balance
            key = f"{tx.bank_name}_{tx.acc_no}"
            external_id_bsm_daily = balance_external_ids.get(key, "")

            # Currency default
            currency = tx.currency if tx.currency else "VND"

            # Date handling
            if tx.date:
                date_str_yyyymmdd = tx.date.strftime("%Y%m%d")
                # Format: M/D/YYYY (no leading zeros)
                formatted_date = f"{tx.date.month}/{tx.date.day}/{tx.date.year}"
            else:
                date_str_yyyymmdd = now.strftime("%Y%m%d")
                formatted_date = f"{now.month}/{now.day}/{now.year}"

            # Bank Statement Daily format: BS/{Bank}/{Currency}-{AccNo}{TxID}/{YYYYMMDD}
            trans_id = tx.transaction_id or ""
            bank_statement_daily = f"BS/{tx.bank_name}/{currency}-{tx.acc_no}{trans_id}/{date_str_yyyymmdd}"

            # Name format: same as Bank Statement Daily with trailing /
            name = f"{bank_statement_daily}/"

            # Debit/Credit as integers, 0 if None/NaN
            debit_val = _safe_int(tx.debit)
            credit_val = _safe_int(tx.credit)

            # Amount and Type
            if debit_val > 0:
                amount = debit_val
                tx_type = "D"
            else:
                amount = credit_val
                tx_type = "C"

            row = [
                external_id,
                external_id_bsm_daily,
                bank_statement_daily,
                name,
                tx.bank_name,
                tx.acc_no,
                trans_id,
                formatted_date,
                tx.description or "",
                currency,
                debit_val,
                credit_val,
                amount,
                tx_type,
                "",  # Balance - leave empty
                tx.beneficiary_acc_name or "",
                tx.beneficiary_acc_no or "",
                tx.beneficiary_bank or ""
            ]
            writer.writerow(row)

        # Get CSV content and encode with UTF-8 BOM
        csv_content = output.getvalue()
        bom = b'\xef\xbb\xbf'
        return bom + csv_content.encode('utf-8')

    def export_to_erp_template_excel(
        self,
        all_transactions: List[BankTransaction],
        all_balances: List[BankBalance]
    ) -> bytes:
        """
        Export to ERP Booking Template Excel format.

        Creates Excel with sheets:
        1. Template balance - Balance information per account/date
        2. Template details - All transactions

        Format follows 'ERP Booking template_NonAPI bank.xlsx'

        Args:
            all_transactions: List of all transactions
            all_balances: List of all balances

        Returns:
            Excel file as bytes
        """
        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Get current date for External ID
            now = datetime.now()
            date_suffix = now.strftime("%m%d%y")  # MMDDYY format

            # ========== Aggregate VALID transactions by bank_name + acc_no ==========
            tx_aggregates = defaultdict(lambda: {
                "total_debit": 0.0,
                "total_credit": 0.0,
                "max_date": None,
                "transactions": []
            })

            for tx in all_transactions:
                # Skip invalid transactions
                if not _is_valid_transaction(tx):
                    continue
                key = f"{tx.bank_name}_{tx.acc_no}"
                tx_aggregates[key]["total_debit"] += _safe_float(tx.debit)
                tx_aggregates[key]["total_credit"] += _safe_float(tx.credit)
                tx_aggregates[key]["transactions"].append(tx)
                if tx.date:
                    if tx_aggregates[key]["max_date"] is None or tx.date > tx_aggregates[key]["max_date"]:
                        tx_aggregates[key]["max_date"] = tx.date

            # Map to store balance external IDs for linking with details
            balance_external_ids = {}

            # ========== Sheet 1: Template balance ==========
            balance_data = []
            seq = 0

            for bal in all_balances:
                key = f"{bal.bank_name}_{bal.acc_no}"

                # Skip accounts with no valid transactions
                # If OCR shows no transactions for an account, don't include it in balance export
                agg_check = tx_aggregates.get(key, {"transactions": []})
                has_transactions = len(agg_check.get("transactions", [])) > 0
                if not has_transactions:
                    continue

                seq += 1

                # External ID format: External ID{MMDDYY}_{SEQ:04d}
                external_id = f"External ID{date_suffix}_{seq:04d}"
                balance_external_ids[key] = external_id

                # Get aggregated totals
                agg = tx_aggregates.get(key, {"total_debit": 0, "total_credit": 0, "max_date": None})

                # Get date from transactions, or statement_date from balance, or use current date
                if agg["max_date"]:
                    tx_date = agg["max_date"]
                    date_str_yyyymmdd = tx_date.strftime("%Y%m%d")
                elif bal.statement_date:
                    # Use statement period end date from parsed file
                    tx_date = bal.statement_date
                    date_str_yyyymmdd = tx_date.strftime("%Y%m%d")
                else:
                    tx_date = now
                    date_str_yyyymmdd = now.strftime("%Y%m%d")

                # Currency default
                currency = bal.currency if bal.currency else "VND"

                # Name format: BS/{BankCode}/{Currency}-{AccNo}/{YYYYMMDD}
                name = f"BS/{bal.bank_name}/{currency}-{bal.acc_no}/{date_str_yyyymmdd}"

                # For USD, keep decimals. For VND, use integers
                if currency == "USD":
                    opening_val = round(_safe_float(bal.opening_balance), 2)
                    closing_val = round(_safe_float(bal.closing_balance), 2)
                    debit_val = round(agg["total_debit"], 2)
                    credit_val = round(agg["total_credit"], 2)
                else:
                    opening_val = _safe_int(bal.opening_balance)
                    closing_val = _safe_int(bal.closing_balance)
                    debit_val = _safe_int(agg["total_debit"])
                    credit_val = int(agg["total_credit"])

                balance_data.append({
                    "External ID": external_id,
                    "Name (*)": name,
                    "Bank Account Number (*)": str(bal.acc_no) if bal.acc_no else "",
                    "Bank code (*)": bal.bank_name,
                    "Openning Balance (*)": opening_val,
                    "Closing Balance (*)": closing_val,
                    "Total Debit (*)": debit_val,
                    "Total Credit (*)": credit_val,
                    "Currency (*)": currency,
                    "Date (*)": tx_date
                })

            if balance_data:
                df_balance = pd.DataFrame(balance_data, columns=[
                    "External ID",
                    "Name (*)",
                    "Bank Account Number (*)",
                    "Bank code (*)",
                    "Openning Balance (*)",
                    "Closing Balance (*)",
                    "Total Debit (*)",
                    "Total Credit (*)",
                    "Currency (*)",
                    "Date (*)"
                ])
                df_balance.to_excel(writer, sheet_name="Template balance", index=False)

            # ========== Sheet 2: Template details ==========
            details_data = []
            seq = 0

            for tx in all_transactions:
                # Skip invalid transactions
                if not _is_valid_transaction(tx):
                    continue

                seq += 1

                # External ID format: External Idline_{MMDDYY}_{SEQ:04d}
                external_id = f"External Idline_{date_suffix}_{seq:04d}"

                # Link to parent Balance
                key = f"{tx.bank_name}_{tx.acc_no}"
                external_id_bsm_daily = balance_external_ids.get(key, "")

                # Currency default
                currency = tx.currency if tx.currency else "VND"

                # Date handling
                if tx.date:
                    tx_date = tx.date
                    date_str_yyyymmdd = tx.date.strftime("%Y%m%d")
                else:
                    tx_date = now
                    date_str_yyyymmdd = now.strftime("%Y%m%d")

                # Bank Statement Daily format: BS/{Bank}/{Currency}-{AccNo}{TxID}/{YYYYMMDD}
                trans_id = tx.transaction_id or ""
                bank_statement_daily = f"BS/{tx.bank_name}/{currency}-{tx.acc_no}{trans_id}/{date_str_yyyymmdd}"

                # Name format: same as Bank Statement Daily with trailing /
                name = f"{bank_statement_daily}/"

                # Debit/Credit - keep decimals for USD, integers for VND
                if currency == "USD":
                    debit_val = round(_safe_float(tx.debit), 2)
                    credit_val = round(_safe_float(tx.credit), 2)
                else:
                    debit_val = _safe_int(tx.debit)
                    credit_val = _safe_int(tx.credit)

                # Amount and Type
                if debit_val > 0:
                    amount = debit_val
                    tx_type = "D"
                else:
                    amount = credit_val
                    tx_type = "C"

                details_data.append({
                    "External ID": external_id,
                    "External ID BSM Daily": external_id_bsm_daily,
                    "Bank Statement Daily": bank_statement_daily,
                    "Name (*)": name,
                    "Bank Code (*)": tx.bank_name,
                    "Bank Account Number (*)": str(tx.acc_no) if tx.acc_no else "",
                    "TRANS ID": trans_id,
                    "Trans Date (*)": tx_date,
                    "Description (*)": tx.description or "",
                    "Currency(*)": currency,
                    "DEBIT (*)": debit_val,
                    "CREDIT (*)": credit_val,
                    "Amount": amount,
                    "Type": tx_type,
                    "Balance": "",
                    "PARTNER": tx.beneficiary_acc_name or "",
                    "PARTNER ACCOUNT": str(tx.beneficiary_acc_no) if tx.beneficiary_acc_no else "",
                    "PARTNER BANK ID": tx.beneficiary_bank or ""
                })

            if details_data:
                df_details = pd.DataFrame(details_data, columns=[
                    "External ID",
                    "External ID BSM Daily",
                    "Bank Statement Daily",
                    "Name (*)",
                    "Bank Code (*)",
                    "Bank Account Number (*)",
                    "TRANS ID",
                    "Trans Date (*)",
                    "Description (*)",
                    "Currency(*)",
                    "DEBIT (*)",
                    "CREDIT (*)",
                    "Amount",
                    "Type",
                    "Balance",
                    "PARTNER",
                    "PARTNER ACCOUNT",
                    "PARTNER BANK ID"
                ])
                df_details.to_excel(writer, sheet_name="Template details", index=False)

            # Handle empty case
            if not balance_data and not details_data:
                df_info = pd.DataFrame([{
                    "Message": "No transactions or balances found",
                    "Possible Reasons": "Bank not recognized or parser doesn't support this format"
                }])
                df_info.to_excel(writer, sheet_name="Info", index=False)

            # ========== Apply Styling ==========
            self._apply_excel_styling(writer)

        output.seek(0)
        return output.read()

    def _apply_excel_styling(self, writer: pd.ExcelWriter) -> None:
        """
        Apply lightweight styling to Excel sheets.
        - Header row: blue background, white bold text
        - Alternating row colors (zebra stripes)
        - Auto column width
        """
        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Alternating row colors (light blue for even rows - matches header)
        alt_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

        workbook = writer.book

        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]

            # Style header row (row 1)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            # Apply alternating row colors (skip header row)
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if row_idx % 2 == 0:  # Even rows get light gray
                    for cell in row:
                        cell.fill = alt_fill

            # Auto-fit column widths (with max limit to keep file light)
            for col_idx, column in enumerate(ws.columns, 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)

                for cell in column:
                    try:
                        cell_value = str(cell.value) if cell.value else ""
                        # Handle multiline cells
                        cell_length = max(len(line) for line in cell_value.split('\n')) if cell_value else 0
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass

                # Set width with min/max limits
                adjusted_width = min(max(max_length + 2, 8), 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Freeze header row
            ws.freeze_panes = "A2"

    def export_to_netsuite_excel(
        self,
        all_transactions: List[BankTransaction],
        all_balances: List[BankBalance]
    ) -> bytes:
        """
        Export to NetSuite Excel format with 2 sheets.

        Creates Excel file with:
        - Sheet 1: Balance (account balance summary)
        - Sheet 2: Details (all transactions)

        Args:
            all_transactions: List of all transactions
            all_balances: List of all balances

        Returns:
            Excel file as bytes
        """
        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Get current date for External ID
            now = datetime.now()
            date_suffix = now.strftime("%m%d%y")  # MMDDYY format

            # ========== Aggregate VALID transactions by bank_name + acc_no ==========
            tx_aggregates = defaultdict(lambda: {
                "total_debit": 0.0,
                "total_credit": 0.0,
                "max_date": None,
                "count": 0
            })

            for tx in all_transactions:
                # Skip invalid transactions
                if not _is_valid_transaction(tx):
                    continue
                key = f"{tx.bank_name}_{tx.acc_no}"
                tx_aggregates[key]["total_debit"] += _safe_float(tx.debit)
                tx_aggregates[key]["total_credit"] += _safe_float(tx.credit)
                tx_aggregates[key]["count"] += 1
                if tx.date:
                    if tx_aggregates[key]["max_date"] is None or tx.date > tx_aggregates[key]["max_date"]:
                        tx_aggregates[key]["max_date"] = tx.date

            # Map to store balance external IDs for linking with details
            balance_external_ids = {}

            # ========== Sheet 1: Balance ==========
            balance_data = []
            seq = 0

            for bal in all_balances:
                key = f"{bal.bank_name}_{bal.acc_no}"

                # Skip accounts with no valid transactions
                # If OCR shows no transactions for an account, don't include it in balance export
                agg_check = tx_aggregates.get(key, {"count": 0})
                has_transactions = agg_check.get("count", 0) > 0
                if not has_transactions:
                    continue

                seq += 1

                # External ID format: External ID{MMDDYY}_{SEQ:04d}
                external_id = f"External ID{date_suffix}_{seq:04d}"
                balance_external_ids[key] = external_id

                # Get aggregated totals
                agg = tx_aggregates.get(key, {"total_debit": 0, "total_credit": 0, "max_date": None})

                # Get date from transactions, or statement_date from balance, or use current date
                if agg["max_date"]:
                    tx_date = agg["max_date"]
                    date_str_yyyymmdd = tx_date.strftime("%Y%m%d")
                    # Format: M/D/YYYY (no leading zeros)
                    formatted_date = f"{tx_date.month}/{tx_date.day}/{tx_date.year}"
                elif bal.statement_date:
                    # Use statement period end date from parsed file
                    tx_date = bal.statement_date
                    date_str_yyyymmdd = tx_date.strftime("%Y%m%d")
                    formatted_date = f"{tx_date.month}/{tx_date.day}/{tx_date.year}"
                else:
                    tx_date = now
                    date_str_yyyymmdd = now.strftime("%Y%m%d")
                    formatted_date = f"{now.month}/{now.day}/{now.year}"

                # Currency default
                currency = bal.currency if bal.currency else "VND"

                # Name format: BS/{BankCode}/{Currency}-{AccNo}/{YYYYMMDD}
                name = f"BS/{bal.bank_name}/{currency}-{bal.acc_no}/{date_str_yyyymmdd}"

                # For USD, keep decimals. For VND, use integers
                if currency == "USD":
                    opening_val = round(_safe_float(bal.opening_balance), 2)
                    closing_val = round(_safe_float(bal.closing_balance), 2)
                    debit_val = round(agg["total_debit"], 2)
                    credit_val = round(agg["total_credit"], 2)
                else:
                    opening_val = _safe_int(bal.opening_balance)
                    closing_val = _safe_int(bal.closing_balance)
                    debit_val = _safe_int(agg["total_debit"])
                    credit_val = _safe_int(agg["total_credit"])

                balance_data.append({
                    "External ID": external_id,
                    "Name (*)": name,
                    "Bank Account Number (*)": str(bal.acc_no) if bal.acc_no else "",
                    "Bank code (*)": bal.bank_name,
                    "Openning Balance (*)": opening_val,
                    "Closing Balance (*)": closing_val,
                    "Total Debit (*)": debit_val,
                    "Total Credit (*)": credit_val,
                    "Currency (*)": currency,
                    "Date (*)": formatted_date
                })

            if balance_data:
                df_balance = pd.DataFrame(balance_data, columns=[
                    "External ID",
                    "Name (*)",
                    "Bank Account Number (*)",
                    "Bank code (*)",
                    "Openning Balance (*)",
                    "Closing Balance (*)",
                    "Total Debit (*)",
                    "Total Credit (*)",
                    "Currency (*)",
                    "Date (*)"
                ])
                df_balance.to_excel(writer, sheet_name="Balance", index=False)

            # ========== Sheet 2: Details ==========
            details_data = []
            seq = 0

            for tx in all_transactions:
                # Skip invalid transactions
                if not _is_valid_transaction(tx):
                    continue

                seq += 1

                # External ID format: External Idline_{MMDDYY}_{SEQ:04d}
                external_id = f"External Idline_{date_suffix}_{seq:04d}"

                # Link to parent Balance
                key = f"{tx.bank_name}_{tx.acc_no}"
                external_id_bsm_daily = balance_external_ids.get(key, "")

                # Currency default
                currency = tx.currency if tx.currency else "VND"

                # Date handling
                if tx.date:
                    tx_date = tx.date
                    date_str_yyyymmdd = tx.date.strftime("%Y%m%d")
                    # Format: M/D/YYYY (no leading zeros)
                    formatted_date = f"{tx.date.month}/{tx.date.day}/{tx.date.year}"
                else:
                    tx_date = now
                    date_str_yyyymmdd = now.strftime("%Y%m%d")
                    formatted_date = f"{now.month}/{now.day}/{now.year}"

                # Bank Statement Daily format: BS/{Bank}/{Currency}-{AccNo}{TxID}/{YYYYMMDD}
                trans_id = tx.transaction_id or ""
                bank_statement_daily = f"BS/{tx.bank_name}/{currency}-{tx.acc_no}{trans_id}/{date_str_yyyymmdd}"

                # Name format: same as Bank Statement Daily with trailing /
                name = f"{bank_statement_daily}/"

                # Debit/Credit - keep decimals for USD, integers for VND
                if currency == "USD":
                    debit_val = round(_safe_float(tx.debit), 2)
                    credit_val = round(_safe_float(tx.credit), 2)
                else:
                    debit_val = _safe_int(tx.debit)
                    credit_val = _safe_int(tx.credit)

                # Amount and Type
                if debit_val > 0:
                    amount = debit_val
                    tx_type = "D"
                else:
                    amount = credit_val
                    tx_type = "C"

                details_data.append({
                    "External ID": external_id,
                    "External ID BSM Daily": external_id_bsm_daily,
                    "Bank Statement Daily": bank_statement_daily,
                    "Name (*)": name,
                    "Bank Code (*)": tx.bank_name,
                    "Bank Account Number (*)": str(tx.acc_no) if tx.acc_no else "",
                    "TRANS ID": trans_id,
                    "Trans Date (*)": formatted_date,
                    "Description (*)": tx.description or "",
                    "Currency(*)": currency,
                    "DEBIT (*)": debit_val,
                    "CREDIT (*)": credit_val,
                    "Amount": amount,
                    "Type": tx_type,
                    "Balance": "",
                    "PARTNER": tx.beneficiary_acc_name or "",
                    "PARTNER ACCOUNT": str(tx.beneficiary_acc_no) if tx.beneficiary_acc_no else "",
                    "PARTNER BANK ID": tx.beneficiary_bank or ""
                })

            if details_data:
                df_details = pd.DataFrame(details_data, columns=[
                    "External ID",
                    "External ID BSM Daily",
                    "Bank Statement Daily",
                    "Name (*)",
                    "Bank Code (*)",
                    "Bank Account Number (*)",
                    "TRANS ID",
                    "Trans Date (*)",
                    "Description (*)",
                    "Currency(*)",
                    "DEBIT (*)",
                    "CREDIT (*)",
                    "Amount",
                    "Type",
                    "Balance",
                    "PARTNER",
                    "PARTNER ACCOUNT",
                    "PARTNER BANK ID"
                ])
                df_details.to_excel(writer, sheet_name="Details", index=False)

            # Handle empty case
            if not balance_data and not details_data:
                df_info = pd.DataFrame([{
                    "Message": "No transactions or balances found",
                    "Possible Reasons": "Bank not recognized or parser doesn't support this format"
                }])
                df_info.to_excel(writer, sheet_name="Info", index=False)

            # ========== Apply Styling ==========
            self._apply_excel_styling(writer)

        output.seek(0)
        return output.read()
