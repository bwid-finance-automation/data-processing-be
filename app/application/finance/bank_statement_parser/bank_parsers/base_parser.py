"""Base class for all bank statement parsers."""

from abc import ABC, abstractmethod
from typing import List, Optional, Tuple
import io

import pandas as pd
from openpyxl import load_workbook

from app.domain.finance.bank_statement_parser.models import BankTransaction, BankBalance


class BaseBankParser(ABC):
    """Abstract base class for bank-specific parsers."""

    # File format detection constants
    XLS_SIGNATURE = b'\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1'  # OLE2 compound document (real .xls)
    XLSX_SIGNATURE = b'PK'  # ZIP archive (xlsx/xlsm/xlsb)
    HTML_MARKERS = [b'<html', b'<!doctype', b'<table', b'<HTML', b'<!DOCTYPE']

    @property
    @abstractmethod
    def bank_name(self) -> str:
        """Return the bank name (e.g., 'ACB', 'OCB')."""
        pass

    @abstractmethod
    def can_parse(self, file_bytes: bytes) -> bool:
        """
        Detect if this parser can handle the given file.

        Args:
            file_bytes: Excel file as binary

        Returns:
            True if this parser recognizes the bank format
        """
        pass

    @abstractmethod
    def parse_transactions(self, file_bytes: bytes, file_name: str) -> List[BankTransaction]:
        """
        Parse transactions from bank statement.

        Args:
            file_bytes: Excel file as binary
            file_name: Original file name

        Returns:
            List of standardized transactions
        """
        pass

    @abstractmethod
    def parse_balances(self, file_bytes: bytes, file_name: str) -> Optional[BankBalance]:
        """
        Parse balance information from bank statement.

        Args:
            file_bytes: Excel file as binary
            file_name: Original file name

        Returns:
            Balance information or None if not found
        """
        pass

    # ========== OCR Text Parsing Methods (Optional - Override in subclass) ==========

    def can_parse_text(self, text: str) -> bool:
        """
        Detect if this parser can handle the given OCR text.

        Args:
            text: OCR extracted text from AI Builder

        Returns:
            True if this parser recognizes the text format.
            Default implementation returns False - override in subclass.
        """
        return False

    def parse_transactions_from_text(self, text: str, file_name: str) -> List[BankTransaction]:
        """
        Parse transactions from OCR text.

        Args:
            text: OCR extracted text
            file_name: Original file name

        Returns:
            List of standardized transactions.
            Default implementation returns empty list - override in subclass.
        """
        return []

    def parse_balances_from_text(self, text: str, file_name: str) -> Optional[BankBalance]:
        """
        Parse balance information from OCR text.

        Args:
            text: OCR extracted text
            file_name: Original file name

        Returns:
            Balance information or None if not found.
            Default implementation returns None - override in subclass.
        """
        return None

    def parse_all_balances_from_text(self, text: str, file_name: str) -> List[BankBalance]:
        """
        Parse ALL balance information from OCR text.

        For PDFs with multiple accounts, this returns all balances.

        Args:
            text: OCR extracted text
            file_name: Original file name

        Returns:
            List of balance information. Empty list if none found.
            Default implementation calls parse_balances_from_text for backward compatibility.
        """
        single_balance = self.parse_balances_from_text(text, file_name)
        return [single_balance] if single_balance else []

    # Helper methods available to all parsers

    @classmethod
    def detect_file_format(cls, file_bytes: bytes) -> str:
        """
        Detect the actual format of an Excel file.

        Returns:
            'xlsx' - Modern Excel format (ZIP-based)
            'xls' - Legacy Excel 97-2003 format (OLE2)
            'html' - HTML table saved as .xls (common from bank exports)
            'unknown' - Could not determine format
        """
        if len(file_bytes) < 8:
            return 'unknown'

        # Check for XLSX (ZIP signature)
        if file_bytes[:2] == cls.XLSX_SIGNATURE:
            return 'xlsx'

        # Check for real XLS (OLE2 signature)
        if file_bytes[:8] == cls.XLS_SIGNATURE:
            return 'xls'

        # Check for HTML (check first 500 bytes for HTML markers)
        header = file_bytes[:500].lower()
        for marker in cls.HTML_MARKERS:
            if marker.lower() in header:
                return 'html'

        return 'unknown'

    @classmethod
    def read_excel_auto(cls, file_bytes: bytes, **kwargs) -> pd.DataFrame:
        """
        Read Excel file with automatic format detection.

        Handles:
        - .xlsx files (modern Excel)
        - .xls files (Excel 97-2003)
        - HTML tables saved as .xls (common bank export format)

        Args:
            file_bytes: File content as bytes
            **kwargs: Additional arguments passed to pd.read_excel/pd.read_html

        Returns:
            DataFrame with file content

        Raises:
            ValueError: If file format cannot be determined or read
        """
        file_format = cls.detect_file_format(file_bytes)

        if file_format == 'html':
            # Read HTML table
            try:
                dfs = pd.read_html(io.BytesIO(file_bytes), **{k: v for k, v in kwargs.items() if k in ['header', 'index_col', 'skiprows']})
                if dfs:
                    return dfs[0]
                raise ValueError("No tables found in HTML file")
            except Exception as e:
                raise ValueError(f"Failed to read HTML table: {e}")

        elif file_format == 'xls':
            # Use xlrd for legacy Excel
            try:
                return pd.read_excel(io.BytesIO(file_bytes), engine='xlrd', **kwargs)
            except Exception as e:
                # Fallback to default engine
                try:
                    return pd.read_excel(io.BytesIO(file_bytes), **kwargs)
                except:
                    raise ValueError(f"Failed to read XLS file: {e}")

        elif file_format == 'xlsx':
            # Use openpyxl for modern Excel
            fixed_bytes = cls._ensure_visible_sheet(file_bytes)
            try:
                return pd.read_excel(io.BytesIO(fixed_bytes), engine='openpyxl', **kwargs)
            except Exception as e:
                if 'at least one sheet must be visible' in str(e).lower():
                    fixed_bytes = cls._ensure_visible_sheet(file_bytes)
                    return pd.read_excel(io.BytesIO(fixed_bytes), engine='openpyxl', **kwargs)
                raise ValueError(f"Failed to read XLSX file: {e}")

        else:
            # Try each engine in order
            engines = ['openpyxl', 'xlrd', None]
            last_error = None

            for engine in engines:
                try:
                    if engine:
                        candidate_bytes = (
                            cls._ensure_visible_sheet(file_bytes)
                            if engine == 'openpyxl'
                            else file_bytes
                        )
                        return pd.read_excel(io.BytesIO(candidate_bytes), engine=engine, **kwargs)
                    else:
                        return pd.read_excel(io.BytesIO(file_bytes), **kwargs)
                except Exception as e:
                    last_error = e
                    continue

            # Try HTML as last resort
            try:
                dfs = pd.read_html(io.BytesIO(file_bytes))
                if dfs:
                    return dfs[0]
            except:
                pass

            raise ValueError(f"Could not read file with any engine. Last error: {last_error}")

    @classmethod
    def get_excel_file(cls, file_bytes: bytes) -> pd.ExcelFile:
        """
        Get ExcelFile object with automatic engine detection.

        Args:
            file_bytes: File content as bytes

        Returns:
            pd.ExcelFile object

        Raises:
            ValueError: If file cannot be opened
        """
        file_format = cls.detect_file_format(file_bytes)

        if file_format == 'html':
            raise ValueError("HTML files cannot be opened as ExcelFile. Use read_excel_auto() instead.")

        if file_format == 'xls':
            try:
                return pd.ExcelFile(io.BytesIO(file_bytes), engine='xlrd')
            except:
                pass

        if file_format == 'xlsx':
            fixed_bytes = cls._ensure_visible_sheet(file_bytes)
            try:
                return pd.ExcelFile(io.BytesIO(fixed_bytes), engine='openpyxl')
            except:
                # Work around files where all sheets are hidden
                fixed_bytes = cls._ensure_visible_sheet(file_bytes)
                return pd.ExcelFile(io.BytesIO(fixed_bytes), engine='openpyxl')

        # Fallback: try default
        try:
            fallback_bytes = file_bytes
            if file_format == 'xlsx':
                fallback_bytes = cls._ensure_visible_sheet(file_bytes)

            return pd.ExcelFile(io.BytesIO(fallback_bytes))
        except Exception as e:
            raise ValueError(f"Could not open Excel file: {e}")

    @classmethod
    def _ensure_visible_sheet(cls, file_bytes: bytes) -> bytes:
        """Ensure workbook has at least one visible sheet.

        Some exported statements hide every sheet which causes ``openpyxl`` to
        raise ``ValueError: At least one sheet must be visible``. To keep the
        parser resilient, we mark the first sheet as visible and return the
        updated workbook bytes. If anything goes wrong, the original bytes are
        returned.
        """
        try:
            workbook = load_workbook(io.BytesIO(file_bytes))
            for sheet in workbook.worksheets:
                if sheet.sheet_state == "visible":
                    return file_bytes

            if workbook.sheetnames:
                first_sheet = workbook[workbook.sheetnames[0]]
                first_sheet.sheet_state = "visible"
                buffer = io.BytesIO()
                workbook.save(buffer)
                return buffer.getvalue()
        except Exception:
            return file_bytes

        return file_bytes

    @staticmethod
    def to_text(value) -> str:
        """Convert any value to text."""
        if value is None or pd.isna(value):
            return ""
        return str(value).strip()

    @staticmethod
    def fix_number(value) -> Optional[float]:
        """Convert value to number, handling Vietnamese formatting."""
        if value is None or pd.isna(value):
            return None

        txt = str(value).strip()
        # Remove commas and spaces
        cleaned = txt.replace(",", "").replace(" ", "")
        # Keep only numbers, decimal, and minus
        selected = ''.join(c for c in cleaned if c.isdigit() or c in ['-', '.'])

        if not selected:
            return None

        try:
            return float(selected)
        except (ValueError, TypeError):
            return None

    @staticmethod
    def fix_date(value):
        """
        Convert value to date.

        Vietnamese date format: DD/MM/YYYY
        Uses dayfirst=True to handle DD/MM correctly.
        """
        if value is None or pd.isna(value):
            return None

        # If already a datetime/date object
        if hasattr(value, 'date'):
            return value.date()

        # Try parsing text
        txt = str(value).strip()
        if ' ' in txt:
            txt = txt.split(' ')[0]  # Take date part only

        try:
            # Use dayfirst=True for Vietnamese DD/MM/YYYY format
            parsed = pd.to_datetime(txt, dayfirst=True, errors='coerce')
            if pd.notna(parsed):
                return parsed.date()
        except:
            pass

        return None
