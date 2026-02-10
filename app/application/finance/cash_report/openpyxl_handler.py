"""
Openpyxl Handler for Cash Report - Cross-platform Excel handler.
Replaces COM automation with optimized openpyxl operations.
Works on Windows, Linux, and macOS.
"""
import io
import math
import re
import zipfile
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import List, Optional, Dict, Any, Tuple
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

from app.shared.utils.logging_config import get_logger

logger = get_logger(__name__)
 

class OpenpyxlHandler:
    """
    Cross-platform Excel handler using openpyxl.
    Optimized for performance with large files.

    Features:
    - Formula template copying with row adjustment
    - Batch operations for better performance
    - Preserve formatting and formulas
    """

    # Column mapping for Movement sheet
    INPUT_COLUMNS = {
        1: 'source',    # A
        2: 'bank',      # B
        3: 'account',   # C
        4: 'date',      # D
        5: 'description',  # E
        6: 'debit',     # F
        7: 'credit',    # G
        9: 'nature',    # I
    }

    FORMULA_COLUMNS = [8, 10, 11, 12, 13, 14, 15, 16]  # H, J, K, L, M, N, O, P

    FIRST_DATA_ROW = 4

    def __init__(self):
        self._formula_cache: Dict[str, Dict[int, str]] = {}

    @staticmethod
    def is_available() -> bool:
        """Always available - openpyxl is cross-platform."""
        return True

    # ========== Byte-level XML helpers for Movement sheet ==========
    # These avoid parsing the full sheet XML (which hangs for 1M-row templates)
    # by scanning/modifying raw bytes and only parsing individual rows as needed.

    def _read_movement_xml(self, file_path: Path) -> Tuple[bytes, str, bytes]:
        """Read ZIP data and extract Movement sheet XML bytes."""
        with open(file_path, "rb") as f:
            zip_data = f.read()
        sheet_paths = self._get_sheet_xml_paths(zip_data)
        movement_path = sheet_paths.get("Movement")
        if not movement_path:
            raise ValueError("Movement sheet not found in workbook")
        with zipfile.ZipFile(io.BytesIO(zip_data), "r") as z:
            mov_xml = z.read(movement_path)
        return zip_data, movement_path, mov_xml

    @classmethod
    def _write_sheet_xml(
        cls, file_path: Path, zip_data: bytes, sheet_path: str, new_xml: bytes,
        styles_xml: Optional[bytes] = None,
        extra_entries: Optional[Dict[str, bytes]] = None,
    ) -> None:
        """Write modified sheet XML back to the ZIP file, removing calcChain.

        Args:
            file_path: Path to Excel file
            zip_data: Original ZIP data
            sheet_path: Path to sheet XML inside ZIP
            new_xml: Modified sheet XML bytes
            styles_xml: Optional modified styles.xml bytes
            extra_entries: Optional dict of {zip_entry_path: new_bytes} for additional files
                           to update in the same pass (e.g. table XML)
        """
        # Strip shared formulas from the modified sheet to prevent Excel corruption
        new_xml = cls._strip_shared_formulas_bytes(new_xml)
        new_xml = cls._sanitize_worksheet_xml_for_download(new_xml)
        if extra_entries is None:
            extra_entries = {}

        output = io.BytesIO()
        with zipfile.ZipFile(io.BytesIO(zip_data), "r") as src_zip:
            with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as dst_zip:
                for entry in src_zip.namelist():
                    # Skip calcChain to force Excel full recalculation
                    if "calcChain" in entry:
                        continue
                    if entry == sheet_path:
                        dst_zip.writestr(entry, new_xml)
                    elif entry in extra_entries:
                        dst_zip.writestr(entry, extra_entries[entry])
                    elif entry == "xl/styles.xml" and styles_xml is not None:
                        dst_zip.writestr(entry, styles_xml)
                    elif entry == "[Content_Types].xml":
                        ct_xml = src_zip.read(entry)
                        ct_xml = re.sub(
                            rb'<Override[^>]*calcChain[^>]*/>', b'', ct_xml
                        )
                        dst_zip.writestr(entry, ct_xml)
                    elif entry == "xl/workbook.xml":
                        wb_xml = src_zip.read(entry)
                        wb_xml = OpenpyxlHandler._set_full_calc_on_load(wb_xml)
                        dst_zip.writestr(entry, wb_xml)
                    elif entry.startswith("xl/worksheets/") and entry.endswith(".xml"):
                        # Strip shared formulas from ALL worksheets to prevent
                        # "Removed Records: Shared formula" errors on any sheet
                        sheet_data = src_zip.read(entry)
                        sheet_data = cls._strip_shared_formulas_bytes(sheet_data)
                        sheet_data = cls._sanitize_worksheet_xml_for_download(sheet_data)
                        dst_zip.writestr(entry, sheet_data)
                    else:
                        dst_zip.writestr(entry, src_zip.read(entry))
        with open(file_path, "wb") as f:
            f.write(output.getvalue())

    @classmethod
    def _write_multiple_sheets(
        cls, file_path: Path, zip_data: bytes,
        modified_sheets: Dict[str, bytes],
        extra_entries: Optional[Dict[str, bytes]] = None,
    ) -> None:
        """Write multiple modified sheet XMLs + extra entries in ONE ZIP pass.

        Like _write_sheet_xml but supports modifying several sheets at once,
        with proper shared-formula stripping for all worksheets.
        """
        if extra_entries is None:
            extra_entries = {}

        # Pre-strip shared formulas from all modified sheets
        for path in list(modified_sheets):
            modified_sheets[path] = cls._strip_shared_formulas_bytes(modified_sheets[path])
            modified_sheets[path] = cls._sanitize_worksheet_xml_for_download(modified_sheets[path])

        output = io.BytesIO()
        with zipfile.ZipFile(io.BytesIO(zip_data), "r") as src_zip:
            with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as dst_zip:
                for entry in src_zip.namelist():
                    if "calcChain" in entry:
                        continue
                    if entry in modified_sheets:
                        dst_zip.writestr(entry, modified_sheets[entry])
                    elif entry in extra_entries:
                        dst_zip.writestr(entry, extra_entries[entry])
                    elif entry == "[Content_Types].xml":
                        ct_xml = src_zip.read(entry)
                        ct_xml = re.sub(rb'<Override[^>]*calcChain[^>]*/>', b'', ct_xml)
                        dst_zip.writestr(entry, ct_xml)
                    elif entry == "xl/workbook.xml":
                        wb_xml = src_zip.read(entry)
                        wb_xml = cls._set_full_calc_on_load(wb_xml)
                        dst_zip.writestr(entry, wb_xml)
                    elif entry.startswith("xl/worksheets/") and entry.endswith(".xml"):
                        sheet_data = src_zip.read(entry)
                        sheet_data = cls._strip_shared_formulas_bytes(sheet_data)
                        sheet_data = cls._sanitize_worksheet_xml_for_download(sheet_data)
                        dst_zip.writestr(entry, sheet_data)
                    else:
                        dst_zip.writestr(entry, src_zip.read(entry))
        with open(file_path, "wb") as f:
            f.write(output.getvalue())

    @staticmethod
    def _set_full_calc_on_load(wb_xml: bytes) -> bytes:
        """Set fullCalcOnLoad='1' in workbook.xml so Excel recalculates all formulas on open."""
        if b"fullCalcOnLoad" in wb_xml:
            # Already has the attribute, ensure it's set to "1"
            wb_xml = re.sub(rb'fullCalcOnLoad="[^"]*"', b'fullCalcOnLoad="1"', wb_xml)
        elif b"<calcPr" in wb_xml:
            # calcPr exists, add the attribute
            wb_xml = wb_xml.replace(b"<calcPr", b'<calcPr fullCalcOnLoad="1"', 1)
        else:
            # No calcPr element, insert one before </workbook>
            wb_xml = wb_xml.replace(
                b"</workbook>",
                b'<calcPr fullCalcOnLoad="1"/></workbook>',
            )
        return wb_xml

    @classmethod
    def _find_row_byte_positions(
        cls, xml_bytes: bytes, target_rows: set
    ) -> Dict[int, Tuple[int, int]]:
        """
        Find byte (start, end) positions for specific row numbers in one scan.
        Stops early once all targets are found (fast for rows near the start).
        """
        positions: Dict[int, Tuple[int, int]] = {}
        target_set = set(target_rows) if not isinstance(target_rows, set) else target_rows
        if not target_set:
            return positions

        for m in re.finditer(rb'<row\s[^>]*r="(\d+)"', xml_bytes):
            row_num = int(m.group(1))
            if row_num not in target_set:
                continue

            start = m.start()
            tag_end = xml_bytes.index(b'>', m.end())

            if xml_bytes[tag_end - 1:tag_end] == b'/':
                end = tag_end + 1
            else:
                close = xml_bytes.find(b'</row>', tag_end)
                end = close + 6  # len('</row>')

            positions[row_num] = (start, end)
            if len(positions) == len(target_set):
                break

        return positions

    @classmethod
    def _parse_single_row(cls, row_bytes: bytes) -> ET.Element:
        """Parse a single <row> element from raw bytes with proper namespace handling."""
        cls._register_all_ns()
        ns_decls = ' '.join(
            f'xmlns:{prefix}="{uri}"' if prefix else f'xmlns="{uri}"'
            for prefix, uri in cls._ALL_NS.items()
        )
        wrapped = f'<_w {ns_decls}>'.encode() + row_bytes + b'</_w>'
        wrapper = ET.fromstring(wrapped)
        return wrapper[0]

    @classmethod
    def _serialize_row(cls, row_el: ET.Element) -> bytes:
        """Serialize a row element, stripping inherited namespace declarations."""
        cls._register_all_ns()
        row_bytes = ET.tostring(row_el, encoding="unicode").encode("utf-8")
        for prefix, uri in cls._ALL_NS.items():
            if prefix:
                decl = f' xmlns:{prefix}="{uri}"'.encode()
            else:
                decl = f' xmlns="{uri}"'.encode()
            row_bytes = row_bytes.replace(decl, b'', 1)
        return row_bytes

    @staticmethod
    def _cell_has_value(xml_bytes: bytes, match_end: int) -> bool:
        """Check if a <c> cell has a <v> or <is> value (not self-closing)."""
        gt = xml_bytes.find(b'>', match_end)
        if gt < 0:
            return False
        # Self-closing tag: <c .../>
        if xml_bytes[gt - 1:gt] == b'/':
            return False
        # Find </c> closing tag
        close = xml_bytes.find(b'</c>', gt)
        if close < 0:
            return False
        content = xml_bytes[gt + 1:close]
        return b'<v>' in content or b'<is>' in content

    def _find_last_data_row_bytes(self, xml_bytes: bytes) -> int:
        """Find last row with data in column C using byte scanning."""
        last_row = self.FIRST_DATA_ROW - 1
        for m in re.finditer(rb'<c\s[^>]*?r="C(\d+)"', xml_bytes):
            row_num = int(m.group(1))
            if self._cell_has_value(xml_bytes, m.end()):
                last_row = max(last_row, row_num)
        return last_row

    def _extract_row4_templates_bytes(
        self, xml_bytes: bytes
    ) -> Tuple[Dict[int, str], Dict[int, str]]:
        """Extract formula and style templates from row 4 via byte-level extraction."""
        positions = self._find_row_byte_positions(xml_bytes, {4})
        if 4 not in positions:
            return {}, {}

        start, end = positions[4]
        row4_el = self._parse_single_row(xml_bytes[start:end])

        ns = self._NS
        formula_templates: Dict[int, str] = {}
        style_templates: Dict[int, str] = {}

        for cell in row4_el:
            ref = cell.get("r", "")
            col_letter = re.match(r"([A-Z]+)", ref)
            if not col_letter:
                continue
            col_num = column_index_from_string(col_letter.group(1))

            f_el = cell.find(f"{{{ns}}}f")
            if f_el is not None and f_el.text:
                formula_templates[col_num] = f_el.text

            style_id = cell.get("s")
            if style_id:
                style_templates[col_num] = style_id

        return formula_templates, style_templates

    # ========== End byte-level helpers ==========

    def update_config(
        self,
        file_path: Path,
        opening_date: date,
        ending_date: date,
        fx_rate: Decimal,
        period_name: str,
    ) -> None:
        """
        Update Summary sheet configuration via direct XML/ZIP manipulation.

        Summary sheet structure:
        - B1: Date (ending date)
        - B3: FX rate (VND/USD)
        - B4: Week/Period name
        - B5: Opening date
        - B6: Ending date
        """
        with open(file_path, "rb") as f:
            zip_data = f.read()

        sheet_paths = self._get_sheet_xml_paths(zip_data)
        summary_path = sheet_paths.get("Summary")
        if not summary_path:
            raise ValueError("Summary sheet not found")

        output = io.BytesIO()
        with zipfile.ZipFile(io.BytesIO(zip_data), "r") as src_zip:
            summary_xml = src_zip.read(summary_path)
            new_summary_xml = self._modify_summary_xml(
                summary_xml, opening_date, ending_date, fx_rate, period_name
            )
            with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as dst_zip:
                for entry in src_zip.namelist():
                    if "calcChain" in entry:
                        continue
                    if entry == summary_path:
                        dst_zip.writestr(entry, self._strip_shared_formulas_bytes(new_summary_xml))
                    elif entry == "[Content_Types].xml":
                        ct_xml = src_zip.read(entry)
                        ct_xml = re.sub(
                            rb'<Override[^>]*calcChain[^>]*/>', b'', ct_xml
                        )
                        dst_zip.writestr(entry, ct_xml)
                    elif entry.startswith("xl/worksheets/") and entry.endswith(".xml"):
                        dst_zip.writestr(entry, self._strip_shared_formulas_bytes(src_zip.read(entry)))
                    else:
                        dst_zip.writestr(entry, src_zip.read(entry))

        with open(file_path, "wb") as f:
            f.write(output.getvalue())
        logger.info(f"Updated config via XML: {file_path.name}")

    def clear_movement_data(self, file_path: Path) -> int:
        """
        Clear Movement sheet data rows (row 5+) via byte-level XML manipulation.
        Keeps rows 1-3 as-is, clears input values from row 4.
        Does NOT parse the full XML tree (handles 1M-row templates efficiently).
        """
        zip_data, movement_path, mov_xml = self._read_movement_xml(file_path)
        ns = self._NS
        self._register_all_ns()

        # Find row 4 for modification
        row4_pos = self._find_row_byte_positions(mov_xml, {4})

        # Find first row >= 5
        first_ge5_start = None
        for m in re.finditer(rb'<row\s[^>]*r="(\d+)"', mov_xml):
            rn = int(m.group(1))
            if rn >= 5:
                first_ge5_start = m.start()
                break

        # Modify row 4 (clear input cells)
        new_row4 = None
        if 4 in row4_pos:
            r4_start, r4_end = row4_pos[4]
            row4_el = self._parse_single_row(mov_xml[r4_start:r4_end])
            input_cols = {"A", "B", "C", "D", "E", "F", "G", "I"}
            for cell in list(row4_el):
                ref = cell.get("r", "")
                col_match = re.match(r"([A-Z]+)", ref)
                if col_match and col_match.group(1) in input_cols:
                    f_el = cell.find(f"{{{ns}}}f")
                    if f_el is None:
                        # Clear value but keep cell with style for template extraction
                        for child in list(cell):
                            cell.remove(child)
                        cell.attrib.pop("t", None)
            new_row4 = self._serialize_row(row4_el)

        if first_ge5_start is None and new_row4 is None:
            return 0

        # Count rows being cleared (fast byte count)
        sd_close = mov_xml.rfind(b'</sheetData>')
        rows_cleared = 0
        if first_ge5_start is not None:
            section = mov_xml[first_ge5_start:sd_close]
            rows_cleared = len(re.findall(rb'<row\s', section))

        # Build result
        parts = []
        if new_row4 and 4 in row4_pos:
            r4_start, r4_end = row4_pos[4]
            parts.append(mov_xml[:r4_start])
            parts.append(new_row4)
            if first_ge5_start is not None:
                parts.append(b'\n')
                parts.append(mov_xml[sd_close:])
            else:
                parts.append(mov_xml[r4_end:])
        elif first_ge5_start is not None:
            parts.append(mov_xml[:first_ge5_start])
            parts.append(mov_xml[sd_close:])

        new_mov_xml = b''.join(parts)
        self._write_sheet_xml(file_path, zip_data, movement_path, new_mov_xml)

        logger.info(f"Cleared {rows_cleared} rows via byte-level XML")
        return rows_cleared

    def append_transactions(
        self,
        file_path: Path,
        transactions: List[Dict[str, Any]],
    ) -> Tuple[int, int]:
        """
        Append transactions to Movement sheet via byte-level XML manipulation.

        Instead of parsing the full sheet XML (which hangs for 1M-row templates),
        this method:
        1. Scans raw bytes to find row 4 (templates) and last data row
        2. Finds only the target rows' byte positions
        3. Parses/modifies individual rows (fast, small XML chunks)
        4. Splices modified row bytes into the original XML
        """
        if not transactions:
            return 0, 0

        zip_data, movement_path, mov_xml = self._read_movement_xml(file_path)
        ns = self._NS
        self._register_all_ns()

        # Step 1: Extract templates from row 4
        formula_templates, style_templates = self._extract_row4_templates_bytes(mov_xml)

        # Always look up correct date style from styles.xml for column D(4).
        # Row 4 may have a style, but it could be a general style, not a date format.
        date_style = self._find_date_style_index(zip_data)
        if date_style:
            style_templates[4] = date_style
            logger.info(f"Applied date style s={date_style} from styles.xml")

        # Step 1.5: Add quotePrefix style for column C (account) - like typing ' in Excel
        # This ensures VLOOKUP matches properly when lookup table has numbers
        with zipfile.ZipFile(io.BytesIO(zip_data), "r") as z:
            styles_xml = z.read("xl/styles.xml")
        base_style_c = style_templates.get(3)  # Column C existing style
        styles_xml, quote_style = self._add_quote_prefix_style(styles_xml, base_style_c)
        style_templates[3] = quote_style
        logger.info(f"Added quotePrefix style s={quote_style} for column C")

        # Step 2: Find last data row (scan column C for values)
        last_data_row = self._find_last_data_row_bytes(mov_xml)
        start_row = max(last_data_row + 1, self.FIRST_DATA_ROW)

        # Step 3: Find byte positions of target rows (only the ones we need)
        target_row_nums = set(range(start_row, start_row + len(transactions)))
        row_positions = self._find_row_byte_positions(mov_xml, target_row_nums)

        # Step 3.5: Collect shared formula masters from target rows
        # Needed to expand dependent shared formulas into regular formulas
        shared_formulas = self._collect_shared_formula_masters(
            mov_xml, ns, row_positions
        )

        # Step 4: Build modifications
        # modifications: (byte_start, byte_end, new_row_bytes) for existing rows
        # new_rows: bytes for rows that don't exist in the template
        modifications = []
        new_rows = []

        for i, tx in enumerate(transactions):
            row_num = start_row + i

            if row_num in row_positions:
                # Existing row (has formula cells) - merge input data
                start, end = row_positions[row_num]
                row_el = self._parse_single_row(mov_xml[start:end])
                self._merge_input_into_row_xml(
                    ns, row_el, row_num, tx, style_templates, shared_formulas,
                    formula_templates=formula_templates
                )
                new_bytes = self._serialize_row(row_el)
                modifications.append((start, end, new_bytes))
            else:
                # Row doesn't exist - build complete new row
                row_el = self._build_movement_row_xml(
                    ns, row_num, tx, formula_templates, style_templates
                )
                new_rows.append(self._serialize_row(row_el))

        # Step 5: Apply modifications in one pass (preserves byte positions)
        modifications.sort(key=lambda x: x[0])
        parts = []
        last_end = 0
        for start, end, new_bytes in modifications:
            parts.append(mov_xml[last_end:start])
            parts.append(new_bytes)
            last_end = end

        # Insert new rows (if any) before </sheetData>
        remaining = mov_xml[last_end:]
        if new_rows:
            sd_close = remaining.rfind(b'</sheetData>')
            if sd_close >= 0:
                parts.append(remaining[:sd_close])
                parts.extend(new_rows)
                parts.append(remaining[sd_close:])
            else:
                parts.append(remaining)
                parts.extend(new_rows)
        else:
            parts.append(remaining)

        new_mov_xml = b''.join(parts)

        # Step 6: Update dimension to reflect actual data range
        last_written_row = start_row + len(transactions) - 1
        new_mov_xml = self._update_dimension(new_mov_xml, last_written_row)

        # Step 7: Write back (with modified styles.xml for quotePrefix)
        self._write_sheet_xml(file_path, zip_data, movement_path, new_mov_xml, styles_xml)

        rows_added = len(transactions)
        total_rows = (start_row - self.FIRST_DATA_ROW) + rows_added
        logger.info(f"Appended {rows_added} transactions via byte-level XML, total: {total_rows}")
        return rows_added, total_rows

    def _collect_shared_formula_masters(
        self,
        xml_bytes: bytes,
        ns: str,
        row_positions: Dict[int, Tuple[int, int]],
    ) -> Dict[str, Tuple[str, int]]:
        """
        Collect shared formula masters from rows.
        Also scans row 4 (template row) for shared formula masters.

        Returns:
            Dict of {si: (formula_text, master_row_num)}
        """
        shared_formulas: Dict[str, Tuple[str, int]] = {}

        # Always include row 4 (template row) as potential master source
        scan_rows = dict(row_positions)
        row4_pos = self._find_row_byte_positions(xml_bytes, {4})
        scan_rows.update(row4_pos)

        for row_num in sorted(scan_rows.keys()):
            start, end = scan_rows[row_num]
            row_el = self._parse_single_row(xml_bytes[start:end])
            for cell in row_el:
                f_el = cell.find(f"{{{ns}}}f")
                if (
                    f_el is not None
                    and f_el.get("t") == "shared"
                    and f_el.text
                ):
                    si = f_el.get("si")
                    if si is not None and si not in shared_formulas:
                        shared_formulas[si] = (f_el.text, row_num)

        return shared_formulas

    @staticmethod
    def _to_excel_number_text(value: Any) -> Optional[str]:
        """
        Convert numeric values to stable Excel number text without float coercion.
        This avoids precision loss/scientific notation for large statement values.
        """
        if value is None:
            return None

        if isinstance(value, bool):
            return "1" if value else "0"

        if isinstance(value, Decimal):
            if not value.is_finite():
                return None
            return format(value, "f")

        if isinstance(value, int):
            return str(value)

        if isinstance(value, float):
            if math.isnan(value) or math.isinf(value):
                return None
            return format(Decimal(str(value)), "f")

        text = str(value).strip()
        if not text:
            return None

        # Normalize common formatted numeric strings.
        text = text.replace(",", "").replace(" ", "")
        try:
            return format(Decimal(text), "f")
        except Exception:
            return None

    def _merge_input_into_row_xml(
        self,
        ns: str,
        row_el: ET.Element,
        row_num: int,
        tx: Dict[str, Any],
        style_templates: Dict[int, str],
        shared_formulas: Optional[Dict[str, Tuple[str, int]]] = None,
        formula_templates: Optional[Dict[int, str]] = None,
    ) -> None:
        """
        Merge input data (columns A-G, I) into an existing row that already
        has formula cells. Preserves all existing formula cells.

        If a cell has no formula but formula_templates has one for that column,
        the formula is added (adjusted for the row number).
        """
        input_data = {
            1: ("str", tx.get("source", "")),
            2: ("str", tx.get("bank", "")),
            3: ("str", str(tx.get("account", ""))),  # quotePrefix applied via style
            4: ("date", tx.get("date")),
            5: ("str", tx.get("description", "")),
            6: ("num", tx.get("debit")),
            7: ("num", tx.get("credit")),
            9: ("str", tx.get("nature", "")),
        }

        # Collect existing cell styles as fallback (before removing them)
        existing_styles: Dict[int, str] = {}
        for cell in row_el:
            ref = cell.get("r", "")
            col_match = re.match(r"([A-Z]+)", ref)
            if col_match:
                col_num = column_index_from_string(col_match.group(1))
                s = cell.get("s")
                if s:
                    existing_styles[col_num] = s

        # Build new input cells
        new_cells = []
        for col_num, (dtype, value) in input_data.items():
            col_letter = get_column_letter(col_num)
            ref = f"{col_letter}{row_num}"
            cell = ET.Element(f"{{{ns}}}c")
            cell.set("r", ref)

            style = style_templates.get(col_num) or existing_styles.get(col_num)
            if style:
                cell.set("s", style)

            if dtype == "str" and value:
                cell.set("t", "inlineStr")
                is_el = ET.SubElement(cell, f"{{{ns}}}is")
                t_el = ET.SubElement(is_el, f"{{{ns}}}t")
                t_el.text = str(value)
            elif dtype == "num" and value is not None:
                v_el = ET.SubElement(cell, f"{{{ns}}}v")
                num_text = self._to_excel_number_text(value)
                if num_text is not None:
                    v_el.text = num_text
                else:
                    continue
            elif dtype == "date" and value is not None:
                v_el = ET.SubElement(cell, f"{{{ns}}}v")
                if isinstance(value, (date, datetime)):
                    v_el.text = str(self._date_to_serial(
                        value if isinstance(value, date) else value.date()
                    ))
                else:
                    v_el.text = str(value)

            new_cells.append((col_num, cell))

        # Collect existing cells, replacing input columns
        existing_by_col: Dict[int, ET.Element] = {}
        for cell in list(row_el):
            ref = cell.get("r", "")
            col_match = re.match(r"([A-Z]+)", ref)
            if col_match:
                col_num = column_index_from_string(col_match.group(1))
                existing_by_col[col_num] = cell

        # Remove existing input column cells (will be replaced)
        input_col_nums = set(input_data.keys())
        for col_num in input_col_nums:
            if col_num in existing_by_col:
                row_el.remove(existing_by_col[col_num])
                del existing_by_col[col_num]

        # Convert shared formulas to regular and remove cached <v>
        for col_num, cell in existing_by_col.items():
            f_el = cell.find(f"{{{ns}}}f")
            if f_el is not None:
                # Convert shared formula → regular formula
                if f_el.get("t") == "shared":
                    si = f_el.get("si")
                    if f_el.text:
                        # Master cell: keep formula text, strip shared attrs
                        pass
                    elif si and shared_formulas and si in shared_formulas:
                        # Dependent cell: expand from master formula
                        master_formula, master_row = shared_formulas[si]
                        f_el.text = self._expand_shared_formula(
                            master_formula, master_row, row_num
                        )
                    # Strip shared formula attributes
                    for attr in ("t", "si", "ref"):
                        if attr in f_el.attrib:
                            del f_el.attrib[attr]

                # Remove cached <v> to force Excel recalculation
                v_el = cell.find(f"{{{ns}}}v")
                if v_el is not None:
                    cell.remove(v_el)

        # For cells that have NO formula but formula_templates has one,
        # replace with a formula cell. This handles template rows with cached values.
        if formula_templates:
            input_col_nums = set(input_data.keys())
            for col_num, formula in formula_templates.items():
                if col_num in input_col_nums:
                    continue  # Skip input columns

                existing_cell = existing_by_col.get(col_num)
                has_formula = False
                if existing_cell is not None:
                    f_el = existing_cell.find(f"{{{ns}}}f")
                    has_formula = f_el is not None

                if not has_formula:
                    # Build a new formula cell
                    col_letter = get_column_letter(col_num)
                    ref = f"{col_letter}{row_num}"
                    cell = ET.Element(f"{{{ns}}}c")
                    cell.set("r", ref)

                    style = style_templates.get(col_num) or (
                        existing_cell.get("s") if existing_cell is not None else None
                    )
                    if style:
                        cell.set("s", style)

                    adjusted = self._adjust_formula(
                        f"={formula}", self.FIRST_DATA_ROW, row_num
                    )
                    f_el = ET.SubElement(cell, f"{{{ns}}}f")
                    f_el.text = adjusted[1:]  # Remove leading '='

                    # Remove old cell if exists
                    if existing_cell is not None and col_num in existing_by_col:
                        del existing_by_col[col_num]

                    # Add new formula cell
                    existing_by_col[col_num] = cell

        # Merge: add new input cells + keep existing formula cells
        all_cells = list(existing_by_col.items()) + new_cells
        all_cells.sort(key=lambda x: x[0])

        # Clear and rebuild row children in order
        for child in list(row_el):
            row_el.remove(child)
        for _, cell in all_cells:
            row_el.append(cell)

    def _build_movement_row_xml(
        self,
        ns: str,
        row_num: int,
        tx: Dict[str, Any],
        formula_templates: Dict[int, str],
        style_templates: Dict[int, str],
    ) -> ET.Element:
        """Build a complete <row> XML element for a transaction."""
        row_el = ET.Element(f"{{{ns}}}row")
        row_el.set("r", str(row_num))

        cells = []

        # Input columns: A(1), B(2), C(3), D(4), E(5), F(6), G(7), I(9)
        input_data = {
            1: ("str", tx.get("source", "")),
            2: ("str", tx.get("bank", "")),
            3: ("str", str(tx.get("account", ""))),  # quotePrefix applied via style
            4: ("date", tx.get("date")),
            5: ("str", tx.get("description", "")),
            6: ("num", tx.get("debit")),
            7: ("num", tx.get("credit")),
            9: ("str", tx.get("nature", "")),
        }

        for col_num, (dtype, value) in input_data.items():
            col_letter = get_column_letter(col_num)
            ref = f"{col_letter}{row_num}"
            cell = ET.Element(f"{{{ns}}}c")
            cell.set("r", ref)

            # Apply style from template
            style = style_templates.get(col_num)
            if style:
                cell.set("s", style)

            if dtype == "str" and value:
                cell.set("t", "inlineStr")
                is_el = ET.SubElement(cell, f"{{{ns}}}is")
                t_el = ET.SubElement(is_el, f"{{{ns}}}t")
                t_el.text = str(value)
            elif dtype == "num" and value is not None:
                v_el = ET.SubElement(cell, f"{{{ns}}}v")
                num_text = self._to_excel_number_text(value)
                if num_text is not None:
                    v_el.text = num_text
                else:
                    continue
            elif dtype == "date" and value is not None:
                v_el = ET.SubElement(cell, f"{{{ns}}}v")
                if isinstance(value, (date, datetime)):
                    v_el.text = str(self._date_to_serial(value if isinstance(value, date) else value.date()))
                else:
                    v_el.text = str(value)

            cells.append((col_num, cell))

        # Formula columns
        for col_num, formula in formula_templates.items():
            col_letter = get_column_letter(col_num)
            ref = f"{col_letter}{row_num}"
            cell = ET.Element(f"{{{ns}}}c")
            cell.set("r", ref)

            style = style_templates.get(col_num)
            if style:
                cell.set("s", style)

            adjusted = self._adjust_formula(
                f"={formula}", self.FIRST_DATA_ROW, row_num
            )
            f_el = ET.SubElement(cell, f"{{{ns}}}f")
            f_el.text = adjusted[1:]  # Remove leading '='

            cells.append((col_num, cell))

        # Sort by column number and add to row
        cells.sort(key=lambda x: x[0])
        for _, cell in cells:
            row_el.append(cell)

        return row_el

    def _adjust_formula(self, formula: str, source_row: int, target_row: int) -> str:
        """
        Adjust formula row references from source_row to target_row.
        Handles both relative and absolute ($) references.

        Examples:
            =F4-G4 -> =F23-G23 (when target_row=23)
            =VLOOKUP(I4,Lookup!$A:$B,2,FALSE) -> =VLOOKUP(I23,Lookup!$A:$B,2,FALSE)
            =$A$1 -> =$A$1 (absolute reference unchanged)
        """
        if not formula or not formula.startswith('='):
            return formula

        # Pattern to match cell references
        pattern = r'(\$?[A-Z]+)(\$?)(\d+)'

        def replace_ref(match):
            col = match.group(1)
            dollar = match.group(2)
            row_num = int(match.group(3))

            # If row is absolute ($), don't adjust
            if dollar == '$':
                return f"{col}${row_num}"

            # Only adjust if row matches source_row
            if row_num == source_row:
                return f"{col}{target_row}"

            return match.group(0)

        return re.sub(pattern, replace_ref, formula)

    @staticmethod
    def _expand_shared_formula(master_formula: str, master_row: int, target_row: int) -> str:
        """
        Expand a shared formula from master_row to target_row.
        Shifts ALL relative row references by (target_row - master_row).
        Absolute references ($) are preserved unchanged.

        Unlike _adjust_formula (which only shifts references matching source_row),
        this shifts every relative reference — required for shared formula expansion
        where cross-row references (e.g. SUM ranges) must also shift.
        """
        if not master_formula:
            return master_formula

        delta = target_row - master_row
        if delta == 0:
            return master_formula

        pattern = r'(\$?[A-Z]+)(\$?)(\d+)'

        def replace_ref(match):
            col = match.group(1)
            dollar = match.group(2)
            row_num = int(match.group(3))

            if dollar == '$':
                return f"{col}${row_num}"

            return f"{col}{row_num + delta}"

        return re.sub(pattern, replace_ref, master_formula)

    @staticmethod
    def _clean_formula_type_attrs(xml_bytes: bytes) -> bytes:
        """Remove t='str|n|b|e|s' from <c> tags that have <f> but no <v>.

        Excel treats ``t="str"`` without a ``<v>`` value as corrupt and
        shows 'Removed Records: Cell information'.  This method strips
        such type attributes from every affected cell.

        **MUST** be called on every worksheet XML written to a ZIP file.
        """
        # Quick check: any type attributes at all?
        if b' t="' not in xml_bytes:
            return xml_bytes

        def _clean(c_match):
            cell = c_match.group(0)
            # Check for <v> or <v  with attributes (e.g. <v xml:space="preserve">)
            has_value = b'<v>' in cell or b'<v ' in cell
            if b'<f' in cell and not has_value:
                cell = re.sub(rb'\s+t="(?:str|n|b|e|s)"', b'', cell)
            return cell

        return re.sub(rb'<c\s[^>]*>.*?</c>', _clean, xml_bytes, flags=re.DOTALL)

    @classmethod
    def _sanitize_worksheet_xml_for_download(cls, xml_bytes: bytes) -> bytes:
        """
        Lightweight worksheet sanitizer for download responses.

        Applies only idempotent safety cleanups:
        - Remove orphan <c> nodes outside <row> blocks
        - Remove invalid t= types from formula cells without cached <v>
        """
        xml_bytes = cls._remove_orphan_cells_outside_rows(xml_bytes)
        xml_bytes = cls._clean_formula_type_attrs(xml_bytes)
        return xml_bytes

    @classmethod
    def _strip_shared_formulas_bytes(cls, xml_bytes: bytes) -> bytes:
        """
        Strip ALL shared formulas from sheet XML, converting to regular formulas.
        Also cleans corrupt t= attributes from formula cells without <v>.
        """
        # Always clean t= attributes regardless of shared formulas
        cleaned = cls._clean_formula_type_attrs(xml_bytes)

        # Quick check: any shared formulas at all?
        if b't="shared"' not in cleaned:
            return cleaned

        # Pass 1: Collect shared formula masters
        masters: Dict[str, Tuple[str, int]] = {}  # si -> (formula, master_row)

        for c_match in re.finditer(rb'<c\s[^>]*?r="[A-Z]+(\d+)"', cleaned):
            c_start = c_match.start()
            row_num = int(c_match.group(1))

            peek = cleaned[c_start:c_start + 500]
            if b't="shared"' not in peek:
                continue

            tag_end = cleaned.find(b'>', c_match.end())
            if tag_end < 0:
                continue
            if cleaned[tag_end - 1:tag_end] == b'/':
                continue

            c_close = cleaned.find(b'</c>', tag_end)
            if c_close < 0:
                continue

            cell_content = cleaned[tag_end + 1:c_close]

            f_m = re.search(
                rb'<f\s[^>]*?t="shared"[^>]*?si="(\d+)"[^>]*?>([^<]+)</f>',
                cell_content,
            )
            if f_m:
                si = f_m.group(1).decode()
                formula = f_m.group(2).decode()
                if si not in masters:
                    masters[si] = (formula, row_num)

        if not masters:
            return cleaned

        logger.info(
            f"Stripping {len(masters)} shared formula groups from sheet XML"
        )

        # Pass 2: Collect all replacements
        replacements: list = []

        for c_match in re.finditer(rb'<c\s[^>]*?r="[A-Z]+(\d+)"', cleaned):
            c_start = c_match.start()
            row_num = int(c_match.group(1))

            peek = cleaned[c_start:c_start + 500]
            if b't="shared"' not in peek:
                continue

            tag_end = cleaned.find(b'>', c_match.end())
            if tag_end < 0:
                continue
            if cleaned[tag_end - 1:tag_end] == b'/':
                continue

            c_close = cleaned.find(b'</c>', tag_end)
            if c_close < 0:
                continue

            cell_content = cleaned[tag_end + 1:c_close]
            content_offset = tag_end + 1

            for f_m in re.finditer(
                rb'<f\s[^>]*?t="shared"[^>]*?si="(\d+)"[^>]*?(?:>([^<]*)</f>|/>)',
                cell_content,
            ):
                si = f_m.group(1).decode()
                formula_text = f_m.group(2)

                if formula_text is not None and formula_text:
                    new_f = b'<f>' + formula_text + b'</f>'
                elif si in masters:
                    master_formula, master_row = masters[si]
                    expanded = cls._expand_shared_formula(
                        master_formula, master_row, row_num
                    )
                    new_f = f'<f>{expanded}</f>'.encode()
                else:
                    new_f = b''

                abs_start = content_offset + f_m.start()
                abs_end = content_offset + f_m.end()
                replacements.append((abs_start, abs_end, new_f))

        if not replacements:
            return cleaned

        replacements.sort(key=lambda x: x[0], reverse=True)
        result = bytearray(cleaned)
        for start, end, new_bytes in replacements:
            result[start:end] = new_bytes
        logger.info(f"Converted {len(replacements)} shared formula elements to regular")

        # Run clean again after shared formula conversion (new regular formulas
        # may have inherited t= from the <c> tag of the shared dependent)
        return cls._clean_formula_type_attrs(bytes(result))

    def _read_sheet_xml(self, file_path: Path, sheet_name: str) -> Tuple[bytes, str, bytes]:
        """Read ZIP data and extract a named sheet's XML bytes."""
        with open(file_path, "rb") as f:
            zip_data = f.read()
        sheet_paths = self._get_sheet_xml_paths(zip_data)
        sheet_path = sheet_paths.get(sheet_name)
        if not sheet_path:
            raise ValueError(f"{sheet_name} sheet not found in workbook")
        with zipfile.ZipFile(io.BytesIO(zip_data), "r") as z:
            sheet_xml = z.read(sheet_path)
        return zip_data, sheet_path, sheet_xml

    def remove_zero_closing_balance_saving_rows(self, file_path: Path) -> int:
        """
        Remove rows from Saving Account sheet where CLOSING BALANCE (VND) = 0.

        Phase 1: Find CLOSING BALANCE (VND) column letter from headers (openpyxl).
        Phase 2: Read XML directly to detect zero values + collect shared formulas.
        Phase 3: Remove rows, convert shared formulas to regular, renumber.

        Returns:
            Number of rows removed
        """
        SAVING_SHEET = "Saving Account"
        DATA_START_ROW = 4

        # Phase 1: Find CLOSING BALANCE (VND) column letter from headers
        wb = load_workbook(file_path, data_only=True, read_only=True)
        try:
            if SAVING_SHEET not in wb.sheetnames:
                logger.warning("Saving Account sheet not found, skipping zero-balance cleanup")
                return 0

            ws = wb[SAVING_SHEET]
            closing_vnd_col = None  # column letter (e.g. "F")
            for row in ws.iter_rows(min_row=1, max_row=3, values_only=False):
                for cell in row:
                    val = str(cell.value or "").upper()
                    if "CLOSING BALANCE" in val and "VND" in val:
                        closing_vnd_col = get_column_letter(cell.column)
                        break
                if closing_vnd_col:
                    break
        finally:
            wb.close()

        if not closing_vnd_col:
            logger.warning("CLOSING BALANCE (VND) column not found in Saving Account headers")
            return 0

        # Phase 2: Read sheet XML, detect zero-balance rows + collect shared formulas
        zip_data, sheet_path, sheet_xml = self._read_sheet_xml(file_path, SAVING_SHEET)
        self._register_all_ns()
        ns = self._NS

        last_data_row = self._find_last_data_row_bytes(sheet_xml)
        if last_data_row < DATA_START_ROW:
            return 0

        target_rows = set(range(DATA_START_ROW, last_data_row + 1))
        row_positions = self._find_row_byte_positions(sheet_xml, target_rows)
        if not row_positions:
            return 0

        rows_to_remove = set()
        shared_formulas = {}  # si -> (formula_text, master_row)

        for row_num in sorted(row_positions.keys()):
            start, end = row_positions[row_num]
            row_el = self._parse_single_row(sheet_xml[start:end])

            has_account = False
            closing_is_zero = False

            for cell in row_el:
                ref = cell.get("r", "")
                col_match = re.match(r"([A-Z]+)", ref)
                if not col_match:
                    continue
                col = col_match.group(1)

                # Check column C for account data
                if col == "C":
                    v_el = cell.find(f"{{{ns}}}v")
                    is_el = cell.find(f"{{{ns}}}is")
                    if (v_el is not None and v_el.text) or is_el is not None:
                        has_account = True

                # Check CLOSING BALANCE (VND) column for zero via <v> element
                if col == closing_vnd_col:
                    v_el = cell.find(f"{{{ns}}}v")
                    if v_el is not None and v_el.text:
                        try:
                            if float(v_el.text) == 0:
                                closing_is_zero = True
                        except (ValueError, TypeError):
                            pass

                # Collect shared formula masters (needed for Phase 3)
                f_el = cell.find(f"{{{ns}}}f")
                if f_el is not None and f_el.get("t") == "shared" and f_el.text:
                    si = f_el.get("si")
                    if si is not None and si not in shared_formulas:
                        shared_formulas[si] = (f_el.text, row_num)

            if has_account and closing_is_zero:
                rows_to_remove.add(row_num)

        if not rows_to_remove:
            logger.info("No zero-closing-balance rows found in Saving Account sheet")
            return 0

        # Phase 3: Remove rows, convert shared formulas, renumber
        remaining_rows = []
        new_row_num = DATA_START_ROW

        for row_num in sorted(row_positions.keys()):
            if row_num in rows_to_remove:
                continue

            start, end = row_positions[row_num]
            row_el = self._parse_single_row(sheet_xml[start:end])

            # Renumber row
            row_el.set("r", str(new_row_num))
            for cell in row_el:
                ref = cell.get("r", "")
                col_match = re.match(r"([A-Z]+)", ref)
                if col_match:
                    cell.set("r", f"{col_match.group(1)}{new_row_num}")

                # Handle formulas
                f_el = cell.find(f"{{{ns}}}f")
                if f_el is not None:
                    if f_el.get("t") == "shared":
                        # Convert shared formula → regular formula
                        si = f_el.get("si")
                        if f_el.text:
                            # Master cell: expand from own row to new row
                            f_el.text = self._expand_shared_formula(
                                f_el.text, row_num, new_row_num
                            )
                        elif si and si in shared_formulas:
                            # Dependent cell: expand from master row to new row
                            master_formula, master_row = shared_formulas[si]
                            f_el.text = self._expand_shared_formula(
                                master_formula, master_row, new_row_num
                            )
                        # Strip shared formula attributes
                        for attr in ("t", "si", "ref"):
                            if attr in f_el.attrib:
                                del f_el.attrib[attr]
                    elif f_el.text and row_num != new_row_num:
                        # Regular formula: adjust row references
                        f_el.text = self._expand_shared_formula(
                            f_el.text, row_num, new_row_num
                        )

            row_bytes = self._serialize_row(row_el)
            remaining_rows.append(row_bytes)
            new_row_num += 1

        # Replace data row region
        sorted_positions = sorted(row_positions.values())
        first_start = sorted_positions[0][0]
        last_end = sorted_positions[-1][1]

        new_sheet_xml = b''.join([
            sheet_xml[:first_start],
            *remaining_rows,
            sheet_xml[last_end:],
        ])

        # Update dimension
        if remaining_rows:
            new_sheet_xml = self._update_dimension(new_sheet_xml, new_row_num - 1)

        self._write_sheet_xml(file_path, zip_data, sheet_path, new_sheet_xml)

        logger.info(f"Removed {len(rows_to_remove)} zero-closing-balance rows from Saving Account sheet")
        return len(rows_to_remove)

    def remove_rows_by_source(self, file_path: Path, source_name: str) -> int:
        """
        Remove rows with matching source name (column A) via byte-level manipulation.
        Only scans data rows (up to last data row), not the full 1M-row template.
        """
        zip_data, movement_path, mov_xml = self._read_movement_xml(file_path)
        self._register_all_ns()

        # Find last data row to limit our scan
        last_data_row = self._find_last_data_row_bytes(mov_xml)
        if last_data_row < self.FIRST_DATA_ROW:
            return 0

        # Find byte positions of all data rows
        target_rows = set(range(self.FIRST_DATA_ROW, last_data_row + 1))
        row_positions = self._find_row_byte_positions(mov_xml, target_rows)

        if not row_positions:
            return 0

        # Check column A for matching source
        rows_to_remove = set()
        for row_num in sorted(row_positions.keys()):
            start, end = row_positions[row_num]
            row_bytes = mov_xml[start:end]

            a_cell = re.search(
                rb'<c\s[^>]*r="A\d+"[^>]*>.*?</c>', row_bytes, re.DOTALL
            )
            if a_cell:
                t_match = re.search(rb'<t[^>]*>(.*?)</t>', a_cell.group(0), re.DOTALL)
                if t_match:
                    val = t_match.group(1).decode("utf-8", errors="replace").strip()
                    if val == source_name.strip():
                        rows_to_remove.add(row_num)

        if not rows_to_remove:
            return 0

        # Collect shared formula masters before modifying rows
        ns = self._NS
        shared_formulas = self._collect_shared_formula_masters(
            mov_xml, ns, row_positions
        )

        # Build remaining data rows with new numbering
        remaining_rows = []
        new_row_num = self.FIRST_DATA_ROW

        for row_num in sorted(row_positions.keys()):
            if row_num in rows_to_remove:
                continue

            start, end = row_positions[row_num]
            row_bytes = mov_xml[start:end]

            # Always parse to handle shared formula conversion + renumbering
            row_el = self._parse_single_row(row_bytes)

            if row_num != new_row_num:
                # Renumber this row
                row_el.set("r", str(new_row_num))
                for cell in row_el:
                    ref = cell.get("r", "")
                    col_match = re.match(r"([A-Z]+)", ref)
                    if col_match:
                        cell.set("r", f"{col_match.group(1)}{new_row_num}")

            # Convert shared formulas to regular
            for cell in row_el:
                f_el = cell.find(f"{{{ns}}}f")
                if f_el is not None and f_el.get("t") == "shared":
                    si = f_el.get("si")
                    if f_el.text:
                        # Master: expand from own row to new row
                        f_el.text = self._expand_shared_formula(
                            f_el.text, row_num, new_row_num
                        )
                    elif si and si in shared_formulas:
                        # Dependent: expand from master
                        master_formula, master_row = shared_formulas[si]
                        f_el.text = self._expand_shared_formula(
                            master_formula, master_row, new_row_num
                        )
                    for attr in ("t", "si", "ref"):
                        if attr in f_el.attrib:
                            del f_el.attrib[attr]
                elif f_el is not None and f_el.text and row_num != new_row_num:
                    # Regular formula: adjust row references
                    f_el.text = self._expand_shared_formula(
                        f_el.text, row_num, new_row_num
                    )

            row_bytes = self._serialize_row(row_el)
            remaining_rows.append(row_bytes)
            new_row_num += 1

        # Replace the data row region in the original XML
        # Everything before first data row + remaining rows + everything after last data row
        sorted_positions = sorted(row_positions.values())
        first_start = sorted_positions[0][0]
        last_end = sorted_positions[-1][1]

        parts = [
            mov_xml[:first_start],
            *remaining_rows,
            mov_xml[last_end:],  # preserves formula-only rows beyond data
        ]

        new_mov_xml = b''.join(parts)
        self._write_sheet_xml(file_path, zip_data, movement_path, new_mov_xml)

        logger.info(f"Removed {len(rows_to_remove)} rows with source '{source_name}' via byte-level XML")
        return len(rows_to_remove)

    def get_movement_row_count(self, file_path: Path) -> int:
        """Get the number of data rows in Movement sheet via byte scanning."""
        try:
            _, _, mov_xml = self._read_movement_xml(file_path)
        except ValueError:
            return 0

        count = 0
        for m in re.finditer(rb'<c\s[^>]*?r="C(\d+)"', mov_xml):
            row_num = int(m.group(1))
            if row_num >= self.FIRST_DATA_ROW:
                if self._cell_has_value(mov_xml, m.end()):
                    count += 1
        return count

    def append_transactions_batch(
        self,
        file_path: Path,
        transactions: List[Dict[str, Any]],
        batch_size: int = 1000
    ) -> Tuple[int, int]:
        """
        Append transactions in batches for very large datasets.

        This method saves periodically to prevent memory issues
        with very large transaction lists.

        Args:
            file_path: Path to Excel file
            transactions: List of transaction dicts
            batch_size: Number of transactions per batch

        Returns:
            Tuple of (rows_added, total_rows)
        """
        if not transactions:
            return 0, 0

        total_added = 0
        total_rows = 0

        # Process in batches
        for i in range(0, len(transactions), batch_size):
            batch = transactions[i:i + batch_size]
            rows_added, total_rows = self.append_transactions(file_path, batch)
            total_added += rows_added
            logger.info(f"Batch {i // batch_size + 1}: added {rows_added} rows")

        return total_added, total_rows

    # Column mapping: Cash Balance -> Cash balance (Prior period)
    # Cash Balance columns (source)
    CB_COL_ENTITY = 1       # A
    CB_COL_BRANCH = 2       # B
    CB_COL_ACCOUNT = 3      # C
    CB_COL_ACCOUNT_TYPE = 4 # D
    CB_COL_CURRENCY = 5     # E
    CB_COL_CLOSING_VND = 12 # L - CLOSING BALANCE (VND)2
    CB_COL_CLOSING_USD = 20 # T - CLOSING BALANCE (USD)
    CB_COL_NAME = 23        # W
    CB_COL_BANK_1 = 24      # X
    CB_COL_BANK = 25        # Y

    # Prior period columns (target)
    PP_COL_ENTITY = 1       # A
    PP_COL_BRANCH = 2       # B
    PP_COL_ACCOUNT = 3      # C
    PP_COL_ACCOUNT_TYPE = 4 # D
    PP_COL_CURRENCY = 5     # E
    PP_COL_CLOSING_VND = 6  # F
    PP_COL_CLOSING_USD = 7  # G
    PP_COL_NAME = 8         # H
    PP_COL_BANK_1 = 9       # I
    PP_COL_BANK = 10        # J

    PP_SHEET_NAME = "Cash balance (Prior period)"
    CB_SHEET_NAME = "Cash Balance"

    def _read_cash_balance_data(self, file_path: Path) -> Tuple[str, List[tuple]]:
        """
        Read computed values from Cash Balance sheet using data_only + read_only mode.

        Returns:
            Tuple of (period_name, list_of_row_tuples)
        """
        wb = load_workbook(file_path, data_only=True, read_only=True)
        try:
            # Read period name from Summary
            summary_ws = wb["Summary"]
            period_name = ""
            for row in summary_ws.iter_rows(min_row=4, max_row=4, min_col=2, max_col=2):
                period_name = row[0].value or ""

            # Read Cash Balance data using iter_rows (fast in read_only mode)
            cb_ws = wb[self.CB_SHEET_NAME]
            rows_data = []

            # Read columns: A,B,C,D,E (1-5), L(12), T(20), W(23), X(24), Y(25)
            for row in cb_ws.iter_rows(min_row=4, max_col=25):
                account = row[self.CB_COL_ACCOUNT - 1].value  # C (index 2)
                if not account or not str(account).strip():
                    continue

                # Closing balances might be non-numeric (e.g. "x" for inactive accounts)
                closing_vnd = row[self.CB_COL_CLOSING_VND - 1].value
                closing_usd = row[self.CB_COL_CLOSING_USD - 1].value
                if not isinstance(closing_vnd, (int, float)):
                    closing_vnd = 0
                if not isinstance(closing_usd, (int, float)):
                    closing_usd = 0

                rows_data.append((
                    row[self.CB_COL_ENTITY - 1].value,
                    row[self.CB_COL_BRANCH - 1].value,
                    row[self.CB_COL_ACCOUNT - 1].value,
                    row[self.CB_COL_ACCOUNT_TYPE - 1].value,
                    row[self.CB_COL_CURRENCY - 1].value,
                    closing_vnd,
                    closing_usd,
                    row[self.CB_COL_NAME - 1].value,
                    row[self.CB_COL_BANK_1 - 1].value,
                    row[self.CB_COL_BANK - 1].value,
                ))
        finally:
            wb.close()

        return period_name, rows_data

    # XML namespace used in xlsx sheet files
    _NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    _NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

    # All namespaces used in Excel sheet XML - must register ALL to prevent
    # ElementTree from mangling prefix names (e.g. xr -> ns2)
    _ALL_NS = {
        "": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "mc": "http://schemas.openxmlformats.org/markup-compatibility/2006",
        "x14ac": "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac",
        "xr": "http://schemas.microsoft.com/office/spreadsheetml/2014/revision",
        "xr2": "http://schemas.microsoft.com/office/spreadsheetml/2015/revision2",
        "xr3": "http://schemas.microsoft.com/office/spreadsheetml/2016/revision3",
    }

    @classmethod
    def _register_all_ns(cls):
        """Register all Excel XML namespaces to preserve prefixes during serialization."""
        for prefix, uri in cls._ALL_NS.items():
            ET.register_namespace(prefix, uri)

    @classmethod
    def _to_xml_bytes(cls, root: ET.Element) -> bytes:
        """Serialize XML element to bytes with proper Excel XML declaration."""
        xml_bytes = ET.tostring(root, xml_declaration=True, encoding="UTF-8")
        # ET uses standalone='no' or omits it; Excel requires standalone="yes"
        xml_bytes = xml_bytes.replace(
            b"<?xml version='1.0' encoding='UTF-8'?>",
            b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
        )
        return xml_bytes

    @classmethod
    def _patch_sheet_data(cls, original_xml: bytes, root: ET.Element) -> bytes:
        """
        Replace only <sheetData>...</sheetData> in the original XML bytes,
        preserving the original XML declaration, worksheet tag, namespaces,
        and all other elements (sheetViews, pageSetup, etc.) byte-for-byte.

        This avoids ElementTree mangling namespace prefixes (xr->ns2, dropping xr2/xr3).
        """
        ns = cls._NS
        # Serialize only the <sheetData> element from the modified tree
        sheet_data = root.find(f"{{{ns}}}sheetData")
        cls._register_all_ns()
        new_sd_bytes = ET.tostring(sheet_data, encoding="unicode").encode("utf-8")

        # Find <sheetData...> and </sheetData> in original bytes
        sd_open = re.search(rb'<sheetData[^>]*/?>', original_xml)
        sd_close_pos = original_xml.rfind(b'</sheetData>')

        if sd_open is None:
            # No sheetData in original - fall back to full serialization
            return cls._to_xml_bytes(root)

        if sd_close_pos > sd_open.start():
            # Normal case: <sheetData>...</sheetData>
            return (
                original_xml[:sd_open.start()]
                + new_sd_bytes
                + original_xml[sd_close_pos + len(b'</sheetData>'):]
            )
        else:
            # Self-closing: <sheetData/>
            return (
                original_xml[:sd_open.start()]
                + new_sd_bytes
                + original_xml[sd_open.end():]
            )

    @staticmethod
    def _date_to_serial(d: date) -> int:
        """Convert Python date to Excel serial number (1900 date system)."""
        return (d - date(1899, 12, 30)).days

    @classmethod
    def _find_date_style_index(cls, zip_data: bytes) -> Optional[str]:
        """
        Find a cell style index (xf) that uses d/m/yyyy date format.
        Falls back to any date format. Returns style index as string, or None.
        """
        ns = cls._NS
        with zipfile.ZipFile(io.BytesIO(zip_data)) as z:
            styles_xml = z.read("xl/styles.xml")

        root = ET.fromstring(styles_xml)

        # Collect custom numFmtIds that look like date formats
        custom_date_ids: Dict[int, str] = {}
        for nf in root.iter(f"{{{ns}}}numFmt"):
            code = nf.get("formatCode", "")
            fid = int(nf.get("numFmtId", "0"))
            if ("d" in code.lower() or "y" in code.lower()) and "m" in code.lower():
                custom_date_ids[fid] = code

        # Built-in date format IDs (14-22)
        builtin_date_ids = set(range(14, 23))

        # Prefer d/m/yyyy (numFmtId=173 in this template)
        preferred_ids = [
            fid for fid, code in custom_date_ids.items()
            if "d/m/yyyy" in code.lower() or "dd/mm/yyyy" in code.lower()
        ]

        logger.info(
            f"Date style lookup: custom_date_ids={custom_date_ids}, "
            f"preferred_ids={preferred_ids}"
        )

        # Search cellXfs for matching style
        cellXfs = root.find(f"{{{ns}}}cellXfs")
        if cellXfs is None:
            return None

        fallback = None
        fallback_any = None
        for idx, xf in enumerate(cellXfs):
            nf_id = int(xf.get("numFmtId", "0"))
            if nf_id in preferred_ids:
                return str(idx)
            if nf_id in builtin_date_ids or nf_id in custom_date_ids:
                if fallback is None and xf.get("applyNumberFormat") == "1":
                    fallback = str(idx)
                elif fallback_any is None:
                    fallback_any = str(idx)

        return fallback or fallback_any

    @staticmethod
    def _update_dimension(xml_bytes: bytes, last_row: int) -> bytes:
        """
        Update <dimension ref="..."> to reflect actual data range.
        Keeps original start ref and end column, updates end row.
        """
        match = re.search(rb'<dimension ref="([A-Z]+\d+):([A-Z]+)\d+"', xml_bytes)
        if match:
            start_ref = match.group(1).decode()
            end_col = match.group(2).decode()
            new_tag = f'<dimension ref="{start_ref}:{end_col}{last_row}"'.encode()
            return xml_bytes[:match.start()] + new_tag + xml_bytes[match.end():]
        return xml_bytes

    @staticmethod
    def _get_sheet_xml_paths(zip_data: bytes) -> Dict[str, str]:
        """
        Map sheet names to their XML file paths inside the xlsx zip.
        Parses workbook.xml and workbook.xml.rels.
        """
        with zipfile.ZipFile(io.BytesIO(zip_data)) as z:
            wb_xml = z.read("xl/workbook.xml")
            wb_rels = z.read("xl/_rels/workbook.xml.rels")

        ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        ns_r = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

        # sheet name -> rId
        wb_root = ET.fromstring(wb_xml)
        sheet_rids = {}
        for sheet_el in wb_root.iter(f"{{{ns}}}sheet"):
            sheet_rids[sheet_el.get("name")] = sheet_el.get(f"{{{ns_r}}}id")

        # rId -> file path
        rels_root = ET.fromstring(wb_rels)
        rid_to_path = {}
        for rel in rels_root:
            rid_to_path[rel.get("Id")] = f"xl/{rel.get('Target')}"

        return {name: rid_to_path[rid] for name, rid in sheet_rids.items()
                if rid in rid_to_path}

    def _build_pp_cell(
        self, col_letter: str, row: int, value,
        is_number: bool = False, style: str = None
    ) -> ET.Element:
        """Build an XML <c> element for the Prior period sheet."""
        ns = self._NS
        ref = f"{col_letter}{row}"
        cell = ET.Element(f"{{{ns}}}c")
        cell.set("r", ref)

        if style:
            cell.set("s", style)

        if value is None:
            return cell

        if is_number:
            v = ET.SubElement(cell, f"{{{ns}}}v")
            v.text = str(value)
        else:
            # Use inline string to avoid shared string table manipulation
            cell.set("t", "inlineStr")
            is_el = ET.SubElement(cell, f"{{{ns}}}is")
            t_el = ET.SubElement(is_el, f"{{{ns}}}t")
            t_el.text = str(value) if value else ""

        return cell

    def _build_catalog_row(self, row_num: int, cell_values: list) -> bytes:
        """Build a <row> XML element for catalog sheets (Acc_Char, Saving Account, Cash Balance).

        Args:
            row_num: 1-based row number
            cell_values: list of tuples: (value, is_number) or (value, is_number, style)

        Returns:
            Serialized row XML bytes
        """
        ns = self._NS
        self._register_all_ns()

        row_el = ET.Element(f"{{{ns}}}row")
        row_el.set("r", str(row_num))

        for col_idx, cell_spec in enumerate(cell_values, start=1):
            col_letter = get_column_letter(col_idx)
            if len(cell_spec) >= 3:
                value, is_number, style = cell_spec[0], cell_spec[1], cell_spec[2]
            else:
                value, is_number = cell_spec[0], cell_spec[1]
                style = None
            cell = self._build_pp_cell(col_letter, row_num, value, is_number=is_number, style=style)
            row_el.append(cell)

        return self._serialize_row(row_el)

    @staticmethod
    def _find_last_xml_row_num(sheet_xml: bytes) -> int:
        """Find the last row number in the sheet XML by scanning <row r='N'> attributes."""
        last_row = 0
        for m in re.finditer(rb'<row[^>]*\sr="(\d+)"', sheet_xml):
            row_num = int(m.group(1))
            if row_num > last_row:
                last_row = row_num
        return last_row

    @staticmethod
    def _remove_orphan_cells_outside_rows(sheet_xml: bytes) -> bytes:
        """
        Remove orphan <c ...> nodes that appear directly under <sheetData>
        (outside any <row ...> block).

        Corruption pattern observed in some generated files:
            </row><c r="I53"...></c><row r="53"...>
        This triggers Excel "Removed Records: Cell information".
        """
        sheet_data_open = re.search(rb'<sheetData[^>]*>', sheet_xml)
        if not sheet_data_open:
            return sheet_xml

        sheet_data_close = sheet_xml.find(b'</sheetData>', sheet_data_open.end())
        if sheet_data_close < 0:
            return sheet_xml

        body = sheet_xml[sheet_data_open.end():sheet_data_close]

        # Fast path: no orphan cell pattern between rows.
        if not re.search(rb'</row>\s*<c\s', body) and not re.search(rb'^\s*<c\s', body):
            return sheet_xml

        row_pattern = (
            rb'(<row[^>]*\sr="(\d+)"[^>]*/\s*>|'
            rb'<row[^>]*\sr="(\d+)"[^>]*>.*?</row>)'
        )
        row_matches = list(re.finditer(row_pattern, body, re.DOTALL))
        if not row_matches:
            return sheet_xml

        rebuilt_rows = [m.group(1) for m in row_matches]
        rebuilt_body = b'\n'.join(rebuilt_rows)
        cleaned = (
            sheet_xml[:sheet_data_open.end()]
            + b'\n'
            + rebuilt_body
            + b'\n'
            + sheet_xml[sheet_data_close:]
        )

        logger.warning("Removed orphan worksheet cell nodes outside <row> blocks")
        return cleaned

    @staticmethod
    def _dedupe_rows_by_number(sheet_xml: bytes) -> bytes:
        """
        Remove duplicate <row r="N"> blocks, keeping the richer row payload.

        Excel can emit "Removed Records: Cell information" when duplicated row
        indices exist in the same worksheet XML. This keeps one row per index:
        - Prefer row with more cell nodes (<c ...>)
        - Tie-breaker: prefer longer XML payload
        """
        # Also remove malformed standalone cells outside row blocks first.
        sheet_xml = OpenpyxlHandler._remove_orphan_cells_outside_rows(sheet_xml)

        row_pattern = rb'(<row[^>]*\sr="(\d+)"[^>]*/\s*>|<row[^>]*\sr="(\d+)"[^>]*>.*?</row>)'
        matches = list(re.finditer(row_pattern, sheet_xml, re.DOTALL))
        if not matches:
            return sheet_xml

        best_for_row: Dict[int, Tuple[int, int, int, int]] = {}
        # value tuple: (start, end, cell_count, byte_len)
        remove_ranges: List[Tuple[int, int]] = []

        for m in matches:
            row_xml = m.group(1)
            row_num_bytes = m.group(2) or m.group(3)
            if not row_num_bytes:
                continue
            row_num = int(row_num_bytes)
            start, end = m.start(1), m.end(1)
            cell_count = len(re.findall(rb'<c\b', row_xml))
            byte_len = end - start

            prev = best_for_row.get(row_num)
            if prev is None:
                best_for_row[row_num] = (start, end, cell_count, byte_len)
                continue

            prev_start, prev_end, prev_cells, prev_len = prev
            take_current = (cell_count > prev_cells) or (
                cell_count == prev_cells and byte_len > prev_len
            )
            if take_current:
                remove_ranges.append((prev_start, prev_end))
                best_for_row[row_num] = (start, end, cell_count, byte_len)
            else:
                remove_ranges.append((start, end))

        if not remove_ranges:
            return sheet_xml

        remove_ranges.sort(key=lambda x: x[0], reverse=True)
        out = bytearray(sheet_xml)
        for start, end in remove_ranges:
            del out[start:end]

        logger.warning(f"Removed {len(remove_ranges)} duplicate row blocks from worksheet XML")
        return bytes(out)

    @staticmethod
    def _xml_has_account(
        sheet_xml: bytes, account_no: str, col_letter: str = "B",
        zip_data: bytes = None,
    ) -> bool:
        """Check if an account number exists in the sheet XML by scanning cell values.

        Handles inline strings, numeric <v> values, AND shared string references.
        When zip_data is provided, resolves shared string indices via sharedStrings.xml.
        """
        account_str = str(account_no).strip()

        # Load shared strings table if available
        shared_strings = None
        if zip_data:
            try:
                with zipfile.ZipFile(io.BytesIO(zip_data), "r") as z:
                    if "xl/sharedStrings.xml" in z.namelist():
                        ss_xml = z.read("xl/sharedStrings.xml")
                        shared_strings = []
                        for t_m in re.finditer(rb'<si[^>]*>.*?</si>', ss_xml, re.DOTALL):
                            si_xml = t_m.group(0)
                            # Collect all <t> text within this <si>
                            parts = re.findall(rb'<t[^>]*>([^<]*)</t>', si_xml)
                            val = "".join(p.decode("utf-8", errors="ignore") for p in parts)
                            shared_strings.append(val)
            except Exception:
                pass

        # Scan target column cells
        pattern = rb'<c\s[^>]*r="' + col_letter.encode() + rb'\d+"[^>]*>.*?</c>'
        for m in re.finditer(pattern, sheet_xml, re.DOTALL):
            cell_xml = m.group(0)
            # Check inline string
            if b't="inlineStr"' in cell_xml:
                t_match = re.search(rb'<t[^>]*>([^<]*)</t>', cell_xml)
                if t_match:
                    val = t_match.group(1).decode("utf-8", errors="ignore").strip()
                    if val == account_str:
                        return True
                continue
            # Check <v> value
            v_match = re.search(rb'<v>([^<]*)</v>', cell_xml)
            if v_match:
                raw_val = v_match.group(1).decode("utf-8", errors="ignore").strip()
                # Shared string reference (t="s")
                if b't="s"' in cell_xml and shared_strings:
                    try:
                        idx = int(raw_val)
                        if 0 <= idx < len(shared_strings):
                            if shared_strings[idx].strip() == account_str:
                                return True
                    except (ValueError, IndexError):
                        pass
                else:
                    # Direct value (numeric account number)
                    if raw_val == account_str:
                        return True
        return False

    def _insert_row_before_sheet_data_close(self, sheet_xml: bytes, row_xml: bytes) -> bytes:
        """Insert a row XML before </sheetData>, handling both normal and self-closing tags."""
        if b'</sheetData>' in sheet_xml:
            return sheet_xml.replace(b'</sheetData>', row_xml + b'\n</sheetData>')
        # Self-closing <sheetData/>
        sd_match = re.search(rb'<sheetData\s*/?>', sheet_xml)
        if sd_match:
            tag = sheet_xml[sd_match.start():sd_match.end()]
            if tag.endswith(b'/>'):
                return (
                    sheet_xml[:sd_match.start()]
                    + b'<sheetData>' + row_xml + b'\n</sheetData>'
                    + sheet_xml[sd_match.end():]
                )
        raise ValueError("Cannot find <sheetData> in sheet XML")

    def _modify_prior_period_xml(
        self, xml_bytes: bytes, old_period_name: str, cb_rows: List[tuple]
    ) -> bytes:
        """
        Directly modify the Prior period sheet XML:
        - Update A2, B2 with period name
        - Replace data rows (4+) with Cash Balance closing balances
        """
        ns = self._NS
        ET.register_namespace("", ns)
        # Preserve other namespaces
        ET.register_namespace("r", self._NS_R)
        ET.register_namespace("mc", "http://schemas.openxmlformats.org/markup-compatibility/2006")
        ET.register_namespace("x14ac", "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac")

        root = ET.fromstring(xml_bytes)
        sheet_data = root.find(f"{{{ns}}}sheetData")

        # Extract styles from existing row 4 (template row) before removing
        pp_col_styles: Dict[str, str] = {}  # col_letter -> style_id
        for row_el in sheet_data:
            if row_el.get("r") == "4":
                for cell in row_el:
                    ref = cell.get("r", "")
                    style = cell.get("s")
                    col = ref.rstrip("0123456789")
                    if style and col:
                        pp_col_styles[col] = style
                break

        # Separate rows: keep 1-3, remove 4+
        rows_to_keep = []
        for row_el in list(sheet_data):
            row_num = int(row_el.get("r"))
            if row_num <= 3:
                rows_to_keep.append(row_el)
            sheet_data.remove(row_el)

        # Modify row 2: update A2 and B2 cells
        for row_el in rows_to_keep:
            if row_el.get("r") == "2":
                for cell in list(row_el):
                    ref = cell.get("r")
                    if ref in ("A2", "B2"):
                        row_el.remove(cell)
                # Add new A2 and B2
                a2 = self._build_pp_cell("A", 2, old_period_name)
                b2 = self._build_pp_cell("B", 2, f"OB{old_period_name}" if old_period_name else "")
                # Insert at the beginning of row
                row_el.insert(0, b2)
                row_el.insert(0, a2)
                break

        # Re-add kept rows
        for row_el in rows_to_keep:
            sheet_data.append(row_el)

        # Column letters for Prior period: A,B,C,D,E,F,G,H,I,J
        pp_cols = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]
        # Which are numbers: F(5) and G(6) are closing balances
        pp_is_number = [False, False, False, False, False, True, True, False, False, False]

        # Add new data rows with preserved styles
        for i, data in enumerate(cb_rows):
            row_num = 4 + i
            row_el = ET.SubElement(sheet_data, f"{{{ns}}}row")
            row_el.set("r", str(row_num))

            for col_idx, (col_letter, is_num) in enumerate(zip(pp_cols, pp_is_number)):
                value = data[col_idx]
                style = pp_col_styles.get(col_letter)
                cell = self._build_pp_cell(
                    col_letter, row_num, value,
                    is_number=is_num, style=style
                )
                row_el.append(cell)

        return self._patch_sheet_data(xml_bytes, root)

    def _modify_summary_xml(
        self,
        xml_bytes: bytes,
        opening_date: date,
        ending_date: date,
        fx_rate: Decimal,
        period_name: str,
    ) -> bytes:
        """
        Directly modify Summary sheet XML:
        - B1: ending_date, B3: fx_rate, B4: period_name
        - B5: opening_date, B6: ending_date
        """
        ns = self._NS
        self._register_all_ns()

        root = ET.fromstring(xml_bytes)
        sheet_data = root.find(f"{{{ns}}}sheetData")

        # Map of cell ref -> (value, is_number, style_id)
        updates = {
            "B1": (self._date_to_serial(ending_date), True),
            "B3": (float(fx_rate), True),
            "B4": (period_name, False),
            "B5": (self._date_to_serial(opening_date), True),
            "B6": (self._date_to_serial(ending_date), True),
        }

        for row_el in sheet_data:
            row_num = row_el.get("r")
            if row_num not in ("1", "2", "3", "4", "5", "6"):
                continue

            for cell in row_el:
                ref = cell.get("r")
                if ref not in updates:
                    continue

                value, is_number = updates[ref]
                style = cell.get("s")  # Preserve style

                # Clear existing content
                for child in list(cell):
                    cell.remove(child)

                if is_number:
                    cell.attrib.pop("t", None)  # Remove type attr for numbers
                    v = ET.SubElement(cell, f"{{{ns}}}v")
                    v.text = str(value)
                else:
                    cell.set("t", "inlineStr")
                    is_el = ET.SubElement(cell, f"{{{ns}}}is")
                    t_el = ET.SubElement(is_el, f"{{{ns}}}t")
                    t_el.text = str(value)

                if style:
                    cell.set("s", style)

        return self._patch_sheet_data(xml_bytes, root)

    def initialize_session_optimized(
        self,
        file_path: Path,
        opening_date: date,
        ending_date: date,
        fx_rate: Decimal,
        period_name: str,
    ) -> int:
        """
        Initialize session: update Summary config via direct XML manipulation.

        Only updates Summary sheet (dates, FX rate, period name).
        Cash Balance copy + Movement clear happen later at upload time
        via prepare_movement_for_writing().

        Returns:
            0
        """
        with open(file_path, "rb") as f:
            zip_data = f.read()

        sheet_paths = self._get_sheet_xml_paths(zip_data)
        summary_path = sheet_paths["Summary"]

        output = io.BytesIO()
        with zipfile.ZipFile(io.BytesIO(zip_data), "r") as src_zip:
            with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as dst_zip:
                for entry in src_zip.namelist():
                    if "calcChain" in entry:
                        continue

                    data = src_zip.read(entry)

                    if entry == summary_path:
                        data = self._modify_summary_xml(
                            data, opening_date, ending_date, fx_rate, period_name
                        )
                    elif entry == "[Content_Types].xml":
                        data = re.sub(
                            rb'<Override[^>]*calcChain[^>]*/>', b'', data
                        )
                    elif entry == "xl/workbook.xml":
                        data = self._set_full_calc_on_load(data)

                    # Strip shared formulas from all worksheets
                    if entry.startswith("xl/worksheets/") and entry.endswith(".xml"):
                        data = self._strip_shared_formulas_bytes(data)

                    dst_zip.writestr(entry, data)

        with open(file_path, "wb") as f:
            f.write(output.getvalue())

        logger.info(f"Session initialized: Summary config updated to {period_name}")
        return 0

    def prepare_movement_for_writing(self, file_path: Path) -> int:
        """
        Prepare working file for writing new transactions:
        1. Copy Cash Balance → Cash balance (Prior period)
        2. Clear Movement sheet data rows (row 5+, clear row 4 inputs)

        Called once before the first upload writes to Movement.

        Returns:
            Number of Movement rows cleared
        """
        # Phase 1: Read Cash Balance computed values
        old_period_name, cb_rows = self._read_cash_balance_data(file_path)

        # Phase 2: Copy to Prior Period + Clear Movement via ZIP/XML
        with open(file_path, "rb") as f:
            zip_data = f.read()

        sheet_paths = self._get_sheet_xml_paths(zip_data)
        pp_path = sheet_paths[self.PP_SHEET_NAME]
        movement_path = sheet_paths["Movement"]

        # Read Movement XML to clear it
        with zipfile.ZipFile(io.BytesIO(zip_data), "r") as z:
            mov_xml = z.read(movement_path)

        # Clear Movement data rows (same logic as clear_movement_data)
        rows_cleared = self._clear_movement_rows(mov_xml)
        new_mov_xml = rows_cleared[1]
        rows_cleared_count = rows_cleared[0]

        # Write all changes in one ZIP pass
        output = io.BytesIO()
        with zipfile.ZipFile(io.BytesIO(zip_data), "r") as src_zip:
            with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as dst_zip:
                for entry in src_zip.namelist():
                    if "calcChain" in entry:
                        continue

                    data = src_zip.read(entry)

                    if entry == pp_path and cb_rows:
                        data = self._modify_prior_period_xml(
                            data, old_period_name, cb_rows
                        )
                    elif entry == movement_path:
                        data = new_mov_xml
                    elif entry == "[Content_Types].xml":
                        data = re.sub(
                            rb'<Override[^>]*calcChain[^>]*/>', b'', data
                        )
                    elif entry == "xl/workbook.xml":
                        data = self._set_full_calc_on_load(data)

                    # Strip shared formulas from all worksheets
                    if entry.startswith("xl/worksheets/") and entry.endswith(".xml"):
                        data = self._strip_shared_formulas_bytes(data)

                    dst_zip.writestr(entry, data)

        with open(file_path, "wb") as f:
            f.write(output.getvalue())

        logger.info(
            f"Prepared for writing: "
            f"{len(cb_rows)} rows copied to Prior period (period: {old_period_name}), "
            f"{rows_cleared_count} Movement rows cleared"
        )
        return rows_cleared_count

    def _clear_movement_rows(self, mov_xml: bytes) -> Tuple[int, bytes]:
        """
        Clear Movement data rows from XML bytes (row 5+ removed, row 4 inputs cleared).

        Returns:
            Tuple of (rows_cleared_count, new_xml_bytes)
        """
        ns = self._NS
        self._register_all_ns()

        row4_pos = self._find_row_byte_positions(mov_xml, {4})

        first_ge5_start = None
        for m in re.finditer(rb'<row\s[^>]*r="(\d+)"', mov_xml):
            rn = int(m.group(1))
            if rn >= 5:
                first_ge5_start = m.start()
                break

        new_row4 = None
        if 4 in row4_pos:
            r4_start, r4_end = row4_pos[4]
            row4_el = self._parse_single_row(mov_xml[r4_start:r4_end])
            input_cols = {"A", "B", "C", "D", "E", "F", "G", "I"}
            for cell in list(row4_el):
                ref = cell.get("r", "")
                col_match = re.match(r"([A-Z]+)", ref)
                if col_match and col_match.group(1) in input_cols:
                    f_el = cell.find(f"{{{ns}}}f")
                    if f_el is None:
                        for child in list(cell):
                            cell.remove(child)
                        cell.attrib.pop("t", None)
            new_row4 = self._serialize_row(row4_el)

        if first_ge5_start is None and new_row4 is None:
            return (0, mov_xml)

        sd_close = mov_xml.rfind(b'</sheetData>')
        rows_cleared = 0
        if first_ge5_start is not None:
            section = mov_xml[first_ge5_start:sd_close]
            rows_cleared = len(re.findall(rb'<row\s', section))

        parts = []
        if new_row4 and 4 in row4_pos:
            r4_start, r4_end = row4_pos[4]
            parts.append(mov_xml[:r4_start])
            parts.append(new_row4)
            if first_ge5_start is not None:
                parts.append(b'\n')
                parts.append(mov_xml[sd_close:])
            else:
                parts.append(mov_xml[r4_end:])
        elif first_ge5_start is not None:
            parts.append(mov_xml[:first_ge5_start])
            parts.append(mov_xml[sd_close:])

        return (rows_cleared, b''.join(parts))

    def modify_cell_values(
        self,
        file_path: Path,
        sheet_name: str,
        modifications: Dict[int, Dict[str, str]],
    ) -> None:
        """
        Modify specific cell values in existing rows via XML byte manipulation.
        Preserves drawings, charts, and all other ZIP entries.

        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name (e.g. "Movement")
            modifications: {row_num: {col_letter: new_value_str}}
                e.g. {30: {"F": "4000000000"}, 45: {"F": "5000000000", "I": "Other receipts"}}
        """
        if not modifications:
            return

        with open(file_path, "rb") as f:
            zip_data = f.read()

        sheet_paths = self._get_sheet_xml_paths(zip_data)
        sheet_path = sheet_paths.get(sheet_name)
        if not sheet_path:
            raise ValueError(f"{sheet_name} sheet not found")

        with zipfile.ZipFile(io.BytesIO(zip_data), "r") as z:
            sheet_xml = z.read(sheet_path)

        self._register_all_ns()
        ns = self._NS

        target_rows = set(modifications.keys())
        row_positions = self._find_row_byte_positions(sheet_xml, target_rows)

        replacements = []
        for row_num in sorted(target_rows):
            if row_num not in row_positions:
                logger.warning(f"modify_cell_values: row {row_num} not found in XML")
                continue
            start, end = row_positions[row_num]
            row_el = self._parse_single_row(sheet_xml[start:end])
            col_mods = modifications[row_num]

            for cell in row_el:
                cell_ref = cell.get("r", "")
                col_letter = re.sub(r'\d+', '', cell_ref)
                if col_letter not in col_mods:
                    continue

                new_val = col_mods[col_letter]

                # Check if this is a numeric or string value
                try:
                    float(new_val)
                    is_numeric = True
                except (ValueError, TypeError):
                    is_numeric = False

                if is_numeric:
                    # Set numeric value: remove t= attribute, set <v>
                    if "t" in cell.attrib:
                        del cell.attrib["t"]
                    v_el = cell.find(f"{{{ns}}}v")
                    if v_el is None:
                        v_el = ET.SubElement(cell, f"{{{ns}}}v")
                    v_el.text = new_val
                    # Remove inline string if present
                    is_el = cell.find(f"{{{ns}}}is")
                    if is_el is not None:
                        cell.remove(is_el)
                else:
                    # Set inline string value
                    cell.set("t", "inlineStr")
                    # Remove existing <v>
                    v_el = cell.find(f"{{{ns}}}v")
                    if v_el is not None:
                        cell.remove(v_el)
                    # Set or create <is><t>value</t></is>
                    is_el = cell.find(f"{{{ns}}}is")
                    if is_el is None:
                        is_el = ET.SubElement(cell, f"{{{ns}}}is")
                    t_el = is_el.find(f"{{{ns}}}t")
                    if t_el is None:
                        t_el = ET.SubElement(is_el, f"{{{ns}}}t")
                    t_el.text = new_val

            new_bytes = self._serialize_row(row_el)
            replacements.append((start, end, new_bytes))

        # Apply replacements
        replacements.sort(key=lambda x: x[0])
        parts = []
        last_end = 0
        for s, e, new_bytes in replacements:
            parts.append(sheet_xml[last_end:s])
            parts.append(new_bytes)
            last_end = e
        parts.append(sheet_xml[last_end:])
        new_sheet_xml = b''.join(parts)

        self._write_sheet_xml(file_path, zip_data, sheet_path, new_sheet_xml)
        logger.info(f"Modified {len(modifications)} cell values in '{sheet_name}'")

    def insert_rows_after(
        self,
        file_path: Path,
        sheet_name: str,
        insertions: Dict[int, Dict[str, Any]],
    ) -> Dict[int, int]:
        """
        Insert new data rows after specified row numbers in the sheet.
        Renumbers all subsequent rows and adjusts formulas.

        Pattern: Same as remove_rows_by_source but with insertions instead of deletions.
        Rebuilds the data row region with new numbering, inserting new rows at the
        correct positions.

        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name (e.g. "Movement")
            insertions: {original_row_num: tx_data_dict}
                The new row is inserted AFTER the specified row.
                tx_data_dict keys: source, bank, account, date, description, debit, credit, nature

        Returns:
            {original_row_num: new_row_num} mapping for all original data rows.
            Callers use this to adjust any cached row indices.
        """
        if not insertions:
            return {}

        with open(file_path, "rb") as f:
            zip_data = f.read()

        sheet_paths = self._get_sheet_xml_paths(zip_data)
        sheet_path = sheet_paths.get(sheet_name)
        if not sheet_path:
            raise ValueError(f"{sheet_name} sheet not found")

        with zipfile.ZipFile(io.BytesIO(zip_data), "r") as z:
            sheet_xml = z.read(sheet_path)

        self._register_all_ns()
        ns = self._NS

        # Find data boundaries
        last_data_row = self._find_last_data_row_bytes(sheet_xml)
        if last_data_row < self.FIRST_DATA_ROW:
            return {}

        # Find byte positions of ALL data rows
        data_rows = set(range(self.FIRST_DATA_ROW, last_data_row + 1))
        row_positions = self._find_row_byte_positions(sheet_xml, data_rows)
        if not row_positions:
            return {}

        # Also find template rows that will be consumed by the expansion.
        # When we insert K rows, the data region grows by K, overlapping
        # template rows at positions (last_data_row+1) to (last_data_row+K).
        num_insertions = len(insertions)
        overlap_rows = set(range(last_data_row + 1, last_data_row + 1 + num_insertions))
        template_positions = self._find_row_byte_positions(sheet_xml, overlap_rows)

        # Collect shared formula masters
        shared_formulas = self._collect_shared_formula_masters(
            sheet_xml, ns, row_positions
        )

        # Extract row 4 templates for building new inserted rows
        formula_templates, style_templates = self._extract_row4_templates_bytes(sheet_xml)

        # Sort insertion points for ordered processing
        insertion_set = set(insertions.keys())

        # Rebuild rows in a single pass (top → bottom), maintaining offset
        mapping: Dict[int, int] = {}
        rebuilt_rows: list = []
        offset = 0  # cumulative shift from insertions above

        first_insert_row = min(insertion_set)
        sorted_data_rows = sorted(row_positions.keys())

        for original_row in sorted_data_rows:
            new_row_num = original_row + offset
            mapping[original_row] = new_row_num

            start, end = row_positions[original_row]
            row_bytes = sheet_xml[start:end]

            if offset == 0 and original_row < first_insert_row:
                # Before any insertion point — copy raw bytes without parsing
                rebuilt_rows.append(row_bytes)
            else:
                # Parse, renumber, convert shared formulas, serialize
                row_el = self._parse_single_row(row_bytes)

                if original_row != new_row_num:
                    row_el.set("r", str(new_row_num))
                    for cell in row_el:
                        ref = cell.get("r", "")
                        col_match = re.match(r"([A-Z]+)", ref)
                        if col_match:
                            cell.set("r", f"{col_match.group(1)}{new_row_num}")

                # Convert shared formulas to regular + adjust row references
                for cell in row_el:
                    f_el = cell.find(f"{{{ns}}}f")
                    if f_el is not None and f_el.get("t") == "shared":
                        si = f_el.get("si")
                        if f_el.text:
                            f_el.text = self._expand_shared_formula(
                                f_el.text, original_row, new_row_num
                            )
                        elif si and si in shared_formulas:
                            master_formula, master_row = shared_formulas[si]
                            f_el.text = self._expand_shared_formula(
                                master_formula, master_row, new_row_num
                            )
                        for attr in ("t", "si", "ref"):
                            if attr in f_el.attrib:
                                del f_el.attrib[attr]
                    elif f_el is not None and f_el.text and original_row != new_row_num:
                        f_el.text = self._expand_shared_formula(
                            f_el.text, original_row, new_row_num
                        )

                rebuilt_rows.append(self._serialize_row(row_el))

            # Insert new row after this one?
            if original_row in insertion_set:
                offset += 1
                insert_row_num = new_row_num + 1
                tx_dict = insertions[original_row]
                new_row_el = self._build_movement_row_xml(
                    ns, insert_row_num, tx_dict, formula_templates, style_templates
                )
                rebuilt_rows.append(self._serialize_row(new_row_el))

        # Replace data region in the XML.
        # Everything before the first data row + rebuilt rows + everything after
        # the last consumed row (data rows + overlapping template rows).
        sorted_positions = sorted(row_positions.values())
        first_start = sorted_positions[0][0]
        last_end = sorted_positions[-1][1]

        # Consume template rows that now overlap with the expanded data region
        if template_positions:
            for tp_row in sorted(template_positions.keys()):
                _, tp_end = template_positions[tp_row]
                if tp_end > last_end:
                    last_end = tp_end

        new_sheet_xml = b''.join([
            sheet_xml[:first_start],
            *rebuilt_rows,
            sheet_xml[last_end:],
        ])

        # Update dimension to reflect expanded range
        new_last_row = last_data_row + num_insertions
        new_sheet_xml = self._update_dimension(new_sheet_xml, new_last_row)

        self._write_sheet_xml(file_path, zip_data, sheet_path, new_sheet_xml)
        logger.info(
            f"Inserted {num_insertions} rows in '{sheet_name}', "
            f"data rows: {last_data_row} -> {new_last_row}"
        )
        return mapping

    def highlight_rows(
        self, file_path: Path, sheet_name: str, row_numbers: List[int],
        fill_xml: Optional[bytes] = None,
    ) -> None:
        """
        Apply highlight to specific rows via XML manipulation.
        Does NOT use openpyxl load/save (which corrupts drawings/charts).

        Args:
            fill_xml: Custom fill XML bytes. Defaults to orange (settlement).
                      Pass green fill for open-new transactions.

        Modifies only:
        - xl/styles.xml: adds a fill + cloned xf entries
        - Sheet XML: updates cell s= attributes on target rows
        All other ZIP entries (drawings, charts, etc.) are preserved byte-for-byte.
        """
        if not row_numbers:
            return

        with open(file_path, "rb") as f:
            zip_data = f.read()

        sheet_paths = self._get_sheet_xml_paths(zip_data)
        sheet_path = sheet_paths.get(sheet_name)
        if not sheet_path:
            raise ValueError(f"{sheet_name} sheet not found")

        with zipfile.ZipFile(io.BytesIO(zip_data), "r") as z:
            styles_xml = z.read("xl/styles.xml")
            sheet_xml = z.read(sheet_path)

        self._register_all_ns()

        # --- Step 1: Collect unique style indices from target rows ---
        row_set = set(row_numbers)
        row_positions = self._find_row_byte_positions(sheet_xml, row_set)

        unique_styles = set()
        for row_num in row_numbers:
            if row_num not in row_positions:
                continue
            start, end = row_positions[row_num]
            for m in re.finditer(rb'<c\s[^>]*?s="(\d+)"', sheet_xml[start:end]):
                unique_styles.add(m.group(1).decode())

        # --- Step 2: Modify styles.xml via byte manipulation ---
        styles_xml, style_map = self._add_highlight_styles(
            styles_xml, unique_styles, fill_xml
        )

        # --- Step 3: Collect shared formula masters from target rows ---
        ns = self._NS
        shared_formulas = self._collect_shared_formula_masters(
            sheet_xml, ns, row_positions
        )

        # --- Step 4: Update cell styles + convert shared formulas in target rows ---
        modifications = []
        for row_num in sorted(row_numbers):
            if row_num not in row_positions:
                continue
            start, end = row_positions[row_num]
            row_el = self._parse_single_row(sheet_xml[start:end])

            for cell in row_el:
                old_s = cell.get("s")
                new_s = style_map.get(old_s, style_map.get(None))
                if new_s:
                    cell.set("s", new_s)

                # Convert shared formulas to regular to prevent chain corruption
                f_el = cell.find(f"{{{ns}}}f")
                if f_el is not None and f_el.get("t") == "shared":
                    si = f_el.get("si")
                    if not f_el.text and si and si in shared_formulas:
                        master_formula, master_row = shared_formulas[si]
                        f_el.text = self._expand_shared_formula(
                            master_formula, master_row, row_num
                        )
                    for attr in ("t", "si", "ref"):
                        if attr in f_el.attrib:
                            del f_el.attrib[attr]

            new_bytes = self._serialize_row(row_el)
            modifications.append((start, end, new_bytes))

        # Apply modifications to sheet XML
        modifications.sort(key=lambda x: x[0])
        parts = []
        last_end = 0
        for start, end, new_bytes in modifications:
            parts.append(sheet_xml[last_end:start])
            parts.append(new_bytes)
            last_end = end
        parts.append(sheet_xml[last_end:])
        new_sheet_xml = b''.join(parts)

        # Strip ALL remaining shared formulas from the entire sheet
        new_sheet_xml = self._strip_shared_formulas_bytes(new_sheet_xml)

        # --- Step 5: Write back to ZIP (preserving drawings etc.) ---
        output = io.BytesIO()
        with zipfile.ZipFile(io.BytesIO(zip_data), "r") as src_zip:
            with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as dst_zip:
                for entry in src_zip.namelist():
                    # Skip calcChain to force Excel full recalculation
                    if "calcChain" in entry:
                        continue
                    if entry == "xl/styles.xml":
                        dst_zip.writestr(entry, styles_xml)
                    elif entry == sheet_path:
                        dst_zip.writestr(entry, new_sheet_xml)
                    elif entry == "[Content_Types].xml":
                        ct_xml = src_zip.read(entry)
                        ct_xml = re.sub(
                            rb'<Override[^>]*calcChain[^>]*/>', b'', ct_xml
                        )
                        dst_zip.writestr(entry, ct_xml)
                    elif entry == "xl/workbook.xml":
                        wb_xml = src_zip.read(entry)
                        wb_xml = OpenpyxlHandler._set_full_calc_on_load(wb_xml)
                        dst_zip.writestr(entry, wb_xml)
                    elif entry.startswith("xl/worksheets/") and entry.endswith(".xml"):
                        # Strip shared formulas from ALL worksheets
                        sheet_data = src_zip.read(entry)
                        sheet_data = self._strip_shared_formulas_bytes(sheet_data)
                        dst_zip.writestr(entry, sheet_data)
                    else:
                        dst_zip.writestr(entry, src_zip.read(entry))

        with open(file_path, "wb") as f:
            f.write(output.getvalue())

        logger.info(f"Highlighted {len(row_numbers)} rows in '{sheet_name}' via XML")

    @staticmethod
    def _add_highlight_styles(
        styles_xml: bytes, unique_styles: set,
        fill_xml: Optional[bytes] = None,
    ) -> tuple:
        """
        Add fill and cloned xf entries to styles.xml via byte manipulation.
        Returns (modified_styles_xml, style_map).

        Args:
            fill_xml: Custom fill XML. Defaults to orange (theme 5, tint 0.6).

        style_map: {old_s_str_or_None: new_s_str}
        """
        # --- Add fill ---
        if fill_xml is None:
            # Default: orange settlement highlight
            fill_xml = (
                b'<fill><patternFill patternType="solid">'
                b'<fgColor theme="5" tint="0.5999938962981048"/>'
                b'</patternFill></fill>'
            )

        fills_count_m = re.search(rb'<fills\s+count="(\d+)"', styles_xml)
        old_fill_count = int(fills_count_m.group(1))
        new_fill_id = old_fill_count

        # Insert fill before </fills>
        fills_close = styles_xml.find(b'</fills>')
        styles_xml = (
            styles_xml[:fills_close]
            + fill_xml
            + styles_xml[fills_close:]
        )
        # Update fills count attribute
        fills_count_m2 = re.search(rb'<fills\s+count="(\d+)"', styles_xml)
        new_fills_tag = f'<fills count="{old_fill_count + 1}"'.encode()
        styles_xml = (
            styles_xml[:fills_count_m2.start()]
            + new_fills_tag
            + styles_xml[fills_count_m2.end():]
        )

        # --- Extract existing xf entries from cellXfs ---
        cellXfs_m = re.search(rb'<cellXfs\s+count="(\d+)"', styles_xml)
        old_xf_count = int(cellXfs_m.group(1))
        cellXfs_start = cellXfs_m.start()
        cellXfs_close = styles_xml.find(b'</cellXfs>', cellXfs_start)
        cellXfs_section = styles_xml[cellXfs_start:cellXfs_close]

        # Extract individual <xf .../> or <xf ...>...</xf> entries
        xf_entries = []
        for m in re.finditer(rb'<xf\s', cellXfs_section):
            xf_start = m.start()
            tag_end = cellXfs_section.find(b'>', m.end())
            if cellXfs_section[tag_end - 1:tag_end] == b'/':
                xf_entries.append(cellXfs_section[xf_start:tag_end + 1])
            else:
                close = cellXfs_section.find(b'</xf>', tag_end)
                xf_entries.append(cellXfs_section[xf_start:close + 5])

        # Build style map and new xf bytes
        style_map = {}
        next_idx = old_xf_count
        new_xf_parts = []

        # Bare xf for cells without any style
        bare_xf = (
            f'<xf numFmtId="0" fontId="0" fillId="{new_fill_id}" '
            f'borderId="0" applyFill="1"/>'
        ).encode()
        new_xf_parts.append(bare_xf)
        style_map[None] = str(next_idx)
        next_idx += 1

        # Clone each unique existing style with the new fill
        for s in sorted(unique_styles):
            s_idx = int(s)
            if s_idx < len(xf_entries):
                cloned = xf_entries[s_idx]
                # Replace or add fillId
                if b'fillId=' in cloned:
                    cloned = re.sub(
                        rb'fillId="\d+"',
                        f'fillId="{new_fill_id}"'.encode(),
                        cloned,
                        count=1,
                    )
                else:
                    cloned = cloned.replace(
                        b'<xf ', f'<xf fillId="{new_fill_id}" '.encode(), 1
                    )
                # Add or update applyFill
                if b'applyFill=' in cloned:
                    cloned = re.sub(
                        rb'applyFill="[^"]*"', b'applyFill="1"', cloned, count=1
                    )
                else:
                    cloned = cloned.replace(
                        b'<xf ', b'<xf applyFill="1" ', 1
                    )
            else:
                cloned = (
                    f'<xf numFmtId="0" fontId="0" fillId="{new_fill_id}" '
                    f'borderId="0" applyFill="1"/>'
                ).encode()
            new_xf_parts.append(cloned)
            style_map[s] = str(next_idx)
            next_idx += 1

        # Insert new xf entries before </cellXfs>
        cellXfs_close_pos = styles_xml.find(b'</cellXfs>')
        insert_bytes = b''.join(new_xf_parts)
        styles_xml = (
            styles_xml[:cellXfs_close_pos]
            + insert_bytes
            + styles_xml[cellXfs_close_pos:]
        )

        # Update cellXfs count
        cellXfs_count_m = re.search(rb'<cellXfs\s+count="(\d+)"', styles_xml)
        new_xfs_tag = f'<cellXfs count="{next_idx}"'.encode()
        styles_xml = (
            styles_xml[:cellXfs_count_m.start()]
            + new_xfs_tag
            + styles_xml[cellXfs_count_m.end():]
        )

        return styles_xml, style_map

    @staticmethod
    def _add_quote_prefix_style(
        styles_xml: bytes, base_style: Optional[str] = None
    ) -> Tuple[bytes, str]:
        """
        Add a cell style with quotePrefix="1" (like typing ' before value in Excel).
        If base_style is provided, clones that style; otherwise creates a basic style.

        Returns (modified_styles_xml, new_style_index_str).
        """
        # --- Extract existing xf entries from cellXfs ---
        cellXfs_m = re.search(rb'<cellXfs\s+count="(\d+)"', styles_xml)
        if not cellXfs_m:
            return styles_xml, "0"

        old_xf_count = int(cellXfs_m.group(1))
        cellXfs_start = cellXfs_m.start()
        cellXfs_close = styles_xml.find(b'</cellXfs>', cellXfs_start)
        cellXfs_section = styles_xml[cellXfs_start:cellXfs_close]

        # Extract individual <xf .../> or <xf ...>...</xf> entries
        xf_entries = []
        for m in re.finditer(rb'<xf\s', cellXfs_section):
            xf_start = m.start()
            tag_end = cellXfs_section.find(b'>', m.end())
            if cellXfs_section[tag_end - 1:tag_end] == b'/':
                xf_entries.append(cellXfs_section[xf_start:tag_end + 1])
            else:
                close = cellXfs_section.find(b'</xf>', tag_end)
                xf_entries.append(cellXfs_section[xf_start:close + 5])

        # Build new xf with quotePrefix
        next_idx = old_xf_count

        if base_style is not None:
            base_idx = int(base_style)
            if base_idx < len(xf_entries):
                cloned = xf_entries[base_idx]
                # Add quotePrefix="1"
                if b'quotePrefix=' in cloned:
                    cloned = re.sub(
                        rb'quotePrefix="[^"]*"', b'quotePrefix="1"', cloned, count=1
                    )
                else:
                    cloned = cloned.replace(b'<xf ', b'<xf quotePrefix="1" ', 1)
                new_xf = cloned
            else:
                new_xf = b'<xf numFmtId="0" fontId="0" fillId="0" borderId="0" quotePrefix="1"/>'
        else:
            new_xf = b'<xf numFmtId="0" fontId="0" fillId="0" borderId="0" quotePrefix="1"/>'

        # Insert new xf before </cellXfs>
        cellXfs_close_pos = styles_xml.find(b'</cellXfs>')
        styles_xml = (
            styles_xml[:cellXfs_close_pos]
            + new_xf
            + styles_xml[cellXfs_close_pos:]
        )

        # Update cellXfs count
        cellXfs_count_m = re.search(rb'<cellXfs\s+count="(\d+)"', styles_xml)
        new_xfs_tag = f'<cellXfs count="{next_idx + 1}"'.encode()
        styles_xml = (
            styles_xml[:cellXfs_count_m.start()]
            + new_xfs_tag
            + styles_xml[cellXfs_count_m.end():]
        )

        return styles_xml, str(next_idx)

    @staticmethod
    def _load_shared_strings(zip_data: bytes) -> list:
        """Load shared strings table from ZIP data."""
        shared_strings = []
        try:
            with zipfile.ZipFile(io.BytesIO(zip_data), "r") as z:
                if "xl/sharedStrings.xml" in z.namelist():
                    ss_xml = z.read("xl/sharedStrings.xml")
                    for t_m in re.finditer(rb'<si[^>]*>.*?</si>', ss_xml, re.DOTALL):
                        parts = re.findall(rb'<t[^>]*>([^<]*)</t>', t_m.group(0))
                        shared_strings.append("".join(
                            p.decode("utf-8", errors="ignore") for p in parts
                        ))
        except Exception:
            pass
        return shared_strings

    @staticmethod
    def _get_cell_value_in_row(row_xml: bytes, col_letter: str, row_num: int,
                                shared_strings: list = None) -> str:
        """Extract cell value for a specific column from a row's XML."""
        cell_ref = f"{col_letter}{row_num}".encode()
        cell_pat = rb'<c\s[^>]*r="' + cell_ref + rb'"[^>]*>.*?</c>'
        cm = re.search(cell_pat, row_xml, re.DOTALL)
        if not cm:
            # Check self-closing cell (empty)
            cell_pat2 = rb'<c\s[^>]*r="' + cell_ref + rb'"[^/>]*\s*/>'
            cm2 = re.search(cell_pat2, row_xml)
            return "" if cm2 else ""
        cell_xml = cm.group(0)
        if b't="inlineStr"' in cell_xml:
            t_match = re.search(rb'<t[^>]*>([^<]*)</t>', cell_xml)
            return t_match.group(1).decode("utf-8", errors="ignore") if t_match else ""
        v_match = re.search(rb'<v>([^<]*)</v>', cell_xml)
        if v_match:
            raw = v_match.group(1).decode("utf-8", errors="ignore").strip()
            if b't="s"' in cell_xml and shared_strings:
                try:
                    return shared_strings[int(raw)]
                except (ValueError, IndexError):
                    return raw
            return raw
        return ""

    @staticmethod
    def _find_x_marker_row(sheet_xml: bytes, col_letter: str, zip_data: bytes = None) -> int:
        """Find the big X marker row in the specified column.

        The big X marker is on a DATA row (column B has a real account number,
        not 'x'). The boundary row (with small 'x' in all columns) is excluded.
        """
        shared_strings = []
        if zip_data:
            shared_strings = OpenpyxlHandler._load_shared_strings(zip_data)

        big_x_row = 0
        row_pattern = rb'<row[^>]*\sr="(\d+)"[^>]*>(.*?)</row>'

        for rm in re.finditer(row_pattern, sheet_xml, re.DOTALL):
            row_num = int(rm.group(1))
            row_content = rm.group(0)

            # Check if column J (or specified col) has "X"
            j_val = OpenpyxlHandler._get_cell_value_in_row(
                row_content, col_letter, row_num, shared_strings
            )
            if j_val.strip().upper() != "X":
                continue

            # Check column B — boundary rows have "x" in B, data rows have account numbers
            b_val = OpenpyxlHandler._get_cell_value_in_row(
                row_content, "B", row_num, shared_strings
            )
            if b_val.strip().lower() == "x":
                # This is the boundary row, skip
                continue

            # This is a data row with X in column J = big X marker
            if row_num > big_x_row:
                big_x_row = row_num

        return big_x_row

    @staticmethod
    def _find_boundary_row(sheet_xml: bytes, after_row: int, zip_data: bytes = None) -> int:
        """Find the boundary row (all 'x' columns) after a given row.

        The boundary row has 'x' in column B (and typically in all columns).
        Returns 0 if not found.
        """
        shared_strings = []
        if zip_data:
            shared_strings = OpenpyxlHandler._load_shared_strings(zip_data)

        row_pattern = rb'<row[^>]*\sr="(\d+)"[^>]*>(.*?)</row>'
        for rm in re.finditer(row_pattern, sheet_xml, re.DOTALL):
            row_num = int(rm.group(1))
            if row_num <= after_row:
                continue
            row_content = rm.group(0)
            b_val = OpenpyxlHandler._get_cell_value_in_row(
                row_content, "B", row_num, shared_strings
            )
            if b_val.strip().lower() == "x":
                return row_num
        return 0

    @staticmethod
    def _shift_rows_from(sheet_xml: bytes, from_row: int, shift: int = 1) -> bytes:
        """Shift all rows with r >= from_row down by shift amount.

        Updates both row r= attribute and all cell r= references within each row.
        Processes in descending order to avoid number collisions.
        """
        rows_to_shift = []
        for m in re.finditer(rb'<row[^>]*\sr="(\d+)"', sheet_xml):
            rn = int(m.group(1))
            if rn >= from_row:
                rows_to_shift.append(rn)

        # Process descending to avoid collisions
        for rn in sorted(rows_to_shift, reverse=True):
            new_rn = rn + shift
            rn_b = str(rn).encode()
            new_rn_b = str(new_rn).encode()

            # Find this row's full XML
            row_pat = rb'(<row[^>]*\sr=")' + rn_b + rb'("[^>]*>.*?</row>)'
            row_match = re.search(row_pat, sheet_xml, re.DOTALL)
            if not row_match:
                # Try self-closing row
                row_pat2 = rb'(<row[^>]*\sr=")' + rn_b + rb'("[^>]*/\s*>)'
                row_match = re.search(row_pat2, sheet_xml)
                if not row_match:
                    continue

            old_row_xml = row_match.group(0)
            new_row_xml = old_row_xml

            # Update row r= attribute
            new_row_xml = re.sub(
                rb'(\sr=")' + rn_b + rb'"',
                rb'\g<1>' + new_rn_b + rb'"',
                new_row_xml, count=1,
            )

            # Update all cell r= references: r="XX{rn}" → r="XX{new_rn}"
            new_row_xml = re.sub(
                rb'(<c\s[^>]*r="[A-Z]{1,3})' + rn_b + rb'"',
                rb'\g<1>' + new_rn_b + rb'"',
                new_row_xml,
            )

            # Update formula references within <f> elements:
            # e.g. $C1754 → $C1755, Acc_Char!$E1754 → Acc_Char!$E1755
            def _fix_formula_refs(f_match):
                f_content = f_match.group(2)
                # Replace COL{old_row} patterns (like $C1754, $E1754)
                # but NOT inside table names (Table1, Table4 etc.)
                f_content = re.sub(
                    rb'(\$?[A-Z]{1,3})' + rn_b + rb'(?=[^0-9]|$)',
                    rb'\g<1>' + new_rn_b,
                    f_content,
                )
                # Also update array ref attribute: ref="I1754" → ref="I1755"
                f_tag = f_match.group(1)
                f_tag = re.sub(
                    rb'(ref="[A-Z]{1,3})' + rn_b + rb'"',
                    rb'\g<1>' + new_rn_b + rb'"',
                    f_tag,
                )
                return f_tag + f_content + f_match.group(3)

            new_row_xml = re.sub(
                rb'(<f[^>]*>)(.*?)(</f>)',
                _fix_formula_refs,
                new_row_xml,
                flags=re.DOTALL,
            )

            sheet_xml = sheet_xml.replace(old_row_xml, new_row_xml, 1)

        return sheet_xml

    @staticmethod
    def _fill_template_row_cells(
        sheet_xml: bytes,
        row_num: int,
        cell_values: Dict[str, str],
    ) -> bytes:
        """Fill values into empty cells of an existing template row.

        For each column in cell_values, finds the empty cell element
        <c r="X{row}" s="NNN"/> and replaces it with an inline-string cell
        that preserves the original style.

        Args:
            sheet_xml: The sheet XML bytes
            row_num: The target row number
            cell_values: Dict mapping column letters to string values, e.g. {"B": "813015095347", "C": "TCS"}
        """
        row_pattern = rb'<row[^>]*\sr="' + str(row_num).encode() + rb'"[^>]*>.*?</row>'
        row_match = re.search(row_pattern, sheet_xml, re.DOTALL)
        if not row_match:
            return sheet_xml

        old_row = row_match.group(0)
        new_row = old_row

        for col, value in cell_values.items():
            col_ref = f"{col}{row_num}"
            # Match empty self-closing cell: <c r="B1754" s="1192"/>
            empty_pattern = (
                rb'<c\s[^>]*r="' + col_ref.encode() + rb'"[^/>]*\s*/>'
            )
            empty_match = re.search(empty_pattern, new_row)
            if empty_match:
                old_cell = empty_match.group(0)
                # Extract style attribute
                s_match = re.search(rb's="(\d+)"', old_cell)
                style_attr = f' s="{s_match.group(1).decode()}"' if s_match else ""
                # Build inline string cell preserving style
                escaped_val = (
                    value.replace("&", "&amp;")
                    .replace("<", "&lt;")
                    .replace(">", "&gt;")
                )
                new_cell = (
                    f'<c r="{col_ref}"{style_attr} t="inlineStr">'
                    f'<is><t>{escaped_val}</t></is></c>'
                ).encode("utf-8")
                new_row = new_row.replace(old_cell, new_cell, 1)

        return sheet_xml.replace(old_row, new_row, 1)

    @staticmethod
    def _find_table_for_sheet(zip_data: bytes, sheet_path: str, table_name: str = "Table2") -> Optional[str]:
        """Find the ZIP path of a named table associated with a sheet.

        Looks in the sheet's .rels file for table references, then checks
        each table's displayName.

        Returns:
            ZIP path to the table XML (e.g. "xl/tables/table7.xml"), or None
        """
        import posixpath
        sheet_dir = posixpath.dirname(sheet_path)   # xl/worksheets
        sheet_base = posixpath.basename(sheet_path)  # sheet17.xml
        rels_path = f"{sheet_dir}/_rels/{sheet_base}.rels"

        with zipfile.ZipFile(io.BytesIO(zip_data), "r") as z:
            if rels_path not in z.namelist():
                return None
            rels_xml = z.read(rels_path)
            # Find all table Target references
            for m in re.finditer(rb'Target="([^"]*table[^"]*\.xml)"', rels_xml, re.IGNORECASE):
                target = m.group(1).decode("utf-8", errors="ignore")
                # Resolve relative path: ../tables/table7.xml -> xl/tables/table7.xml
                table_path = posixpath.normpath(posixpath.join(sheet_dir, target))
                if table_path in z.namelist():
                    table_xml = z.read(table_path)
                    if f'displayName="{table_name}"'.encode() in table_xml:
                        return table_path
        return None

    @staticmethod
    def _expand_table_ref(table_xml: bytes, new_last_row: int) -> bytes:
        """Expand table ref and autoFilter ref to include up to new_last_row.

        Changes ref="B1:I{old}" -> ref="B1:I{new}" (preserving column letters).
        """
        row_bytes = str(new_last_row).encode()

        def replace_ref(match):
            prefix = match.group(1)   # bytes: e.g. b'ref="B1:'
            col = match.group(2)      # bytes: e.g. b'I'
            return prefix + col + row_bytes + b'"'

        # Update ref="..." on <table> element
        table_xml = re.sub(
            rb'(ref="[A-Z]+\d+:)([A-Z]+)\d+"',
            replace_ref,
            table_xml,
        )
        # Update autoFilter ref="..."
        table_xml = re.sub(
            rb'(<autoFilter[^>]*ref="[A-Z]+\d+:)([A-Z]+)\d+"',
            replace_ref,
            table_xml,
        )
        return table_xml

    @staticmethod
    def _build_acc_char_row_xml(row_num: int, account_no: str, code: str,
                                 branch: str, currency: str,
                                 account_type: str) -> bytes:
        """Build Acc_Char row XML with data cells AND formula cells.

        Data cells (inline string): B, C, E, F, G
        Formula cells: D (VLOOKUP entity), H (VLOOKUP name), I (XLOOKUP bank)
        Styles match the template rows in the original file.
        """
        rn = str(row_num)

        def _esc(v):
            return (v.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))

        cells = []
        # B: Account No. (s=1192)
        cells.append(
            f'<c r="B{rn}" s="1192" t="inlineStr"><is><t>{_esc(account_no)}</t></is></c>'
        )
        # C: CODE (s=995)
        cells.append(
            f'<c r="C{rn}" s="995" t="inlineStr"><is><t>{_esc(code)}</t></is></c>'
        )
        # D: ENTITY formula (s=991) — VLOOKUP from CODE
        cells.append(
            f'<c r="D{rn}" s="991" t="str">'
            f'<f>VLOOKUP($C{rn},Table1[#All],2,0)</f><v></v></c>'
        )
        # E: BRANCH (s=991)
        cells.append(
            f'<c r="E{rn}" s="991" t="inlineStr"><is><t>{_esc(branch)}</t></is></c>'
        )
        # F: CURRENCY (s=991)
        cells.append(
            f'<c r="F{rn}" s="991" t="inlineStr"><is><t>{_esc(currency)}</t></is></c>'
        )
        # G: ACCOUNT TYPE (s=991)
        cells.append(
            f'<c r="G{rn}" s="991" t="inlineStr"><is><t>{_esc(account_type)}</t></is></c>'
        )
        # H: NAME formula (s=991) — VLOOKUP from CODE
        cells.append(
            f'<c r="H{rn}" s="991" t="str">'
            f'<f>VLOOKUP($C{rn},Table1[#All],3,0)</f><v></v></c>'
        )
        # I: Bank formula (s=924) — XLOOKUP from BRANCH
        cells.append(
            f'<c r="I{rn}" s="924" t="str" cm="1">'
            f'<f t="array" ref="I{rn}">'
            f'_xlfn.XLOOKUP(Acc_Char!$E{rn},Table4[[#All],[Branch]],Table4[[#All],[Bank]],"recheck",0)'
            f'</f><v></v></c>'
        )

        row_xml = f'<row r="{rn}" spans="2:9">{"".join(cells)}</row>'
        return row_xml.encode("utf-8")

    def add_row_to_acc_char(
        self,
        file_path: Path,
        account_no: str,
        code: str,
        entity: str,
        branch: str,
        currency: str = "VND",
        account_type: str = "Saving Account",
        name: str = "",
        bank: str = "",
    ) -> int:
        """
        Add a new row to Acc_Char sheet after the Big X marker row.

        Strategy:
        1. Find Big X row (column J) — the last original data row.
        2. Find first template row after Big X (formula in D, empty B).
        3. If template found → OVERWRITE it with new data row (no shifting).
        4. If no template → shift boundary row down and insert before it.
        5. Expand Table2 only when used_row exceeds current table range.

        Returns:
            Row number where the account was added, or 0 if already exists.
        """
        with open(file_path, "rb") as f:
            zip_data = f.read()

        sheet_paths = self._get_sheet_xml_paths(zip_data)
        sheet_path = sheet_paths.get("Acc_Char")
        if not sheet_path:
            raise ValueError("Acc_Char sheet not found")

        with zipfile.ZipFile(io.BytesIO(zip_data), "r") as z:
            sheet_xml = z.read(sheet_path)

        # Check duplicate via XML scan (column B = Account No.)
        if self._xml_has_account(sheet_xml, account_no, col_letter="B", zip_data=zip_data):
            logger.debug(f"Account {account_no} already exists in Acc_Char")
            return 0

        # Find the Big X marker row (data row with X in J)
        x_row = self._find_x_marker_row(sheet_xml, "J", zip_data=zip_data)
        if not x_row:
            logger.warning("No X marker found in Acc_Char column J, falling back to end")
            x_row = self._find_last_xml_row_num(sheet_xml)

        # Find a source data row to clone
        source_row = self._find_last_data_row_before(sheet_xml, x_row + 1, data_col="B")
        if not source_row:
            source_row = 4

        # Select the nearest template row after Big X:
        # formula-empty rows first by position, with small-x rows as additional candidates.
        formula_templates = self._find_all_template_rows(sheet_xml, "D", "B", start_row=x_row + 1)
        small_x_rows = self._find_small_x_rows(sheet_xml, "B", start_row=x_row + 1, zip_data=zip_data)
        template_rows = sorted(set(formula_templates + small_x_rows))
        first_tpl = template_rows[0] if template_rows else 0

        extra_entries = {}

        def _esc(v):
            return v.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

        # Extract styles from source row
        src_pat = rb'<row[^>]*\sr="' + str(source_row).encode() + rb'"[^>]*>.*?</row>'
        src_match = re.search(src_pat, sheet_xml, re.DOTALL)
        src_xml = src_match.group(0) if src_match else b""

        def _get_s(col):
            m = re.search(rb'<c\s[^>]*r="' + col.encode() + rb'\d+"[^>]*s="(\d+)"', src_xml)
            return m.group(1).decode() if m else "999"

        def _build_overrides(rn_str):
            return {
                "B": f'<c r="B{rn_str}" s="{_get_s("B")}" t="inlineStr"><is><t>{_esc(account_no)}</t></is></c>'.encode(),
                "C": f'<c r="C{rn_str}" s="{_get_s("C")}" t="inlineStr"><is><t>{_esc(code)}</t></is></c>'.encode(),
                "E": f'<c r="E{rn_str}" s="{_get_s("E")}" t="inlineStr"><is><t>{_esc(branch)}</t></is></c>'.encode(),
                "F": f'<c r="F{rn_str}" s="{_get_s("F")}" t="inlineStr"><is><t>{_esc(currency)}</t></is></c>'.encode(),
                "G": f'<c r="G{rn_str}" s="{_get_s("G")}" t="inlineStr"><is><t>{_esc(account_type)}</t></is></c>'.encode(),
                # Clear marker column so cloned source rows do not propagate big "X".
                "J": f'<c r="J{rn_str}" s="{_get_s("J")}"/>'.encode(),
            }

        if first_tpl:
            insert_row_num = first_tpl
            data_overrides = _build_overrides(str(insert_row_num))
            row_xml = self._clone_row_for_insert(
                sheet_xml, source_row, insert_row_num, data_overrides
            )
            old_row_pat = rb'<row[^>]*\sr="' + str(first_tpl).encode() + rb'"[^>]*>.*?</row>'
            old_row_match = re.search(old_row_pat, sheet_xml, re.DOTALL)
            if old_row_match:
                sheet_xml = sheet_xml[:old_row_match.start()] + row_xml + sheet_xml[old_row_match.end():]
            used_row = insert_row_num
            logger.info(f"Cloned row {source_row} → Acc_Char row {insert_row_num} with account {account_no}")
        else:
            # No template available — insert before boundary (shift it down)
            boundary = self._find_boundary_row(sheet_xml, x_row, zip_data=zip_data)
            if boundary:
                insert_row_num = boundary
                sheet_xml = self._shift_rows_from(sheet_xml, boundary, shift=1)
                data_overrides = _build_overrides(str(insert_row_num))
                row_xml = self._clone_row_for_insert(
                    sheet_xml, source_row, insert_row_num, data_overrides
                )
                shifted_pat = rb'<row[^>]*\sr="' + str(boundary + 1).encode() + rb'"'
                bm = re.search(shifted_pat, sheet_xml)
                if bm:
                    sheet_xml = sheet_xml[:bm.start()] + row_xml + b'\n' + sheet_xml[bm.start():]
                else:
                    sheet_xml = self._insert_row_before_sheet_data_close(sheet_xml, row_xml)
                sheet_xml = self._update_dimension(sheet_xml, self._find_last_xml_row_num(sheet_xml))
            else:
                last_row = self._find_last_xml_row_num(sheet_xml)
                insert_row_num = last_row + 1
                row_xml = self._build_acc_char_row_xml(
                    insert_row_num, account_no, code, branch, currency, account_type
                )
                sheet_xml = self._insert_row_before_sheet_data_close(sheet_xml, row_xml)
                sheet_xml = self._update_dimension(sheet_xml, insert_row_num)
            used_row = insert_row_num
            logger.info(f"Inserted account {account_no} at Acc_Char row {insert_row_num} (shifted boundary)")

        # Expand Table2 only if used_row exceeds current table range
        table_path = self._find_table_for_sheet(zip_data, sheet_path, "Table2")
        if table_path:
            with zipfile.ZipFile(io.BytesIO(zip_data), "r") as z:
                table_xml = z.read(table_path)
            ref_m = re.search(rb'ref="[A-Z]+\d+:[A-Z]+(\d+)"', table_xml)
            if ref_m:
                current_last = int(ref_m.group(1))
                if used_row > current_last:
                    table_xml = self._expand_table_ref(table_xml, used_row)
                    extra_entries[table_path] = table_xml
                    logger.info(f"Expanded Table2 ref to row {used_row} (was {current_last})")

        self._write_sheet_xml(file_path, zip_data, sheet_path, sheet_xml,
                              extra_entries=extra_entries)
        return used_row

    def mark_acc_char_x_marker(self, file_path: Path, row_num: int) -> None:
        """Add a red 'X' marker in column J of the specified Acc_Char row.

        Uses style 1142 (Calibri 11pt red, center-aligned) for the red X.
        This marks the last row inserted in the current automation run.
        """
        RED_CENTER_STYLE = "1142"  # Calibri 11pt FF0000 red, center aligned

        with open(file_path, "rb") as f:
            zip_data = f.read()

        sheet_paths = self._get_sheet_xml_paths(zip_data)
        sheet_path = sheet_paths.get("Acc_Char")
        if not sheet_path:
            return

        with zipfile.ZipFile(io.BytesIO(zip_data), "r") as z:
            sheet_xml = z.read(sheet_path)

        row_pat = rb'<row[^>]*\sr="' + str(row_num).encode() + rb'"[^>]*>.*?</row>'
        row_match = re.search(row_pat, sheet_xml, re.DOTALL)
        if not row_match:
            logger.warning(f"Row {row_num} not found in Acc_Char for X marker")
            return

        old_row = row_match.group(0)
        cell_ref = f"J{row_num}"

        # Check if J cell already exists (self-closing empty or with value)
        j_cell_pat = rb'<c\s[^>]*r="' + cell_ref.encode() + rb'"[^>]*>.*?</c>'
        j_cell_empty = rb'<c\s[^>]*r="' + cell_ref.encode() + rb'"[^/>]*\s*/>'
        jm = re.search(j_cell_pat, old_row, re.DOTALL)
        jm_empty = re.search(j_cell_empty, old_row)

        new_j_cell = (
            f'<c r="{cell_ref}" s="{RED_CENTER_STYLE}" t="inlineStr">'
            f'<is><t>X</t></is></c>'
        ).encode("utf-8")

        if jm:
            new_row = old_row.replace(jm.group(0), new_j_cell, 1)
        elif jm_empty:
            new_row = old_row.replace(jm_empty.group(0), new_j_cell, 1)
        else:
            # No J cell exists — insert before </row>
            new_row = old_row.replace(b'</row>', new_j_cell + b'</row>', 1)

        sheet_xml = sheet_xml.replace(old_row, new_row, 1)
        self._write_sheet_xml(file_path, zip_data, sheet_path, sheet_xml)
        logger.info(f"Added red X marker at Acc_Char J{row_num}")

    @staticmethod
    def _clone_row_for_insert(
        sheet_xml: bytes,
        source_row_num: int,
        target_row_num: int,
        data_overrides: Dict[str, bytes],
    ) -> bytes:
        """Clone an existing row's XML, updating row number and replacing specific data cells.

        Preserves ALL formulas, styles, and columns from the source row.
        Formulas with row references (e.g. $C4) are updated to the new row number.

        Args:
            sheet_xml: Full sheet XML bytes
            source_row_num: Row number to clone from (e.g., last data row)
            target_row_num: Row number for the new row
            data_overrides: Dict of column letter -> cell XML bytes to replace.
                           e.g. {"C": b'<c r="C10" s="1142" t="inlineStr"><is><t>ACC123</t></is></c>'}

        Returns:
            New row XML bytes
        """
        # Extract source row XML
        src_pat = rb'<row[^>]*\sr="' + str(source_row_num).encode() + rb'"[^>]*>.*?</row>'
        src_match = re.search(src_pat, sheet_xml, re.DOTALL)
        if not src_match:
            raise ValueError(f"Source row {source_row_num} not found in sheet XML")

        row_xml = src_match.group(0)
        src = str(source_row_num)
        tgt = str(target_row_num)

        # Update row number in <row r="N" ...>
        row_xml = re.sub(
            rb'(<row[^>]*\sr=")' + src.encode() + rb'(")',
            rb'\g<1>' + tgt.encode() + rb'\2',
            row_xml,
        )

        # Update all cell references: r="XX{src}" -> r="XX{tgt}"
        row_xml = re.sub(
            rb'(r="[A-Z]+)' + src.encode() + rb'"',
            lambda m: m.group(1) + tgt.encode() + b'"',
            row_xml,
        )

        # Update array formula ref attributes: ref="XX{src}" -> ref="XX{tgt}"
        # (e.g. <f t="array" ref="I1754"> must become ref="I1755" for row 1755)
        row_xml = re.sub(
            rb'(<f[^>]*\sref="[A-Z]+)' + src.encode() + rb'(")',
            lambda m: m.group(1) + tgt.encode() + m.group(2),
            row_xml,
        )

        # Shift ALL relative formula row references by row delta.
        # This matches Excel copy behavior for formulas like "=A1753+1" and
        # avoids stale refs such as "=A1752+1" after row cloning.
        if target_row_num != source_row_num:
            def update_formula_block(f_match):
                """Update row references inside any <f ...>formula</f> block."""
                full = f_match.group(0)
                inner = re.search(rb'>([^<]*)</f>', full)
                if not inner:
                    return full
                formula_text = inner.group(1).decode("utf-8", errors="ignore")
                shifted = OpenpyxlHandler._expand_shared_formula(
                    formula_text, source_row_num, target_row_num
                )
                return (
                    full[:inner.start(1)]
                    + shifted.encode("utf-8")
                    + full[inner.end(1):]
                )

            row_xml = re.sub(rb'<f[^>]*>[^<]*</f>', update_formula_block, row_xml)

        # Replace data cells from overrides
        for col, new_cell_xml in data_overrides.items():
            col_bytes = col.encode()
            # Try to replace existing cell for this column
            cell_pat = rb'<c\s[^>]*r="' + col_bytes + tgt.encode() + rb'"[^>]*>.*?</c>'
            cell_self_pat = rb'<c\s[^>]*r="' + col_bytes + tgt.encode() + rb'"[^/>]*/\s*>'
            if re.search(cell_pat, row_xml, re.DOTALL):
                row_xml = re.sub(cell_pat, new_cell_xml, row_xml, count=1, flags=re.DOTALL)
            elif re.search(cell_self_pat, row_xml):
                row_xml = re.sub(cell_self_pat, new_cell_xml, row_xml, count=1)
            # If cell doesn't exist, insert before </row>
            else:
                row_xml = row_xml.replace(b'</row>', new_cell_xml + b'</row>', 1)

        # Clear cached values in formula cells (let Excel recalculate)
        # Remove <v>...</v> AND t="str|n|e|b|s" from cells that have <f>
        # Leaving t= without <v> causes "Removed Records: Cell information"
        def clear_formula_values(cell_match):
            cell = cell_match.group(0)
            if b'<f' in cell:
                cell = re.sub(rb'<v>.*?</v>', b'', cell)
                cell = re.sub(rb'\s+t="(?:str|n|b|e|s)"', b'', cell)
            return cell
        row_xml = re.sub(rb'<c\s[^>]*>.*?</c>', clear_formula_values, row_xml, flags=re.DOTALL)

        return row_xml

    @staticmethod
    def _find_last_data_row_before(sheet_xml: bytes, before_row: int, data_col: str = "C") -> int:
        """Find the last row with actual data in data_col before a given row number."""
        last_data = 0
        row_pattern = rb'<row[^>]*\sr="(\d+)"[^>]*>(.*?)</row>'
        for rm in re.finditer(row_pattern, sheet_xml, re.DOTALL):
            rn = int(rm.group(1))
            if rn >= before_row or rn < 4:
                continue
            c_ref = f"{data_col}{rn}".encode()
            c_cell = re.search(
                rb'<c\s[^>]*r="' + c_ref + rb'"[^>]*>.*?</c>',
                rm.group(0), re.DOTALL,
            )
            if c_cell and (b't="inlineStr"' in c_cell.group(0) or b'<v>' in c_cell.group(0)):
                last_data = rn
        return last_data

    @staticmethod
    def _build_saving_account_row_xml(
        row_num: int, account_no: str, currency: str,
        closing_balance: float = 0,
        term_months: Optional[int] = None,
        term_days: Optional[int] = None,
        opening_date: Optional[date] = None,
        maturity_date: Optional[date] = None,
        interest_rate: Optional[float] = None,
    ) -> bytes:
        """Build Saving Account row XML — LEGACY fallback.
        Used only when no source row is available for cloning.
        Prefer _clone_row_for_insert() with _build_saving_data_overrides().
        """
        rn = str(row_num)

        def _esc(v):
            return (v.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))

        cells = []
        cells.append(
            f'<c r="A{rn}" s="920" t="str">'
            f'<f>IFERROR(VLOOKUP($C{rn},Table2[#All],3,0),0)</f></c>'
        )
        cells.append(
            f'<c r="B{rn}" s="921" t="str">'
            f'<f>_xlfn.XLOOKUP(Table1218[[#This Row],[ACCOUNT NUMBER]],Table2[Account No.],Table2[BRANCH],"recheck",0)</f></c>'
        )
        cells.append(
            f'<c r="C{rn}" s="1142" t="inlineStr"><is><t>{_esc(account_no)}</t></is></c>'
        )
        cells.append(
            f'<c r="D{rn}" s="922" t="str">'
            f'<f>_xlfn.XLOOKUP(Table1218[[#This Row],[ACCOUNT NUMBER]],Table2[Account No.],Table2[ACCOUNT TYPE],"recheck",0)</f></c>'
        )
        cells.append(
            f'<c r="E{rn}" s="1130" t="inlineStr"><is><t>{_esc(currency)}</t></is></c>'
        )
        cells.append(
            f'<c r="F{rn}" s="926">'
            f'<f>SUMIFS(Table11[CLOSING BALANCE (VND)2],Table11[ACCOUNT NUMBER],'
            f'Table1218[[#This Row],[ACCOUNT NUMBER]],Table11[CURRENCY],'
            f'Table1218[[#This Row],[CURRENCY]])</f></c>'
        )
        cells.append(f'<c r="G{rn}" s="926"/>')
        if term_months:
            cells.append(f'<c r="H{rn}" s="1130" t="inlineStr"><is><t>{term_months} months</t></is></c>')
        else:
            cells.append(f'<c r="H{rn}" s="1130"/>')
        if term_days:
            cells.append(f'<c r="I{rn}" s="1130"><v>{term_days}</v></c>')
        else:
            cells.append(f'<c r="I{rn}" s="1130"/>')
        if opening_date and isinstance(opening_date, date):
            serial = OpenpyxlHandler._date_to_serial(opening_date)
            cells.append(f'<c r="J{rn}" s="1130"><v>{serial}</v></c>')
        else:
            cells.append(f'<c r="J{rn}" s="1130"/>')
        if maturity_date and isinstance(maturity_date, date):
            serial = OpenpyxlHandler._date_to_serial(maturity_date)
            cells.append(f'<c r="K{rn}" s="1131"><v>{serial}</v></c>')
        else:
            cells.append(f'<c r="K{rn}" s="1131"/>')
        if interest_rate:
            cells.append(f'<c r="L{rn}" s="927"><v>{interest_rate}</v></c>')
        else:
            cells.append(f'<c r="L{rn}" s="927"/>')

        row_xml = f'<row r="{rn}" spans="1:30">{"".join(cells)}</row>'
        return row_xml.encode("utf-8")

    @staticmethod
    def _build_saving_data_overrides(
        row_num: int, account_no: str, currency: str,
        closing_balance: float = 0,
        term_months: Optional[int] = None,
        term_days: Optional[int] = None,
        opening_date: Optional[date] = None,
        maturity_date: Optional[date] = None,
        interest_rate: Optional[float] = None,
        source_row_xml: bytes = b"",
    ) -> Dict[str, bytes]:
        """Build data cell overrides for Saving Account row cloning.
        Only overrides cells that contain input data (not formula cells).
        """
        rn = str(row_num)

        def _esc(v):
            return (v.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))

        def _get_style(col: str) -> str:
            """Extract style index from source row for a given column."""
            pat = rb'<c\s[^>]*r="' + col.encode() + rb'\d+"[^>]*'
            m = re.search(pat, source_row_xml)
            if m:
                s = re.search(rb's="(\d+)"', m.group(0))
                return s.group(1).decode() if s else ""
            return ""

        overrides = {}

        # C: Account number (data cell)
        s_c = _get_style("C") or "1142"
        overrides["C"] = f'<c r="C{rn}" s="{s_c}" t="inlineStr"><is><t>{_esc(account_no)}</t></is></c>'.encode()

        # H: Term months (data cell)
        s_h = _get_style("H") or "1130"
        if term_months:
            overrides["H"] = f'<c r="H{rn}" s="{s_h}" t="inlineStr"><is><t>{term_months} months</t></is></c>'.encode()
        else:
            overrides["H"] = f'<c r="H{rn}" s="{s_h}"/>'.encode()

        # I: Term days (data cell)
        s_i = _get_style("I") or "1130"
        if term_days:
            overrides["I"] = f'<c r="I{rn}" s="{s_i}"><v>{term_days}</v></c>'.encode()

        # J: Opening date (data cell)
        s_j = _get_style("J") or "1130"
        if opening_date and isinstance(opening_date, date):
            serial = OpenpyxlHandler._date_to_serial(opening_date)
            overrides["J"] = f'<c r="J{rn}" s="{s_j}"><v>{serial}</v></c>'.encode()

        # K: Maturity date (data cell)
        s_k = _get_style("K") or "1131"
        if maturity_date and isinstance(maturity_date, date):
            serial = OpenpyxlHandler._date_to_serial(maturity_date)
            overrides["K"] = f'<c r="K{rn}" s="{s_k}"><v>{serial}</v></c>'.encode()

        # L: Interest rate (data cell)
        s_l = _get_style("L") or "927"
        if interest_rate:
            overrides["L"] = f'<c r="L{rn}" s="{s_l}"><v>{interest_rate}</v></c>'.encode()

        # AD: Clear marker column so cloned source rows do not propagate big "X".
        s_ad = _get_style("AD") or "1130"
        overrides["AD"] = f'<c r="AD{rn}" s="{s_ad}"/>'.encode()

        return overrides

    @staticmethod
    def _build_cash_balance_row_xml(row_num: int, account_no: str) -> bytes:
        """Build Cash Balance row XML with account number + ALL formula cells.

        Only C (Account No.) is data. Everything else is formula-derived:
        A (ENTITY), B (BRANCH), D (ACCOUNT TYPE), E (CURRENCY) from Acc_Char via Table2,
        F (OB_BS), G (Dif), H (Opening Balance), I (Cash In), J (Cash Out),
        K (Net Amount), L (Closing Balance).
        Styles match existing data rows in the file.
        """
        rn = str(row_num)

        def _esc(v):
            return (v.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))

        cells = []
        # A: ENTITY formula (s=1134)
        cells.append(
            f'<c r="A{rn}" s="1134" t="str">'
            f"<f>VLOOKUP('Cash Balance'!$C{rn},Table2[#All],MATCH('Cash Balance'!A$3,Acc_Char!$B$1:$I$1,0),0)</f><v></v></c>"
        )
        # B: BRANCH formula (s=1135)
        cells.append(
            f'<c r="B{rn}" s="1135" t="str">'
            f"<f>VLOOKUP('Cash Balance'!$C{rn},Table2[#All],MATCH('Cash Balance'!B$3,Acc_Char!$B$1:$I$1,0),0)</f><v></v></c>"
        )
        # C: Account number (s=1306, inline string) — the KEY data input
        cells.append(
            f'<c r="C{rn}" s="1306" t="inlineStr"><is><t>{_esc(account_no)}</t></is></c>'
        )
        # D: Account type formula (s=1001)
        cells.append(
            f'<c r="D{rn}" s="1001" t="str">'
            f"<f>VLOOKUP('Cash Balance'!$C{rn},Table2[#All],MATCH('Cash Balance'!D$3,Acc_Char!$B$1:$I$1,0),0)</f><v></v></c>"
        )
        # E: Currency formula (s=1182)
        cells.append(
            f'<c r="E{rn}" s="1182" t="str">'
            f"<f>VLOOKUP('Cash Balance'!$C{rn},Table2[#All],MATCH('Cash Balance'!E$3,Acc_Char!$B$1:$I$1,0),0)</f><v></v></c>"
        )
        # F: OB_BS formula (s=1074)
        cells.append(
            f'<c r="F{rn}" s="1074" t="str">'
            f"<f>VLOOKUP(C{rn},'Cash balance (BS)'!A:C,3,)</f><v></v></c>"
        )
        # G: Dif = F - H (s=1172)
        cells.append(
            f'<c r="G{rn}" s="1172"><f>F{rn}-H{rn}</f><v>0</v></c>'
        )
        # H: Opening balance formula from Prior period (s=1304)
        cells.append(
            f'<c r="H{rn}" s="1304">'
            f"<f>SUMIFS('Cash balance (Prior period)'!F:F,"
            f"'Cash balance (Prior period)'!C:C,Table11[[#This Row],[ACCOUNT NUMBER]],"
            f"'Cash balance (Prior period)'!E:E,Table11[[#This Row],[CURRENCY]])</f><v>0</v></c>"
        )
        # I: Cash In formula from Movement (s=1305)
        cells.append(
            f'<c r="I{rn}" s="1305">'
            f'<f>SUMIFS(Movement!$F$4:$F$2999,Movement!$M$4:$M$2999,"VND",'
            f"Movement!$C$4:$C$2999,Table11[[#This Row],[ACCOUNT NUMBER]],"
            f"Movement!$J$4:$J$2999,Table11[[#This Row],[ENTITY]],"
            f"Movement!$O$4:$O$2999,'Cash Balance'!$A$2)</f><v>0</v></c>"
        )
        # J: Cash Out formula from Movement (s=1298)
        cells.append(
            f'<c r="J{rn}" s="1298">'
            f'<f>SUMIFS(Movement!$G$4:$G$2999,Movement!$M$4:$M$2999,"VND",'
            f"Movement!$C$4:$C$2999,Table11[[#This Row],[ACCOUNT NUMBER]],"
            f"Movement!$J$4:$J$2999,Table11[[#This Row],[ENTITY]],"
            f"Movement!$O$4:$O$2999,'Cash Balance'!$A$2)</f><v>0</v></c>"
        )
        # K: Net amount = Cash In - Cash Out (s=1235)
        cells.append(
            f'<c r="K{rn}" s="1235">'
            f'<f>Table11[[#This Row],[CASH IN - CREDIT (VND)]]-Table11[[#This Row],[CASH OUT - DEBIT (VND)]]</f><v>0</v></c>'
        )
        # L: Closing balance = Opening + Cash In - Cash Out (s=1115)
        cells.append(
            f'<c r="L{rn}" s="1115"><f>H{rn}+I{rn}-J{rn}</f><v>0</v></c>'
        )

        row_xml = f'<row r="{rn}" spans="1:12">{"".join(cells)}</row>'
        return row_xml.encode("utf-8")

    @staticmethod
    def _find_first_template_row(
        sheet_xml: bytes, formula_col: str = "A", data_col: str = "C",
        start_row: int = 4,
    ) -> int:
        """Find the first template row in a sheet (has formula in formula_col but empty data_col).

        Template rows are formula-only placeholder rows (no real account data in C).
        Data rows have actual account numbers in C (shared string or inline string).

        Args:
            start_row: Only scan rows >= this number (default 4, skip headers).

        Returns row number of first template, or 0 if not found.
        """
        row_pattern = rb'<row[^>]*\sr="(\d+)"[^>]*>(.*?)</row>'
        for rm in re.finditer(row_pattern, sheet_xml, re.DOTALL):
            rn = int(rm.group(1))
            if rn < start_row:
                continue
            row_content = rm.group(0)

            # Check if formula_col (A) has a <f> element
            a_ref = f"{formula_col}{rn}".encode()
            a_cell = re.search(
                rb'<c\s[^>]*r="' + a_ref + rb'"[^>]*>.*?</c>',
                row_content, re.DOTALL,
            )
            if not a_cell or b'<f' not in a_cell.group(0):
                continue

            # Check if data_col (C) is empty (self-closing or no cell)
            c_ref = f"{data_col}{rn}".encode()
            c_cell = re.search(
                rb'<c\s[^>]*r="' + c_ref + rb'"[^>]*>.*?</c>',
                row_content, re.DOTALL,
            )
            c_cell_empty = re.search(
                rb'<c\s[^>]*r="' + c_ref + rb'"[^/>]*/\s*>',
                row_content,
            )
            # First check self-closing (truly empty C cell)
            if c_cell_empty:
                # C is self-closing — no data, this is a template
                return rn
            if c_cell:
                # C has content — check if it has a real value
                # (inline string, shared string, or numeric <v>)
                cc = c_cell.group(0)
                if b't="inlineStr"' in cc or b'<v>' in cc:
                    # C has real data — this is a data row, skip
                    continue
            # C has no cell at all — this is a template
            return rn
        return 0

    def add_row_to_saving_account(
        self,
        file_path: Path,
        entity: str,
        branch: str,
        account_no: str,
        currency: str = "VND",
        closing_balance: float = 0,
        term_months: Optional[int] = None,
        term_days: Optional[int] = None,
        opening_date: Optional[date] = None,
        maturity_date: Optional[date] = None,
        interest_rate: Optional[float] = None,
    ) -> int:
        """
        Add a new row to Saving Account sheet after the Big X marker row.

        Strategy:
        1. Find Big X row (column AD).
        2. Find first template row after Big X (formula in A, empty C).
        3. If template found → OVERWRITE it with new data row.
        4. If no template → append after last data row in section.
        5. Expand Table1218 only when used_row exceeds current table range.

        Returns:
            Row number where account was added, or 0 if already exists.
        """
        with open(file_path, "rb") as f:
            zip_data = f.read()

        sheet_paths = self._get_sheet_xml_paths(zip_data)
        sheet_path = sheet_paths.get("Saving Account")
        if not sheet_path:
            raise ValueError("Saving Account sheet not found")

        with zipfile.ZipFile(io.BytesIO(zip_data), "r") as z:
            sheet_xml = z.read(sheet_path)

        # Check duplicate via XML scan (column C = ACCOUNT NUMBER)
        if self._xml_has_account(sheet_xml, account_no, col_letter="C", zip_data=zip_data):
            logger.debug(f"Account {account_no} already exists in Saving Account")
            return 0

        # Find Big X row (column AD)
        x_row = self._find_x_marker_row(sheet_xml, "AD", zip_data=zip_data)
        if not x_row:
            x_row = self._find_last_xml_row_num(sheet_xml)

        # Find a source data row to clone (last data row before or at Big X)
        source_row = self._find_last_data_row_before(sheet_xml, x_row + 1, data_col="C")
        if not source_row:
            source_row = 4  # fallback to row 4

        # Select the nearest template row after Big X:
        # formula-empty rows first by position, with small-x rows as additional candidates.
        formula_templates = self._find_all_template_rows(sheet_xml, "A", "C", start_row=x_row + 1)
        small_x_rows = self._find_small_x_rows(sheet_xml, "C", start_row=x_row + 1, zip_data=zip_data)
        template_rows = sorted(set(formula_templates + small_x_rows))
        first_tpl = template_rows[0] if template_rows else 0

        extra_entries = {}

        # Extract source row XML for style extraction
        src_pat = rb'<row[^>]*\sr="' + str(source_row).encode() + rb'"[^>]*>.*?</row>'
        src_match = re.search(src_pat, sheet_xml, re.DOTALL)
        source_row_xml = src_match.group(0) if src_match else b""

        if first_tpl:
            # OVERWRITE the template row — clone source row with new data
            insert_row_num = first_tpl
            data_overrides = self._build_saving_data_overrides(
                insert_row_num, account_no, currency,
                closing_balance=closing_balance,
                term_months=term_months, term_days=term_days,
                opening_date=opening_date, maturity_date=maturity_date,
                interest_rate=interest_rate,
                source_row_xml=source_row_xml,
            )
            row_xml = self._clone_row_for_insert(
                sheet_xml, source_row, insert_row_num, data_overrides
            )
            old_row_pat = rb'<row[^>]*\sr="' + str(first_tpl).encode() + rb'"[^>]*>.*?</row>'
            old_row_match = re.search(old_row_pat, sheet_xml, re.DOTALL)
            if old_row_match:
                sheet_xml = sheet_xml[:old_row_match.start()] + row_xml + sheet_xml[old_row_match.end():]
            used_row = insert_row_num
            logger.info(f"Cloned row {source_row} → Saving Account row {insert_row_num} with account {account_no}")
        else:
            # No template — find last data row after Big X, then insert after it
            last_data = x_row
            row_pattern = rb'<row[^>]*\sr="(\d+)"[^>]*>(.*?)</row>'
            for rm in re.finditer(row_pattern, sheet_xml, re.DOTALL):
                rn = int(rm.group(1))
                if rn <= x_row:
                    continue
                c_ref = f"C{rn}".encode()
                c_cell = re.search(
                    rb'<c\s[^>]*r="' + c_ref + rb'"[^>]*>.*?</c>',
                    rm.group(0), re.DOTALL,
                )
                if c_cell and (b't="inlineStr"' in c_cell.group(0) or b'<v>' in c_cell.group(0)):
                    last_data = rn
                else:
                    break
            insert_row_num = last_data + 1
            data_overrides = self._build_saving_data_overrides(
                insert_row_num, account_no, currency,
                closing_balance=closing_balance,
                term_months=term_months, term_days=term_days,
                opening_date=opening_date, maturity_date=maturity_date,
                interest_rate=interest_rate,
                source_row_xml=source_row_xml,
            )
            row_xml = self._clone_row_for_insert(
                sheet_xml, source_row, insert_row_num, data_overrides
            )
            # Find position in XML to insert (before first row >= insert_row_num)
            insert_pos = None
            for rm in re.finditer(rb'<row[^>]*\sr="(\d+)"', sheet_xml):
                if int(rm.group(1)) >= insert_row_num:
                    insert_pos = rm.start()
                    break
            if insert_pos:
                sheet_xml = sheet_xml[:insert_pos] + row_xml + b'\n' + sheet_xml[insert_pos:]
            else:
                sheet_xml = self._insert_row_before_sheet_data_close(sheet_xml, row_xml)
            sheet_xml = self._update_dimension(sheet_xml, self._find_last_xml_row_num(sheet_xml))
            used_row = insert_row_num
            logger.info(f"Inserted account {account_no} at Saving Account row {insert_row_num} (after last data)")

        # Expand Table1218 only if used_row exceeds current table range
        table_path = self._find_table_for_sheet(zip_data, sheet_path, "Table1218")
        if table_path:
            with zipfile.ZipFile(io.BytesIO(zip_data), "r") as z:
                table_xml = z.read(table_path)
            ref_m = re.search(rb'ref="[A-Z]+\d+:[A-Z]+(\d+)"', table_xml)
            if ref_m:
                current_last = int(ref_m.group(1))
                if used_row > current_last:
                    table_xml = self._expand_table_ref(table_xml, used_row)
                    extra_entries[table_path] = table_xml
                    logger.info(f"Expanded Table1218 ref to row {used_row} (was {current_last})")

        self._write_sheet_xml(file_path, zip_data, sheet_path, sheet_xml,
                              extra_entries=extra_entries)
        return used_row

    def add_row_to_cash_balance(
        self,
        file_path: Path,
        entity: str,
        branch: str,
        account_no: str,
        currency: str = "VND",
        opening_balance: float = 0,
    ) -> int:
        """
        Add a new row to Cash Balance sheet after the Big X marker row.

        Strategy:
        1. Find Big X row (column Z for saving section).
        2. Find first template row after Big X (formula in A, empty C).
        3. If template found → OVERWRITE it with new data row.
        4. If no template → shift boundary row down and insert before it.
        5. Expand Table11 only when used_row exceeds current table range.

        Returns:
            Row number where account was added, or 0 if already exists.
        """
        with open(file_path, "rb") as f:
            zip_data = f.read()

        sheet_paths = self._get_sheet_xml_paths(zip_data)
        sheet_path = sheet_paths.get("Cash Balance")
        if not sheet_path:
            raise ValueError("Cash Balance sheet not found")

        with zipfile.ZipFile(io.BytesIO(zip_data), "r") as z:
            sheet_xml = z.read(sheet_path)

        # Check duplicate via XML scan (column C = ACCOUNT NUMBER)
        if self._xml_has_account(sheet_xml, account_no, col_letter="C", zip_data=zip_data):
            logger.debug(f"Account {account_no} already exists in Cash Balance")
            return 0

        # Find Big X row (column Z for saving section)
        x_row = self._find_x_marker_row(sheet_xml, "Z", zip_data=zip_data)
        if not x_row:
            x_row = self._find_last_xml_row_num(sheet_xml)

        # Find a source data row to clone
        source_row = self._find_last_data_row_before(sheet_xml, x_row + 1, data_col="C")
        if not source_row:
            source_row = 4

        # Select the nearest template row after Big X:
        # formula-empty rows first by position, with small-x rows as additional candidates.
        formula_templates = self._find_all_template_rows(sheet_xml, "A", "C", start_row=x_row + 1)
        small_x_rows = self._find_small_x_rows(sheet_xml, "C", start_row=x_row + 1, zip_data=zip_data)
        template_rows = sorted(set(formula_templates + small_x_rows))
        first_tpl = template_rows[0] if template_rows else 0

        extra_entries = {}

        # Data override: only C (account number) is input data, all else are formulas
        def _esc(v):
            return v.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

        # Extract C style from source row
        src_pat = rb'<row[^>]*\sr="' + str(source_row).encode() + rb'"[^>]*>.*?</row>'
        src_match = re.search(src_pat, sheet_xml, re.DOTALL)
        if src_match:
            s_m = re.search(rb'<c\s[^>]*r="C\d+"[^>]*s="(\d+)"', src_match.group(0))
            s_c = s_m.group(1).decode() if s_m else "1306"
            s_z_m = re.search(rb'<c\s[^>]*r="Z\d+"[^>]*s="(\d+)"', src_match.group(0))
            s_z = s_z_m.group(1).decode() if s_z_m else "759"
        else:
            s_c = "1306"
            s_z = "759"

        rn = str(0)  # placeholder, will be set below

        if first_tpl:
            insert_row_num = first_tpl
            rn = str(insert_row_num)
            data_overrides = {
                "C": f'<c r="C{rn}" s="{s_c}" t="inlineStr"><is><t>{_esc(account_no)}</t></is></c>'.encode(),
                # Clear marker column so cloned source rows do not propagate big "X".
                "Z": f'<c r="Z{rn}" s="{s_z}"/>'.encode(),
            }
            row_xml = self._clone_row_for_insert(
                sheet_xml, source_row, insert_row_num, data_overrides
            )
            old_row_pat = rb'<row[^>]*\sr="' + str(first_tpl).encode() + rb'"[^>]*>.*?</row>'
            old_row_match = re.search(old_row_pat, sheet_xml, re.DOTALL)
            if old_row_match:
                sheet_xml = sheet_xml[:old_row_match.start()] + row_xml + sheet_xml[old_row_match.end():]
            used_row = insert_row_num
            logger.info(f"Cloned row {source_row} → Cash Balance row {insert_row_num} with account {account_no}")
        else:
            # No template — insert before boundary (shift it down)
            boundary = self._find_boundary_row(sheet_xml, x_row, zip_data=zip_data)
            if boundary:
                insert_row_num = boundary
                sheet_xml = self._shift_rows_from(sheet_xml, boundary, shift=1)
                rn = str(insert_row_num)
                data_overrides = {
                    "C": f'<c r="C{rn}" s="{s_c}" t="inlineStr"><is><t>{_esc(account_no)}</t></is></c>'.encode(),
                    "Z": f'<c r="Z{rn}" s="{s_z}"/>'.encode(),
                }
                row_xml = self._clone_row_for_insert(
                    sheet_xml, source_row, insert_row_num, data_overrides
                )
                shifted_pat = rb'<row[^>]*\sr="' + str(boundary + 1).encode() + rb'"'
                bm = re.search(shifted_pat, sheet_xml)
                if bm:
                    sheet_xml = sheet_xml[:bm.start()] + row_xml + b'\n' + sheet_xml[bm.start():]
                else:
                    sheet_xml = self._insert_row_before_sheet_data_close(sheet_xml, row_xml)
                sheet_xml = self._update_dimension(sheet_xml, self._find_last_xml_row_num(sheet_xml))
            else:
                last_row = self._find_last_xml_row_num(sheet_xml)
                insert_row_num = last_row + 1
                rn = str(insert_row_num)
                data_overrides = {
                    "C": f'<c r="C{rn}" s="{s_c}" t="inlineStr"><is><t>{_esc(account_no)}</t></is></c>'.encode(),
                    "Z": f'<c r="Z{rn}" s="{s_z}"/>'.encode(),
                }
                row_xml = self._clone_row_for_insert(
                    sheet_xml, source_row, insert_row_num, data_overrides
                )
                sheet_xml = self._insert_row_before_sheet_data_close(sheet_xml, row_xml)
                sheet_xml = self._update_dimension(sheet_xml, insert_row_num)
            used_row = insert_row_num
            logger.info(f"Cloned row {source_row} → Cash Balance row {insert_row_num} with account {account_no}")

        # Expand Table11 only if used_row exceeds current table range
        table_path = self._find_table_for_sheet(zip_data, sheet_path, "Table11")
        if table_path:
            with zipfile.ZipFile(io.BytesIO(zip_data), "r") as z:
                table_xml = z.read(table_path)
            ref_m = re.search(rb'ref="[A-Z]+\d+:[A-Z]+(\d+)"', table_xml)
            if ref_m:
                current_last = int(ref_m.group(1))
                if used_row > current_last:
                    table_xml = self._expand_table_ref(table_xml, used_row)
                    extra_entries[table_path] = table_xml
                    logger.info(f"Expanded Table11 ref to row {used_row} (was {current_last})")

        self._write_sheet_xml(file_path, zip_data, sheet_path, sheet_xml,
                              extra_entries=extra_entries)
        return used_row

    # ── Batch methods ──────────────────────────────────────────────────

    @staticmethod
    def _find_all_template_rows(
        sheet_xml: bytes, formula_col: str, data_col: str, start_row: int = 4,
    ) -> list:
        """Find ALL template rows (has formula in formula_col, empty data_col).

        Returns sorted list of row numbers.
        """
        results = []
        row_pattern = rb'<row[^>]*\sr="(\d+)"[^>]*>(.*?)</row>'
        for rm in re.finditer(row_pattern, sheet_xml, re.DOTALL):
            rn = int(rm.group(1))
            if rn < start_row:
                continue
            row_content = rm.group(0)

            # formula_col must have <f>
            a_ref = f"{formula_col}{rn}".encode()
            a_cell = re.search(
                rb'<c\s[^>]*r="' + a_ref + rb'"[^>]*>.*?</c>',
                row_content, re.DOTALL,
            )
            if not a_cell or b'<f' not in a_cell.group(0):
                continue

            # data_col must be empty
            c_ref = f"{data_col}{rn}".encode()
            c_cell_empty = re.search(
                rb'<c\s[^>]*r="' + c_ref + rb'"[^/>]*/\s*>',
                row_content,
            )
            if c_cell_empty:
                results.append(rn)
                continue
            c_cell = re.search(
                rb'<c\s[^>]*r="' + c_ref + rb'"[^>]*>.*?</c>',
                row_content, re.DOTALL,
            )
            if c_cell:
                cc = c_cell.group(0)
                if b't="inlineStr"' in cc or b'<v>' in cc:
                    continue  # has data, skip
            results.append(rn)
        return sorted(results)

    @staticmethod
    def _find_small_x_rows(
        sheet_xml: bytes,
        data_col: str,
        start_row: int = 4,
        zip_data: bytes = None,
    ) -> list:
        """
        Find rows where ``data_col`` contains lowercase/uppercase ``x`` marker.

        These rows are explicit template markers configured by finance users.
        They are treated as additional template candidates together with
        formula-empty rows.
        """
        shared_strings = []
        if zip_data:
            shared_strings = OpenpyxlHandler._load_shared_strings(zip_data)

        results = []
        row_pattern = rb'<row[^>]*\sr="(\d+)"[^>]*>(.*?)</row>'
        for rm in re.finditer(row_pattern, sheet_xml, re.DOTALL):
            rn = int(rm.group(1))
            if rn < start_row:
                continue
            row_content = rm.group(0)
            data_val = OpenpyxlHandler._get_cell_value_in_row(
                row_content, data_col, rn, shared_strings
            )
            if (data_val or "").strip().lower() == "x":
                results.append(rn)
        return sorted(results)

    def add_rows_to_sheets_batch(
        self,
        file_path: Path,
        entries: List[Dict[str, Any]],
    ) -> Dict[str, int]:
        """Batch-add rows to Acc_Char, Saving Account, Cash Balance in ONE ZIP pass.

        Each entry dict must have keys:
            saving_acc, entity, bank, code, currency,
            closing_balance, term_info, opening_date

        Returns dict: {acc_char_added, saving_added, cash_balance_added}
        """
        if not entries:
            return {"acc_char_added": 0, "saving_added": 0, "cash_balance_added": 0}

        with open(file_path, "rb") as f:
            zip_data = f.read()

        sheet_paths = self._get_sheet_xml_paths(zip_data)
        ac_path = sheet_paths.get("Acc_Char")
        sa_path = sheet_paths.get("Saving Account")
        cb_path = sheet_paths.get("Cash Balance")

        with zipfile.ZipFile(io.BytesIO(zip_data), "r") as z:
            ac_xml = z.read(ac_path) if ac_path else None
            sa_xml = z.read(sa_path) if sa_path else None
            cb_xml = z.read(cb_path) if cb_path else None

        def _esc(v):
            return v.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

        def _get_style(src_xml: bytes, col: str, fallback: str = "999") -> str:
            m = re.search(rb'<c\s[^>]*r="' + col.encode() + rb'\d+"[^>]*s="(\d+)"', src_xml)
            return m.group(1).decode() if m else fallback

        extra_entries: Dict[str, bytes] = {}
        ac_added = sa_added = cb_added = 0

        # ── Acc_Char ──
        if ac_xml and ac_path:
            x_row = self._find_x_marker_row(ac_xml, "J", zip_data=zip_data)
            if not x_row:
                x_row = self._find_last_xml_row_num(ac_xml)
            source_row = self._find_last_data_row_before(ac_xml, x_row + 1, "B")
            if not source_row:
                source_row = 4

            src_pat = rb'<row[^>]*\sr="' + str(source_row).encode() + rb'"[^>]*>.*?</row>'
            src_m = re.search(src_pat, ac_xml, re.DOTALL)
            src_xml = src_m.group(0) if src_m else b""

            small_x_rows = self._find_small_x_rows(ac_xml, "B", x_row + 1, zip_data=zip_data)
            formula_templates = self._find_all_template_rows(ac_xml, "D", "B", x_row + 1)
            templates = sorted(set(formula_templates + small_x_rows))
            tpl_idx = 0
            seen = set()
            max_row = 0
            current_last_row = self._find_last_xml_row_num(ac_xml)

            for entry in entries:
                acc = entry["saving_acc"]
                if acc in seen:
                    continue
                if self._xml_has_account(ac_xml, acc, "B", zip_data):
                    seen.add(acc)
                    continue
                seen.add(acc)

                is_tpl = tpl_idx < len(templates)
                if is_tpl:
                    row_num = templates[tpl_idx]
                    tpl_idx += 1
                else:
                    # No template rows left - insert after current last row
                    current_last_row += 1
                    row_num = current_last_row
                    logger.info(f"No template row left in Acc_Char; inserting {acc} at new row {row_num}")

                rn = str(row_num)
                overrides = {
                    "B": f'<c r="B{rn}" s="{_get_style(src_xml, "B")}" t="inlineStr"><is><t>{_esc(acc)}</t></is></c>'.encode(),
                    "C": f'<c r="C{rn}" s="{_get_style(src_xml, "C")}" t="inlineStr"><is><t>{_esc(entry.get("code", ""))}</t></is></c>'.encode(),
                    "E": f'<c r="E{rn}" s="{_get_style(src_xml, "E")}" t="inlineStr"><is><t>{_esc(entry.get("bank", ""))}</t></is></c>'.encode(),
                    "F": f'<c r="F{rn}" s="{_get_style(src_xml, "F")}" t="inlineStr"><is><t>{_esc(entry.get("currency", "VND"))}</t></is></c>'.encode(),
                    "G": f'<c r="G{rn}" s="{_get_style(src_xml, "G")}" t="inlineStr"><is><t>Saving Account</t></is></c>'.encode(),
                    "J": f'<c r="J{rn}" s="{_get_style(src_xml, "J", "1142")}"/>'.encode(),
                }
                row_xml = self._clone_row_for_insert(ac_xml, source_row, row_num, overrides)

                if is_tpl:
                    old_pat = rb'<row[^>]*\sr="' + rn.encode() + rb'"[^>]*>.*?</row>'
                    old_m = re.search(old_pat, ac_xml, re.DOTALL)
                    if old_m:
                        ac_xml = ac_xml[:old_m.start()] + row_xml + ac_xml[old_m.end():]
                else:
                    ins = None
                    for rm in re.finditer(rb'<row[^>]*\sr="(\d+)"', ac_xml):
                        if int(rm.group(1)) >= row_num:
                            ins = rm.start()
                            break
                    if ins:
                        ac_xml = ac_xml[:ins] + row_xml + b'\n' + ac_xml[ins:]
                    else:
                        ac_xml = self._insert_row_before_sheet_data_close(ac_xml, row_xml)
                    ac_xml = self._update_dimension(ac_xml, self._find_last_xml_row_num(ac_xml))

                max_row = max(max_row, row_num)
                current_last_row = max(current_last_row, row_num)
                ac_added += 1
                logger.info(f"Batch Acc_Char: added {acc} at row {row_num}")

            # Expand Table2 if needed
            if max_row:
                tp = self._find_table_for_sheet(zip_data, ac_path, "Table2")
                if tp:
                    with zipfile.ZipFile(io.BytesIO(zip_data), "r") as z:
                        txml = z.read(tp)
                    ref_m = re.search(rb'ref="[A-Z]+\d+:[A-Z]+(\d+)"', txml)
                    if ref_m and max_row > int(ref_m.group(1)):
                        extra_entries[tp] = self._expand_table_ref(txml, max_row)

        # ── Saving Account ──
        if sa_xml and sa_path:
            x_row = self._find_x_marker_row(sa_xml, "AD", zip_data=zip_data)
            if not x_row:
                x_row = self._find_last_xml_row_num(sa_xml)
            source_row = self._find_last_data_row_before(sa_xml, x_row + 1, "C")
            if not source_row:
                source_row = 4

            src_pat = rb'<row[^>]*\sr="' + str(source_row).encode() + rb'"[^>]*>.*?</row>'
            src_m = re.search(src_pat, sa_xml, re.DOTALL)
            src_xml = src_m.group(0) if src_m else b""

            small_x_rows = self._find_small_x_rows(sa_xml, "C", x_row + 1, zip_data=zip_data)
            formula_templates = self._find_all_template_rows(sa_xml, "A", "C", x_row + 1)
            templates = sorted(set(formula_templates + small_x_rows))
            tpl_idx = 0
            seen = set()
            max_row = 0
            current_last_row = self._find_last_xml_row_num(sa_xml)

            for entry in entries:
                acc = entry["saving_acc"]
                if acc in seen:
                    continue
                if self._xml_has_account(sa_xml, acc, "C", zip_data):
                    seen.add(acc)
                    continue
                seen.add(acc)

                is_tpl = tpl_idx < len(templates)
                if is_tpl:
                    row_num = templates[tpl_idx]
                    tpl_idx += 1
                else:
                    # No template rows left - insert after current last row
                    current_last_row += 1
                    row_num = current_last_row
                    logger.info(f"No template row left in Saving Account; inserting {acc} at new row {row_num}")

                rn = str(row_num)
                currency = entry.get("currency", "VND")
                term = entry.get("term_info", {}) or {}
                overrides = {
                    "C": f'<c r="C{rn}" s="{_get_style(src_xml, "C", "1142")}" t="inlineStr"><is><t>{_esc(acc)}</t></is></c>'.encode(),
                    "E": f'<c r="E{rn}" s="{_get_style(src_xml, "E", "1130")}" t="inlineStr"><is><t>{_esc(currency)}</t></is></c>'.encode(),
                }
                if term.get("term_months"):
                    overrides["H"] = f'<c r="H{rn}" s="{_get_style(src_xml, "H", "1130")}" t="inlineStr"><is><t>{term["term_months"]} months</t></is></c>'.encode()
                if term.get("term_days"):
                    overrides["I"] = f'<c r="I{rn}" s="{_get_style(src_xml, "I", "1130")}"><v>{term["term_days"]}</v></c>'.encode()
                op_date = entry.get("opening_date")
                if op_date and isinstance(op_date, date):
                    overrides["J"] = f'<c r="J{rn}" s="{_get_style(src_xml, "J", "1130")}"><v>{self._date_to_serial(op_date)}</v></c>'.encode()
                mat_date = term.get("maturity_date")
                if mat_date and isinstance(mat_date, date):
                    overrides["K"] = f'<c r="K{rn}" s="{_get_style(src_xml, "K", "1131")}"><v>{self._date_to_serial(mat_date)}</v></c>'.encode()
                if term.get("interest_rate"):
                    overrides["L"] = f'<c r="L{rn}" s="{_get_style(src_xml, "L", "927")}"><v>{term["interest_rate"]}</v></c>'.encode()
                overrides["AD"] = f'<c r="AD{rn}" s="{_get_style(src_xml, "AD", "1130")}"/>'.encode()

                row_xml = self._clone_row_for_insert(sa_xml, source_row, row_num, overrides)

                if is_tpl:
                    old_pat = rb'<row[^>]*\sr="' + rn.encode() + rb'"[^>]*>.*?</row>'
                    old_m = re.search(old_pat, sa_xml, re.DOTALL)
                    if old_m:
                        sa_xml = sa_xml[:old_m.start()] + row_xml + sa_xml[old_m.end():]
                else:
                    ins = None
                    for rm in re.finditer(rb'<row[^>]*\sr="(\d+)"', sa_xml):
                        if int(rm.group(1)) >= row_num:
                            ins = rm.start()
                            break
                    if ins:
                        sa_xml = sa_xml[:ins] + row_xml + b'\n' + sa_xml[ins:]
                    else:
                        sa_xml = self._insert_row_before_sheet_data_close(sa_xml, row_xml)
                    sa_xml = self._update_dimension(sa_xml, self._find_last_xml_row_num(sa_xml))

                max_row = max(max_row, row_num)
                current_last_row = max(current_last_row, row_num)
                sa_added += 1
                logger.info(f"Batch Saving: added {acc} at row {row_num}")

            # Expand Table1218 if needed
            if max_row:
                tp = self._find_table_for_sheet(zip_data, sa_path, "Table1218")
                if tp:
                    with zipfile.ZipFile(io.BytesIO(zip_data), "r") as z:
                        txml = z.read(tp)
                    ref_m = re.search(rb'ref="[A-Z]+\d+:[A-Z]+(\d+)"', txml)
                    if ref_m and max_row > int(ref_m.group(1)):
                        extra_entries[tp] = self._expand_table_ref(txml, max_row)

        # ── Cash Balance ──
        if cb_xml and cb_path:
            x_row = self._find_x_marker_row(cb_xml, "Z", zip_data=zip_data)
            if not x_row:
                x_row = self._find_last_xml_row_num(cb_xml)
            source_row = self._find_last_data_row_before(cb_xml, x_row + 1, "C")
            if not source_row:
                source_row = 4

            src_pat = rb'<row[^>]*\sr="' + str(source_row).encode() + rb'"[^>]*>.*?</row>'
            src_m = re.search(src_pat, cb_xml, re.DOTALL)
            src_xml = src_m.group(0) if src_m else b""

            small_x_rows = self._find_small_x_rows(cb_xml, "C", x_row + 1, zip_data=zip_data)
            formula_templates = self._find_all_template_rows(cb_xml, "A", "C", x_row + 1)
            templates = sorted(set(formula_templates + small_x_rows))
            tpl_idx = 0
            seen = set()
            max_row = 0
            current_last_row = self._find_last_xml_row_num(cb_xml)

            for entry in entries:
                acc = entry["saving_acc"]
                if acc in seen:
                    continue
                if self._xml_has_account(cb_xml, acc, "C", zip_data):
                    seen.add(acc)
                    continue
                seen.add(acc)

                is_tpl = tpl_idx < len(templates)
                if is_tpl:
                    row_num = templates[tpl_idx]
                    tpl_idx += 1
                else:
                    # No template rows left - insert after current last row
                    current_last_row += 1
                    row_num = current_last_row
                    logger.info(f"No template row left in Cash Balance; inserting {acc} at new row {row_num}")

                rn = str(row_num)
                s_c = _get_style(src_xml, "C", "1306") if src_xml else "1306"
                overrides = {
                    "C": f'<c r="C{rn}" s="{s_c}" t="inlineStr"><is><t>{_esc(acc)}</t></is></c>'.encode(),
                    "Z": f'<c r="Z{rn}" s="{_get_style(src_xml, "Z", "759")}"/>'.encode(),
                }
                row_xml = self._clone_row_for_insert(cb_xml, source_row, row_num, overrides)

                if is_tpl:
                    old_pat = rb'<row[^>]*\sr="' + rn.encode() + rb'"[^>]*>.*?</row>'
                    old_m = re.search(old_pat, cb_xml, re.DOTALL)
                    if old_m:
                        cb_xml = cb_xml[:old_m.start()] + row_xml + cb_xml[old_m.end():]
                else:
                    ins = None
                    for rm in re.finditer(rb'<row[^>]*\sr="(\d+)"', cb_xml):
                        if int(rm.group(1)) >= row_num:
                            ins = rm.start()
                            break
                    if ins:
                        cb_xml = cb_xml[:ins] + row_xml + b'\n' + cb_xml[ins:]
                    else:
                        cb_xml = self._insert_row_before_sheet_data_close(cb_xml, row_xml)
                    cb_xml = self._update_dimension(cb_xml, self._find_last_xml_row_num(cb_xml))

                max_row = max(max_row, row_num)
                current_last_row = max(current_last_row, row_num)
                cb_added += 1
                logger.info(f"Batch Cash Balance: added {acc} at row {row_num}")

            # Expand Table11 if needed
            if max_row:
                tp = self._find_table_for_sheet(zip_data, cb_path, "Table11")
                if tp:
                    with zipfile.ZipFile(io.BytesIO(zip_data), "r") as z:
                        txml = z.read(tp)
                    ref_m = re.search(rb'ref="[A-Z]+\d+:[A-Z]+(\d+)"', txml)
                    if ref_m and max_row > int(ref_m.group(1)):
                        extra_entries[tp] = self._expand_table_ref(txml, max_row)

        # ── Write all changes in ONE pass ──
        modified_sheets: Dict[str, bytes] = {}
        if ac_xml and ac_path:
            modified_sheets[ac_path] = self._dedupe_rows_by_number(ac_xml)
        if sa_xml and sa_path:
            modified_sheets[sa_path] = self._dedupe_rows_by_number(sa_xml)
        if cb_xml and cb_path:
            modified_sheets[cb_path] = self._dedupe_rows_by_number(cb_xml)

        if modified_sheets:
            self._write_multiple_sheets(file_path, zip_data, modified_sheets, extra_entries)

        logger.info(f"Batch complete: Acc_Char +{ac_added}, Saving +{sa_added}, Cash Balance +{cb_added}")
        return {
            "acc_char_added": ac_added,
            "saving_added": sa_added,
            "cash_balance_added": cb_added,
        }


# Singleton instance for reuse
_handler: Optional[OpenpyxlHandler] = None


def get_openpyxl_handler() -> OpenpyxlHandler:
    """Get singleton OpenpyxlHandler instance."""
    global _handler
    if _handler is None:
        _handler = OpenpyxlHandler()
    return _handler
