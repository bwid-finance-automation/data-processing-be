"""
Master Template Manager for Cash Report automation.
Handles copying, configuring, and managing the master Excel template.
"""
import os
import shutil
import uuid
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Optional, Dict, Any

import openpyxl
from openpyxl.utils import get_column_letter

from app.shared.utils.logging_config import get_logger
from .openpyxl_handler import OpenpyxlHandler, get_openpyxl_handler

logger = get_logger(__name__)

# Always use OpenpyxlHandler for cross-platform compatibility
# COM automation is no longer used - openpyxl works on all platforms

# Base paths
BASE_DIR = Path(__file__).resolve().parent.parent.parent.parent.parent
TEMPLATES_DIR = BASE_DIR / "templates" / "cash_report"
WORKING_DIR = BASE_DIR / "working" / "cash_report"

# Master template filename - use original file directly to preserve all formulas
MASTER_TEMPLATE_FILENAME = "master_data.xlsx"

# In-memory template cache (shared across all instances)
_template_cache: Optional[bytes] = None
_template_mtime: Optional[float] = None


def _load_template_to_cache() -> bytes:
    """Load master template into memory cache (lazy loading, auto-invalidates on file change)."""
    global _template_cache, _template_mtime
    template_path = TEMPLATES_DIR / MASTER_TEMPLATE_FILENAME
    if not template_path.exists():
        raise FileNotFoundError(f"Master template not found at {template_path}")

    current_mtime = template_path.stat().st_mtime
    if _template_cache is None or _template_mtime != current_mtime:
        with open(template_path, 'rb') as f:
            _template_cache = f.read()
        _template_mtime = current_mtime

        size_mb = len(_template_cache) / (1024 * 1024)
        logger.info(f"Loaded master template into memory cache ({size_mb:.2f} MB)")

    return _template_cache


def clear_template_cache() -> None:
    """Clear the template cache to force reload on next use."""
    global _template_cache, _template_mtime
    _template_cache = None
    _template_mtime = None
    logger.info("Template cache cleared")


class MasterTemplateManager:
    """
    Manages the master Excel template for cash report generation.

    Responsibilities:
    - Copy master template to working directory for each session
    - Update configuration (dates, FX rate, period name)
    - Clear Movement sheet data rows
    - Reset session to clean state
    """

    def __init__(self):
        self.templates_dir = TEMPLATES_DIR
        self.working_dir = WORKING_DIR
        self.master_template_path = self.templates_dir / MASTER_TEMPLATE_FILENAME

        # Ensure directories exist
        self.templates_dir.mkdir(parents=True, exist_ok=True)
        self.working_dir.mkdir(parents=True, exist_ok=True)

    def _get_session_dir(self, session_id: str) -> Path:
        """Get the working directory for a specific session."""
        return self.working_dir / session_id

    def _get_working_file_path(self, session_id: str) -> Path:
        """Get the path to the working Excel file for a session."""
        return self._get_session_dir(session_id) / "working_master.xlsx"

    def create_session(
        self,
        opening_date: date,
        ending_date: date,
        fx_rate: Decimal = Decimal("26175"),
        period_name: str = "",
    ) -> Dict[str, Any]:
        """
        Create a new session with a fresh copy of the master template.

        Args:
            opening_date: Report period start date
            ending_date: Report period end date
            fx_rate: VND/USD exchange rate
            period_name: Period name (e.g., "W3-4Jan26")

        Returns:
            Dict with session_id and file path
        """
        # Generate session ID
        session_id = str(uuid.uuid4())

        # Create session directory
        session_dir = self._get_session_dir(session_id)
        session_dir.mkdir(parents=True, exist_ok=True)

        # Write template from memory cache (fast)
        working_file = self._get_working_file_path(session_id)

        template_bytes = _load_template_to_cache()
        with open(working_file, 'wb') as f:
            f.write(template_bytes)

        logger.info(f"Created session {session_id}, wrote template from memory cache")

        # Initialize session in single optimized operation
        # (combines: step 0 prior period copy, update_config, clear_movement_data)
        handler = get_openpyxl_handler()
        handler.initialize_session_optimized(
            file_path=working_file,
            opening_date=opening_date,
            ending_date=ending_date,
            fx_rate=fx_rate,
            period_name=period_name,
        )

        return {
            "session_id": session_id,
            "working_file": str(working_file),
            "config": {
                "opening_date": opening_date.isoformat(),
                "ending_date": ending_date.isoformat(),
                "fx_rate": float(fx_rate),
                "period_name": period_name,
            }
        }

    def _update_config(
        self,
        session_id: str,
        opening_date: date,
        ending_date: date,
        fx_rate: Decimal,
        period_name: str,
    ) -> None:
        """
        Update the Summary sheet configuration.

        Summary sheet structure:
        - B1: Date (ending date)
        - B3: FX rate (VND/USD)
        - B4: Week/Period name
        - B5: Opening date
        - B6: Ending date
        """
        working_file = self._get_working_file_path(session_id)

        handler = get_openpyxl_handler()
        handler.update_config(
            file_path=working_file,
            opening_date=opening_date,
            ending_date=ending_date,
            fx_rate=fx_rate,
            period_name=period_name,
        )
        logger.info(f"Updated config for session {session_id}")

    def _clear_movement_data(self, session_id: str) -> None:
        """
        Clear Movement sheet data rows (row 4+), keeping:
        - Row 1: Legend
        - Row 2: Subtotals
        - Row 3: Headers
        - Row 4: Keep as formula template (but clear values)
        """
        working_file = self._get_working_file_path(session_id)

        handler = get_openpyxl_handler()
        rows_cleared = handler.clear_movement_data(working_file)
        logger.info(f"Cleared {rows_cleared} Movement rows for session {session_id}")

    def reset_session(self, session_id: str) -> Dict[str, Any]:
        """
        Reset a session to clean state (re-copy from master template).
        Preserves the session configuration.
        """
        working_file = self._get_working_file_path(session_id)

        if not working_file.exists():
            raise FileNotFoundError(f"Session {session_id} not found")

        # Read current config before reset
        wb = openpyxl.load_workbook(working_file, data_only=True)
        try:
            ws = wb["Summary"]
            opening_date = ws["B5"].value
            ending_date = ws["B6"].value
            fx_rate = ws["B3"].value
            period_name = ws["B4"].value
        finally:
            wb.close()

        # Convert to proper types
        if isinstance(opening_date, datetime):
            opening_date = opening_date.date()
        elif isinstance(opening_date, str):
            try:
                opening_date = date.fromisoformat(opening_date[:10])
            except:
                opening_date = date.today()
        elif opening_date is None:
            opening_date = date.today()

        if isinstance(ending_date, datetime):
            ending_date = ending_date.date()
        elif isinstance(ending_date, str):
            try:
                ending_date = date.fromisoformat(ending_date[:10])
            except:
                ending_date = date.today()
        elif ending_date is None:
            ending_date = date.today()

        if fx_rate is None:
            fx_rate = 26175
        if period_name is None:
            period_name = ""

        # Re-copy master template from memory cache
        template_bytes = _load_template_to_cache()
        with open(working_file, 'wb') as f:
            f.write(template_bytes)

        # Initialize session in single optimized operation
        # (combines: step 0 prior period copy, update_config, clear_movement_data)
        handler = get_openpyxl_handler()
        handler.initialize_session_optimized(
            file_path=working_file,
            opening_date=opening_date,
            ending_date=ending_date,
            fx_rate=Decimal(str(fx_rate)),
            period_name=period_name,
        )

        logger.info(f"Reset session {session_id}")

        return {
            "session_id": session_id,
            "status": "reset",
            "config": {
                "opening_date": opening_date.isoformat() if opening_date else None,
                "ending_date": ending_date.isoformat() if ending_date else None,
                "fx_rate": float(fx_rate),
                "period_name": period_name,
            }
        }

    def delete_session(self, session_id: str) -> bool:
        """Delete a session and its working files."""
        session_dir = self._get_session_dir(session_id)

        if session_dir.exists():
            shutil.rmtree(session_dir)
            logger.info(f"Deleted session {session_id}")
            return True

        return False

    def get_session_info(self, session_id: str, include_movement_count: bool = True) -> Optional[Dict[str, Any]]:
        """
        Get information about a session.

        Args:
            session_id: The session ID
            include_movement_count: If True, count Movement rows (slower). Set False for listing.
        """
        working_file = self._get_working_file_path(session_id)

        if not working_file.exists():
            return None

        # Use read_only mode for faster loading
        wb = openpyxl.load_workbook(working_file, data_only=True, read_only=True)
        try:
            # Get config from Summary
            summary_ws = wb["Summary"]
            # In read_only mode, we need to iterate to specific cells
            opening_date = None
            ending_date = None
            fx_rate = None
            period_name = None

            for row in summary_ws.iter_rows(min_row=1, max_row=6, min_col=2, max_col=2):
                for cell in row:
                    if cell.row == 3:
                        fx_rate = cell.value
                    elif cell.row == 4:
                        period_name = cell.value
                    elif cell.row == 5:
                        opening_date = cell.value
                    elif cell.row == 6:
                        ending_date = cell.value

            # Count Movement data rows only if requested
            movement_rows = 0
            if include_movement_count:
                movement_ws = wb["Movement"]
                for row in movement_ws.iter_rows(min_row=4, min_col=3, max_col=3):
                    for cell in row:
                        if cell.value:
                            movement_rows += 1

            # Helper to extract date properly (handle timezone offset from Excel)
            def extract_date(value) -> str:
                if value is None:
                    return None
                if isinstance(value, datetime):
                    # Excel stores dates as datetime. If time is 17:00:00 (UTC),
                    # it's actually midnight in Vietnam (UTC+7). Add offset to get correct date.
                    if value.hour == 17 and value.minute == 0 and value.second == 0:
                        from datetime import timedelta
                        corrected = value + timedelta(hours=7)
                        return corrected.date().isoformat()
                    return value.date().isoformat()
                if isinstance(value, date):
                    return value.isoformat()
                return str(value) if value else None

            return {
                "session_id": session_id,
                "working_file": str(working_file),
                "file_size_mb": round(working_file.stat().st_size / (1024 * 1024), 2),
                "movement_rows": movement_rows,
                "config": {
                    "opening_date": extract_date(opening_date),
                    "ending_date": extract_date(ending_date),
                    "fx_rate": float(fx_rate) if fx_rate else None,
                    "period_name": period_name,
                }
            }

        finally:
            wb.close()

    def get_working_file_path(self, session_id: str) -> Optional[Path]:
        """Get the working file path for a session (if exists)."""
        working_file = self._get_working_file_path(session_id)
        return working_file if working_file.exists() else None

    def list_sessions(self) -> list:
        """List all active sessions (fast - skips Movement row counting)."""
        sessions = []

        if not self.working_dir.exists():
            return sessions

        for session_dir in self.working_dir.iterdir():
            if session_dir.is_dir():
                session_id = session_dir.name
                # Skip Movement row counting for faster listing
                info = self.get_session_info(session_id, include_movement_count=False)
                if info:
                    sessions.append(info)

        return sessions
