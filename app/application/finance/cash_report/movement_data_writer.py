"""
Movement Data Writer for Cash Report automation.
Handles writing transaction data to the Movement sheet.
"""
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import List, Optional, Tuple

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import Cell

from app.shared.utils.logging_config import get_logger
from .openpyxl_handler import OpenpyxlHandler, get_openpyxl_handler

logger = get_logger(__name__)

# Always use OpenpyxlHandler for cross-platform compatibility
# COM automation is no longer used - openpyxl works on all platforms


@dataclass
class MovementTransaction:
    """Represents a transaction to write to Movement sheet."""
    source: str  # Column A - Source filename
    bank: str  # Column B - Bank name
    account: str  # Column C - Account number
    date: date  # Column D - Transaction date
    description: str  # Column E - Bank description
    debit: Optional[Decimal]  # Column F - Nợ (Debit)
    credit: Optional[Decimal]  # Column G - Có (Credit)
    nature: str  # Column I - Nature (classified)
    _classified_by: str = ""  # "rule" or "ai" - not written to Excel

    def __post_init__(self):
        # Ensure account is string
        if self.account:
            self.account = str(self.account)
        # Ensure debit/credit are Decimal or None
        if self.debit is not None and not isinstance(self.debit, Decimal):
            self.debit = Decimal(str(self.debit))
        if self.credit is not None and not isinstance(self.credit, Decimal):
            self.credit = Decimal(str(self.credit))


class MovementDataWriter:
    """
    Writes transaction data to the Movement sheet.

    Movement sheet structure:
    - Row 1: Legend
    - Row 2: Subtotals (formulas)
    - Row 3: Headers
    - Row 4+: Data rows

    Columns:
    - A: Source (INPUT)
    - B: Bank (INPUT)
    - C: Account (INPUT)
    - D: Date (INPUT)
    - E: Bank description (INPUT)
    - F: Nợ/Debit (INPUT)
    - G: Có/Credit (INPUT)
    - H: Net (FORMULA: =F-G)
    - I: Nature (INPUT - AI classified)
    - J: Entity (FORMULA: VLOOKUP)
    - K: Grouping (FORMULA: VLOOKUP)
    - L: Key payment (FORMULA: =I)
    - M: Currency (FORMULA: VLOOKUP)
    - N: Account type (FORMULA: VLOOKUP)
    - O: Period (FORMULA: =Summary!$B$4)
    - P: Text (FORMULA: =TEXT(C,"0"))
    """

    # Column mapping
    INPUT_COLUMNS = {
        'A': 'source',
        'B': 'bank',
        'C': 'account',
        'D': 'date',
        'E': 'description',
        'F': 'debit',
        'G': 'credit',
        'I': 'nature',
    }

    # Formula columns (will be copied from row 4 template)
    FORMULA_COLUMNS = ['H', 'J', 'K', 'L', 'M', 'N', 'O', 'P']

    # First data row
    FIRST_DATA_ROW = 4

    # Header row
    HEADER_ROW = 3

    def __init__(self, working_file_path: Path):
        """
        Initialize with path to working Excel file.

        Args:
            working_file_path: Path to the working Excel file
        """
        self.working_file_path = working_file_path

        if not self.working_file_path.exists():
            raise FileNotFoundError(f"Working file not found: {working_file_path}")

    def get_last_data_row(self) -> int:
        """
        Find the last row with data in the Movement sheet.

        Returns:
            The last row number with data, or FIRST_DATA_ROW-1 if empty
        """
        wb = openpyxl.load_workbook(self.working_file_path, data_only=True)
        try:
            ws = wb["Movement"]

            last_row = self.FIRST_DATA_ROW - 1  # Start from row 3 (before first data)

            # Scan column C (Account) to find last row with data
            for row in range(self.FIRST_DATA_ROW, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=3).value  # Column C
                if cell_value is not None and str(cell_value).strip():
                    last_row = row

            return last_row

        finally:
            wb.close()

    def get_row_count(self) -> int:
        """Get the number of data rows in Movement sheet."""
        last_row = self.get_last_data_row()
        if last_row < self.FIRST_DATA_ROW:
            return 0
        return last_row - self.FIRST_DATA_ROW + 1

    def _get_formula_templates(self, wb: openpyxl.Workbook) -> dict:
        """
        Extract formula templates from row 4.

        Returns:
            Dict mapping column letter to formula template string
        """
        ws = wb["Movement"]
        templates = {}

        for col in self.FORMULA_COLUMNS:
            cell = ws[f"{col}{self.FIRST_DATA_ROW}"]
            if cell.value and str(cell.value).startswith('='):
                templates[col] = str(cell.value)

        return templates

    def _adjust_formula_for_row(self, formula: str, source_row: int, target_row: int) -> str:
        """
        Adjust a formula's row references from source_row to target_row.

        Args:
            formula: The formula string
            source_row: The row the formula was copied from
            target_row: The row to adjust to

        Returns:
            Adjusted formula string
        """
        if not formula or not formula.startswith('='):
            return formula

        # Pattern to match cell references like A4, $A4, A$4, $A$4
        # We want to adjust only non-absolute row references
        pattern = r'(\$?[A-Z]+)(\$?)(\d+)'

        def replace_ref(match):
            col = match.group(1)
            dollar = match.group(2)  # $ before row number (if absolute)
            row_num = int(match.group(3))

            # If row is absolute ($), don't adjust
            if dollar == '$':
                return f"{col}${row_num}"

            # Only adjust if the row matches source_row
            if row_num == source_row:
                return f"{col}{target_row}"

            return match.group(0)

        return re.sub(pattern, replace_ref, formula)

    def append_transactions(self, transactions: List[MovementTransaction]) -> Tuple[int, int]:
        """
        Append transactions to the Movement sheet.

        Args:
            transactions: List of transactions to append

        Returns:
            Tuple of (rows_added, total_rows)
        """
        if not transactions:
            return 0, self.get_row_count()

        # Use OpenpyxlHandler for cross-platform compatibility
        handler = get_openpyxl_handler()

        # Convert MovementTransaction to dict for handler
        tx_dicts = [
            {
                'source': tx.source,
                'bank': tx.bank,
                'account': tx.account,
                'date': tx.date,
                'description': tx.description,
                'debit': tx.debit if tx.debit is not None else 0,
                'credit': tx.credit if tx.credit is not None else 0,
                'nature': tx.nature,
            }
            for tx in transactions
        ]

        rows_added, total_rows = handler.append_transactions(
            self.working_file_path, tx_dicts
        )
        logger.info(f"Appended {rows_added} transactions, total: {total_rows}")
        return rows_added, total_rows

    def modify_cell_values(self, modifications: dict) -> None:
        """
        Modify specific cell values in existing Movement rows.
        Uses byte-level XML manipulation to preserve drawings/charts.

        Args:
            modifications: {row_num: {col_letter: new_value_str}}
                e.g. {30: {"F": "4000000000"}}
        """
        if not modifications:
            return
        handler = get_openpyxl_handler()
        handler.modify_cell_values(self.working_file_path, "Movement", modifications)

    def insert_rows_after(self, insertions: dict) -> dict:
        """
        Insert new rows after specified row numbers in Movement sheet.
        Used to place interest split rows directly below their settlement row.

        Args:
            insertions: {original_row_num: tx_data_dict}

        Returns:
            {original_row_num: new_row_num} mapping for all data rows
        """
        if not insertions:
            return {}
        handler = get_openpyxl_handler()
        return handler.insert_rows_after(self.working_file_path, "Movement", insertions)

    def highlight_settlement_rows(self, row_numbers: List[int]) -> None:
        """
        Apply settlement highlight (orange/theme 5 tint 0.6) to specific rows in Movement sheet.
        Uses byte-level XML manipulation to avoid corrupting drawings/charts.

        Args:
            row_numbers: List of 1-based row numbers to highlight
        """
        if not row_numbers:
            return

        handler = get_openpyxl_handler()
        handler.highlight_rows(self.working_file_path, "Movement", row_numbers)

    def highlight_open_new_rows(self, row_numbers: List[int]) -> None:
        """
        Apply green highlight to open-new (mở mới) rows in Movement sheet.
        Uses byte-level XML manipulation to avoid corrupting drawings/charts.

        Args:
            row_numbers: List of 1-based row numbers to highlight
        """
        if not row_numbers:
            return

        green_fill = (
            b'<fill><patternFill patternType="solid">'
            b'<fgColor rgb="FF92D050"/>'
            b'</patternFill></fill>'
        )

        handler = get_openpyxl_handler()
        handler.highlight_rows(self.working_file_path, "Movement", row_numbers, fill_xml=green_fill)

    def remove_transactions_by_source(self, source_name: str) -> int:
        """
        Remove all transactions with matching source name (column A).
        Used to implement "overwrite on duplicate file upload".

        Args:
            source_name: Source name to match

        Returns:
            Number of rows removed
        """
        handler = get_openpyxl_handler()
        rows_removed = handler.remove_rows_by_source(self.working_file_path, source_name)
        return rows_removed

    def clear_all_data(self) -> int:
        """
        Clear all data rows from Movement sheet (row 5+, keep row 4 as template).

        Returns:
            Number of rows cleared
        """
        handler = get_openpyxl_handler()
        rows_cleared = handler.clear_movement_data(self.working_file_path)
        logger.info(f"Cleared {rows_cleared} rows")
        return rows_cleared

    def get_all_transactions(self) -> List[MovementTransaction]:
        """
        Read all transactions from Movement sheet.

        Returns:
            List of MovementTransaction objects
        """
        from datetime import datetime

        wb = openpyxl.load_workbook(self.working_file_path, data_only=True, read_only=True)

        try:
            ws = wb["Movement"]
            transactions = []

            for row_idx, row_data in enumerate(ws.iter_rows(min_row=self.FIRST_DATA_ROW, max_col=9), start=self.FIRST_DATA_ROW):
                account = row_data[2].value if len(row_data) > 2 else None
                if not account:
                    continue

                # Parse date
                date_val = row_data[3].value if len(row_data) > 3 else None
                if isinstance(date_val, datetime):
                    date_val = date_val.date()
                elif isinstance(date_val, str):
                    try:
                        date_val = datetime.strptime(date_val[:10], "%Y-%m-%d").date()
                    except:
                        date_val = None

                # Parse amounts
                debit = row_data[5].value if len(row_data) > 5 else None
                credit = row_data[6].value if len(row_data) > 6 else None

                if debit is not None:
                    debit = Decimal(str(debit)) if debit else None
                if credit is not None:
                    credit = Decimal(str(credit)) if credit else None

                transactions.append(MovementTransaction(
                    source=row_data[0].value if len(row_data) > 0 else "",
                    bank=row_data[1].value if len(row_data) > 1 else "",
                    account=str(account),
                    date=date_val,
                    description=str(row_data[4].value) if len(row_data) > 4 and row_data[4].value else "",
                    debit=debit,
                    credit=credit,
                    nature=row_data[8].value if len(row_data) > 8 else "",
                ))

            return transactions

        finally:
            wb.close()

    def get_data_preview(self, limit: int = 10) -> List[dict]:
        """
        Get a preview of the Movement data (fast - read_only mode).

        Args:
            limit: Maximum number of rows to return

        Returns:
            List of dicts with row data
        """
        # Use read_only mode for much faster loading
        wb = openpyxl.load_workbook(self.working_file_path, data_only=True, read_only=True)

        try:
            ws = wb["Movement"]
            rows = []
            row_count = 0

            # In read_only mode, iterate rows instead of direct cell access
            for row_idx, row_data in enumerate(ws.iter_rows(min_row=self.FIRST_DATA_ROW, max_col=9), start=self.FIRST_DATA_ROW):
                if row_count >= limit:
                    break

                # row_data is tuple of cells
                account = row_data[2].value if len(row_data) > 2 else None  # Column C (index 2)
                if not account:
                    continue

                rows.append({
                    "row": row_idx,
                    "source": row_data[0].value if len(row_data) > 0 else None,  # Column A
                    "bank": row_data[1].value if len(row_data) > 1 else None,  # Column B
                    "account": account,
                    "date": row_data[3].value if len(row_data) > 3 else None,  # Column D
                    "description": str(row_data[4].value)[:100] if len(row_data) > 4 and row_data[4].value else None,  # Column E
                    "debit": row_data[5].value if len(row_data) > 5 else None,  # Column F
                    "credit": row_data[6].value if len(row_data) > 6 else None,  # Column G
                    "nature": row_data[8].value if len(row_data) > 8 else None,  # Column I
                })
                row_count += 1

            return rows

        finally:
            wb.close()
