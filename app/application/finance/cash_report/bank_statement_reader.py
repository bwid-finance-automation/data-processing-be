"""
Bank Statement Reader for Cash Report automation.
Reads parsed Excel files from the bank statement parser.
"""
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import List, Optional, Tuple, Union
import math
import re

import openpyxl
import pandas as pd

from app.shared.utils.logging_config import get_logger
from .movement_data_writer import MovementTransaction

logger = get_logger(__name__)


class BankStatementReader:
    """
    Reads parsed bank statement Excel files and converts to MovementTransaction objects.

    Expected input: Excel file with "Template details" sheet from bank statement parser.

    Sheet columns (0-indexed):
    0: External ID
    1: External ID BSM Daily
    2: Bank Statement Daily
    3: Name (*)
    4: Bank Code (*)
    5: Bank Account Number (*)
    6: TRANS ID
    7: Trans Date (*)
    8: Description (*)
    9: Currency(*)
    10: DEBIT (*)
    11: CREDIT (*)
    12: Amount
    13: Type
    14: Balance
    15: PARTNER
    16: PARTNER ACCOUNT
    17: PARTNER BANK ID
    """

    # Column indices in Template details sheet
    COL_BANK_CODE = 4  # Bank Code (*)
    COL_ACCOUNT_NUMBER = 5  # Bank Account Number (*)
    COL_TRANS_DATE = 7  # Trans Date (*)
    COL_DESCRIPTION = 8  # Description (*)
    COL_CURRENCY = 9  # Currency(*)
    COL_DEBIT = 10  # DEBIT (*)
    COL_CREDIT = 11  # CREDIT (*)

    # Column indices for Movement NS/Manual file (0-indexed)
    COL_MOV_SOURCE = 0       # A - Source (Automation/NS/Manual)
    COL_MOV_BANK = 1         # B - Bank
    COL_MOV_ACCOUNT = 2      # C - Account
    COL_MOV_DATE = 3         # D - Date
    COL_MOV_DESCRIPTION = 4  # E - Bank description
    COL_MOV_DEBIT = 5        # F - Nợ (Debit)
    COL_MOV_CREDIT = 6       # G - Có (Credit)
    # Skip 7 = H (Net formula)
    COL_MOV_NATURE = 8       # I - Nature
    COL_MOV_ENTITY = 9       # J - Entity (VLOOKUP cached)
    COL_MOV_GROUPING = 10    # K - Grouping (VLOOKUP cached)
    COL_MOV_KEY_PAYMENT = 11 # L - Key payment (=I cached)
    COL_MOV_CURRENCY = 12    # M - Currency (VLOOKUP cached)
    COL_MOV_ACCOUNT_TYPE = 13 # N - Account type (VLOOKUP cached)

    # Sheet name
    SHEET_NAME = "Template details"

    # Required headers in Template details sheet (0-indexed column -> expected keyword)
    REQUIRED_HEADERS = {
        4: "bank code",
        5: "bank account number",
        7: "trans date",
        8: "description",
        10: "debit",
        11: "credit",
    }

    def __init__(self):
        pass

    def read_from_file(
        self,
        file_path: Union[str, Path],
        source_name: Optional[str] = None,
    ) -> List[MovementTransaction]:
        """
        Read transactions from a parsed bank statement Excel file.

        Args:
            file_path: Path to the Excel file
            source_name: Source name to use (default: filename)

        Returns:
            List of MovementTransaction objects
        """
        file_path = Path(file_path)

        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        if source_name is None:
            source_name = file_path.name

        with open(file_path, 'rb') as f:
            content = f.read()

        return self.read_from_bytes(content, source_name)

    def read_from_bytes(
        self,
        content: bytes,
        source_name: str = "uploaded_file",
    ) -> List[MovementTransaction]:
        """
        Read transactions from Excel bytes content.

        Args:
            content: Excel file content as bytes
            source_name: Source name for the transactions

        Returns:
            List of MovementTransaction objects (without Nature - to be classified later)
        """
        try:
            # Try reading with openpyxl first for better date handling
            wb = openpyxl.load_workbook(BytesIO(content), data_only=True)

            # Find the Template details sheet
            sheet_name = self._find_sheet(wb)
            if not sheet_name:
                wb.close()
                raise ValueError(f"Sheet '{self.SHEET_NAME}' not found in the Excel file")

            ws = wb[sheet_name]
            self._validate_template_details_schema(ws)
            transactions = []

            # Skip header row (row 1)
            for row_num in range(2, ws.max_row + 1):
                try:
                    tx = self._parse_row_openpyxl(ws, row_num, source_name)
                    if tx:
                        transactions.append(tx)
                except Exception as e:
                    logger.warning(f"Error parsing row {row_num}: {e}")
                    continue

            wb.close()

            logger.info(f"Read {len(transactions)} transactions from {source_name}")
            return transactions

        except Exception as e:
            logger.error(f"Error reading Excel file: {e}")
            raise

    # ------------------------------------------------------------------ #
    #  Movement NS/Manual file reader                                      #
    # ------------------------------------------------------------------ #

    # Required headers for Movement file (0-indexed column -> expected keyword)
    MOVEMENT_REQUIRED_HEADERS = {
        0: "source",
        1: "bank",
        2: "account",
        3: "date",
        4: "description",
    }

    def read_movement_file(
        self,
        content: bytes,
        source_name: str = "movement_upload",
    ) -> List[MovementTransaction]:
        """
        Read pre-classified transactions from Movement Excel file.

        Expected structure:
        - Sheet: "Sheet1" (or first sheet)
        - Columns: A=Source, B=Bank, C=Account, D=Date, E=Description,
                   F=Debit, G=Credit, I=Nature
        - All rows are imported with their original source (Automation/NS/Manual)

        Args:
            content: Excel file content as bytes
            source_name: Source name for tracking

        Returns:
            List of MovementTransaction with pre-populated .nature field
        """
        try:
            wb = openpyxl.load_workbook(BytesIO(content), data_only=True)

            # Find sheet (try "Sheet1" first, else use first sheet)
            sheet_name = self._find_movement_sheet(wb)
            if not sheet_name:
                wb.close()
                raise ValueError("No valid sheet found in Movement file")

            ws = wb[sheet_name]

            # Validate headers
            self._validate_movement_schema(ws)

            transactions = []
            skipped_invalid = 0

            # Skip header row (row 1), parse data starting from row 2
            for row_num in range(2, ws.max_row + 1):
                try:
                    tx = self._parse_movement_row(ws, row_num, source_name)
                    if tx is None:
                        continue
                    transactions.append(tx)
                except Exception as e:
                    skipped_invalid += 1
                    logger.warning(f"Movement row {row_num}: {e}")
                    continue

            wb.close()

            logger.info(
                f"Read {len(transactions)} transactions from {source_name} "
                f"(skipped {skipped_invalid} invalid)"
            )
            return transactions

        except Exception as e:
            logger.error(f"Error reading Movement file: {e}")
            raise

    def _find_movement_sheet(self, wb: openpyxl.Workbook) -> Optional[str]:
        """Find Sheet1 or return first sheet name."""
        for sheet_name in wb.sheetnames:
            if sheet_name.lower() == "sheet1":
                return sheet_name
        # Fallback to first sheet
        return wb.sheetnames[0] if wb.sheetnames else None

    def _validate_movement_schema(self, ws) -> None:
        """
        Validate Movement file headers (fuzzy match).
        Checks columns A-E for expected keywords.
        """
        for col_idx, keyword in self.MOVEMENT_REQUIRED_HEADERS.items():
            header_value = ws.cell(row=1, column=col_idx + 1).value
            header_norm = self._normalize_header(header_value)
            if keyword not in header_norm:
                raise ValueError(
                    f"Invalid Movement file format: "
                    f"column {col_idx + 1} header '{header_value}' does not contain '{keyword}'"
                )

    def _parse_movement_row(
        self,
        ws,
        row_num: int,
        source_name: str,
    ) -> Optional[MovementTransaction]:
        """
        Parse a single row from Movement file.

        Returns None if row is empty/invalid.
        All rows (Automation, NS, Manual) are kept with their original source.
        """
        # Read Source
        source_val = str(ws.cell(row=row_num, column=self.COL_MOV_SOURCE + 1).value or "").strip()

        # Read other fields
        bank = str(ws.cell(row=row_num, column=self.COL_MOV_BANK + 1).value or "").strip()
        account_raw = ws.cell(row=row_num, column=self.COL_MOV_ACCOUNT + 1)
        account = self._parse_text_value(account_raw.value, account_raw.number_format)
        date_val = ws.cell(row=row_num, column=self.COL_MOV_DATE + 1).value
        description = str(ws.cell(row=row_num, column=self.COL_MOV_DESCRIPTION + 1).value or "").strip()
        debit_val = ws.cell(row=row_num, column=self.COL_MOV_DEBIT + 1).value
        credit_val = ws.cell(row=row_num, column=self.COL_MOV_CREDIT + 1).value
        nature = str(ws.cell(row=row_num, column=self.COL_MOV_NATURE + 1).value or "").strip()

        # Read formula/derived columns (cached values for VLOOKUP columns)
        entity_val = ws.cell(row=row_num, column=self.COL_MOV_ENTITY + 1).value
        grouping_val = ws.cell(row=row_num, column=self.COL_MOV_GROUPING + 1).value
        key_payment_val = ws.cell(row=row_num, column=self.COL_MOV_KEY_PAYMENT + 1).value
        currency_val = ws.cell(row=row_num, column=self.COL_MOV_CURRENCY + 1).value
        account_type_val = ws.cell(row=row_num, column=self.COL_MOV_ACCOUNT_TYPE + 1).value

        # Skip empty rows (no account or no description)
        if not account:
            return None

        # Skip rows with no financial data
        if not date_val and not debit_val and not credit_val:
            return None

        # Parse date
        parsed_date = self._parse_date(date_val)
        if not parsed_date:
            return None

        # Parse amounts
        parsed_debit = self._parse_amount(debit_val)
        parsed_credit = self._parse_amount(credit_val)

        # Skip if both debit and credit are 0/None
        if not parsed_debit and not parsed_credit:
            return None

        # Helper to clean cached values (skip #N/A, empty strings)
        def _clean_cached(val):
            if val is None:
                return None
            s = str(val).strip()
            if not s or s.upper() in ("#N/A", "#REF!", "#VALUE!", "#NAME?", "N/A"):
                return None
            return s

        # Use the Source value from file (NS/Manual) as the source column
        return MovementTransaction(
            source=source_val,
            bank=bank,
            account=account,
            date=parsed_date,
            description=description,
            debit=parsed_debit,
            credit=parsed_credit,
            nature=nature,
            _classified_by="file",
            entity=_clean_cached(entity_val),
            grouping=_clean_cached(grouping_val),
            key_payment=_clean_cached(key_payment_val),
            currency=_clean_cached(currency_val),
            account_type=_clean_cached(account_type_val),
        )

    # ------------------------------------------------------------------ #

    def _find_sheet(self, wb: openpyxl.Workbook) -> Optional[str]:
        """Find the Template details sheet (case-insensitive)."""
        for sheet_name in wb.sheetnames:
            if sheet_name.lower() == self.SHEET_NAME.lower():
                return sheet_name
        return None

    @staticmethod
    def _normalize_header(value) -> str:
        if value is None:
            return ""
        # Keep alnum and spaces; normalize to lowercase single-space text.
        text = re.sub(r"[^a-zA-Z0-9 ]+", " ", str(value)).lower().strip()
        return re.sub(r"\s+", " ", text)

    def _validate_template_details_schema(self, ws) -> None:
        """
        Ensure we are reading the expected ERP statement structure from Template details.
        This prevents importing data from any non-standard worksheet layout.
        """
        for col_idx, keyword in self.REQUIRED_HEADERS.items():
            header_value = ws.cell(row=1, column=col_idx + 1).value
            header_norm = self._normalize_header(header_value)
            if keyword not in header_norm:
                raise ValueError(
                    f"Invalid statement format in '{self.SHEET_NAME}': "
                    f"column {col_idx + 1} header '{header_value}' does not contain '{keyword}'"
                )

    def _parse_row_openpyxl(
        self,
        ws,
        row_num: int,
        source_name: str,
    ) -> Optional[MovementTransaction]:
        """
        Parse a single row from openpyxl worksheet.

        Args:
            ws: openpyxl worksheet
            row_num: Row number (1-indexed)
            source_name: Source name

        Returns:
            MovementTransaction or None if row is empty/invalid
        """
        # Get cells (column numbers are 1-indexed in openpyxl)
        bank_cell = ws.cell(row=row_num, column=self.COL_BANK_CODE + 1)
        account_cell = ws.cell(row=row_num, column=self.COL_ACCOUNT_NUMBER + 1)
        date_cell = ws.cell(row=row_num, column=self.COL_TRANS_DATE + 1)
        description_cell = ws.cell(row=row_num, column=self.COL_DESCRIPTION + 1)
        debit_cell = ws.cell(row=row_num, column=self.COL_DEBIT + 1)
        credit_cell = ws.cell(row=row_num, column=self.COL_CREDIT + 1)

        bank_code = bank_cell.value
        account_number = self._parse_text_value(account_cell.value, account_cell.number_format)
        trans_date = date_cell.value
        description = self._parse_text_value(description_cell.value, description_cell.number_format)
        debit = debit_cell.value
        credit = credit_cell.value

        # Skip empty rows
        if not account_number:
            return None

        # Skip if no transaction data
        if not trans_date and not debit and not credit:
            return None

        # Parse date
        parsed_date = self._parse_date(trans_date)
        if not parsed_date:
            return None

        # Parse amounts
        parsed_debit = self._parse_amount(debit)
        parsed_credit = self._parse_amount(credit)

        # Skip if both debit and credit are 0/None
        if not parsed_debit and not parsed_credit:
            return None

        return MovementTransaction(
            source=source_name,
            bank=str(bank_code) if bank_code else "",
            account=account_number,
            date=parsed_date,
            description=description,
            debit=parsed_debit,
            credit=parsed_credit,
            nature="",  # Will be classified later
        )

    @staticmethod
    def _extract_zero_pad_width(number_format: Optional[str]) -> int:
        """
        Extract zero-padding width from formats like "00000000000000".
        Returns 0 when not a pure zero-pad format.
        """
        if not number_format:
            return 0
        primary = number_format.split(";")[0].strip()
        return len(primary) if re.fullmatch(r"0+", primary) else 0

    def _parse_text_value(self, value, number_format: Optional[str] = None) -> str:
        """
        Parse text-like fields (account/description) and preserve numeric identifiers.
        Handles Excel numeric conversion that may append ".0" or apply scientific notation.
        """
        if value is None:
            return ""

        if isinstance(value, str):
            return value.strip()

        if isinstance(value, Decimal):
            if value == value.to_integral_value():
                return str(value.to_integral_value())
            return format(value.normalize(), "f")

        if isinstance(value, int):
            text = str(value)
            width = self._extract_zero_pad_width(number_format)
            return text.zfill(width) if width and len(text) < width else text

        if isinstance(value, float):
            if math.isnan(value) or math.isinf(value):
                return ""
            dec = Decimal(str(value))
            if dec == dec.to_integral_value():
                text = str(dec.to_integral_value())
                width = self._extract_zero_pad_width(number_format)
                return text.zfill(width) if width and len(text) < width else text
            return format(dec.normalize(), "f")

        return str(value).strip()

    @staticmethod
    def _looks_like_thousands_format(text: str, sep: str) -> bool:
        """
        True when separators likely represent thousands groups:
        e.g., 1.234.567 or 1,234,567
        """
        parts = text.split(sep)
        if len(parts) <= 1:
            return False
        if not all(part.isdigit() for part in parts):
            return False
        if len(parts[0]) < 1 or len(parts[0]) > 3:
            return False
        return all(len(part) == 3 for part in parts[1:])

    def _parse_date(self, value) -> Optional[date]:
        """Parse date from various formats."""
        if value is None:
            return None

        if isinstance(value, datetime):
            return value.date()

        if isinstance(value, date):
            return value

        if isinstance(value, str):
            value = value.strip()
            if not value:
                return None

            # Try various date formats (with and without time component)
            formats = [
                "%Y-%m-%d",
                "%d/%m/%Y",
                "%m/%d/%Y",
                "%Y/%m/%d",
                "%d-%m-%Y",
                "%m-%d-%Y",
                "%d-%m-%Y %H:%M:%S",
                "%Y-%m-%d %H:%M:%S",
                "%d/%m/%Y %H:%M:%S",
                "%m/%d/%Y %H:%M:%S",
            ]

            for fmt in formats:
                try:
                    return datetime.strptime(value, fmt).date()
                except ValueError:
                    continue

            logger.warning(f"Could not parse date: {value}")
            return None

        return None

    def _parse_amount(self, value) -> Optional[Decimal]:
        """Parse amount to Decimal."""
        if value is None:
            return None

        if isinstance(value, (int, float)):
            if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
                return None
            if value == 0:
                return None
            return Decimal(str(value))

        if isinstance(value, Decimal):
            if value == 0:
                return None
            return value

        if isinstance(value, str):
            value = value.strip()
            if not value or value == "0":
                return None

            # Support negatives in parentheses: (123) => -123
            is_negative = value.startswith("(") and value.endswith(")")
            if is_negative:
                value = value[1:-1].strip()

            # Normalize spaces and NBSP
            value = value.replace("\u00A0", "").replace(" ", "")

            # Locale-aware normalization:
            # - 1.234.567,89 -> 1234567.89
            # - 1,234,567.89 -> 1234567.89
            # - 1.234.567 -> 1234567
            # - 1,234,567 -> 1234567
            if "." in value and "," in value:
                if value.rfind(",") > value.rfind("."):
                    # Comma is decimal separator
                    value = value.replace(".", "").replace(",", ".")
                else:
                    # Dot is decimal separator
                    value = value.replace(",", "")
            elif "," in value:
                if self._looks_like_thousands_format(value, ","):
                    value = value.replace(",", "")
                else:
                    value = value.replace(",", ".")
            elif "." in value and self._looks_like_thousands_format(value, "."):
                value = value.replace(".", "")

            try:
                result = Decimal(value)
                if is_negative:
                    result = -result
                return result if result != 0 else None
            except:
                return None

        return None

    def filter_by_date_range(
        self,
        transactions: List[MovementTransaction],
        start_date: date,
        end_date: date,
    ) -> Tuple[List[MovementTransaction], int]:
        """
        Filter transactions by date range.

        Args:
            transactions: List of transactions
            start_date: Start date (inclusive)
            end_date: End date (inclusive)

        Returns:
            Tuple of (filtered_transactions, skipped_count)
        """
        filtered = []
        skipped = 0
        skipped_details = []

        logger.info(f"Filtering transactions with date range: {start_date} to {end_date}")

        for tx in transactions:
            if tx.date is None:
                skipped += 1
                skipped_details.append(f"No date: {tx.description[:50] if tx.description else 'N/A'}")
                continue

            if start_date <= tx.date <= end_date:
                filtered.append(tx)
            else:
                skipped += 1
                skipped_details.append(f"{tx.date.strftime('%d/%m/%Y')} (outside range): {tx.description[:50] if tx.description else 'N/A'}")

        logger.info(f"Filtered {len(filtered)} transactions, skipped {skipped} outside date range")
        if skipped_details:
            logger.info(f"Skipped transactions: {skipped_details[:10]}")  # Log first 10
        return filtered, skipped

    def get_summary(self, transactions: List[MovementTransaction]) -> dict:
        """
        Get summary statistics for a list of transactions.

        Returns:
            Dict with summary statistics
        """
        if not transactions:
            return {
                "total_transactions": 0,
                "total_debit": 0,
                "total_credit": 0,
                "unique_accounts": 0,
                "unique_banks": 0,
                "date_range": None,
            }

        total_debit = sum(tx.debit or Decimal(0) for tx in transactions)
        total_credit = sum(tx.credit or Decimal(0) for tx in transactions)
        unique_accounts = len(set(tx.account for tx in transactions))
        unique_banks = len(set(tx.bank for tx in transactions))
        dates = [tx.date for tx in transactions if tx.date]
        date_range = (min(dates), max(dates)) if dates else None

        return {
            "total_transactions": len(transactions),
            "total_debit": float(total_debit),
            "total_credit": float(total_credit),
            "unique_accounts": unique_accounts,
            "unique_banks": unique_banks,
            "date_range": {
                "start": date_range[0].isoformat() if date_range else None,
                "end": date_range[1].isoformat() if date_range else None,
            }
        }
