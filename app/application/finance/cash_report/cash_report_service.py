"""
Cash Report Service - Main service for cash report automation.
Integrates all components: MasterTemplateManager, MovementDataWriter, BankStatementReader, AI Classifier.
"""
import asyncio
import csv
import re
import shutil
import uuid
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from difflib import SequenceMatcher
from decimal import Decimal
from pathlib import Path
from typing import List, Optional, Dict, Any, Tuple

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, update, delete
from sqlalchemy.orm import selectinload

from app.shared.utils.logging_config import get_logger
from app.infrastructure.database.models.cash_report_session import (
    CashReportSessionModel,
    CashReportUploadedFileModel,
    CashReportSessionStatus,
)
from app.domain.finance.cash_report.services.ai_classifier import (
    AITransactionClassifier,
    ALL_PAYMENT_CATEGORIES,
    ALL_RECEIPT_CATEGORIES,
)
from app.domain.finance.cash_report.services.tfidf_classifier import TfidfNatureClassifier
from app.domain.finance.cash_report.models.key_payment import KeyPaymentClassifier

from .master_template_manager import MasterTemplateManager
from .movement_data_writer import MovementDataWriter, MovementTransaction
from .bank_statement_reader import BankStatementReader
from .progress_store import ProgressEvent

logger = get_logger(__name__)

# Fallback mapping when Def!Bank/Abb table cannot be read.
# Keys are normalized by removing spaces/punctuation and uppercasing.
DEFAULT_BANK_ALIAS_TO_FULL_NAME = {
    "VCB": "VIETCOMBANK",
    "VIETCOMBANK": "VIETCOMBANK",
    "VTB": "VIETINBANK",
    "VIETINBANK": "VIETINBANK",
    "BIDV": "BIDV",
    "OCB": "OCB",
    "VIB": "VIB",
    "MB": "MB",
    "MBB": "MB",
    "ACB": "ACB",
    "OTHERS": "Others",
    "SINOPAC": "SINOPAC",
    "SNP": "SINOPAC",
    "TECHCOMBANK": "TECHCOMBANK",
    "TCB": "TECHCOMBANK",
    "UOB": "UOB",
    "STANDARDCHARTERED": "STANDARD CHARTERED",
    "KEBHANA": "KEB Hana",
    "VPBANK": "VPBank",
    "KBANK": "KBANK",
    "MAYBANK": "MAYBANK",
    "CTBC": "CTBC",
    "WOORIBANK": "Woori Bank",
    "WOORI": "Woori Bank",
}

# â"€â"€ Settlement detection patterns (Group A - Táº¥t toÃ¡n / RÃºt tiá»n) â"€â"€
# Used to detect settlement-eligible transactions by description.
# Requires BOTH: Nature = "Internal transfer in" AND keyword/pattern match.

# Regex patterns for settlement detection (compiled at module load)
SETTLEMENT_PATTERNS = [
    # GROUP A -- Táº¥t toÃ¡n / RÃºt tiá»n (Current Account â† Savings)
    # 1-4: Táº¥t toÃ¡n / Terminate
    r"TAT\s*TOAN",                              # Táº¥t toÃ¡n
    r"TT\s+HDTG",                               # Viáº¿t táº¯t: Táº¥t toÃ¡n HDTG
    r"TERMINATE\s*(PARTIAL\s*)?SAVING",         # Terminate saving -- VTB
    r"CA\s*-?\s*TARGET",                        # CA - TARGET -- SINOPAC settlement
    # 5-10: Withdrawal / RÃºt tiá»n (EN)
    r"WITHDRAWAL\s*(OF\s*)?(THE\s*)?TERM\s*DEPOSIT",  # Withdrawal term deposit -- VTB
    r"FULL\s*WITHDRAWAL\s*FROM\s*SAVINGS",      # Full withdrawal -- VTB
    r"WITHDRAW\s*FIXED\s*DEPOSIT",              # Withdraw fixed deposit -- BIDV
    r"PARTIALLY\s*WITHDRAW\s*TERM\s*DEPOSIT",   # Partially withdraw -- BIDV
    r"CLOSE\s*TD\s*ACC",                        # Close Term Deposit -- Kbank
    r"CLOSING\s*TERM\s*DEPOSIT",                # Closing term deposit -- Escrow
    r"CLOSING\s*TDA",                           # Closing TDA -- Woori
    # 11-17: RÃºt tiá»n (VI)
    r"RUT\s*TIEN\s*GUI",                        # RÃºt tiá»n gá»­i online -- BIDV
    r"RUT\s*TIEN.*TIET\s*KIEM",                 # RÃºt tiá»n tá»« TK tiáº¿t kiá»‡m
    r"RUT\s*TIEN.*HDTG",                        # RÃºt tiá»n theo HDTG
    r"RUT\s*MOT\s*PHAN",                        # RÃºt má»™t pháº§n tiá»n gá»­i
    r"RUT\s*1\s*PHAN",                          # RÃºt 1 pháº§n -- variant sá»'
    r"RUT\s*GOC\s*MOT\s*PHAN",                  # RÃºt gá»'c má»™t pháº§n HDTG
    r"RUT\s*TRUOC\s*HAN",                       # RÃºt trÆ°á»›c háº¡n
    # 18-22: ÄÃ³ng TK
    r"DONG\s*TKKH",                             # ÄÃ³ng TK ká»³ háº¡n
    r"DONG\s*TK.*KY\s*HAN",                     # ÄÃ³ng TK + ká»³ háº¡n
    r"DONG\s*TK.*TIET\s*KIEM",                  # ÄÃ³ng TK tiáº¿t kiá»‡m
    r"DONG\s*TK\s*(BIDV|THEO\s*DE|\d{3,})",     # ÄÃ³ng TK + mÃ£ NH/sá»' TK
    r"DONG\s*HDTG",                             # ÄÃ³ng HDTG
    # 23-24: Tráº£ gá»'c
    r"TRA\s*GOC",                               # Tráº£ gá»'c (standalone -- catches all TRA GOC variants)
    # 25: Bare account number (SINOPAC-style, description = only digits)
    r"^\s*\d{10,}(?:\.0+)?\s*$",                # Pure account number (10+ digits, optional .0) -- SINOPAC
]

# Simple keyword matching (faster, for common patterns)
# NOTE: Only include keywords specific to settling/closing saving deposits.
# Do NOT include intercompany transfer keywords (SHL, BCC, CHUYEN TIEN, etc.)
# -- those are normal "Internal transfer in" but NOT settlement of savings.
SETTLEMENT_KEYWORDS = [
    "tat toan",         # Táº¥t toÃ¡n (lowercase fallback)
    "tra goc",          # Tráº£ gá»'c (lowercase fallback)
    "rut tien gui",     # RÃºt tiá»n gá»­i (lowercase fallback)
]

# Compile regex patterns for efficiency
import re as _re
SETTLEMENT_PATTERNS_COMPILED = [_re.compile(p, _re.IGNORECASE) for p in SETTLEMENT_PATTERNS]

# Broad HDTG/deposit-related keywords for large-receipt guardrail.
# When a RECEIPT has description matching these keywords AND debit > 1B VND,
# it is almost certainly an Internal transfer in (deposit maturity / withdrawal).
# This catches cases not covered by specific SETTLEMENT_PATTERNS regex.
HDTG_DEPOSIT_KEYWORDS = [
    "hdtg",             # Hop dong tien gui
    "tien gui",         # Tien gui (deposit)
    "tiet kiem",        # Tiet kiem (savings)
    "timemo",           # VCB time deposit (memo)
    "timect",           # VCB time deposit (CT)
    "ky han",           # Ky han (term)
    "saving",           # Saving (EN)
    "term deposit",     # Term deposit (EN)
    "fixed deposit",    # Fixed deposit (EN)
]
_UNIT_1B = Decimal("1000000000")  # 1 billion VND


@dataclass(frozen=True)
class ExternalKeywordRule:
    """One keyword-classification rule loaded from Key Words CSV."""
    keyword_norm: str
    category: str
    is_receipt: Optional[bool]
    amount_operator: Optional[str]
    amount_threshold: Optional[Decimal]

# Targeted settlement exceptions observed in production data.
# Keep these narrow to avoid broad side effects.
CA_TARGET_NORMALIZED = "ca target"
BIDV_IBANK_WITHDRAWAL_NORMALIZED = "rut tien gui online tren bidv ibank"
BIDV_IBANK_WITHDRAWAL_RECEIPT_ACCOUNTS = frozenset({"2221133456"})

# Natures eligible for settlement detection.
SETTLEMENT_ELIGIBLE_NATURES = frozenset({
    "internal transfer in",   # Only Internal transfer in triggers settlement
})

# Natures that are NEVER settlement candidates, regardless of description.
# Counter entries and automation-generated interest rows must be excluded.
SETTLEMENT_BLOCKED_NATURES = frozenset({
    "internal transfer out",
    "other receipts",
})

# Special handling for numeric-only bank descriptions (account-like values).
# Flexible rule: any numeric-only description in this shape is processed.
NUMERIC_ACCOUNT_DESC_RE = _re.compile(r"^\d{10,20}$")

# Ground-truth transaction mapping file: description -> Nature.
# This file is curated by finance team and should be treated as highest-priority
# reference when an exact description+direction match exists.
TRANSACTIONS_REFERENCE_FILE = Path(__file__).resolve().parents[4] / "movement_nature_filter" / "Transactions.csv"
SETTLEMENT_REFERENCE_FILE = Path(__file__).resolve().parents[4] / "movement_nature_filter" / "Settlement.csv"
OPEN_NEW_REFERENCE_FILE = Path(__file__).resolve().parents[4] / "movement_nature_filter" / "OpenNew.csv"

# Keep these exclusions even when loading patterns from OpenNew.csv.
# They are known non-open-new movements and caused false positives in production.
OPEN_NEW_EXCLUDED_REGEXES = frozenset({
    r"AUTO\s*ROLLOVER",
    r"^WITHDRAWAL$",
})

# Intercompany profit-distribution payments are operational transfers,
# not dividend expense and never open-new savings.
PROFIT_DISTRIBUTION_KEYWORDS = (
    "profit distribution",
    "advance for profit distribution",
    "phan loi nhuan",
    "phan phoi loi nhuan",
    "tam ung phan loi nhuan",
    "tam ung phan phoi",
    "chia loi nhuan",
    "co tuc",
)

# Targeted open-new exception observed for Woori statements where
# savings placement transactions are labeled as "Withdrawal - Withdrawal".
WOORI_WITHDRAWAL_WITHDRAWAL_NORMALIZED = "withdrawal withdrawal"

# Normalize known legacy/variant category names from reference files.
TRANSACTION_NATURE_ALIASES = {
    "Operating Expense": "Operating expense",
    "Dividend paid (inside group)": "Dividend paid (inside group)",
    "Dividend paid (inside\xa0group)": "Dividend paid (inside group)",
    "Internal contribution": "Internal transfer in",
    "Internal Contribution": "Internal transfer in",
    "Internal contribution in": "Internal transfer in",
    "Internal Contribution in": "Internal transfer in",
    "Internal contribution out": "Internal transfer out",
    "Internal Contribution out": "Internal transfer out",
    "Refund land/deal deposit payment": "Receipt from tenants",
}

# Entity codes for dual-entity transfer detection.
# If description mentions 2+ DIFFERENT entity codes -> internal transfer between entities.
# E.g., "VC3_DC2_TRANSFER MONEY" contains VC3 and DC2 -> settlement candidate.
ENTITY_CODES = [
    "bwd", "bhd", "bbn", "hd2", "bhp", "bnt", "hd3", "bth",
    "th1hc", "th2hc", "th2", "btu", "bsg", "bnc", "bsl", "nt2",
    "bwb", "bwp", "dvl", "ndv", "tdh", "dal", "bla", "sgl",
    "hnm", "ire", "tdm", "ktr", "scc",
    "vc1", "vc2", "vc3",
    "bb3", "tpt", "bb4", "stl", "bha",
    "spv4a", "spv4b", "spv4c", "spv4d", "spv4e",
    "spv5a", "spv5b", "spv5c", "spv5d", "spv5e",
    "plc", "gtc", "bci", "bdg",
    "h4f", "h4h", "h4g", "h4i", "h4j",
    "h5f", "h5g", "h5h", "h5i", "h5j",
    "dvc", "byp", "bb5", "bb6", "blb",
    "xa1", "xa2", "wl1", "wl2",
    "bnh", "bdh", "btp", "mp3", "bba",
    "h4k", "h4l", "h4m", "h4n", "h4o", "h4p",
    "h5k", "h5l", "h5m",
    "wj1", "wj2", "wj3", "t3b", "msp", "tsp", "sc2",
    "h4q", "h4r", "h4s", "h4t", "h4u", "h4v", "h4w", "h4x", "h4y", "h4z",
    "h5q", "h5r", "h5s", "h5t", "h5u",
    "bhi", "btd", "w2a", "w2b",
    "hj1", "hj2", "yph", "hqv", "tha", "hhl", "bhl",
    "jyp", "jhp", "dtp",
    "h4c", "tcs", "h4d", "bmp", "h4a", "h5a", "h4b", "h5b", "h4e",
    "qv2", "w3a", "w3b",
    "pna", "bpd", "tna",
    "n4a", "n4b", "n4c", "n4d", "n4e",
    "lmx", "dc2", "pat", "st2", "tde", "mpl", "nas", "lbn",
    # Additional codes found in settlement descriptions
    "pae", "pdhd", "pde", "wl2b", "shl",
]

# â"€â"€ Open New Saving Account patterns (Group B - Gá»­i tiá»n / Má»Ÿ HDTG) â"€â"€
# Used to detect transactions that open new saving accounts.
# Requires BOTH: Nature = "Internal transfer out" AND pattern match.

OPEN_NEW_PATTERNS = [
    # GROUP B -- Gá»­i tiá»n / Má»Ÿ HDTG (Current Account -> Saving Account)
    # Tiá»n gá»­i cÃ³ ká»³ háº¡n (reversed word order: "TIEN GUI CO KY HAN" not "GUI TIEN")
    r"TIEN\s*GUI\s*CO\s*KY\s*HAN",               # Tiá»n gá»­i cÃ³ ká»³ háº¡n -- BIDV eFAST
    # Há»£p Ä'á»"ng tiá»n gá»­i (without MO prefix)
    r"HOP\s*DONG\s*TIEN\s*GUI",                   # Há»£p Ä'á»"ng tiá»n gá»­i sá»' -- BIDV
    # Gá»­i tiá»n patterns
    r"GUI\s*TIEN.*HDTG",                          # Gá»­i tiá»n theo HDTG
    r"GUI\s*TIEN.*HOP\s*DONG",                    # Gá»­i tiá»n theo há»£p Ä'á»"ng
    r"GUI\s*TIEN\s*VAO\s*HOP\s*DONG",             # Gá»­i tiá»n vÃ o HDTG
    r"GUI\s*TIEN\s*GUI\s*KY\s*HAN",               # Gá»­i tiá»n gá»­i ká»³ háº¡n
    r"GUI\s*TIEN\s*VAO\s*TAI\s*KHOAN\s*TIEN\s*GUI",  # Gá»­i tiá»n vÃ o TKTG cÃ³ ká»³ háº¡n
    r"GUI\s*TIEN\s*TK",                           # Gá»­i tiá»n TK -- BIDV
    r"GUI\s*TIEN\s*GUI\s*CKH",                    # Gá»­i tiá»n gá»­i CKH -- ACB
    r"GUI\s*TIEN\s*NGAY",                         # Gá»­i tiá»n ngÃ y -- TCB
    # 32-38: Gá»­i HDTG/TK
    r"GUI\s*HDTG",                                # Gá»­i HDTG
    r"GUI\s*TK\s*THEO\s*HDTG",                    # Gá»­i TK theo HDTG
    r"GUI\s*TIET\s*KIEM",                         # Gá»­i tiáº¿t kiá»‡m
    r"GUI\s*KY\s*HAN",                            # Gá»­i ká»³ háº¡n
    r"GUI\s*TK\s*\d+\s*THANG",                    # Gá»­i TK + thÃ¡ng -- VTB
    r"GUI\s*TKLH",                                # Gá»­i TK linh hoáº¡t -- BIDV
    r"GUI\s*TGCKH",                               # Gá»­i tiá»n gá»­i CKH -- ACB
    # 39-44: Má»Ÿ HDTG
    r"MO\s*HDTG",                                 # Má»Ÿ HDTG
    r"MO\s*HOP\s*DONG\s*TIEN\s*GUI",              # Má»Ÿ há»£p Ä'á»"ng tiá»n gá»­i
    r"MO\s*HGTG",                                 # Má»Ÿ HGTG -- typo = MO HDTG
    r"TRICH\s*TK\s*MO.*HDTG",                     # TrÃ­ch TK má»Ÿ HDTG
    r"HACH\s*TOAN\s*HDTG",                        # Háº¡ch toÃ¡n HDTG
    r"HT\s+HDTG",                                 # Viáº¿t táº¯t: Háº¡ch toÃ¡n HDTG
    # 45-54: HDTG patterns
    r"CHUYEN\s*KHOAN\s*VAO\s*HDTG",               # Chuyá»ƒn khoáº£n vÃ o HDTG
    r"HDTG\s*KH\s*\d",                            # HDTG ká»³ háº¡n + sá»'
    r"HDTG\s*RGLH",                               # HDTG rÃºt gá»'c linh hoáº¡t
    r"HDTG\s*\d+\s*(THANG|NGAY|TH|THT)",          # HDTG + ká»³ háº¡n thÃ¡ng/ngÃ y
    r"HDTG\s*SO\s*\d",                            # HDTG sá»' -- BIDV
    r"HDTG\s+\d{3}[\s/]",                         # HDTG + mÃ£ chi nhÃ¡nh
    r"HDTG:\s*\d{3}",                             # HDTG: + mÃ£ HÄ
    r"HDTG\s+CKH",                                # HDTG cÃ³ ká»³ háº¡n
    r"HGTD\s*SO",                                 # HGTD -- typo = HDTG
    r"\d+/\d+/HDTG",                              # MÃ£ HÄ dáº¡ng sá»'/nÄƒm/HDTG
    # 55-57: VCB Time Deposit
    r"CK\s*SANG\s*TK\s*TIME",                     # CK sang TK TIME -- VCB
    r"TIMEMO",                                    # TK TIMEMO -- VCB savings
    r"TIMECT",                                    # TK TIMECT -- VCB savings
    # 58-63: English patterns
    r"DEPOSIT\s*FOR\s*NEXT\s*PRINCIPAL",          # KEB Hana escrow savings
    r"OPEN\s*TD",                                 # Open Term Deposit -- Kbank
    r"COMPLETED\s*TRANSFER\s*TO\s*BIDV\s*CD",     # Transfer to BIDV CD account
    r"SETTLEMENT\s*CONTRACT",                     # Settlement contract -- VIB
    # NOTE: AUTO ROLLOVER removed -- rollover renews existing deposit, not open new
    # Ká»³ háº¡n / RGLH / ÄÃ¡o háº¡n
    r"KY\s*HAN.*RGLH",                            # Ká»³ háº¡n + RGLH
    r"RGLH.*KY\s*HAN",                            # RGLH + ká»³ háº¡n
    r"DAO\s*HAN.*QUAY\s*VONG",                    # ÄÃ¡o háº¡n quay vÃ²ng
]

# Compile open new patterns for efficiency
OPEN_NEW_PATTERNS_COMPILED = [_re.compile(p, _re.IGNORECASE) for p in OPEN_NEW_PATTERNS]


class CashReportService:
    """
    Main service for cash report automation.

    Orchestrates the flow:
    1. Create session with config
    2. Upload parsed bank statements
    3. Classify transactions (rule-based first, AI for leftovers)
    4. Review/confirm before writing
    5. Write to Movement sheet
    6. Download result
    """

    def __init__(self, db_session: Optional[AsyncSession] = None):
        """
        Initialize the service.

        Args:
            db_session: Optional async database session for persistence
        """
        self.db_session = db_session
        self.template_manager = MasterTemplateManager()
        self.statement_reader = BankStatementReader()
        self.ai_classifier = AITransactionClassifier()
        self.rule_classifier = KeyPaymentClassifier()
        self._rules_file_mtime = self._get_rules_file_mtime()
        self._external_keyword_index = self._build_external_keyword_index()
        self._settlement_rules_mtime: Optional[float] = None
        self._open_new_rules_mtime: Optional[float] = None
        self._settlement_patterns_compiled = SETTLEMENT_PATTERNS_COMPILED
        self._open_new_patterns_compiled = OPEN_NEW_PATTERNS_COMPILED
        self._transactions_reference_mtime: Optional[float] = None
        self._transactions_reference_exact: Dict[Tuple[str, bool], str] = {}
        self._transactions_reference_majority: Dict[Tuple[str, bool], str] = {}
        self._transactions_reference_fuzzy_exact: Dict[Tuple[str, bool], str] = {}
        self._transactions_reference_fuzzy_majority: Dict[Tuple[str, bool], str] = {}
        self._transactions_reference_template_exact: Dict[Tuple[str, bool], str] = {}
        self._transactions_reference_template_majority: Dict[Tuple[str, bool], str] = {}
        self._tfidf_classifier = TfidfNatureClassifier(min_similarity=0.55)
        self._tfidf_mtime: Optional[float] = None
        self._reload_detection_pattern_indexes()
        self._rebuild_transactions_reference_index()
        self._rebuild_tfidf_index()

    def _log_progress_event(
        self,
        *,
        process: str,
        event: ProgressEvent,
        session_id: Optional[str] = None,
    ) -> None:
        """Mirror progress events to terminal logs for easier debugging."""
        detail = (event.detail or "").strip()
        data_keys = ",".join(sorted(str(k) for k in event.data.keys())) if event.data else "-"
        logger.info(
            "[%s] event=%s step=%s pct=%s session=%s message=%s%s data_keys=%s",
            process,
            event.event_type,
            event.step,
            event.percentage,
            session_id or "-",
            event.message,
            f" detail={detail}" if detail else "",
            data_keys,
        )

    def _emit_progress(
        self,
        *,
        process: str,
        progress_callback: Optional[callable],
        event_type: str,
        step: str,
        message: str,
        detail: str = "",
        percentage: int = 0,
        data: Optional[Dict[str, Any]] = None,
        session_id: Optional[str] = None,
        message_key: str = "",
        message_params: Optional[Dict[str, Any]] = None,
        detail_key: str = "",
        detail_params: Optional[Dict[str, Any]] = None,
    ) -> None:
        """Build, log, and optionally emit a progress event."""
        event = ProgressEvent(
            event_type=event_type,
            step=step,
            message=message,
            detail=detail,
            percentage=percentage,
            data=data or {},
            message_key=message_key,
            message_params=message_params,
            detail_key=detail_key,
            detail_params=detail_params,
        )
        self._log_progress_event(process=process, event=event, session_id=session_id)
        if progress_callback:
            progress_callback(event)

    async def _verify_session_owner(self, session_id: str, user_id: int) -> None:
        """
        Verify that the session belongs to the given user.
        Raises PermissionError if the session belongs to another user.
        Raises ValueError if session not found.
        """
        if not self.db_session:
            return

        result = await self.db_session.execute(
            select(CashReportSessionModel).where(
                CashReportSessionModel.session_id == session_id
            )
        )
        session = result.scalar_one_or_none()

        if not session:
            raise ValueError(f"Session {session_id} not found")

        if session.user_id is not None and session.user_id != user_id:
            raise PermissionError("You do not have access to this session")

    async def _ensure_working_file(
        self, session_id: str
    ) -> Tuple[Dict[str, Any], Path]:
        """
        Validate that both session info AND working file exist on disk.

        Returns (session_info, working_file_path).

        Raises ValueError with a clear message distinguishing:
        - Session not found at all (not in DB, not on disk)
        - Working file expired (session in DB but file lost after server restart)
        """
        session_info = self.template_manager.get_session_info(session_id)
        working_file = self.template_manager.get_working_file_path(session_id)

        if session_info and working_file and working_file.exists():
            return session_info, working_file

        # File missing — check DB to give a clear error
        if self.db_session:
            result = await self.db_session.execute(
                select(CashReportSessionModel).where(
                    CashReportSessionModel.session_id == session_id
                )
            )
            db_session = result.scalar_one_or_none()
            if db_session:
                logger.warning(
                    "Session %s exists in DB but working file is missing on disk "
                    "(likely lost after server restart/deploy)",
                    session_id,
                )
                raise ValueError(
                    f"Working file for session {session_id} has expired "
                    f"(server was restarted). Please create a new session."
                )

        raise ValueError(f"Session {session_id} not found")

    async def get_or_create_session(
        self,
        opening_date: date,
        ending_date: date,
        fx_rate: Decimal = Decimal("26175"),
        period_name: str = "",
        user_id: Optional[int] = None,
        template_bytes: Optional[bytes] = None,
    ) -> Dict[str, Any]:
        """
        Get existing active session or create a new one.
        Only 1 session allowed per user.

        Args:
            opening_date: Report period start date
            ending_date: Report period end date
            fx_rate: VND/USD exchange rate
            period_name: Period name (e.g., "W3-4Jan26")
            user_id: Owner user ID

        Returns:
            Session info dict with 'is_existing' flag
        """
        # Check for existing active session for this user
        if self.db_session:
            query = select(CashReportSessionModel).where(
                CashReportSessionModel.status == CashReportSessionStatus.ACTIVE
            ).options(selectinload(CashReportSessionModel.uploaded_files))
            if user_id is not None:
                query = query.where(CashReportSessionModel.user_id == user_id)
            query = query.options(selectinload(CashReportSessionModel.uploaded_files))
            query = query.order_by(CashReportSessionModel.created_at.desc())

            result = await self.db_session.execute(query)
            existing_session = result.scalar_one_or_none()

            if existing_session:
                # Return existing session info
                working_file = self.template_manager.get_working_file_path(existing_session.session_id)
                file_size_mb = 0
                if working_file and working_file.exists():
                    file_size_mb = round(working_file.stat().st_size / (1024 * 1024), 2)

                uploaded_transactions = sum(
                    int(f.transactions_added or 0)
                    for f in (existing_session.uploaded_files or [])
                )
                breakdown = self._session_breakdown(
                    existing_session.total_transactions,
                    existing_session.metadata_json,
                    original_transactions=uploaded_transactions,
                )

                logger.info(f"Returning existing session: {existing_session.session_id}")
                return {
                    "session_id": existing_session.session_id,
                    "is_existing": True,
                    "working_file": str(working_file) if working_file else None,
                    "file_size_mb": file_size_mb,
                    "movement_rows": breakdown["transactions"],
                    "movement_rows_total": existing_session.total_transactions,
                    "breakdown": breakdown,
                    "config": {
                        "opening_date": existing_session.opening_date.isoformat() if existing_session.opening_date else None,
                        "ending_date": existing_session.ending_date.isoformat() if existing_session.ending_date else None,
                        "fx_rate": float(existing_session.fx_rate) if existing_session.fx_rate else None,
                        "period_name": existing_session.period_name,
                    },
                }

        # No existing session, create new one
        return await self._create_new_session(opening_date, ending_date, fx_rate, period_name, user_id=user_id, template_bytes=template_bytes)

    async def _create_new_session(
        self,
        opening_date: date,
        ending_date: date,
        fx_rate: Decimal,
        period_name: str,
        user_id: Optional[int] = None,
        template_bytes: Optional[bytes] = None,
    ) -> Dict[str, Any]:
        """Internal method to create a new session."""
        # Create session with template manager
        session_info = self.template_manager.create_session(
            opening_date=opening_date,
            ending_date=ending_date,
            fx_rate=fx_rate,
            period_name=period_name,
            template_bytes=template_bytes,
        )

        # Persist to database if session available
        if self.db_session:
            db_model = CashReportSessionModel(
                session_id=session_info["session_id"],
                status=CashReportSessionStatus.ACTIVE,
                period_name=period_name,
                opening_date=opening_date,
                ending_date=ending_date,
                fx_rate=fx_rate,
                working_file_path=session_info["working_file"],
                total_transactions=0,
                total_files_uploaded=0,
                user_id=user_id,
                # Track whether Movement prep already ran (user-uploaded template)
                metadata_json={
                    "movement_prepared": session_info.get("movement_prepared", False),
                    "settlement_counter_entries": 0,
                    "open_new_counter_entries": 0,
                    "settlement_interest_splits": 0,
                },
            )
            self.db_session.add(db_model)
            await self.db_session.commit()

        logger.info(f"Created new cash report session: {session_info['session_id']}")
        return {
            **session_info,
            "is_existing": False,
            "breakdown": {
                "transactions": 0,
                "settlement": 0,
                "open_new": 0,
                "interest_splits": 0,
                "total_rows": 0,
            },
        }

    async def upload_bank_statements(
        self,
        session_id: str,
        files: List[Tuple[str, bytes]],  # List of (filename, content)
        filter_by_date: bool = True,
        progress_callback: Optional[callable] = None,
        user_id: Optional[int] = None,
    ) -> Dict[str, Any]:
        """
        Upload and process parsed bank statement files.

        Args:
            session_id: The session ID
            files: List of (filename, file_content) tuples
            filter_by_date: Whether to filter transactions by session date range
            user_id: Owner user ID for access control

        Returns:
            Processing result summary
        """
        import time as _time

        _started_at = _time.monotonic()
        logger.info(
            "Upload bank statements started: session_id=%s files=%s filter_by_date=%s has_progress_callback=%s user_id=%s",
            session_id,
            len(files),
            filter_by_date,
            bool(progress_callback),
            user_id,
        )

        def _finish(result: Dict[str, Any]) -> Dict[str, Any]:
            elapsed_ms = int((_time.monotonic() - _started_at) * 1000)
            logger.info(
                "Upload bank statements finished: session_id=%s added=%s skipped=%s files_processed=%s elapsed_ms=%s",
                session_id,
                result.get("total_transactions_added", 0),
                result.get("total_transactions_skipped", 0),
                result.get("files_processed", len(files)),
                elapsed_ms,
            )
            return result

        def emit(
            event_type: str,
            step: str,
            message: str,
            *,
            detail: str = "",
            percentage: int = 0,
            data: Optional[Dict[str, Any]] = None,
            message_key: str = "",
            message_params: Optional[Dict[str, Any]] = None,
            detail_key: str = "",
            detail_params: Optional[Dict[str, Any]] = None,
        ) -> None:
            self._emit_progress(
                process="upload_bank_statements",
                progress_callback=progress_callback,
                session_id=session_id,
                event_type=event_type,
                step=step,
                message=message,
                detail=detail,
                percentage=percentage,
                data=data,
                message_key=message_key,
                message_params=message_params,
                detail_key=detail_key,
                detail_params=detail_params,
            )

        if user_id is not None:
            await self._verify_session_owner(session_id, user_id)

        # Get session info + validate working file exists
        session_info, working_file = await self._ensure_working_file(session_id)

        # Get date range for filtering
        opening_date = None
        ending_date = None
        if filter_by_date and session_info.get("config"):
            config = session_info["config"]
            if config.get("opening_date"):
                date_str = config["opening_date"]
                # Handle both date and datetime strings
                if "T" in date_str:
                    opening_date = datetime.fromisoformat(date_str).date()
                else:
                    opening_date = date.fromisoformat(date_str)
            if config.get("ending_date"):
                date_str = config["ending_date"]
                if "T" in date_str:
                    ending_date = datetime.fromisoformat(date_str).date()
                else:
                    ending_date = date.fromisoformat(date_str)

        # Process each file
        all_transactions: List[MovementTransaction] = []
        file_results = []
        total_skipped = 0
        total_found = 0

        for file_idx, (filename, content) in enumerate(files):
            try:
                # Emit progress: reading file
                emit(
                    "step_start",
                    "reading",
                    f"Reading {filename}...",
                    detail=f"Parsing transactions from file {file_idx + 1}/{len(files)}",
                    percentage=int((file_idx) / len(files) * 20),
                    message_key="progress.reading_file",
                    message_params={"filename": filename},
                    detail_key="progress.parsing_file_n_of_total",
                    detail_params={"n": file_idx + 1, "total": len(files)},
                )

                # Read transactions from parsed Excel
                transactions = self.statement_reader.read_from_bytes(content, "Automation")
                found_count = len(transactions)
                total_found += found_count

                # Emit progress: file read complete
                emit(
                    "step_complete",
                    "reading",
                    f"Found {found_count} transactions in {filename}",
                    percentage=int((file_idx + 1) / len(files) * 20),
                    data={"filename": filename, "count": found_count},
                    message_key="progress.found_transactions_in_file",
                    message_params={"count": found_count, "filename": filename},
                )

                # Collect file date range before filtering
                file_dates = [tx.date for tx in transactions if tx.date]
                file_date_range = None
                if file_dates:
                    file_date_range = {
                        "start": min(file_dates).isoformat(),
                        "end": max(file_dates).isoformat(),
                    }

                # Filter by date if enabled
                skipped = 0
                if filter_by_date and opening_date and ending_date:
                    transactions, skipped = self.statement_reader.filter_by_date_range(
                        transactions, opening_date, ending_date
                    )

                all_transactions.extend(transactions)
                total_skipped += skipped

                file_result = {
                    "filename": filename,
                    "status": "success",
                    "transactions_found": found_count,
                    "transactions_added": len(transactions),
                    "transactions_skipped": skipped,
                }
                if file_date_range:
                    file_result["file_date_range"] = file_date_range

                file_results.append(file_result)

                # Track in database only when transactions were added
                if self.db_session and len(transactions) > 0:
                    await self._track_uploaded_file(
                        session_id=session_id,
                        filename=filename,
                        file_size=len(content),
                        transactions_count=len(transactions) + skipped,
                        transactions_added=len(transactions),
                        transactions_skipped=skipped,
                    )

            except Exception as e:
                logger.error(f"Error processing file {filename}: {e}")
                file_results.append({
                    "filename": filename,
                    "status": "error",
                    "error": str(e),
                })

        # Emit filtering summary
        if filter_by_date and opening_date and ending_date:
            emit(
                "step_start",
                "filtering",
                f"Filtering {total_found} transactions by date range",
                detail=f"{opening_date.strftime('%d/%m/%Y')} - {ending_date.strftime('%d/%m/%Y')}",
                percentage=25,
                message_key="progress.filtering_by_date",
                message_params={"count": total_found},
                detail_key="progress.date_range",
                detail_params={"start": opening_date.strftime('%d/%m/%Y'), "end": ending_date.strftime('%d/%m/%Y')},
            )
            emit(
                "step_complete",
                "filtering",
                f"{len(all_transactions)} transactions within range, {total_skipped} skipped",
                percentage=30,
                data={"kept": len(all_transactions), "skipped": total_skipped},
                message_key="progress.filter_result",
                message_params={"kept": len(all_transactions), "skipped": total_skipped},
            )

        if not all_transactions:
            # Emit completion even when no transactions
            emit(
                "complete",
                "done",
                "No transactions to process",
                percentage=100,
                message_key="progress.no_transactions",
            )
            # Build a helpful warning message
            if total_skipped > 0 and total_found > 0:
                message = (
                    f"All {total_skipped} transactions were outside the session period "
                    f"({opening_date.strftime('%d/%m/%Y')} - {ending_date.strftime('%d/%m/%Y')}). "
                    f"Please check if the correct period was selected."
                )
                warning = "date_mismatch"
            else:
                message = "No valid transactions found in uploaded files"
                warning = None

            result = {
                "session_id": session_id,
                "files_processed": len(files),
                "total_transactions_added": 0,
                "total_transactions_found": total_found,
                "total_transactions_skipped": total_skipped,
                "file_results": file_results,
                "message": message,
            }
            if warning:
                result["warning"] = warning
                if opening_date and ending_date:
                    result["session_period"] = {
                        "start": opening_date.isoformat(),
                        "end": ending_date.isoformat(),
                    }
            return _finish(result)

        # Classify transactions (rule-based first, AI for leftovers)
        emit(
            "step_start",
            "classifying",
            f"Classifying {len(all_transactions)} transactions (rule-based + AI)...",
            detail="Rule-based keywords first, then AI for unmatched",
            percentage=35,
            message_key="progress.classifying",
            message_params={"count": len(all_transactions)},
            detail_key="progress.classifying_detail",
        )
        if progress_callback:
            await asyncio.sleep(0)  # Flush SSE

        logger.info(f"Classifying {len(all_transactions)} transactions...")
        _classify_start = _time.monotonic()
        classified_transactions = await self._classify_transactions(
            all_transactions,
            progress_callback=progress_callback,
        )
        _classify_elapsed_ms = (_time.monotonic() - _classify_start) * 1000

        # Collect classification stats
        rule_count = sum(1 for tx in classified_transactions if getattr(tx, '_classified_by', '') == 'rule')
        ai_count = sum(1 for tx in classified_transactions if getattr(tx, '_classified_by', '') == 'ai')
        ai_cached_count = sum(1 for tx in classified_transactions if getattr(tx, '_classified_by', '') == 'ai_cached')
        unclassified_count = len(classified_transactions) - rule_count - ai_count - ai_cached_count

        logger.info(f"Classification stats: rule={rule_count}, ai={ai_count}, "
                     f"ai_cached={ai_cached_count}, unclassified={unclassified_count}")

        emit(
            "step_complete",
            "classifying",
            f"Classified {len(classified_transactions)} transactions (rule: {rule_count}, AI: {ai_count}, cached: {ai_cached_count})",
            percentage=80,
            data={"rule_classified": rule_count, "ai_classified": ai_count, "unclassified": unclassified_count},
            message_key="progress.classified_result",
            message_params={"count": len(classified_transactions), "rule": rule_count, "ai": ai_count, "cached": ai_cached_count},
        )
        if progress_callback:
            await asyncio.sleep(0)  # Flush SSE

        # Prepare Movement sheet (copy Cash Balance -> Prior Period + clear old data)
        # Only on first upload (check DB total_transactions, not Excel rows which include template data)
        is_first_upload = await self._is_first_upload(session_id)
        if is_first_upload:
            emit(
                "step_start",
                "preparing",
                "Copying Cash Balance to Prior Period and clearing Movement...",
                percentage=82,
                message_key="progress.preparing_first_upload",
            )
            if progress_callback:
                await asyncio.sleep(0)

            from .openpyxl_handler import get_openpyxl_handler
            handler = get_openpyxl_handler()
            rows_cleared = await asyncio.to_thread(
                handler.prepare_movement_for_writing, working_file
            )

            emit(
                "step_complete",
                "preparing",
                f"Cleared {rows_cleared} old rows, Cash Balance copied",
                percentage=84,
                message_key="progress.prepared_result",
                message_params={"rows_cleared": rows_cleared},
            )

        # Write to Movement sheet
        emit(
            "step_start",
            "writing",
            f"Writing {len(classified_transactions)} transactions to Movement sheet...",
            detail="Appending data to Excel workbook",
            percentage=85,
            message_key="progress.writing",
            message_params={"count": len(classified_transactions)},
            detail_key="progress.writing_detail",
        )
        if progress_callback:
            await asyncio.sleep(0)  # Flush SSE

        writer = MovementDataWriter(working_file)
        # Run blocking Excel write in thread to not block event loop
        rows_added, total_rows = await asyncio.to_thread(
            writer.append_transactions, classified_transactions
        )

        emit(
            "step_complete",
            "writing",
            f"Written {rows_added} rows (total: {total_rows})",
            percentage=95,
            message_key="progress.written_result",
            message_params={"rows_added": rows_added, "total": total_rows},
        )

        # Update session stats in database
        if self.db_session:
            await self._update_session_stats(session_id, rows_added, len(files))

        # Emit completion
        emit(
            "complete",
            "done",
            f"Upload complete! {rows_added} transactions processed.",
            percentage=100,
            data={
                "files_processed": len(files),
                "total_transactions_added": rows_added,
                "total_transactions_skipped": total_skipped,
                "total_rows_in_movement": total_rows,
            },
            message_key="progress.upload_complete",
            message_params={"count": rows_added},
        )

        # Collect AI usage from classifier
        ai_usage = self.ai_classifier.get_and_reset_usage()
        ai_usage["processing_time_ms"] = _classify_elapsed_ms

        final_result = {
            "session_id": session_id,
            "files_processed": len(files),
            "total_transactions_added": rows_added,
            "total_transactions_skipped": total_skipped,
            "total_rows_in_movement": total_rows,
            "file_results": file_results,
            "ai_usage": ai_usage,
            "classification_stats": {
                "rule_based": rule_count,
                "ai": ai_count,
                "ai_cached": ai_cached_count,
                "unclassified": unclassified_count,
            },
        }

        return _finish(final_result)

    # ------------------------------------------------------------------ #
    #  Upload Movement NS/Manual file                                      #
    # ------------------------------------------------------------------ #

    async def upload_movement_file(
        self,
        session_id: str,
        file: Tuple[str, bytes],  # (filename, content)
        filter_by_date: bool = True,
        progress_callback: Optional[callable] = None,
        user_id: Optional[int] = None,
    ) -> Dict[str, Any]:
        """
        Upload Movement data file to a session.

        Supports two file formats (auto-detected):
        - **7-column** (Movement Data): Source, Bank, Account, Date, Description,
          Debit, Credit. Nature is classified via AI automatically.
        - **16-column** (NS/Manual): Pre-classified with Nature already populated.
          No AI classification needed.

        Workflow:
        1. Read Movement file
        2. Auto-detect format: if no Nature values → run AI classification
        3. Filter by session date range (if enabled)
        4. Prepare Movement sheet on first upload (copy Cash Balance → Prior Period)
        5. Append transactions to Movement sheet
        6. Update session stats

        Args:
            session_id: The session ID
            file: Tuple of (filename, file_content)
            filter_by_date: Whether to filter transactions by session date range
            progress_callback: SSE progress callback
            user_id: Owner user ID for access control

        Returns:
            Upload result summary (includes AI usage if classification was needed)
        """
        filename, content = file

        if user_id is not None:
            await self._verify_session_owner(session_id, user_id)

        # Get session info + validate working file exists
        session_info, working_file = await self._ensure_working_file(session_id)

        # Get date range for filtering
        opening_date = None
        ending_date = None
        if filter_by_date and session_info.get("config"):
            config = session_info["config"]
            if config.get("opening_date"):
                date_str = config["opening_date"]
                if "T" in date_str:
                    opening_date = datetime.fromisoformat(date_str).date()
                else:
                    opening_date = date.fromisoformat(date_str)
            if config.get("ending_date"):
                date_str = config["ending_date"]
                if "T" in date_str:
                    ending_date = datetime.fromisoformat(date_str).date()
                else:
                    ending_date = date.fromisoformat(date_str)

        # --- Step 1: Read Movement file ---
        if progress_callback:
            progress_callback(ProgressEvent(
                event_type="step_start",
                step="reading",
                message=f"Reading Movement file: {filename}...",
                percentage=10,
                message_key="progress.reading_movement_file",
                message_params={"filename": filename},
            ))
            await asyncio.sleep(0)

        transactions = self.statement_reader.read_movement_file(content, filename)
        found_count = len(transactions)

        if progress_callback:
            progress_callback(ProgressEvent(
                event_type="step_complete",
                step="reading",
                message=f"Found {found_count} Movement transactions",
                percentage=20,
                data={"count": found_count},
                message_key="progress.found_movement_transactions",
                message_params={"count": found_count},
            ))

        # --- Step 2: Auto-detect format and classify if needed ---
        # If none of the transactions have a Nature value, this is a 7-col
        # Movement Data file → classify via AI (rule-based + Gemini).
        needs_classification = (
            len(transactions) > 0
            and all(not tx.nature for tx in transactions)
        )

        if needs_classification:
            if progress_callback:
                progress_callback(ProgressEvent(
                    event_type="step_start",
                    step="classifying",
                    message=f"7-col format detected — classifying {len(transactions)} transactions via AI...",
                    percentage=25,
                    message_key="progress.classifying_movement_ai",
                    message_params={"count": len(transactions)},
                ))
                await asyncio.sleep(0)

            transactions = await self._classify_transactions(
                transactions, progress_callback=None,
            )

            if progress_callback:
                classified_count = sum(1 for tx in transactions if tx.nature)
                progress_callback(ProgressEvent(
                    event_type="step_complete",
                    step="classifying",
                    message=f"Classified {classified_count}/{len(transactions)} transactions",
                    percentage=40,
                    data={"classified": classified_count},
                    message_key="progress.classified_count",
                    message_params={"classified": classified_count, "total": len(transactions)},
                ))

        # --- Step 3: Filter by date range ---
        total_skipped = 0
        if filter_by_date and opening_date and ending_date:
            if progress_callback:
                progress_callback(ProgressEvent(
                    event_type="step_start",
                    step="filtering",
                    message=f"Filtering by date range {opening_date} - {ending_date}...",
                    percentage=45,
                    message_key="progress.filtering_movement_by_date",
                    message_params={"start": str(opening_date), "end": str(ending_date)},
                ))
            transactions, skipped = self.statement_reader.filter_by_date_range(
                transactions, opening_date, ending_date
            )
            total_skipped = skipped
            if progress_callback:
                progress_callback(ProgressEvent(
                    event_type="step_complete",
                    step="filtering",
                    message=f"{len(transactions)} within range, {skipped} skipped",
                    percentage=55,
                    message_key="progress.filter_result",
                    message_params={"kept": len(transactions), "skipped": skipped},
                ))

        if not transactions:
            if progress_callback:
                progress_callback(ProgressEvent(
                    event_type="complete",
                    step="done",
                    message="No valid Movement transactions to process",
                    percentage=100,
                    message_key="progress.no_movement_transactions",
                ))
            return {
                "session_id": session_id,
                "total_transactions_found": found_count,
                "total_transactions_added": 0,
                "total_transactions_skipped": total_skipped,
                "message": "No valid Movement transactions found after filtering",
            }

        # --- Step 4: Prepare Movement sheet on first upload ---
        is_first_upload = await self._is_first_upload(session_id)
        if is_first_upload:
            if progress_callback:
                progress_callback(ProgressEvent(
                    event_type="step_start",
                    step="preparing",
                    message="First upload: Copying Cash Balance to Prior Period...",
                    percentage=60,
                    message_key="progress.preparing_first_upload",
                ))
                await asyncio.sleep(0)

            from .openpyxl_handler import get_openpyxl_handler
            handler = get_openpyxl_handler()
            rows_cleared = await asyncio.to_thread(
                handler.prepare_movement_for_writing, working_file
            )

            if progress_callback:
                progress_callback(ProgressEvent(
                    event_type="step_complete",
                    step="preparing",
                    message=f"Cleared {rows_cleared} old rows, Cash Balance copied",
                    percentage=70,
                    message_key="progress.prepared_result",
                    message_params={"rows_cleared": rows_cleared},
                ))

        # --- Step 5: Write to Movement sheet ---
        if progress_callback:
            progress_callback(ProgressEvent(
                event_type="step_start",
                step="writing",
                message=f"Writing {len(transactions)} transactions to Movement sheet...",
                percentage=75,
                message_key="progress.writing",
                message_params={"count": len(transactions)},
            ))
            await asyncio.sleep(0)

        writer = MovementDataWriter(working_file)
        rows_added, total_rows = await asyncio.to_thread(
            writer.append_transactions, transactions
        )

        if progress_callback:
            progress_callback(ProgressEvent(
                event_type="step_complete",
                step="writing",
                message=f"Written {rows_added} rows (total: {total_rows})",
                percentage=90,
                message_key="progress.written_result",
                message_params={"rows_added": rows_added, "total": total_rows},
            ))

        # --- Step 6: Update session stats & track file ---
        if self.db_session:
            await self._update_session_stats(session_id, rows_added, 1)
            if rows_added > 0:
                await self._track_uploaded_file(
                    session_id=session_id,
                    filename=filename,
                    file_size=len(content),
                    transactions_count=found_count,
                    transactions_added=rows_added,
                    transactions_skipped=total_skipped,
                    file_type="movement",
                )

        # Emit completion
        if progress_callback:
            progress_callback(ProgressEvent(
                event_type="complete",
                step="done",
                message=f"Upload complete! {rows_added} Movement transactions processed (from {found_count} found)."
                        + (f" AI classified {sum(1 for tx in transactions if tx.nature)}/{len(transactions)}." if needs_classification else ""),
                percentage=100,
                data={
                    "total_transactions_found": found_count,
                    "total_transactions_added": rows_added,
                    "total_transactions_skipped": total_skipped,
                    "total_rows_in_movement": total_rows,
                    "ai_classified": needs_classification,
                },
                message_key="progress.movement_upload_complete",
                message_params={"count": rows_added, "found": found_count},
            ))

        result = {
            "session_id": session_id,
            "total_transactions_found": found_count,
            "total_transactions_added": rows_added,
            "total_transactions_skipped": total_skipped,
            "total_rows_in_movement": total_rows,
            "ai_classified": needs_classification,
            "message": f"Successfully uploaded {rows_added} Movement transactions"
                       + (" (AI classified)" if needs_classification else " (pre-classified)"),
        }
        return result

    # ------------------------------------------------------------------ #

    def _rule_based_classify(self, description: str, is_receipt: bool) -> Optional[str]:
        """
        Rule-based classification using KeyPaymentClassifier.

        Returns:
            Nature category string if confidently matched, None if needs AI.
        """
        key_payment, category, _ = self.rule_classifier.classify(description, is_receipt)
        # Default fallbacks mean no specific keyword matched -> needs AI
        if key_payment in ("Other receipts", "Other payment"):
            return None
        return category

    @staticmethod
    def _normalize_text_for_match(text: str) -> str:
        """Normalize free text for robust case-insensitive keyword matching."""
        normalized = (text or "").lower().replace("_", " ").strip()
        normalized = re.sub(r"[^\w]+", " ", normalized, flags=re.UNICODE)
        return re.sub(r"\s+", " ", normalized).strip()

    @staticmethod
    def _normalize_text_template(text: str) -> str:
        """
        Normalize text into a number-stripped template for pattern matching.

        Strips ALL digits so that descriptions differing only by dates,
        account numbers, invoice numbers, or amounts collapse into one key.
        E.g. "TAT TOAN HDTG RGLH SO 1209/2025/VCBBD-BWTU KY NGAY 12.09.2025 (TK 1060462381)"
           → "tat toan hdtg rglh so vcbbd bwtu ky ngay tk"
        """
        normalized = (text or "").lower().replace("_", " ").strip()
        normalized = re.sub(r"[^\w]+", " ", normalized, flags=re.UNICODE)
        normalized = re.sub(r"\d+", "", normalized)
        normalized = re.sub(r"\s+", " ", normalized).strip()
        return normalized

    @staticmethod
    def _canonical_nature_for_direction(raw_category: str, is_receipt: bool) -> Optional[str]:
        """
        Map external keyword category to canonical Nature for receipt/payment direction.
        Returns None when the category is not valid for the transaction direction.
        """
        category = (raw_category or "").replace("\u00A0", " ").strip()
        if not category:
            return None

        category = TRANSACTION_NATURE_ALIASES.get(category, category)
        category_lower = category.lower()
        if category_lower == "internal transfer":
            return "Internal transfer in" if is_receipt else "Internal transfer out"
        if category_lower == "other payment":
            category = "Operating expense"

        target = ALL_RECEIPT_CATEGORIES if is_receipt else ALL_PAYMENT_CATEGORIES
        for valid_category in target:
            if valid_category.lower() == category.lower():
                return valid_category
        return None

    @staticmethod
    def _normalize_reference_description(text: str) -> str:
        """Normalize bank description for exact reference matching (whitespace-insensitive)."""
        normalized = (text or "").replace("\u00A0", " ").strip().lower()
        return re.sub(r"\s+", " ", normalized)

    @staticmethod
    def _canonical_reference_nature(raw_nature: str) -> Optional[str]:
        """Normalize Nature values from reference CSV to canonical categories."""
        nature = (raw_nature or "").replace("\u00A0", " ").strip()
        if not nature:
            return None
        nature = TRANSACTION_NATURE_ALIASES.get(nature, nature)

        all_categories = ALL_PAYMENT_CATEGORIES | ALL_RECEIPT_CATEGORIES
        if nature in all_categories:
            return nature

        nature_lower = nature.lower()
        for valid_category in all_categories:
            if valid_category.lower() == nature_lower:
                return valid_category
        return None

    @staticmethod
    def _is_receipt_nature(nature: str) -> bool:
        """True if Nature belongs to receipt direction."""
        return nature in ALL_RECEIPT_CATEGORIES

    @staticmethod
    def _choose_majority_nature(counter: Counter) -> Optional[str]:
        """
        Resolve ambiguous reference labels by majority vote.
        Returns None when no clear winner exists.
        """
        if not counter:
            return None
        return counter.most_common(1)[0][0]

    def _rebuild_transactions_reference_index(self) -> None:
        """
        Build exact, fuzzy, and template description->nature lookup tables from Transactions.csv.
        Direction is derived from nature (receipt/payment) to avoid cross-direction leakage.

        Three normalization levels (each broader than the previous):
        - exact: whitespace-normalized, keeps punctuation and digits
        - fuzzy: strips punctuation/underscores, keeps digits
        - template: strips ALL digits so recurring descriptions with different
          dates/account-numbers/amounts collapse into one key
        """
        path = TRANSACTIONS_REFERENCE_FILE
        self._transactions_reference_exact = {}
        self._transactions_reference_majority = {}
        self._transactions_reference_fuzzy_exact = {}
        self._transactions_reference_fuzzy_majority = {}
        self._transactions_reference_template_exact = {}
        self._transactions_reference_template_majority = {}
        self._transactions_reference_mtime = None

        if not path.exists():
            logger.warning(f"Transactions reference file not found: {path}")
            return

        exact_counters: Dict[Tuple[str, bool], Counter] = defaultdict(Counter)
        fuzzy_counters: Dict[Tuple[str, bool], Counter] = defaultdict(Counter)
        template_counters: Dict[Tuple[str, bool], Counter] = defaultdict(Counter)
        total_rows = 0
        skipped_rows = 0

        try:
            with path.open("r", encoding="utf-8-sig", newline="") as f:
                reader = csv.reader(f, delimiter=";")
                next(reader, None)  # header

                for row in reader:
                    if not row or len(row) < 2:
                        continue

                    desc_raw = row[0]
                    # File format: desc;debit;credit;nature (nature is LAST column)
                    nature_raw = row[-1]
                    nature = self._canonical_reference_nature(nature_raw)
                    if not nature:
                        skipped_rows += 1
                        continue

                    is_receipt = self._is_receipt_nature(nature)
                    desc_exact = self._normalize_reference_description(desc_raw)
                    desc_fuzzy = self._normalize_text_for_match(desc_raw)
                    desc_template = self._normalize_text_template(desc_raw)
                    if not desc_exact and not desc_fuzzy:
                        skipped_rows += 1
                        continue

                    if desc_exact:
                        exact_counters[(desc_exact, is_receipt)][nature] += 1
                    if desc_fuzzy:
                        fuzzy_counters[(desc_fuzzy, is_receipt)][nature] += 1
                    if desc_template:
                        template_counters[(desc_template, is_receipt)][nature] += 1
                    total_rows += 1
        except Exception as e:
            logger.warning(f"Failed to load Transactions.csv reference data: {e}")
            return

        exact_ambiguous = 0
        for key, counts in exact_counters.items():
            if len(counts) == 1:
                self._transactions_reference_exact[key] = counts.most_common(1)[0][0]
                continue
            exact_ambiguous += 1
            majority_nature = self._choose_majority_nature(counts)
            if majority_nature:
                self._transactions_reference_majority[key] = majority_nature

        fuzzy_ambiguous = 0
        for key, counts in fuzzy_counters.items():
            if len(counts) == 1:
                self._transactions_reference_fuzzy_exact[key] = counts.most_common(1)[0][0]
                continue
            fuzzy_ambiguous += 1
            majority_nature = self._choose_majority_nature(counts)
            if majority_nature:
                self._transactions_reference_fuzzy_majority[key] = majority_nature

        template_ambiguous = 0
        for key, counts in template_counters.items():
            if len(counts) == 1:
                self._transactions_reference_template_exact[key] = counts.most_common(1)[0][0]
                continue
            template_ambiguous += 1
            majority_nature = self._choose_majority_nature(counts)
            if majority_nature:
                self._transactions_reference_template_majority[key] = majority_nature

        try:
            self._transactions_reference_mtime = path.stat().st_mtime
        except OSError:
            self._transactions_reference_mtime = None

        logger.info(
            "Transactions reference loaded: rows=%s, exact=%s, exact_majority=%s, "
            "fuzzy=%s, fuzzy_majority=%s, template=%s, template_majority=%s, "
            "ambiguous_exact=%s, ambiguous_fuzzy=%s, ambiguous_template=%s, skipped=%s",
            total_rows,
            len(self._transactions_reference_exact),
            len(self._transactions_reference_majority),
            len(self._transactions_reference_fuzzy_exact),
            len(self._transactions_reference_fuzzy_majority),
            len(self._transactions_reference_template_exact),
            len(self._transactions_reference_template_majority),
            exact_ambiguous,
            fuzzy_ambiguous,
            template_ambiguous,
            skipped_rows,
        )

    def _get_transactions_reference_mtime(self) -> Optional[float]:
        """Get modification time for Transactions.csv reference file."""
        try:
            return TRANSACTIONS_REFERENCE_FILE.stat().st_mtime
        except OSError:
            return None

    def _refresh_transactions_reference_index_if_needed(self) -> None:
        """Reload Transactions.csv reference when file content changes."""
        current_mtime = self._get_transactions_reference_mtime()
        if current_mtime is None or current_mtime == self._transactions_reference_mtime:
            return
        self._rebuild_transactions_reference_index()

    def _rebuild_tfidf_index(self) -> None:
        """Fit the TF-IDF classifier from Transactions.csv."""
        path = TRANSACTIONS_REFERENCE_FILE
        count = self._tfidf_classifier.load_from_csv(path)
        try:
            self._tfidf_mtime = path.stat().st_mtime if path.exists() else None
        except OSError:
            self._tfidf_mtime = None
        logger.info("TF-IDF index rebuilt: %d corpus entries", count)

    def _refresh_tfidf_index_if_needed(self) -> None:
        """Reload TF-IDF index when Transactions.csv changes."""
        current_mtime = self._get_transactions_reference_mtime()
        if current_mtime is None or current_mtime == self._tfidf_mtime:
            return
        self._rebuild_tfidf_index()

    def _classify_from_tfidf(self, tx: MovementTransaction) -> Optional[str]:
        """
        Classify using TF-IDF similarity against Transactions.csv corpus.

        Business rules enforced BEFORE TF-IDF similarity:
        - Receipt + settlement/deposit keywords + debit >= 1B → "Internal transfer in"
        - Receipt + settlement regex patterns → "Internal transfer in"
        - Payment + open-new regex patterns → "Internal transfer out"

        Returns Nature string if confident match, else None.
        """
        is_receipt = bool(tx.debit)
        description = (tx.description or "").strip()

        # ── Hard business rules: SHL/BCC repayment, LOAN AGREEMENT, KLHT ──
        # Must be checked BEFORE settlement patterns to avoid false override.
        desc_lower_tfidf = description.lower()
        _shl_bcc_kws = (
            "repayment of shl", "shl repayment", "shl interest",
            "lai vay shl", "tra mot phan lai vay",
            "repayment bcc", "tra lai bcc",
        )
        if is_receipt and any(kw in desc_lower_tfidf for kw in _shl_bcc_kws):
            return "Loan receipts"
        if is_receipt and "loan agreement" in desc_lower_tfidf:
            return "Loan receipts"
        if "klht" in desc_lower_tfidf:
            return "Refinancing" if is_receipt else "Construction expense"

        # ── Settlement patterns: roundness + amount-based nature ──
        # >= 1B (any shape)  → Internal transfer in (settlement flow separates principal/interest)
        # < 1B + round (÷100M) → Internal transfer in (small round deposit)
        # < 1B + non-round:
        #   >= 100M          → Receipt from tenants (interest portion)
        #   < 100M           → Other receipts (interest portion)
        # NOTE: pre-split interest (non-round >= 1B with companion round tx on same
        #       account+date) is reclassified in _reclassify_presplit_interest().
        if is_receipt and any(p.search(description) for p in self._settlement_patterns_compiled):
            amount = tx.debit or Decimal("0")
            return self._classify_settlement_receipt_nature(
                description=description,
                amount=amount,
                account=tx.account,
            )

        # ── HDTG/deposit keywords + debit >= 1B → Internal transfer in ──
        if is_receipt and (tx.debit or Decimal("0")) >= _UNIT_1B:
            desc_lower = description.lower()
            if any(kw in desc_lower for kw in HDTG_DEPOSIT_KEYWORDS):
                return "Internal transfer in"

        # ── Open-new patterns for payments → Internal transfer out ──
        if not is_receipt:
            if any(p.search(description) for p in self._open_new_patterns_compiled):
                return "Internal transfer out"

        # ── TF-IDF similarity fallback ──
        if not self._tfidf_classifier.is_fitted:
            return None
        return self._tfidf_classifier.predict(description, is_receipt)

    def _classify_settlement_receipt_nature(
        self,
        description: str,
        amount: Decimal,
        account: Optional[str] = None,
    ) -> str:
        """
        Amount-based nature for receipt-side settlement patterns.
        Supports targeted non-round exceptions to avoid false Internal transfer in.
        """
        _UNIT_100M = Decimal("100000000")
        amount = amount or Decimal("0")
        is_round = amount >= _UNIT_100M and amount % _UNIT_100M == 0
        desc_norm = self._normalize_text_for_match(description or "")

        # Align with keyword CSV SUB_Category:
        # [Receipt] CA - TARGET + amount < 1B -> Other receipts.
        if CA_TARGET_NORMALIZED in desc_norm and amount < _UNIT_1B:
            return "Other receipts"

        if not is_round:
            # CA - TARGET:
            # - non-round >= 1B => Receipt from tenants
            # - non-round < 1B  => Other receipts (handled above)
            if CA_TARGET_NORMALIZED in desc_norm:
                return "Receipt from tenants" if amount >= _UNIT_1B else "Other receipts"

            # BIDV iBank withdrawal for known account 2221133456 behaves as
            # receipt-side interest on non-round amounts.
            account_norm = re.sub(r"\D+", "", str(account or ""))
            if (
                BIDV_IBANK_WITHDRAWAL_NORMALIZED in desc_norm
                and account_norm in BIDV_IBANK_WITHDRAWAL_RECEIPT_ACCOUNTS
            ):
                return "Receipt from tenants"

        if amount >= _UNIT_1B:
            return "Internal transfer in"
        if is_round:
            return "Internal transfer in"
        if amount >= _UNIT_100M:
            return "Receipt from tenants"
        return "Other receipts"

    def _classify_from_transactions_reference(self, tx: MovementTransaction) -> Tuple[Optional[str], Optional[str]]:
        """
        Classify by ground-truth Transactions.csv.
        Returns tuple (nature, source_tag).

        Lookup order (first match wins):
        1. Exact normalized match (whitespace-insensitive)
        2. Fuzzy match (strips punctuation, keeps digits)
        3. Template match (strips ALL digits — catches recurring descriptions
           with different dates/account-numbers/amounts)
        """
        is_receipt = bool(tx.debit)
        description = tx.description or ""

        desc_exact = self._normalize_reference_description(description)
        if desc_exact:
            key = (desc_exact, is_receipt)
            if key in self._transactions_reference_exact:
                return self._transactions_reference_exact[key], "reference_exact"
            if key in self._transactions_reference_majority:
                return self._transactions_reference_majority[key], "reference_majority"

        desc_fuzzy = self._normalize_text_for_match(description)
        if desc_fuzzy:
            key = (desc_fuzzy, is_receipt)
            if key in self._transactions_reference_fuzzy_exact:
                return self._transactions_reference_fuzzy_exact[key], "reference_fuzzy_exact"
            if key in self._transactions_reference_fuzzy_majority:
                return self._transactions_reference_fuzzy_majority[key], "reference_fuzzy_majority"

        desc_template = self._normalize_text_template(description)
        if desc_template:
            key = (desc_template, is_receipt)
            if key in self._transactions_reference_template_exact:
                return self._transactions_reference_template_exact[key], "reference_template"
            if key in self._transactions_reference_template_majority:
                return self._transactions_reference_template_majority[key], "reference_template_majority"

        return None, None

    @staticmethod
    def _extract_external_rule_column(
        row: Dict[str, Any],
        *column_names: str,
    ) -> str:
        """Read CSV column by candidate names (case-insensitive)."""
        if not row:
            return ""
        lowered = {
            (str(key).strip().lower() if key is not None else ""): (value or "")
            for key, value in row.items()
        }
        for name in column_names:
            value = lowered.get(name.strip().lower(), "")
            if str(value).strip():
                return str(value).strip()
        return ""

    @staticmethod
    def _parse_external_rule_type(raw_type: str) -> Optional[bool]:
        """
        Parse rule Type column to cash-flow direction.
        Returns:
            True  -> Receipt
            False -> Payment
            None  -> no direction constraint
        """
        type_norm = (raw_type or "").strip().lower()
        if type_norm == "receipt":
            return True
        if type_norm == "payment":
            return False
        return None

    @staticmethod
    def _parse_amount_condition(raw_condition: str) -> Tuple[Optional[str], Optional[Decimal]]:
        """
        Parse SUB_Category amount condition.
        Examples:
            "> 1.000.000.000" -> (">", Decimal("1000000000"))
            "< 500.000.000"   -> ("<", Decimal("500000000"))
        """
        text = (raw_condition or "").strip()
        if not text:
            return None, None

        matched = re.match(r"^(<=|>=|<|>)\s*([0-9][0-9\.,\s]*)$", text)
        if not matched:
            return None, None

        operator = matched.group(1)
        digits = re.sub(r"\D+", "", matched.group(2))
        if not digits:
            return None, None
        return operator, Decimal(digits)

    @staticmethod
    def _amount_condition_matches(
        amount: Decimal,
        operator: Optional[str],
        threshold: Optional[Decimal],
    ) -> bool:
        """Evaluate amount against parsed SUB_Category condition."""
        if operator is None or threshold is None:
            return True
        if operator == ">":
            return amount > threshold
        if operator == ">=":
            return amount >= threshold
        if operator == "<":
            return amount < threshold
        if operator == "<=":
            return amount <= threshold
        return True

    def _build_external_keyword_index(self) -> List[ExternalKeywordRule]:
        """
        Build keyword-rule index from movement_nature_filter/Key Words CSV.csv.
        Supports Type + SUB_Category amount conditions for deterministic matching.
        """
        rules_file = getattr(self.ai_classifier, "_rules_file", None)
        if not rules_file:
            return []

        path = Path(rules_file)
        if not path.exists():
            logger.warning("Keyword rules file not found: %s", path)
            return []

        index: List[ExternalKeywordRule] = []
        try:
            with path.open("r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f, delimiter=";")
                for row in reader:
                    category = self._extract_external_rule_column(
                        row,
                        "Category Cash consol",
                        "Category",
                    )
                    keyword = self._extract_external_rule_column(
                        row,
                        "Key word",
                        "Keyword",
                    )
                    if not category or not keyword:
                        continue

                    keyword_norm = self._normalize_text_for_match(keyword)
                    if not keyword_norm:
                        continue

                    is_receipt = self._parse_external_rule_type(
                        self._extract_external_rule_column(row, "Type")
                    )
                    operator, threshold = self._parse_amount_condition(
                        self._extract_external_rule_column(row, "SUB_Category")
                    )

                    index.append(
                        ExternalKeywordRule(
                            keyword_norm=keyword_norm,
                            category=category,
                            is_receipt=is_receipt,
                            amount_operator=operator,
                            amount_threshold=threshold,
                        )
                    )
        except Exception as e:
            logger.warning("Failed to load keyword rules from %s: %s", path, e)
            return []

        # Priority: longer keyword first, then direction-specific, then amount-constrained.
        index.sort(
            key=lambda rule: (
                len(rule.keyword_norm),
                1 if rule.is_receipt is not None else 0,
                1 if rule.amount_operator else 0,
            ),
            reverse=True,
        )
        return index

    def _get_rules_file_mtime(self) -> Optional[float]:
        """Get current modification time for keyword-rules CSV used by AI classifier."""
        rules_file = getattr(self.ai_classifier, "_rules_file", None)
        if not rules_file:
            return None
        try:
            return Path(rules_file).stat().st_mtime
        except OSError:
            return None

    def _refresh_external_keyword_index_if_needed(self) -> None:
        """
        Reload keyword rules when keyword CSV changes and rebuild the matching index.
        This keeps the flow aligned with latest CSV without restarting the service.
        """
        current_mtime = self._get_rules_file_mtime()
        if current_mtime is None or current_mtime == self._rules_file_mtime:
            return

        try:
            self.ai_classifier.reload_rules()
        except Exception as e:
            logger.warning(f"Failed to reload keyword rules file: {e}")
            return

        self._external_keyword_index = self._build_external_keyword_index()
        self._rules_file_mtime = current_mtime
        logger.info(
            f"Reloaded keyword rules and rebuilt index: {len(self._external_keyword_index)} keyword patterns"
        )

    @staticmethod
    def _normalize_regex_pattern(pattern: str) -> str:
        """Normalize regex text for dedupe/exclusion checks."""
        normalized = (pattern or "").replace("\u00A0", " ").strip()
        normalized = re.sub(r"\s+", " ", normalized)
        return normalized.lower()

    def _load_detection_patterns_from_csv(
        self,
        file_path: Path,
        fallback_patterns: List[str],
        flow_name: str,
        excluded_patterns: Optional[set] = None,
    ) -> List[Any]:
        """
        Load regex patterns from CSV file and merge with fallback patterns in code.
        CSV rows are expected to use semicolon separator with a `Regex` column.
        """
        csv_patterns: List[str] = []

        if file_path.exists():
            try:
                with file_path.open("r", encoding="utf-8-sig", newline="") as f:
                    reader = csv.DictReader(f, delimiter=";")
                    for row in reader:
                        regex_value = ((row or {}).get("Regex") or (row or {}).get("regex") or "").strip()
                        if regex_value:
                            csv_patterns.append(regex_value)
            except Exception as e:
                logger.warning(f"Failed to load {flow_name} patterns from {file_path}: {e}")
        else:
            logger.warning(f"{flow_name} pattern file not found: {file_path}")

        merged_patterns: List[str] = []
        seen_norm = set()

        def _append_unique(pattern: str) -> None:
            norm = self._normalize_regex_pattern(pattern)
            if not norm or norm in seen_norm:
                return
            seen_norm.add(norm)
            merged_patterns.append(pattern.strip())

        for pattern in csv_patterns:
            _append_unique(pattern)
        for pattern in fallback_patterns:
            _append_unique(pattern)

        excluded_norm = {
            self._normalize_regex_pattern(pattern)
            for pattern in (excluded_patterns or set())
            if pattern
        }
        if excluded_norm:
            merged_patterns = [
                pattern for pattern in merged_patterns
                if self._normalize_regex_pattern(pattern) not in excluded_norm
            ]

        compiled_patterns: List[Any] = []
        for pattern in merged_patterns:
            try:
                compiled_patterns.append(_re.compile(pattern, _re.IGNORECASE))
            except _re.error as e:
                logger.warning(f"Invalid {flow_name} regex '{pattern}' in {file_path}: {e}")

        if not compiled_patterns:
            for pattern in fallback_patterns:
                norm = self._normalize_regex_pattern(pattern)
                if norm in excluded_norm:
                    continue
                try:
                    compiled_patterns.append(_re.compile(pattern, _re.IGNORECASE))
                except _re.error:
                    continue

        logger.info(
            "%s patterns loaded: compiled=%s, from_csv=%s, merged_total=%s, file=%s",
            flow_name,
            len(compiled_patterns),
            len(csv_patterns),
            len(merged_patterns),
            file_path.name,
        )
        return compiled_patterns

    @staticmethod
    def _get_file_mtime(path: Path) -> Optional[float]:
        """Get file modification time, or None when missing/unreadable."""
        try:
            return path.stat().st_mtime
        except OSError:
            return None

    def _reload_detection_pattern_indexes(self) -> None:
        """Reload settlement/open-new detection regex from CSV and merge with in-code fallbacks."""
        self._settlement_patterns_compiled = self._load_detection_patterns_from_csv(
            file_path=SETTLEMENT_REFERENCE_FILE,
            fallback_patterns=SETTLEMENT_PATTERNS,
            flow_name="settlement",
        )
        self._open_new_patterns_compiled = self._load_detection_patterns_from_csv(
            file_path=OPEN_NEW_REFERENCE_FILE,
            fallback_patterns=OPEN_NEW_PATTERNS,
            flow_name="open_new",
            excluded_patterns=OPEN_NEW_EXCLUDED_REGEXES,
        )
        self._settlement_rules_mtime = self._get_file_mtime(SETTLEMENT_REFERENCE_FILE)
        self._open_new_rules_mtime = self._get_file_mtime(OPEN_NEW_REFERENCE_FILE)

    def _refresh_detection_pattern_indexes_if_needed(self) -> None:
        """Reload settlement/open-new detection regex when CSV files change."""
        settlement_mtime = self._get_file_mtime(SETTLEMENT_REFERENCE_FILE)
        open_new_mtime = self._get_file_mtime(OPEN_NEW_REFERENCE_FILE)
        if (
            settlement_mtime == self._settlement_rules_mtime
            and open_new_mtime == self._open_new_rules_mtime
        ):
            return
        self._reload_detection_pattern_indexes()

    @staticmethod
    def _keyword_matches_description(
        description_norm: str,
        description_tokens: set,
        keyword_norm: str,
    ) -> bool:
        """Match keyword in normalized description with boundary-safe handling for short tokens."""
        if not keyword_norm:
            return False
        if len(keyword_norm) <= 3 and " " not in keyword_norm:
            return keyword_norm in description_tokens
        return keyword_norm in description_norm

    def _classify_from_external_keyword_rules(
        self,
        description: str,
        is_receipt: bool,
        amount: Optional[Decimal],
    ) -> Optional[str]:
        """
        Deterministic classification using external keyword CSV rules.
        Returns canonical Nature for direction or None if no confident match.
        """
        if not self._external_keyword_index:
            return None

        description_norm = self._normalize_text_for_match(description or "")
        if not description_norm:
            return None
        description_tokens = set(description_norm.split())
        tx_amount = amount or Decimal("0")

        for rule in self._external_keyword_index:
            if rule.is_receipt is not None and rule.is_receipt != is_receipt:
                continue
            if not self._keyword_matches_description(description_norm, description_tokens, rule.keyword_norm):
                continue
            if not self._amount_condition_matches(tx_amount, rule.amount_operator, rule.amount_threshold):
                continue

            mapped = self._canonical_nature_for_direction(rule.category, is_receipt)
            if mapped:
                return mapped
        return None

    def _apply_classification_guardrails(self, tx: MovementTransaction) -> None:
        """
        Post-classification guardrails to prevent obvious misclassifications.
        - Numeric account-like descriptions follow deterministic round/non-round rule.
        - Settlement/open-new patterns enforce transfer directions.
        - Receipt amount constraints:
          * Round receipts => Internal transfer in
          * Other receipts must be non-round and < 100M
        - Empty nature always falls back by cash direction.
        """
        is_receipt = bool(tx.debit)
        description = (tx.description or "").strip()
        nature_now = (tx.nature or "").strip()

        special_nature = self._classify_special_bank_description(tx)
        if special_nature:
            tx.nature = special_nature
            tx._classified_by = "rule_guardrail"
            return

        # Don't override ground-truth classifications from Transactions.csv reference
        classified_by = getattr(tx, "_classified_by", "") or ""
        is_reference = classified_by.startswith("reference")

        desc_lower = description.lower()

        if is_receipt:
            # ── Hard business rules: SHL/BCC repayment, LOAN AGREEMENT, KLHT ──
            # These override ALL other classification (including settlement patterns)
            # because they are deterministic business rules, not heuristics.
            _shl_bcc_repay_kws = (
                "repayment of shl", "shl repayment", "shl interest",
                "lai vay shl", "tra mot phan lai vay",
                "repayment bcc", "tra lai bcc",
            )
            if any(kw in desc_lower for kw in _shl_bcc_repay_kws):
                tx.nature = "Loan receipts"
                tx._classified_by = "rule_guardrail"
                return

            if "loan agreement" in desc_lower:
                tx.nature = "Loan receipts"
                tx._classified_by = "rule_guardrail"
                return

            if "klht" in desc_lower:
                tx.nature = "Refinancing"
                tx._classified_by = "rule_guardrail"
                return

            # Settlement patterns (from Settlement.csv) — explicit curated rules, always override.
            # Roundness + amount-based nature:
            #   >= 1B (any shape)  → Internal transfer in (settlement flow separates)
            #   < 1B + round (÷100M) → Internal transfer in (small round deposit)
            #   < 1B + non-round:
            #     >= 100M          → Receipt from tenants (interest, NOT settlement)
            #     < 100M           → Other receipts (interest, NOT settlement)
            # NOTE: pre-split interest reclassified in _reclassify_presplit_interest().
            if any(pattern.search(description) for pattern in self._settlement_patterns_compiled):
                amount = tx.debit or Decimal("0")
                tx.nature = self._classify_settlement_receipt_nature(
                    description=description,
                    amount=amount,
                    account=tx.account,
                )
                tx._classified_by = "rule_guardrail"
                return

            # Broad HDTG/deposit guardrail: receipt with deposit-related keywords
            # AND debit > 1B VND → Internal transfer in (settlement candidate).
            # Catches cases not covered by specific settlement regex patterns,
            # e.g. "GUI HDTG SO ...", "HOP DONG TIEN GUI ...", "SAVING ACCOUNT ..."
            if (tx.debit or Decimal("0")) >= _UNIT_1B:
                if any(kw in desc_lower for kw in HDTG_DEPOSIT_KEYWORDS):
                    tx.nature = "Internal transfer in"
                    tx._classified_by = "rule_guardrail"
                    return
        else:
            # ── Hard business rule: KLHT payment → Construction expense ──
            if "klht" in desc_lower:
                tx.nature = "Construction expense"
                tx._classified_by = "rule_guardrail"
                return

            # Intercompany profit-distribution payments should be transfer-out.
            if (
                self._has_dual_entity_transfer(desc_lower)
                and self._is_profit_distribution_transfer(description)
            ):
                tx.nature = "Internal transfer out"
                tx._classified_by = "rule_guardrail"
                return

            # Open-new patterns (from OpenNew.csv) — explicit curated rules, always override
            if any(pattern.search(description) for pattern in self._open_new_patterns_compiled):
                tx.nature = "Internal transfer out"
                tx._classified_by = "rule_guardrail"
                return

        if not nature_now:
            tx.nature = "Receipt from tenants" if is_receipt else "Operating expense"
            tx._classified_by = "fallback"
            nature_now = tx.nature

        # Don't override ground-truth reference classifications with generic amount rules
        # EXCEPT for very large round receipts (>= 1B) — at that scale the round amount
        # is an overwhelming signal for Internal transfer in, stronger than any reference.
        if is_reference:
            if is_receipt and (tx.debit or Decimal("0")) >= _UNIT_1B:
                unit = Decimal("100000000")
                if tx.debit >= unit and tx.debit % unit == 0:
                    tx.nature = "Internal transfer in"
                    tx._classified_by = "rule_guardrail"
            return

        adjusted_nature = self._apply_receipt_amount_nature_constraints(tx, nature_now)
        if adjusted_nature and adjusted_nature != nature_now:
            tx.nature = adjusted_nature
            tx._classified_by = "rule_guardrail"

    @staticmethod
    def _is_round_receipt_amount(amount: Decimal) -> bool:
        """
        Treat round receipts as internal transfer amounts.
        Round = >= 100M and divisible by 100M.
        """
        if not amount or amount <= 0:
            return False
        unit = Decimal("100000000")
        return amount >= unit and amount % unit == 0

    def _apply_receipt_amount_nature_constraints(self, tx: MovementTransaction, current_nature: str) -> Optional[str]:
        """
        Additional user-defined constraints for receipt-side natures:
        - Round (>=100M, divisible by 100M) → Internal transfer in
        - Non-round + >=100M                → Receipt from tenants
        - Internal transfer in + <100M      → Other receipts
        """
        if not tx.debit or tx.debit <= 0:
            return None

        # Keep explicit numeric-account rule behavior unchanged.
        desc_norm = self._normalize_bank_description(tx.description or "")
        if self._is_numeric_account_like_description(desc_norm):
            return None

        nature = (current_nature or "").strip().lower()
        if nature not in {"receipt from tenants", "other receipts", "internal transfer in"}:
            return None

        amount = tx.debit
        UNIT_100M = Decimal("100000000")
        is_round = amount >= UNIT_100M and amount % UNIT_100M == 0

        # Rule 1: Round → Internal transfer in
        if is_round:
            return "Internal transfer in"

        # Below: non-round

        # Rule 2: >=100M → Receipt from tenants
        if amount >= UNIT_100M:
            return "Receipt from tenants"

        # Rule 3: Internal transfer in + <100M → Other receipts
        if nature == "internal transfer in":
            return "Other receipts"

        return None

    @staticmethod
    def _normalize_bank_description(description: str) -> str:
        """Normalize numeric bank descriptions that may come as '<digits>.0'."""
        return re.sub(r"\.0+$", "", (description or "").strip())

    @staticmethod
    def _is_numeric_account_like_description(desc_norm: str) -> bool:
        """True for account-like numeric descriptions (10-20 digits)."""
        return bool(NUMERIC_ACCOUNT_DESC_RE.fullmatch(desc_norm or ""))

    def _classify_special_bank_description(self, tx: MovementTransaction) -> Optional[str]:
        """
        Special rules for account-like numeric bank descriptions.
        - round amount (>= 100M and divisible by 100M) => Internal transfer in
        - non-round amount => Other receipts
        """
        desc_norm = self._normalize_bank_description(tx.description or "")
        if not self._is_numeric_account_like_description(desc_norm):
            return None
        if not tx.debit or tx.debit <= 0:
            return None

        unit = Decimal("100000000")
        if tx.debit >= unit and tx.debit % unit == 0:
            return "Internal transfer in"
        return "Other receipts"

    async def _classify_transactions(
        self,
        transactions: List[MovementTransaction],
        progress_callback: Optional[callable] = None,
    ) -> List[MovementTransaction]:
        """
        Classify transactions using hybrid approach:
        Phase 1: Rule-based (keyword matching) - fast, deterministic
        Phase 2: AI (Gemini) - only for transactions rule-based couldn't classify

        Args:
            transactions: List of transactions without Nature
            progress_callback: Optional callback for progress updates

        Returns:
            List of transactions with Nature classified and classification_stats dict
        """
        if not transactions:
            return []

        self._refresh_detection_pattern_indexes_if_needed()
        self._refresh_external_keyword_index_if_needed()
        self._refresh_transactions_reference_index_if_needed()
        self._refresh_tfidf_index_if_needed()

        # â"€â"€ Phase 1: Rule-based classification â"€â"€
        needs_ai_indices = []
        rule_classified = 0

        for i, tx in enumerate(transactions):
            reference_nature, reference_source = self._classify_from_transactions_reference(tx)
            if reference_nature:
                tx.nature = reference_nature
                tx._classified_by = reference_source or "reference"
                rule_classified += 1
                continue

            special_nature = self._classify_special_bank_description(tx)
            if special_nature:
                tx.nature = special_nature
                tx._classified_by = "rule"
                rule_classified += 1
                continue

            is_receipt = bool(tx.debit)  # Debit = money in = Receipt
            tx_amount = tx.debit if is_receipt else tx.credit
            external_rule_nature = self._classify_from_external_keyword_rules(
                tx.description or "",
                is_receipt,
                tx_amount,
            )
            if external_rule_nature:
                tx.nature = external_rule_nature
                tx._classified_by = "rule"
                rule_classified += 1
                continue

            # TF-IDF similarity against Transactions.csv reference corpus
            tfidf_nature = self._classify_from_tfidf(tx)
            if tfidf_nature:
                tx.nature = tfidf_nature
                tx._classified_by = "tfidf"
                rule_classified += 1
                continue

            nature = self._rule_based_classify(tx.description or "", is_receipt)
            if nature:
                tx.nature = nature
                tx._classified_by = "rule"
                rule_classified += 1
            else:
                needs_ai_indices.append(i)

        logger.info(f"Rule-based classified {rule_classified}/{len(transactions)}, "
                     f"{len(needs_ai_indices)} need AI")

        if progress_callback:
            progress_callback(ProgressEvent(
                event_type="step_update",
                step="classifying",
                message=f"Rule-based: {rule_classified} classified, {len(needs_ai_indices)} need AI...",
                percentage=40,
                message_key="progress.rule_based_result",
                message_params={"classified": rule_classified, "need_ai": len(needs_ai_indices)},
            ))
            await asyncio.sleep(0)

        # â"€â"€ Phase 2: AI classification for remaining â"€â"€
        if not needs_ai_indices:
            logger.info("All transactions classified by rules, no AI needed")
            for tx in transactions:
                self._apply_classification_guardrails(tx)
            return transactions

        if not getattr(self.ai_classifier, "client", None):
            logger.warning("AI classifier is unavailable; applying deterministic fallback for remaining transactions")
            for tx in transactions:
                self._apply_classification_guardrails(tx)
            return transactions

        # Prepare batch for AI (only unclassified)
        ai_batch = []
        for idx in needs_ai_indices:
            tx = transactions[idx]
            is_receipt = bool(tx.debit)
            ai_batch.append((tx.description, is_receipt))
        # Classify with per-batch progress updates
        try:
            batch_size = 50
            all_natures = []
            total_batches = (len(ai_batch) + batch_size - 1) // batch_size

            for i in range(0, len(ai_batch), batch_size):
                chunk = ai_batch[i:i + batch_size]
                batch_num = i // batch_size + 1

                if progress_callback:
                    pct = 45 + int((batch_num / total_batches) * 35)  # 45-80%
                    progress_callback(ProgressEvent(
                        event_type="step_update",
                        step="classifying",
                        message=f"AI batch {batch_num}/{total_batches} ({len(chunk)} transactions)...",
                        percentage=min(pct, 79),
                        message_key="progress.ai_batch",
                        message_params={"batch": batch_num, "total": total_batches, "count": len(chunk)},
                    ))
                    await asyncio.sleep(0)

                batch_natures = await asyncio.to_thread(
                    self.ai_classifier._classify_batch_internal, chunk
                )
                all_natures.extend(batch_natures)
                logger.info(f"AI classified batch {batch_num}/{total_batches}: {len(chunk)} transactions")

            # Apply AI classifications to the correct transactions
            ai_classified = 0
            ai_cached = 0
            for j, (nature, was_cached) in enumerate(all_natures):
                if j < len(needs_ai_indices):
                    idx = needs_ai_indices[j]
                    transactions[idx].nature = nature
                    transactions[idx]._classified_by = "ai_cached" if was_cached else "ai"
                    if nature:
                        ai_classified += 1
                        if was_cached:
                            ai_cached += 1

            logger.info(f"AI classified {ai_classified}/{len(needs_ai_indices)} "
                         f"(cached: {ai_cached}, API calls: {ai_classified - ai_cached})")

        except Exception as e:
            logger.error(f"AI classification error: {e}")
            # Leave nature empty for failed AI classifications

        for tx in transactions:
            self._apply_classification_guardrails(tx)

        # ── Phase 3: Detect pre-split interest rows ──
        # When the bank already split principal/interest into separate rows,
        # the non-round row is interest and should NOT be "Internal transfer in".
        # Detected by grouping settlement transactions by (account, date):
        # if a group has BOTH round and non-round amounts, the non-round ones
        # are reclassified as interest.
        self._reclassify_presplit_interest(transactions)

        return transactions

    def _reclassify_presplit_interest(self, transactions: List[MovementTransaction]) -> None:
        """
        Detect pre-split interest rows among settlement-classified transactions.

        When the bank already separated principal + interest into two rows
        (same account, same date, same settlement keyword), the non-round row
        is interest — reclassify it:
          - >= 100M → Receipt from tenants
          - < 100M  → Other receipts

        Only reclassifies when a companion ROUND transaction exists in the
        same (account, date) group, proving the bank already split them.
        """
        _UNIT_100M = Decimal("100000000")

        # Collect settlement "Internal transfer in" receipts with their indices
        settlement_indices: Dict[Tuple[str, Optional[date]], List[int]] = defaultdict(list)
        for i, tx in enumerate(transactions):
            if not tx.debit or tx.debit <= 0:
                continue
            nature = (tx.nature or "").strip()
            if nature != "Internal transfer in":
                continue
            description = (tx.description or "").strip()
            if not any(p.search(description) for p in self._settlement_patterns_compiled):
                continue
            key = (tx.account or "", tx.date)
            settlement_indices[key].append(i)

        reclassified = 0
        for key, indices in settlement_indices.items():
            if len(indices) < 2:
                continue  # Single transaction — no companion, keep as-is

            # Check if group has at least one round amount (the principal)
            has_round = False
            for idx in indices:
                amount = transactions[idx].debit or Decimal("0")
                if amount >= _UNIT_100M and amount % _UNIT_100M == 0:
                    has_round = True
                    break

            if not has_round:
                continue  # No round principal — all are unsplit, keep as-is

            # Reclassify non-round amounts as interest
            for idx in indices:
                tx = transactions[idx]
                amount = tx.debit or Decimal("0")
                is_round = amount >= _UNIT_100M and amount % _UNIT_100M == 0
                if is_round:
                    continue  # Keep round amounts as "Internal transfer in"

                if amount >= _UNIT_100M:
                    tx.nature = "Receipt from tenants"
                else:
                    tx.nature = "Other receipts"
                tx._classified_by = "presplit_interest"
                reclassified += 1

        if reclassified > 0:
            logger.info(f"Reclassified {reclassified} pre-split interest rows")

    def _get_pending_file(self, session_id: str) -> Path:
        """Get path for pending classifications JSON file."""
        working_file = self.template_manager.get_working_file_path(session_id)
        if not working_file:
            raise ValueError(f"Session {session_id} not found")
        return working_file.parent / f"{session_id}_pending.json"

    def _save_pending(self, session_id: str, transactions: List[MovementTransaction], file_results: list, stats: dict) -> None:
        """Save pending classified transactions to JSON for review."""
        import json
        pending_file = self._get_pending_file(session_id)
        data = {
            "transactions": [
                {
                    "index": i,
                    "source": tx.source,
                    "bank": tx.bank,
                    "account": tx.account,
                    "date": tx.date.isoformat() if tx.date else None,
                    "description": tx.description,
                    "debit": float(tx.debit) if tx.debit else None,
                    "credit": float(tx.credit) if tx.credit else None,
                    "nature": tx.nature,
                    "classified_by": getattr(tx, '_classified_by', ''),
                }
                for i, tx in enumerate(transactions)
            ],
            "file_results": file_results,
            "stats": stats,
        }
        pending_file.write_text(json.dumps(data, ensure_ascii=False), encoding='utf-8')
        logger.info(f"Saved {len(transactions)} pending classifications to {pending_file}")

    def _load_pending(self, session_id: str) -> Optional[dict]:
        """Load pending classified transactions from JSON."""
        import json
        pending_file = self._get_pending_file(session_id)
        if not pending_file.exists():
            return None
        data = json.loads(pending_file.read_text(encoding='utf-8'))
        return data

    def _clear_pending(self, session_id: str) -> None:
        """Delete pending classifications file."""
        try:
            pending_file = self._get_pending_file(session_id)
            if pending_file.exists():
                pending_file.unlink()
        except Exception as e:
            logger.warning(f"Failed to clear pending file: {e}")

    async def upload_and_preview(
        self,
        session_id: str,
        files: List[Tuple[str, bytes]],
        filter_by_date: bool = True,
        progress_callback: Optional[callable] = None,
        user_id: Optional[int] = None,
    ) -> Dict[str, Any]:
        """
        Upload files, classify transactions, and return preview for review.
        Does NOT write to Excel yet. Call confirm_classifications() to write.

        Returns:
            Preview with classified transactions and stats
        """
        import time as _time

        _started_at = _time.monotonic()
        logger.info(
            "Upload preview started: session_id=%s files=%s filter_by_date=%s has_progress_callback=%s user_id=%s",
            session_id,
            len(files),
            filter_by_date,
            bool(progress_callback),
            user_id,
        )

        def _finish(result: Dict[str, Any]) -> Dict[str, Any]:
            elapsed_ms = int((_time.monotonic() - _started_at) * 1000)
            logger.info(
                "Upload preview finished: session_id=%s status=%s total_transactions=%s skipped=%s elapsed_ms=%s",
                session_id,
                result.get("status"),
                result.get("total_transactions", 0),
                result.get("total_transactions_skipped", 0),
                elapsed_ms,
            )
            return result

        def emit(
            event_type: str,
            step: str,
            message: str,
            *,
            detail: str = "",
            percentage: int = 0,
            data: Optional[Dict[str, Any]] = None,
        ) -> None:
            self._emit_progress(
                process="upload_preview",
                progress_callback=progress_callback,
                session_id=session_id,
                event_type=event_type,
                step=step,
                message=message,
                detail=detail,
                percentage=percentage,
                data=data,
            )

        if user_id is not None:
            await self._verify_session_owner(session_id, user_id)

        # Get session info + validate working file exists
        session_info, _ = await self._ensure_working_file(session_id)

        # Get date range for filtering
        opening_date = None
        ending_date = None
        if filter_by_date and session_info.get("config"):
            config = session_info["config"]
            if config.get("opening_date"):
                date_str = config["opening_date"]
                if "T" in date_str:
                    opening_date = datetime.fromisoformat(date_str).date()
                else:
                    opening_date = date.fromisoformat(date_str)
            if config.get("ending_date"):
                date_str = config["ending_date"]
                if "T" in date_str:
                    ending_date = datetime.fromisoformat(date_str).date()
                else:
                    ending_date = date.fromisoformat(date_str)

        # Process each file
        all_transactions: List[MovementTransaction] = []
        file_results = []
        total_skipped = 0
        total_found = 0

        for file_idx, (filename, content) in enumerate(files):
            try:
                emit(
                    "step_start",
                    "reading",
                    f"Reading {filename}...",
                    percentage=int((file_idx) / len(files) * 20),
                )

                transactions = self.statement_reader.read_from_bytes(content, "Automation")
                found_count = len(transactions)
                total_found += found_count

                emit(
                    "step_complete",
                    "reading",
                    f"Found {found_count} transactions in {filename}",
                    percentage=int((file_idx + 1) / len(files) * 20),
                )

                # Filter by date if enabled
                skipped = 0
                if filter_by_date and opening_date and ending_date:
                    transactions, skipped = self.statement_reader.filter_by_date_range(
                        transactions, opening_date, ending_date
                    )

                all_transactions.extend(transactions)
                total_skipped += skipped
                file_results.append({
                    "filename": filename,
                    "status": "success",
                    "transactions_found": found_count,
                    "transactions_added": len(transactions),
                    "transactions_skipped": skipped,
                })

                # Track in database
                if self.db_session and len(transactions) > 0:
                    await self._track_uploaded_file(
                        session_id=session_id, filename=filename,
                        file_size=len(content), transactions_count=len(transactions) + skipped,
                        transactions_added=len(transactions), transactions_skipped=skipped,
                    )

            except Exception as e:
                logger.error(f"Error processing file {filename}: {e}")
                file_results.append({"filename": filename, "status": "error", "error": str(e)})

        if not all_transactions:
            emit("complete", "done", "No transactions to process", percentage=100)
            return _finish({
                "session_id": session_id,
                "status": "no_transactions",
                "files_processed": len(files),
                "total_transactions_found": total_found,
                "total_transactions_skipped": total_skipped,
                "file_results": file_results,
                "transactions": [],
            })

        # Classify (rule-based + AI)
        emit(
            "step_start",
            "classifying",
            f"Classifying {len(all_transactions)} transactions...",
            percentage=35,
        )
        if progress_callback:
            await asyncio.sleep(0)

        _classify_start = _time.monotonic()
        classified = await self._classify_transactions(all_transactions, progress_callback)
        _classify_elapsed_ms = (_time.monotonic() - _classify_start) * 1000

        # Collect stats
        rule_count = sum(1 for tx in classified if getattr(tx, '_classified_by', '') == 'rule')
        ai_count = sum(1 for tx in classified if getattr(tx, '_classified_by', '') == 'ai')
        ai_cached_count = sum(1 for tx in classified if getattr(tx, '_classified_by', '') == 'ai_cached')
        unclassified_count = len(classified) - rule_count - ai_count - ai_cached_count
        ai_usage = self.ai_classifier.get_and_reset_usage()
        ai_usage["processing_time_ms"] = _classify_elapsed_ms

        stats = {
            "rule_based": rule_count,
            "ai": ai_count,
            "ai_cached": ai_cached_count,
            "unclassified": unclassified_count,
        }

        logger.info(f"Classification stats: rule={rule_count}, ai={ai_count}, "
                     f"ai_cached={ai_cached_count}, unclassified={unclassified_count}")

        # Save pending (don't write to Excel yet)
        self._save_pending(session_id, classified, file_results, stats)
        self._emit_progress(
            process="upload_preview",
            progress_callback=None,
            session_id=session_id,
            event_type="complete",
            step="preview_ready",
            message=f"Classification complete - review {len(classified)} transactions before confirming",
            percentage=100,
            data={"classification_stats": stats},
        )

        if progress_callback:
            progress_callback(ProgressEvent(
                event_type="complete", step="preview_ready",
                message=f"Classification complete -- review {len(classified)} transactions before confirming",
                percentage=100,
                data={"classification_stats": stats},
            ))

        # Build preview response
        preview_transactions = [
            {
                "index": i,
                "source": tx.source,
                "bank": tx.bank,
                "account": tx.account,
                "date": tx.date.isoformat() if tx.date else None,
                "description": tx.description,
                "debit": float(tx.debit) if tx.debit else None,
                "credit": float(tx.credit) if tx.credit else None,
                "nature": tx.nature,
                "classified_by": getattr(tx, '_classified_by', ''),
            }
            for i, tx in enumerate(classified)
        ]

        return _finish({
            "session_id": session_id,
            "status": "pending_review",
            "files_processed": len(files),
            "total_transactions": len(classified),
            "total_transactions_skipped": total_skipped,
            "file_results": file_results,
            "classification_stats": stats,
            "ai_usage": ai_usage,
            "transactions": preview_transactions,
        })

    async def confirm_classifications(
        self,
        session_id: str,
        modifications: Optional[List[Dict[str, Any]]] = None,
        user_id: Optional[int] = None,
    ) -> Dict[str, Any]:
        """
        Confirm and write pending classifications to Excel.

        Args:
            session_id: The session ID
            modifications: Optional list of {index, nature} to override AI/rule classifications
            user_id: Owner user ID for access control

        Returns:
            Write result summary
        """
        import time as _time

        _started_at = _time.monotonic()
        logger.info(
            "Confirm classifications started: session_id=%s modifications=%s user_id=%s",
            session_id,
            len(modifications or []),
            user_id,
        )

        def _finish(result: Dict[str, Any]) -> Dict[str, Any]:
            elapsed_ms = int((_time.monotonic() - _started_at) * 1000)
            logger.info(
                "Confirm classifications finished: session_id=%s status=%s written=%s modifications_applied=%s elapsed_ms=%s",
                session_id,
                result.get("status"),
                result.get("total_transactions_written", 0),
                result.get("modifications_applied", 0),
                elapsed_ms,
            )
            return result

        if user_id is not None:
            await self._verify_session_owner(session_id, user_id)

        # Load pending data
        pending = self._load_pending(session_id)
        if not pending:
            raise ValueError("No pending classifications found. Please upload and preview first.")

        _, working_file = await self._ensure_working_file(session_id)

        # Apply user modifications
        modifications_applied = 0
        if modifications:
            tx_map = {tx["index"]: tx for tx in pending["transactions"]}
            for mod in modifications:
                idx = mod.get("index")
                new_nature = mod.get("nature")
                if idx is not None and new_nature and idx in tx_map:
                    tx_map[idx]["nature"] = new_nature
                    tx_map[idx]["classified_by"] = "manual"
                    modifications_applied += 1
            logger.info(f"Applied {modifications_applied} manual modifications")

        # Convert back to MovementTransaction objects
        transactions = []
        for tx_data in pending["transactions"]:
            tx_date = None
            if tx_data["date"]:
                tx_date = date.fromisoformat(tx_data["date"])
            transactions.append(MovementTransaction(
                source=tx_data["source"],
                bank=tx_data["bank"],
                account=tx_data["account"],
                date=tx_date,
                description=tx_data["description"],
                debit=Decimal(str(tx_data["debit"])) if tx_data["debit"] else None,
                credit=Decimal(str(tx_data["credit"])) if tx_data["credit"] else None,
                nature=tx_data["nature"],
            ))

        # Prepare Movement (copy Cash Balance -> Prior Period + clear old data) on first write
        is_first_upload = await self._is_first_upload(session_id)
        if is_first_upload:
            from .openpyxl_handler import get_openpyxl_handler
            handler = get_openpyxl_handler()
            await asyncio.to_thread(handler.prepare_movement_for_writing, working_file)

        # Write to Excel
        writer = MovementDataWriter(working_file)
        rows_added, total_rows = await asyncio.to_thread(
            writer.append_transactions, transactions
        )

        # Update session stats
        if self.db_session:
            await self._update_session_stats(session_id, rows_added, len(pending["file_results"]))

        # Clear pending file
        self._clear_pending(session_id)

        result = {
            "session_id": session_id,
            "status": "confirmed",
            "total_transactions_written": rows_added,
            "total_rows_in_movement": total_rows,
            "modifications_applied": modifications_applied,
            "classification_stats": pending.get("stats", {}),
        }

        return _finish(result)

    async def _is_first_upload(self, session_id: str) -> bool:
        """Check if this is the first upload for a session.

        Returns False (skip prep) if:
        - total_transactions > 0 (already uploaded before), OR
        - metadata_json.movement_prepared == True (prep already ran at session init
          because user uploaded their own template)
        """
        if not self.db_session:
            return True  # No DB = always prepare (safe default)
        result = await self.db_session.execute(
            select(
                CashReportSessionModel.total_transactions,
                CashReportSessionModel.metadata_json,
            ).where(CashReportSessionModel.session_id == session_id)
        )
        row = result.one_or_none()
        if row is None:
            return True
        total, metadata = row
        # Skip prep if Movement was already prepared at session init
        if metadata and metadata.get("movement_prepared"):
            return False
        return total is None or total == 0

    async def _track_uploaded_file(
        self,
        session_id: str,
        filename: str,
        file_size: int,
        transactions_count: int,
        transactions_added: int,
        transactions_skipped: int,
        file_type: str = "bank_statement",
    ) -> None:
        """Track uploaded file in database."""
        if not self.db_session:
            return

        # Find session in database
        result = await self.db_session.execute(
            select(CashReportSessionModel).where(
                CashReportSessionModel.session_id == session_id
            )
        )
        db_session = result.scalar_one_or_none()

        if db_session:
            file_model = CashReportUploadedFileModel(
                session_id=db_session.id,
                original_filename=filename,
                file_size=file_size,
                transactions_count=transactions_count,
                transactions_added=transactions_added,
                transactions_skipped=transactions_skipped,
                processed_at=datetime.utcnow(),
                file_type=file_type,
            )
            self.db_session.add(file_model)
            await self.db_session.commit()

    async def _update_session_stats(
        self,
        session_id: str,
        rows_added: int,
        files_added: int,
    ) -> None:
        """Update session statistics in database."""
        if not self.db_session:
            return

        await self.db_session.execute(
            update(CashReportSessionModel)
            .where(CashReportSessionModel.session_id == session_id)
            .values(
                total_transactions=CashReportSessionModel.total_transactions + rows_added,
                total_files_uploaded=CashReportSessionModel.total_files_uploaded + files_added,
            )
        )
        await self.db_session.commit()

    @staticmethod
    def _session_breakdown(
        total_transactions: int,
        metadata: Optional[Dict[str, Any]],
        original_transactions: Optional[int] = None,
    ) -> Dict[str, int]:
        """Build UI-friendly transaction breakdown for session cards/status."""
        meta = metadata or {}

        def _as_int(value: Any) -> int:
            try:
                return int(value or 0)
            except (TypeError, ValueError):
                return 0

        settlement = _as_int(meta.get("settlement_counter_entries"))
        open_new = _as_int(meta.get("open_new_counter_entries"))
        interest_splits = _as_int(meta.get("settlement_interest_splits"))
        if original_transactions is None:
            original = _as_int(total_transactions) - settlement - open_new - interest_splits
            if original < 0:
                original = 0
        else:
            original = _as_int(original_transactions)
        return {
            "transactions": original,
            "settlement": settlement,
            "open_new": open_new,
            "interest_splits": interest_splits,
            "total_rows": _as_int(total_transactions),
        }

    async def get_session_status(self, session_id: str, user_id: Optional[int] = None) -> Optional[Dict[str, Any]]:
        """
        Get session status and statistics (fast - uses database).

        Args:
            session_id: The session ID
            user_id: Owner user ID for access control

        Returns:
            Session info dict or None if not found
        """
        if user_id is not None:
            await self._verify_session_owner(session_id, user_id)

        # Get from database first (fast)
        if self.db_session:
            result = await self.db_session.execute(
                select(CashReportSessionModel)
                .options(selectinload(CashReportSessionModel.uploaded_files))
                .where(CashReportSessionModel.session_id == session_id)
            )
            db_session = result.scalar_one_or_none()

            if db_session:
                # Get file size without opening Excel
                working_file = self.template_manager.get_working_file_path(session_id)
                file_size_mb = 0
                if working_file and working_file.exists():
                    file_size_mb = round(working_file.stat().st_size / (1024 * 1024), 2)

                uploaded_transactions = sum(
                    int(f.transactions_added or 0)
                    for f in (db_session.uploaded_files or [])
                )
                breakdown = self._session_breakdown(
                    db_session.total_transactions,
                    db_session.metadata_json,
                    original_transactions=uploaded_transactions,
                )

                return {
                    "session_id": session_id,
                    "status": db_session.status.value,
                    "movement_rows": breakdown["transactions"],
                    "movement_rows_total": db_session.total_transactions,
                    "breakdown": breakdown,
                    "file_size_mb": file_size_mb,
                    "total_files_uploaded": db_session.total_files_uploaded,
                    "config": {
                        "opening_date": db_session.opening_date.isoformat() if db_session.opening_date else None,
                        "ending_date": db_session.ending_date.isoformat() if db_session.ending_date else None,
                        "fx_rate": float(db_session.fx_rate) if db_session.fx_rate else None,
                        "period_name": db_session.period_name,
                    },
                    "uploaded_files": [
                        {
                            "filename": f.original_filename,
                            "file_size": f.file_size,
                            "transactions_added": f.transactions_added,
                            "transactions_skipped": f.transactions_skipped,
                            "processed_at": f.processed_at.isoformat() if f.processed_at else None,
                            "file_type": getattr(f, "file_type", "bank_statement"),
                        }
                        for f in db_session.uploaded_files
                    ],
                }

        # Fallback to reading file (slower)
        info = self.template_manager.get_session_info(session_id)
        if info:
            info["breakdown"] = self._session_breakdown(
                info.get("movement_rows", 0),
                None,
            )
        return info

    async def reset_session(self, session_id: str, user_id: Optional[int] = None) -> Dict[str, Any]:
        """
        Reset session to clean state.

        Args:
            session_id: The session ID
            user_id: Owner user ID for access control

        Returns:
            Reset result
        """
        if user_id is not None:
            await self._verify_session_owner(session_id, user_id)

        result = self.template_manager.reset_session(session_id)

        # Update database
        if self.db_session:
            # Delete uploaded files records
            db_result = await self.db_session.execute(
                select(CashReportSessionModel).where(
                    CashReportSessionModel.session_id == session_id
                )
            )
            db_session = db_result.scalar_one_or_none()

            if db_session:
                # Delete uploaded files
                await self.db_session.execute(
                    delete(CashReportUploadedFileModel).where(
                        CashReportUploadedFileModel.session_id == db_session.id
                    )
                )
                # Reset stats
                db_session.total_transactions = 0
                db_session.total_files_uploaded = 0
                metadata = dict(db_session.metadata_json or {})
                metadata["settlement_counter_entries"] = 0
                metadata["open_new_counter_entries"] = 0
                metadata["settlement_interest_splits"] = 0
                db_session.metadata_json = metadata
                await self.db_session.commit()

        logger.info(f"Reset session {session_id}")
        return result


    async def delete_session(self, session_id: str, user_id: Optional[int] = None) -> bool:
        """
        Delete a session.

        Args:
            session_id: The session ID
            user_id: Owner user ID for access control

        Returns:
            True if deleted, False if not found
        """
        if user_id is not None:
            await self._verify_session_owner(session_id, user_id)

        # Delete from file system
        file_deleted = self.template_manager.delete_session(session_id)

        # Delete from database
        db_deleted = False
        if self.db_session:
            result = await self.db_session.execute(
                delete(CashReportSessionModel).where(
                    CashReportSessionModel.session_id == session_id
                )
            )
            await self.db_session.commit()
            db_deleted = result.rowcount > 0

        return file_deleted or db_deleted

    def _save_step_snapshot(self, session_id: str, step_name: str) -> None:
        """Save a snapshot of the working file after a step completes.

        Creates a copy like ``snapshot_settlement.xlsx`` in the session directory
        so the user can download the result of each step independently.
        """
        working_file = self.template_manager.get_working_file_path(session_id)
        if not working_file or not working_file.exists():
            return
        snapshot_path = working_file.parent / f"snapshot_{step_name}.xlsx"
        import shutil
        shutil.copy2(working_file, snapshot_path)
        logger.info(f"Saved step snapshot: {snapshot_path.name}")

    async def get_working_file_path(
        self, session_id: str, user_id: Optional[int] = None, step: Optional[str] = None,
    ) -> Optional[Path]:
        """Get the working file path for download.

        Args:
            step: Optional step name (``settlement``, ``open_new``).
                  If given, returns the snapshot from that step.
        """
        if user_id is not None:
            await self._verify_session_owner(session_id, user_id)

        _, working_file = await self._ensure_working_file(session_id)

        if step:
            snapshot = working_file.parent / f"snapshot_{step}.xlsx"
            if snapshot.exists():
                return snapshot
            # Fallback to current working file if snapshot not found
        return working_file

    async def list_sessions(self, user_id: Optional[int] = None) -> List[Dict[str, Any]]:
        """List active sessions from database, filtered by user."""
        if self.db_session:
            query = select(CashReportSessionModel).where(
                CashReportSessionModel.status == CashReportSessionStatus.ACTIVE
            )
            if user_id is not None:
                query = query.where(CashReportSessionModel.user_id == user_id)
            query = query.options(selectinload(CashReportSessionModel.uploaded_files))
            query = query.order_by(CashReportSessionModel.created_at.desc())

            result = await self.db_session.execute(query)
            db_sessions = result.scalars().all()

            sessions: List[Dict[str, Any]] = []
            for s in db_sessions:
                uploaded_transactions = sum(
                    int(f.transactions_added or 0)
                    for f in (s.uploaded_files or [])
                )
                breakdown = self._session_breakdown(
                    s.total_transactions,
                    s.metadata_json,
                    original_transactions=uploaded_transactions,
                )
                sessions.append({
                    "session_id": s.session_id,
                    "movement_rows": breakdown["transactions"],
                    "movement_rows_total": s.total_transactions,
                    "breakdown": breakdown,
                    "file_size_mb": 0,  # Don't read file for listing
                    "config": {
                        "opening_date": s.opening_date.isoformat() if s.opening_date else None,
                        "ending_date": s.ending_date.isoformat() if s.ending_date else None,
                        "fx_rate": float(s.fx_rate) if s.fx_rate else None,
                        "period_name": s.period_name,
                    },
                })
            return sessions

        # Fallback to file system (slower)
        sessions = self.template_manager.list_sessions()
        for s in sessions:
            s["breakdown"] = self._session_breakdown(
                s.get("movement_rows", 0),
                None,
            )
        return sessions

    async def get_data_preview(self, session_id: str, limit: int = 20, user_id: Optional[int] = None) -> List[dict]:
        """
        Get preview of Movement data.

        Args:
            session_id: The session ID
            limit: Maximum number of rows
            user_id: Owner user ID for access control

        Returns:
            List of row dicts
        """
        if user_id is not None:
            await self._verify_session_owner(session_id, user_id)
        try:
            _, working_file = await self._ensure_working_file(session_id)
        except ValueError:
            return []

        writer = MovementDataWriter(working_file)
        return writer.get_data_preview(limit)

    async def run_reconcile_checks(
        self,
        session_id: str,
        user_id: Optional[int] = None,
    ) -> Dict[str, Any]:
        """
        Run Step 8 + Step 9 reconciliation checks with real data.

        Step 8:
        - Internal transfer in (debit) must equal internal transfer out (credit)

        Step 9:
        - Per account/currency closing balance reconciliation versus
          ``Cash balance (BS)`` end-of-period balance.
        """
        if user_id is not None:
            await self._verify_session_owner(session_id, user_id)

        _, working_file = await self._ensure_working_file(session_id)

        def _normalize_account_text(value: Any) -> str:
            text = str(value or "").strip().replace("'", "")
            if not text:
                return ""
            try:
                if "." in text and float(text) == int(float(text)):
                    return str(int(float(text)))
            except (ValueError, OverflowError):
                pass
            return text

        def _parse_decimal(value: Any, *, default_none: Optional[Decimal] = Decimal("0")) -> Optional[Decimal]:
            if value is None:
                return default_none
            if isinstance(value, Decimal):
                return value
            if isinstance(value, (int, float)):
                return Decimal(str(value))

            raw = str(value).strip().replace("\u00A0", "").replace(" ", "")
            if not raw:
                return default_none

            negative = False
            if raw.startswith("(") and raw.endswith(")"):
                negative = True
                raw = raw[1:-1]

            if re.match(r"^\d{1,3}(?:\.\d{3})+(?:,\d+)?$", raw):
                raw = raw.replace(".", "").replace(",", ".")
            else:
                raw = raw.replace(",", "")

            try:
                parsed = Decimal(raw)
            except Exception:
                return default_none
            return -parsed if negative else parsed

        def _to_float(value: Optional[Decimal]) -> Optional[float]:
            if value is None:
                return None
            try:
                return float(value)
            except Exception:
                return None

        writer = MovementDataWriter(working_file)
        transactions = writer.get_all_transactions()
        acc_char_data = self._read_acc_char_full_data(working_file)

        # account -> canonical currency fallback for rows where formula cache is empty
        account_currency_map: Dict[str, str] = {}
        for acc, info in acc_char_data.items():
            cur = str((info or {}).get("currency") or "").strip().upper()
            if cur:
                account_currency_map[_normalize_account_text(acc)] = cur

        # Movement aggregates by (account, currency)
        movement_by_key: Dict[Tuple[str, str], Dict[str, Decimal]] = defaultdict(
            lambda: {"debit": Decimal("0"), "credit": Decimal("0")}
        )
        internal_by_currency: Dict[str, Dict[str, Decimal]] = defaultdict(
            lambda: {"in": Decimal("0"), "out": Decimal("0")}
        )

        for tx in transactions:
            account = _normalize_account_text(tx.account)
            if not account:
                continue
            currency = account_currency_map.get(account, "VND")
            key = (account, currency)

            debit = tx.debit or Decimal("0")
            credit = tx.credit or Decimal("0")
            movement_by_key[key]["debit"] += debit
            movement_by_key[key]["credit"] += credit

            nature = (tx.nature or "").strip().lower()
            if nature == "internal transfer in":
                internal_by_currency[currency]["in"] += debit
            elif nature == "internal transfer out":
                internal_by_currency[currency]["out"] += credit

        # Read prior-period opening balances (by account/currency)
        prior_opening_by_key: Dict[Tuple[str, str], Decimal] = defaultdict(lambda: Decimal("0"))
        # Read bank statement balances from "Cash balance (BS)"
        bank_bs_by_key: Dict[Tuple[str, str], Dict[str, Any]] = {}
        # Read metadata from current "Cash Balance" table
        cash_balance_meta_by_key: Dict[Tuple[str, str], Dict[str, Any]] = {}

        import openpyxl
        wb = openpyxl.load_workbook(working_file, data_only=True, read_only=True)
        try:
            # Step 9 input A: prior period closing balances (opening for current period)
            prior_sheet = None
            for sheet_name in wb.sheetnames:
                name_lower = sheet_name.lower()
                if "prior" in name_lower and "cash" in name_lower:
                    prior_sheet = wb[sheet_name]
                    break
            if prior_sheet is not None:
                for row in prior_sheet.iter_rows(min_row=4, max_col=7):
                    account = _normalize_account_text(row[2].value if len(row) > 2 else None)  # C
                    currency = str(row[4].value or "").strip().upper() if len(row) > 4 else ""  # E
                    if not account or not currency:
                        continue
                    value_raw = (row[5].value if currency == "VND" else row[6].value) if len(row) > 6 else None
                    opening_val = _parse_decimal(value_raw)
                    if opening_val is None:
                        continue
                    prior_opening_by_key[(account, currency)] += opening_val

            # Step 9 input B: bank statement balances
            if "Cash balance (BS)" in wb.sheetnames:
                ws_bs = wb["Cash balance (BS)"]
                for row in ws_bs.iter_rows(min_row=2, max_col=5):
                    account = _normalize_account_text(row[0].value if len(row) > 0 else None)  # A
                    currency = str(row[1].value or "").strip().upper() if len(row) > 1 else ""  # B
                    if not account or not currency:
                        continue
                    bop = _parse_decimal(row[2].value if len(row) > 2 else None)  # C
                    eop = _parse_decimal(row[3].value if len(row) > 3 else None, default_none=None)  # D
                    bank_name = str(row[4].value or "").strip() if len(row) > 4 else ""  # E
                    bank_bs_by_key[(account, currency)] = {
                        "bop": bop,
                        "eop": eop,
                        "bank_name": bank_name,
                    }

            # Step 9 metadata: account display info from current Cash Balance sheet
            if "Cash Balance" in wb.sheetnames:
                ws_cb = wb["Cash Balance"]
                for row_num, row in enumerate(ws_cb.iter_rows(min_row=4, max_col=25), start=4):
                    account = _normalize_account_text(row[2].value if len(row) > 2 else None)  # C
                    if not account or account.lower() == "x":
                        continue
                    currency = str(row[4].value or "").strip().upper() if len(row) > 4 else ""  # E
                    if not currency:
                        currency = account_currency_map.get(account, "VND")
                    key = (account, currency)
                    if key in cash_balance_meta_by_key:
                        continue

                    entity = str(row[0].value or "").strip() if len(row) > 0 else ""  # A
                    branch = str(row[1].value or "").strip() if len(row) > 1 else ""  # B
                    bank_name = str(row[24].value or "").strip() if len(row) > 24 else ""  # Y
                    bank_1 = str(row[23].value or "").strip() if len(row) > 23 else ""  # X

                    # Fallback to Acc_Char if formula cache on Cash Balance is empty
                    acc_info = acc_char_data.get(account, {})
                    if not entity:
                        entity = str(acc_info.get("entity") or "").strip()
                    if not branch:
                        branch = str(acc_info.get("branch") or "").strip()

                    cash_balance_meta_by_key[key] = {
                        "row": row_num,
                        "entity": entity,
                        "branch": branch,
                        "bank": bank_name,
                        "bank_1": bank_1,
                    }
        finally:
            wb.close()

        # Step 8: internal transfer in/out balance
        tolerance_by_currency = {
            "VND": Decimal("1"),
            "USD": Decimal("0.01"),
        }
        internal_rows = []
        internal_balanced = True
        for currency in sorted(internal_by_currency.keys()):
            total_in = internal_by_currency[currency]["in"]
            total_out = internal_by_currency[currency]["out"]
            diff = total_in - total_out
            tolerance = tolerance_by_currency.get(currency, Decimal("1"))
            is_balanced = abs(diff) <= tolerance
            if not is_balanced:
                internal_balanced = False
            internal_rows.append({
                "currency": currency,
                "total_in": _to_float(total_in),
                "total_out": _to_float(total_out),
                "difference": _to_float(diff),
                "tolerance": _to_float(tolerance),
                "is_balanced": is_balanced,
            })

        # Step 9: account-level cash balance reconciliation
        all_keys = (
            set(movement_by_key.keys())
            | set(prior_opening_by_key.keys())
            | set(bank_bs_by_key.keys())
            | set(cash_balance_meta_by_key.keys())
        )

        account_rows: List[Dict[str, Any]] = []
        matched_bank_count = 0
        reconciled_count = 0
        missing_bank_count = 0
        not_reconciled_count = 0
        total_abs_diff = Decimal("0")

        for account, currency in sorted(all_keys, key=lambda x: (x[0], x[1])):
            opening = prior_opening_by_key.get((account, currency), Decimal("0"))
            movement = movement_by_key.get((account, currency), {"debit": Decimal("0"), "credit": Decimal("0")})
            total_debit = movement["debit"]
            total_credit = movement["credit"]
            calculated_closing = opening + total_debit - total_credit

            bs_info = bank_bs_by_key.get((account, currency), {})
            bank_eop = bs_info.get("eop")
            bank_bop = bs_info.get("bop")
            bank_name = str(bs_info.get("bank_name") or "").strip()

            meta = cash_balance_meta_by_key.get((account, currency), {})
            entity = str(meta.get("entity") or "").strip()
            branch = str(meta.get("branch") or "").strip()
            bank = str(meta.get("bank") or "").strip() or bank_name

            tolerance = tolerance_by_currency.get(currency, Decimal("1"))
            issues: List[str] = []
            difference: Optional[Decimal] = None
            is_reconciled = False

            if bank_eop is None:
                missing_bank_count += 1
                issues.append("Missing EOP in Cash balance (BS)")
            else:
                matched_bank_count += 1
                difference = calculated_closing - bank_eop
                total_abs_diff += abs(difference)
                if abs(difference) <= tolerance:
                    is_reconciled = True
                    reconciled_count += 1
                else:
                    not_reconciled_count += 1
                    issues.append(f"Difference exceeds tolerance ({tolerance})")

            account_rows.append({
                "account": account,
                "currency": currency,
                "entity": entity,
                "branch": branch,
                "bank": bank,
                "opening_balance": _to_float(opening),
                "movement_debit": _to_float(total_debit),
                "movement_credit": _to_float(total_credit),
                "calculated_closing": _to_float(calculated_closing),
                "bank_statement_bop": _to_float(bank_bop),
                "bank_statement_eop": _to_float(bank_eop),
                "difference": _to_float(difference),
                "tolerance": _to_float(tolerance),
                "is_reconciled": is_reconciled,
                "issues": issues,
            })

        cash_balance_balanced = (not_reconciled_count == 0 and missing_bank_count == 0)

        result = {
            "session_id": session_id,
            "status": "success",
            "step8_internal_transfer_reconcile": {
                "is_balanced": internal_balanced,
                "rows": internal_rows,
                "movement_rows_scanned": len(transactions),
            },
            "step9_cash_balance_reconcile": {
                "is_balanced": cash_balance_balanced,
                "total_accounts": len(account_rows),
                "accounts_with_bank_balance": matched_bank_count,
                "reconciled_count": reconciled_count,
                "not_reconciled_count": not_reconciled_count,
                "missing_bank_balance_count": missing_bank_count,
                "total_absolute_difference": _to_float(total_abs_diff),
                "rows": account_rows,
            },
        }

        logger.info(
            "Reconcile checks finished: session=%s step8_balanced=%s step9_balanced=%s accounts=%s not_reconciled=%s missing_bank=%s",
            session_id,
            internal_balanced,
            cash_balance_balanced,
            len(account_rows),
            not_reconciled_count,
            missing_bank_count,
        )
        return result


    def _read_saving_accounts(self, working_file: str) -> List[Dict[str, Any]]:
        """
        Read Saving Account sheet from working file.

        Returns:
            List of dicts with keys: row, account, bank_1, bank, entity, branch,
            term_months, closing_balance_vnd, maturity_date, interest_rate
        """
        import openpyxl
        from datetime import datetime
        wb = openpyxl.load_workbook(working_file, data_only=True, read_only=True)
        saving_accounts = []
        try:
            if "Saving Account" not in wb.sheetnames:
                logger.warning("Saving Account sheet not found")
                return []

            ws = wb["Saving Account"]
            # Row 3 = headers, Row 4+ = data
            # Col A=Entity, B=Branch, C=Account Number, D=Type, E=Currency,
            # Col F=CLOSING BALANCE (VND), Col H=Term (months), Col K=Maturity date,
            # Col N(14)=Bank_1 (short code like TCB, BIDV), Col O(15)=Bank (full name)
            for row_num, row_data in enumerate(ws.iter_rows(min_row=4, values_only=False), start=4):
                account = row_data[2].value if len(row_data) > 2 else None
                if not account:
                    continue
                account_str = str(account).strip()
                if not account_str or account_str.lower() == "x":
                    continue

                # Col H (index 7): Term (months), usually "1 month"/"3 months"
                term_months = None
                term_raw = row_data[7].value if len(row_data) > 7 else None
                if term_raw is not None and term_raw != "":
                    if isinstance(term_raw, (int, float)):
                        term_months = int(term_raw)
                    else:
                        text = str(term_raw).strip().upper()
                        m = re.search(r"(\d{1,2})\s*(?:THANG|MONTHS?|M)\b", text)
                        if m:
                            term_months = int(m.group(1))
                        elif text.isdigit():
                            term_months = int(text)

                # Col F (index 5): CLOSING BALANCE (VND)
                closing_raw = row_data[5].value if len(row_data) > 5 else None
                closing_balance = Decimal("0")
                if closing_raw is not None:
                    try:
                        closing_balance = Decimal(str(closing_raw))
                    except Exception:
                        pass

                # Col K (index 10): Maturity date
                maturity_raw = row_data[10].value if len(row_data) > 10 else None
                maturity_date = None
                if isinstance(maturity_raw, datetime):
                    maturity_date = maturity_raw.date()
                elif maturity_raw:
                    try:
                        maturity_date = datetime.strptime(str(maturity_raw).strip(), "%Y-%m-%d").date()
                    except Exception:
                        try:
                            maturity_date = datetime.strptime(str(maturity_raw).strip(), "%d/%m/%Y").date()
                        except Exception:
                            pass

                # Col L (index 11): Interest rate
                rate_raw = row_data[11].value if len(row_data) > 11 else None
                interest_rate = None
                if rate_raw is not None:
                    try:
                        rv = float(rate_raw)
                        # Stored as decimal (e.g. 0.0475) or percentage (4.75)
                        interest_rate = rv if rv < 1 else rv / 100
                    except (ValueError, TypeError):
                        pass

                saving_accounts.append({
                    "row": row_num,
                    "entity": str(row_data[0].value or "").strip(),
                    "branch": str(row_data[1].value or "").strip(),
                    "account": account_str,
                    "bank_1": str(row_data[13].value or "").strip() if len(row_data) > 13 else "",
                    "bank": str(row_data[14].value or "").strip() if len(row_data) > 14 else "",
                    "term_months": term_months,
                    "closing_balance_vnd": closing_balance,
                    "maturity_date": maturity_date,
                    "interest_rate": interest_rate,
                })
        finally:
            wb.close()

        logger.info(f"Loaded {len(saving_accounts)} saving accounts")
        return saving_accounts

    async def _sync_saving_account_lookup_fields(
        self,
        working_file: Path,
        lookup_account_details: Dict[str, Dict[str, Any]],
    ) -> Dict[str, int]:
        """
        Sync all Saving Account rows with lookup metadata by account number.

        Updates existing rows (not only newly inserted rows) for:
        - H: Term (months)
        - K: Maturity date
        - L: Interest rate
        """
        if not lookup_account_details:
            return {
                "rows_checked": 0,
                "rows_matched_lookup": 0,
                "rows_updated": 0,
                "term_months_updated": 0,
                "maturity_date_updated": 0,
                "interest_rate_updated": 0,
                "provider_mismatch_skipped": 0,
            }

        def _normalize_account_text(value: Any) -> str:
            text = str(value or "").strip().replace("'", "")
            if not text:
                return ""
            try:
                if "." in text and float(text) == int(float(text)):
                    return str(int(float(text)))
            except (ValueError, OverflowError):
                pass
            return text

        def _excel_serial(dt: date) -> int:
            # Excel 1900 date system (same convention used by openpyxl_handler)
            base = date(1899, 12, 30)
            return (dt - base).days

        saving_rows = self._read_saving_accounts(str(working_file))
        modifications: Dict[int, Dict[str, str]] = {}

        rows_matched_lookup = 0
        rows_updated = 0
        term_months_updated = 0
        maturity_date_updated = 0
        interest_rate_updated = 0
        provider_mismatch_skipped = 0
        rate_changed_cells: List[Tuple[int, str]] = []  # (row, "L")

        for row in saving_rows:
            row_num = int(row.get("row") or 0)
            if row_num <= 0:
                continue

            account = _normalize_account_text(row.get("account"))
            if not account:
                continue

            lookup_meta = lookup_account_details.get(account)
            if not lookup_meta:
                continue
            rows_matched_lookup += 1

            provider = str(lookup_meta.get("provider") or "").strip().upper()
            bank_1 = str(row.get("bank_1") or row.get("bank") or "").strip().upper()
            if provider and bank_1 and not self._bank_matches(bank_1, provider):
                provider_mismatch_skipped += 1
                continue

            row_mod: Dict[str, str] = {}

            new_term_months = lookup_meta.get("term_months")
            if new_term_months is not None:
                try:
                    new_term_int = int(new_term_months)
                except (ValueError, TypeError):
                    new_term_int = None
                if new_term_int and new_term_int > 0:
                    current_term = row.get("term_months")
                    if current_term != new_term_int:
                        suffix = "month" if new_term_int == 1 else "months"
                        row_mod["H"] = f"{new_term_int} {suffix}"
                        term_months_updated += 1

            new_maturity = lookup_meta.get("maturity_date")
            if isinstance(new_maturity, datetime):
                new_maturity = new_maturity.date()
            if isinstance(new_maturity, date):
                current_maturity = row.get("maturity_date")
                if current_maturity != new_maturity:
                    row_mod["K"] = str(_excel_serial(new_maturity))
                    maturity_date_updated += 1

            new_rate = lookup_meta.get("interest_rate")
            if new_rate is not None:
                try:
                    new_rate_f = float(new_rate)
                except (ValueError, TypeError):
                    new_rate_f = None
                if new_rate_f is not None and new_rate_f > 0:
                    current_rate = row.get("interest_rate")
                    if current_rate is None or abs(float(current_rate) - new_rate_f) > 1e-9:
                        row_mod["L"] = str(new_rate_f)
                        interest_rate_updated += 1
                        rate_changed_cells.append((row_num, "L"))

            if row_mod:
                modifications[row_num] = row_mod
                rows_updated += 1

        if modifications:
            from .openpyxl_handler import get_openpyxl_handler
            handler = get_openpyxl_handler()
            await asyncio.to_thread(
                handler.modify_cell_values,
                Path(working_file),
                "Saving Account",
                modifications,
            )
            logger.info(
                "Saving Account lookup sync: updated_rows=%s term_months=%s maturity_dates=%s rates=%s",
                rows_updated,
                term_months_updated,
                maturity_date_updated,
                interest_rate_updated,
            )

            # Highlight cells where interest rate was changed
            if rate_changed_cells:
                light_blue_fill = (
                    b'<fill><patternFill patternType="solid">'
                    b'<fgColor rgb="FF00B0F0"/>'
                    b'</patternFill></fill>'
                )
                await asyncio.to_thread(
                    handler.highlight_cells,
                    Path(working_file),
                    "Saving Account",
                    rate_changed_cells,
                    light_blue_fill,
                )
                logger.info(
                    "Highlighted %d rate-changed cells in Saving Account",
                    len(rate_changed_cells),
                )
        else:
            logger.info("Saving Account lookup sync: no field changes detected")

        return {
            "rows_checked": len(saving_rows),
            "rows_matched_lookup": rows_matched_lookup,
            "rows_updated": rows_updated,
            "term_months_updated": term_months_updated,
            "maturity_date_updated": maturity_date_updated,
            "interest_rate_updated": interest_rate_updated,
            "rate_changed_cells": len(rate_changed_cells),
            "provider_mismatch_skipped": provider_mismatch_skipped,
        }

    @staticmethod
    def _read_acc_char_account_to_code(working_file) -> Dict[str, str]:
        """
        Read account->code mapping from Acc_Char sheet columns B (Account No.) and C (CODE).
        Both are value columns (not formulas), so always reliable.

        Returns:
            {account_no: code} e.g. {"8680028808": "TCS", "1037150624": "T3B"}
        """
        import openpyxl
        account_to_code: Dict[str, str] = {}
        try:
            wb = openpyxl.load_workbook(working_file, data_only=True, read_only=True)
            if "Acc_Char" not in wb.sheetnames:
                wb.close()
                return {}
            ws = wb["Acc_Char"]
            for row in ws.iter_rows(min_row=2, max_col=3):
                acc = row[1].value if len(row) > 1 else None   # B: Account No.
                code = row[2].value if len(row) > 2 else None  # C: CODE
                if acc and code:
                    acc_str = str(acc).strip()
                    try:
                        if "." in acc_str and float(acc_str) == int(float(acc_str)):
                            acc_str = str(int(float(acc_str)))
                    except (ValueError, OverflowError):
                        pass
                    account_to_code[acc_str] = str(code).strip()
            wb.close()
        except Exception as e:
            logger.warning(f"Failed to read Acc_Char account map: {e}")
        return account_to_code

    @staticmethod
    def _read_acc_char_full_data(working_file) -> Dict[str, Dict[str, str]]:
        """
        Read full account data from Acc_Char sheet (columns B-H).

        Returns:
            {account_no: {
                "code": str, "entity": str, "branch": str,
                "currency": str, "account_type": str, "grouping": str
            }}
        """
        import openpyxl
        acc_data: Dict[str, Dict[str, str]] = {}
        try:
            wb = openpyxl.load_workbook(working_file, data_only=True, read_only=True)
            if "Acc_Char" not in wb.sheetnames:
                wb.close()
                return {}
            ws = wb["Acc_Char"]
            for row in ws.iter_rows(min_row=2, max_col=8):  # B..H = indices 1..7
                acc = row[1].value if len(row) > 1 else None   # B: Account No.
                if not acc:
                    continue
                acc_str = str(acc).strip()
                # Normalize float account numbers
                try:
                    if "." in acc_str and float(acc_str) == int(float(acc_str)):
                        acc_str = str(int(float(acc_str)))
                except (ValueError, OverflowError):
                    pass
                if not acc_str or acc_str.lower() == "x":
                    continue

                code_val = row[2].value if len(row) > 2 else None       # C: CODE
                entity_val = row[3].value if len(row) > 3 else None     # D: ENTITY (formula)
                branch_val = row[4].value if len(row) > 4 else None     # E: BRANCH
                currency_val = row[5].value if len(row) > 5 else None   # F: CURRENCY
                acc_type_val = row[6].value if len(row) > 6 else None   # G: ACCOUNT TYPE
                grouping_val = row[7].value if len(row) > 7 else None   # H: NAME/Grouping (formula)

                acc_data[acc_str] = {
                    "code": str(code_val).strip() if code_val else "",
                    "entity": str(entity_val).strip() if entity_val else "",
                    "branch": str(branch_val).strip() if branch_val else "",
                    "currency": str(currency_val).strip() if currency_val else "",
                    "account_type": str(acc_type_val).strip() if acc_type_val else "",
                    "grouping": str(grouping_val).strip() if grouping_val else "",
                }
            wb.close()
        except Exception as e:
            logger.warning(f"Failed to read Acc_Char full data: {e}")
        logger.info(f"Read Acc_Char full data: {len(acc_data)} accounts")
        return acc_data

    @staticmethod
    def _resolve_acc_char_for_saving_account(
        saving_acc: str,
        acc_char_data: Dict[str, Dict[str, str]],
        max_suffix: int = 20,
    ) -> Optional[Dict[str, str]]:
        """
        Look up saving account in Acc_Char data, trying suffix variations.

        Search order:
        1. Exact match: saving_acc
        2. Suffixed: saving_acc_1, saving_acc_2, ..., saving_acc_{max_suffix}

        Returns:
            The Acc_Char data dict if found, or None.
        """
        # 1) Exact match
        if saving_acc in acc_char_data:
            return acc_char_data[saving_acc]
        # 2) Suffixed variations
        base_acc = saving_acc.split("_")[0]
        for suffix in range(1, max_suffix + 1):
            suffixed = f"{base_acc}_{suffix}"
            if suffixed in acc_char_data:
                return acc_char_data[suffixed]
        return None

    @staticmethod
    def _build_account_entity_map(working_file: str) -> Dict[str, str]:
        """
        Build current_account -> entity_name mapping from Cash Balance sheet.

        Cash Balance columns: A=Entity, C=Account, X=Bank_1.
        These are value columns (not formulas), so always available.

        This replaces reliance on the Movement sheet's Entity formula column
        (col J), which has no cached values when the file hasn't been opened
        in Excel.

        Returns:
            {account_number: entity_name} e.g. {"110002915294": "BW BAU BANG 01"}
        """
        import openpyxl
        account_entity: Dict[str, str] = {}
        try:
            wb = openpyxl.load_workbook(working_file, data_only=True, read_only=True)
            sheet_name = None
            for sn in wb.sheetnames:
                if sn == "Cash Balance":
                    sheet_name = sn
                    break
            if not sheet_name:
                wb.close()
                return {}

            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=4, max_col=25):
                account = row[2].value if len(row) > 2 else None  # C: Account
                entity = row[0].value if len(row) > 0 else None   # A: Entity
                if account and entity:
                    acc_str = str(account).strip()
                    # Normalize float account numbers (Excel sometimes stores as float)
                    try:
                        if "." in acc_str and float(acc_str) == int(float(acc_str)):
                            acc_str = str(int(float(acc_str)))
                    except (ValueError, OverflowError):
                        pass
                    account_entity[acc_str] = str(entity).strip()
            wb.close()
        except Exception as e:
            logger.warning(f"Failed to build account->entity map from Cash Balance: {e}")
        logger.info(f"Built account->entity map: {len(account_entity)} entries")
        return account_entity

    @staticmethod
    def _read_def_entity_to_code(working_file) -> Dict[str, str]:
        """
        Read entity->code mapping from the Def sheet.
        Def sheet: Row 3 = headers, Row 4+ = data.
        Column A (0) = Abb. (CODE), Column B (1) = Entities (ENTITY).

        Returns:
            {entity_upper: code} e.g. {"TCS JSC": "TCS", "THUAN THANH 3B": "T3B"}
        """
        import openpyxl
        entity_to_code: Dict[str, str] = {}
        try:
            wb = openpyxl.load_workbook(working_file, data_only=True, read_only=True)
            if "Def" not in wb.sheetnames:
                wb.close()
                return {}
            ws = wb["Def"]
            for row in ws.iter_rows(min_row=4, max_col=2):
                code = row[0].value if len(row) > 0 else None   # A: Abb. (CODE)
                entity = row[1].value if len(row) > 1 else None  # B: Entities
                if code and entity:
                    entity_to_code[str(entity).strip().upper()] = str(code).strip()
            wb.close()
        except Exception as e:
            logger.warning(f"Failed to read Def entity map: {e}")
        return entity_to_code

    @staticmethod
    def _normalize_bank_key(value: str) -> str:
        """Normalize bank text for dictionary matching."""
        return re.sub(r"[^A-Z0-9]+", "", str(value or "").strip().upper())

    @staticmethod
    def _read_def_bank_alias_to_name(working_file) -> Dict[str, str]:
        """
        Read bank alias mapping from Def sheet Bank/Abb table.

        Def layout:
        - Row 3: headers
        - One table has "Bank" + adjacent "Abb."

        Returns:
            {normalized_alias_or_name: canonical_bank_name}
            e.g. {"TCB": "TECHCOMBANK", "TECHCOMBANK": "TECHCOMBANK"}
        """
        import openpyxl

        alias_to_bank: Dict[str, str] = {}
        wb = None
        try:
            wb = openpyxl.load_workbook(working_file, data_only=True, read_only=True)
            if "Def" not in wb.sheetnames:
                return {}
            ws = wb["Def"]

            bank_col = 7
            abb_col = 8
            max_scan_col = min(ws.max_column, 60)
            for col_idx in range(1, max_scan_col):
                h_bank = str(ws.cell(3, col_idx).value or "").strip().lower()
                h_abb = str(ws.cell(3, col_idx + 1).value or "").strip().lower().replace(".", "")
                if h_bank == "bank" and h_abb in {"abb", "abbr"}:
                    bank_col = col_idx
                    abb_col = col_idx + 1
                    break

            for row_idx in range(4, ws.max_row + 1):
                bank_name = str(ws.cell(row_idx, bank_col).value or "").strip()
                abb = str(ws.cell(row_idx, abb_col).value or "").strip()
                if not bank_name:
                    continue

                bank_key = CashReportService._normalize_bank_key(bank_name)
                if bank_key:
                    alias_to_bank[bank_key] = bank_name

                abb_key = CashReportService._normalize_bank_key(abb)
                if abb_key:
                    alias_to_bank[abb_key] = bank_name
        except Exception as e:
            logger.warning(f"Failed to read Def bank map: {e}")
        finally:
            if wb:
                wb.close()

        return alias_to_bank

    @classmethod
    def _normalize_bank_name(
        cls,
        bank_value: str,
        def_bank_alias_to_name: Optional[Dict[str, str]] = None,
    ) -> str:
        """
        Normalize bank text to canonical bank name using Def map first,
        then fallback aliases.
        """
        raw = str(bank_value or "").strip()
        if not raw:
            return ""

        parts = raw.split("-", 1)
        root = parts[0].strip()
        suffix = parts[1].strip() if len(parts) > 1 else ""

        root_key = cls._normalize_bank_key(root)
        full_key = cls._normalize_bank_key(raw)

        resolved = ""
        if def_bank_alias_to_name:
            resolved = (
                def_bank_alias_to_name.get(root_key)
                or def_bank_alias_to_name.get(full_key)
                or ""
            )
        if not resolved:
            resolved = (
                DEFAULT_BANK_ALIAS_TO_FULL_NAME.get(root_key)
                or DEFAULT_BANK_ALIAS_TO_FULL_NAME.get(full_key)
                or ""
            )

        if not resolved:
            return raw
        if suffix:
            return f"{resolved} - {suffix}"
        return resolved

    @staticmethod
    def _read_prior_period_branches(working_file) -> Dict[str, List[str]]:
        """
        Read entity->branches mapping from 'Cash balance (Prior period)' sheet.
        Row 3 = headers, Row 4+ = data.
        Column A (0) = ENTITY, Column B (1) = BRANCH.

        Returns:
            {entity_upper: [branch1, branch2, ...]} e.g. {"TCS JSC": ["BIDV - THANH XUAN", "VIETINBANK - QUANG MINH"]}
        """
        import openpyxl
        entity_branches: Dict[str, List[str]] = {}
        try:
            wb = openpyxl.load_workbook(working_file, data_only=True, read_only=True)
            sheet_name = None
            for sn in wb.sheetnames:
                if "prior" in sn.lower() and "cash" in sn.lower():
                    sheet_name = sn
                    break
            if not sheet_name:
                wb.close()
                return {}
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=4, max_col=2):
                entity = row[0].value if len(row) > 0 else None
                branch = row[1].value if len(row) > 1 else None
                if entity and branch:
                    ent_key = str(entity).strip().upper()
                    br_val = str(branch).strip()
                    if ent_key not in entity_branches:
                        entity_branches[ent_key] = []
                    if br_val not in entity_branches[ent_key]:
                        entity_branches[ent_key].append(br_val)
            wb.close()
        except Exception as e:
            logger.warning(f"Failed to read prior period branches: {e}")
        return entity_branches

    @staticmethod
    def _determine_saving_branch(
        saving_acc: str,
        entity: str,
        entity_branches: Dict[str, List[str]],
        original_bank: str,
        def_bank_alias_to_name: Optional[Dict[str, str]] = None,
        provider_hint: str = "",
    ) -> str:
        """
        Determine BRANCH for a new saving account by looking up the entity's
        branches in Cash Balance (Prior period), then matching with the saving
        account's bank prefix.

        Logic:
        1. If provider_hint is given (from lookup metadata), use it as primary
        2. Otherwise determine bank from saving account prefix
        3. Pick the branch that matches the bank
        4. Fallback: first branch for entity, or normalized original_bank
        """
        acc = str(saving_acc).strip()
        normalized_original_bank = CashReportService._normalize_bank_name(
            original_bank, def_bank_alias_to_name,
        )

        # Determine bank keyword — prefer provider_hint from lookup metadata
        bank_keyword = ""
        hint_norm = str(provider_hint or "").strip().upper()
        if hint_norm:
            # Map provider to keyword used in branch names
            _PROVIDER_TO_KEYWORD = {
                "BIDV": "BIDV",
                "VCB": "VIETCOMBANK",
                "VTB": "VIETINBANK",
                "WOORI": "WOORI",
                "SNP": "SINOPAC",
                "TCB": "TECHCOMBANK",
                "TECHCOMBANK": "TECHCOMBANK",
                "VIETCOMBANK": "VIETCOMBANK",
                "VIETINBANK": "VIETINBANK",
            }
            bank_keyword = _PROVIDER_TO_KEYWORD.get(hint_norm, hint_norm)

        # Fallback: determine bank from account prefix
        if not bank_keyword and len(acc) >= 10:
            if acc.startswith("8"):
                bank_keyword = "BIDV"
            elif acc.startswith("1") and len(acc) in (10, 13):
                bank_keyword = "VIETCOMBANK"
            elif acc.startswith("2007") and len(acc) == 12:
                # Woori accounts: 200700xxxxx pattern
                bank_keyword = "WOORI"
            elif acc.startswith("2") and len(acc) == 12:
                bank_keyword = "VIETINBANK"

        # Look up entity's branches
        branches = entity_branches.get((entity or "").upper(), [])
        if branches and bank_keyword:
            for br in branches:
                if bank_keyword in br.upper():
                    return br

        # Fallback: first branch for this entity
        if branches:
            return branches[0]

        return normalized_original_bank

    @staticmethod
    def _extract_saving_account_from_description(description: str) -> Optional[str]:
        """
        Extract saving account number from description for settlement transactions.

        Patterns:
        - "TAT TOAN TIEN GUI SO 14501110378000" -> 14501110378000
        - "TAT TOAN HDTG RGLH SO... (TK 1059960714)" -> 1059960714
        - "TAT TOAN TAI KHOAN TIEN GUI CO KY HAN SO 218000472157" -> 218000472157
        - "Tra goc TK tien gui 217000486074" -> 217000486074
        - "07003600017772" (bare account number) -> 07003600017772

        Returns:
            Account number string or None
        """
        if not description:
            return None

        desc_stripped = description.strip()

        # Pattern 0: Bare account number (SINOPAC-style, description = only digits)
        # Accept optional ".0" suffix from Excel numeric string conversion.
        bare_acc_match = re.match(r'^(\d{10,})(?:\.0+)?$', desc_stripped)
        if bare_acc_match:
            return bare_acc_match.group(1)

        # Pattern 1: "(TK xxxxxx)" - account in parentheses
        match = re.search(r'\(TK\s*(\d{6,20})\)', description, re.IGNORECASE)
        if match:
            return match.group(1)

        # Pattern 2: "SO xxxxxx" at end - "TAT TOAN ... SO 14501110378000"
        # Must have "TAT TOAN" or "TIEN GUI" or "HDTG" context
        # NOTE: Only match pure-digit accounts. Dot-separated numbers like
        # "640.2025.41808" are HDTG contract numbers, NOT saving account numbers.
        if re.search(r'TAT\s*TOAN|TIEN\s*GUI|HDTG', description, re.IGNORECASE):
            match = re.search(r'SO\s*(\d{6,20})\s*$', description, re.IGNORECASE)
            if match:
                return match.group(1)
            # Also try without $ anchor if there's trailing text
            match = re.search(r'SO\s*(\d{10,20})(?:\s|$)', description, re.IGNORECASE)
            if match:
                return match.group(1)

        # Pattern 3: "tien gui XXXXXXXX" - account after "tien gui" (VTB "Tra goc" pattern)
        # E.g., "Tra goc TK tien gui 217000486074" -> 217000486074
        match = re.search(r'tien\s*gui\s+(\d{10,20})', description, re.IGNORECASE)
        if match:
            return match.group(1)

        return None

    @staticmethod
    def _split_settlement_amount(amount: Decimal) -> tuple:
        """
        Split a settlement amount into principal (round) and interest.

        Round (divisible by 100M) → no split (full amount is principal)
        >= 1B non-round           → split at 1B boundary (principal = floor(amount/1B)*1B)
        < 1B & divisible by 100M → no split (round, full amount is principal)
        < 1B & not divisible by 100M → principal=0, pure interest

        Returns:
            (principal, interest) — both Decimal.
        """
        UNIT_100M = Decimal("100000000")   # 100M VND
        UNIT_1B = Decimal("1000000000")    # 1B VND
        # Round totals (e.g. 3.8B) are principal-only and should not be split.
        if amount >= UNIT_100M and amount % UNIT_100M == 0:
            return amount, Decimal("0")
        # Always split at 1B boundary first
        principal = (amount // UNIT_1B) * UNIT_1B
        interest = amount - principal
        return principal, interest

    @staticmethod
    def _bank_matches(tx_bank: str, sa_bank_1: str) -> bool:
        """
        Check if a Movement bank code matches a Saving Account bank_1.

        Handles mismatches like:
        - "WOORI" vs "Woori Bank" (abbreviation vs full name)
        - "SINOPAC" vs "SNP" (different abbreviations)
        """
        a = CashReportService._normalize_bank_key(
            CashReportService._normalize_bank_name(tx_bank),
        )
        b = CashReportService._normalize_bank_key(
            CashReportService._normalize_bank_name(sa_bank_1),
        )
        if not a or not b:
            return False
        if a == b:
            return True
        # Partial match: one contains the other
        if a in b or b in a:
            return True
        return False

    def _find_saving_account(
        self,
        bank: str,
        entity: str,
        saving_accounts: List[Dict[str, Any]],
        description: str = "",
        principal: Decimal = None,
        tx_date=None,
    ) -> Optional[str]:
        """
        Find saving account for settlement transaction.

        Priority:
        1. Extract account number from description (for TAT TOAN transactions)
        2. Match by CLOSING BALANCE (VND) = principal AND Maturity date ±3 days
           (filtered by bank + entity first)
        3. Fallback: bank + entity only (single match)

        Returns:
            Saving account number or None
        """
        from datetime import timedelta

        # First try to extract from description
        extracted = self._extract_saving_account_from_description(description)
        if extracted:
            logger.info(f"Settlement: extracted account {extracted} from description")
            return extracted

        entity_upper = entity.upper().strip() if entity else ""

        # Filter by bank + entity
        bank_entity_matches = [
            sa for sa in saving_accounts
            if self._bank_matches(bank, sa["bank_1"])
            and sa["entity"].upper() == entity_upper
        ]

        # Try matching by CLOSING BALANCE + Maturity date ±3 days
        if principal and tx_date and bank_entity_matches:
            amount_date_matches = []
            for sa in bank_entity_matches:
                # Check closing balance matches principal
                if sa.get("closing_balance_vnd") != principal:
                    continue
                # Check maturity date within ±3 days
                mat_date = sa.get("maturity_date")
                if mat_date is None:
                    continue
                delta = abs((mat_date - tx_date).days)
                if delta <= 3:
                    amount_date_matches.append((sa, delta))

            if amount_date_matches:
                # Pick the closest maturity date
                amount_date_matches.sort(key=lambda x: x[1])
                best = amount_date_matches[0][0]
                logger.info(
                    f"Settlement: found saving account {best['account']} by "
                    f"closing_balance={principal}, maturity_date={best['maturity_date']}, "
                    f"tx_date={tx_date}, bank={bank}, entity={entity}"
                )
                return best["account"]

            # Partial withdrawal fallback: principal < closing_balance + maturity ±3 days
            # For "TAT TOAN MOT PHAN" / partial withdrawals, the principal is less than
            # the full deposit balance, so exact match fails.
            partial_matches = []
            for sa in bank_entity_matches:
                bal = sa.get("closing_balance_vnd") or Decimal("0")
                if bal <= 0 or principal >= bal:
                    continue
                mat_date = sa.get("maturity_date")
                if mat_date is None:
                    continue
                delta = abs((mat_date - tx_date).days)
                if delta <= 3:
                    partial_matches.append((sa, delta))

            if len(partial_matches) == 1:
                best = partial_matches[0][0]
                logger.info(
                    f"Settlement: found saving account {best['account']} by "
                    f"partial withdrawal (principal={principal} < balance={best['closing_balance_vnd']}), "
                    f"maturity_date={best['maturity_date']}, tx_date={tx_date}, "
                    f"bank={bank}, entity={entity}"
                )
                return best["account"]
            elif len(partial_matches) > 1:
                logger.info(
                    f"Settlement: {len(partial_matches)} partial-withdrawal matches for "
                    f"bank={bank}, entity={entity}, principal={principal}, tx_date={tx_date}: "
                    f"{[(m[0]['account'], m[0]['closing_balance_vnd']) for m in partial_matches]}"
                )

        # Fallback: match by bank + entity + exact closing balance (no maturity required)
        # When multiple accounts share the same balance, use maturity date as
        # tiebreaker — pick the one whose maturity is closest to tx_date,
        # preferring maturity <= tx_date (already matured = most likely settled).
        if principal and bank_entity_matches:
            balance_only_matches = [
                sa for sa in bank_entity_matches
                if sa.get("closing_balance_vnd") == principal
            ]
            if len(balance_only_matches) == 1:
                best = balance_only_matches[0]
                logger.info(
                    f"Settlement: found saving account {best['account']} by "
                    f"closing_balance={principal} (no maturity constraint), "
                    f"bank={bank}, entity={entity}"
                )
                return best["account"]
            elif len(balance_only_matches) > 1 and tx_date:
                # Tiebreak by maturity: prefer matured (mat <= tx_date), then closest
                def _maturity_sort_key(sa):
                    mat = sa.get("maturity_date")
                    if mat is None:
                        return (1, 9999)  # no maturity → lowest priority
                    delta = (mat - tx_date).days  # negative = already matured
                    already_matured = 0 if delta <= 0 else 1
                    return (already_matured, abs(delta))

                balance_only_matches.sort(key=_maturity_sort_key)
                best = balance_only_matches[0]
                logger.info(
                    f"Settlement: found saving account {best['account']} by "
                    f"closing_balance={principal}, maturity tiebreak "
                    f"(maturity={best.get('maturity_date')}, tx_date={tx_date}), "
                    f"bank={bank}, entity={entity}, "
                    f"candidates={[(m['account'], m.get('maturity_date')) for m in balance_only_matches]}"
                )
                return best["account"]
            elif len(balance_only_matches) > 1:
                logger.info(
                    f"Settlement: {len(balance_only_matches)} balance-only matches for "
                    f"bank={bank}, entity={entity}, principal={principal}: "
                    f"{[m['account'] for m in balance_only_matches]}"
                )

        # Fallback: single match by bank + entity
        if len(bank_entity_matches) == 1:
            logger.info(f"Settlement: found saving account {bank_entity_matches[0]['account']} by bank={bank} + entity={entity}")
            return bank_entity_matches[0]["account"]
        elif len(bank_entity_matches) > 1:
            logger.warning(
                f"Settlement: multiple saving accounts for bank={bank}, entity={entity}, "
                f"principal={principal}, tx_date={tx_date}: {[m['account'] for m in bank_entity_matches]} — no amount/date match"
            )

        # Fallback: match by bank only
        matches_bank = [sa for sa in saving_accounts if self._bank_matches(bank, sa["bank_1"])]
        if len(matches_bank) == 1:
            logger.info(f"Settlement fallback: found saving account {matches_bank[0]['account']} by bank={bank}")
            return matches_bank[0]["account"]

        logger.warning(f"Settlement: no saving account found for bank={bank}, entity={entity}")
        return None

    @staticmethod
    def _has_dual_entity_transfer(desc_lower: str) -> bool:
        """
        Check if description mentions 2+ different entity codes -> internal transfer.
        Example: "VC3_DC2_TRANSFER MONEY" -> VC3 + DC2 = 2 entities.

        Handles substring overlap: if both "th2" and "th2hc" match,
        "th2" is removed (it's a substring of the longer match).
        """
        matched = {code for code in ENTITY_CODES if code in desc_lower}
        if len(matched) < 2:
            return False
        # Remove codes that are substrings of longer matched codes
        # e.g., "th2" inside "th2hc" -> only keep "th2hc"
        filtered = {
            code for code in matched
            if not any(code != other and code in other for other in matched)
        }
        return len(filtered) >= 2

    def _is_profit_distribution_transfer(self, description: str) -> bool:
        """Detect profit-distribution wording in transfer descriptions."""
        desc_norm = self._normalize_text_for_match(description or "")
        if not desc_norm:
            return False
        return any(keyword in desc_norm for keyword in PROFIT_DISTRIBUTION_KEYWORDS)

    def _is_settlement_candidate(self, tx) -> bool:
        """
        Detect if a transaction is a settlement candidate.
        Requires BOTH conditions:
        1. Nature == "Internal transfer in"  (mandatory)
        2. Description matches settlement patterns/keywords OR contains dual entity codes

        A transaction is a settlement candidate if:
        - debit > 0 (cash in)
        - NOT an existing counter entry
        - Nature is "Internal transfer in"
        - AND (regex pattern match OR keyword match OR dual-entity transfer)
        """
        # Must have debit > 0 (cash in)
        if not tx.debit or tx.debit <= 0:
            return False

        nature = (tx.nature or "").strip().lower()

        # Hard-block counter entries and interest rows — these are NEVER settlement
        # candidates regardless of description content.
        if nature in SETTLEMENT_BLOCKED_NATURES:
            return False

        # Only "Internal transfer in" is eligible for settlement.
        # Respect AI classification — if AI says "Receipt from tenants" etc.,
        # it's not a settlement even if description matches patterns.
        if nature not in SETTLEMENT_ELIGIBLE_NATURES:
            return False

        description = (tx.description or "").strip()
        desc_lower = description.lower()
        desc_norm = self._normalize_bank_description(description)

        # Description-level overrides to prevent false settlement detection.
        if self._is_numeric_account_like_description(desc_norm):
            _, interest = self._split_settlement_amount(tx.debit)
            if interest > 0:
                return False

        # Check settlement regex patterns (Group A patterns)
        for pattern in self._settlement_patterns_compiled:
            if pattern.search(description):
                return True

        # Check simple settlement keywords
        for keyword in SETTLEMENT_KEYWORDS:
            if keyword in desc_lower:
                return True

        # HDTG/deposit keywords with large amount (>= 1B) → settlement candidate
        if (tx.debit or Decimal("0")) >= _UNIT_1B:
            if any(kw in desc_lower for kw in HDTG_DEPOSIT_KEYWORDS):
                return True

        # NOTE: Dual-entity check removed -- it caused false positives on
        # intercompany transfers (e.g., "INTERNAL TRANSFER FROM VC3 TO DC2").
        # All legitimate settlements are caught by patterns above.

        return False

    def _is_settlement_interest_row(self, tx) -> bool:
        """
        Detect if a transaction is an interest row accompanying a settlement.
        These have Nature="Other receipts" (interest payment) but their description
        matches settlement patterns (e.g., "TAT TOAN HDTG SO ...").
        They should be highlighted but NOT generate counter entries.
        """
        if not tx.debit or tx.debit <= 0:
            return False
        nature = (tx.nature or "").strip().lower()
        if nature != "other receipts":
            return False
        description = (tx.description or "").strip()
        desc_lower = description.lower()
        for pattern in self._settlement_patterns_compiled:
            if pattern.search(description):
                return True
        for keyword in SETTLEMENT_KEYWORDS:
            if keyword in desc_lower:
                return True
        return False

    def _is_open_new_candidate(self, tx) -> bool:
        """
        Detect if a transaction is an "open new saving account" candidate.
        Requires BOTH conditions:
        1. Nature == "Internal transfer out" (mandatory)
        2. Description matches GROUP B patterns (Gá»­i tiá»n / Má»Ÿ HDTG)

        A transaction is an open-new candidate if:
        - credit > 0 (cash out from current account)
        - NOT an existing counter entry (created by previous automation)
        - Nature is "Internal transfer out"
        - AND regex pattern match (GROUP B)
        """
        # Must have credit > 0 (cash out)
        if not tx.credit or tx.credit <= 0:
            return False

        # Nature must be "Internal transfer out" (mandatory)
        # This implicitly skips counter entries ("Internal transfer in")
        # regardless of source (Automation, NS, Manual)
        nature = (tx.nature or "").strip().lower()
        if nature != "internal transfer out":
            return False

        description = tx.description or ""

        # Profit-distribution transfers are not savings open-new transactions.
        if self._is_profit_distribution_transfer(description):
            return False

        # Woori-specific override:
        # "Withdrawal - Withdrawal" with large round credit is a known open-new flow.
        credit_amount = tx.credit or Decimal("0")
        desc_norm = self._normalize_text_for_match(description)
        unit_100m = Decimal("100000000")
        if (
            self._bank_matches(tx.bank or "", "WOORI")
            and desc_norm == WOORI_WITHDRAWAL_WITHDRAWAL_NORMALIZED
            and credit_amount >= unit_100m
            and credit_amount % unit_100m == 0
        ):
            return True

        # Check open-new regex patterns (Group B patterns)
        for pattern in self._open_new_patterns_compiled:
            if pattern.search(description):
                return True

        return False

    @staticmethod
    def _extract_saving_account_for_open_new(description: str) -> Optional[str]:
        """
        Extract saving account number from description for open-new transactions.

        Patterns for HDTG/saving account:
        - "CK den tai khoan 813015095347" -> 813015095347
        - "HDTG SO 123456789" -> 123456789
        - "(TK 123456789)" -> 123456789
        - "GUI HDTG ... SO 123456789" -> 123456789

        Returns:
            Account number string or None
        """
        if not description:
            return None

        # Pattern 1: "(TK xxxxxx)" - account in parentheses
        match = re.search(r'\(TK\s*(\d{6,20})\)', description, re.IGNORECASE)
        if match:
            return match.group(1)

        # Pattern 2: "tai khoan XXXXXX" or "account XXXXXX" - BIDV/VCB style
        match = re.search(r'(?:tai\s*khoan|account)\s*(\d{6,20})', description, re.IGNORECASE)
        if match:
            return match.group(1)

        # Pattern 3: "HDTG SO xxxxxx" or "SO xxxxxx" in HDTG context
        if re.search(r'HDTG|GUI\s*TIEN|MO\s*HD', description, re.IGNORECASE):
            match = re.search(r'SO\s*(\d{8,20})', description, re.IGNORECASE)
            if match:
                return match.group(1)

        # Pattern 4: Long number sequence that looks like account (10+ digits)
        # Used when description is mostly numeric (SNP/SINOPAC case)
        if re.match(r'^[\d,.\sE+]+$', description.strip()):
            match = re.search(r'(\d{10,20})', description)
            if match:
                return match.group(1)

        return None

    @staticmethod
    def _extract_term_info(description: str) -> dict:
        """
        Extract term, interest rate, and maturity date from description.

        Examples:
        - "HDTG 900/2026/47248, KH 01 THANG, LS 4.75%/NAM" -> term="1 months", rate=0.0475
        - "KY HAN 175 NGAY, DEN HAN 24/07/2026" -> term="175 days", maturity="2026-07-24"
        - "TKKH 46 NGAY THEO HD SO 0034.2026" -> term="46 days"
        - "1 THANG THEO HD" -> term="1 months"
        - "LS 4.75%" or "LAI SUAT 4.75%" -> rate=0.0475

        Returns:
            dict with keys: term_months, term_days, interest_rate, maturity_date (all optional)
        """
        result = {
            "term_months": None,
            "term_days": None,
            "interest_rate": None,
            "maturity_date": None,
        }

        if not description:
            return result

        desc_upper = description.upper()

        # Extract term in months: "KH 01 THANG", "1 THANG", "KY HAN 3 THANG", "3M", "01M"
        month_patterns = [
            r'(?:KH|KY\s*HAN|TKKH)?\s*(\d{1,2})\s*(?:THANG|M(?:ONTH)?S?)\b',
            r'\b(\d{1,2})\s*THANG\b',
        ]
        for pattern in month_patterns:
            match = re.search(pattern, desc_upper)
            if match:
                result["term_months"] = int(match.group(1))
                break

        # Extract term in days: "175 NGAY", "KY HAN 84 NGAY"
        day_patterns = [
            r'(?:KY\s*HAN|TKKH)?\s*(\d{1,3})\s*NGAY\b',
            r'\b(\d{2,3})\s*(?:NGAY|DAY)\b',
        ]
        for pattern in day_patterns:
            match = re.search(pattern, desc_upper)
            if match:
                days = int(match.group(1))
                if days > 1:  # Avoid matching single digit dates
                    result["term_days"] = days
                    break

        # Extract interest rate: "LS 4.75%", "LAI SUAT 4.75%/NAM", "4.75%"
        rate_patterns = [
            r'(?:LS|LAI\s*SUAT)\s*(\d+[.,]\d+)\s*%',
            r'\b(\d+[.,]\d{1,2})\s*%\s*(?:/\s*NAM)?',
        ]
        for pattern in rate_patterns:
            match = re.search(pattern, desc_upper)
            if match:
                rate_str = match.group(1).replace(',', '.')
                result["interest_rate"] = float(rate_str) / 100  # Convert to decimal
                break

        # Extract maturity date: "DEN HAN 24/07/2026", "24.07.2026"
        date_patterns = [
            r'(?:DEN\s*HAN|DAO\s*HAN)\s*(\d{1,2})[/.](\d{1,2})[/.](\d{4})',
            r'\bNGAY\s*(\d{1,2})[/.](\d{1,2})[/.](\d{4})',
        ]
        for pattern in date_patterns:
            match = re.search(pattern, desc_upper)
            if match:
                try:
                    day, month, year = int(match.group(1)), int(match.group(2)), int(match.group(3))
                    from datetime import date as dt_date
                    result["maturity_date"] = dt_date(year, month, day)
                except (ValueError, TypeError):
                    pass
                break

        return result

    async def run_settlement_automation(
        self,
        session_id: str,
        user_id: Optional[int] = None,
        progress_callback=None,
        dry_run: bool = False,
    ) -> Dict[str, Any]:
        """
        Run settlement (táº¥t toÃ¡n) automation on Movement data.

        Logic:
        1. Detect settlement transactions by keyword patterns
        2. Lookup saving account from Saving Account sheet (by account number in description + bank match)
        3. Create counter entry with saving account, reversed debit/credit, nature = Internal transfer out

        Args:
            session_id: The session ID
            user_id: Owner user ID for access control
            progress_callback: Optional callable for SSE progress events
            dry_run: If True, run detect+lookup but do NOT write to Excel.
                     Returns candidate preview data instead.

        Returns:
            Dict with results summary (or preview data when dry_run=True)
        """
        import time as _time

        _started_at = _time.monotonic()
        logger.info(
            "Settlement automation started: session_id=%s dry_run=%s has_progress_callback=%s user_id=%s",
            session_id,
            dry_run,
            bool(progress_callback),
            user_id,
        )

        def _finish(result: Dict[str, Any]) -> Dict[str, Any]:
            elapsed_ms = int((_time.monotonic() - _started_at) * 1000)
            logger.info(
                "Settlement automation finished: session_id=%s status=%s counter_entries=%s interest_splits=%s elapsed_ms=%s",
                session_id,
                result.get("status"),
                result.get("counter_entries_created", 0),
                result.get("interest_splits_created", 0),
                elapsed_ms,
            )
            return result

        def emit(event_type, step, message, detail="", percentage=0, data=None):
            self._emit_progress(
                process="settlement",
                progress_callback=progress_callback,
                session_id=session_id,
                event_type=event_type,
                step=step,
                message=message,
                detail=detail,
                percentage=percentage,
                data=data,
            )

        if user_id is not None:
            await self._verify_session_owner(session_id, user_id)

        _, working_file = await self._ensure_working_file(session_id)

        writer = MovementDataWriter(working_file)
        self._refresh_detection_pattern_indexes_if_needed()

        # -- Step 1: Scan Movement data --
        emit("step_start", "scanning", "Scanning Movement data...", percentage=0)
        await asyncio.sleep(0.3)  # Allow SSE to deliver step_start

        transactions = writer.get_all_transactions()
        if not transactions:
            emit("complete", "done", "No transactions found", percentage=100,
                 data={"status": "no_transactions"})
            return _finish({
                "session_id": session_id,
                "status": "no_transactions",
                "message": "No transactions found in Movement sheet",
                "counter_entries_created": 0,
            })

        # Build current_account -> entity mapping from Cash Balance sheet.
        # The Movement sheet's Entity column (J) is formula-based and has no
        # cached values when the file hasn't been opened in Excel, so we
        # resolve entity from Cash Balance (which has value columns).
        account_entity_map = self._build_account_entity_map(working_file)

        # Build comprehensive Acc_Char lookup for populating cached VLOOKUP values
        # on settlement counter entries (prevents #N/A in Entity/Currency/etc.)
        acc_char_data = self._read_acc_char_full_data(working_file)

        # Load entity→code and branch mappings for inserting new Acc_Char rows
        def_entity_to_code = self._read_def_entity_to_code(working_file)
        def_bank_alias_to_name = self._read_def_bank_alias_to_name(working_file)
        entity_branches = self._read_prior_period_branches(working_file)

        emit("step_complete", "scanning",
             f"Found {len(transactions)} transactions",
             percentage=20,
             data={"total_transactions": len(transactions)})
        await asyncio.sleep(0.4)  # Allow SSE to deliver step transition

        # â"€â"€ Step 2: Detect settlement transactions â"€â"€
        emit("step_start", "detecting", "Detecting settlement transactions...", percentage=20)
        await asyncio.sleep(0.3)

        saving_accounts = self._read_saving_accounts(working_file)

        # Detect by keyword patterns + amount splitting upfront.
        # Amount splitting BEFORE detection ensures pure-interest rows
        # (< 100M, no round principal) are classified correctly from the start.
        settlement_transactions = []
        interest_row_indices = []  # Interest rows — nature correction only, NO counter
        # Track cells that need modification (amount splitting + nature correction)
        cell_modifications = {}  # {row_num: {"F": new_debit_str, "I": new_nature}}
        # Interest rows to INSERT right below their settlement row
        interest_inserts = {}  # {row_idx: tx_data_dict}
        _skip_reasons = {"no_debit": 0, "wrong_nature": 0, "no_pattern": 0}

        for idx, tx in enumerate(transactions):
            row_idx = idx + 4

            if self._is_settlement_interest_row(tx):
                interest_row_indices.append(row_idx)
                continue

            if not self._is_settlement_candidate(tx):
                if not tx.debit or tx.debit <= 0:
                    _skip_reasons["no_debit"] += 1
                elif (tx.nature or "").strip().lower() in SETTLEMENT_BLOCKED_NATURES:
                    _skip_reasons["wrong_nature"] += 1
                else:
                    _skip_reasons["no_pattern"] += 1
                continue

            # — Amount splitting: separate principal from interest —
            # Bank deposits have round principal amounts (divisible by 100M).
            # If the total debit is NOT round, the bank combined principal + interest.
            # Settlement prerequisite: principal must be round AND >= 100M.
            principal, interest = self._split_settlement_amount(tx.debit)

            if principal <= 0:
                # Pure interest (no round principal, amount < 1B) → NOT a settlement.
                # Always "Other receipts".
                interest_row_indices.append(row_idx)
                nature_now = (tx.nature or "").strip().lower()
                if nature_now != "other receipts":
                    cell_modifications[row_idx] = {"I": "Other receipts"}
                logger.info(f"Settlement: pure interest row {row_idx}, amount={tx.debit}, nature fixed to Other receipts")
                continue

            if interest > 0:
                # Combined amount → split: modify original row debit to principal,
                # insert interest row DIRECTLY BELOW the settlement row.
                # Interest part is always "Other receipts".
                cell_modifications[row_idx] = {"F": str(principal)}
                interest_inserts[row_idx] = {
                    'source': tx.source or "Automation",
                    'bank': tx.bank or "",
                    'account': tx.account or "",
                    'date': tx.date,
                    'description': tx.description or "",
                    'debit': interest,
                    'credit': 0,
                    'nature': "Other receipts",
                }
                logger.info(
                    f"Settlement split row {row_idx}: "
                    f"principal={principal}, interest={interest} (total={tx.debit})"
                )

            # Store principal for counter entry creation in Step 3
            tx._principal = principal
            settlement_transactions.append((tx, row_idx))

        logger.info(f"Settlement detect: {len(settlement_transactions)} candidates, {len(interest_row_indices)} interest rows, skip reasons: {_skip_reasons}")

        if not settlement_transactions:
            emit("complete", "done", "No settlement transactions found", percentage=100,
                 data={"status": "no_settlements", "total_transactions_scanned": len(transactions)})
            return _finish({
                "session_id": session_id,
                "status": "no_settlements",
                "message": "No settlement transactions found",
                "counter_entries_created": 0,
                "total_transactions_scanned": len(transactions),
            })

        emit("step_complete", "detecting",
             f"Found {len(settlement_transactions)} settlement transactions",
             percentage=40,
             data={"settlement_found": len(settlement_transactions)})
        await asyncio.sleep(0.4)

        # â"€â"€ Step 3: Lookup saving accounts â"€â"€
        emit("step_start", "lookup", "Looking up saving accounts...", percentage=40)
        await asyncio.sleep(0.3)

        existing_descs = set()
        for tx in transactions:
            if tx.nature and "transfer out" in tx.nature.lower():
                existing_descs.add(tx.description.strip() if tx.description else "")

        counter_entries = []
        original_row_indices = []  # Rows to highlight (exclude nature="Other receipts")
        skipped_no_account = []
        skipped_duplicate = []
        acc_char_inserts = []  # Saving accounts to add to Acc_Char (not found anywhere)

        for tx, row_idx in settlement_transactions:
            desc = tx.description.strip() if tx.description else ""
            if desc in existing_descs:
                skipped_duplicate.append(desc)
                continue

            # Principal was computed in Step 2 (amount splitting)
            principal = tx._principal

            # Resolve entity from Cash Balance account→entity map
            tx_account = str(tx.account).strip() if tx.account else ""
            entity = account_entity_map.get(tx_account, "")

            # Normalize settlement rows to Internal transfer in even if classifier mislabeled.
            nature_now = (tx.nature or "").strip().lower()
            if nature_now != "internal transfer in":
                cell_modifications.setdefault(row_idx, {})["I"] = "Internal transfer in"

            saving_acc = self._find_saving_account(
                bank=tx.bank or "",
                entity=entity,
                saving_accounts=saving_accounts,
                description=desc,
                principal=principal,
                tx_date=tx.date if isinstance(tx.date, date) else None,
            )

            if not saving_acc:
                skipped_no_account.append(tx.description)
                logger.warning(f"Settlement: no saving account found for '{tx.description}', bank={tx.bank}, entity={entity}")
                continue

            # --- Resolve Acc_Char data for the saving account ---
            # Populate cached VLOOKUP values to prevent #N/A in Movement formula columns
            acc_data = self._resolve_acc_char_for_saving_account(saving_acc, acc_char_data)

            # Determine the account number to use in the Movement counter entry
            counter_account = saving_acc  # Default: use saving account from description

            if acc_data:
                # Found in Acc_Char (exact or suffixed) — use saving_acc as-is
                counter_entity = acc_data.get("entity") or account_entity_map.get(saving_acc, "") or entity
                counter_grouping = acc_data.get("grouping") or "Subsidiaries"
                counter_currency = acc_data.get("currency") or "VND"
                counter_account_type = acc_data.get("account_type") or "Saving Account"
                logger.debug(f"Settlement: found Acc_Char data for {saving_acc}: entity={counter_entity}")
            else:
                # Not found by saving_acc — check current_account_1, _2, ... in Acc_Char
                # If _1 exists → reuse it (no insert needed)
                # If _1 NOT exist → insert _1. If _1 taken → insert _2, etc.
                found_existing_suffix = None
                for s in range(1, 21):
                    candidate = f"{tx_account}_{s}"
                    if candidate in acc_char_data:
                        found_existing_suffix = (candidate, acc_char_data[candidate])
                        break

                if found_existing_suffix:
                    # Reuse existing suffixed account — no Acc_Char insert needed
                    counter_account, existing_data = found_existing_suffix
                    counter_entity = existing_data.get("entity") or entity
                    counter_grouping = existing_data.get("grouping") or "Subsidiaries"
                    counter_currency = existing_data.get("currency") or "VND"
                    counter_account_type = existing_data.get("account_type") or "Saving Account"
                    logger.debug(
                        f"Settlement: {saving_acc} not in Acc_Char, "
                        f"reusing existing {counter_account} (entity={counter_entity})"
                    )
                else:
                    # No suffixed account exists — create _1 (or next available)
                    suffix = 1
                    while f"{tx_account}_{suffix}" in acc_char_data:
                        suffix += 1
                    counter_account = f"{tx_account}_{suffix}"

                    current_acc_data = acc_char_data.get(tx_account, {})
                    code = current_acc_data.get("code", "")
                    if not code and entity:
                        code = def_entity_to_code.get(entity.upper(), "")
                    branch = self._determine_saving_branch(
                        saving_acc,
                        entity,
                        entity_branches,
                        tx.bank or "",
                        def_bank_alias_to_name,
                    )
                    counter_currency = current_acc_data.get("currency") or "VND"

                    acc_char_inserts.append({
                        "saving_acc": counter_account,
                        "code": code,
                        "entity": entity,
                        "branch": branch,
                        "currency": counter_currency,
                        "account_type": "Saving Account",
                    })

                    # Register so next iteration sees it
                    acc_char_data[counter_account] = {
                        "code": code,
                        "entity": entity,
                        "branch": branch,
                        "currency": counter_currency,
                        "account_type": "Saving Account",
                        "grouping": "Subsidiaries",
                    }

                    counter_entity = entity
                    counter_grouping = "Subsidiaries"
                    counter_account_type = "Saving Account"
                    logger.info(
                        f"Settlement: {saving_acc} not in Acc_Char, "
                        f"will insert as {counter_account} (code={code}, entity={entity})"
                    )

            counter = MovementTransaction(
                source=tx.source or "Automation",
                bank=tx.bank,
                account=counter_account,  # May be saving_acc or current_account_N
                date=tx.date,
                description=tx.description or "",
                debit=Decimal("0"),
                credit=principal,  # Use principal (round part), NOT full amount
                nature="Internal transfer out",
                entity=counter_entity,
                grouping=counter_grouping,
                currency=counter_currency,
                account_type=counter_account_type,
            )

            counter_entries.append(counter)
            final_nature = cell_modifications.get(row_idx, {}).get("I", tx.nature or "")
            if final_nature.strip().lower() != "other receipts":
                original_row_indices.append(row_idx)

        emit("step_complete", "lookup",
             f"Matched {len(counter_entries)} accounts, {len(skipped_no_account)} not found, {len(skipped_duplicate)} duplicates",
             percentage=60,
             data={
                 "matched": len(counter_entries),
                 "skipped_no_account": len(skipped_no_account),
                 "skipped_duplicate": len(skipped_duplicate),
             })
        await asyncio.sleep(0.4)

        # ── Dry-run: return preview without writing ──
        if dry_run:
            nature_corrections_count = len([m for m in cell_modifications.values() if "I" in m])
            candidates_preview = []
            for tx, row_idx in settlement_transactions:
                principal = getattr(tx, '_principal', tx.debit)
                # Find the matched counter entry for this tx (by description)
                matched_counter = next(
                    (c for c in counter_entries if c.description == (tx.description or "").strip()),
                    None
                )
                desc = (tx.description or "").strip()
                candidates_preview.append({
                    "row": row_idx,
                    "date": tx.date.isoformat() if tx.date else None,
                    "bank": tx.bank or "",
                    "account": str(tx.account or ""),
                    "description": desc,
                    "debit": float(principal),
                    "nature": tx.nature or "",
                    "status": (
                        "duplicate" if desc in [d for d in skipped_duplicate]
                        else "no_account" if desc in skipped_no_account
                        else "matched"
                    ),
                    "counter_account": str(matched_counter.account) if matched_counter else None,
                    "has_interest_split": row_idx in interest_inserts,
                    "interest_amount": float(interest_inserts[row_idx]["debit"]) if row_idx in interest_inserts else None,
                })
            return _finish({
                "session_id": session_id,
                "dry_run": True,
                "status": "preview",
                "total_transactions_scanned": len(transactions),
                "settlement_candidates": len(settlement_transactions),
                "counter_entries_would_create": len(counter_entries),
                "interest_splits_would_create": len(interest_inserts),
                "nature_corrections_would_apply": nature_corrections_count,
                "skipped_no_account": skipped_no_account,
                "skipped_duplicate": skipped_duplicate,
                "acc_char_inserts_would_add": [i["saving_acc"] for i in acc_char_inserts],
                "candidates": candidates_preview,
            })

        if not counter_entries:
            # Even without counter entries, apply modifications and insertions
            if cell_modifications:
                try:
                    writer.modify_cell_values(cell_modifications)
                    logger.info(f"Applied {len(cell_modifications)} cell modifications (no counter entries)")
                except Exception as e:
                    logger.warning(f"Failed to apply cell modifications: {e}")
            if interest_inserts:
                try:
                    writer.insert_rows_after(interest_inserts)
                except Exception as e:
                    logger.warning(f"Failed to insert interest rows: {e}")
            self._save_step_snapshot(session_id, "settlement")

            msg = "No counter entries created"
            if skipped_no_account:
                msg += f". {len(skipped_no_account)} skipped (no matching saving account)"
            if skipped_duplicate:
                msg += f". {len(skipped_duplicate)} skipped (already exists)"
            emit("complete", "done", msg, percentage=100,
                 data={"status": "no_counter_entries"})
            return _finish({
                "session_id": session_id,
                "status": "no_counter_entries",
                "message": msg,
                "counter_entries_created": 0,
                "interest_splits_created": len(interest_inserts),
                "nature_corrections": len(cell_modifications),
                "skipped_no_account": len(skipped_no_account),
                "skipped_duplicate": len(skipped_duplicate),
                "total_transactions_scanned": len(transactions),
            })

        # â"€â"€ Step 4: Write changes â"€â"€
        emit("step_start", "writing",
             f"Creating {len(counter_entries)} counter entries"
             + (f", {len(interest_inserts)} interest splits" if interest_inserts else "")
             + "...", percentage=60)
        await asyncio.sleep(0.3)

        # Phase A: Apply cell modifications (debit splits + nature corrections)
        # Uses original row numbers (before any insertion shifts rows)
        if cell_modifications:
            try:
                writer.modify_cell_values(cell_modifications)
                logger.info(f"Modified {len(cell_modifications)} rows (debit splits + nature fixes)")
            except Exception as e:
                logger.warning(f"Failed to modify cells for splits/nature: {e}")

        # Phase B: Insert interest rows directly below their settlement rows
        # Returns mapping {original_row: new_row} for all data rows
        row_mapping = {}
        if interest_inserts:
            try:
                row_mapping = writer.insert_rows_after(interest_inserts)
                logger.info(f"Inserted {len(interest_inserts)} interest rows, mapping has {len(row_mapping)} entries")
            except Exception as e:
                logger.warning(f"Failed to insert interest rows: {e}")

        # Adjust tracked row indices using the mapping (insertions shifted rows)
        if row_mapping:
            original_row_indices = [row_mapping.get(r, r) for r in original_row_indices]

        # Phase C: Append ONLY counter entries at the end
        rows_added, total_rows = writer.append_transactions(counter_entries)

        # Calculate counter entry row indices for highlighting
        append_start_row = total_rows - rows_added + 4
        counter_row_indices = list(range(
            append_start_row, append_start_row + len(counter_entries)
        ))

        # Phase D: Add missing saving accounts to Acc_Char
        acc_char_added = 0
        if acc_char_inserts:
            from .openpyxl_handler import get_openpyxl_handler
            handler_early = get_openpyxl_handler()
            for insert_info in acc_char_inserts:
                try:
                    row_added = handler_early.add_row_to_acc_char(
                        Path(working_file),
                        account_no=insert_info["saving_acc"],
                        code=insert_info["code"],
                        entity=insert_info["entity"],
                        branch=insert_info["branch"],
                        currency=insert_info["currency"],
                        account_type=insert_info["account_type"],
                    )
                    if row_added > 0:
                        acc_char_added += 1
                except Exception as e:
                    logger.warning(f"Failed to add {insert_info['saving_acc']} to Acc_Char: {e}")
            logger.info(f"Settlement: added {acc_char_added}/{len(acc_char_inserts)} accounts to Acc_Char")

        # Highlight: original settlement rows (excluding nature="Other receipts")
        # + counter entries ONLY. "Other receipts" rows are NEVER highlighted.
        all_highlight_rows = original_row_indices + counter_row_indices
        try:
            writer.highlight_settlement_rows(all_highlight_rows)
        except Exception as e:
            logger.warning(f"Failed to highlight settlement rows: {e}")

        emit("step_complete", "writing",
             f"Created {len(counter_entries)} counter entries"
             + (f", {len(interest_inserts)} interest splits" if interest_inserts else "")
             + (f", added {acc_char_added} accounts to Acc_Char" if acc_char_added else "")
             + f", highlighted {len(all_highlight_rows)} rows",
             percentage=80,
             data={
                 "counter_entries": len(counter_entries),
                 "interest_splits": len(interest_inserts),
                 "rows_appended": rows_added,
                 "rows_inserted": len(interest_inserts),
                 "acc_char_accounts_added": acc_char_added,
             })
        await asyncio.sleep(0.4)

        # â"€â"€ Step 5: Cleanup & Finalize â"€â"€
        emit("step_start", "cleanup", "Cleaning up zero-balance saving accounts...", percentage=80)
        await asyncio.sleep(0.3)

        from .openpyxl_handler import get_openpyxl_handler
        handler = get_openpyxl_handler()
        saving_rows_removed = 0
        # Collect settled saving account numbers from counter entries
        settled_saving_accounts = {
            str(c.account).strip() for c in counter_entries if c.account
        }
        try:
            saving_rows_removed = handler.remove_settled_saving_account_rows(
                Path(working_file), settled_saving_accounts,
            )
        except Exception as e:
            logger.warning(f"Failed to remove settled saving account rows: {e}")

        total_new_rows = rows_added + len(interest_inserts)
        if self.db_session:
            db_result = await self.db_session.execute(
                select(CashReportSessionModel).where(
                    CashReportSessionModel.session_id == session_id
                )
            )
            db_session = db_result.scalar_one_or_none()
            if db_session:
                db_session.total_transactions = int(db_session.total_transactions or 0) + total_new_rows
                metadata = dict(db_session.metadata_json or {})
                metadata["settlement_counter_entries"] = (
                    int(metadata.get("settlement_counter_entries") or 0) + len(counter_entries)
                )
                metadata.setdefault("open_new_counter_entries", int(metadata.get("open_new_counter_entries") or 0))
                metadata["settlement_interest_splits"] = (
                    int(metadata.get("settlement_interest_splits") or 0) + len(interest_inserts)
                )
                db_session.metadata_json = metadata
            await self.db_session.commit()

        nature_corrections = len([m for m in cell_modifications.values() if "I" in m])
        result = {
            "session_id": session_id,
            "status": "success",
            "message": f"Created {len(counter_entries)} counter entries"
                       + (f", {len(interest_inserts)} interest splits" if interest_inserts else "")
                       + (f", added {acc_char_added} accounts to Acc_Char" if acc_char_added else ""),
            "counter_entries_created": len(counter_entries),
            "interest_splits_created": len(interest_inserts),
            "nature_corrections": nature_corrections,
            "rows_appended": rows_added,
            "rows_inserted": len(interest_inserts),
            "total_rows_in_movement": total_rows,
            "settlement_transactions_found": len(settlement_transactions),
            "skipped_no_account": len(skipped_no_account),
            "skipped_duplicate": len(skipped_duplicate),
            "total_transactions_scanned": len(transactions),
            "saving_rows_removed": saving_rows_removed,
            "acc_char_accounts_added": acc_char_added,
        }

        emit("step_complete", "cleanup",
             f"Removed {saving_rows_removed} zero-balance rows",
             percentage=95,
             data={"saving_rows_removed": saving_rows_removed})
        await asyncio.sleep(0.4)

        # Save snapshot so this step's result can be downloaded independently
        self._save_step_snapshot(session_id, "settlement")

        emit("complete", "done",
             f"Settlement complete -- {len(counter_entries)} counter entries"
             + (f", {len(interest_inserts)} interest splits" if interest_inserts else ""),
             percentage=100, data=result)

        return _finish(result)

    async def run_open_new_automation(
        self,
        session_id: str,
        lookup_file_contents: Optional[List[bytes]] = None,
        user_id: Optional[int] = None,
        progress_callback=None,
        dry_run: bool = False,
    ) -> Dict[str, Any]:
        """
        Run "má»Ÿ má»›i" (open new saving account) automation on Movement data.

        Logic:
        1. Detect open-new transactions by GROUP B patterns (Nature = Internal transfer out)
        2. Extract or lookup saving account number from:
           - Description (HDTG SO xxx, TK xxx, etc.)
           - Lookup file(s) (VTB Saving style) if provided
        3. Create counter entry with saving account, reversed credit/debit, nature = Internal transfer in

        Args:
            session_id: The session ID
            lookup_file_contents: Optional list of lookup file bytes (.xls/.xlsx)
            user_id: Owner user ID for access control
            progress_callback: Optional callable for SSE progress events
            dry_run: If True, run detect+lookup but do NOT write to Excel.
                     Returns candidate preview data instead.

        Returns:
            Dict with results summary (or preview data when dry_run=True)
        """
        import time as _time

        _started_at = _time.monotonic()
        logger.info(
            "Open-new automation started: session_id=%s dry_run=%s lookup_files=%s has_progress_callback=%s user_id=%s",
            session_id,
            dry_run,
            len(lookup_file_contents or []),
            bool(progress_callback),
            user_id,
        )

        def _finish(result: Dict[str, Any]) -> Dict[str, Any]:
            elapsed_ms = int((_time.monotonic() - _started_at) * 1000)
            logger.info(
                "Open-new automation finished: session_id=%s status=%s counter_entries=%s acc_char_added=%s elapsed_ms=%s",
                session_id,
                result.get("status"),
                result.get("counter_entries_created", 0),
                result.get("acc_char_added", 0),
                elapsed_ms,
            )
            return result

        def emit(event_type, step, message, detail="", percentage=0, data=None):
            self._emit_progress(
                process="open_new",
                progress_callback=progress_callback,
                session_id=session_id,
                event_type=event_type,
                step=step,
                message=message,
                detail=detail,
                percentage=percentage,
                data=data,
            )

        if user_id is not None:
            await self._verify_session_owner(session_id, user_id)

        _, working_file = await self._ensure_working_file(session_id)

        writer = MovementDataWriter(working_file)
        self._refresh_detection_pattern_indexes_if_needed()

        # â"€â"€ Step 1: Scan Movement data â"€â"€
        emit("step_start", "scanning", "Scanning Movement data for open-new transactions...", percentage=0)
        await asyncio.sleep(0.3)

        transactions = writer.get_all_transactions()
        if not transactions:
            emit("complete", "done", "No transactions found", percentage=100,
                 data={"status": "no_transactions"})
            return _finish({
                "session_id": session_id,
                "status": "no_transactions",
                "message": "No transactions found in Movement sheet",
                "counter_entries_created": 0,
            })

        # Build current_account -> entity mapping from Cash Balance sheet
        # (same approach as settlement -- avoids reliance on formula columns)
        account_entity_map = self._build_account_entity_map(working_file)

        emit("step_complete", "scanning",
             f"Found {len(transactions)} transactions",
             percentage=20,
             data={"total_transactions": len(transactions)})
        await asyncio.sleep(0.4)

        # â"€â"€ Step 2: Detect open-new transactions â"€â"€
        emit("step_start", "detecting", "Detecting open-new transactions (GROUP B patterns)...", percentage=20)
        await asyncio.sleep(0.3)

        open_new_transactions = []
        _on_skip = {"no_credit": 0, "wrong_nature": 0, "no_pattern": 0}
        for idx, tx in enumerate(transactions):
            row_idx = idx + 4
            if self._is_open_new_candidate(tx):
                open_new_transactions.append((tx, row_idx))
            else:
                if not tx.credit or tx.credit <= 0:
                    _on_skip["no_credit"] += 1
                elif (tx.nature or "").strip().lower() != "internal transfer out":
                    _on_skip["wrong_nature"] += 1
                else:
                    _on_skip["no_pattern"] += 1
        logger.info(f"Open-new detect: {len(open_new_transactions)} candidates, skip reasons: {_on_skip}")

        if not open_new_transactions:
            saving_lookup_sync = {
                "rows_checked": 0,
                "rows_matched_lookup": 0,
                "rows_updated": 0,
                "term_months_updated": 0,
                "maturity_date_updated": 0,
                "interest_rate_updated": 0,
                "provider_mismatch_skipped": 0,
            }

            # Even when there is no open-new candidate, allow lookup-based
            # whole-sheet sync for Saving Account term/maturity/rate.
            if lookup_file_contents and not dry_run:
                emit(
                    "step_start",
                    "sync_lookup",
                    "No open-new candidates. Syncing Saving Account from lookup metadata...",
                    percentage=80,
                )
                await asyncio.sleep(0.2)
                try:
                    lookup_only_details: Dict[str, Dict[str, Any]] = {}
                    for file_content in lookup_file_contents:
                        _, parsed_details = self._parse_saving_lookup_file_with_metadata(file_content)
                        for acc, meta in parsed_details.items():
                            if not acc:
                                continue
                            dst = lookup_only_details.setdefault(acc, {})
                            for mk, mv in meta.items():
                                if mv in (None, ""):
                                    continue
                                if dst.get(mk) in (None, ""):
                                    dst[mk] = mv

                    saving_lookup_sync = await self._sync_saving_account_lookup_fields(
                        Path(working_file),
                        lookup_only_details,
                    )
                except Exception as e:
                    logger.warning(f"Saving Account lookup sync failed (no_open_new path): {e}")

                emit(
                    "step_complete",
                    "sync_lookup",
                    f"Saving Account sync updated {saving_lookup_sync.get('rows_updated', 0)} rows",
                    percentage=95,
                    data={"saving_account_lookup_sync": saving_lookup_sync},
                )
                await asyncio.sleep(0.2)

            emit("complete", "done", "No open-new transactions found", percentage=100,
                 data={"status": "no_open_new", "total_transactions_scanned": len(transactions)})
            return _finish({
                "session_id": session_id,
                "status": "no_open_new",
                "message": "No open-new transactions found (GROUP B patterns)",
                "counter_entries_created": 0,
                "total_transactions_scanned": len(transactions),
                "saving_account_lookup_sync": saving_lookup_sync,
            })

        emit("step_complete", "detecting",
             f"Found {len(open_new_transactions)} open-new transactions",
             percentage=40,
             data={"open_new_found": len(open_new_transactions)})
        await asyncio.sleep(0.4)

        # â"€â"€ Step 3: Load lookup data if provided â"€â"€
        emit("step_start", "lookup", "Looking up saving accounts...", percentage=40)
        await asyncio.sleep(0.3)

        lookup_accounts: Dict[Tuple[str, float], List[str]] = {}
        lookup_account_details: Dict[str, Dict[str, Any]] = {}
        if lookup_file_contents:
            for i, file_content in enumerate(lookup_file_contents):
                parsed, parsed_details = self._parse_saving_lookup_file_with_metadata(file_content)
                loaded_count = 0
                for key, accounts in parsed.items():
                    if not accounts:
                        continue
                    bucket = lookup_accounts.setdefault(key, [])
                    for account in accounts:
                        if account not in bucket:
                            bucket.append(account)
                            loaded_count += 1
                for acc, meta in parsed_details.items():
                    if not acc:
                        continue
                    dst = lookup_account_details.setdefault(acc, {})
                    for mk, mv in meta.items():
                        if mv in (None, ""):
                            continue
                        if dst.get(mk) in (None, ""):
                            dst[mk] = mv
                logger.info(f"Lookup file {i+1}: loaded {loaded_count} accounts")
            total_lookup_accounts = sum(len(v) for v in lookup_accounts.values())
            logger.info(
                f"Total lookup accounts from {len(lookup_file_contents)} file(s): "
                f"{total_lookup_accounts} across {len(lookup_accounts)} (entity,amount) keys"
            )
        lookup_providers = {
            str(meta.get("provider") or "").strip().upper()
            for meta in lookup_account_details.values()
            if str(meta.get("provider") or "").strip()
        }
        if lookup_providers:
            logger.info(
                "Open-new lookup providers loaded: %s",
                ", ".join(sorted(lookup_providers)),
            )

        # Load existing saving accounts to filter out non-new accounts
        saving_accounts = self._read_saving_accounts(working_file)
        existing_saving_acc_set = {sa["account"] for sa in saving_accounts}
        # Map account → old interest rate (for rate-decrease detection)
        existing_acc_rates: Dict[str, float] = {
            sa["account"]: sa["interest_rate"]
            for sa in saving_accounts
            if sa.get("interest_rate") is not None
        }
        logger.info(f"Existing saving accounts in template: {len(existing_saving_acc_set)}")

        # Load full Acc_Char data for suffix search (current_account_1, _2...)
        acc_char_data = self._read_acc_char_full_data(working_file)

        # Load lookup maps for CODE and BRANCH determination
        # 1) account->code from Acc_Char B->C (reliable fallback)
        account_to_code = self._read_acc_char_account_to_code(working_file)
        # 2) entity->code from Def sheet (primary CODE source per user spec)
        def_entity_to_code = self._read_def_entity_to_code(working_file)
        # 3) bank alias->bank name from Def sheet (for BRANCH normalization)
        def_bank_alias_to_name = self._read_def_bank_alias_to_name(working_file)
        # 4) entity->branches from Cash balance (Prior period) for BRANCH
        entity_branches = self._read_prior_period_branches(working_file)
        logger.info(
            f"Lookups: {len(account_to_code)} acc->code, "
            f"{len(def_entity_to_code)} ent->code (Def), "
            f"{len(def_bank_alias_to_name)} bank aliases (Def), "
            f"{len(entity_branches)} ent->branches (Prior)",
        )

        # Check for existing counter entries to avoid duplicates
        # Counter entries have nature "Internal transfer in" (from any source)
        existing_descs = set()
        existing_counter_accounts: Dict[str, str] = {}  # desc → saving account
        for tx in transactions:
            if tx.nature and "transfer in" in tx.nature.lower():
                desc_key = tx.description.strip() if tx.description else ""
                existing_descs.add(desc_key)
                # Map description to the counter entry's account (the saving account)
                if desc_key and tx.account:
                    existing_counter_accounts[desc_key] = str(tx.account).strip()

        counter_entries = []
        counter_entry_info = []  # Additional info for B1-B3 steps
        original_row_indices = []
        skipped_no_account = []
        skipped_duplicate = []
        skipped_existing = []  # Accounts that already exist (not genuinely new)
        def _normalize_account_text(value: str) -> str:
            text = str(value or "").strip().replace("'", "")
            if not text:
                return ""
            try:
                if "." in text and float(text) == int(float(text)):
                    return str(int(float(text)))
            except (ValueError, OverflowError):
                pass
            return text

        # Normalize lookup details/account values once so downstream matching
        # and metadata lookups remain stable even if source files contain
        # numeric strings with decimal artifacts.
        normalized_lookup_account_details: Dict[str, Dict[str, Any]] = {}
        for acc, meta in lookup_account_details.items():
            acc_norm = _normalize_account_text(acc)
            if not acc_norm:
                continue
            dst = normalized_lookup_account_details.setdefault(acc_norm, {})
            for mk, mv in (meta or {}).items():
                if mv in (None, ""):
                    continue
                if dst.get(mk) in (None, ""):
                    dst[mk] = mv
        lookup_account_details = normalized_lookup_account_details

        for key, accounts in list(lookup_accounts.items()):
            normalized_accounts: List[str] = []
            for account in accounts:
                acc_norm = _normalize_account_text(account)
                if acc_norm and acc_norm not in normalized_accounts:
                    normalized_accounts.append(acc_norm)
            lookup_accounts[key] = normalized_accounts

        # Normalize existing account set once to prevent same-run duplicates.
        existing_saving_acc_set = {_normalize_account_text(acc) for acc in existing_saving_acc_set if acc}

        # Index lookup accounts by amount for fallback matching (entity may be abbreviated in template).
        def _amount_key(value: Any) -> int:
            try:
                return int(round(float(value or 0)))
            except (ValueError, TypeError):
                return 0

        lookup_accounts_by_amount: Dict[int, List[str]] = {}
        for (_, amt), accounts in lookup_accounts.items():
            amt_key = _amount_key(amt)
            if amt_key <= 0:
                continue
            bucket = lookup_accounts_by_amount.setdefault(amt_key, [])
            for account in accounts:
                acc_norm = _normalize_account_text(account)
                if acc_norm and acc_norm not in bucket:
                    bucket.append(acc_norm)

        lookup_accounts_by_opening_date: Dict[date, List[str]] = {}
        for account, meta in lookup_account_details.items():
            opening_date = meta.get("opening_date")
            if not isinstance(opening_date, date):
                continue
            acc_norm = _normalize_account_text(account)
            if not acc_norm:
                continue
            bucket = lookup_accounts_by_opening_date.setdefault(opening_date, [])
            if acc_norm not in bucket:
                bucket.append(acc_norm)

        account_like_pattern = re.compile(r"\d{8,20}")

        def _parse_raw_lookup_amount(raw: Any) -> Optional[float]:
            if raw is None or raw == "":
                return None
            if isinstance(raw, (int, float)):
                value = float(raw)
                if value <= 0:
                    return None
                if abs(value - round(value)) <= 1e-6:
                    serial = int(round(value))
                    if 20000 <= serial <= 80000:
                        return None
                return value

            text = str(raw).strip().replace(" ", "")
            if not text:
                return None

            if re.match(r"^\d{1,3}(?:\.\d{3})+(?:,\d+)?$", text):
                text = text.replace(".", "").replace(",", ".")
            else:
                text = text.replace(",", "")

            try:
                value = float(text)
            except (TypeError, ValueError):
                return None
            if value <= 0:
                return None
            return value

        def _parse_raw_lookup_date(raw: Any) -> Optional[date]:
            if raw is None or raw == "":
                return None
            if isinstance(raw, datetime):
                return raw.date()
            if isinstance(raw, date):
                return raw
            if isinstance(raw, (int, float)):
                try:
                    serial = int(float(raw))
                    if 20000 <= serial <= 80000:
                        base_ord = datetime(1899, 12, 30).toordinal()
                        return datetime.fromordinal(base_ord + serial).date()
                except Exception:
                    return None
                return None

            text = str(raw).strip()
            if not text:
                return None
            text = text.split()[0]
            if text.replace(".", "", 1).isdigit():
                try:
                    serial = int(float(text))
                    if 20000 <= serial <= 80000:
                        base_ord = datetime(1899, 12, 30).toordinal()
                        return datetime.fromordinal(base_ord + serial).date()
                except Exception:
                    pass
            for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y", "%Y/%m/%d", "%Y-%m-%d", "%d/%m/%y", "%d-%m-%y"):
                try:
                    return datetime.strptime(text, fmt).date()
                except ValueError:
                    continue
            return None

        def _extract_accounts_from_raw_cell(raw: Any) -> List[str]:
            accounts: List[str] = []
            if raw is None or raw == "":
                return accounts

            if isinstance(raw, (int, float)):
                try:
                    value = float(raw)
                    if abs(value - int(value)) <= 1e-6:
                        token = str(int(value))
                        if account_like_pattern.fullmatch(token):
                            accounts.append(token)
                except (TypeError, ValueError, OverflowError):
                    pass

            text = str(raw).strip().replace("'", "")
            if text:
                for token in account_like_pattern.findall(text):
                    if token not in accounts:
                        accounts.append(token)
            return accounts

        def _collect_raw_lookup_rows(file_content: bytes) -> List[Dict[str, Any]]:
            rows: List[Dict[str, Any]] = []

            def _append_row(cells: List[Any]) -> None:
                text_parts: List[str] = []
                accounts: List[str] = []
                dates: List[date] = []
                amounts: List[float] = []

                for cell in cells:
                    if cell in (None, ""):
                        continue
                    cell_text = str(cell).strip()
                    if not cell_text:
                        continue
                    text_parts.append(cell_text.upper())

                    for account in _extract_accounts_from_raw_cell(cell):
                        account_norm = _normalize_account_text(account)
                        if account_norm and account_norm not in accounts:
                            accounts.append(account_norm)

                    parsed_date = _parse_raw_lookup_date(cell)
                    if parsed_date and parsed_date not in dates:
                        dates.append(parsed_date)

                    parsed_amount = _parse_raw_lookup_amount(cell)
                    if parsed_amount is not None and parsed_amount not in amounts:
                        amounts.append(parsed_amount)

                if not accounts:
                    return

                rows.append({
                    "accounts": accounts,
                    "dates": dates,
                    "amounts": amounts,
                    "text": " ".join(text_parts),
                })

            # Try xlsx
            try:
                import io
                import openpyxl

                wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True, read_only=True)
                for ws in wb.worksheets:
                    for row in ws.iter_rows(values_only=True):
                        _append_row(list(row))
                if rows:
                    return rows
            except Exception:
                pass

            # Try xls
            try:
                import xlrd

                wb = xlrd.open_workbook(file_contents=file_content)
                for ws in wb.sheets():
                    for row_idx in range(ws.nrows):
                        row_values = [ws.cell_value(row_idx, c) for c in range(ws.ncols)]
                        _append_row(row_values)
                if rows:
                    return rows
            except Exception:
                pass

            # Try PDF (Woori-style)
            try:
                parsed_pdf = self._parse_woori_pdf_lookup(file_content)
                for entry in parsed_pdf:
                    _append_row([
                        entry.get("account"),
                        entry.get("entity"),
                        entry.get("amount"),
                        entry.get("opening_date"),
                        entry.get("maturity_date"),
                        entry.get("rate"),
                        entry.get("currency"),
                        entry.get("term"),
                    ])
            except Exception:
                pass

            return rows

        raw_lookup_rows: List[Dict[str, Any]] = []
        if lookup_file_contents:
            for file_content in lookup_file_contents:
                raw_lookup_rows.extend(_collect_raw_lookup_rows(file_content))

        entity_noise_tokens = {
            "CT", "CTY", "CONG", "TY", "TNHH", "MTV", "CP", "CO", "PHAN", "VA",
            "PHAT", "TRIEN", "NGHIEP", "DAU", "TU", "DU", "AN", "MOT", "THANH",
            "VIEN", "HUU", "HAN", "JSC", "PTCN", "BW",
        }

        def _normalize_entity_text(value: str) -> str:
            text = str(value or "").upper()
            text = text.replace("BWID", "BW")
            text = re.sub(r'[^A-Z0-9]+', ' ', text)
            tokens = [tok for tok in text.split() if tok and tok not in entity_noise_tokens]
            return "".join(tokens)

        def _entity_similarity_score(left: str, right: str) -> float:
            left_norm = _normalize_entity_text(left)
            right_norm = _normalize_entity_text(right)
            if not left_norm or not right_norm:
                return 0.0
            if left_norm == right_norm:
                return 1.0
            base = SequenceMatcher(None, left_norm, right_norm).ratio()
            if left_norm in right_norm or right_norm in left_norm:
                base = max(base, 0.85)
            return base

        def _available_account(candidate: str) -> bool:
            acc_norm = _normalize_account_text(candidate)
            if not acc_norm:
                return False
            return True

        def _select_lookup_account(candidates: List[str]) -> Optional[str]:
            """
            Pick a lookup account, preferring genuinely new accounts first.
            If only existing accounts are available, still return one so caller
            can create a suffixed variant based on lookup account (not current account).
            """
            normalized = []
            for candidate in candidates:
                acc_norm = _normalize_account_text(candidate)
                if acc_norm and acc_norm not in normalized:
                    normalized.append(acc_norm)

            # Pass 1: prefer non-existing accounts.
            for acc_norm in normalized:
                if not _available_account(acc_norm):
                    continue
                if acc_norm in existing_saving_acc_set:
                    continue
                return acc_norm

            # Pass 2: allow existing accounts (caller will suffix if needed).
            for acc_norm in normalized:
                if _available_account(acc_norm):
                    return acc_norm
            return None

        def _opening_date_match(
            candidates: List[str],
            tx_date: Optional[date],
            *,
            strict: bool,
        ) -> Optional[List[str]]:
            if not tx_date:
                return None if strict else candidates
            matched = [
                acc for acc in candidates
                if lookup_account_details.get(acc, {}).get("opening_date") == tx_date
            ]
            if matched:
                return matched
            return None if strict else candidates

        def _bank_match(
            candidates: List[str],
            tx_bank: str,
            *,
            strict: bool,
        ) -> Optional[List[str]]:
            bank_norm = str(tx_bank or "").strip().upper()
            if not bank_norm:
                return candidates
            matched = [
                acc for acc in candidates
                if self._bank_matches(
                    bank_norm,
                    str(lookup_account_details.get(acc, {}).get("provider") or "").strip().upper(),
                )
            ]
            if matched:
                return matched
            return None if strict else candidates

        def _lookup_amount_for_account(account: str) -> Optional[float]:
            meta = lookup_account_details.get(account, {})
            raw_amount = meta.get("amount")
            try:
                amount_value = float(raw_amount)
            except (TypeError, ValueError):
                return None
            return amount_value if amount_value > 0 else None

        def _filter_candidates_by_lookup_amount(
            candidates: List[str],
            tx_amount: float,
        ) -> List[str]:
            """
            Amount selection rule for open-new lookup:
            1. Prefer exact amount match (lookup == movement).
            2. If no exact match, accept lookup amount > Movement amount.
               Pick the closest greater amount (smallest lookup-transaction gap).
            """
            if not candidates:
                return []
            if tx_amount <= 0:
                return []

            exact_matches = []
            greater_than = []
            for acc in candidates:
                lookup_amount = _lookup_amount_for_account(acc)
                if lookup_amount is None:
                    continue
                if abs(lookup_amount - tx_amount) <= 1:
                    exact_matches.append(acc)
                    continue
                if lookup_amount > tx_amount + 1:
                    greater_than.append((lookup_amount - tx_amount, acc))

            if exact_matches:
                return exact_matches
            if not greater_than:
                return []

            min_gap = min(item[0] for item in greater_than)
            return [acc for gap, acc in greater_than if abs(gap - min_gap) <= 1]

        def _filter_candidates_by_entity(
            candidates: List[str],
            entity_name: str,
            *,
            require_match: bool,
        ) -> List[str]:
            """
            Narrow candidates to the same entity when lookup rows provide entity text.
            If entity cannot be matched confidently, return empty list when required.
            """
            if not candidates:
                return []
            unique_candidates = list(dict.fromkeys([acc for acc in candidates if acc]))
            if not unique_candidates or not entity_name:
                return unique_candidates

            scored: List[Tuple[float, str]] = []
            has_lookup_entity = False
            for acc in unique_candidates:
                lookup_entity = str(lookup_account_details.get(acc, {}).get("entity") or "")
                if _normalize_entity_text(lookup_entity):
                    has_lookup_entity = True
                score = _entity_similarity_score(entity_name, lookup_entity)
                scored.append((score, acc))

            if not has_lookup_entity:
                return unique_candidates

            scored.sort(key=lambda item: item[0], reverse=True)
            top_score = scored[0][0] if scored else 0.0
            if top_score <= 0:
                return [] if require_match else unique_candidates

            # Keep a compact band near top score to avoid drifting to another entity.
            cutoff = max(0.35, top_score - 0.12)
            filtered = [acc for score, acc in scored if score >= cutoff and score > 0]
            if filtered:
                return filtered
            return [] if require_match else unique_candidates

        def _select_by_entity_similarity(candidates: List[str], entity_name: str) -> Optional[str]:
            if not candidates:
                return None
            if len(candidates) == 1 or not entity_name:
                return _select_lookup_account(candidates)
            entity_candidates = _filter_candidates_by_entity(
                candidates,
                entity_name,
                require_match=False,
            )
            if entity_candidates:
                selected = _select_lookup_account(entity_candidates)
                if selected:
                    return selected
            return _select_lookup_account(candidates)

        def _match_lookup_by_opening_date(
            *,
            tx_bank: str,
            tx_date: Optional[date],
            amount_value: float,
            entity_name: str,
            term_info: Optional[Dict[str, Any]] = None,
        ) -> Optional[str]:
            """
            Primary lookup for open-new:
            - Opening date must match.
            - Amount selection:
              * exact amount first
              * else closest amount > Movement amount
            """
            if not tx_date:
                return None
            date_candidates = lookup_accounts_by_opening_date.get(tx_date, [])
            candidates = [
                _normalize_account_text(acc)
                for acc in date_candidates
                if _available_account(acc)
            ]
            candidates = list(dict.fromkeys([acc for acc in candidates if acc]))
            if not candidates:
                return None
            bank_candidates = _bank_match(candidates, tx_bank, strict=True)
            if not bank_candidates:
                return None
            candidates = bank_candidates

            if entity_name:
                entity_candidates = _filter_candidates_by_entity(
                    candidates,
                    entity_name,
                    require_match=True,
                )
                if not entity_candidates:
                    return None
                candidates = entity_candidates

            amount_candidates = _filter_candidates_by_lookup_amount(candidates, amount_value)
            if not amount_candidates:
                return None
            candidates = amount_candidates

            info = term_info or {}
            maturity = info.get("maturity_date")
            if maturity:
                maturity_candidates = [
                    acc for acc in candidates
                    if lookup_account_details.get(acc, {}).get("maturity_date") == maturity
                ]
                if maturity_candidates:
                    candidates = maturity_candidates

            term_days = info.get("term_days")
            if term_days:
                day_candidates = [
                    acc for acc in candidates
                    if lookup_account_details.get(acc, {}).get("term_days") == term_days
                ]
                if day_candidates:
                    candidates = day_candidates

            term_months = info.get("term_months")
            if term_months:
                month_candidates = [
                    acc for acc in candidates
                    if lookup_account_details.get(acc, {}).get("term_months") == term_months
                ]
                if month_candidates:
                    candidates = month_candidates

            rate = info.get("interest_rate")
            if rate is not None:
                rate_candidates = []
                for acc in candidates:
                    lookup_rate = lookup_account_details.get(acc, {}).get("interest_rate")
                    if lookup_rate is None:
                        continue
                    if abs(float(lookup_rate) - float(rate)) <= 0.0005:
                        rate_candidates.append(acc)
                if rate_candidates:
                    candidates = rate_candidates

            return _select_by_entity_similarity(candidates, entity_name)

        def _has_lookup_for_bank(tx_bank: str) -> bool:
            """Check whether uploaded lookup files contain data for tx_bank."""
            if not lookup_providers:
                return False
            bank_norm = str(tx_bank or "").strip().upper()
            if not bank_norm:
                return True
            return any(self._bank_matches(bank_norm, provider) for provider in lookup_providers)

        def _match_lookup_by_metadata(
            *,
            tx_bank: str,
            tx_date: Optional[date],
            amount_value: float,
            entity_name: str,
            term_info: Dict[str, Any],
            require_opening_date: bool = False,
        ) -> Optional[str]:
            """
            Fallback matcher when description has no explicit account and entity text is abbreviated.
            Uses bank + opening date + term/maturity + fuzzy entity score,
            with amount rule: exact first, else closest amount > Movement amount.
            """
            candidates = [
                _normalize_account_text(acc)
                for acc in lookup_account_details.keys()
                if _available_account(acc)
            ]
            # Keep order stable and unique.
            candidates = list(dict.fromkeys([acc for acc in candidates if acc]))
            if not candidates:
                return None

            bank_candidates = _bank_match(candidates, tx_bank, strict=False)
            if bank_candidates is None:
                return None
            candidates = bank_candidates

            date_candidates = _opening_date_match(candidates, tx_date, strict=require_opening_date)
            if date_candidates is None:
                return None
            candidates = date_candidates

            if entity_name:
                entity_candidates = _filter_candidates_by_entity(
                    candidates,
                    entity_name,
                    require_match=True,
                )
                if not entity_candidates:
                    return None
                candidates = entity_candidates

            amount_candidates = _filter_candidates_by_lookup_amount(candidates, amount_value)
            if not amount_candidates:
                return None
            candidates = amount_candidates

            maturity = term_info.get("maturity_date")
            if maturity:
                maturity_candidates = [
                    acc for acc in candidates
                    if lookup_account_details.get(acc, {}).get("maturity_date") == maturity
                ]
                if maturity_candidates:
                    candidates = maturity_candidates

            term_days = term_info.get("term_days")
            if term_days:
                day_candidates = [
                    acc for acc in candidates
                    if lookup_account_details.get(acc, {}).get("term_days") == term_days
                ]
                if day_candidates:
                    candidates = day_candidates

            term_months = term_info.get("term_months")
            if term_months:
                month_candidates = [
                    acc for acc in candidates
                    if lookup_account_details.get(acc, {}).get("term_months") == term_months
                ]
                if month_candidates:
                    candidates = month_candidates

            return _select_by_entity_similarity(candidates, entity_name)

        def _match_lookup_by_raw_cells(
            *,
            tx_bank: str,
            tx_date: Optional[date],
            amount_value: float,
            entity_name: str,
        ) -> Optional[str]:
            """
            Rescue matcher: scan all raw lookup cells (not only structured columns).
            This is used before suffix fallback so current_account_N is created only
            when no account can be found from any lookup cell by date+amount.
            Amount rule remains: exact first, else closest amount > Movement amount.
            """
            if not raw_lookup_rows or not tx_date or amount_value <= 0:
                return None

            bank_norm = str(tx_bank or "").strip().upper()
            ranked_rows: List[Tuple[int, float, float, List[str]]] = []
            min_entity_score = 0.35

            for row in raw_lookup_rows:
                row_dates = row.get("dates") or []
                if tx_date not in row_dates:
                    continue

                row_accounts = [_normalize_account_text(acc) for acc in (row.get("accounts") or [])]
                row_accounts = list(dict.fromkeys([acc for acc in row_accounts if acc and _available_account(acc)]))
                if not row_accounts:
                    continue

                # If row carries known provider(s), at least one must match tx bank.
                if bank_norm:
                    row_has_known_provider = False
                    row_has_matching_provider = False
                    filtered_accounts: List[str] = []
                    for acc in row_accounts:
                        provider = str(lookup_account_details.get(acc, {}).get("provider") or "").strip().upper()
                        if provider:
                            row_has_known_provider = True
                            if self._bank_matches(bank_norm, provider):
                                row_has_matching_provider = True
                                filtered_accounts.append(acc)
                        else:
                            filtered_accounts.append(acc)
                    if row_has_known_provider and not row_has_matching_provider:
                        continue
                    row_accounts = list(dict.fromkeys(filtered_accounts))
                    if not row_accounts:
                        continue

                row_amounts: List[float] = []
                for raw_amount in (row.get("amounts") or []):
                    try:
                        amount_float = float(raw_amount)
                    except (TypeError, ValueError):
                        continue
                    if amount_float > 0:
                        row_amounts.append(amount_float)
                if not row_amounts:
                    continue

                exact = any(abs(row_amount - amount_value) <= 1 for row_amount in row_amounts)
                gt_gaps = [
                    row_amount - amount_value
                    for row_amount in row_amounts
                    if row_amount > amount_value + 1
                ]
                if not exact and not gt_gaps:
                    continue

                amount_priority = 0 if exact else 1
                amount_gap = 0.0 if exact else min(gt_gaps)

                entity_score = 0.0
                if entity_name:
                    row_text = str(row.get("text") or "")
                    entity_score = _entity_similarity_score(entity_name, row_text)
                    for acc in row_accounts:
                        lookup_entity = str(lookup_account_details.get(acc, {}).get("entity") or "")
                        entity_score = max(entity_score, _entity_similarity_score(entity_name, lookup_entity))
                    # Do not accept rows that cannot be tied to the same entity.
                    if entity_score < min_entity_score:
                        continue

                ranked_rows.append((amount_priority, amount_gap, -entity_score, row_accounts))

            if not ranked_rows:
                return None

            ranked_rows.sort(key=lambda item: (item[0], item[1], item[2]))
            best_key = ranked_rows[0][:3]
            pooled_accounts: List[str] = []
            for item in ranked_rows:
                if item[:3] != best_key:
                    break
                pooled_accounts.extend(item[3])

            candidates = list(dict.fromkeys([acc for acc in pooled_accounts if acc]))
            if entity_name:
                entity_candidates = _filter_candidates_by_entity(
                    candidates,
                    entity_name,
                    require_match=True,
                )
                if not entity_candidates:
                    return None
                candidates = entity_candidates

            return _select_by_entity_similarity(candidates, entity_name)

        # Track catalog-only entries: pre-existing counter entries that need
        # Cash Balance / Saving Account / Acc_Char rows but NOT new Movement rows.
        catalog_only_entries: List[Dict[str, Any]] = []

        for tx, row_idx in open_new_transactions:
            desc = tx.description.strip() if tx.description else ""
            if desc in existing_descs:
                # Counter entry already exists in Movement — but we may still need
                # to create catalog rows (Acc_Char, Saving Account, Cash Balance)
                # if the saving account hasn't been registered yet.
                pre_existing_acc = _normalize_account_text(
                    existing_counter_accounts.get(desc, "")
                )
                if pre_existing_acc and pre_existing_acc not in existing_saving_acc_set:
                    # Skip Current Account entries — they are NOT saving accounts.
                    # This happens when the "Internal transfer in" counter is a
                    # current account receiving funds from a settled time deposit
                    # (e.g., "TAT TOAN HDTG" flows: saving → current).
                    _acc_type = (
                        acc_char_data.get(pre_existing_acc, {}).get("account_type", "")
                    )
                    if _acc_type.lower() == "current account":
                        logger.info(
                            "Open-new: skipping pre-existing counter '%s' — "
                            "account %s is a Current Account, not Saving Account",
                            desc[:60], pre_existing_acc,
                        )
                        skipped_duplicate.append(desc)
                        continue
                    # Resolve entity/code/bank for catalog creation
                    _tx_acc = str(tx.account).strip() if tx.account else ""
                    _entity = account_entity_map.get(_tx_acc, "")
                    if not _entity:
                        _code = account_to_code.get(_tx_acc, "")
                        if _code:
                            for ent_name, c in def_entity_to_code.items():
                                if c == _code:
                                    _entity = ent_name
                                    break
                            if not _entity:
                                _entity = _code
                    _code = ""
                    if _entity:
                        _code = def_entity_to_code.get(_entity.upper(), "")
                        if not _code:
                            _ent_norm = re.sub(r'\s+', '', _entity.upper())
                            for ent_name, c in def_entity_to_code.items():
                                ent_n = re.sub(r'\s+', '', ent_name)
                                if ent_n == _ent_norm or ent_n in _ent_norm or _ent_norm in ent_n:
                                    _code = c
                                    break
                    if not _code:
                        _code = account_to_code.get(_tx_acc, "")
                    # Determine bank from lookup or original bank
                    _lookup_key = _normalize_account_text(pre_existing_acc)
                    _lmeta = lookup_account_details.get(_lookup_key, {})
                    _lprov = str(_lmeta.get("provider") or "").strip().upper()
                    _saving_bank = self._determine_saving_branch(
                        pre_existing_acc, _entity, entity_branches,
                        tx.bank or "", def_bank_alias_to_name,
                        provider_hint=_lprov,
                    )
                    _term_info = {"term_months": None, "term_days": None,
                                  "interest_rate": None, "maturity_date": None}
                    _currency = str(_lmeta.get("currency") or "VND").strip().upper() or "VND"
                    _no_lookup = not _lmeta
                    if _lmeta:
                        if _lmeta.get("maturity_date"):
                            _term_info["maturity_date"] = _lmeta["maturity_date"]
                        if _lmeta.get("term_months"):
                            _term_info["term_months"] = _lmeta["term_months"]
                        elif _lmeta.get("term_days"):
                            _term_info["term_days"] = _lmeta["term_days"]
                        if _lmeta.get("interest_rate") is not None:
                            _term_info["interest_rate"] = _lmeta["interest_rate"]
                    _opening = _lmeta.get("opening_date") or tx.date
                    catalog_only_entries.append({
                        "saving_acc": pre_existing_acc,
                        "entity": _entity,
                        "bank": _saving_bank,
                        "code": _code,
                        "amount": float(tx.credit) if tx.credit else 0,
                        "currency": _currency,
                        "term_info": _term_info,
                        "opening_date": _opening,
                        "rate_decreased": False,
                        "account_type": "Saving Account",
                        "no_lookup_data": _no_lookup,
                    })
                    existing_saving_acc_set.add(pre_existing_acc)
                    logger.info(
                        "Open-new: pre-existing counter entry for '%s' → saving_acc=%s, "
                        "will create catalog entries only",
                        desc[:60], pre_existing_acc,
                    )
                else:
                    skipped_duplicate.append(desc)
                continue

            # Resolve entity from Cash Balance account->entity map
            tx_account = str(tx.account).strip() if tx.account else ""
            entity = account_entity_map.get(tx_account, "")

            # If still empty, try Acc_Char -> Def reverse lookup
            if not entity:
                code_for_entity = account_to_code.get(tx_account, "")
                if code_for_entity:
                    for ent_name, c in def_entity_to_code.items():
                        if c == code_for_entity:
                            entity = ent_name
                            break
                    if not entity:
                        entity = code_for_entity
                logger.debug(f"Derived entity '{entity}' from current account {tx_account}")
            amount = float(tx.credit) if tx.credit else 0
            tx_date = tx.date if isinstance(tx.date, date) else None
            term_info = self._extract_term_info(desc)

            # Try to find saving account:
            # 1. Extract from description
            saving_acc = _normalize_account_text(self._extract_saving_account_for_open_new(desc))
            matched_lookup_account = (
                saving_acc
                if saving_acc and saving_acc in lookup_account_details
                else None
            )
            tx_bank_norm = str(tx.bank or "").strip().upper()
            bank_lookup_available = _has_lookup_for_bank(tx_bank_norm)
            if tx_bank_norm and lookup_file_contents and not bank_lookup_available:
                logger.info(
                    "Open-new: no lookup provider for bank %s; skip lookup matching for this transaction",
                    tx_bank_norm,
                )

            # 2. Lookup from uploaded file
            if (
                not saving_acc
                and lookup_account_details
                and (not tx_bank_norm or bank_lookup_available)
            ):
                # 2a. Primary rule: opening date + amount from lookup file.
                saving_acc = _match_lookup_by_opening_date(
                    tx_bank=str(tx.bank or ""),
                    tx_date=tx_date,
                    amount_value=amount,
                    entity_name=entity,
                    term_info=term_info,
                )
                if not saving_acc:
                    logger.info(
                        "Open-new: no opening-date match found for bank=%s date=%s entity=%s",
                        tx.bank,
                        tx_date,
                        entity,
                    )
                else:
                    matched_lookup_account = saving_acc
                    logger.info(
                        "Open-new: matched by opening date: %s (bank=%s, date=%s, entity=%s)",
                        saving_acc,
                        tx.bank,
                        tx_date,
                        entity,
                    )

                # 2b. Try entity-aligned candidates from lookup map, then apply amount rule.
                if not saving_acc and lookup_accounts and entity:
                    entity_norm = re.sub(r'\s+', '', entity.upper())
                    entity_candidates: List[str] = []
                    for (ent, _amt), accounts in lookup_accounts.items():
                        ent_norm = re.sub(r'\s+', '', ent)
                        if not ent_norm:
                            continue
                        if ent_norm == entity_norm or ent_norm in entity_norm or entity_norm in ent_norm:
                            for acc in accounts:
                                acc_norm = _normalize_account_text(acc)
                                if acc_norm and _available_account(acc_norm) and acc_norm not in entity_candidates:
                                    entity_candidates.append(acc_norm)
                    entity_candidates = _opening_date_match(
                        entity_candidates,
                        tx_date,
                        strict=bool(tx_date),
                    ) or []
                    bank_candidates = _bank_match(entity_candidates, str(tx.bank or ""), strict=False)
                    if bank_candidates:
                        entity_candidates = list(dict.fromkeys(bank_candidates))
                    entity_candidates = _filter_candidates_by_lookup_amount(entity_candidates, amount)
                    if entity_candidates:
                        saving_acc = _select_by_entity_similarity(entity_candidates, entity)
                        if saving_acc:
                            matched_lookup_account = saving_acc

                # 2c. Try partial entity match, with opening date guard.
                if not saving_acc and lookup_accounts and entity:
                    entity_norm = re.sub(r'\s+', '', entity.upper())
                    partial_candidates: List[str] = []
                    for (ent, _amt), accounts in lookup_accounts.items():
                        ent_norm = re.sub(r'\s+', '', ent)
                        if ent_norm in entity_norm or entity_norm in ent_norm:
                            for acc in accounts:
                                acc_norm = _normalize_account_text(acc)
                                if acc_norm and _available_account(acc_norm) and acc_norm not in partial_candidates:
                                    partial_candidates.append(acc_norm)
                    partial_candidates = _opening_date_match(
                        partial_candidates,
                        tx_date,
                        strict=bool(tx_date),
                    ) or []
                    bank_candidates = _bank_match(partial_candidates, str(tx.bank or ""), strict=False)
                    if bank_candidates:
                        partial_candidates = list(dict.fromkeys(bank_candidates))
                    partial_candidates = _filter_candidates_by_lookup_amount(partial_candidates, amount)
                    if partial_candidates:
                        saving_acc = _select_by_entity_similarity(partial_candidates, entity)
                        if saving_acc:
                            matched_lookup_account = saving_acc

                # 2d. Fallback: infer from amount/bank/date/term metadata.
                if not saving_acc and lookup_account_details:
                    saving_acc = _match_lookup_by_metadata(
                        tx_bank=str(tx.bank or ""),
                        tx_date=tx_date,
                        amount_value=amount,
                        entity_name=entity,
                        term_info=term_info,
                        require_opening_date=bool(tx_date),
                    )
                    if saving_acc:
                        matched_lookup_account = saving_acc
                        logger.info(
                            f"Open-new: matched by metadata: {saving_acc} "
                            f"(bank={tx.bank}, amount={amount}, date={tx.date})"
                        )

                # 2e. Last fallback: amount only, but opening date must still match when available.
                if not saving_acc and lookup_account_details:
                    available_unique = [
                        _normalize_account_text(candidate)
                        for candidate in lookup_account_details.keys()
                        if _available_account(candidate)
                    ]
                    available_unique = list(dict.fromkeys([acc for acc in available_unique if acc]))
                    available_unique = _opening_date_match(
                        available_unique,
                        tx_date,
                        strict=bool(tx_date),
                    ) or []
                    bank_candidates = _bank_match(available_unique, str(tx.bank or ""), strict=False)
                    if bank_candidates:
                        available_unique = list(dict.fromkeys(bank_candidates))
                    if entity:
                        entity_candidates = _filter_candidates_by_entity(
                            available_unique,
                            entity,
                            require_match=True,
                        )
                        available_unique = entity_candidates or []
                    available_unique = _filter_candidates_by_lookup_amount(available_unique, amount)
                    if available_unique:
                        saving_acc = _select_by_entity_similarity(available_unique, entity)
                        if saving_acc:
                            matched_lookup_account = saving_acc
                            logger.info(f"Open-new: matched by amount only: {saving_acc}")

                # 2f. Rescue by full-cell scan across lookup files.
                # Suffix current account is only allowed after this step also fails.
                if not saving_acc and raw_lookup_rows:
                    saving_acc = _match_lookup_by_raw_cells(
                        tx_bank=str(tx.bank or ""),
                        tx_date=tx_date,
                        amount_value=amount,
                        entity_name=entity,
                    )
                    if saving_acc:
                        matched_lookup_account = _normalize_account_text(saving_acc)
                        logger.info(
                            "Open-new: matched by raw lookup cell scan: %s (bank=%s, date=%s, amount=%s, entity=%s)",
                            saving_acc,
                            tx.bank,
                            tx_date,
                            amount,
                            entity,
                        )

            # NOTE: Do NOT lookup from existing Saving Account sheet for mở mới.
            # Step 3 from settlement flow is intentionally omitted here because
            # mở mới only creates counter entries for genuinely NEW accounts.

            # Fallback: if no saving account found, create current_account_{next}.
            # Find the next available suffix: if _1 exists → _2, if _1,_2 exist → _3.
            is_current_account_fallback = False
            if not saving_acc:
                if not tx_account:
                    skipped_no_account.append(tx.description)
                    logger.warning(f"Open-new: no saving account and no current account for '{tx.description}'")
                    continue

                # Find next available suffix (_1, _2, _3, ...)
                # Check both Acc_Char and already-reserved accounts in this run
                suffix = 1
                while (f"{tx_account}_{suffix}" in acc_char_data
                       or f"{tx_account}_{suffix}" in existing_saving_acc_set):
                    suffix += 1
                saving_acc = f"{tx_account}_{suffix}"
                logger.info(f"Open-new: created suffixed account {saving_acc} (no lookup match)")
                is_current_account_fallback = True

            # If saving account already exists, create a suffixed version (_1, _2, _3, ...)
            # This handles multiple deposits to the same base account in the same period.
            if saving_acc in existing_saving_acc_set:
                base_acc = saving_acc.split('_')[0]
                suffix = 1
                while f"{base_acc}_{suffix}" in existing_saving_acc_set:
                    suffix += 1
                saving_acc = f"{base_acc}_{suffix}"
                logger.info(f"Open-new: created suffixed account {saving_acc} (base already exists)")

            # Reserve this account immediately to prevent duplicate inserts in same run.
            existing_saving_acc_set.add(saving_acc)

            lookup_meta_key = _normalize_account_text(matched_lookup_account or saving_acc)
            lookup_meta = lookup_account_details.get(lookup_meta_key, {})
            lookup_provider = str(lookup_meta.get("provider") or "").strip().upper()
            if (
                lookup_meta
                and tx_bank_norm
                and lookup_provider
                and not self._bank_matches(tx_bank_norm, lookup_provider)
            ):
                logger.info(
                    "Open-new: provider mismatch for %s (tx_bank=%s, lookup_provider=%s); treating as no lookup data",
                    lookup_meta_key or saving_acc,
                    tx_bank_norm,
                    lookup_provider,
                )
                lookup_meta = {}
            # Flag: no lookup data → insert row but leave term/rate empty, highlight yellow
            # Always True for current_account fallback (suffixed accounts never in lookup)
            no_lookup_data = is_current_account_fallback or not lookup_meta
            if no_lookup_data:
                # Clear all term info — cells will be left empty in Saving Account
                term_info = {"term_months": None, "term_days": None, "interest_rate": None, "maturity_date": None}
                logger.info(f"Open-new: no lookup data for {saving_acc}, will insert with empty term/rate and highlight yellow")
            else:
                # Lookup file is authoritative for dates and term — override
                # description-extracted values when lookup provides them.
                if lookup_meta.get("maturity_date"):
                    term_info["maturity_date"] = lookup_meta["maturity_date"]
                if lookup_meta.get("term_months"):
                    term_info["term_months"] = lookup_meta["term_months"]
                elif lookup_meta.get("term_days"):
                    term_info["term_days"] = lookup_meta["term_days"]
                # Interest rate: ONLY use lookup value. Description-extracted rates
                # are unreliable — clear them so the cell shows "check web" instead.
                if lookup_meta.get("interest_rate") is not None:
                    term_info["interest_rate"] = lookup_meta["interest_rate"]
                else:
                    term_info["interest_rate"] = None

            opening_date_for_insert = tx.date
            if lookup_meta.get("opening_date"):
                opening_date_for_insert = lookup_meta["opening_date"]

            currency_for_insert = str(lookup_meta.get("currency") or "VND").strip().upper() or "VND"

            # For missing lookup data, keep source bank intent but normalize alias
            # via Def Bank/Abb table (e.g. TCB -> TECHCOMBANK).
            if no_lookup_data:
                saving_bank = self._normalize_bank_name(tx.bank or "", def_bank_alias_to_name)
                if not saving_bank:
                    saving_bank = self._determine_saving_branch(
                        saving_acc,
                        entity,
                        entity_branches,
                        tx.bank or "",
                        def_bank_alias_to_name,
                    )
            else:
                # Determine BRANCH from Cash balance (Prior period) by entity + bank prefix.
                # Pass lookup provider as hint so we don't rely solely on account prefix
                # (e.g. Woori accounts 200700xxxxx also start with "2" like VietinBank).
                saving_bank = self._determine_saving_branch(
                    saving_acc,
                    entity,
                    entity_branches,
                    tx.bank or "",
                    def_bank_alias_to_name,
                    provider_hint=lookup_provider,
                )

            # Create counter entry: swap credit to debit, nature = Internal transfer in
            counter = MovementTransaction(
                source=tx.source or "Automation",
                bank=tx.bank,
                account=saving_acc,
                date=tx.date,
                description=tx.description or "",
                debit=tx.credit,  # Swap: original credit -> counter debit
                credit=Decimal("0"),
                nature="Internal transfer in",
            )
            counter_entries.append(counter)
            original_row_indices.append(row_idx)

            # Look up CODE from Def sheet (primary): entity -> code
            # Fallback: Acc_Char account->code using CURRENT account (tx.account)
            code = ""
            if entity:
                code = def_entity_to_code.get(entity.upper(), "")
                # Partial match with whitespace normalization (for shared-string artifacts)
                if not code:
                    entity_norm = re.sub(r'\s+', '', entity.upper())
                    for ent_name, c in def_entity_to_code.items():
                        ent_norm = re.sub(r'\s+', '', ent_name)
                        if ent_norm == entity_norm or ent_norm in entity_norm or entity_norm in ent_norm:
                            code = c
                            break
            if not code:
                current_acc = str(tx.account).strip() if tx.account else ""
                code = account_to_code.get(current_acc, "")
            if not code:
                logger.warning(f"Open-new: no CODE found for entity={entity}, current_acc={str(tx.account).strip() if tx.account else ''}")

            # Detect interest rate decrease (1-2%) vs same account's old rate
            rate_decreased = False
            new_rate = term_info.get("interest_rate")
            if new_rate is not None:
                # For suffixed accounts (e.g. 1064099857_1), check the base account
                base_acc = saving_acc.split("_")[0]
                old_rate = existing_acc_rates.get(base_acc)
                if old_rate is not None:
                    drop = old_rate - new_rate
                    if 0.01 <= drop <= 0.02:  # 1-2% decrease (stored as decimal)
                        rate_decreased = True
                        logger.info(
                            f"Open-new: rate decrease detected for {saving_acc}: "
                            f"{old_rate:.4f} → {new_rate:.4f} (drop {drop:.4f})"
                        )

            # Store additional info for B1-B3 steps
            # Even for current-account fallback, the suffixed account (e.g. 116632406789_1)
            # is treated as a Saving Account entry — it goes into all 3 sheets.
            account_type = "Saving Account"
            counter_entry_info.append({
                "saving_acc": saving_acc,
                "entity": entity,
                "bank": saving_bank,
                "code": code,
                "amount": amount,
                "currency": currency_for_insert,
                "term_info": term_info,
                "opening_date": opening_date_for_insert,
                "rate_decreased": rate_decreased,
                "account_type": account_type,
                "no_lookup_data": no_lookup_data,
            })

        emit("step_complete", "lookup",
             f"Matched {len(counter_entries)} new accounts, {len(skipped_existing)} existing, {len(skipped_no_account)} not found",
             percentage=60,
             data={
                 "matched": len(counter_entries),
                 "skipped_existing": len(skipped_existing),
                 "skipped_no_account": len(skipped_no_account),
                 "skipped_duplicate": len(skipped_duplicate),
             })
        await asyncio.sleep(0.4)

        saving_lookup_sync = {
            "rows_checked": 0,
            "rows_matched_lookup": 0,
            "rows_updated": 0,
            "term_months_updated": 0,
            "maturity_date_updated": 0,
            "interest_rate_updated": 0,
            "provider_mismatch_skipped": 0,
        }

        # ── Dry-run: return preview without writing ──
        if dry_run:
            candidates_preview = []
            for (tx, row_idx), info in zip(open_new_transactions, counter_entry_info):
                desc = (tx.description or "").strip()
                candidates_preview.append({
                    "row": row_idx,
                    "date": tx.date.isoformat() if tx.date else None,
                    "bank": tx.bank or "",
                    "account": str(tx.account or ""),
                    "description": desc,
                    "credit": float(tx.credit) if tx.credit else 0,
                    "nature": tx.nature or "",
                    "status": (
                        "duplicate" if desc in [d for d in skipped_duplicate]
                        else "no_account" if desc in skipped_no_account
                        else "matched"
                    ),
                    "saving_account": info["saving_acc"],
                    "entity": info.get("entity", ""),
                    "code": info.get("code", ""),
                    "currency": info.get("currency", "VND"),
                    "no_lookup_data": info.get("no_lookup_data", False),
                    "term_months": info.get("term_info", {}).get("term_months"),
                    "interest_rate": info.get("term_info", {}).get("interest_rate"),
                    "maturity_date": (
                        info.get("term_info", {}).get("maturity_date").isoformat()
                        if info.get("term_info", {}).get("maturity_date") else None
                    ),
                })
            # Also add skipped rows
            for desc in skipped_no_account:
                candidates_preview.append({
                    "description": desc,
                    "status": "no_account",
                    "saving_account": None,
                })
            for desc in skipped_duplicate:
                candidates_preview.append({
                    "description": desc,
                    "status": "duplicate",
                    "saving_account": None,
                })
            return _finish({
                "session_id": session_id,
                "dry_run": True,
                "status": "preview",
                "total_transactions_scanned": len(transactions),
                "open_new_candidates": len(open_new_transactions),
                "counter_entries_would_create": len(counter_entries),
                "skipped_no_account": skipped_no_account,
                "skipped_duplicate": skipped_duplicate,
                "skipped_existing": skipped_existing,
                "candidates": candidates_preview,
            })

        # Merge catalog-only entries (pre-existing counter entries that need
        # Acc_Char / Saving Account / Cash Balance rows but NOT Movement rows).
        if catalog_only_entries:
            counter_entry_info.extend(catalog_only_entries)
            logger.info(
                "Open-new: added %d catalog-only entries to counter_entry_info",
                len(catalog_only_entries),
            )

        if not counter_entries and not catalog_only_entries:
            if lookup_account_details:
                emit(
                    "step_start",
                    "sync_lookup",
                    "Syncing all Saving Account rows with lookup metadata...",
                    percentage=85,
                )
                await asyncio.sleep(0.2)
                try:
                    saving_lookup_sync = await self._sync_saving_account_lookup_fields(
                        Path(working_file),
                        lookup_account_details,
                    )
                except Exception as e:
                    logger.warning(f"Saving Account lookup sync failed (no_counter_entries path): {e}")
                emit(
                    "step_complete",
                    "sync_lookup",
                    f"Saving Account sync updated {saving_lookup_sync.get('rows_updated', 0)} rows",
                    percentage=95,
                    data={"saving_account_lookup_sync": saving_lookup_sync},
                )
                await asyncio.sleep(0.2)

            msg = "No counter entries created"
            if skipped_no_account:
                msg += f". {len(skipped_no_account)} skipped (no matching saving account)"
            if skipped_duplicate:
                msg += f". {len(skipped_duplicate)} skipped (already exists)"
            emit("complete", "done", msg, percentage=100,
                 data={"status": "no_counter_entries"})
            return _finish({
                "session_id": session_id,
                "status": "no_counter_entries",
                "message": msg,
                "counter_entries_created": 0,
                "skipped_no_account": len(skipped_no_account),
                "skipped_duplicate": len(skipped_duplicate),
                "total_transactions_scanned": len(transactions),
                "saving_account_lookup_sync": saving_lookup_sync,
            })

        # â"€â"€ Step 4: Create counter entries â"€â"€
        rows_added = 0
        if counter_entries:
            emit("step_start", "writing", f"Creating {len(counter_entries)} counter entries...", percentage=60)
            await asyncio.sleep(0.3)

            rows_added, total_rows = writer.append_transactions(counter_entries)

            # Highlight both original and counter rows
            counter_start_row = total_rows - rows_added + 4
            counter_row_indices = list(range(counter_start_row, counter_start_row + rows_added))
            all_highlight_rows = original_row_indices + counter_row_indices
            try:
                writer.highlight_open_new_rows(all_highlight_rows)
            except Exception as e:
                logger.warning(f"Failed to highlight open-new rows: {e}")

            emit("step_complete", "writing",
                 f"Created {rows_added} counter entries, highlighted {len(all_highlight_rows)} rows",
                 percentage=70,
                 data={"rows_added": rows_added})
            await asyncio.sleep(0.4)

        # â"€â"€ Step 4.5: Add rows to Acc_Char, Saving Account, Cash Balance (B1-B3) â"€â"€
        # Batch: ONE ZIP read/write for all 3 sheets
        emit("step_start", "catalog",
             "Adding new saving accounts to catalog sheets (Acc_Char, Saving Account, Cash Balance)...",
             percentage=70)
        await asyncio.sleep(0.2)

        from .openpyxl_handler import get_openpyxl_handler
        handler = get_openpyxl_handler()

        sa_acc_to_row: Dict[str, int] = {}
        try:
            batch_result = await asyncio.to_thread(
                handler.add_rows_to_sheets_batch,
                working_file,
                counter_entry_info,
            )
            acc_char_added = batch_result["acc_char_added"]
            saving_account_added = batch_result["saving_added"]
            cash_balance_added = batch_result["cash_balance_added"]
            sa_acc_to_row = batch_result.get("sa_acc_to_row", {})
        except Exception as e:
            logger.error(f"Batch add to catalog sheets failed: {e}", exc_info=True)
            acc_char_added = saving_account_added = cash_balance_added = 0

        logger.info(f"B1-B3 completed: Acc_Char +{acc_char_added}, Saving Account +{saving_account_added}, Cash Balance +{cash_balance_added}")

        # Highlight interest-rate cells in red for accounts with rate decrease
        rate_decrease_cells: List[Tuple[int, str]] = []
        for info in counter_entry_info:
            if info.get("rate_decreased") and info["saving_acc"] in sa_acc_to_row:
                rate_decrease_cells.append((sa_acc_to_row[info["saving_acc"]], "L"))
        if rate_decrease_cells:
            try:
                await asyncio.to_thread(
                    handler.highlight_cells,
                    working_file,
                    "Saving Account",
                    rate_decrease_cells,
                )
                logger.info(f"Highlighted {len(rate_decrease_cells)} rate-decrease cells in red")
            except Exception as e:
                logger.warning(f"Failed to highlight rate-decrease cells: {e}")

        # Highlight yellow: rows with no lookup data (missing term/rate info)
        no_lookup_rows: List[int] = []
        for info in counter_entry_info:
            if info.get("no_lookup_data") and info["saving_acc"] in sa_acc_to_row:
                no_lookup_rows.append(sa_acc_to_row[info["saving_acc"]])
        if no_lookup_rows:
            yellow_fill = (
                b'<fill><patternFill patternType="solid">'
                b'<fgColor rgb="FFFFFF00"/>'
                b'</patternFill></fill>'
            )
            try:
                await asyncio.to_thread(
                    handler.highlight_rows,
                    Path(working_file),
                    "Saving Account",
                    no_lookup_rows,
                    yellow_fill,
                )
                logger.info(f"Highlighted {len(no_lookup_rows)} no-lookup-data rows yellow in Saving Account")
            except Exception as e:
                logger.warning(f"Failed to highlight no-lookup-data rows yellow: {e}")

        emit("step_complete", "catalog",
             f"Added to catalogs: Acc_Char +{acc_char_added}, Saving Account +{saving_account_added}, Cash Balance +{cash_balance_added}",
             percentage=85,
             data={
                 "acc_char_added": acc_char_added,
                 "saving_account_added": saving_account_added,
                 "cash_balance_added": cash_balance_added,
             })
        await asyncio.sleep(0.3)

        # Step 4.6: sync ALL Saving Account rows by lookup metadata.
        if lookup_account_details:
            emit(
                "step_start",
                "sync_lookup",
                "Syncing all Saving Account rows with lookup metadata...",
                percentage=85,
            )
            await asyncio.sleep(0.2)
            try:
                saving_lookup_sync = await self._sync_saving_account_lookup_fields(
                    Path(working_file),
                    lookup_account_details,
                )
            except Exception as e:
                logger.warning(f"Saving Account lookup sync failed: {e}")
            emit(
                "step_complete",
                "sync_lookup",
                f"Saving Account sync updated {saving_lookup_sync.get('rows_updated', 0)} rows",
                percentage=93,
                data={"saving_account_lookup_sync": saving_lookup_sync},
            )
            await asyncio.sleep(0.2)

        # â"€â"€ Step 5: Finalize â"€â"€
        if self.db_session:
            db_result = await self.db_session.execute(
                select(CashReportSessionModel).where(
                    CashReportSessionModel.session_id == session_id
                )
            )
            db_session = db_result.scalar_one_or_none()
            if db_session:
                db_session.total_transactions = int(db_session.total_transactions or 0) + rows_added
                metadata = dict(db_session.metadata_json or {})
                metadata["open_new_counter_entries"] = (
                    int(metadata.get("open_new_counter_entries") or 0) + rows_added
                )
                metadata.setdefault("settlement_counter_entries", int(metadata.get("settlement_counter_entries") or 0))
                metadata.setdefault("settlement_interest_splits", int(metadata.get("settlement_interest_splits") or 0))
                db_session.metadata_json = metadata
            await self.db_session.commit()

        result = {
            "session_id": session_id,
            "status": "success",
            "message": f"Created {rows_added} counter entries for open-new transactions",
            "counter_entries_created": rows_added,
            "total_rows_in_movement": total_rows,
            "open_new_transactions_found": len(open_new_transactions),
            "candidates_found": len(open_new_transactions),
            "skipped_existing": len(skipped_existing),
            "skipped_no_account": len(skipped_no_account),
            "skipped_duplicate": len(skipped_duplicate),
            "total_transactions_scanned": len(transactions),
            # B1-B3 stats
            "acc_char_added": acc_char_added,
            "saving_account_added": saving_account_added,
            "cash_balance_added": cash_balance_added,
            # Whole-sheet sync stats from lookup metadata
            "saving_account_lookup_sync": saving_lookup_sync,
        }

        # Save snapshot so this step's result can be downloaded independently
        self._save_step_snapshot(session_id, "open_new")

        emit("complete", "done",
             f"Open-new complete -- {rows_added} entries, {acc_char_added} new accounts added",
             percentage=100, data=result)

        return _finish(result)

    def _parse_saving_lookup_file(self, file_content: bytes) -> Dict[Tuple[str, float], List[str]]:
        """Backward-compatible wrapper returning only entity+amount -> accounts mapping."""
        lookup, _ = self._parse_saving_lookup_file_with_metadata(file_content)
        return lookup

    def _parse_saving_lookup_file_with_metadata(
        self,
        file_content: bytes,
    ) -> Tuple[Dict[Tuple[str, float], List[str]], Dict[str, Dict[str, Any]]]:
        """
        Parse lookup file and return:
        - lookup map: ``(ENTITY_UPPER, AMOUNT) -> [ACCOUNT_NO, ...]``
        - account metadata: ``ACCOUNT_NO -> {provider, amount, maturity_date, opening_date, term, rate, ...}``
        """
        import io
        import zipfile as _zf
        import xml.etree.ElementTree as _ET

        lookup: Dict[Tuple[str, float], List[str]] = {}
        account_meta: Dict[str, Dict[str, Any]] = {}

        def _normalize_lookup_account(raw: Any) -> str:
            acc_str = str(raw or "").strip().replace("'", "")
            if not acc_str:
                return ""
            try:
                if "." in acc_str and float(acc_str) == int(float(acc_str)):
                    acc_str = str(int(float(acc_str)))
            except (ValueError, OverflowError):
                pass
            return acc_str

        def _is_probable_lookup_account(account: str) -> bool:
            """
            Keep only plausible saving account values.
            Prevent header/text noise or mapped amount values from being treated
            as account metadata.
            """
            acc = str(account or "").strip()
            if not acc:
                return False
            if not re.fullmatch(r"\d{8,20}", acc):
                return False
            return True

        def _parse_lookup_amount(raw: Any) -> Optional[float]:
            if raw is None:
                return None
            if isinstance(raw, (int, float)):
                return float(raw)
            s = str(raw).strip().replace(" ", "")
            if not s:
                return None
            if re.match(r"^\d{1,3}(?:\.\d{3})+(?:,\d+)?$", s):
                s = s.replace(".", "").replace(",", ".")
            else:
                s = s.replace(",", "")
            try:
                return float(s)
            except (ValueError, TypeError):
                return None

        def _parse_lookup_rate(raw: Any) -> Optional[float]:
            if raw is None or raw == "":
                return None
            if isinstance(raw, (int, float)):
                v = float(raw)
                if v <= 0:
                    return None
                return v / 100 if v > 1 else v
            s = str(raw).strip().replace("%", "").replace(",", ".")
            if not s:
                return None
            try:
                v = float(s)
            except (ValueError, TypeError):
                return None
            if v <= 0:
                return None
            return v / 100 if v > 1 else v

        def _prefer_lookup_amount_raw(*raw_values: Any) -> Any:
            """
            Select the first positive amount-like raw value.
            Used for bank exports that expose multiple amount columns
            (e.g., principal/opening amount and current balance).
            """
            fallback = None
            for raw in raw_values:
                if fallback is None and raw not in (None, ""):
                    fallback = raw
                amount_val = _parse_lookup_amount(raw)
                if amount_val is not None and amount_val > 0:
                    return raw
            return fallback

        def _parse_lookup_date(raw: Any) -> Optional[date]:
            if raw is None or raw == "":
                return None
            if isinstance(raw, datetime):
                return raw.date()
            if isinstance(raw, date):
                return raw
            if isinstance(raw, (int, float)):
                try:
                    serial = int(float(raw))
                    if 20000 <= serial <= 80000:
                        base_ord = datetime(1899, 12, 30).toordinal()
                        return datetime.fromordinal(base_ord + serial).date()
                except Exception:
                    return None
                return None

            s = str(raw).strip()
            if not s:
                return None
            s = s.split()[0]
            # Handle string-encoded Excel serial numbers (e.g., "46051" from XML)
            if s.replace('.', '', 1).isdigit():
                try:
                    serial = int(float(s))
                    if 20000 <= serial <= 80000:
                        base_ord = datetime(1899, 12, 30).toordinal()
                        return datetime.fromordinal(base_ord + serial).date()
                except Exception:
                    pass
            for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y", "%Y/%m/%d", "%Y-%m-%d", "%d/%m/%y", "%d-%m-%y"):
                try:
                    return datetime.strptime(s, fmt).date()
                except ValueError:
                    continue
            return None

        def _parse_term_info(term_raw: Any, product_raw: Any = None) -> Tuple[Optional[int], Optional[int]]:
            term_months = None
            term_days = None

            def _consume(text: str) -> None:
                nonlocal term_months, term_days
                if not text:
                    return
                up = text.upper()
                m_month = re.search(r'(\d{1,2})\s*(THANG|THÁNG|MONTHS?|M)\b', up)
                if m_month:
                    term_months = int(m_month.group(1))
                    return
                m_day = re.search(r'(\d{1,3})\s*(NGAY|NGÀY|DAYS?|D)\b', up)
                if m_day:
                    term_days = int(m_day.group(1))

            _consume(str(term_raw or ""))
            if term_months is None and term_days is None:
                _consume(str(product_raw or ""))
            return term_months, term_days

        def _merge_meta(
            account: str,
            *,
            entity: str = "",
            currency: str = "",
            provider: str = "",
            amount: Optional[float] = None,
            opening_date: Optional[date] = None,
            maturity_date: Optional[date] = None,
            interest_rate: Optional[float] = None,
            term_months: Optional[int] = None,
            term_days: Optional[int] = None,
        ) -> None:
            if not account:
                return
            meta = account_meta.setdefault(account, {
                "entity": "",
                "currency": "",
                "provider": "",
                "amount": None,
                "opening_date": None,
                "maturity_date": None,
                "interest_rate": None,
                "term_months": None,
                "term_days": None,
            })
            if entity and not meta["entity"]:
                meta["entity"] = entity
            if currency and not meta["currency"]:
                meta["currency"] = currency
            if provider and not meta["provider"]:
                meta["provider"] = provider
            if amount is not None and amount > 0 and meta["amount"] is None:
                meta["amount"] = amount
            if opening_date and not meta["opening_date"]:
                meta["opening_date"] = opening_date
            if maturity_date and not meta["maturity_date"]:
                meta["maturity_date"] = maturity_date
            if interest_rate is not None and meta["interest_rate"] is None:
                meta["interest_rate"] = interest_rate
            if term_months and not meta["term_months"]:
                meta["term_months"] = term_months
            if term_days and not meta["term_days"]:
                meta["term_days"] = term_days

        def _add_lookup_account(
            entity_raw: Any,
            amount_raw: Any,
            account_raw: Any,
            *,
            term_raw: Any = None,
            rate_raw: Any = None,
            opening_date_raw: Any = None,
            maturity_date_raw: Any = None,
            currency_raw: Any = None,
            product_raw: Any = None,
            provider_raw: Any = None,
        ) -> None:
            acc_str = _normalize_lookup_account(account_raw)
            if not _is_probable_lookup_account(acc_str):
                return
            ent_str = str(entity_raw or "").strip().upper()
            amount_val = _parse_lookup_amount(amount_raw)
            if amount_val is None or amount_val <= 0:
                return
            currency_str = str(currency_raw or "").strip().upper()
            provider_str = str(provider_raw or "").strip().upper()
            opening_date_val = _parse_lookup_date(opening_date_raw)
            maturity_date_val = _parse_lookup_date(maturity_date_raw)
            interest_rate_val = _parse_lookup_rate(rate_raw)
            term_months, term_days = _parse_term_info(term_raw, product_raw)

            # Compute term_months from opening/maturity dates when not
            # explicitly provided (VCB lookup files have dates but no term column).
            if term_months is None and term_days is None and opening_date_val and maturity_date_val:
                # Count complete calendar months between the two dates.
                m_diff = (maturity_date_val.year - opening_date_val.year) * 12 + (maturity_date_val.month - opening_date_val.month)
                if maturity_date_val.day < opening_date_val.day:
                    m_diff -= 1  # not a full month yet
                if m_diff > 0:
                    term_months = m_diff
                elif m_diff == 0 and maturity_date_val > opening_date_val:
                    # Less than a month apart → compute days instead
                    term_days = (maturity_date_val - opening_date_val).days

            _merge_meta(
                acc_str,
                entity=ent_str,
                currency=currency_str,
                provider=provider_str,
                amount=amount_val,
                opening_date=opening_date_val,
                maturity_date=maturity_date_val,
                interest_rate=interest_rate_val,
                term_months=term_months,
                term_days=term_days,
            )

            if not ent_str:
                return

            key = (ent_str, amount_val)
            bucket = lookup.setdefault(key, [])
            if acc_str not in bucket:
                bucket.append(acc_str)

        # Try .xlsx
        try:
            zdata = io.BytesIO(file_content)
            with _zf.ZipFile(zdata) as z:
                if "xl/workbook.xml" not in z.namelist():
                    raise ValueError("Not xlsx")

                shared_strings = []
                if "xl/sharedStrings.xml" in z.namelist():
                    ss_xml = z.read("xl/sharedStrings.xml")
                    for m in re.finditer(rb'<si[^>]*>.*?</si>', ss_xml, re.DOTALL):
                        parts = re.findall(rb'<t[^>]*>([^<]*)</t>', m.group(0))
                        shared_strings.append("".join(
                            p.decode("utf-8", errors="replace") for p in parts
                        ))

                worksheet_paths = sorted(
                    n for n in z.namelist()
                    if n.startswith("xl/worksheets/sheet") and n.endswith(".xml")
                )
                if not worksheet_paths:
                    raise ValueError("No worksheet XML found")
                sheet_xml = z.read(worksheet_paths[0])

            ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
            root = _ET.fromstring(sheet_xml)
            sheet_data = root.find(f"{{{ns}}}sheetData")
            if sheet_data is None:
                raise ValueError("No sheetData")

            for row_el in sheet_data:
                row: Dict[str, Any] = {}
                for cell in row_el:
                    ref = cell.get("r", "")
                    col = re.match(r"([A-Z]+)", ref)
                    if not col:
                        continue
                    col_letter = col.group(1)
                    t = cell.get("t", "")
                    v_el = cell.find(f"{{{ns}}}v")
                    val = v_el.text if v_el is not None else None
                    is_el = cell.find(f"{{{ns}}}is")
                    if is_el is not None:
                        t_el = is_el.find(f"{{{ns}}}t")
                        val = t_el.text if t_el is not None else ""
                    elif t == "s" and val and shared_strings:
                        try:
                            idx = int(val)
                            val = shared_strings[idx] if idx < len(shared_strings) else val
                        except (ValueError, TypeError):
                            pass
                    row[col_letter] = val

                # VCB style
                _add_lookup_account(
                    row.get("C"), row.get("E"), row.get("D"),
                    currency_raw=row.get("F"),
                    opening_date_raw=row.get("G"),
                    maturity_date_raw=row.get("H"),
                    product_raw=row.get("I"),
                    provider_raw="VCB",
                )
                # VCB alternate layout (common export):
                # B=entity, C=account, D=amount, E=currency, F=open, G=maturity, H=term/rate note
                _add_lookup_account(
                    row.get("B"), row.get("D"), row.get("C"),
                    currency_raw=row.get("E"),
                    opening_date_raw=row.get("F"),
                    maturity_date_raw=row.get("G"),
                    product_raw=row.get("H"),
                    provider_raw="VCB",
                )
                # VTB detail style
                _add_lookup_account(
                    row.get("E"),
                    _prefer_lookup_amount_raw(row.get("N"), row.get("L")),
                    row.get("B"),
                    term_raw=row.get("G"),
                    rate_raw=row.get("H"),
                    currency_raw=row.get("I"),
                    opening_date_raw=row.get("J"),
                    maturity_date_raw=row.get("K"),
                    product_raw=row.get("F"),
                    provider_raw="VTB",
                )
                # BIDV-like style
                _add_lookup_account(
                    row.get("F") or row.get("E"), row.get("G"), row.get("C"),
                    term_raw=row.get("N"),
                    rate_raw=row.get("M"),
                    currency_raw=row.get("L"),
                    opening_date_raw=row.get("O"),
                    maturity_date_raw=row.get("P"),
                    provider_raw="BIDV",
                )

            if lookup or account_meta:
                return lookup, account_meta

        except Exception as xlsx_err:
            logger.debug(f"xlsx XML parse failed: {xlsx_err}")

        # Try .xls
        try:
            import xlrd
            wb = xlrd.open_workbook(file_contents=file_content)
            ws = wb.sheet_by_index(0)

            # SNP: extract entity name from header blob in row 1
            # Format: "\xa0Tên khách hàng / tên tài khoản\xa0:0315694672 CONG TY TNHH BW SAI GON\n..."
            snp_entity = ""
            for r in range(min(3, ws.nrows)):
                cell_val = str(ws.cell_value(r, 0) if ws.ncols > 0 else "")
                # Match "ten khach hang" or "Tên khách hàng" with any whitespace/\xa0
                m = re.search(r'[Tt][eê]n[\s\xa0]+kh[aá]ch[\s\xa0]+h[aà]ng.*?:\s*\d+\s+(.+?)(?:\n|$)', cell_val)
                if m:
                    snp_entity = m.group(1).strip()
                    break

            for row_idx in range(ws.nrows):
                row = [ws.cell_value(row_idx, c) for c in range(ws.ncols)]

                def _v(i: int) -> Any:
                    return row[i] if i < len(row) else None

                # VTB detail block
                _add_lookup_account(
                    _v(4),
                    _prefer_lookup_amount_raw(_v(13), _v(11)),
                    _v(1),
                    term_raw=_v(6),
                    rate_raw=_v(7),
                    currency_raw=_v(8),
                    opening_date_raw=_v(9),
                    maturity_date_raw=_v(10),
                    product_raw=_v(5),
                    provider_raw="VTB",
                )
                # BIDV style
                _add_lookup_account(
                    _v(5), _v(6), _v(2),
                    term_raw=_v(13),
                    rate_raw=_v(12),
                    currency_raw=_v(11),
                    opening_date_raw=_v(14),
                    maturity_date_raw=_v(15),
                    provider_raw="BIDV",
                )
                # VCB-like fallback (when saved to xls)
                _add_lookup_account(
                    _v(2), _v(4), _v(3),
                    currency_raw=_v(5),
                    opening_date_raw=_v(6),
                    maturity_date_raw=_v(7),
                    product_raw=_v(8),
                    provider_raw="VCB",
                )
                # VCB alternate layout in .xls:
                # B=entity, C=account, D=amount, E=currency, F=open, G=maturity, H=term/rate note
                _add_lookup_account(
                    _v(1), _v(3), _v(2),
                    currency_raw=_v(4),
                    opening_date_raw=_v(5),
                    maturity_date_raw=_v(6),
                    product_raw=_v(7),
                    provider_raw="VCB",
                )
                # SNP (SinoPac) style: cols A=account, B=currency, D=amount,
                # E=opening_date, F=maturity_date, H=rate
                _add_lookup_account(
                    snp_entity,  # entity from header blob (parsed below)
                    _v(3), _v(0),
                    currency_raw=_v(1),
                    rate_raw=_v(7),
                    opening_date_raw=_v(4),
                    maturity_date_raw=_v(5),
                    provider_raw="SNP",
                )

            return lookup, account_meta

        except Exception as xls_err:
            logger.debug(f"xls parse failed: {xls_err}")

        # ── Try PDF (Woori-style: one deposit per file) ──
        try:
            parsed_pdf = self._parse_woori_pdf_lookup(file_content)
            if parsed_pdf:
                for entry in parsed_pdf:
                    _add_lookup_account(
                        entry.get("entity"), entry.get("amount"), entry.get("account"),
                        currency_raw=entry.get("currency"),
                        rate_raw=entry.get("rate"),
                        term_raw=entry.get("term"),
                        opening_date_raw=entry.get("opening_date"),
                        maturity_date_raw=entry.get("maturity_date"),
                        provider_raw="WOORI",
                    )
                if lookup or account_meta:
                    return lookup, account_meta
        except Exception as pdf_err:
            logger.debug(f"PDF parse failed: {pdf_err}")

        logger.warning("Failed to parse lookup file (tried xlsx, xls, and PDF)")
        return {}, {}

    @staticmethod
    def _parse_woori_pdf_lookup(file_content: bytes) -> List[Dict[str, Any]]:
        """
        Parse Woori Bank PDF lookup files (one deposit per file).

        Extracts key-value pairs from the PDF text:
        - Account number, Account holder (entity), Opening amount, Balance
        - Account Opening Date, Due date (maturity), Deposit Period (term)
        - Applied interest rate (= Basic + Incremental)

        Returns a list with one entry dict per file (Woori = 1 deposit per PDF).
        """
        try:
            import fitz  # PyMuPDF
        except ImportError:
            logger.warning("PyMuPDF (fitz) not installed - cannot parse Woori PDF lookup files")
            return []

        results: List[Dict[str, Any]] = []

        try:
            doc = fitz.open(stream=file_content, filetype="pdf")
        except Exception:
            return []

        full_text = ""
        for page in doc:
            full_text += page.get_text() + "\n"
        doc.close()

        if not full_text.strip():
            return []

        # Helper: extract value after a label (label\nvalue pattern)
        def _extract(label: str) -> str:
            pattern = re.compile(
                rf'{re.escape(label)}\s*\n\s*(.+?)(?:\n|$)',
                re.IGNORECASE,
            )
            m = pattern.search(full_text)
            return m.group(1).strip() if m else ""

        # Account number: "200700045749 [VND]"
        account_raw = _extract("Account number")
        account = re.sub(r'\s*\[.*?\]\s*$', '', account_raw).strip()
        # Currency from account suffix
        currency_m = re.search(r'\[(\w+)\]', account_raw)
        currency = currency_m.group(1) if currency_m else "VND"

        entity = _extract("Account holder")

        # Opening amount: "VND 17,000,000,000" or "VND 17,000,000,000"
        opening_raw = _extract("Opening amount")
        amount_str = re.sub(r'^[A-Z]{3}\s*', '', opening_raw).replace(' ', '').replace(',', '').strip()

        opening_date = _extract("Account Opening Date")
        maturity_date = _extract("Due date")
        term_raw = _extract("Deposit Period")
        rate_raw = _extract("Applied interest rate")

        # Use opening amount (principal) as the lookup amount
        try:
            amount = float(amount_str) if amount_str else None
        except (ValueError, TypeError):
            amount = None

        if account:
            results.append({
                "account": account,
                "entity": entity,
                "amount": amount,
                "currency": currency,
                "opening_date": opening_date,
                "maturity_date": maturity_date,
                "term": term_raw,
                "rate": rate_raw,
            })

        return results


    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # Test Automation (settlement + open-new using test template)
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

    async def run_test_automation(
        self,
        lookup_file_contents: Optional[List[bytes]] = None,
        progress_callback=None,
    ) -> Dict[str, Any]:
        """
        Run settlement + open-new automation using the test template
        (cash_report_for_test.xlsx). No AI calls needed -- natures are pre-classified.

        Creates a temporary session, runs both automations, returns combined results.
        The test session stays available for download until manually deleted.

        Args:
            lookup_file_contents: Optional list of lookup file bytes for open-new account matching.
            progress_callback: Optional callback for progress events.

        Returns:
            Dict with combined settlement + open-new results and session_id for download.
        """
        from .master_template_manager import TEMPLATES_DIR, WORKING_DIR
        import time as _time

        _started_at = _time.monotonic()
        logger.info(
            "Test automation started: lookup_files=%s has_progress_callback=%s",
            len(lookup_file_contents or []),
            bool(progress_callback),
        )
        test_session_id: Optional[str] = None

        def _finish(result: Dict[str, Any]) -> Dict[str, Any]:
            elapsed_ms = int((_time.monotonic() - _started_at) * 1000)
            logger.info(
                "Test automation finished: status=%s test_session_id=%s settlement_entries=%s open_new_entries=%s elapsed_ms=%s",
                result.get("status"),
                result.get("test_session_id"),
                ((result.get("settlement") or {}).get("counter_entries_created", 0) if isinstance(result.get("settlement"), dict) else 0),
                ((result.get("open_new") or {}).get("counter_entries_created", 0) if isinstance(result.get("open_new"), dict) else 0),
                elapsed_ms,
            )
            return result

        def emit(event_type, step, message, detail="", percentage=0, data=None):
            self._emit_progress(
                process="test_automation",
                progress_callback=progress_callback,
                session_id=test_session_id,
                event_type=event_type,
                step=step,
                message=message,
                detail=detail,
                percentage=percentage,
                data=data,
            )

        test_template = TEMPLATES_DIR / "cash_report_for_test.xlsx"
        if not test_template.exists():
            raise FileNotFoundError(f"Test template not found: {test_template}")

        # Create temporary test session
        test_session_id = f"test-{uuid.uuid4().hex[:8]}"
        working_dir = WORKING_DIR / test_session_id
        working_dir.mkdir(parents=True, exist_ok=True)
        working_file = working_dir / "working_master.xlsx"
        shutil.copy2(test_template, working_file)

        emit("progress", "setup", "Test session created", percentage=5)

        try:
            # Step 1: Settlement
            emit("progress", "settlement", "Running settlement automation...", percentage=10)
            settlement_result = await self.run_settlement_automation(
                session_id=test_session_id,
                user_id=None,
                progress_callback=None,
            )

            emit("progress", "settlement_done",
                 f"Settlement: {settlement_result.get('counter_entries_created', 0)} entries created",
                 percentage=50)

            # Step 2: Open-new
            emit("progress", "open_new", "Running open-new automation...", percentage=55)
            open_new_result = await self.run_open_new_automation(
                session_id=test_session_id,
                lookup_file_contents=lookup_file_contents,
                user_id=None,
                progress_callback=None,
            )

            emit("progress", "open_new_done",
                 f"Open-new: {open_new_result.get('counter_entries_created', 0)} entries created",
                 percentage=95)

            combined = {
                "status": "success",
                "test_session_id": test_session_id,
                "settlement": {
                    "status": settlement_result.get("status"),
                    "counter_entries_created": settlement_result.get("counter_entries_created", 0),
                    "settlement_transactions_found": settlement_result.get("settlement_transactions_found", 0),
                    "skipped_no_account": settlement_result.get("skipped_no_account", 0),
                    "skipped_duplicate": settlement_result.get("skipped_duplicate", 0),
                    "saving_rows_removed": settlement_result.get("saving_rows_removed", 0),
                },
                "open_new": {
                    "status": open_new_result.get("status"),
                    "counter_entries_created": open_new_result.get("counter_entries_created", 0),
                    "candidates_found": open_new_result.get("candidates_found", 0),
                    "skipped_no_account": open_new_result.get("skipped_no_account", 0),
                    "skipped_duplicate": open_new_result.get("skipped_duplicate", 0),
                    "acc_char_added": open_new_result.get("acc_char_added", 0),
                    "saving_account_added": open_new_result.get("saving_account_added", 0),
                    "cash_balance_added": open_new_result.get("cash_balance_added", 0),
                },
            }

            emit("complete", "done", "Test automation completed", percentage=100, data=combined)
            return _finish(combined)

        except Exception as e:
            # Clean up on failure
            if working_dir.exists():
                shutil.rmtree(working_dir, ignore_errors=True)
            emit("error", "error", str(e))
            logger.exception("Test automation failed")
            raise

    # ==================== Approve & Export Transactions ====================

    def _check_summary_validation(self, working_file: Path) -> Dict[str, Any]:
        """
        Read Summary sheet row 109 "Test E.O.P" cached values.
        Columns: B=BWID JSC, C=VC3, D=Subsidiaries, E=Total.

        Returns dict: {"valid": bool, "details": {col_name: value}}

        NOTE: Cached values may be stale if the file was never opened in Excel
        after data was written. Use the manual approve endpoint with force=true
        as fallback.
        """
        import openpyxl as _openpyxl

        wb = _openpyxl.load_workbook(working_file, data_only=True, read_only=True)
        try:
            ws = wb["Summary"]
            col_names = {2: "BWID JSC", 3: "VC3", 4: "Subsidiaries", 5: "Total"}
            details = {}
            all_true = True
            for row in ws.iter_rows(min_row=109, max_row=109, min_col=2, max_col=5):
                for cell in row:
                    col_name = col_names.get(cell.column, f"Col{cell.column}")
                    val = cell.value
                    # Accept: True, 1, "TRUE", "true"
                    is_true = val is True or val == 1 or str(val).strip().upper() == "TRUE"
                    details[col_name] = val
                    if not is_true:
                        all_true = False
        finally:
            wb.close()

        return {"valid": all_true, "details": details}

    def _check_all_natures_valid(
        self, transactions: List["MovementTransaction"]
    ) -> Dict[str, Any]:
        """
        Validate that all transactions have valid, known natures.
        This is a necessary condition for correct classification —
        if any transaction has empty/unknown nature, data quality is suspect.

        Returns: {"valid": bool, "invalid_count": int, "invalid_natures": list}
        """
        all_categories = ALL_PAYMENT_CATEGORIES | ALL_RECEIPT_CATEGORIES
        invalid_natures = []
        for tx in transactions:
            nature = (tx.nature or "").strip()
            if not nature or nature not in all_categories:
                invalid_natures.append(
                    f"{(tx.description or '')[:50]}... → '{nature}'"
                )
        return {
            "valid": len(invalid_natures) == 0,
            "invalid_count": len(invalid_natures),
            "invalid_natures": invalid_natures[:10],  # limit for readability
        }

    async def approve_and_export_transactions(
        self,
        session_id: str,
        user_id: Optional[int] = None,
        force: bool = False,
    ) -> Dict[str, Any]:
        """
        Approve a session's classifications and append verified transactions
        to Transactions.csv for future ground-truth classification.

        Validates Summary "Test E.O.P" = TRUE for all entities before proceeding.
        Deduplicates by (description, direction).
        If same description exists with different nature, updates it.

        Args:
            session_id: The session ID to approve
            user_id: Owner user ID for access control
            force: Skip Summary validation check

        Returns:
            Dict with stats: appended, updated, skipped, total
        """
        if user_id is not None:
            await self._verify_session_owner(session_id, user_id)

        _, working_file = await self._ensure_working_file(session_id)

        writer = MovementDataWriter(working_file)
        transactions = writer.get_all_transactions()

        # Validate: all natures must be valid known categories
        if not force:
            nature_check = self._check_all_natures_valid(transactions)
            if not nature_check["valid"]:
                raise ValueError(
                    f"{nature_check['invalid_count']} transactions have invalid natures: "
                    f"{nature_check['invalid_natures']}. Use force=true to override."
                )
            summary_check = self._check_summary_validation(working_file)
            if not summary_check["valid"]:
                failed = {
                    k: v for k, v in summary_check["details"].items()
                    if not (v is True or v == 1 or str(v).strip().upper() == "TRUE")
                }
                raise ValueError(
                    f"Summary validation failed. These columns are not TRUE: "
                    f"{failed}. Use force=true to override."
                )

        if not transactions:
            return {"appended": 0, "updated": 0, "skipped": 0, "total": 0}

        # Filter: only require non-empty description and nature
        valid_transactions = [
            tx for tx in transactions
            if (tx.description or "").strip()
            and (tx.nature or "").strip()
        ]

        if not valid_transactions:
            return {"appended": 0, "updated": 0, "skipped": 0, "total": 0}

        # Load existing CSV into lookup dict
        csv_path = TRANSACTIONS_REFERENCE_FILE
        existing: Dict[Tuple[str, bool], Tuple[int, str, str]] = {}
        # {(normalized_desc, is_receipt): (line_index, raw_line, nature)}
        all_lines: list = []

        if csv_path.exists():
            with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
                all_lines = f.readlines()

            reader = csv.reader(all_lines[1:], delimiter=";")  # skip header
            for i, row in enumerate(reader):
                if not row or len(row) < 2:
                    continue
                desc_raw = row[0]
                # File format: desc;debit;credit;nature (nature is LAST column)
                nature_raw = row[-1]
                nature = self._canonical_reference_nature(nature_raw)
                if not nature:
                    continue
                is_receipt = self._is_receipt_nature(nature)
                desc_norm = self._normalize_reference_description(desc_raw)
                if desc_norm:
                    existing[(desc_norm, is_receipt)] = (i + 1, all_lines[i + 1], nature)

        appended = 0
        updated = 0
        skipped = 0
        lines_to_update: Dict[int, str] = {}  # line_index → new line content
        lines_to_append: list = []

        for tx in valid_transactions:
            # Strip newlines — multi-line descriptions break CSV row structure
            desc = re.sub(r"[\r\n]+", " ", (tx.description or "")).strip()
            nature = (tx.nature or "").strip()
            canonical = self._canonical_reference_nature(nature)
            if not canonical or not desc:
                skipped += 1
                continue

            is_receipt = bool(tx.debit)
            desc_norm = self._normalize_reference_description(desc)
            key = (desc_norm, is_receipt)

            debit_str = str(tx.debit) if tx.debit else ""
            credit_str = str(tx.credit) if tx.credit else ""

            if key in existing:
                _, _, existing_nature = existing[key]
                if existing_nature == canonical:
                    skipped += 1
                    continue
                # Nature changed → update
                line_idx = existing[key][0]
                # Format: desc;debit;credit;nature (matching existing file format)
                new_line = f"{desc};{debit_str};{credit_str};{canonical}\n"
                lines_to_update[line_idx] = new_line
                existing[key] = (line_idx, new_line, canonical)
                updated += 1
            else:
                new_line = f"{desc};{debit_str};{credit_str};{canonical}\n"
                lines_to_append.append(new_line)
                existing[key] = (-1, new_line, canonical)
                appended += 1

        # Write atomically: update in-place + append new
        if lines_to_update or lines_to_append:
            # Apply in-place updates
            for idx, new_content in lines_to_update.items():
                if idx < len(all_lines):
                    all_lines[idx] = new_content

            # Ensure header exists (preserve original if already valid)
            if not all_lines:
                all_lines = ["Bank description;Debit;Credit;Nature\n"]

            # Append new lines
            all_lines.extend(lines_to_append)

            # Atomic write: temp file → rename
            import tempfile
            tmp_fd, tmp_path = tempfile.mkstemp(
                dir=csv_path.parent, suffix=".csv.tmp"
            )
            try:
                with open(tmp_fd, "w", encoding="utf-8-sig", newline="") as f:
                    f.writelines(all_lines)
                Path(tmp_path).replace(csv_path)
            except Exception:
                Path(tmp_path).unlink(missing_ok=True)
                raise

            # Rebuild index to pick up new entries
            self._rebuild_transactions_reference_index()

            logger.info(
                "Approved transactions for session %s: appended=%d, updated=%d, "
                "skipped=%d, total_user_tx=%d",
                session_id, appended, updated, skipped, len(valid_transactions),
            )

        return {
            "appended": appended,
            "updated": updated,
            "skipped": skipped,
            "total": len(valid_transactions),
        }

