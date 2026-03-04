#!/usr/bin/env python3
"""
Script to search for specific account numbers in Movement sheet.
Searches Cash_Report_OpenNew_Wrong.xlsx and Cash_Report_OpenNew_Correct.xlsx
"""

from openpyxl import load_workbook
from pathlib import Path

# Target accounts to search
TARGET_ACCOUNTS = {
    "211000511419",
    "213000511417",
    "212000511418",
    "14537134677054",
    "14540114651017",
    "200700052415",
    "213000511348",
}

FILES_TO_SEARCH = [
    "D:/Project/data-processing-be/sample_for_test/cash_report_automation/file_compare/Cash_Report_OpenNew_Wrong.xlsx",
    "D:/Project/data-processing-be/sample_for_test/cash_report_automation/file_compare/Cash_Report_OpenNew_Correct.xlsx",
]

def search_movement_sheet(file_path):
    """Search Movement sheet in Excel file for target accounts."""
    print(f"\n{'='*100}")
    print(f"Searching: {Path(file_path).name}")
    print(f"{'='*100}")

    try:
        # Load with read_only=False to get actual values (including from formulas)
        wb = load_workbook(file_path, data_only=True)

        # Find Movement sheet
        movement_sheet = None
        for sheet_name in wb.sheetnames:
            if "Movement" in sheet_name:
                movement_sheet = wb[sheet_name]
                break

        if not movement_sheet:
            print("ERROR: Movement sheet not found!")
            return

        print(f"Found Movement sheet: {movement_sheet.title}")

        # Column indices (letter for openpyxl)
        # A=Source, B=Bank, C=Account, D=Date, E=Description, F=Debit, G=Credit, H=?, I=Nature
        col_map = {
            'source': 'A',
            'bank': 'B',
            'account': 'C',
            'date': 'D',
            'description': 'E',
            'debit': 'F',
            'credit': 'G',
            'nature': 'I',
        }

        # Track results
        found_accounts = {acc: [] for acc in TARGET_ACCOUNTS}

        # Get max row to avoid reading too far
        max_row = movement_sheet.max_row
        print(f"Sheet max_row: {max_row}")

        # Read data (header is row 3, so start from row 4)
        for row_num in range(4, max_row + 1):
            account_cell = movement_sheet[f"{col_map['account']}{row_num}"]
            account_value = account_cell.value

            # Skip empty rows
            if not account_value:
                continue

            # Get account value and convert to string
            account = str(account_value).strip()

            # Check if this account is in our target list
            if account in TARGET_ACCOUNTS:
                source = movement_sheet[f"{col_map['source']}{row_num}"].value
                bank = movement_sheet[f"{col_map['bank']}{row_num}"].value
                date = movement_sheet[f"{col_map['date']}{row_num}"].value
                description = movement_sheet[f"{col_map['description']}{row_num}"].value
                debit = movement_sheet[f"{col_map['debit']}{row_num}"].value
                credit = movement_sheet[f"{col_map['credit']}{row_num}"].value
                nature = movement_sheet[f"{col_map['nature']}{row_num}"].value

                found_accounts[account].append({
                    'row': row_num,
                    'source': source,
                    'bank': bank,
                    'account': account,
                    'date': date,
                    'description': description,
                    'debit': debit,
                    'credit': credit,
                    'nature': nature,
                })

        # Print results
        total_found = 0
        for account in sorted(TARGET_ACCOUNTS):
            records = found_accounts[account]
            total_found += len(records)

            if records:
                print(f"\nAccount: {account}")
                print("-" * 100)
                for rec in records:
                    print(f"  Row {rec['row']}: Source={rec['source']}, Bank={rec['bank']}, Account={rec['account']}")
                    print(f"           Date={rec['date']}, Description={rec['description']}")
                    print(f"           Debit={rec['debit']}, Credit={rec['credit']}, Nature={rec['nature']}")
                    print()
            else:
                print(f"\nAccount: {account} - NOT FOUND")

        print(f"\nTotal records found: {total_found}")
        wb.close()

    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()

def main():
    """Main entry point."""
    print("\nSearching for specific account numbers in Movement sheet")
    print(f"Target accounts: {sorted(TARGET_ACCOUNTS)}\n")

    for file_path in FILES_TO_SEARCH:
        search_movement_sheet(file_path)

    print(f"\n{'='*100}")
    print("Search complete")
    print(f"{'='*100}\n")

if __name__ == "__main__":
    main()
