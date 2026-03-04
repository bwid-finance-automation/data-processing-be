import openpyxl
import os
import sys
import io

# Force UTF-8 output on Windows to handle Vietnamese characters
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

KEYWORDS = [
    "sao hoa",
    "supply chain",
    "suplly chain",
    "my phuoc",
    "19037134677017",
    "19040114651012",
    "100700488704",
    "200700052415",
]

FILES = [
    r"D:\Project\data-processing-be\sample_for_test\cash_report_automation\file_compare\Cash_Report_OpenNew_Correct.xlsx",
    r"D:\Project\data-processing-be\sample_for_test\cash_report_automation\file_compare\Cash_Report_OpenNew_Wrong.xlsx",
]

COLUMNS = {
    "A": "Source",
    "B": "Bank",
    "C": "Account",
    "D": "Date",
    "E": "Description",
    "F": "Debit",
    "G": "Credit",
    "H": "Net",
    "I": "Nature",
    "J": "Entity",
}

# Column indices (1-based): A=1, B=2, ..., P=16
COL_A = 1
COL_P = 16
COL_J = 10  # Entity is column J


def cell_contains_keyword(cell_value, keywords):
    if cell_value is None:
        return False
    val_lower = str(cell_value).lower()
    for kw in keywords:
        if kw in val_lower:
            return True
    return False


def search_movement_sheet(filepath):
    filename = os.path.basename(filepath)
    print(f"\n{'='*80}")
    print(f"FILE: {filename}")
    print(f"{'='*80}")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    # Find Movement sheet
    sheet_names = wb.sheetnames
    movement_sheet = None
    for name in sheet_names:
        if "movement" in name.lower():
            movement_sheet = name
            break

    if movement_sheet is None:
        print(f"  ERROR: No 'Movement' sheet found. Available sheets: {sheet_names}")
        wb.close()
        return

    print(f"  Sheet found: '{movement_sheet}'")
    ws = wb[movement_sheet]

    matching_rows = []

    for row in ws.iter_rows(min_row=4, max_col=COL_P):  # Skip header rows 1-3, check cols A-P
        # Get row number from first non-empty cell, or skip if row is empty
        row_num = None
        for c in row:
            if hasattr(c, 'row') and c.row is not None:
                row_num = c.row
                break
        if row_num is None:
            continue
        # Check if any cell in columns A-P contains a keyword
        matched = False
        for cell in row:
            if cell_contains_keyword(cell.value, KEYWORDS):
                matched = True
                break
        if matched:
            # Extract columns A through J for display
            row_data = {}
            for col_idx, col_letter in enumerate(["A","B","C","D","E","F","G","H","I","J"], start=1):
                cell = row[col_idx - 1] if col_idx - 1 < len(row) else None
                row_data[col_letter] = cell.value if cell is not None else None
            matching_rows.append((row_num, row_data))

    wb.close()

    if not matching_rows:
        print(f"  No matching rows found.")
        return

    print(f"  Found {len(matching_rows)} matching rows:\n")

    for row_num, data in matching_rows:
        print(f"  --- Row {row_num} ---")
        print(f"    Source (A)     : {data['A']}")
        print(f"    Bank (B)       : {data['B']}")
        print(f"    Account (C)    : {data['C']}")
        print(f"    Date (D)       : {data['D']}")
        print(f"    Description (E): {data['E']}")
        print(f"    Debit (F)      : {data['F']}")
        print(f"    Credit (G)     : {data['G']}")
        print(f"    Net (H)        : {data['H']}")
        print(f"    Nature (I)     : {data['I']}")
        print(f"    Entity (J)     : {data['J']}")
        print()


if __name__ == "__main__":
    for filepath in FILES:
        search_movement_sheet(filepath)
    print("\nDone.")
