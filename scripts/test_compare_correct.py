"""
Compare CORRECT vs WRONG open-new output files for specific target accounts.

Checks Saving Account, Cash Balance, Acc_Char, and Movement sheets
to show what the correct code produced vs what the buggy code produced.
"""

import openpyxl

CORRECT_FILE = r"D:\Project\data-processing-be\sample_for_test\cash_report_automation\file_compare\Cash_Report_OpenNew_Correct.xlsx"
WRONG_FILE   = r"D:\Project\data-processing-be\sample_for_test\cash_report_automation\file_compare\Cash_Report_OpenNew_Wrong.xlsx"

# Target accounts and their expected entity labels.
# Key = saving account string (as it appears in Saving Account col C),
# Value = human-readable entity label used in printout.
TARGET_ACCOUNTS = {
    "114628748888_1":   "WINLOCK 2A",
    "213000511348":     "THUAN DAO",
    "116646026868_1":   "XENIA 1",
    "118625469999_1":   "XENIA 2",
    "19037134677017_1": "SAO HOA",
    "19040114651012_1": "BW SUPLLY CHAIN CITY",
    "200700052415":     "BW MY PHUOC 3",
}

# Also keep a set of bare account numbers (without _suffix) for Cash Balance
# and Acc_Char lookups, since those sheets use the raw account numbers.
TARGET_BASE_ACCOUNTS = {
    "114628748888":   "WINLOCK 2A",
    "213000511348":   "THUAN DAO",
    "116646026868":   "XENIA 1",
    "118625469999":   "XENIA 2",
    "19037134677017": "SAO HOA",
    "19040114651012": "BW SUPLLY CHAIN CITY",
    "200700052415":   "BW MY PHUOC 3",
}

# ---------------------------------------------------------------------------
# Helper utilities
# ---------------------------------------------------------------------------

def _str(v) -> str:
    """Safely convert a cell value to a stripped string."""
    if v is None:
        return ""
    return str(v).strip()


def _find_sheet(wb, *keywords):
    """
    Return the first sheet whose name contains ALL of the given keywords
    (case-insensitive).  Returns None if not found.
    """
    for name in wb.sheetnames:
        lower = name.lower()
        if all(kw.lower() in lower for kw in keywords):
            return wb[name]
    return None


def _header(title: str):
    print()
    print("=" * 90)
    print(f"  {title}")
    print("=" * 90)


def _section(title: str):
    print()
    print(f"  --- {title} ---")


# ---------------------------------------------------------------------------
# Sheet readers
# ---------------------------------------------------------------------------

def read_saving_account(wb):
    """
    Read the Saving Account sheet.
    Returns a dict: saving_account_str -> {field: value, ...}

    Column mapping (1-based openpyxl index):
      A=1  entity
      C=3  account
      D=4  account type
      E=5  currency
      H=8  opening date
      I=9  term months
      K=11 interest rate
      L=12 maturity date
      Y=25 bank name
    """
    ws = _find_sheet(wb, "saving")
    if ws is None:
        print("  [WARNING] Saving Account sheet not found")
        return {}

    data = {}
    for row in ws.iter_rows(min_row=2, max_col=25, values_only=True):
        # row is a tuple indexed 0-based
        entity       = _str(row[0])   # A
        account      = _str(row[2])   # C
        acc_type     = _str(row[3])   # D
        currency     = _str(row[4])   # E
        opening_date = row[7]          # H  (keep raw for date display)
        term_months  = row[8]          # I
        interest_rate = row[10]        # K
        maturity_date = row[11]        # L  (keep raw)
        bank_name    = _str(row[24])  # Y

        if not account:
            continue

        data[account] = {
            "entity":        entity,
            "account":       account,
            "acc_type":      acc_type,
            "currency":      currency,
            "opening_date":  opening_date,
            "term_months":   term_months,
            "interest_rate": interest_rate,
            "maturity_date": maturity_date,
            "bank":          bank_name,
        }
    return data


def read_cash_balance(wb):
    """
    Read the Cash Balance sheet (excluding Prior Period).
    Returns a dict: account_str -> {field: value, ...}

    Column mapping (1-based openpyxl index):
      A=1  entity
      C=3  account
      Y=25 bank
    """
    ws = _find_sheet(wb, "cash", "balance")
    # Exclude "Prior Period" sheets that also contain "cash balance"
    if ws is None or "prior" in ws.title.lower():
        # Try harder: iterate manually
        for name in wb.sheetnames:
            low = name.lower()
            if "cash" in low and "balance" in low and "prior" not in low:
                ws = wb[name]
                break
    if ws is None:
        print("  [WARNING] Cash Balance sheet not found")
        return {}

    data = {}
    for row in ws.iter_rows(min_row=2, max_col=25, values_only=True):
        entity  = _str(row[0])   # A
        account = _str(row[2])   # C
        bank    = _str(row[24])  # Y

        if not account:
            continue

        data[account] = {
            "entity":  entity,
            "account": account,
            "bank":    bank,
        }
    return data


def read_acc_char(wb):
    """
    Read the Acc_Char sheet.
    Returns a dict: account_str -> {field: value, ...}

    Column mapping (1-based openpyxl index):
      B=2  account
      C=3  code
    """
    ws = _find_sheet(wb, "acc")
    if ws is None:
        print("  [WARNING] Acc_Char sheet not found")
        return {}

    data = {}
    for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
        account = _str(row[1])  # B
        code    = _str(row[2])  # C

        if not account:
            continue

        data[account] = {
            "account": account,
            "code":    code,
        }
    return data


def read_movement_internal_transfers(wb):
    """
    Read the Movement sheet and return all rows whose nature column contains
    'Internal transfer in'.

    We scan the header row (row 1) to locate the Nature and Account columns
    dynamically, falling back to known positions if not found.

    Returns a list of dicts with keys: row_num, account, nature, description, amount
    """
    ws = _find_sheet(wb, "movement")
    if ws is None:
        print("  [WARNING] Movement sheet not found")
        return []

    # Attempt to locate header columns dynamically in the first row
    nature_col   = None
    account_col  = None
    desc_col     = None
    amount_col   = None

    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        for idx, cell in enumerate(row):
            cv = _str(cell).lower()
            if "nature" in cv and nature_col is None:
                nature_col = idx
            if ("account" in cv or "acc" in cv) and account_col is None:
                account_col = idx
            if ("description" in cv or "desc" in cv) and desc_col is None:
                desc_col = idx
            if ("debit" in cv or "amount" in cv or "credit" in cv) and amount_col is None:
                amount_col = idx
        break

    # Fallback column indices (0-based) if header scan failed
    # Typical Movement layout based on the codebase context:
    #   A=0 date, B=1 entity, C=2 branch, D=3 transaction_date, E=4 description,
    #   F=5 debit, G=6 credit, H=7 ..., J=9 entity (formula), K=10 nature
    if nature_col  is None: nature_col  = 10   # K
    if account_col is None: account_col = 4    # E  (description often contains account refs)
    if desc_col    is None: desc_col    = 4    # E
    if amount_col  is None: amount_col  = 5    # F

    results = []
    row_num = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_num += 1
        if not row or len(row) <= nature_col:
            continue

        nature = _str(row[nature_col])
        if "internal transfer in" not in nature.lower():
            continue

        account = _str(row[account_col]) if account_col < len(row) else ""
        desc    = _str(row[desc_col])    if desc_col    < len(row) else ""
        amount  = row[amount_col]         if amount_col  < len(row) else None

        results.append({
            "row_num":     row_num,
            "nature":      nature,
            "account_col": account,
            "description": desc,
            "amount":      amount,
        })

    return results


# ---------------------------------------------------------------------------
# Printing helpers
# ---------------------------------------------------------------------------

SAVING_FIELDS = [
    ("entity",        "Entity (A)"),
    ("account",       "Account (C)"),
    ("acc_type",      "Acc Type (D)"),
    ("currency",      "Currency (E)"),
    ("opening_date",  "Opening Date (H)"),
    ("term_months",   "Term Months (I)"),
    ("interest_rate", "Interest Rate (K)"),
    ("maturity_date", "Maturity Date (L)"),
    ("bank",          "Bank Name (Y)"),
]

CB_FIELDS = [
    ("entity",  "Entity (A)"),
    ("account", "Account (C)"),
    ("bank",    "Bank (Y)"),
]

AC_FIELDS = [
    ("account", "Account (B)"),
    ("code",    "Code (C)"),
]


def _fmt_val(v) -> str:
    if v is None:
        return "(empty)"
    s = str(v).strip()
    return s if s else "(empty)"


def _diff_marker(correct_val, wrong_val) -> str:
    """Return ' *** MISMATCH ***' when values differ, otherwise ''."""
    c = _fmt_val(correct_val)
    w = _fmt_val(wrong_val)
    return " *** MISMATCH ***" if c != w else ""


def print_comparison_table(
    label: str,
    saving_acc: str,
    entity_hint: str,
    correct_record,
    wrong_record,
    fields,
):
    """Print a side-by-side comparison for one account."""
    print(f"\n  [{label}]  saving_account={saving_acc!r}  ({entity_hint})")
    print(f"  {'Field':<25}  {'CORRECT':^35}  {'WRONG':^35}  {'Status'}")
    print(f"  {'-'*25}  {'-'*35}  {'-'*35}  {'-'*15}")

    for field_key, field_label in fields:
        c_val = correct_record.get(field_key) if correct_record else None
        w_val = wrong_record.get(field_key)   if wrong_record   else None

        c_str = _fmt_val(c_val)[:35]
        w_str = _fmt_val(w_val)[:35]
        diff  = _diff_marker(c_val, w_val)

        status = "MATCH" if not diff else "MISMATCH"
        print(f"  {field_label:<25}  {c_str:<35}  {w_str:<35}  {status}{diff}")

    if correct_record is None:
        print(f"  !! Account NOT FOUND in CORRECT file")
    if wrong_record is None:
        print(f"  !! Account NOT FOUND in WRONG file")


# ---------------------------------------------------------------------------
# Main comparison logic
# ---------------------------------------------------------------------------

def main():
    print()
    print("=" * 90)
    print("  OPEN-NEW OUTPUT FILE COMPARISON")
    print(f"  CORRECT: {CORRECT_FILE}")
    print(f"  WRONG  : {WRONG_FILE}")
    print("=" * 90)

    # -----------------------------------------------------------------------
    # Load workbooks (read-only for performance)
    # -----------------------------------------------------------------------
    print("\nLoading workbooks (read_only=True) ...")
    wb_correct = openpyxl.load_workbook(CORRECT_FILE, read_only=True, data_only=True)
    wb_wrong   = openpyxl.load_workbook(WRONG_FILE,   read_only=True, data_only=True)

    print(f"  CORRECT sheets: {wb_correct.sheetnames}")
    print(f"  WRONG   sheets: {wb_wrong.sheetnames}")

    # -----------------------------------------------------------------------
    # Read all relevant sheets
    # -----------------------------------------------------------------------
    print("\nReading Saving Account sheets ...")
    sa_correct = read_saving_account(wb_correct)
    sa_wrong   = read_saving_account(wb_wrong)
    print(f"  CORRECT: {len(sa_correct)} saving accounts found")
    print(f"  WRONG  : {len(sa_wrong)}   saving accounts found")

    print("\nReading Cash Balance sheets ...")
    cb_correct = read_cash_balance(wb_correct)
    cb_wrong   = read_cash_balance(wb_wrong)
    print(f"  CORRECT: {len(cb_correct)} cash balance entries found")
    print(f"  WRONG  : {len(cb_wrong)}   cash balance entries found")

    print("\nReading Acc_Char sheets ...")
    ac_correct = read_acc_char(wb_correct)
    ac_wrong   = read_acc_char(wb_wrong)
    print(f"  CORRECT: {len(ac_correct)} Acc_Char entries found")
    print(f"  WRONG  : {len(ac_wrong)}   Acc_Char entries found")

    print("\nReading Movement sheets (Internal transfer in rows) ...")
    mv_correct = read_movement_internal_transfers(wb_correct)
    mv_wrong   = read_movement_internal_transfers(wb_wrong)
    print(f"  CORRECT: {len(mv_correct)} 'Internal transfer in' rows")
    print(f"  WRONG  : {len(mv_wrong)}   'Internal transfer in' rows")

    # -----------------------------------------------------------------------
    # SECTION 1: Saving Account comparison
    # -----------------------------------------------------------------------
    _header("SECTION 1: SAVING ACCOUNT SHEET")
    print("  Checking target accounts in Saving Account sheet (col C = account).")
    print("  Note: saving account keys include _1 suffix for new open-new rows.")

    for sa_key, entity_hint in TARGET_ACCOUNTS.items():
        correct_rec = sa_correct.get(sa_key)
        wrong_rec   = sa_wrong.get(sa_key)

        # If the exact key is missing, try alternative: with/without suffix
        if correct_rec is None:
            # Try bare account (without suffix)
            base = sa_key.split("_")[0]
            correct_rec = sa_correct.get(base)
            if correct_rec:
                sa_key_display = f"{sa_key} (found as '{base}')"
            else:
                sa_key_display = sa_key
        else:
            sa_key_display = sa_key

        if wrong_rec is None:
            base = sa_key.split("_")[0]
            wrong_rec = sa_wrong.get(base)

        print_comparison_table(
            label="Saving Account",
            saving_acc=sa_key_display,
            entity_hint=entity_hint,
            correct_record=correct_rec,
            wrong_record=wrong_rec,
            fields=SAVING_FIELDS,
        )

    # -----------------------------------------------------------------------
    # Extra: dump all saving account keys that contain any target fragment
    # -----------------------------------------------------------------------
    _section("All Saving Account keys in each file that partially match targets")
    target_fragments = set()
    for k in TARGET_ACCOUNTS:
        target_fragments.add(k.split("_")[0])  # bare number
    target_fragments.update(TARGET_ACCOUNTS.keys())  # with suffix

    def _is_target_account(acc_str):
        for frag in target_fragments:
            if frag in acc_str or acc_str in frag:
                return True
        return False

    correct_matching = {k: v for k, v in sa_correct.items() if _is_target_account(k)}
    wrong_matching   = {k: v for k, v in sa_wrong.items()   if _is_target_account(k)}

    print(f"\n  Saving Account keys in CORRECT file matching any target:")
    for k, v in sorted(correct_matching.items()):
        print(f"    {k:<30}  entity={v['entity']!r:<35}  bank={v['bank']!r}")

    print(f"\n  Saving Account keys in WRONG file matching any target:")
    for k, v in sorted(wrong_matching.items()):
        print(f"    {k:<30}  entity={v['entity']!r:<35}  bank={v['bank']!r}")

    # -----------------------------------------------------------------------
    # SECTION 2: Cash Balance comparison
    # -----------------------------------------------------------------------
    _header("SECTION 2: CASH BALANCE SHEET")
    print("  Checking base accounts (without _suffix) in Cash Balance sheet (col C = account).")

    for base_acc, entity_hint in TARGET_BASE_ACCOUNTS.items():
        correct_rec = cb_correct.get(base_acc)
        wrong_rec   = cb_wrong.get(base_acc)

        print_comparison_table(
            label="Cash Balance",
            saving_acc=base_acc,
            entity_hint=entity_hint,
            correct_record=correct_rec,
            wrong_record=wrong_rec,
            fields=CB_FIELDS,
        )

    # -----------------------------------------------------------------------
    # Extra: Cash Balance keys partially matching any target
    # -----------------------------------------------------------------------
    _section("Cash Balance keys in each file partially matching targets")

    correct_cb_matching = {k: v for k, v in cb_correct.items() if _is_target_account(k)}
    wrong_cb_matching   = {k: v for k, v in cb_wrong.items()   if _is_target_account(k)}

    print(f"\n  Cash Balance keys in CORRECT file matching any target:")
    for k, v in sorted(correct_cb_matching.items()):
        print(f"    {k:<30}  entity={v['entity']!r:<35}  bank={v['bank']!r}")

    print(f"\n  Cash Balance keys in WRONG file matching any target:")
    for k, v in sorted(wrong_cb_matching.items()):
        print(f"    {k:<30}  entity={v['entity']!r:<35}  bank={v['bank']!r}")

    # -----------------------------------------------------------------------
    # SECTION 3: Acc_Char comparison
    # -----------------------------------------------------------------------
    _header("SECTION 3: ACC_CHAR SHEET")
    print("  Checking saving account keys (with _suffix) in Acc_Char sheet (col B = account).")
    print("  Also checking base accounts as fallback.")

    for sa_key, entity_hint in TARGET_ACCOUNTS.items():
        correct_rec = ac_correct.get(sa_key)
        wrong_rec   = ac_wrong.get(sa_key)

        # Fallback to bare account
        if correct_rec is None:
            base = sa_key.split("_")[0]
            correct_rec = ac_correct.get(base)
        if wrong_rec is None:
            base = sa_key.split("_")[0]
            wrong_rec = ac_wrong.get(base)

        print_comparison_table(
            label="Acc_Char",
            saving_acc=sa_key,
            entity_hint=entity_hint,
            correct_record=correct_rec,
            wrong_record=wrong_rec,
            fields=AC_FIELDS,
        )

    # -----------------------------------------------------------------------
    # Extra: dump Acc_Char keys matching any target
    # -----------------------------------------------------------------------
    _section("Acc_Char keys in each file partially matching targets")

    correct_ac_matching = {k: v for k, v in ac_correct.items() if _is_target_account(k)}
    wrong_ac_matching   = {k: v for k, v in ac_wrong.items()   if _is_target_account(k)}

    print(f"\n  Acc_Char keys in CORRECT file matching any target:")
    for k, v in sorted(correct_ac_matching.items()):
        print(f"    {k:<30}  code={v['code']!r}")

    print(f"\n  Acc_Char keys in WRONG file matching any target:")
    for k, v in sorted(wrong_ac_matching.items()):
        print(f"    {k:<30}  code={v['code']!r}")

    # -----------------------------------------------------------------------
    # SECTION 4: Movement sheet — Internal transfer in rows
    # -----------------------------------------------------------------------
    _header("SECTION 4: MOVEMENT SHEET — 'Internal transfer in' ROWS")
    print("  Listing all counter-entry rows (nature = 'Internal transfer in').")
    print("  These are the saving-account credit rows inserted by the settlement step.")

    _section("CORRECT file — Internal transfer in rows")
    if mv_correct:
        print(f"  {'Row':<6}  {'Nature':<25}  {'Acc/Desc col':<35}  {'Amount'}")
        print(f"  {'-'*6}  {'-'*25}  {'-'*35}  {'-'*20}")
        for rec in mv_correct:
            amount_str = f"{rec['amount']:>20,.0f}" if isinstance(rec['amount'], (int, float)) else _fmt_val(rec['amount'])
            print(f"  {rec['row_num']:<6}  {rec['nature']:<25}  {rec['account_col'][:35]:<35}  {amount_str}")
    else:
        print("  (none found)")

    _section("WRONG file — Internal transfer in rows")
    if mv_wrong:
        print(f"  {'Row':<6}  {'Nature':<25}  {'Acc/Desc col':<35}  {'Amount'}")
        print(f"  {'-'*6}  {'-'*25}  {'-'*35}  {'-'*20}")
        for rec in mv_wrong:
            amount_str = f"{rec['amount']:>20,.0f}" if isinstance(rec['amount'], (int, float)) else _fmt_val(rec['amount'])
            print(f"  {rec['row_num']:<6}  {rec['nature']:<25}  {rec['account_col'][:35]:<35}  {amount_str}")
    else:
        print("  (none found)")

    # Count difference
    _section("Movement counter-entry count comparison")
    print(f"  CORRECT: {len(mv_correct)} 'Internal transfer in' rows")
    print(f"  WRONG  : {len(mv_wrong)}   'Internal transfer in' rows")
    if len(mv_correct) != len(mv_wrong):
        print(f"  *** COUNT MISMATCH: {len(mv_correct)} vs {len(mv_wrong)} ***")
    else:
        print(f"  Counts match.")

    # -----------------------------------------------------------------------
    # SECTION 5: Full Saving Account dump for ALL target fragments
    # -----------------------------------------------------------------------
    _header("SECTION 5: FULL SAVING ACCOUNT ROW DUMP FOR TARGET ACCOUNTS")
    print("  All fields for each matched account, CORRECT vs WRONG side by side.")

    all_target_keys = set(TARGET_ACCOUNTS.keys())
    # Also include bare numbers in case the file uses those
    for k in list(all_target_keys):
        all_target_keys.add(k.split("_")[0])

    for sa_key in sorted(all_target_keys):
        entity_hint = TARGET_ACCOUNTS.get(sa_key, TARGET_BASE_ACCOUNTS.get(sa_key.split("_")[0], ""))
        c_rec = sa_correct.get(sa_key)
        w_rec = sa_wrong.get(sa_key)
        if c_rec is None and w_rec is None:
            continue  # Skip keys that appear in neither file

        print(f"\n  Account: {sa_key!r}  ({entity_hint})")
        print(f"  {'Field':<25}  {'CORRECT':^40}  {'WRONG':^40}")
        print(f"  {'-'*25}  {'-'*40}  {'-'*40}")
        for field_key, field_label in SAVING_FIELDS:
            c_val = c_rec.get(field_key) if c_rec else "(NOT IN FILE)"
            w_val = w_rec.get(field_key) if w_rec else "(NOT IN FILE)"
            c_str = _fmt_val(c_val)[:40]
            w_str = _fmt_val(w_val)[:40]
            marker = " <-- DIFF" if _fmt_val(c_val) != _fmt_val(w_val) else ""
            print(f"  {field_label:<25}  {c_str:<40}  {w_str:<40}{marker}")

    # -----------------------------------------------------------------------
    # Close workbooks
    # -----------------------------------------------------------------------
    wb_correct.close()
    wb_wrong.close()

    print()
    print("=" * 90)
    print("  DONE")
    print("=" * 90)


if __name__ == "__main__":
    main()
