"""
Debug script: trace the open_new matching for problem entities.
"""
import openpyxl
import re
from difflib import SequenceMatcher

WRONG_FILE = r"D:\Project\data-processing-be\sample_for_test\cash_report_automation\file_compare\Cash_Report_OpenNew_Wrong.xlsx"

ENTITY_NOISE = {
    "CT", "CTY", "CONG", "TY", "TNHH", "MTV", "CP", "CO", "PHAN", "VA",
    "PHAT", "TRIEN", "NGHIEP", "DAU", "TU", "DU", "AN", "MOT", "THANH",
    "VIEN", "HUU", "HAN", "JSC", "PTCN", "BW",
}

def normalize_entity(value):
    text = str(value or "").upper()
    text = text.replace("BWID", "BW")
    text = re.sub(r'[^A-Z0-9]+', ' ', text)
    tokens = [tok for tok in text.split() if tok and tok not in ENTITY_NOISE]
    return "".join(tokens)

def entity_score(left, right):
    l = normalize_entity(left)
    r = normalize_entity(right)
    if not l or not r:
        return 0.0
    if l == r:
        return 1.0
    base = SequenceMatcher(None, l, r).ratio()
    if l in r or r in l:
        base = max(base, 0.85)
    return base

TARGETS = {
    "WINLOCK 2A": {"account": "114628748888", "bank": "VTB"},
    "THUAN DAO": {"account": "116002949864", "bank": "VTB"},
    "XENIA 1": {"account": "116646026868", "bank": "VTB"},
    "XENIA 2": {"account": "118625469999", "bank": "VTB"},
    "BW SUPLLY CHAIN CITY": {"account": "19040114651012", "bank": "TCB"},
    "SAO HOA": {"account": "19037134677017", "bank": "TCB"},
    "BW MY PHUOC 3": {"account": "100700488704", "bank": "WOORI"},
}

print("=" * 80)
print("STEP 1: Check Cash Balance for target entity accounts")
print("=" * 80)

wb = openpyxl.load_workbook(WRONG_FILE, data_only=True, read_only=True)

for name in wb.sheetnames:
    if "cash balance" in name.lower() and "prior" not in name.lower():
        print(f"Cash Balance sheet: '{name}'")
        ws = wb[name]
        account_entity_map = {}
        for row in ws.iter_rows(min_row=2, max_col=30, values_only=False):
            entity_val = row[0].value  # A
            account_val = row[2].value  # C
            if account_val:
                acc_str = str(account_val).strip()
                if acc_str:
                    account_entity_map[acc_str] = str(entity_val or "").strip()

        print(f"Total accounts in Cash Balance: {len(account_entity_map)}")

        for entity_name, info in TARGETS.items():
            tx_acc = info["account"]
            found = account_entity_map.get(tx_acc, "")
            print(f"  {entity_name}: account={tx_acc} -> entity='{found}' ({'FOUND' if found else 'NOT FOUND'})")

        print("\n  Searching for WINLOCK in Cash Balance:")
        for acc, ent in account_entity_map.items():
            if "WINLOCK" in str(ent).upper() or "WIN LOCK" in str(ent).upper():
                print(f"    account={acc} -> entity='{ent}'")

        print("\n  Searching for XENIA in Cash Balance:")
        for acc, ent in account_entity_map.items():
            if "XENIA" in str(ent).upper():
                print(f"    account={acc} -> entity='{ent}'")

        print("\n  Searching for THUAN DAO in Cash Balance:")
        for acc, ent in account_entity_map.items():
            if "THUAN" in str(ent).upper() and "DAO" in str(ent).upper():
                print(f"    account={acc} -> entity='{ent}'")

        print("\n  Searching for SAO HOA in Cash Balance:")
        for acc, ent in account_entity_map.items():
            if "SAO" in str(ent).upper() and "HOA" in str(ent).upper():
                print(f"    account={acc} -> entity='{ent}'")

        print("\n  Searching for SUPPLY CHAIN in Cash Balance:")
        for acc, ent in account_entity_map.items():
            if "SUPPLY" in str(ent).upper() or "SUPLLY" in str(ent).upper():
                print(f"    account={acc} -> entity='{ent}'")

        print("\n  Searching for MY PHUOC in Cash Balance:")
        for acc, ent in account_entity_map.items():
            if "MY PHUOC" in str(ent).upper() or "MYPHUOC" in str(ent).upper():
                print(f"    account={acc} -> entity='{ent}'")
        break

wb.close()

print("\n" + "=" * 80)
print("STEP 2: Entity similarity scores")
print("=" * 80)

vtb_entities = [
    "CT CP PHAT TRIEN CONG NGHIEP THUAN DAO",
    "CONG TY TNHH PHAT TRIEN CN BW TAN PHU TRUNG",
    "CTCP PHAT TRIEN CONG NGHIEP PHO NOI A",
    "CT TNHH PHAT TRIEN CONG NGHIEP LE MINH XUAN 3 SAIGON",
    "CONG TY TNHH WIN LOCK 2B",
    "CT TNHH PHAT TRIEN CONG NGHIEP BW BAU BANG BB06",
    "CTY CP SAO HOA TOAN QUOC",
]

for target_name in ["WINLOCK 2A", "THUAN DAO", "XENIA 1", "XENIA 2"]:
    print(f"\n  Scores for '{target_name}' (norm: '{normalize_entity(target_name)}'):")
    scores = []
    for vtb_ent in vtb_entities:
        score = entity_score(target_name, vtb_ent)
        scores.append((score, vtb_ent, normalize_entity(vtb_ent)))
    scores.sort(reverse=True)
    for score, ent, norm in scores[:5]:
        print(f"    {score:.4f}: '{ent}' (norm: '{norm}')")

print("\n" + "=" * 80)
print("STEP 3: Description extraction check")
print("=" * 80)

descriptions = {
    "WINLOCK 2A": "GUI TIEN THEO HOP DONG TIEN GUI CO KY HAN SO 285.2026.49236 NGAY 12.02.2026",
    "THUAN DAO": "HDTG 900/2026/49150",
    "XENIA 1": "HOP DONG TIEN GUI SO 285/2026/49235",
    "XENIA 2": "GUI TIEN THEO HDTG CO KY HAN SO 285.2026.49237 NGAY 12.02.2026",
    "BW SUPLLY CHAIN CITY": "CT TNHH BW SUPPLY CHAIN CITY GUI TIEN NGAY 12/02/2026 THEO HDTG",
    "SAO HOA row604": "CONG TY CP SAO HOA TOAN QUOC GUI TIEN NGAY 09/02/2026 THEO",
    "BW MY PHUOC 3 row880": "Withdrawal - Withdrawal",
}

for name, desc in descriptions.items():
    result = None
    m = re.search(r'\(TK\s*(\d{6,20})\)', desc, re.IGNORECASE)
    if m:
        result = ("Pattern1_TK", m.group(1))
    if not result:
        m = re.search(r'(?:tai\s*khoan|account)\s*(\d{6,20})', desc, re.IGNORECASE)
        if m:
            result = ("Pattern2_taikhoan", m.group(1))
    if not result:
        ctx = re.search(r'HDTG|GUI\s*TIEN|MO\s*HD', desc, re.IGNORECASE)
        if ctx:
            m = re.search(r'SO\s*(\d{8,20})', desc, re.IGNORECASE)
            if m:
                result = ("Pattern3_SO", m.group(1))
            else:
                result = ("Pattern3_ctx_NO_SO", None)
        else:
            result = ("No_ctx", None)
    if not result or result[1] is None:
        if re.match(r'^[\d,.\sE+]+$', desc.strip()):
            m = re.search(r'(\d{10,20})', desc)
            if m:
                result = ("Pattern4_bare", m.group(1))

    print(f"  {name}: {result}")

print("\n" + "=" * 80)
print("STEP 4: Check Acc_Char for target accounts")
print("=" * 80)

wb = openpyxl.load_workbook(WRONG_FILE, data_only=True, read_only=True)
for name in wb.sheetnames:
    if "acc_char" in name.lower() or "acc char" in name.lower():
        print(f"Acc_Char sheet: '{name}'")
        ws = wb[name]
        account_to_code = {}
        for row in ws.iter_rows(min_row=2, max_col=11, values_only=False):
            account = row[1].value
            code = row[2].value
            if account and code:
                account_to_code[str(account).strip()] = str(code).strip()

        print(f"Total Acc_Char entries: {len(account_to_code)}")
        for entity_name, info in TARGETS.items():
            tx_acc = info["account"]
            code = account_to_code.get(tx_acc, "")
            print(f"  {entity_name}: current_acc={tx_acc} -> code='{code}'")

        print("\n  Codes for WINLOCK/XENIA/THUAN:")
        for acc, code in account_to_code.items():
            if any(x in code.upper() for x in ["W2A", "WL1", "WL2", "XA1", "XA2", "XEN", "WIN", "TDA"]):
                print(f"    account={acc} -> code='{code}'")
        break
wb.close()
