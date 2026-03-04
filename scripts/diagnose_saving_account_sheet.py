"""
Diagnostic script: inspect the Saving Account sheet in the open-new output
and dump relevant rows + lookup data for the target accounts.
"""
import sys
import os

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl

XLSX_FILE = r"D:\Project\data-processing-be\sample_for_test\cash_report_automation\Cash_Report_0ea06c34_open_new.xlsx"
LOOKUP_DIR = r"D:\Project\data-processing-be\sample_for_test\cash_report_automation\lookup_file"

TARGET_ACCOUNTS = {"213000511348", "200700052415"}
# Also match anything ending with _1 or _2

# ── 1. Open the workbook ─────────────────────────────────────────────────────

print("Opening workbook (read_only, data_only)...")
wb = openpyxl.load_workbook(XLSX_FILE, read_only=True, data_only=True)
print(f"Sheet names: {wb.sheetnames}\n")

# Find the Saving Account sheet
saving_sheet_name = None
for name in wb.sheetnames:
    if "saving" in name.lower():
        saving_sheet_name = name
        break

if not saving_sheet_name:
    print("ERROR: Could not find 'Saving Account' sheet!")
    sys.exit(1)

print(f"Using sheet: '{saving_sheet_name}'\n")
ws = wb[saving_sheet_name]

# ── 2. Dump header row ───────────────────────────────────────────────────────

print("=" * 100)
print("HEADER ROW (row 3, which is usually the column header row in the template)")
print("=" * 100)

rows_iter = ws.iter_rows(values_only=True)

# Read all rows into memory to find headers and data
all_rows = []
for row in rows_iter:
    all_rows.append(row)

print(f"Total rows read by openpyxl: {len(all_rows)}\n")

# Print the first few rows to find the header
for idx, row in enumerate(all_rows[:6], start=1):
    non_empty = [(i, v) for i, v in enumerate(row) if v is not None]
    if non_empty:
        print(f"Row {idx}: {non_empty[:30]}")

print()

# Find the header row (look for "Entity" or "Account" in column A or B area)
header_row_idx = None
header_row = None
for idx, row in enumerate(all_rows):
    row_vals = [str(v).strip().lower() if v else "" for v in row[:10]]
    if any("entity" in v or "account" in v for v in row_vals):
        header_row_idx = idx
        header_row = row
        print(f"Found header at row index {idx + 1}: {list(row[:30])}")
        break

if header_row is None:
    print("Could not auto-detect header row. Using row 3 as header.")
    header_row_idx = 2
    header_row = all_rows[2] if len(all_rows) > 2 else []

# Print full header with column letters
print("\n" + "=" * 100)
print("FULL COLUMN LAYOUT (header row):")
print("=" * 100)
import string

def col_letter(idx_0based):
    """Convert 0-based column index to Excel column letter."""
    n = idx_0based + 1
    result = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result

for i, val in enumerate(header_row):
    if val is not None:
        print(f"  Col {col_letter(i):>3} (col {i+1:>3}): {val!r}")

# ── 3. Find and print target rows ────────────────────────────────────────────

print("\n" + "=" * 100)
print("SCANNING ALL DATA ROWS FOR TARGET ACCOUNTS")
print("=" * 100)

# Determine data start row (one after header)
data_start_idx = header_row_idx + 1 if header_row_idx is not None else 3

# Column indices (0-based) — default from standard cash report template
# We'll try to detect them from header, falling back to known positions
def find_col(header, *names):
    """Find 0-based column index by matching header names (case-insensitive)."""
    names_lower = [n.lower() for n in names]
    for i, v in enumerate(header):
        if v and any(n in str(v).lower() for n in names_lower):
            return i
    return None

col_entity = find_col(header_row, "entity") or 0       # A
col_branch = find_col(header_row, "branch") or 1       # B
col_account = find_col(header_row, "account no", "account") or 2  # C
col_acc_type = find_col(header_row, "acc type", "type") or 3      # D
col_currency = find_col(header_row, "currency", "curr") or 4      # E

print(f"\nDetected column indices (0-based): entity={col_entity}, branch={col_branch}, "
      f"account={col_account}, acc_type={col_acc_type}, currency={col_currency}")

# Print ALL columns for found rows
matched_rows = []

for row_num, row in enumerate(all_rows[data_start_idx:], start=data_start_idx + 1):
    if not any(v is not None for v in row):
        continue

    # Extract account value
    account_val = row[col_account] if len(row) > col_account else None
    acc_str = str(account_val).strip().replace("'", "") if account_val is not None else ""

    # Check for match
    is_target = acc_str in TARGET_ACCOUNTS
    is_suffix = (
        "_1" in acc_str or "_2" in acc_str or
        any(acc_str.startswith(ta) for ta in TARGET_ACCOUNTS)
    )

    if is_target or is_suffix:
        matched_rows.append((row_num, row, acc_str))

print(f"\nMatched {len(matched_rows)} row(s):\n")

for row_num, row, acc_str in matched_rows:
    print(f"--- Row {row_num} | Account: {acc_str} ---")
    for i, val in enumerate(row):
        if val is not None or i < 30:
            letter = col_letter(i)
            # Determine if this is a "known" column
            label = ""
            if i == col_entity:       label = " [Entity]"
            elif i == col_branch:     label = " [Branch]"
            elif i == col_account:    label = " [Account]"
            elif i == col_acc_type:   label = " [Acc Type]"
            elif i == col_currency:   label = " [Currency]"
            if val is not None:
                print(f"  Col {letter:>3} ({i+1:>3}){label}: {val!r}")
    print()

# ── 4. Print last 60 rows (the "bottom area" likely containing new rows) ─────

print("=" * 100)
print("BOTTOM AREA OF SHEET (last 60 data rows) — newly inserted rows are here")
print("=" * 100)

data_rows = [(i + data_start_idx + 1, r) for i, r in enumerate(all_rows[data_start_idx:])
             if any(v is not None for v in r)]

bottom_rows = data_rows[-60:] if len(data_rows) > 60 else data_rows

print(f"Showing rows {bottom_rows[0][0] if bottom_rows else '?'} to "
      f"{bottom_rows[-1][0] if bottom_rows else '?'} of the sheet:\n")

for row_num, row in bottom_rows:
    account_val = row[col_account] if len(row) > col_account else None
    entity_val  = row[col_entity]  if len(row) > col_entity  else None
    acc_str = str(account_val).strip().replace("'", "") if account_val is not None else ""
    ent_str = str(entity_val).strip() if entity_val is not None else ""

    # Print key columns compactly
    key_cols = {}
    for i, v in enumerate(row):
        if v is not None:
            key_cols[col_letter(i)] = v

    # Compact summary line
    non_empty_pairs = [(col_letter(i), v) for i, v in enumerate(row) if v is not None]
    print(f"  Row {row_num:>4} | Acct={acc_str:<25} | Entity={ent_str:<30} | "
          f"Cols({len(non_empty_pairs)}): {non_empty_pairs[:15]}")

wb.close()

# ── 5. Parse lookup files ─────────────────────────────────────────────────────

print("\n" + "=" * 100)
print("PARSING LOOKUP FILES")
print("=" * 100)

from app.application.finance.cash_report.cash_report_service import CashReportService

service = CashReportService.__new__(CashReportService)

all_account_details = {}

for fname in os.listdir(LOOKUP_DIR):
    fpath = os.path.join(LOOKUP_DIR, fname)
    print(f"\nParsing: {fname}")
    try:
        with open(fpath, "rb") as f:
            content = f.read()
        parsed, details = service._parse_saving_lookup_file_with_metadata(content)
        print(f"  -> {len(details)} accounts found")
        for acc, meta in details.items():
            if acc:
                dst = all_account_details.setdefault(acc, {})
                for mk, mv in meta.items():
                    if mv not in (None, ""):
                        if dst.get(mk) in (None, ""):
                            dst[mk] = mv
    except Exception as e:
        print(f"  ERROR: {e}")

print(f"\nTotal unique accounts in all lookups: {len(all_account_details)}")

# ── 6. Print lookup data for target accounts ──────────────────────────────────

print("\n" + "=" * 100)
print("LOOKUP DATA FOR TARGET ACCOUNTS")
print("=" * 100)

def normalize_acc(raw):
    s = str(raw or "").strip().replace("'", "")
    try:
        if "." in s and float(s) == int(float(s)):
            s = str(int(float(s)))
    except (ValueError, OverflowError):
        pass
    return s

# Normalize lookup keys
normalized_details = {}
for acc, meta in all_account_details.items():
    acc_n = normalize_acc(acc)
    if acc_n:
        dst = normalized_details.setdefault(acc_n, {})
        for mk, mv in (meta or {}).items():
            if mv not in (None, ""):
                if dst.get(mk) in (None, ""):
                    dst[mk] = mv

print(f"\nAll accounts in lookup (normalized):")
for acc in sorted(normalized_details.keys()):
    meta = normalized_details[acc]
    print(f"  {acc}: provider={meta.get('provider')}, entity={meta.get('entity')}, "
          f"amount={meta.get('amount')}, rate={meta.get('interest_rate')}, "
          f"term={meta.get('term_months')}M/{meta.get('term_days')}D, "
          f"opening={meta.get('opening_date')}, maturity={meta.get('maturity_date')}")

for target_acc in sorted(TARGET_ACCOUNTS):
    print(f"\n{'='*60}")
    print(f"TARGET ACCOUNT: {target_acc}")
    print(f"{'='*60}")

    meta = normalized_details.get(target_acc)
    if meta is None:
        print(f"  NOT FOUND in any lookup file")
        # Try partial match
        partial = [acc for acc in normalized_details if target_acc in acc or acc in target_acc]
        if partial:
            print(f"  Partial matches: {partial}")
            for p in partial:
                print(f"    {p}: {normalized_details[p]}")
    else:
        print(f"  provider      : {meta.get('provider')}")
        print(f"  entity        : {meta.get('entity')}")
        print(f"  amount        : {meta.get('amount')}")
        print(f"  interest_rate : {meta.get('interest_rate')}")
        print(f"  term_months   : {meta.get('term_months')}")
        print(f"  term_days     : {meta.get('term_days')}")
        print(f"  opening_date  : {meta.get('opening_date')}")
        print(f"  maturity_date : {meta.get('maturity_date')}")
        print(f"  Full meta     : {meta}")

print("\nDone.")
